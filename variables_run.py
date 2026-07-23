"""
variables_run.py - фоновый прогон «Проверки КП» (пункт 1.4).

Для каждого поддомена из «Карты присутствия» (catalogs/{proj}-kp.csv) качает
главную страницу и сверяет с КП:
  • город / страна - нет ли чужого (region_checker);
  • телефоны (поиск/реклама/общий) - номер на сайте входит в набор КП города;
  • почта, адрес, Telegram, WhatsApp - совпадают с КП.
Результат пишется в cache/variables/<proj>/variables.xlsx (лист «Проверка КП» +
лист «Расхождения»). Прогресс идёт в stdout, откуда его читает вкладка.

Запуск:
    python variables_run.py --project smu
    python variables_run.py --project imp --cities "Москва,Казань"
Прокси (для проектов, блокирующих зарубежный IP) - через env proxy_url.
"""
import argparse
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

ROOT = Path(__file__).parent
WORK_ROOT = ROOT / 'cache' / 'variables'

PROJECT_NAMES = {
    'smu': 'СМУ - Стальметурал', 'imp': 'ИМП - Инметпром',
    'mpe': 'МПЭ - Мепэн', 'avia': 'АПС - Авиапромсталь',
}

# Порядок и подписи переменных-колонок. Телефоны - с префиксом «Тел.» (чтобы не
# путать с колонкой «Город»), в порядке КП: общий → реклама → SEO. «Страна»
# убрана по просьбе заказчика; справочная колонка «Страна(КП)» остаётся.
VAR_COLUMNS = ["Город", "Тел. Общий Город", "Тел. Реклама Город", "Тел. SEO Город",
               "Почта", "Адрес", "Telegram", "WhatsApp"]

_SYMBOL = {"ok": "✓", "ok_set": "✓", "bug": "✗", "warn": "⚠", "na": "–"}
_COLOR = {"ok": "1E8E3E", "ok_set": "1E8E3E", "bug": "C62828",
          "warn": "B26A00", "na": "9E9E9E"}


def _stamp(msg):
    print(f'[{datetime.now().strftime("%H:%M:%S")}] {msg}', flush=True)


def _use_proxy(project: str) -> bool:
    p = ROOT / 'projects' / f'{project}.json'
    try:
        return bool(json.loads(p.read_text(encoding='utf-8')).get('use_proxy'))
    except Exception:
        return False


_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

# Для повторов «как из другого браузера»: разные User-Agent + всегда без кеша.
_UA_POOL = [
    _UA,
    ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 "
     "(KHTML, like Gecko) Version/17.4 Safari/605.1.15"),
    ("Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) "
     "Gecko/20100101 Firefox/125.0"),
    ("Mozilla/5.0 (Linux; Android 14; SM-S918B) AppleWebKit/537.36 "
     "(KHTML, like Gecko) Chrome/124.0.0.0 Mobile Safari/537.36"),
]


def _proxy_parts(proxy):
    """(proxy_host, proxy_port, proxy_headers|None) из proxy-URL. () если нет."""
    if not proxy:
        return None
    from urllib.parse import urlparse
    pr = urlparse(proxy if '://' in proxy else 'http://' + proxy)
    if not pr.hostname:
        return None
    headers = {}
    if pr.username:
        import base64
        tok = base64.b64encode(
            f"{pr.username}:{pr.password or ''}".encode()).decode()
        headers['Proxy-Authorization'] = f'Basic {tok}'
    return pr.hostname, pr.port or 8080, headers


def _fetch_one(dom, proxy_parts, ua=None, path='/'):
    """Скачивает https://<dom><path> через http.client (CONNECT-туннель с
    Proxy-Authorization в CONNECT-запросе - надёжный способ прокси-авторизации
    для HTTPS, в отличие от aiohttp, который упорно отдавал 407). Один редирект
    в пределах того же/родственного хоста поддерживаем. Каждый вызов - свежее
    соединение без кеша/куки (по сути «инкогнито»); ua позволяет притвориться
    другим браузером на повторах. → (html, ошибка)."""
    import http.client
    import ssl
    ua = ua or _UA

    def _get(host, path, depth=0):
        conn = None
        try:
            if proxy_parts:
                phost, pport, phdrs = proxy_parts
                conn = http.client.HTTPSConnection(
                    phost, pport, timeout=30, context=ssl.create_default_context())
                conn.set_tunnel(host, 443, headers=dict(phdrs))
            else:
                conn = http.client.HTTPSConnection(host, 443, timeout=30)
            conn.request('GET', path or '/', headers={
                'User-Agent': ua, 'Accept-Encoding': 'identity',
                'Accept': 'text/html,application/xhtml+xml',
                'Cache-Control': 'no-cache', 'Pragma': 'no-cache'})
            resp = conn.getresponse()
            if resp.status in (301, 302, 303, 307, 308) and depth < 3:
                loc = resp.getheader('Location') or ''
                resp.read()
                conn.close()
                from urllib.parse import urlparse, urljoin
                nu = urlparse(urljoin(f'https://{host}{path or "/"}', loc))
                return _get(nu.hostname or host,
                            (nu.path or '/') + (f'?{nu.query}' if nu.query else ''),
                            depth + 1)
            if resp.status >= 400:
                resp.read()
                return '', f'HTTP {resp.status}'
            # Сервер иногда обрывает ответ (IncompleteRead: не дослал все байты
            # из Content-Length). Прочитанную часть сохраняем, но помечаем как
            # НЕПОЛНУЮ - тогда fetch_all повторит (адрес/контакты в ПОДВАЛЕ, а он
            # в конце документа: у обрезанной страницы его нет). Если и повторы
            # оборвутся - используем самый полный partial (лучше, чем провал).
            truncated = False
            try:
                data = resp.read()
            except http.client.IncompleteRead as ie:
                data = ie.partial or b''
                truncated = True
            html = data.decode('utf-8', 'replace')
            # Антибот/заглушка вместо страницы (частый ответ при частых запросах):
            # либо совсем короткий ответ, либо явные маркеры проверки браузера.
            # Помечаем ошибкой - тогда fetch_all повторит попытку свежим соединением.
            low = html[:5000].lower()
            _block = ('ddos-guard', 'challenge-platform', 'attention required',
                      'checking your browser', 'проверяем ваш браузер',
                      'проверка вашего браузера', 'запрос отправили вы, а не робот',
                      'доступ ограничен', 'captcha-delivery')
            if len(html) < 1500 or (len(html) < 25000 and any(m in low for m in _block)):
                return '', 'похоже на антибот/капчу'
            if truncated:
                return html, 'неполная страница (обрыв соединения)'
            return html, ''
        finally:
            try:
                if conn:
                    conn.close()
            except Exception:
                pass

    try:
        return _get(dom, path or '/')
    except Exception as e:  # noqa: BLE001
        return '', (str(e)[:200] or e.__class__.__name__)


# Ссылка на «Контакты» в навигации главной - чтобы догрузить эту страницу для
# сверки адреса (у части проектов - МПЭ/mepen - адрес выводится только там, в
# карточке «Адрес: …», а в подвале главной его нет). Берём <a> с текстом
# «Контакты» либо href с /contacts//kontakty/ на этом же хосте.
_CONTACTS_LINK_RE = re.compile(
    r'<a\b[^>]*?href=["\']([^"\']+)["\'][^>]*>(?:(?!</a>)[\s\S]){0,80}?контакт',
    re.I)
_CONTACTS_PATH_RE = re.compile(
    r'href=["\']([^"\']*/(?:contacts?|kontakt\w*)/[^"\']*)["\']', re.I)


def _find_contacts_path(html, dom):
    """Путь страницы «Контакты» на этом же хосте (из навигации главной) или ''."""
    from urllib.parse import urlparse, urljoin
    from kp import _norm_host
    host = _norm_host(dom)
    cands = ([m.group(1) for m in list(_CONTACTS_LINK_RE.finditer(html or ''))[:6]]
             + [m.group(1) for m in list(_CONTACTS_PATH_RE.finditer(html or ''))[:6]])
    for href in cands:
        href = (href or '').strip()
        if not href or href.startswith(('#', 'tel:', 'mailto:', 'javascript:')):
            continue
        p = urlparse(urljoin(f'https://{dom}/', href))
        if (p.hostname and _norm_host(p.hostname) == host
                and p.path and p.path != '/'):
            return p.path + (f'?{p.query}' if p.query else '')
    return ''


def _проверить_ссылку(url, base_dom, proxy_parts):
    """HTTP-код ссылки кнопки (для WhatsApp с битой ссылкой). Возвращает int-код,
    'stub' для пустых ссылок (#/javascript/mailto/tel) или None, если не проверить.
    Относительные ссылки раскрываем относительно домена страницы."""
    from urllib.parse import urljoin, urlparse
    import http.client
    import ssl
    u = (url or '').strip()
    if not u or u.startswith('#') or u.lower().startswith(('javascript:', 'mailto:', 'tel:')):
        return 'stub'
    full = urljoin(f'https://{base_dom}/', u)
    p = urlparse(full)
    if not p.hostname:
        return 'stub'
    try:
        if proxy_parts:
            phost, pport, phdrs = proxy_parts
            conn = http.client.HTTPSConnection(phost, pport, timeout=20,
                                               context=ssl.create_default_context())
            conn.set_tunnel(p.hostname, 443, headers=dict(phdrs))
        else:
            conn = http.client.HTTPSConnection(p.hostname, 443, timeout=20)
        conn.request('GET', (p.path or '/') + (f'?{p.query}' if p.query else ''),
                     headers={'User-Agent': _UA, 'Accept-Encoding': 'identity'})
        resp = conn.getresponse()
        resp.read()
        conn.close()
        return resp.status
    except Exception:
        return None


def fetch_all(domains, proxy, log, retries=3):
    """Качает главные всех поддоменов МЯГКО, чтобы сервер не сыпал 500: малый
    пул (3 потока) + разнесённые старты запросов (pace), не бьём залпом. Bitrix
    у mepen обслуживает все поддомены одним бэкендом и отдаёт 500/рвёт соединение,
    когда в него летит много параллельных запросов. Не загрузившиеся повторяет
    ещё `retries` раз - только их, «как из другого браузера» (свежее соединение,
    другой User-Agent), с РАСТУЩЕЙ паузой (бэкофф 5→10→20с) и ещё меньшим пулом.
    Успешные повторно не трогает. → {domain: (html, ошибка)}."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    parts = _proxy_parts(proxy)
    out: dict = {}
    N = len(domains)

    def _pass(items, workers, counting, ua=None, pace=0.0):
        with ThreadPoolExecutor(max_workers=workers) as ex:
            # Разносим СТАРТЫ запросов (pace) - не бьём по серверу залпом. Bitrix
            # у mepen отдаёт 500, когда все поддомены (один бэкенд) дёргают разом.
            futs = {}
            for d, row in items:
                futs[ex.submit(_fetch_one, d, parts, ua)] = (d, row.city)
                if pace:
                    time.sleep(pace)
            for k, fut in enumerate(as_completed(futs), 1):
                dom, city = futs[fut]
                try:
                    html, err = fut.result()
                except Exception as e:  # noqa: BLE001
                    html, err = '', str(e)[:200]
                # Держим ЛУЧШИЙ результат: успешный (без ошибки) в приоритете,
                # иначе - самый ПОЛНЫЙ partial (у обрыва подвал может быть длиннее
                # на другом заходе). Так повтор не затрёт больший кусок меньшим.
                prev = out.get(dom)
                if (not err) or prev is None or len(html) > len(prev[0] or ''):
                    out[dom] = (html, err)
                if counting:
                    log(f'  [{k}/{N}] {dom} ({city}): '
                        + (f'ошибка загрузки - {err}' if err else 'загружено'))
                else:
                    log(f'    повтор {dom} ({city}): '
                        + (f'снова ошибка - {err}' if err else 'загружено ✓'))

    _pass(domains, 3, counting=True, pace=0.15)
    # Повторяем ТОЛЬКО упавшие: сервер часто отдаёт 500 при частых параллельных
    # запросах - на повторе ещё меньшим пулом, с РАСТУЩЕЙ паузой (бэкофф, даём
    # серверу остыть) и от имени другого браузера большинство доходит.
    for attempt in range(1, retries + 1):
        failed = [(d, row) for d, row in domains if out.get(d, ('', ''))[1]]
        if not failed:
            break
        ua = _UA_POOL[attempt % len(_UA_POOL)]
        _back = min(30, 5 * (2 ** (attempt - 1)))   # 5с → 10с → 20с
        log(f'↻ Повтор {attempt}/{retries}: заново пробуем {len(failed)} '
            f'не загрузившихся, помягче (пауза {_back}с, 2 потока, другой браузер)…')
        time.sleep(_back)
        _pass(failed, 2, counting=False, ua=ua, pace=0.35)
    n_ok = sum(1 for v in out.values() if not v[1])
    n_fail = N - n_ok
    log(f'Итог загрузки: {n_ok} из {N}'
        + (f' (осталось с ошибкой: {n_fail})' if n_fail else ' (все загрузились)'))
    return out


def _own_city_regex(city):
    """Regex «свой город на странице», устойчивый к склонениям. ё→е; для каждого
    слова несколько основ: само слово, без конечной гласной/й/ь, без прилагат.
    окончания (-ый/-ий→«Новый»→«нов»), с беглой гласной (-ец/-ёл: «Череповец»→
    «череповц», «Орёл»→«орл»). Для СВОЕГО города на его поддомене (омонимы не
    страшны, поэтому берём широко). None - если город пуст."""
    import re as _re
    city = (city or "").replace("ё", "е").strip()
    if not city:
        return None
    parts = []
    for w in _re.split(r"[\s-]+", city):
        wl = w.lower()
        if len(wl) < 2:
            parts.append(_re.escape(wl))
            continue
        stems = {wl}
        if len(wl) > 4 and wl[-2:] in (
                "ый", "ий", "ой", "ая", "яя", "ое", "ее", "ей"):
            stems.add(wl[:-2])                     # прилагательное: Новый→нов
        if wl[-1] in "аяоеиыуюьй":
            stems.add(wl[:-1])                     # конечная гласная/й/ь
        if len(wl) > 3 and wl[-1] not in "аяоеиыуюё" and wl[-2] in "ео":
            stems.add(wl[:-2] + wl[-1])            # беглая гласная: Череповец→череповц
        alt = "|".join(_re.escape(s) for s in sorted(stems, key=len))
        parts.append(r"(?:" + alt + r")[а-яё]{0,3}")
    return _re.compile(r"(?<![а-яё])" + r"[\s-]+".join(parts) + r"(?![а-яё])",
                       _re.I)


# Виджет выбора города на сайте: «Город: Нижний Новгород», «Ваш город: Казань»,
# «Ваш регион — Москва». Берём город из этой подписи (именительный падеж), чтобы
# показать, что выведено на сайте, когда в КП города нет.
_ГОРОД_НА_САЙТЕ_RE = re.compile(
    r'(?:ваш\s+)?(?:город|регион)\s*[:\-–—]?\s*'
    r'([А-ЯЁ][а-яё]+(?:[-\s]+[А-ЯЁ][а-яё]+){0,2})', re.I)
# Служебные хвосты после названия города, которые надо отсечь, если склеились.
_ГОРОД_ХВОСТ_RE = re.compile(
    r'\s+(?:изменить|сменить|выбрать|другой|верно|да|нет|это)\b.*', re.I)


def _город_на_сайте(html: str) -> str:
    """Город, ВЫВЕДЕННЫЙ на сайте (виджет «Город: …» / «Ваш город: …»), в
    именительном падеже. '' - если такой подписи на странице нет."""
    try:
        from text_checker import html_to_visible_text
        text = html_to_visible_text(html or "")
    except Exception:
        text = html or ""
    m = _ГОРОД_НА_САЙТЕ_RE.search(text)
    if not m:
        return ""
    cand = re.sub(r"\s+", " ", m.group(1)).strip(" .,;·|")
    cand = _ГОРОД_ХВОСТ_RE.sub("", cand).strip(" .,;·|")
    return cand


def _регион_статусы(html, host, ctx):
    """Город/страна через region_checker → (город_dict, страна_dict) в формате
    check_variables-поля {field, expected, found, status, note}."""
    import region_checker as rc
    свой = (ctx.host_city.get(host, "") or "").strip()
    # Город в КП убрали/не задали (пусто или прочерк «-») - проверять нечего.
    # Раньше искали на странице сам прочерк «-» и выдавали ложное «не найден» ✗;
    # теперь это «нет в КП» (–), как у пустых телефона/почты. Сравнивать не с чем.
    if свой in ("-", "–", "—"):
        свой = ""
    город = {"field": "Город", "expected": свой or "–", "found": "–",
             "status": "na", "note": "в КП город не задан" if not свой else ""}
    страна = {"field": "Страна", "expected": ctx.host_country.get(host, "–"),
              "found": "–", "status": "na", "note": ""}
    # В КП города нет, но на сайте он ВЫВЕДЕН - показываем какой (⚠), чтобы было
    # видно: сайт город показывает, а в КП он не заполнен (по просьбе заказчика).
    if not свой:
        _site_city = _город_на_сайте(html)
        if _site_city:
            город.update(found=_site_city, status="warn",
                         note="на сайте выведен город, а в КП город не указан")
    # Город: СТРОГО - город из КП должен быть выведен на странице (шапка/H1/
    # title/текст). Нет = ✗ (переменная не подставилась / другой город). Плюс
    # ловим ЧУЖОЙ город проекта (затесался номер/город другого поддомена).
    if свой:
        try:
            zones = rc.извлечь_зоны(html)
            # ё→е: на сайте пишут и «Могилёв», и «Могилеве»; нормализуем обе
            # стороны, чтобы ё/е не давало ложного расхождения.
            hay = " ".join([zones.get("title", ""), zones.get("h1", ""),
                            zones.get("description", ""),
                            (zones.get("текст", "") or "")[:8000]]).replace("ё", "е")
            # СВОЙ город на его же поддомене ищем устойчиво к склонениям (Тула→
            # Туле, Череповец→Череповце, Орёл→Орле, Новый Уренгой→Новом Уренгое).
            rx = _own_city_regex(свой)
            found_self = bool(rx and rc._city_match_propernoun(rx, hay))
            rv = rc.check_region_vars(html, host, ctx)
            foreign = ([i for i in (rv.get("issues") or [])
                        if i.get("тип") == "город"] if rv else [])
            if found_self and not foreign:
                город.update(found="есть на сайте", status="ok",
                             note="совпадает с КП")
            elif not found_self:
                город.update(found="не найден на сайте", status="bug",
                             note="город из КП на сайте не найден "
                             "(не подставился или выведен другой)")
            else:
                город.update(found="есть чужой город", status="bug",
                             note=foreign[0].get("пояснение", ""))
        except Exception:  # noqa: BLE001
            pass
    try:
        cm = rc.check_cis_mentions(html, host, ctx)
        if cm is None:
            страна.update(status="na", note="РФ - проверка чужих стран не нужна"
                          if ctx.host_country.get(host) == "Россия" else "")
        else:
            iss = cm.get("issues", [])
            if iss:
                _zru = {'title': 'title', 'description': 'description',
                        'h1': 'H1', 'текст': 'текст страницы'}
                parts = []
                for it in iss[:2]:
                    z = _zru.get(it.get('зона', ''), it.get('зона', ''))
                    ctxt = (it.get('контекст') or '').strip()
                    parts.append(f'«{it.get("найдено", "?")}» в {z}: …{ctxt}…'
                                 if ctxt else f'«{it.get("найдено", "?")}» в {z}')
                note = '; '.join(parts)
                if len(iss) > 2:
                    note += f' (и ещё {len(iss) - 2})'
                страна.update(found="есть чужая страна", status="bug", note=note)
            else:
                страна.update(found="чисто", status="ok", note="")
    except Exception:  # noqa: BLE001
        pass
    return город, страна


def _только_почта_для_перевода(city: str, fields: list) -> list:
    """Переводная копия сайта (город в КП помечен «(перевод)», напр. steelgroup.az
    → «Азербайджан (перевод)») - это дубль основного домена с теми же переменными.
    На такой строке проверяем ТОЛЬКО «Почту», остальные колонки помечаем «–» (не
    проверяем): иначе плодятся ложные ✗ вроде «города „Азербайджан (перевод)“ нет
    на странице» (это метка КП, а не реальный город). ЧИСТАЯ функция (юнит-тест)."""
    if "(перевод)" not in (city or "").lower():
        return fields
    for f in fields:
        if f.get("field") != "Почта":
            f.update(status="na", found="–",
                     note="переводная версия сайта - проверяем только почту")
    return fields


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('--project', required=True, choices=list(PROJECT_NAMES))
    ap.add_argument('--cities', default='', help='города через запятую (пусто = все)')
    a = ap.parse_args()

    sys.path.insert(0, str(ROOT))
    import kp as kpmod
    from region_checker import build_region_context

    # Источник КП: Google-таблица или снапшот CSV. Логируем ЯВНО - иначе не
    # видно, подтянулись ли правки из таблицы (частая причина «поменял данные в
    # Google, а проверка их не заметила» = обновление молча не прошло).
    _kp_url = ''
    try:
        import kp_sheets as _kps
        _kp_url = _kps.kp_sheet_url(a.project)
    except Exception:
        _kps = None
    if _kp_url and _kps:
        try:
            _ok, _msg = _kps.refresh_project(a.project, log=lambda *x, **k: None)
            _stamp('КП ← Google-таблица: '
                   + ('обновлено из таблицы' if _ok
                      else f'НЕ удалось ({_msg}) - беру прежний снапшот CSV'))
        except Exception as _e:
            _stamp(f'КП ← Google: ошибка обновления ({_e}) - беру снапшот CSV')
    else:
        _stamp(f'⚠️ КП: ссылка на Google-таблицу НЕ задана (секрет '
               f'kp_sheet_url_{a.project}) - беру СНАПШОТ '
               f'catalogs/{a.project}-kp.csv. Правки в Google так НЕ '
               f'подхватятся - обнови снапшот или задай секрет!')

    # Города КП списком - по одному городу-владельцу на сайт. У СНГ-стран все
    # города делят один сайт (stalmetural.kz/.by/.uz - поддоменов нет): в отчёт
    # берём только город со своей ссылкой, безссылочные города-спутники убираем
    # (иначе сверялись бы с чужим городским сайтом и давали ложные ошибки).
    kp_rows = kpmod.load_kp_rows(a.project)   # уже обновили выше
    if not kp_rows:
        _stamp(f'✗ Нет базы КП catalogs/{a.project}-kp.csv')
        return 2
    try:
        _csvp = ROOT / 'catalogs' / f'{a.project}-kp.csv'
        _mt = (datetime.fromtimestamp(_csvp.stat().st_mtime).strftime('%d.%m.%Y %H:%M')
               if _csvp.exists() else '–')
    except Exception:
        _mt = '–'
    _stamp(f'КП загружена: {len(kp_rows)} городов, снапшот обновлён {_mt}')

    wanted = {c.strip().lower() for c in a.cities.split(',') if c.strip()}
    domains = [(row.domain, row) for row in kp_rows
               if not wanted or (row.city or '').lower() in wanted]
    # Порядок как в КП: страны в порядке появления в КП, но Россия первой;
    # внутри страны сохраняем исходный порядок КП (сортировка стабильная).
    _country_seq = []
    for _row in kp_rows:
        _c = (_row.country or '').strip()
        if _c and _c not in _country_seq:
            _country_seq.append(_c)

    def _crank(row):
        c = (row.country or '').strip()
        if c.lower() in ('россия', 'рф'):
            return -1
        return _country_seq.index(c) if c in _country_seq else 10 ** 6
    domains.sort(key=lambda x: _crank(x[1]))

    # Прокси используем ТОЛЬКО для проектов с use_proxy=true (напр. ИМП, который
    # блокирует зарубежный IP). СМУ/МПЭ (use_proxy=false) качаем напрямую - им
    # прокси не нужен, а сломанный proxy_url иначе давал бы им ложный 407.
    proxy = (os.environ.get('proxy_url') or '').strip() or None
    if _use_proxy(a.project):
        if not proxy:
            _stamp('⚠️ У проекта use_proxy=true, а proxy_url не задан - '
                   'зарубежный IP может блокироваться (будут ошибки загрузки).')
    else:
        if proxy:
            _stamp(f'Проект {a.project}: use_proxy=false - страницы качаем '
                   'напрямую, без прокси.')
        proxy = None
    # Диагностика прокси (без вывода самих логина/пароля).
    _pp = _proxy_parts(proxy)
    if _pp:
        _ph, _pport, _phdrs = _pp
        _stamp(f'Прокси: {_ph}:{_pport}; авторизация в proxy_url: '
               + ('есть' if _phdrs.get('Proxy-Authorization')
                  else 'НЕТ - в ссылке нет логина:пароля (будет 407)'))
    elif proxy:
        _stamp('⚠️ proxy_url задан, но не разобрался '
               '(ожидается http://логин:пароль@хост:порт).')

    # Регион-контекст строим из dict-КП (по одному городу на домен) - ему нужен
    # набор городов/телефонов для сверки «чужой город на странице».
    kp = kpmod.load_kp(a.project, refresh=False)
    ctx = build_region_context(
        kp, [SimpleNamespace(host=d, city=row.city, country=row.country)
             for d, row in kp.items()])

    # Один сайт качаем ОДИН раз, даже если на нём несколько городов: СНГ-страны
    # делят сайт (stalmetural.kz/.by/.uz), не бьём сервер повторами.
    _seen_dom = set()
    _fetch_list = []
    for _d, _r in domains:
        if _d not in _seen_dom:
            _seen_dom.add(_d)
            _fetch_list.append((_d, _r))
    _stamp(f'ПРОВЕРКА КП (1.4) - {PROJECT_NAMES[a.project]} - '
           f'городов: {len(domains)} (уникальных сайтов: {len(_fetch_list)})')

    html_map = fetch_all(_fetch_list, proxy, _stamp)
    _n407 = sum(1 for h, e in html_map.values() if '407' in (e or ''))
    if _n407 and _n407 == len(html_map):
        _stamp('⚠️ ВСЕ страницы вернули 407 Proxy Authentication Required - '
               'прокси отклонил авторизацию. Проверь логин:пароль в секрете '
               'proxy_url (формат http://логин:пароль@хост:порт).')
    _stamp('Загрузка завершена, сверяю с КП …')
    # Диагностика: какие значения КП реально сравниваем (видно, дошли ли правки
    # из Google в нужную строку - частая путаница «поменял в таблице, а тут старое»).
    if len(domains) <= 15:
        for dom, row in domains:
            _stamp(f'  КП[{dom}]: почта={row.email!r}, тел.поиск='
                   f'{row.phone_seo!r}, тел.общий={row.phone_common!r}, '
                   f'город={row.city!r}, адрес={(row.address or "")[:30]!r}')
    результаты = []
    _n_fail = 0
    for dom, row in domains:
        html, err = html_map.get(dom, ("", "не загружено"))
        # ЛЮБАЯ ошибка загрузки (HTTP 500 / обрыв соединения / таймаут / неполная
        # страница) = сайт не отдал страницу целиком. НЕ сверяем «что успели» -
        # это давало мешанину (шапка ✓, адрес ⚠), хотя по факту сайт упал.
        # Помечаем весь город ✗ с причиной (по просьбе заказчика).
        if err:
            _le = err.lower()
            if ('name or service not known' in _le or 'errno -2' in _le
                    or 'getaddrinfo' in _le or 'nodename nor servname' in _le):
                err = ('домен не существует (DNS не находит) - проверьте адрес '
                       'в КП, обычно опечатка')
            _n_fail += 1
            результаты.append({"domain": dom, "city": row.city,
                               "country": row.country, "error": err, "fields": []})
            continue
        var = kpmod.check_variables(html, dom, row=row)
        # Адрес не нашли в шапке/подвале главной? У части проектов (МПЭ/mepen) он
        # только на «Контактах». Догружаем эту страницу и пересверяем адрес.
        # ВАЖНО: только если ГЛАВНАЯ загрузилась чисто (not err). Если главная
        # упала (500/обрыв), «Контакты» тоже не откроются - незачем их дёргать:
        # иначе на сотне не загрузившихся доменов фаза «сверяю с КП» висела
        # минутами на ретраях. Один лёгкий повтор - главная-то грузится нормально.
        _af = next((f for f in var["fields"] if f.get("field") == "Адрес"), None)
        if html and not err and _af and _af.get("status") == "warn":
            _cpath = _find_contacts_path(html, dom)
            if _cpath:
                _ch, _cerr = '', 'not tried'
                for _try in range(2):
                    _ch, _cerr = _fetch_one(dom, _proxy_parts(proxy), path=_cpath)
                    if _ch and not _cerr:
                        break
                    time.sleep(1.0)
                if _ch and not _cerr:
                    var = kpmod.check_variables(html, dom, contacts_html=_ch, row=row)
                    _stamp(f'    {dom}: адрес не в подвале - догрузил «Контакты» '
                           f'({_cpath})')
        # WhatsApp: кнопка есть, но ссылка не ведёт в WhatsApp - проверяем ссылку
        # вживую и уточняем код (404 и т.п.), чтобы в примечании был точный ответ.
        for f in var.get("fields", []):
            if f.get("field") == "WhatsApp" and f.get("check_url"):
                код = _проверить_ссылку(f["check_url"], dom, _proxy_parts(proxy))
                if код == 'stub':
                    f["found"] = "ссылка-заглушка (#/пусто)"
                    f["note"] = ("кнопка «Чат в WhatsApp» есть, но ссылка пустая "
                                 "(#/javascript) - при переходе ничего не откроется")
                elif isinstance(код, int) and код >= 400:
                    f["found"] = f"при переходе ошибка {код}"
                    f["note"] = (f"кнопка «Чат в WhatsApp» есть, но при переходе "
                                 f"ошибка {код} (ссылка не ведёт в WhatsApp)")
        город, _страна = _регион_статусы(html, kpmod._norm_host(dom), ctx)
        var["fields"] = [город] + var["fields"]   # «Страна» убрана из отчёта
        # Переводная копия сайта: проверяем только почту, остальное «–».
        var["fields"] = _только_почта_для_перевода(row.city, var["fields"])
        var["error"] = ""
        результаты.append(var)

    if _n_fail:
        _stamp(f'⚠️ Не загрузилось (сайт отдал 500 / оборвал соединение / таймаут): '
               f'{_n_fail} из {len(domains)}. В отчёте они помечены ✗ по всей строке '
               'с причиной - это НЕ ошибки КП, а недоступность сайта. '
               'Перезапусти позже или проверь, открывается ли сайт в браузере.')
    work = WORK_ROOT / a.project
    work.mkdir(parents=True, exist_ok=True)
    xlsx = work / 'variables.xlsx'
    _записать_xlsx(xlsx, PROJECT_NAMES[a.project], результаты)
    _stamp(f'Отчёт сохранён: {xlsx}')
    # Telegram: отчёт КП получателям проекта (креды - в окружении, их проставляет
    # страница из секретов). Подпись унифицирована с формами и целями. Без
    # настроенного TG - тихо пропуск.
    try:
        import telegram_notify as tn
        from telegram_notify import escape_html
        import datetime as _dt
        _бренд = PROJECT_NAMES[a.project].split(' - ')[0].strip()
        _дата = _dt.datetime.now(_dt.timezone(_dt.timedelta(hours=5))).strftime('%d.%m.%Y')
        _города_список = [c.strip() for c in (a.cities or '').split(',') if c.strip()]
        _части = [f'<b>Проверка КП {escape_html(_бренд)}</b>']
        if _города_список:
            _части.append(f'Города: {escape_html(", ".join(_города_список))}')
        else:
            _части.append('Проверены все домены и поддомены проекта')
        _части.append('📎 Полный отчёт - в прикреплённом xlsx-файле')
        _текст = '\n\n'.join(_части)
        _res = tn.send_report_from_env(
            project_name=PROJECT_NAMES[a.project], summary_text=_текст,
            report_file=xlsx if xlsx.is_file() else None,
            report_filename=f'KP-{a.project}-{_дата}.xlsx',
            log=lambda lvl, msg: _stamp(msg))
        if not _res.get('skipped'):
            _stamp(f'✓ Telegram: отправлено {_res.get("sent", 0)}, '
                   f'не доставлено {_res.get("failed", 0)}')
    except Exception as e:  # noqa: BLE001
        _stamp(f'⚠ Telegram-отправка не удалась ({e}) - отчёт всё равно готов.')
    _stamp('✅ ВСЁ ГОТОВО')
    return 0


_ЛЕГЕНДА = [
    ("Как читать результат", True),
    ("", False),
    ("✓  – значение на сайте совпадает с КП "
     "(для телефона: номер входит в набор номеров города из КП).", False),
    ("✗  – расхождение. В примечании ячейки: «КП / На сайте». "
     "Для адреса ✗ – на сайте найден ДРУГОЙ адрес (не совпадает с КП).", False),
    ("⚠  – на сайте не найдено (телефон / почта / адрес / мессенджер).", False),
    ("И ✗, и ⚠ также собраны списком на листе «Расхождения».", False),
    ("–  – в КП этого поля нет (проверять не с чем).", False),
    ("✗ по ВСЕЙ строке – сайт этого города не загрузился (HTTP 500 / обрыв / "
     "таймаут). В примечании ячейки – причина. Это НЕ ошибка КП, а недоступность "
     "сайта: перезапусти позже или проверь, открывается ли сайт в браузере.", False),
]


def _написать_легенду(ws) -> None:
    from openpyxl.styles import Font, Alignment
    ws.column_dimensions["A"].width = 100
    for i, (text, bold) in enumerate(_ЛЕГЕНДА, 1):
        cell = ws.cell(i, 1, text)
        cell.font = Font(bold=bold, size=14 if (bold and i == 1) else 11)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _записать_xlsx(path: Path, proj_name: str, результаты: list) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    # Лист-подсказка «Как читать результат» - первым (перед данными). Просьба заказчика.
    _написать_легенду(wb.active)
    wb.active.title = "Как читать результат"
    ws = wb.create_sheet("Проверка КП")
    hdr_fill = PatternFill("solid", fgColor="EEF3FB")
    # Порядок по просьбе заказчика: Страна(КП), затем Город(КП) со ссылкой на
    # домен/поддомен, дальше проверяемые переменные.
    headers = ["Страна(КП)", "Город(КП)"] + VAR_COLUMNS
    for c, t in enumerate(headers, 1):
        cell = ws.cell(1, c, t)
        cell.font = Font(bold=True)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "C2"

    from openpyxl.comments import Comment
    LINK_FONT = Font(color="1155CC", underline="single")
    # Заливка только у проблемных ячеек, чтобы зелёные ✓ оставались чистыми
    # (без «тревожного» красного уголка-примечания на каждой ячейке).
    BUG_FILL = PatternFill("solid", fgColor="FDE3E3")   # мягкий красный
    WARN_FILL = PatternFill("solid", fgColor="FFF2DA")  # мягкий оранжевый
    _FIRST_VAR_COL = 3   # A=Страна(КП), B=Город(КП), переменные с C

    расхождения = []
    r = 2
    for res in результаты:
        ws.cell(r, 1, res.get("country", ""))
        # Город(КП) - текст города, кликом ведёт на домен/поддомен.
        gcell = ws.cell(r, 2, res.get("city", "") or res["domain"])
        gcell.hyperlink = f'https://{res["domain"]}'
        gcell.font = LINK_FONT
        by = {f["field"]: f for f in res.get("fields", [])}
        if res.get("error"):
            # Сайт не загрузился - ✗ по ВСЕМ колонкам + причина в примечании
            # каждой ячейки (не путаем с «телефон ✓, адрес ⚠» у частично
            # загруженной страницы: тут упал весь сайт).
            reason = f"Сайт не загрузился: {res['error']}"
            for c in range(_FIRST_VAR_COL, _FIRST_VAR_COL + len(VAR_COLUMNS)):
                cell = ws.cell(r, c, "✗")
                cell.font = Font(color=_COLOR.get("bug", "C62828"), bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = BUG_FILL
                cm = Comment(reason, "1.4")
                cm.width, cm.height = 340, 90
                cell.comment = cm
            r += 1
            continue
        for c, name in enumerate(VAR_COLUMNS, _FIRST_VAR_COL):
            f = by.get(name)
            if not f:
                ws.cell(r, c, "–")
                continue
            status = f["status"]
            cell = ws.cell(r, c, _SYMBOL.get(status, "?"))
            cell.font = Font(color=_COLOR.get(status, "000000"), bold=True)
            cell.alignment = Alignment(horizontal="center")
            # Примечание + заливку вешаем ТОЛЬКО на проблемные ячейки (✗ и ⚠) -
            # тогда зелёные ✓ чистые, а красные сразу видно (красная заливка +
            # уголок-примечание с деталями). Просьба заказчика.
            if status in ("bug", "warn"):
                # Примечание - ВВЕРХУ (сразу видно, в чём дело), ниже - что ждали
                # и что по факту на сайте.
                note = (f.get("note") or "").strip()
                подпись = f"Примечание: {note}\n\n" if note else ""
                подпись += (f"КП: {f['expected']}\n"
                            f"На сайте: {f['found']}")
                # Длинное не расписываем в ячейке - отсылаем на лист «Расхождения».
                if len(подпись) > 220:
                    подпись = подпись[:210].rstrip() + "…\n→ см. лист «Расхождения»"
                cm = Comment(подпись, "1.4")
                cm.width, cm.height = 340, 170   # чтобы текст влезал в окошко
                cell.comment = cm
                cell.fill = BUG_FILL if status == "bug" else WARN_FILL
            # На лист «Расхождения» выводим И красные (✗ bug), И жёлтые (⚠ warn) -
            # по просьбе заказчика: жёлтые «проверьте» тоже нужно видеть списком,
            # а не только уголком-примечанием в ячейке.
            if status in ("bug", "warn"):
                расхождения.append((res["domain"], res.get("city", ""), name,
                                    f["expected"], f["found"], f.get("note", "")))
        r += 1

    ws.column_dimensions["A"].width = 16   # Страна(КП)
    ws.column_dimensions["B"].width = 26   # Город(КП) со ссылкой

    # Лист «Расхождения» - только проблемные ячейки, для быстрого разбора.
    ws2 = wb.create_sheet("Расхождения")
    for c, t in enumerate(["Поддомен", "Город", "Переменная", "КП",
                           "На сайте", "Примечание"], 1):
        cell = ws2.cell(1, c, t)
        cell.font = Font(bold=True)
        cell.fill = hdr_fill
    for i, row in enumerate(расхождения, 2):
        for c, v in enumerate(row, 1):
            cell = ws2.cell(i, c, v)
            if c in (4, 5, 6):     # «Ожидалось», «На сайте», «Примечание» - переносим
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in (("A", 32), ("B", 16), ("C", 14), ("D", 34), ("E", 34), ("F", 70)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"
    if not расхождения:
        ws2.cell(2, 1, "Расхождений не найдено 🎉")

    wb.save(path)


if __name__ == '__main__':
    raise SystemExit(main())
