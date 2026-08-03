"""
convert_catalog.py - конвертация xlsx-каталога проекта в компактные
catalogs/{proj}-catalog.csv (+ отдельные каталоги стран) и сборка
catalogs/{proj}-subdomains.csv из уже готовой КП.

Запуск:
    python convert_catalog.py mpi /путь/к/Каталог_МПИ.xlsx

Зачем отдельный модуль (а не convert_catalogs.py): тот скрипт - разовая
миграция из старого Node.js-проекта, с жёстко прошитыми путями чужой машины
и запуском прямо на импорте. Здесь - CLI по образцу convert_kp.py: заказчик
присылает обновлённый каталог, мы прогоняем команду и коммитим CSV.

Поддомены берём НЕ из xlsx, а из catalogs/{proj}-kp.csv: КП - эталон (она сама
обновляется из Google-таблицы перед каждым прогоном), и так список доменов
гарантированно совпадает с тем, по которому идут остальные проверки.
"""
import argparse
import csv
import re
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

from kp import CATALOGS_DIR

# Раскладка каталога по проектам: какой лист xlsx в какой CSV кладём.
# Ключ sheets - {имя листа: имя файла в catalogs/}. Несколько листов = у части
# доменов свой каталог (см. host_catalogs в projects/<proj>.json).
CATALOG_LAYOUT = {
    # МПИ (МетПромИнтекс). Каталог плоский: у всех разделов путь /catalog/<slug>
    # без вложенности, уровни 1-5 живут в отдельных колонках и в URL не видны.
    # Листы «РФ» и «СНГ» расходятся (~30 слагов из 1259), поэтому СНГ-домены
    # ходят по своему каталогу - иначе чек-лист ловил бы ложные 404.
    'mpi': {
        'sheets': {
            'РФ':  'mpi-catalog.csv',
            'СНГ': 'mpi-cat-cis.csv',
        },
        # Каким листом дополняем лонгтейл из sitemap (sitemap у нас один -
        # главного домена) и какой префикс считаем каталогом.
        'sitemap_sheet': 'РФ',
        'catalog_prefix': '/catalog/',
        # Раздел услуг (резка, гибка, цинкование - 45 страниц). В структурной
        # выгрузке его нет, и до сих пор он не проверялся ничем. Кладём типом
        # «категория»: это такие же разводящие посадочные, и проверки категории
        # (крошки, H1, подзаголовки, SEO-текст) к ним подходят один в один.
        # В тех. страницы не идут намеренно - те проверяются ВСЕ и на каждом
        # прогоне, а 45 страниц раздела правильнее просматривать выборкой.
        'extra_category_prefix': '/services/',
    },
}


def _url_column(ws) -> int:
    """Индекс колонки со ссылкой на страницу. Ищем по заголовку («url-адрес
    страницы»), а не по позиции: в присылаемых таблицах слева могут добавиться
    служебные колонки. Шапку ищем в первых 10 строках - выше неё бывают
    объединённые заголовки блоков («Теги», «Частные данные проекта»)."""
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for i, cell in enumerate(row):
            if cell and 'url' in str(cell).lower():
                return i
    raise RuntimeError('в листе не найдена колонка со ссылкой («url-адрес страницы»)')


def _convert_sheet(ws, dst: Path) -> int:
    """Лист каталога → CSV (url, type). Строки без http пропускаем - так
    отсеиваются шапка, строка счётчиков и «ссылка на правила» под ней.

    Тип строк из xlsx - «категория»: это структурная выгрузка (уровни 1-5).
    Лонгтейл-страницы в неё не попадают и добираются из sitemap - см.
    _sitemap_tags."""
    col = _url_column(ws)
    rows, seen = [], set()
    for row in ws.iter_rows(values_only=True):
        url = row[col] if len(row) > col else None
        if not url or not str(url).strip().lower().startswith('http'):
            continue
        u = str(url).strip()
        if u in seen:
            continue
        seen.add(u)
        rows.append([u, 'категория'])
    return rows


def _write_catalog(rows: list, dst: Path) -> None:
    with open(dst, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerow(['url', 'type'])
        w.writerows(rows)


def _sitemap_urls(sitemap_path: str) -> list[str]:
    """Все <loc> из сохранённого sitemap.xml. Файл берём с диска - на сайте
    может стоять антибот, который отдаёт нам challenge вместо XML."""
    text = Path(sitemap_path).read_text(encoding='utf-8', errors='replace')
    return [u.strip() for u in re.findall(r'<loc>\s*([^<]+?)\s*</loc>', text)]


def _sitemap_by_prefix(sitemap_path: str, prefix: str, type_name: str,
                       known: set) -> list:
    """Строки [url, type] из sitemap по префиксу раздела, минус уже известные."""
    rows, seen = [], set()
    for u in _sitemap_urls(sitemap_path):
        path = urlparse(u).path
        if not path.startswith(prefix) or path in known or u in seen:
            continue
        seen.add(u)
        rows.append([u, type_name])
    return rows


def _sitemap_tags(sitemap_path: str, known: set, prefix: str) -> list:
    """Страницы каталога из sitemap, которых НЕТ в структурной выгрузке.

    Это лонгтейл-посадочные («/catalog/abraziv-r40», «/catalog/32-mm-truba-
    dlya-kanalizatsii»): в xlsx их не отдают, а в индексе они и составляют
    основную массу. Помечаем типом «тег» - ровно как у ИМП, где таких страниц
    14555 против 1454 категорий. Без этого чек-лист видел бы пятую часть сайта.

    known - множество УЖЕ известных путей (категории из xlsx), prefix -
    раздел каталога ('/catalog/')."""
    return _sitemap_by_prefix(sitemap_path, prefix, 'тег', known)


def convert(project_id: str, xlsx_path: str,
            sitemap_path: str = '') -> list[Path]:
    """Каталог проекта из xlsx (+ необязательный sitemap) → CSV в catalogs/."""
    if project_id not in CATALOG_LAYOUT:
        raise SystemExit(f'Нет раскладки каталога для проекта «{project_id}». '
                         f'Известные: {", ".join(CATALOG_LAYOUT)}')
    layout = CATALOG_LAYOUT[project_id]
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    # Sitemap у нас один - главного домена, поэтому дополняем им только
    # основной каталог. Каталог СНГ живёт на других хостах, и его лонгтейл
    # добавится, когда придёт sitemap СНГ-домена.
    main_sheet = layout.get('sitemap_sheet')

    out = []
    for sheet, filename in layout['sheets'].items():
        if sheet not in wb.sheetnames:
            raise SystemExit(f'в таблице нет листа «{sheet}» '
                             f'(есть: {", ".join(wb.sheetnames)})')
        dst = CATALOGS_DIR / filename
        rows = _convert_sheet(wb[sheet], dst)
        n_cat, n_tag, n_svc = len(rows), 0, 0
        if sitemap_path and sheet == main_sheet:
            known = {urlparse(u).path for u, _ in rows}
            tags = _sitemap_tags(sitemap_path, known, layout['catalog_prefix'])
            rows += tags
            n_tag = len(tags)
            svc_prefix = layout.get('extra_category_prefix')
            if svc_prefix:
                svc = _sitemap_by_prefix(sitemap_path, svc_prefix,
                                         'категория', known)
                rows += svc
                n_svc = len(svc)
        _write_catalog(rows, dst)
        tail = ''.join((f', тегов из sitemap: {n_tag}' if n_tag else '',
                        f', услуг: {n_svc}' if n_svc else ''))
        print(f'{project_id}: лист «{sheet}» → {dst.name}, '
              f'категорий: {n_cat}{tail}')
        out.append(dst)
    return out


def build_subdomains(project_id: str) -> Path:
    """catalogs/{proj}-subdomains.csv из catalogs/{proj}-kp.csv.

    В КП домен лежит без схемы (metpromintex.ru, spb.metpromintex.ru) - здесь
    приводим к виду, который ждёт sources.parse_subdomains: https://<хост>/."""
    kp_path = CATALOGS_DIR / f'{project_id}-kp.csv'
    if not kp_path.is_file():
        raise SystemExit(f'нет базы КП {kp_path} - сначала прогоните convert_kp.py')

    rows, seen = [], set()
    with open(kp_path, encoding='utf-8') as f:
        for r in csv.DictReader(f):
            host = (r.get('domain') or '').strip().strip('/').lower()
            if not host or host in seen:
                continue
            seen.add(host)
            rows.append([f'https://{host}/', (r.get('city') or '').strip(),
                         (r.get('country') or '').strip()])

    dst = CATALOGS_DIR / f'{project_id}-subdomains.csv'
    with open(dst, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerow(['url', 'city', 'country'])
        w.writerows(rows)
    print(f'{project_id}: КП → {dst.name}, доменов: {len(rows)}')
    return dst


def main():
    ap = argparse.ArgumentParser(
        description='Каталог проекта из xlsx (+ sitemap) в catalogs/*.csv')
    ap.add_argument('project', choices=sorted(CATALOG_LAYOUT))
    ap.add_argument('xlsx', help='структурная выгрузка каталога')
    ap.add_argument('--sitemap', default='',
                    help='сохранённый sitemap.xml - добавит лонгтейл-страницы '
                         'каталога, которых нет в структурной выгрузке')
    a = ap.parse_args()
    convert(a.project, a.xlsx, a.sitemap)
    build_subdomains(a.project)


if __name__ == '__main__':
    main()
