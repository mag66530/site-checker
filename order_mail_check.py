"""
order_mail_check.py - проверка, что письмо об оформлении заказа дошло покупателю.

Пункт чек-листа 2.9: «Письмо с оформлением заказа приходит пользователю».
При оформлении заказа через корзину движок указывает почту покупателя ПОЧТА
(тестовый ящик, по умолчанию test111@yandex.ru). Bitrix присылает на неё письмо-
подтверждение заказа. Здесь мы заходим в этот ящик по IMAP и проверяем, что
такое письмо реально пришло вскоре после нашего тест-заказа.

Тот же модуль годится и для пункта 2.10 (письма менеджерам/руководителю проекта):
логика та же - меняются только креды ящика (ORDER_MAIL_EMAIL/ORDER_MAIL_PASSWORD)
и, при желании, маркеры темы письма.

IMAP-клиент переиспользуем из metrika_404 (там уже есть подключение через
HTTP-CONNECT-прокси и кодек IMAP UTF-7 для русских папок).

Креды и прокси берём из окружения (их передаёт страница «Проверка форм»):
  ORDER_MAIL_EMAIL     - ящик покупателя (куда падает подтверждение заказа)
  ORDER_MAIL_PASSWORD  - пароль приложения этого ящика
  ORDER_MAIL_FOLDER    - папка (по умолчанию INBOX)
  proxy_url            - прокси (нужен, если IMAP с датацентр-IP блокируется)
Если кред нет - проверка тихо пропускается (как и проверка админки).
"""
from __future__ import annotations

import json
import os
import re
from datetime import datetime, timedelta
from email.header import decode_header
from email.utils import parseaddr, parsedate_to_datetime
from pathlib import Path
from urllib.parse import urlparse


# IMAP-сервер Яндекса (тестовые ящики @yandex.ru). Для других провайдеров
# хост можно передать через ORDER_MAIL_IMAP_HOST.
DEFAULT_IMAP_HOST = "imap.yandex.ru"
DEFAULT_IMAP_PORT = 993

# Слова-маркеры письма-подтверждения заказа (тема или отправитель). Совпадение
# нестрогое (регистр/ё игнорируем). Список намеренно широкий - у разных CMS
# формулировки разные («Ваш заказ принят», «Заказ №… оформлен», «Спасибо за заказ»).
МАРКЕРЫ_ЗАКАЗА = [
    "заказ", "оформлен", "принят", "спасибо за", "order", "покупк",
]

# Окно поиска письма относительно момента оформления заказа: письмо приходит
# ПОСЛЕ заказа, но часы клиента и сервера могут расходиться, а доставка занимать
# минуты - поэтому берём небольшой запас назад и щедрый вперёд.
ОКНО_НАЗАД = timedelta(minutes=5)
ОКНО_ВПЕРЁД = timedelta(minutes=45)

_МЕСЯЦЫ = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


# ── Чистые функции (легко тестируются без сети) ──────────────────────


def _norm(s: str) -> str:
    """Нормализация для нестрогого сравнения: lower + ё→е."""
    return (s or "").lower().replace("ё", "е")


def извлечь_адрес(header_value: str) -> str:
    """Из заголовка From/To достаём голый e-mail (без имени)."""
    return (parseaddr(header_value or "")[1] or "").strip().lower()


def декодировать_заголовок(raw) -> str:
    """MIME-заголовок (=?utf-8?B?…?=) → обычная строка. Терпимо к мусору."""
    if raw is None:
        return ""
    if isinstance(raw, bytes):
        try:
            raw = raw.decode("utf-8", errors="replace")
        except Exception:
            raw = raw.decode("latin-1", errors="replace")
    parts = []
    try:
        for txt, enc in decode_header(raw):
            if isinstance(txt, bytes):
                parts.append(txt.decode(enc or "utf-8", errors="replace"))
            else:
                parts.append(txt)
    except Exception:
        return str(raw)
    return "".join(parts).strip()


def дата_имап(dt: datetime) -> str:
    """Дата для IMAP-команды SEARCH SINCE в формате DD-Mon-YYYY с АНГЛИЙСКИМ
    месяцем (не зависит от локали машины). Берём на день раньше - у SINCE
    гранулярность в сутки, а мы хотим точно захватить день заказа."""
    d = dt - timedelta(days=1)
    return f"{d.day:02d}-{_МЕСЯЦЫ[d.month - 1]}-{d.year}"


def разобрать_дату_письма(date_header: str):
    """Заголовок Date письма → naive datetime в локальной зоне (для сравнения с
    моментом заказа, который тоже пишется как datetime.now()). None, если не разобрать."""
    if not date_header:
        return None
    try:
        dt = parsedate_to_datetime(date_header)
    except Exception:
        return None
    if dt is None:
        return None
    if dt.tzinfo is not None:
        # приводим к локальному времени и снимаем tz, чтобы сравнивать с naive
        dt = dt.astimezone().replace(tzinfo=None)
    return dt


def похоже_на_письмо_заказа(subject: str, sender: str, markers=None) -> bool:
    """True, если тема или отправитель содержат маркер заказа."""
    markers = markers or МАРКЕРЫ_ЗАКАЗА
    hay = _norm(subject) + " " + _norm(sender)
    return any(_norm(m) in hay for m in markers)


def _хост_бренд(домен: str) -> str:
    """https://stalmetural.uz/... → 'stalmetural' (второй уровень без зоны) -
    мягкий признак принадлежности письма нужному проекту."""
    host = re.sub(r"^https?://", "", (домен or "").strip()).split("/")[0]
    host = host.split(":")[0]
    parts = [p for p in host.split(".") if p]
    if len(parts) >= 2:
        return parts[-2].lower()
    return host.lower()


def выбрать_подтверждение(письма: list, момент_заказа: datetime,
                          домен: str = "", markers=None):
    """Из списка писем выбирает то, что похоже на подтверждение НАШЕГО заказа:
    - похоже на письмо заказа (тема/отправитель), и
    - пришло в окне [момент-5мин, момент+45мин].
    Из подходящих предпочитаем письмо, где в теме/отправителе есть бренд домена;
    при равенстве - ближайшее по времени к моменту заказа. None, если нет.

    письма: [{subject, from, date(datetime|None)}, ...]
    """
    бренд = _хост_бренд(домен) if домен else ""
    рано = момент_заказа - ОКНО_НАЗАД
    поздно = момент_заказа + ОКНО_ВПЕРЁД
    кандидаты = []
    for e in письма:
        dt = e.get("date")
        subj = e.get("subject", "")
        frm = e.get("from", "")
        if not похоже_на_письмо_заказа(subj, frm, markers):
            continue
        if dt is not None and not (рано <= dt <= поздно):
            continue
        совпал_бренд = bool(бренд and (бренд in _norm(subj) or бренд in _norm(frm)))
        разница = abs((dt - момент_заказа).total_seconds()) if dt else 1e9
        # ключ сортировки: сперва совпавший бренд, затем ближе по времени
        кандидаты.append(((0 if совпал_бренд else 1, разница), e))
    if not кандидаты:
        return None
    кандидаты.sort(key=lambda p: p[0])
    return кандидаты[0][1]


# ── IMAP (сеть) ─────────────────────────────────────────────────────


def загрузить_креды() -> dict | None:
    """Логин/пароль/папка/прокси ящика покупателя из окружения. None, если
    не заданы e-mail и пароль (тогда проверку пропускаем)."""
    email_addr = (os.environ.get("ORDER_MAIL_EMAIL") or "").strip()
    password = os.environ.get("ORDER_MAIL_PASSWORD") or ""
    # Игнорируем незаполненный шаблон.
    if not email_addr or not password or "ВПИШИ" in email_addr.upper():
        return None
    return {
        "email": email_addr,
        "password": password,
        "folder": (os.environ.get("ORDER_MAIL_FOLDER") or "INBOX").strip() or "INBOX",
        "imap_host": (os.environ.get("ORDER_MAIL_IMAP_HOST") or DEFAULT_IMAP_HOST).strip(),
        "proxy_url": (os.environ.get("proxy_url") or "").strip() or None,
    }


def _подключиться(host: str, port: int, proxy_url, timeout: int = 60):
    """IMAP4_SSL напрямую или через HTTP-CONNECT-прокси (класс из metrika_404)."""
    import imaplib
    import ssl as _ssl

    ssl_ctx = _ssl.create_default_context()
    if proxy_url:
        from metrika_404 import IMAP4_SSL_via_Proxy
        return IMAP4_SSL_via_Proxy(host, port, proxy_url=proxy_url,
                                   ssl_context=ssl_ctx, timeout=timeout)
    return imaplib.IMAP4_SSL(host, port, ssl_context=ssl_ctx, timeout=timeout)


def _имя_папки_imap(folder: str) -> bytes:
    """Имя папки для команды SELECT: русские имена - в IMAP UTF-7."""
    if all(ord(ch) < 128 for ch in folder):
        return folder.encode("ascii")
    from metrika_404 import _imap_utf7_encode
    return _imap_utf7_encode(folder)


def забрать_письма(email_addr: str, password: str, since: datetime,
                   folder: str = "INBOX", imap_host: str = DEFAULT_IMAP_HOST,
                   proxy_url=None, log=print) -> list:
    """Заходит в ящик по IMAP, забирает заголовки писем начиная с даты `since`
    (гранулярность - сутки). Возвращает [{subject, from, to, date}]. Тело письма
    не тянем: подтверждение заказа почти всегда узнаётся по теме/отправителю."""
    import imaplib

    M = _подключиться(imap_host, DEFAULT_IMAP_PORT, proxy_url)
    письма = []
    try:
        try:
            M.login(email_addr, password)
        except imaplib.IMAP4.error as e:
            raise RuntimeError(
                f"не удалось войти в почту {email_addr}: {e}. Проверьте, что задан "
                f"ПАРОЛЬ ПРИЛОЖЕНИЯ и включён IMAP-доступ в настройках ящика."
            ) from e

        M.select(_имя_папки_imap(folder), readonly=True)
        typ, data = M.search(None, "SINCE", дата_имап(since))
        if typ != "OK" or not data or not data[0]:
            return []
        uids = data[0].split()
        # Ограничим разбор разумным числом последних писем (заказных за день - единицы).
        for uid in uids[-200:]:
            try:
                t2, d2 = M.fetch(
                    uid, "(BODY.PEEK[HEADER.FIELDS (SUBJECT FROM TO DATE)])"
                )
            except Exception:  # noqa: BLE001
                continue
            if t2 != "OK" or not d2:
                continue
            raw = b""
            for part in d2:
                if isinstance(part, tuple) and len(part) > 1 and part[1]:
                    raw = part[1]
                    break
            письма.append(_разобрать_заголовки(raw))
    finally:
        try:
            M.logout()
        except Exception:
            pass
    return письма


def _разобрать_заголовки(raw: bytes) -> dict:
    """RFC822-заголовки (bytes) → {subject, from, to, date(datetime|None)}."""
    import email as _email

    msg = _email.message_from_bytes(raw or b"")
    subj = декодировать_заголовок(msg.get("Subject", ""))
    frm = декодировать_заголовок(msg.get("From", ""))
    to = декодировать_заголовок(msg.get("To", ""))
    dt = разобрать_дату_письма(msg.get("Date", ""))
    return {"subject": subj, "from": frm, "to": to, "date": dt}


# ── Оркестратор ─────────────────────────────────────────────────────


def _parse_ts(ts: str):
    try:
        return datetime.strptime((ts or "")[:19], "%Y-%m-%dT%H:%M:%S")
    except Exception:
        return None


def проверить_заказы(orders: list, log=print) -> list:
    """По каждому оформленному заказу проверяет, пришло ли покупателю письмо.

    orders: [{город, почта, домен|url, ts, название}, ...] (из placed_orders.json).
    Возвращает [{город, название, статус, письмо, примечание}]. Заходит в каждый
    уникальный ящик (почту покупателя) один раз."""
    creds = загрузить_креды()
    if not creds:
        log("ℹ️ Проверка письма о заказе пропущена: не заданы ящик/пароль покупателя "
            "(введите их на странице проверки форм или в секретах).")
        return []
    if not orders:
        log("ℹ️ Проверка письма о заказе: заказов за прогон не было.")
        return []

    # Все тест-заказы уходят на один тестовый ящик (ПОЧТА), но на всякий случай
    # группируем по адресу покупателя и явно заданному ящику из кредов.
    ящик = creds["email"]
    since = min((_parse_ts(o.get("ts")) or datetime.now()) for o in orders)

    log(f"📧 Проверка писем о заказе: захожу в {ящик}"
        + (f" через прокси" if creds["proxy_url"] else "") + " …")
    try:
        письма = забрать_письма(
            ящик, creds["password"], since,
            folder=creds["folder"], imap_host=creds["imap_host"],
            proxy_url=creds["proxy_url"], log=log,
        )
    except Exception as e:  # noqa: BLE001
        log(f"⚠️ Проверка письма о заказе не выполнена: {e}")
        return []

    log(f"   писем в ящике с {since.strftime('%d.%m.%Y')}: {len(письма)}")
    результаты = []
    есть = нет = 0
    for o in orders:
        момент = _parse_ts(o.get("ts")) or datetime.now()
        домен = o.get("домен") or o.get("url") or ""
        письмо = выбрать_подтверждение(письма, момент, домен=домен)
        база = {"город": o.get("город", ""), "название": o.get("название", ""),
                "почта": o.get("почта", ящик)}
        if письмо:
            есть += 1
            результаты.append({**база, "статус": "Письмо получено", "письмо": письмо,
                               "примечание": ""})
        else:
            нет += 1
            результаты.append({**база, "статус": "Письмо НЕ найдено", "письмо": None,
                               "примечание": "письмо-подтверждение заказа не найдено в ящике "
                                             "покупателя за окно после оформления"})
    log(f"✅ Проверка письма о заказе: получено {есть}, НЕ найдено {нет}.")
    return результаты


def записать_в_логи(excel_path: str, результаты: list) -> None:
    """Дописывает результат в лист «Логи»: колонка «Письмо покупателю» у строк
    оформления заказа (сопоставление по городу + названию сценария), детали - в
    «Комментарий». По образцу admin_check.записать_в_логи."""
    if not результаты:
        return
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_path)
    if "Логи" not in wb.sheetnames:
        return
    ws = wb["Логи"]
    headers = [str(c.value or "").strip() for c in ws[1]]

    def col(name):
        for i, h in enumerate(headers):
            if h.lower() == name.lower():
                return i + 1
        return -1

    i_city, i_name = col("Город"), col("Название")
    i_comment = col("Комментарий")
    if i_city == -1 or i_name == -1:
        return

    i_mail = col("Письмо покупателю")
    if i_mail == -1:
        i_adm = col("Статус в админке")
        i_st = col("Статус")
        i_mail = (i_adm + 1) if i_adm != -1 else ((i_st + 1) if i_st != -1 else len(headers) + 1)
        ws.insert_cols(i_mail)
        hc = ws.cell(1, i_mail, "Письмо покупателю")
        hc.font = Font(bold=True)
        hc.fill = PatternFill("solid", fgColor="EEF3FB")
        ws.column_dimensions[get_column_letter(i_mail)].width = 20
        headers = [str(c.value or "").strip() for c in ws[1]]
        i_comment = col("Комментарий")

    def norm(v):
        return _norm(str(v or "")).strip()

    строки = {}
    for r in range(2, ws.max_row + 1):
        key = (norm(ws.cell(r, i_city).value), norm(ws.cell(r, i_name).value))
        строки.setdefault(key, []).append(r)
    занято = set()

    for res in результаты:
        key = (norm(res.get("город", "")), norm(res.get("название", "")))
        rows = [r for r in строки.get(key, []) if r not in занято]
        if not rows:
            continue
        r = rows[0]
        занято.add(r)
        статус = str(res.get("статус", ""))
        есть = статус.startswith("Письмо получено")
        ручная = статус.startswith("Проверьте вручную")
        if есть:
            текст, цвет = "Получено", "1E8E3E"
        elif ручная:
            текст, цвет = "Проверьте вручную", "B26A00"
        else:
            текст, цвет = "НЕ найдено", "C62828"
        cell = ws.cell(r, i_mail, текст)
        cell.font = Font(color=цвет, bold=True)
        if i_comment != -1:
            п = res.get("письмо") or {}
            if есть:
                деталь = (f"письмо «{(п.get('subject') or '')[:60]}» "
                          f"({(п.get('date').strftime('%d.%m %H:%M') if п.get('date') else '')})")
            else:
                деталь = res.get("примечание", "") or ("письмо не найдено" if not ручная else "")
            if деталь:
                prev = str(ws.cell(r, i_comment).value or "").strip()
                ws.cell(r, i_comment, (prev + "; " if prev else "") + деталь)

    wb.save(excel_path)


def _загрузить_заказы(orders_path: str) -> list:
    p = Path(orders_path)
    if not p.is_file():
        return []
    try:
        return [o for o in (json.loads(p.read_text(encoding="utf-8")) or []) if o]
    except Exception:
        return []


def отметить_ручную_проверку(orders: list, buyer: str, log=print) -> list:
    """Ручной режим (п.2.9, вариант «своя почта»): IMAP не проверяем, а в отчёте
    просто напоминаем, на какую почту должно прийти письмо о заказе, чтобы человек
    проверил его глазами. Возвращает результаты со статусом «Проверьте вручную»."""
    if not orders:
        return []
    log(f"📧 Письмо о заказе: заказ(ы) оформлены на {buyer} - "
        f"проверьте письмо-подтверждение в этом ящике вручную.")
    результаты = []
    for o in orders:
        результаты.append({
            "город": o.get("город", ""),
            "название": o.get("название", ""),
            "почта": o.get("почта", buyer),
            "статус": "Проверьте вручную",
            "письмо": None,
            "примечание": f"письмо о заказе должно прийти на {o.get('почта', buyer)} - "
                          f"проверьте вручную",
        })
    return результаты


def выполнить_проверку(orders_path: str = "placed_orders.json",
                       excel_path: str = "log_forms.xlsx", log=print) -> bool:
    """Читает placed_orders.json и отмечает письма покупателю в Excel.

    Режим определяется по окружению:
      • есть ORDER_MAIL_EMAIL+ORDER_MAIL_PASSWORD → АВТОпроверка по IMAP (заходим
        в ящик и ищем письмо-подтверждение);
      • иначе, но задан ORDER_BUYER_EMAIL → РУЧНОЙ режим (напоминание в отчёте);
      • ничего не задано → тихо пропускаем.
    По образцу admin_check.выполнить_проверку - вызывается из forms_run после прогона."""
    orders = _загрузить_заказы(orders_path)
    creds = загрузить_креды()
    if creds:
        результаты = проверить_заказы(orders, log=log)
    else:
        buyer = (os.environ.get("ORDER_BUYER_EMAIL") or "").strip()
        if buyer and orders:
            результаты = отметить_ручную_проверку(orders, buyer, log=log)
        else:
            if orders:
                log("ℹ️ Проверка письма о заказе пропущена: не задана почта покупателя "
                    "(укажите её на странице проверки форм).")
            результаты = []
    if результаты:
        try:
            записать_в_логи(excel_path, результаты)
        except Exception as e:  # noqa: BLE001
            log(f"⚠️ Не удалось записать результат письма о заказе в Excel: {e}")
    return bool(результаты)
