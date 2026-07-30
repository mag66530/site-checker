# -*- coding: utf-8 -*-
"""Site Checker — эталон нового проекта.xlsx.

Единый шаблон, который заполняет тот, кто добавляет проект. Все примеры —
вымышленные (домен primer.ru), реальные данные наших проектов не раскрываются.
Заполняющий заменяет строки-примеры своими и удаляет пометку «(пример)».
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INK   = "1A1A1A"
ACCENT= "2B3A55"          # тёмно-синяя шапка
EX    = "8A867D"          # серый — строки-примеры
NOTE  = "5B5853"
HFILL = PatternFill("solid", fgColor=ACCENT)
EXFILL= PatternFill("solid", fgColor="F4F3EF")
thin  = Side(style="thin", color="D9D6CE")
BORDER= Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def _style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill = HFILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER

def _sheet(title, headers, widths, examples, note=None):
    ws = wb.create_sheet(title)
    r = 1
    if note:
        ws.cell(row=1, column=1, value=note).font = Font(italic=True, color=NOTE, size=10)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.row_dimensions[1].height = 30
        ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="center")
        r = 2
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    _style_header(ws, len(headers), row=r)
    ws.row_dimensions[r].height = 30
    for ex in examples:
        r += 1
        for c, v in enumerate(ex, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(italic=True, color=EX, size=10)
            cell.fill = EXFILL
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    header_row = 2 if note else 1
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return ws

# ── Инструкция ──────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Инструкция"
ws.column_dimensions["A"].width = 110
lines = [
    ("Site Checker — эталон нового проекта", True, 15),
    ("", False, 11),
    ("Это единый шаблон для добавления проекта в Site Checker. Заполняйте его вместо разрозненных "
     "выгрузок из CMS — тогда данные попадут в сервис без ручной подгонки.", False, 11),
    ("", False, 11),
    ("Как заполнять:", True, 12),
    ("1. На каждом листе первая строка (серая, курсив) — пример. Замените её своими данными и удалите пометку «(пример)».", False, 11),
    ("2. Не переименовывайте столбцы — по их названиям сервис распознаёт данные.", False, 11),
    ("3. Один город/поддомен = одна строка. Заполняйте по всем странам проекта (РФ, РБ, КЗ и т.д.).", False, 11),
    ("4. Пустая ячейка допустима, если данных нет (например, у города нет Telegram).", False, 11),
    ("", False, 11),
    ("Листы книги:", True, 12),
    ("• Паспорт проекта — базовые сведения о сайте (название, домен, sitemap, прокси, фильтры).", False, 11),
    ("• Поддомены — список городских поддоменов проекта.", False, 11),
    ("• Каталог — адреса каталога, категорий и фильтров (тегов).", False, 11),
    ("• КП (Карта присутствия) — контакты по каждому городу: телефоны, почта, адрес, мессенджеры. Главный лист.", False, 11),
    ("• Формы — список форм по страницам (что и где проверять при отправке заявок).", False, 11),
    ("• Цели — цели Яндекс.Метрики по каждому домену/стране.", False, 11),
    ("", False, 11),
    ("Подробное описание — в документе «Добавление проекта» и в разделе «Добавление проекта» гайда Site Checker.", False, 11),
]
for i, (txt, bold, sz) in enumerate(lines, 1):
    c = ws.cell(row=i, column=1, value=txt)
    c.font = Font(bold=bold, size=sz, color=(ACCENT if bold and sz >= 12 else INK))
    c.alignment = Alignment(wrap_text=True, vertical="top")

# ── Паспорт проекта ─────────────────────────────────────────────────────
ws = wb.create_sheet("Паспорт проекта")
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 60
hdr = ["Параметр", "Значение (пример)", "Пояснение"]
for c, h in enumerate(hdr, 1):
    ws.cell(row=1, column=1 + c - 1, value=h)
_style_header(ws, 3)
ws.row_dimensions[1].height = 24
passport = [
    ("Код проекта (латиницей)", "primer", "Короткий код, латиница/цифры. Станет именем файлов проекта."),
    ("Название", "Пример — Примерсталь", "Как показывать в списках сервиса."),
    ("Главный домен", "primer.ru", "Без http:// и без слэша."),
    ("Главная страница", "https://primer.ru/", "Полный адрес главной со слэшем."),
    ("Sitemap", "https://primer.ru/sitemap.xml", "Карта сайта (если есть)."),
    ("Есть фильтры / теги?", "да", "«да», если в каталоге есть страницы фильтров/тегов."),
    ("Блокирует зарубежный IP?", "нет", "«да», если сайт открывается только из РФ — тогда нужен прокси."),
    ("Ресурс Google Search Console", "sc-domain:primer.ru", "Если подключён GSC (можно оставить пустым)."),
]
for i, (a, b, cc) in enumerate(passport, 2):
    ws.cell(row=i, column=1, value=a).font = Font(bold=True, size=10)
    vb = ws.cell(row=i, column=2, value=b); vb.font = Font(italic=True, color=EX, size=10); vb.fill = EXFILL
    ws.cell(row=i, column=3, value=cc).font = Font(color=NOTE, size=10)
    for c in range(1, 4):
        ws.cell(row=i, column=c).border = BORDER
        ws.cell(row=i, column=c).alignment = Alignment(wrap_text=True, vertical="center")
ws.freeze_panes = "A2"

# ── Поддомены ───────────────────────────────────────────────────────────
_sheet(
    "Поддомены",
    ["url", "city", "country"],
    [40, 26, 18],
    [
        ["https://primer.ru/", "Москва (пример)", "Россия"],
        ["https://spb.primer.ru/", "Санкт-Петербург", "Россия"],
        ["https://minsk.primer.by/", "Минск", "Беларусь"],
    ],
    note="Городские поддомены проекта. url — полный адрес со слэшем; city — город; "
         "country — страна (нужна для СНГ-доменов).")

# ── Каталог ─────────────────────────────────────────────────────────────
_sheet(
    "Каталог",
    ["url", "type"],
    [60, 18],
    [
        ["https://primer.ru/catalog/ (пример)", "категория"],
        ["https://primer.ru/catalog/armatura/", "категория"],
        ["https://primer.ru/catalog/armatura/filter/gost-5781/", "тег"],
    ],
    note="Адреса каталога и категорий. type — ровно одно из двух значений: «категория» или «тег» "
         "(тег = страница фильтра; нужны, только если в «Паспорте» указано, что фильтры есть).")

# ── КП ──────────────────────────────────────────────────────────────────
_sheet(
    "КП (Карта присутствия)",
    ["Страна", "Город", "URL (домен города)", "Тел. SEO Город", "Тел. Реклама Город",
     "Тел. Общий Город", "Почта", "Адрес", "Telegram", "WhatsApp"],
    [16, 20, 30, 20, 20, 20, 26, 34, 18, 20],
    [
        ["Россия", "Москва (пример)", "https://primer.ru", "7 (499) 000-00-01", "7 (499) 000-00-02",
         "7 (499) 000-00-03", "msk@primer.ru", "улица Примерная, 1", "primer_manager", "7-900-000-00-03"],
        ["Россия", "Санкт-Петербург", "https://spb.primer.ru", "7 (812) 000-00-01", "7 (812) 000-00-02",
         "7 (812) 000-00-03", "spb@primer.ru", "проспект Невский, 2", "primer_manager_spb", "7-900-000-00-13"],
    ],
    note="Главный лист. Одна строка = один город/поддомен. Телефоны: SEO — для поиска/органики, "
         "Реклама — подменный (коллтрекинг), Общий — основной. Если телефон один — впишите его в «Общий», "
         "остальные оставьте пустыми. Telegram — логин менеджера без @.")

# ── Формы ───────────────────────────────────────────────────────────────
_sheet(
    "Формы",
    ["Страница (тип)", "Название формы", "Где находится", "URL страницы", "Где есть (домены/страны)"],
    [20, 30, 34, 40, 24],
    [
        ["Главная (пример)", "Заказать звонок", "Кнопка в шапке, открывает окно", "https://primer.ru/", "все"],
        ["Главная", "Обратная связь", "Внизу главной, блок с картой и контактами", "https://primer.ru/", "все"],
        ["Товар", "Купить в один клик", "Кнопка у товара, открывает окно", "https://primer.ru/catalog/armatura/tovar-1/", "РФ"],
        ["Оформление заказа", "Оформление заказа через корзину", "Товар в корзину → оформление заказа", "https://primer.ru/personal/cart/", "РФ, РБ"],
    ],
    note="Список форм, которые проверяем при отправке заявок. Для каждой — на какой странице, как называется, "
         "где на странице искать (шапка / подвал / кнопка-модалка) и в каких доменах/странах присутствует.")

# ── Цели ────────────────────────────────────────────────────────────────
_sheet(
    "Цели",
    ["Страна", "Домен", "Счётчик Метрики", "№ цели", "Название", "Тип", "Идентификатор", "Условие"],
    [16, 26, 18, 12, 30, 12, 24, 34],
    [
        ["Россия (пример)", "https://primer.ru", "12345678", "101", "Клик по телефону", "auto", "", "Клик по номеру tel:"],
        ["Россия", "https://primer.ru", "12345678", "102", "Отправка формы «Звонок»", "js", "call-back", "reachGoal('call-back')"],
        ["Беларусь", "https://minsk.primer.by", "87654321", "201", "Клик по WhatsApp", "js", "wa-click", "reachGoal('wa-click')"],
    ],
    note="Цели Яндекс.Метрики по каждому домену/стране. Счётчик — ID счётчика Метрики. "
         "Тип: js (событие reachGoal), auto (клик tel/mailto — Метрика считает сама), url / url_re (переход на страницу), "
         "jivo (чат — проверяется вручную). Идентификатор — id кнопки/события из reachGoal.")

_out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                    "templates", "Site Checker — эталон нового проекта.xlsx")
os.makedirs(os.path.dirname(_out), exist_ok=True)
wb.save(_out)
print("OK saved →", _out)
