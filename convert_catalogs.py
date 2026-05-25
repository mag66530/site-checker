"""
Конвертация xlsx-каталогов в компактный CSV для Streamlit.

Запускается один раз для подготовки данных:
    python convert_catalogs.py

На вход — папка с исходными xlsx (из старого проекта Node.js).
На выход — папка catalogs/ с CSV-файлами по 3 проектам.

Каждый проект → 3 файла:
    {proj}-subdomains.csv  — поддомены (url, city)
    {proj}-catalog.csv     — каталог (url, type)  (только для smu, imp)
    {proj}-categories.csv  — категории (url)      (только для smu)
"""
import csv
import os
from openpyxl import load_workbook

SRC_DIR = '/home/claude/site-checker/catalogs'
DST_DIR = '/home/claude/site-checker-py/catalogs'

os.makedirs(DST_DIR, exist_ok=True)


def first_sheet_matching(wb, *keywords):
    """Найти первый лист, в имени которого есть один из ключевых слов."""
    for name in wb.sheetnames:
        lower = name.lower()
        for kw in keywords:
            if kw.lower() in lower:
                return name
    return None


def convert_subdomains(src_path, dst_path, sheet_name, url_col, city_col):
    """Поддомены: 1 строка = 1 URL поддомена + город."""
    wb = load_workbook(src_path, read_only=True, data_only=True)
    sheet = sheet_name if sheet_name in wb.sheetnames else first_sheet_matching(
        wb, 'поддомен', 'тех', 'карта', 'кп'
    )
    if not sheet:
        raise Exception(f"В {src_path} не найден подходящий лист")

    ws = wb[sheet]
    seen = set()
    rows = []
    for row in ws.iter_rows(values_only=True):
        url = row[url_col] if len(row) > url_col else None
        city = row[city_col] if len(row) > city_col else None
        if not url or not isinstance(url, str) or not url.strip().startswith('http'):
            continue
        u = url.strip()
        if u in seen:
            continue
        seen.add(u)
        rows.append([u, str(city).strip() if city else ''])

    with open(dst_path, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerow(['url', 'city'])
        w.writerows(rows)
    return len(rows)


def convert_catalog_structured(src_path, dst_path, sheet_name):
    """СМУ/ИМП каталог: ищем колонки 'url-адрес страницы' + 'сущность страницы'."""
    wb = load_workbook(src_path, read_only=True, data_only=True)
    sheet = sheet_name if sheet_name in wb.sheetnames else first_sheet_matching(
        wb, 'без дублей', 'каталог'
    )
    if not sheet:
        raise Exception(f"В {src_path} не найден подходящий лист")

    ws = wb[sheet]

    # Ищем строку с заголовками (обычно в первых 10 строках)
    header_row_idx = None
    headers = None
    rows_buffer = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows_buffer.append(row)
        if i >= 10 and header_row_idx is None:
            break
        clean = [str(c).strip() if c else '' for c in row]
        if 'url-адрес страницы' in clean and 'сущность страницы' in clean:
            header_row_idx = i
            headers = clean
            break

    if header_row_idx is None:
        raise Exception("В каталоге нет колонок 'url-адрес страницы' и 'сущность страницы'")

    url_idx = headers.index('url-адрес страницы')
    type_idx = headers.index('сущность страницы')

    # Считываем все строки после header
    rows = []
    seen = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row_idx:
            continue
        url = row[url_idx] if len(row) > url_idx else None
        typ = row[type_idx] if len(row) > type_idx else None
        if not url or not isinstance(url, str) or not url.strip().startswith('http'):
            continue
        if typ not in ('категория', 'тег'):
            continue
        u = url.strip()
        key = (u, typ)
        if key in seen:
            continue
        seen.add(key)
        rows.append([u, typ])

    with open(dst_path, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerow(['url', 'type'])
        w.writerows(rows)
    return len(rows)


def convert_catalog_flat(src_path, dst_path, sheet_name, url_col=0):
    """МПЭ структура: все строки — категории, нет колонки 'сущность'."""
    wb = load_workbook(src_path, read_only=True, data_only=True)
    sheet = sheet_name if sheet_name in wb.sheetnames else first_sheet_matching(
        wb, 'структура', 'каталог'
    )
    if not sheet:
        raise Exception(f"В {src_path} не найден подходящий лист")

    ws = wb[sheet]
    seen = set()
    rows = []
    for row in ws.iter_rows(values_only=True):
        url = row[url_col] if len(row) > url_col else None
        if not url or not isinstance(url, str) or not url.strip().startswith('http'):
            continue
        u = url.strip()
        if u in seen:
            continue
        seen.add(u)
        rows.append([u, 'категория'])

    with open(dst_path, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerow(['url', 'type'])
        w.writerows(rows)
    return len(rows)


def convert_categories(src_path, dst_path):
    """Отдельный файл актуальных категорий (только для СМУ)."""
    wb = load_workbook(src_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    seen = set()
    for row in ws.iter_rows(values_only=True):
        url = row[0]
        if not url or not isinstance(url, str) or not url.strip().startswith('http'):
            continue
        u = url.strip()
        if u in seen:
            continue
        seen.add(u)
        rows.append([u])

    with open(dst_path, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerow(['url'])
        w.writerows(rows)
    return len(rows)


# ── СМУ ─────────────────────────────────────────────────────────────
print('═══ СМУ ═══')
n = convert_subdomains(
    f'{SRC_DIR}/smu-subdomains.xlsx',
    f'{DST_DIR}/smu-subdomains.csv',
    sheet_name='Тех инфо. поддомены',
    url_col=0, city_col=2,
)
print(f'  Поддомены: {n}')

n = convert_catalog_structured(
    f'{SRC_DIR}/smu-catalog.xlsx',
    f'{DST_DIR}/smu-catalog.csv',
    sheet_name='РФ без дублей',
)
print(f'  Каталог (категории+теги): {n}')

n = convert_categories(
    f'{SRC_DIR}/smu-categories.xlsx',
    f'{DST_DIR}/smu-categories.csv',
)
print(f'  Категории (свежий файл): {n}')

# ── ИМП ─────────────────────────────────────────────────────────────
print('\n═══ ИМП ═══')
n = convert_subdomains(
    f'{SRC_DIR}/imp-subdomains.xlsx',
    f'{DST_DIR}/imp-subdomains.csv',
    sheet_name='Карта присутствия',
    url_col=3, city_col=1,
)
print(f'  Поддомены: {n}')

n = convert_catalog_structured(
    f'{SRC_DIR}/imp-catalog.xlsx',
    f'{DST_DIR}/imp-catalog.csv',
    sheet_name='Каталог без дублей',
)
print(f'  Каталог: {n}')

# ── МПЭ ─────────────────────────────────────────────────────────────
print('\n═══ МПЭ ═══')
n = convert_subdomains(
    f'{SRC_DIR}/mpe-subdomains.xlsx',
    f'{DST_DIR}/mpe-subdomains.csv',
    sheet_name='КП',
    url_col=2, city_col=1,
)
print(f'  Поддомены: {n}')

n = convert_catalog_flat(
    f'{SRC_DIR}/mpe-catalog.xlsx',
    f'{DST_DIR}/mpe-catalog.csv',
    sheet_name='структура',
    url_col=0,
)
print(f'  Категории (flat): {n}')

# ── Сводка ──────────────────────────────────────────────────────────
print('\n═══ Размеры файлов ═══')
total = 0
for f in sorted(os.listdir(DST_DIR)):
    size = os.path.getsize(os.path.join(DST_DIR, f))
    total += size
    print(f'  {f}: {size/1024:.0f} КБ')
print(f'\nИтого: {total/1024/1024:.2f} МБ (было 33 МБ в xlsx)')
