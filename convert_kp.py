"""
convert_kp.py - разовая конвертация «Карты присутствия» (xlsx) в компактный
catalogs/{proj}-kp.csv для сверки контактов.

Запуск:
    python convert_kp.py smu /путь/к/КП_СМУ.xlsx
    python convert_kp.py imp /путь/к/КП_ИМП.xlsx
    python convert_kp.py mpe /путь/к/КП_МПЭ.xlsx

В CSV кладём ТОЛЬКО контактные поля (домен, город, телефоны SEO/реклама/общий,
почта, адрес) - исходный xlsx с внутренними данными в репозиторий не идёт.
"""
import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from kp import KP_LAYOUT, CATALOGS_DIR, _norm_host
from kp import split_phones as _split_phones
from kp import phones_in_cell as _phones_in_cell

# Домены, которые НЕ проверяем нигде (по просьбе заказчика) - даже если они есть
# в КП-таблице. Строку с таким доменом пропускаем при сборке CSV.
# inmetprom.am / mepen.am - армянских сайтов НЕТ (домен не резолвится), в КП строка
# «Армения/Ереван» есть, но проверять нечего - иначе город всегда ✗ «сайт не
# загрузился». Если в таблице у Еревана другой домен - добавить его сюда же.
_EXCLUDE_HOSTS = {'steemet.uz', 'inmetprom.am', 'mepen.am'}

# Переподпись города для сайтов-«дублей». У СМУ азербайджанский сайт живёт в двух
# версиях: основной smg.az и переводная копия steelgroup.az. На обоих сайтах город
# подписан одинаково («Баку»), поэтому в «Проверке КП» две строки сливались в
# визуальный дубль (два «Баку» подряд). Переводную версию подписываем как в
# проверке целей (az2 → «Азербайджан (перевод)»), чтобы строки различались.
_CITY_OVERRIDE = {'steelgroup.az': 'Азербайджан (перевод)'}


def _phone_columns(headers):
    """Индексы всех телефонных колонок (Общий/Реклама/SEO/Сотовый/основной/
    подменные/ватсап). Городскую колонку «Город» (название города) исключаем."""
    out = []
    for i, h in enumerate(headers):
        if h is None:
            continue
        ht = str(h).lower().replace('\n', ' ').strip()
        if ht == 'город':
            continue
        # Колонки блока проверки («Контактный телефон/ошибка», «Ватсап/ошибка»…)
        # содержат «ок»/«ошибка», а не номер - в телефоны их не берём (у МПК такой
        # аудит-блок стоит рядом с контактами и раньше давал мусор в all_phones).
        if 'ошибка' in ht:
            continue
        if any(k in ht for k in ('город', 'сотов', 'мобильн', 'основн',
                                 'подменн', 'ватсап', 'для ватсап')):
            out.append(i)
    return out


def _find_header_row(ws, max_scan=6):
    """Найти строку заголовков - где встречаются 'город' и 'адрес'."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        cells = [str(c).lower() if c else '' for c in row]
        joined = ' '.join(cells)
        if 'город' in joined and 'адрес' in joined:
            return i, row
    # запасной вариант - первая строка
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return 1, first


def _col(headers, *keywords, exact=None):
    """Индекс колонки, чей заголовок содержит все keywords (или равен exact).
    Колонки блока проверки («Почта/ошибка», «Адрес/ошибка», …) пропускаем: у МПК
    аудит-столбцы соседствуют с контактными и по слову «почта»/«адрес» перебивали
    настоящую колонку (в почту попадало «ок» вместо e-mail)."""
    for i, h in enumerate(headers):
        if h is None:
            continue
        ht = str(h).lower().replace('\n', ' ').strip()
        if 'ошибка' in ht:
            continue
        if exact is not None and ht == exact:
            return i
        if keywords and all(k in ht for k in keywords):
            return i
    return None


def _sheet_has_kp_header(ws, max_scan=6):
    """Похож ли лист на КП: в первых строках есть и 'город', и 'адрес'."""
    for row in ws.iter_rows(min_row=1, max_row=max_scan, values_only=True):
        joined = ' '.join(str(c).lower() for c in row if c)
        if 'город' in joined and 'адрес' in joined:
            return True
    return False


def _pick_sheet(wb, preferred: str):
    """Выбрать лист КП. Приоритет:
      1) точное имя из KP_LAYOUT;
      2) то же имя без учёта регистра/пробелов (в таблице могли переименовать
         «КП » с пробелом, «кп» строчными и т.п.);
      3) любой лист, похожий на КП (в шапке есть «город» и «адрес»).
    Если ничего не подошло - понятная ошибка со списком листов таблицы,
    чтобы сразу было видно, как называется нужная вкладка."""
    names = wb.sheetnames
    if preferred in names:
        return wb[preferred]
    norm = lambda s: str(s).lower().replace('\n', ' ').strip()
    want = norm(preferred)
    for n in names:
        if norm(n) == want:
            return wb[n]
    for n in names:
        if _sheet_has_kp_header(wb[n]):
            return wb[n]
    raise RuntimeError(
        f'в таблице нет листа «{preferred}» (и ни один лист не похож на КП). '
        f'Листы таблицы: {", ".join(names)}. Переименуйте вкладку с КП '
        f'в «{preferred}» либо укажите её название.')


def convert(project_id: str, xlsx_path: str) -> Path:
    layout = KP_LAYOUT[project_id]
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = _pick_sheet(wb, layout['sheet'])

    hdr_row_idx, headers = _find_header_row(ws)

    ci_city = _col(headers, exact='город') or _col(headers, 'город')
    ci_addr = _col(headers, exact='адрес') or _col(headers, 'адрес')
    # Если у колонок «город»/«страна» ПУСТОЙ заголовок (как у АПС - первые два
    # столбца без шапки) - берём их по позиции из layout (city_col/country_col).
    if ci_city is None and layout.get('city_col') is not None:
        ci_city = layout['city_col']
    # Почта: сначала ОСНОВНАЯ колонка «Почта» (контактный блок), и только потом
    # «E-Mail». У ИМП в таблице ДВЕ почтовые колонки: «Почта» (заполнена у всех
    # городов) и вторичная «E-Mail» (в блоке карт, заполнена лишь у части). Раньше
    # брали «E-Mail» первой - и у ~половины городов почта выходила пустой.
    ci_email = (_col(headers, exact='почта') or _col(headers, 'почта')
                or _col(headers, 'e-mail') or _col(headers, 'email'))
    # ВАЖНО: сначала ТОЧНАЯ колонка «url» - иначе _col(...,'ссылка') цеплял
    # «Ссылка для яндекс-карт» (iframe карты) вместо адреса сайта, и домены
    # городов (особенно поддомены СНГ) не читались - города выпадали из проверки.
    ci_url = (_col(headers, exact='url') or _col(headers, 'url', 'магазин')
              or _col(headers, 'домен') or _col(headers, 'ссылка')
              or _col(headers, 'url'))
    # Если колонку ссылки по заголовку не нашли (у МПИ она без шапки - пустой
    # заголовок после «Численность») - берём её по позиции из layout.
    if ci_url is None and layout.get('url_col') is not None:
        ci_url = layout['url_col']
    ci_seo = _col(headers, *layout['phone_seo'])
    ci_ad = _col(headers, *layout['phone_ad'])
    ci_common = _col(headers, *layout['phone_common'])
    phone_cols = _phone_columns(headers)
    # Доп. переменные (пункт 1.4): страна, Telegram, WhatsApp.
    ci_country = _col(headers, exact='страна') or _col(headers, 'страна')
    if ci_country is None and layout.get('country_col') is not None:
        ci_country = layout['country_col']
    # Telegram/WhatsApp: читаем ПЕРВУЮ одноимённую колонку (её и редактируют в
    # КП). Нераспознанное значение (напр. «2») дальше отфильтрует check_variables
    # как ошибку КП - так же, как у телефонов.
    ci_tg = _col(headers, 'telegram') or _col(headers, 'телеграм')
    ci_wa = (_col(headers, 'whatsapp') or _col(headers, 'ватсап')
             or _col(headers, 'вацап') or _col(headers, 'ватсапп'))
    # Колонка «приоритет ресурса» (у МПК): строк на город несколько - свой сайт
    # и маркетплейсы (pulscen/all.biz/…). В проверку КП берём ТОЛЬКО свой сайт -
    # строки, где приоритет == priority_keep. Для остальных проектов не задано.
    ci_priority = layout.get('priority_col')
    priority_keep = layout.get('priority_keep')
    # per_city (МПК): один сайт на много городов - дедуп по (домен, ГОРОД), а не по
    # одному домену, иначе все города metpromko.ru схлопнулись бы в один.
    per_city = layout.get('per_city')

    def cell(row, idx):
        if idx is None or idx >= len(row):
            return ''
        v = row[idx]
        if v is None:
            return ''
        # openpyxl отдаёт число из ячейки как float: «2» → 2.0, а телефон-числом →
        # 74991234567.0 (а то и «7.5e+10»). Целое показываем без «.0» и без
        # экспоненты - иначе в КП телефон бился, а мусор «2» выводился как «2.0».
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    def norm_country(v):
        """Приводим синонимы страны к единому виду (в КП АПС встречается и «РФ»,
        и «Россия» - это одна страна, иначе в отчёте две)."""
        s = (v or '').strip()
        return 'Россия' if s.lower() in ('рф', 'россия', 'russia') else s

    def clean_msgr(v):
        """Мусорные значения мессенджеров в КП (#N/A, «нет», «подтвердить») → пусто."""
        v = (v or '').strip()
        return '' if v.lower() in ('нет', '-', '#n/a', 'подтвердить',
                                   'подтвердить телефон') else v

    # Предпроход: домены, у которых ЕСТЬ своя ссылка (непустая колонка url) -
    # это города-«владельцы» своего сайта. Нужен, чтобы город без своей ссылки,
    # чей домен уже принадлежит владельцу, не попал в отчёт (см. ниже).
    url_hosts = set()
    for row in ws.iter_rows(min_row=hdr_row_idx + 1, values_only=True):
        if row and any(row):
            h = _norm_host(cell(row, ci_url))
            if h:
                url_hosts.add(h)

    rows_out = []
    seen = set()
    for row in ws.iter_rows(min_row=hdr_row_idx + 1, values_only=True):
        if not row or not any(row):
            continue
        # Фильтр по приоритету ресурса (МПК): оставляем только строки своего сайта
        # (priority == priority_keep), маркетплейсы/пустой приоритет пропускаем.
        if ci_priority is not None:
            try:
                pv = int(float(cell(row, ci_priority)))
            except (TypeError, ValueError):
                pv = None
            if pv != priority_keep:
                continue
        city = cell(row, ci_city)
        url = cell(row, ci_url)
        # Домен берём из колонки url (это «своя ссылка» города).
        host = _norm_host(url)
        if not host:
            # Своей ссылки нет. Пробуем достать домен сети из строки (для проектов
            # без явной колонки url, напр. АПС). НО: если этот домен уже есть у
            # города-владельца ссылки - значит это город-«спутник» без своего сайта
            # (у СНГ одна ссылка на всю страну: stalmetural.kz/.by/.uz). Такие в
            # отчёт НЕ берём - иначе они сверялись бы с чужим городским сайтом и
            # давали ложные ошибки. В отчёте остаётся один город на ссылку.
            joined = ' '.join(str(c) for c in row if c)
            m = re.search(r'([a-z0-9-]+\.)*(?:inmetprom|stalmetural|mepen|aviastal|smg)\.(?:ru|uz|kz|by|az|kg|am)', joined)
            host = _norm_host(m.group(0)) if m else ''
            if host and host in url_hosts:
                continue                     # спутник владельца ссылки - пропускаем
        if host in _EXCLUDE_HOSTS:          # исключённый домен - не проверяем нигде
            continue
        # Дедуп: обычно по ДОМЕНУ (один город-владелец ссылки на сайт), а для
        # per_city (МПК) - по (ДОМЕН, ГОРОД), чтобы все города общего сайта остались.
        _dkey = (host, city) if per_city else host
        if not host or _dkey in seen:
            continue
        seen.add(_dkey)
        # Сайт-«дубль» (переводная версия) - подписываем страной с пометкой,
        # чтобы в отчёте он не сливался с основным городом.
        city = _CITY_OVERRIDE.get(host, city)
        # Все телефоны города (нормализованные, 10 цифр) из всех тел. колонок -
        # сайт может статически показывать любой из них (Общий/SEO/Сотовый).
        all_norm = []
        for idx in phone_cols:
            for n in _split_phones(cell(row, idx)):
                if n not in all_norm:
                    all_norm.append(n)
        # Назначенные слоты (SEO/Реклама/Общий) - это выверенные телефонные ячейки:
        # разбираем их мягким phones_in_cell, который понимает «голый» местный номер
        # без кода страны («(861) 944-63-99»), тогда как строгий split_phones (для
        # произвольных колонок) его пропускает. Иначе у городов с местным номером
        # all_phones выходил пустым (и ложно срабатывало предупреждение о «съехавших
        # колонках»), хотя номер в КП валиден.
        for idx in (ci_seo, ci_ad, ci_common):
            if idx is None:
                continue
            for n in _phones_in_cell(cell(row, idx)):
                if n not in all_norm:
                    all_norm.append(n)
        rows_out.append({
            'domain': host,
            'city': city,
            'phone_seo': cell(row, ci_seo),
            'phone_ad': cell(row, ci_ad),
            'phone_common': cell(row, ci_common),
            'all_phones': ';'.join(all_norm),
            'email': cell(row, ci_email),
            'address': cell(row, ci_addr),
            'country': norm_country(cell(row, ci_country)),
            'telegram': clean_msgr(cell(row, ci_tg)),
            'whatsapp': clean_msgr(cell(row, ci_wa)),
        })
    wb.close()

    CATALOGS_DIR.mkdir(parents=True, exist_ok=True)
    out = CATALOGS_DIR / f'{project_id}-kp.csv'
    with open(out, 'w', encoding='utf-8', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['domain', 'city', 'phone_seo',
                                          'phone_ad', 'phone_common', 'all_phones',
                                          'email', 'address',
                                          'country', 'telegram', 'whatsapp'])
        w.writeheader()
        w.writerows(rows_out)
    print(f'{project_id}: {len(rows_out)} городов → {out}')
    # маленькая сводка качества
    no_phone = sum(1 for r in rows_out if not (r['phone_seo'] or r['phone_ad'] or r['phone_common']))
    with_addr = sum(1 for r in rows_out if r['address'])
    print(f'  без телефона в КП: {no_phone}, с адресом: {with_addr}')
    # Диагностика «съехавших» колонок: ячейки телефонов НЕ пустые, но НИ ОДНОГО
    # валидного номера не разобралось (напр. везде «2»/«2.0» - в колонку телефона
    # попало что-то не то). Если таких почти вся таблица - структура КП, скорее
    # всего, сломана: громко предупреждаем, чтобы это не выглядело как «сайт не
    # виден» в отчёте.
    _garbage = sum(1 for r in rows_out
                   if (r['phone_seo'] or r['phone_ad'] or r['phone_common'])
                   and not r['all_phones'])
    if rows_out and _garbage >= max(3, len(rows_out) // 2):
        print(f'  ⚠️ ВНИМАНИЕ: у {_garbage} из {len(rows_out)} городов в колонках '
              f'телефона стоит НЕ номер (не разобрался ни один). Похоже, в таблице '
              f'КП съехали/переименовались колонки - проверьте лист '
              f'«{layout["sheet"]}»: телефоны/почта/адрес должны быть в своих '
              f'столбцах, а не, например, «2». Иначе в отчёте всё уйдёт в ✗ '
              f'«в КП не распознан», хотя на сайте данные есть.')
    return out


def main():
    if len(sys.argv) != 3:
        projects = '|'.join(KP_LAYOUT)
        print(f'Использование: python convert_kp.py <{projects}> <путь_к_xlsx>')
        sys.exit(1)
    convert(sys.argv[1], sys.argv[2])


if __name__ == '__main__':
    main()
