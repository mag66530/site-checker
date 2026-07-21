"""Тест Excel-отчёта «Все прогоны» (кнопка «Скачать все прогоны»)."""
import io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl  # noqa: E402

import pagespeed_history as H  # noqa: E402
import pagespeed_report as R  # noqa: E402
from pagespeed_checker import MetricSet, PageResult  # noqa: E402


def _mk(u, tc, ds, ms):
    return PageResult(url=u, type_code=tc,
                      desktop=MetricSet(score=ds), mobile=MetricSet(score=ms))


def test_all_runs_workbook_two_sheets_newest_first(monkeypatch, tmp_path):
    monkeypatch.setattr(H, "DATA_DIR", tmp_path)
    H.append_run("smu", "2026-07-01 10:00:00", [_mk("https://a/", "main", 90, 40)])
    H.append_run("smu", "2026-07-15 10:00:00",
                 [_mk("https://a/", "main", 88, 50), _mk("https://a/p/", "product", 77, 60)])
    runs = H.all_run_aggregates("smu")

    blob = R.all_runs_to_bytes(project_name="СМУ", runs_agg=runs)
    assert blob[:2] == b"PK"                       # валидный xlsx (zip)
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert wb.sheetnames == ["Все прогоны", "По типам"]

    ws = wb["Все прогоны"]
    dates = [r[0] for r in ws.iter_rows(values_only=True)
             if r[0] and str(r[0])[:2].isdigit()]
    assert dates[0].startswith("15.07.2026")       # новые прогоны сверху
    assert dates[-1].startswith("01.07.2026")

    # «По типам»: 2 типа (15.07) + 1 тип (01.07) = 3 строки данных
    assert wb["По типам"].max_row - 1 == 3


def test_all_runs_workbook_empty_history():
    """Пустая история - книга всё равно собирается (без падения)."""
    wb = R.build_all_runs_workbook(project_name="СМУ", runs_agg=[])
    assert wb.sheetnames == ["Все прогоны", "По типам"]
    assert str(wb["Все прогоны"]["A1"].value).startswith("Скорость страниц")
