# -*- coding: utf-8 -*-
"""Лист «Техничка» убран - всё, что было на нём, должно быть в «Проблемах».

Раньше девять детальных листов (Индексация, Метаданные, Ошибки JavaScript,
Валидация и скорость, Страница 404, Дубли главной, Арсенкин, Фильтры ПС,
Нагрузка и парсинг) склеивались в групповой лист. Часть их находок жила
ТОЛЬКО там и при удалении листа потерялась бы - эти тесты держат перенос."""
from report_priorities import (
    Finding, classify, content_sections_findings, interlinking_note,
    robots_hygiene_findings, static_delivery_findings, url_format_findings,
    ux_interactive_findings,
)


class _Рез:
    """Минимальный CheckResult для перелинковки."""

    def __init__(self, int_links, tech_targets=None):
        self.indexing = {'int_links': int_links,
                         'tech_targets': tech_targets or {}}


def _тексты(находки):
    return ' | '.join(f.problem for f in находки)


# ── ЧПУ и формат адресов: строка на КАЖДЫЙ адрес ──────────────────────

def test_чпу_одна_строка_на_адрес():
    summary = {'host': 'a.ru', 'url_format': {
        'checked': 120, 'total_bad': 3,
        'non_sef': ['/catalog/?ID=1', '/index.php'], 'non_sef_n': 2,
        'cyrillic': ['/каталог/'], 'cyrillic_n': 1}}
    находки = url_format_findings(summary)
    assert len(находки) == 3                       # ровно по адресу на строку
    адреса = {f.url for f in находки}
    assert 'https://a.ru/catalog/?ID=1' in адреса
    assert 'https://a.ru/каталог/' in адреса
    assert all(f.level == 'Предупреждение' for f in находки)


def test_чпу_справочные_цифры_идут_в_как_исправить():
    summary = {'host': 'a.ru', 'url_format': {
        'checked': 120, 'total_bad': 1,
        'uppercase': ['/Catalog/'], 'uppercase_n': 1}}
    f = url_format_findings(summary)[0]
    assert 'Проверено путей каталога: 120' in f.fix_note
    assert classify(f)['title']                    # правило нашлось, не дефолт


def test_чпу_обрезанный_список_не_прячет_остаток():
    """Список адресов ограничен потолком - остаток должен быть виден строкой."""
    summary = {'host': 'a.ru', 'url_format': {
        'checked': 900, 'total_bad': 700,
        'underscore': ['/a_b/'], 'underscore_n': 700}}
    находки = url_format_findings(summary)
    assert any('ещё 699' in f.detail for f in находки)


def test_чпу_нет_проблем_нет_находок():
    assert url_format_findings({'url_format': {'checked': 10, 'total_bad': 0}}) == []


# ── Гигиена robots.txt ────────────────────────────────────────────────

def test_группы_user_agent_становятся_предупреждениями():
    summary = {'host': 'a.ru',
               'ua_groups': {'yandex': False, 'google': True, 'star': False}}
    находки = robots_hygiene_findings(summary)
    т = _тексты(находки)
    assert 'Yandex' in т and 'Googlebot' not in т   # Google есть - не жалуемся
    assert 'User-agent: *' in т
    assert all(f.level == 'Предупреждение' for f in находки)


def test_робots_всё_на_месте_молчим():
    assert robots_hygiene_findings(
        {'host': 'a.ru',
         'ua_groups': {'yandex': True, 'google': True, 'star': True}}) == []


# ── Отгрузки и даты статей ────────────────────────────────────────────

def test_отгрузки_без_ссылок_на_каталог():
    находки = content_sections_findings(
        {'host': 'a.ru', 'otgruzki': {'found': '/otgruzki/', 'catalog_links': 0}})
    assert 'Отгрузки' in _тексты(находки)


def test_отгрузки_со_ссылками_не_находка():
    assert content_sections_findings(
        {'host': 'a.ru',
         'otgruzki': {'found': '/otgruzki/', 'catalog_links': 12}}) == []


def test_статья_без_даты_обновления():
    находки = content_sections_findings(
        {'host': 'a.ru', 'news_dates': {'article': 'https://a.ru/news/1/',
                                        'published': True, 'modified': False}})
    assert 'dateModified' in _тексты(находки)
    assert classify(находки[0])['title'] == 'Проставить даты у статей и новостей'


def test_статья_без_даты_публикации():
    находки = content_sections_findings(
        {'host': 'a.ru', 'news_dates': {'article': 'https://a.ru/news/1/',
                                        'published': False, 'modified': False}})
    assert 'datePublished' in _тексты(находки)
    assert len(находки) == 1                        # не обе сразу
    assert classify(находки[0])['title'] == 'Проставить даты у статей и новостей'


# ── Сжатие, кеш и скорость ресурсов (п.1.16-1.17) ─────────────────────

def _w3c(**timings):
    return {'available': True, 'show': {'valid': True, 'static': True},
            'pages': [{'url': 'https://a.ru/', 'timings': timings}]}


def test_сжатие_и_кеш_по_строке_на_страницу():
    check = _w3c(compression={'checked': 8, 'ok': 0, 'missing': ['/app.js']},
                 caching={'checked': 8, 'ok': 3, 'missing': ['/style.css']})
    находки = static_delivery_findings(check)
    т = _тексты(находки)
    assert 'сжатие CSS/JS не включено' in т
    assert 'кеш статики настроен не для всех' in т
    assert all(f.url == 'https://a.ru/' for f in находки)
    assert 'сжато 0 из 8' in ' '.join(f.detail for f in находки)


def test_сжатие_и_кеш_в_порядке_молчим():
    assert static_delivery_findings(
        _w3c(compression={'checked': 8, 'ok': 8},
             caching={'checked': 8, 'ok': 8})) == []


def test_медленные_ресурсы_за_порогом_8_секунд():
    находки = static_delivery_findings(_w3c(
        total_ms=12000, html_ms=400,
        by_type={'js': {'ms': 6000, 'count': 12, 'kb': 900}},
        slowest={'url': 'https://a.ru/app.js', 'ms': 5000, 'kind': 'js'}))
    assert 'дольше 8 секунд' in _тексты(находки)
    f = находки[0]
    assert 'JS 6000мс/12шт/900КБ' in f.detail        # замеры не потеряны
    assert 'app.js' in f.fix_note                    # самый долгий - в подсказке


def test_быстрая_страница_не_находка():
    assert static_delivery_findings(_w3c(total_ms=3000)) == []


# ── Интерактив: слайдер, меню, cookie, модальная форма ────────────────

def test_интерактив_уходит_в_верстку():
    check = {'pages': [{
        'url': 'https://a.ru/',
        'ux': {'slider': 'fail', 'dropdown': 'fail',
               'cookie': {'status': 'not_remembered'}},
        'mobile': {'form_close_m': {'status': 'not_closed', 'name': 'Звонок'}},
    }]}
    находки = ux_interactive_findings(check)
    assert len(находки) == 4
    assert {f.section for f in находки} == {'Вёрстка'}
    т = _тексты(находки)
    for кусок in ('слайдер', 'выпадающее меню', 'cookie-баннер', 'модальная форма'):
        assert кусок in т
    # у каждой находки есть готовая рекомендация, а не дефолт «проверить»
    assert all(classify(f)['title'] for f in находки)


def test_рабочий_интерактив_не_попадает_в_проблемы():
    check = {'pages': [{'url': 'https://a.ru/',
                        'ux': {'slider': 'ok', 'dropdown': 'ok',
                               'cookie': {'status': 'ok', 'days': 30}},
                        'mobile': {'form_close': {'status': 'ok'}}}]}
    assert ux_interactive_findings(check) == []


def test_cookie_короткий_срок_с_числом_дней():
    check = {'pages': [{'url': 'https://a.ru/',
                        'ux': {'cookie': {'status': 'short', 'days': 1}}}]}
    assert '1 дн.' in ux_interactive_findings(check)[0].problem


# ── Перелинковка: блок под таблицей, только при проблеме ──────────────

def test_перелинковка_молчит_когда_каталог_получает_больше():
    assert interlinking_note([_Рез({'home': 5, 'catalog': 40, 'tech': 10,
                                    'other': 2})]) is None


def test_перелинковка_показывается_когда_вес_уходит_в_тех_страницы():
    note = interlinking_note([_Рез({'home': 5, 'catalog': 3, 'tech': 30,
                                    'other': 2}, {'/policy/': 12})])
    assert note and 'тех/инфо' in note['text']
    assert 'на тех/инфо 75%' in note['detail']
    assert '/policy/ (12)' in note['detail']


def test_перелинковка_без_данных():
    assert interlinking_note([]) is None


# ── Колонка «Как исправить» несёт справочные цифры ────────────────────

def test_как_исправить_дописывает_fix_note():
    from openpyxl import Workbook

    import reporter
    wb = Workbook()
    wb.remove(wb.active)
    reporter._build_problems_sheet(wb, [Finding(
        'Предупреждение', 'Индексация', 'кириллица в адресе',
        url='https://a.ru/каталог/', fix_note='Проверено путей каталога: 120.')])
    значения = [str(c.value) for row in wb['Проблемы'].iter_rows()
                for c in row if c.value]
    assert any('Проверено путей каталога: 120.' in v for v in значения)


def test_блок_перелинковки_рисуется_под_таблицей():
    from openpyxl import Workbook

    import reporter
    wb = Workbook()
    wb.remove(wb.active)
    reporter._build_problems_sheet(
        wb, [Finding('Ошибка', 'Индексация', 'что-то', url='https://a.ru/')],
        interlinking={'text': 'вес уходит в тех/инфо', 'detail': 'тех/инфо 75%'})
    ws = wb['Проблемы']
    значения = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    assert any('Перелинковка (внутренний вес)' in v for v in значения)
    assert any('тех/инфо 75%' in v for v in значения)
    # блок идёт ПОСЛЕ таблицы с автофильтром, а не внутрь неё
    последняя_фильтра = int(ws.auto_filter.ref.split(':')[1][1:])
    строка_блока = next(c.row for row in ws.iter_rows() for c in row
                        if c.value and 'Перелинковка' in str(c.value))
    assert строка_блока > последняя_фильтра
