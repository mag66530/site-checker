"""Тест потокобезопасности append_log_row()/record_submitted_form() в
forms_tester/test_all.py.

Формы теперь проверяются ПАРАЛЛЕЛЬНО (до 3 одновременно, run_test() внутри
себя гоняет их через ThreadPoolExecutor) - несколько потоков пишут в ОДИН
log_forms.xlsx / submitted_forms.json. Без блокировки это read-modify-write
на общем файле: конкурентные вызовы теряли бы строки друг друга (кто сохранил
последним - тот и победил). Проверяем реальными потоками, не моками - на
временном файле, что ни одна запись не потерялась."""
import sys
import threading
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "forms_tester"))

import test_all as ta


def test_concurrent_append_log_row_loses_no_rows(tmp_path, monkeypatch):
    path = str(tmp_path / "log_forms.xlsx")
    monkeypatch.chdir(tmp_path)
    ta.init_excel_log(path, очистить=True)

    N = 24     # больше реальных 6 форм - нагрузочный запас
    threads = []

    def _writer(i):
        ta.append_log_row(path, {
            "тип": "PLAYWRIGHT-FORM", "страница": "Тест", "url": f"https://x/{i}",
            "тип_селектора": "текст", "ид": f"id{i}", "название": f"Форма{i}",
            "имя": f"Форма{i}", "статус": "Успешно", "код": "",
        })

    for i in range(N):
        t = threading.Thread(target=_writer, args=(i,))
        threads.append(t)
        t.start()
    for t in threads:
        t.join()

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["Логи"] if "Логи" in wb.sheetnames else wb.active
    data_rows = ws.max_row - 1     # минус строка заголовка
    assert data_rows == N, f"ожидали {N} строк, получили {data_rows} - записи потерялись"
    print(f'✓ {N} конкурентных записей - все {N} строк на месте, ни одна не потеряна')


def test_concurrent_record_submitted_form_loses_no_entries(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(ta, "SUBMITTED_FORMS_FILE", str(tmp_path / "submitted_forms.json"))
    ta.reset_submitted_forms()

    N = 24
    threads = []

    def _writer(i):
        ta.record_submitted_form({"название": f"Форма{i}", "город": "Тест"})

    for i in range(N):
        t = threading.Thread(target=_writer, args=(i,))
        threads.append(t)
        t.start()
    for t in threads:
        t.join()

    import json
    with open(ta.SUBMITTED_FORMS_FILE, encoding="utf-8") as f:
        data = json.load(f)
    assert len(data) == N, f"ожидали {N} записей, получили {len(data)} - записи потерялись"
    print(f'✓ {N} конкурентных отправок в submitted_forms.json - все на месте')
