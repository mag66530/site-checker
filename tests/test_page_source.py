# -*- coding: utf-8 -*-
"""Источник адреса: страница из карты сайта должна быть видна в отчёте.

Раньше в прогон попадали случайные URL из sitemap, и в таблицах они выглядели
как обычные страницы каталога - понять, откуда они взялись, было нельзя."""
from openpyxl import Workbook

import reporter
from http_checker import CheckResult
from sources import Sources, build_custom_tasks_typed


def _результат(url, источник):
    r = CheckResult(url=url, city='Москва', subdomain='shopmet.ru',
                    type_code='category', type_label='Категория', source=источник)
    r.http_code, r.status = 200, 'ok'
    r.is_ok, r.is_error = True, False
    return r


def test_задачи_помечаются_источником():
    src = Sources(subdomains=[], categories=[], filters=[])
    из_карты = build_custom_tasks_typed(['https://shopmet.ru/catalog/mramor'],
                                        src, source='Карта сайта')
    свой = build_custom_tasks_typed(['https://shopmet.ru/catalog/armatura'], src)
    assert из_карты[0].source == 'Карта сайта'
    assert свой[0].source == 'Свой список URL'      # значение по умолчанию


def test_лист_страницы_показывает_источник():
    wb = Workbook()
    wb.remove(wb.active)
    reporter._build_pages_overview_sheet(wb, [
        _результат('https://shopmet.ru/catalog/armatura', 'Каталог проекта'),
        _результат('https://shopmet.ru/catalog/mramor', 'Карта сайта'),
    ], {}, None)
    ws = wb['Страницы']
    значения = [c.value for row in ws.iter_rows() for c in row if c.value]
    assert 'Источник' in значения                    # колонка есть
    assert 'Карта сайта' in значения                 # и заполнена


def test_проблемы_подписывают_страницу_из_карты_сайта():
    class Находка:
        level, section, problem = 'Ошибка', 'Техничка', 'Битая ссылка'
        city, page_type, detail = 'Москва', 'Категория', ''
        url = 'https://shopmet.ru/catalog/mramor'

    wb = Workbook()
    wb.remove(wb.active)
    reporter._build_problems_sheet(wb, [Находка()],
                                   {'https://shopmet.ru/catalog/mramor': 'Карта сайта'})
    ws = wb['Проблемы']
    значения = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    assert any('Карта сайта' in v for v in значения)
