"""Тесты metrika_404.py — парсинг тем писем и xlsx-таблиц."""
import sys
import tempfile
from pathlib import Path
sys.path.insert(0, '/home/claude/site-checker-py')

from metrika_404 import (
    parse_subject, parse_table_xlsx, is_table_attachment,
    Page404, Report404, save_report, load_report, list_stored_reports,
    COUNTRY_LABELS,
    _imap_utf7_encode, _imap_utf7_decode,
)


def test_imap_utf7_ascii_unchanged():
    """ASCII-имена не меняются."""
    assert _imap_utf7_encode('INBOX') == b'INBOX'
    assert _imap_utf7_encode('Sent Items') == b'Sent Items'
    assert _imap_utf7_decode(b'INBOX') == 'INBOX'
    print('✓ IMAP UTF-7: ASCII не меняется')


def test_imap_utf7_ampersand():
    """Символ & экранируется в &-."""
    assert _imap_utf7_encode('A&B') == b'A&-B'
    assert _imap_utf7_decode(b'A&-B') == 'A&B'
    print('✓ IMAP UTF-7: & → &-')


def test_imap_utf7_russian():
    """Кириллица кодируется/декодируется правильно."""
    # Реальная папка Яндекса
    encoded = _imap_utf7_encode('Я.Метрика 404 и др')
    # Должно начинаться с & и кончаться -
    assert encoded.startswith(b'&'), f'Не закодировалось: {encoded}'
    # Декод обратно
    decoded = _imap_utf7_decode(encoded)
    assert decoded == 'Я.Метрика 404 и др', f'Roundtrip сломан: {decoded!r}'
    print(f'✓ IMAP UTF-7: «Я.Метрика 404 и др» → {encoded.decode()}')


def test_imap_utf7_known_yandex_folders():
    """Roundtrip для типичных имён папок Яндекса."""
    test_cases = [
        'Входящие',
        'Отправленные',
        'Спам',
        'Корзина',
        'Я.Метрика 404 и др',
        'A&B & test',  # сочетание ASCII, & и не-ASCII
    ]
    for name in test_cases:
        encoded = _imap_utf7_encode(name)
        decoded = _imap_utf7_decode(encoded)
        assert decoded == name, f'Roundtrip сломан для {name!r}: got {decoded!r}'
    print(f'✓ IMAP UTF-7: {len(test_cases)} имён roundtrip успешно')


def test_parse_subject_basic():
    """Простой случай — Отчёт «АЗ 404 отчет» за 25.05.2026."""
    result = parse_subject('Отчёт «АЗ 404 отчет» за 25.05.2026')
    assert result == {'country': 'АЗ', 'date': '2026-05-25'}
    print('✓ parse_subject: АЗ за 25.05.2026')


def test_parse_subject_perevod():
    """Сложный случай с (перевод)."""
    result = parse_subject('Отчёт «АЗ (перевод) 404 отчет» за 25.05.2026')
    assert result is not None
    assert 'перевод' in result['country']
    assert result['date'] == '2026-05-25'
    print(f'✓ parse_subject с (перевод): {result["country"]}')


def test_parse_subject_all_countries():
    """Все страны из реальных писем."""
    for code in ['РФ', 'КЗ', 'РБ', 'УЗ', 'АЗ', 'АМ', 'КГ']:
        subj = f'Отчёт «{code} 404 отчет» за 22.05.2026'
        result = parse_subject(subj)
        assert result is not None, f'Не распарсилось: {subj}'
        assert result['country'] == code
        assert result['date'] == '2026-05-22'
    print('✓ parse_subject: все 7 стран распарсились')


def test_parse_subject_not_metrika():
    """Не наше письмо — None."""
    cases = [
        'Re: Какая-то рабочая переписка',
        '',
        'Скидки и акции в Яндекс',
        'Отчёт за прошлый месяц',  # нет «404 отчет»
    ]
    for s in cases:
        assert parse_subject(s) is None, f'Должно быть None для: {s!r}'
    print('✓ parse_subject: чужие письма игнорируются')


def test_is_table_attachment():
    assert is_table_attachment('АЗ_404_отчет_за_25_05_2026__таблица.xlsx') is True
    assert is_table_attachment('АЗ_404_отчет_за_25_05_2026__график.xlsx') is False
    assert is_table_attachment('report.xlsx') is False
    assert is_table_attachment(None) is False
    print('✓ is_table_attachment: «таблица» детектится')


def test_parse_empty_table_xlsx():
    """Пустой отчёт (как АЗ за 25.05.2026 — 0 строк)."""
    with open('/mnt/user-data/uploads/АЗ_404_отчет_за_25_05_2026__таблица.xlsx', 'rb') as f:
        pages = parse_table_xlsx(f.read())
    assert pages == [], f'Ожидался пустой список, получили {len(pages)} страниц'
    print('✓ Пустая таблица возвращает []')


def test_parse_real_table_xlsx_structure():
    """Реальный отчёт хотя бы не падает на парсинге."""
    with open('/mnt/user-data/uploads/АЗ_404_отчет_за_24_05_2026__таблица__1_.xlsx', 'rb') as f:
        pages = parse_table_xlsx(f.read())
    # Не должно упасть, может вернуть пустой список или данные
    assert isinstance(pages, list)
    print(f'✓ Реальный xlsx распарсился: {len(pages)} страниц')


def test_parse_table_with_synthetic_data():
    """Создам синтетический xlsx с данными — чтобы покрыть случай когда 404 есть."""
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Отчет за период с 2026-05-25 по 2026-05-25'
    ws['A2'] = 'Фильтры: Заголовок страницы содержит "404"'
    ws['A3'] = 'Атрибуция: Последний значимый переход'
    # Строка 4 пустая
    ws['A5'] = 'Заголовок страницы'
    ws['B5'] = 'Просмотры'
    ws['C5'] = 'Посетители'
    # Данные
    ws['A6'] = 'Страница не найдена | Стальметурал'
    ws['B6'] = 15
    ws['C6'] = 12
    ws['A7'] = 'Страница не найдена https://stalmetural.ru/catalog/broken/'
    ws['B7'] = 3
    ws['C7'] = 3

    buf = BytesIO()
    wb.save(buf)
    pages = parse_table_xlsx(buf.getvalue())

    assert len(pages) == 2
    assert pages[0].page_title == 'Страница не найдена | Стальметурал'
    assert pages[0].views == 15
    assert pages[0].visitors == 12
    assert pages[0].page_url is None  # нет URL в строке
    assert pages[1].page_url == 'https://stalmetural.ru/catalog/broken/'  # URL вытащился
    print(f'✓ Синтетический xlsx с данными: распарсилось {len(pages)} страниц')


def test_save_and_load_report():
    """Сохранение и загрузка отчёта."""
    import metrika_404
    original = metrika_404.CACHE_DIR
    with tempfile.TemporaryDirectory() as tmp:
        metrika_404.CACHE_DIR = Path(tmp)

        report = Report404(
            project_id='smu',
            country_code='РФ',
            country_name='Россия',
            report_date='2026-05-25',
            received_at='2026-05-25T10:00:00',
            pages=[
                Page404(page_title='Страница не найдена', page_url='https://example.com', views=10, visitors=8),
            ],
            total_views=10,
            total_pages=1,
        )
        save_report(report)

        loaded = load_report('smu', 'РФ', '2026-05-25')
        assert loaded is not None
        assert loaded.country_code == 'РФ'
        assert loaded.country_name == 'Россия'
        assert len(loaded.pages) == 1
        assert loaded.pages[0].views == 10

    metrika_404.CACHE_DIR = original
    print('✓ Сохранение и загрузка отчёта')


def test_list_stored_reports():
    """Листинг сохранённых отчётов."""
    import metrika_404
    original = metrika_404.CACHE_DIR
    with tempfile.TemporaryDirectory() as tmp:
        metrika_404.CACHE_DIR = Path(tmp)

        # Сохраняем 3 отчёта: 2 разных страны и 2 даты
        for code, date, views in [
            ('РФ', '2026-05-25', 50),
            ('РФ', '2026-05-24', 40),
            ('КЗ', '2026-05-25', 20),
        ]:
            r = Report404(
                project_id='smu', country_code=code, country_name=COUNTRY_LABELS.get(code, code),
                report_date=date, received_at='2026-05-25T10:00:00',
                pages=[Page404(page_title='x', page_url=None, views=views, visitors=views)],
                total_views=views, total_pages=1,
            )
            save_report(r)

        listing = list_stored_reports('smu')
        assert len(listing) == 3
        # Сортировка: сначала свежие даты
        assert listing[0]['date'] == '2026-05-25'

    metrika_404.CACHE_DIR = original
    print(f'✓ Листинг: {len(listing)} отчётов')


def test_get_stored_dates():
    """Получение множества дат, за которые есть отчёты."""
    import metrika_404
    from metrika_404 import get_stored_dates
    original = metrika_404.CACHE_DIR
    with tempfile.TemporaryDirectory() as tmp:
        metrika_404.CACHE_DIR = Path(tmp)
        for code, date in [('РФ', '2026-05-25'), ('КЗ', '2026-05-25'), ('РФ', '2026-05-24')]:
            r = Report404(
                project_id='smu', country_code=code, country_name=COUNTRY_LABELS.get(code, code),
                report_date=date, received_at='x',
                pages=[], total_views=0, total_pages=0,
            )
            save_report(r)
        dates = get_stored_dates('smu')
        assert dates == {'2026-05-25', '2026-05-24'}
    metrika_404.CACHE_DIR = original
    print('✓ get_stored_dates: {2026-05-25, 2026-05-24}')


def test_get_latest_available_date():
    """Самая свежая дата."""
    import metrika_404
    from metrika_404 import get_latest_available_date
    original = metrika_404.CACHE_DIR
    with tempfile.TemporaryDirectory() as tmp:
        metrika_404.CACHE_DIR = Path(tmp)
        # Пусто
        assert get_latest_available_date('smu') is None

        # Добавляем
        for date in ['2026-05-24', '2026-05-23', '2026-05-25']:
            r = Report404(
                project_id='smu', country_code='РФ', country_name='Россия',
                report_date=date, received_at='x',
                pages=[], total_views=0, total_pages=0,
            )
            save_report(r)
        assert get_latest_available_date('smu') == '2026-05-25'
    metrika_404.CACHE_DIR = original
    print('✓ get_latest_available_date: max из {24, 23, 25} = 25')


def test_load_reports_for_date():
    """Загрузка всех отчётов за одну дату."""
    import metrika_404
    from metrika_404 import load_reports_for_date
    original = metrika_404.CACHE_DIR
    with tempfile.TemporaryDirectory() as tmp:
        metrika_404.CACHE_DIR = Path(tmp)
        # 25 числа: РФ и КЗ, 24 числа: только РФ
        for code, date in [('РФ', '2026-05-25'), ('КЗ', '2026-05-25'), ('РФ', '2026-05-24')]:
            r = Report404(
                project_id='smu', country_code=code, country_name=COUNTRY_LABELS.get(code, code),
                report_date=date, received_at='x',
                pages=[Page404(page_title=f'{code} test', page_url=None, views=10, visitors=8)],
                total_views=10, total_pages=1,
            )
            save_report(r)

        reports_25 = load_reports_for_date('smu', '2026-05-25')
        assert len(reports_25) == 2
        assert {r.country_code for r in reports_25} == {'РФ', 'КЗ'}

        reports_24 = load_reports_for_date('smu', '2026-05-24')
        assert len(reports_24) == 1

        reports_22 = load_reports_for_date('smu', '2026-05-22')
        assert len(reports_22) == 0
    metrika_404.CACHE_DIR = original
    print('✓ load_reports_for_date: правильная фильтрация по дате')


def test_load_reports_for_period():
    """Загрузка за период N дней (от вчера на N дней назад)."""
    import metrika_404
    from metrika_404 import load_reports_for_period
    from datetime import datetime, timedelta
    original = metrika_404.CACHE_DIR
    with tempfile.TemporaryDirectory() as tmp:
        metrika_404.CACHE_DIR = Path(tmp)
        today = datetime.now().date()
        # Создаём отчёты на сегодня, вчера, 5 дней назад, 10 дней назад
        for offset in [0, 1, 5, 10]:
            d = (today - timedelta(days=offset)).strftime('%Y-%m-%d')
            r = Report404(
                project_id='smu', country_code='РФ', country_name='Россия',
                report_date=d, received_at='x',
                pages=[], total_views=0, total_pages=0,
            )
            save_report(r)

        # days=1 → только вчера, без сегодня
        period_1 = load_reports_for_period('smu', 1)
        yesterday_str = (today - timedelta(days=1)).strftime('%Y-%m-%d')
        assert len(period_1) == 1
        assert period_1[0].report_date == yesterday_str

        # days=7 → вчера + 5 дней назад (но не сегодня)
        period_7 = load_reports_for_period('smu', 7)
        assert len(period_7) == 2
        assert all(r.report_date != today.strftime('%Y-%m-%d') for r in period_7)

        # days=14 → ещё 10 дней назад
        period_14 = load_reports_for_period('smu', 14)
        assert len(period_14) == 3

        # days=30 → все кроме сегодня
        period_30 = load_reports_for_period('smu', 30)
        assert len(period_30) == 3
    metrika_404.CACHE_DIR = original
    print('✓ load_reports_for_period: семантика «N дней до вчера»')


if __name__ == '__main__':
    test_imap_utf7_ascii_unchanged()
    test_imap_utf7_ampersand()
    test_imap_utf7_russian()
    test_imap_utf7_known_yandex_folders()
    test_parse_subject_basic()
    test_parse_subject_perevod()
    test_parse_subject_all_countries()
    test_parse_subject_not_metrika()
    test_is_table_attachment()
    test_parse_empty_table_xlsx()
    test_parse_real_table_xlsx_structure()
    test_parse_table_with_synthetic_data()
    test_save_and_load_report()
    test_list_stored_reports()
    test_get_stored_dates()
    test_get_latest_available_date()
    test_load_reports_for_date()
    test_load_reports_for_period()
    print('\n✅ Все тесты metrika_404.py прошли')
