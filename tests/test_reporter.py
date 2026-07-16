"""Тесты reporter - генерация xlsx и проверка содержимого."""
import sys
import tempfile
import time
from pathlib import Path
from dataclasses import dataclass, field

sys.path.insert(0, '/home/claude/site-checker-py')

from reporter import build_report, make_report_filename
from text_checker import TextIssue
from http_checker import CheckResult, STATUS, SPEED
from sources import Subdomain


def make_result(**kw):
    """Хелпер для создания тестового CheckResult."""
    defaults = {
        'url': 'https://example.com/page',
        'city': 'Москва',
        'subdomain': 'example.com',
        'type_code': 'category',
        'type_label': 'Категория',
        'http_code': 200,
        'status': 'ok',
        'is_ok': True,
        'is_warning': False,
        'is_error': False,
        'elapsed_ms': 1200,
        'body_size': 50000,
        'speed_rating': 'fast',
        'attempts': 1,
        'final_url': None,
        'redirect_chain': [],
        'error_kind': None,
        'error_message': None,
        'text_issues': [],
        'has_text_issues': False,
    }
    defaults.update(kw)
    return CheckResult(**defaults)


def test_basic_report_creation():
    """Базовый случай - отчёт создаётся и открывается."""
    results = [
        make_result(url='https://stalmetural.ru/', type_label='Главная'),
        make_result(url='https://stalmetural.ru/catalog/', type_label='Каталог', elapsed_ms=2800, speed_rating='normal'),
        make_result(url='https://kazan.stalmetural.ru/catalog/dead/',
                    city='Казань', subdomain='kazan.stalmetural.ru',
                    http_code=404, status='not_found',
                    is_ok=False, is_error=True, speed_rating=None,
                    elapsed_ms=100, body_size=1000),
    ]
    selected = [
        Subdomain(url='https://stalmetural.ru/', city='Москва', host='stalmetural.ru'),
        Subdomain(url='https://kazan.stalmetural.ru/', city='Казань', host='kazan.stalmetural.ru'),
    ]
    
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / 'test.xlsx'
        build_report(
            project_name='Тест проект',
            started_at_ms=int(time.time() * 1000) - 30000,
            finished_at_ms=int(time.time() * 1000),
            selected_subdomains=selected,
            results=results,
            output_path=out,
        )
        assert out.exists()
        assert out.stat().st_size > 5000  # явно не пустой
        
        # Открываем и проверяем содержимое
        from openpyxl import load_workbook
        wb = load_workbook(out)
        assert 'Обзор' in wb.sheetnames
        assert 'Все детали' in wb.sheetnames
        assert 'Битые тексты' not in wb.sheetnames  # нет находок
        
        # На листе «Все детали» - 3 строки данных + 1 шапка = 4
        ws = wb['Все детали']
        rows_with_data = sum(1 for r in ws.iter_rows(values_only=True) if r[0])
        assert rows_with_data == 4
    print('✓ Базовый отчёт создаётся')


def test_report_with_text_issues():
    """Если есть битые тексты - добавляется третий лист."""
    issue1 = TextIssue(pattern='{{...}}', match='{{city}}',
                       context='Купить трубу в {{city}} с доставкой')
    issue2 = TextIssue(pattern='%переменная%', match='%price%',
                       context='Цена от %price% рублей')
    
    results = [
        make_result(url='https://stalmetural.ru/cat-a',
                    text_issues=[issue1], has_text_issues=True),
        make_result(url='https://stalmetural.ru/cat-b',
                    text_issues=[issue2], has_text_issues=True),
    ]
    selected = [Subdomain(url='https://stalmetural.ru/', city='Москва', host='stalmetural.ru')]
    
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / 'test.xlsx'
        build_report(
            project_name='Тест', started_at_ms=int(time.time() * 1000) - 5000,
            finished_at_ms=int(time.time() * 1000),
            selected_subdomains=selected, results=results, output_path=out,
        )
        from openpyxl import load_workbook
        wb = load_workbook(out)
        # «Битые тексты» теперь - секция внутри группового листа «Техничка».
        assert 'Техничка' in wb.sheetnames
        assert 'Битые тексты' not in wb.sheetnames
        ws = wb['Техничка']
        all_cells = []
        for row in ws.iter_rows(values_only=True):
            all_cells.extend(c for c in row if c)
        assert '{{city}}' in all_cells
        assert '%price%' in all_cells
        assert '▸ Битые тексты' in all_cells   # полоса-разделитель секции
    print('✓ Секция «Битые тексты» в листе «Техничка»')


def test_redirect_chain_in_path_column():
    """Цепочка редиректов корректно отображается в «Откуда перешли»."""
    chain = [
        {'from': 'https://stalmetural.ru/old/', 'to': 'https://stalmetural.ru/new/', 'code': 301},
    ]
    results = [
        # 404 после редиректа
        make_result(url='https://stalmetural.ru/old/',
                    http_code=404, status='not_found',
                    is_ok=False, is_error=True, speed_rating=None,
                    redirect_chain=chain),
        # 404 без редиректа (прямая ссылка из каталога)
        make_result(url='https://stalmetural.ru/dead/',
                    http_code=404, status='not_found',
                    is_ok=False, is_error=True, speed_rating=None),
    ]
    selected = [Subdomain(url='https://stalmetural.ru/', city='Москва', host='stalmetural.ru')]
    
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / 'test.xlsx'
        build_report(
            project_name='Тест', started_at_ms=int(time.time() * 1000) - 5000,
            finished_at_ms=int(time.time() * 1000),
            selected_subdomains=selected, results=results, output_path=out,
        )
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb['Все детали']

        # Находим колонку «Откуда перешли» по заголовку, а не по фикс. индексу -
        # в лист «Все детали» со временем добавляли колонки (напр. «Отдел»),
        # из-за чего хардкод column=10 ломался.
        header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        path_col = header_row.index('Откуда перешли') + 1
        paths = [ws.cell(row=r, column=path_col).value for r in range(2, 4)]
        # Должна быть и цепочка и «Прямая ссылка»
        assert any('301:' in p for p in paths if p)
        assert any('Прямая ссылка' in p for p in paths if p)
    print('✓ Колонка «Откуда перешли» правильная')


def test_speed_with_comma():
    """Скорость отображается с запятой (для русского Excel)."""
    results = [make_result(elapsed_ms=2340)]  # 2,34 сек
    selected = [Subdomain(url='https://stalmetural.ru/', city='Москва', host='stalmetural.ru')]
    
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / 'test.xlsx'
        build_report(
            project_name='Тест', started_at_ms=int(time.time() * 1000) - 5000,
            finished_at_ms=int(time.time() * 1000),
            selected_subdomains=selected, results=results, output_path=out,
        )
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb['Все детали']
        # Колонка G - «Скорость, с»
        speed = ws.cell(row=2, column=7).value
        assert speed == '2,34', f'Ожидалось "2,34", получили {speed!r}'
    print('✓ Скорость с запятой')


def test_make_report_filename():
    """Имена файлов: smu-21.05.2026, smu-21.05.2026_2, ..."""
    with tempfile.TemporaryDirectory() as tmp:
        d = Path(tmp)
        # дата фиксированная для воспроизводимости
        ts = int(time.mktime(time.strptime('2026-05-21 10:00', '%Y-%m-%d %H:%M')) * 1000)
        
        n1 = make_report_filename('smu', ts, d)
        assert n1 == 'smu-21.05.2026.xlsx'
        (d / n1).write_text('x')
        
        n2 = make_report_filename('smu', ts, d)
        assert n2 == 'smu-21.05.2026_2.xlsx'
        (d / n2).write_text('x')
        
        n3 = make_report_filename('smu', ts, d)
        assert n3 == 'smu-21.05.2026_3.xlsx'
        
        # Другой проект - без суффикса
        n4 = make_report_filename('mpe', ts, d)
        assert n4 == 'mpe-21.05.2026.xlsx'
    print('✓ Имена файлов: smu-21.05.2026 с инкрементом')


def test_custom_run_no_subdomains():
    """Custom-прогон: пустой selected_subdomains не ломает отчёт."""
    results = [
        make_result(url='https://example.com/page1', city='', subdomain='example.com',
                    type_code='custom', type_label='URL'),
        make_result(url='https://example.com/page2', city='', subdomain='example.com',
                    type_code='custom', type_label='URL',
                    http_code=404, status='not_found',
                    is_ok=False, is_error=True, speed_rating=None),
    ]
    
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / 'custom.xlsx'
        build_report(
            project_name='Свой список URL',
            started_at_ms=int(time.time() * 1000) - 3000,
            finished_at_ms=int(time.time() * 1000),
            selected_subdomains=[],  # пусто
            results=results,
            output_path=out,
        )
        from openpyxl import load_workbook
        wb = load_workbook(out)
        # Должен открыться без ошибок
        ws = wb['Обзор']
        # На листе «Обзор» в параметрах прогона - только «Длительность», без «Поддоменов»
        text_in_overview = []
        for row in ws.iter_rows(values_only=True):
            text_in_overview.extend(str(c) for c in row if c)
        assert 'Длительность' in text_in_overview
        assert 'Поддоменов' not in text_in_overview
    print('✓ Custom-прогон без поддоменов: отчёт корректный')


def test_tech_section_mandatory_bug_and_broken_links():
    """Тех. страница с обязательным багом (нет карты на «Контактах») и битой
    ссылкой (404) корректно отображается в листе «Структура страниц»."""
    from content_checker import check_content
    base = ('<header><a href="tel:1">т</a></header><h1>Контакты</h1><p>'
            + ('Адрес и режим работы. ' * 30) + '</p>')
    tech_content = check_content(base, 'tech', url='https://inmetprom.ru/contact/')
    # Нужна хотя бы одна НЕ тех. страница с контентом, иначе лист структуры
    # не строится (тех. секция живёт внутри него).
    cat_html = ('<header><a href="tel:+74951234567">+7 (495) 123-45-67</a></header>'
                '<div class="breadcrumb">x</div><h1>Категория</h1>'
                '<div class="catalog-product-card-item"><a href="/catalog/c/t/">Т</a>'
                '<span>1 200 ₽</span><span class="an-ico-basket"></span></div>')
    cat_content = check_content(cat_html, 'category', url='https://inmetprom.ru/catalog/truby/')
    results = [
        make_result(url='https://inmetprom.ru/catalog/truby/', type_code='category',
                    type_label='Категория', content=cat_content,
                    content_bugs=cat_content.bug_count,
                    has_content_bugs=cat_content.has_bugs),
        make_result(url='https://inmetprom.ru/contact/', type_code='tech',
                    type_label='Тех. страница', content=tech_content,
                    content_bugs=tech_content.bug_count,
                    has_content_bugs=tech_content.has_bugs,
                    broken_links={'checked': 5, 'broken': [
                        {'url': 'https://inmetprom.ru/dead/', 'code': 404}]}),
    ]
    selected = [Subdomain(url='https://inmetprom.ru/', city='Москва', host='inmetprom.ru')]
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / 'tech.xlsx'
        build_report(project_name='ИМП', started_at_ms=int(time.time() * 1000) - 5000,
                     finished_at_ms=int(time.time() * 1000),
                     selected_subdomains=selected, results=results, output_path=out)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        # «Структура страниц» теперь - секция внутри группового «Техничка».
        assert 'Техничка' in wb.sheetnames
        assert 'Структура страниц' not in wb.sheetnames
        ws = wb['Техничка']
        blob = ' | '.join(str(c) for row in ws.iter_rows(values_only=True)
                          for c in row if c)
        assert 'Карта: БАГ' in blob       # обязательный спец-элемент отсутствует
        assert 'битых' in blob            # битая ссылка (404)
        # В текстах отчёта не должно быть ДЛИННЫХ тире («—») - проект перешёл
        # на дефисы (коммит «Длинные тире -> дефисы»). Проверяем именно «—»:
        # раньше тут стояло «—», но замена тире прошлась и по самому тесту,
        # превратив проверку в «нет дефиса» - и тест ложно падал.
        assert '—' not in blob
    print('✓ Тех. секция: обязательный баг + битые ссылки в отчёте')


def test_cell_state_bug_shows_count():
    """Ячейка грида: обязательный блок с числом (напр. «Фото товаров» - сколько
    без фото) показывает «БАГ (N)», а без числа - просто «БАГ»."""
    from reporter import _cell_state
    from types import SimpleNamespace as NS
    col = {'kind': 'block', 'key': 'photos', 'label': 'Фото товаров'}
    by_key = {'photos': NS(key='photos', label='Фото товаров', required=True,
                           present=False, count=23)}
    assert _cell_state(col, by_key) == ('БАГ (23)', 'bug')
    # обязательный без числа (count=None) - просто «БАГ»
    col2 = {'kind': 'block', 'key': 'h1', 'label': 'H1'}
    bk2 = {'h1': NS(key='h1', label='H1', required=True, present=False, count=None)}
    assert _cell_state(col2, bk2) == ('БАГ', 'bug')
    # обязательный с count=0 (напр. «Карточки товаров» = 0) - «БАГ», не «БАГ (0)»
    col3 = {'kind': 'block', 'key': 'product_cards', 'label': 'Карточки'}
    bk3 = {'product_cards': NS(key='product_cards', label='Карточки', required=True,
                              present=False, count=0)}
    assert _cell_state(col3, bk3) == ('БАГ', 'bug')
    print('✓ Ячейка «БАГ (N)» показывает число, где оно есть')


def test_problem_text_human_phrases():
    """«Что чинить» пишет по-человечески: «нет цены», а не «нет: Цена (есть)»."""
    from reporter import _problem_text
    from content_checker import check_content
    from types import SimpleNamespace as NS
    html = ('<header><a href="tel:1">т</a></header><div class="breadcrumb">x</div>'
            '<h1>Категория</h1><div class="catalog-product-card-item">'
            '<a href="/c/t/">Товар</a><img src="/i/a.jpg"></div>')
    content = check_content(html, 'category', url='https://x.ru/catalog/c/')
    txt = _problem_text(NS(content=content, broken_links=None, contacts_addr=None,
                           page_phone=None, kp_result=None))
    assert 'нет цены' in txt and 'нет кнопки заказа' in txt
    assert 'Цена (есть)' not in txt and 'нет:' not in txt
    print('✓ «Что чинить»: человеческие формулировки')


def test_metrika_404_goal_есть():
    """Цель на 404 найдена - строка появляется на листе «404 из Метрики»,
    даже если 404-страниц за период не было (сам факт проверки цели не
    должен теряться, когда сшивка сама по себе пустая)."""
    selected = [Subdomain(url='https://stalmetural.ru/', city='Москва', host='stalmetural.ru')]
    results = [make_result(url='https://stalmetural.ru/')]

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / 'test.xlsx'
        build_report(
            project_name='Тест', started_at_ms=int(time.time() * 1000) - 5000,
            finished_at_ms=int(time.time() * 1000),
            selected_subdomains=selected, results=results, output_path=out,
            metrika_reports=None,
            metrika_404_goal={'есть': True, 'счётчики':
                               {'15630172': {'есть': True, 'название': '404'}}},
        )
        from openpyxl import load_workbook
        wb = load_workbook(out)
        # «404 из Метрики» теперь - секция внутри группового «Аналитика».
        assert 'Аналитика' in wb.sheetnames
        assert '404 из Метрики' not in wb.sheetnames
        ws = wb['Аналитика']
        text = ' '.join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        assert 'Цель на 404' in text
        assert 'есть' in text.lower()
    print('✓ Секция «404 из Метрики» в листе «Аналитика»: «цель есть»')


def test_metrika_404_goal_не_найдена():
    selected = [Subdomain(url='https://mepen.ru/', city='Москва', host='mepen.ru')]
    results = [make_result(url='https://mepen.ru/')]

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / 'test.xlsx'
        build_report(
            project_name='Тест', started_at_ms=int(time.time() * 1000) - 5000,
            finished_at_ms=int(time.time() * 1000),
            selected_subdomains=selected, results=results, output_path=out,
            metrika_reports=None,
            metrika_404_goal={'есть': False, 'счётчики':
                               {'99551890': {'есть': False, 'название': None}}},
        )
        from openpyxl import load_workbook
        wb = load_workbook(out)
        assert 'Аналитика' in wb.sheetnames
        assert '404 из Метрики' not in wb.sheetnames
        ws = wb['Аналитика']
        text = ' '.join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        assert 'не найдена' in text.lower()
        assert 'стоит создать' in text.lower()
    print('✓ Секция «404 из Метрики» в «Аналитике»: «цель не найдена»')


def test_metrika_404_goal_none_без_данных_лист_не_создаётся():
    """metrika_404_goal не передан (сбор 404 выключен) и отчётов о 404 тоже
    нет - лист не появляется (прежнее поведение не сломано)."""
    selected = [Subdomain(url='https://stalmetural.ru/', city='Москва', host='stalmetural.ru')]
    results = [make_result(url='https://stalmetural.ru/')]

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / 'test.xlsx'
        build_report(
            project_name='Тест', started_at_ms=int(time.time() * 1000) - 5000,
            finished_at_ms=int(time.time() * 1000),
            selected_subdomains=selected, results=results, output_path=out,
        )
        from openpyxl import load_workbook
        wb = load_workbook(out)
        assert '404 из Метрики' not in wb.sheetnames
    print('✓ Без данных о 404 (ни отчётов, ни цели) лист не создаётся')


if __name__ == '__main__':
    test_basic_report_creation()
    test_report_with_text_issues()
    test_redirect_chain_in_path_column()
    test_speed_with_comma()
    test_make_report_filename()
    test_custom_run_no_subdomains()
    test_metrika_404_goal_есть()
    test_metrika_404_goal_не_найдена()
    test_metrika_404_goal_none_без_данных_лист_не_создаётся()
    print('\n✅ Все тесты reporter.py прошли')
