# -*- coding: utf-8 -*-
"""Лист «Аналитика»: события из сервисов одной таблицей.

Раньше это были три листа с семью разными наборами колонок; на общем листе
они склеивались подряд, и колонка «C» в одном блоке значила «Серьёзность», в
другом «Оценка», в третьем ничего. У писем шапка повторялась для каждой
серьёзности - блок почты выглядел длиннее всего остального листа."""
from openpyxl import Workbook

import reporter
from webmaster_api import ServiceIssue
from webmaster_notify import WebmasterNotification as N


def _письмо(**kw):
    основа = dict(msg_id='1', project_id='p', source='yandex_webmaster',
                  date='2026-08-10', subject='Тема', body_preview='текст',
                  priority='important', category='indexing')
    основа.update(kw)
    return N(**основа)


def _лист(письма=(), проблемы=(), автокликер=None):
    wb = Workbook()
    wb.remove(wb.active)
    reporter._build_notifications_sheet(wb, list(письма), list(проблемы),
                                        автокликер)
    return wb['Аналитика']


def _строки(ws):
    """Строки таблицы: {заголовок: значение}."""
    шапка = [c.value for c in ws[5]]
    out = []
    for row in ws.iter_rows(min_row=6):
        знач = [c.value for c in row]
        if isinstance(знач[1], int):        # № строки таблицы
            out.append({h: v for h, v in zip(шапка, знач) if h})
    return out


def test_одна_шапка_на_весь_лист():
    ws = _лист([_письмо(), _письмо(msg_id='2', source='gsc', subject='Другое')])
    заголовки = [c.value for c in ws[5] if c.value]
    assert заголовки == ['№', 'Уровень', 'Источник', 'Дата', 'Событие',
                         'Сайты', 'Кол-во', 'Подробности', 'Отдел']
    # шапка больше нигде не повторяется
    повторы = sum(1 for row in ws.iter_rows()
                  for c in row if c.value == 'Уровень')
    assert повторы == 1


def test_таблица_фильтруется():
    ws = _лист([_письмо()])
    assert ws.auto_filter.ref and ws.freeze_panes


def test_письма_и_api_в_одной_таблице():
    ws = _лист([_письмо(subject='Письмо про индексацию')],
               [ServiceIssue(project_id='p', service='webmaster', host='a.ru',
                             severity='fatal', code='X', title='Сайт не грузится')])
    источники = {r['Источник'] for r in _строки(ws)}
    assert источники == {'Вебмастер (почта)', 'Вебмастер (API)'}


def test_почта_и_api_различимы():
    """Письмо часто дублирует то, что API уже отдал - надо видеть, откуда."""
    ws = _лист([_письмо()], [ServiceIssue(project_id='p', service='webmaster',
                                          host='a.ru', severity='fatal',
                                          code='X', title='Т')])
    for r in _строки(ws):
        assert '(почта)' in r['Источник'] or '(API)' in r['Источник']


def test_одинаковые_письма_по_сайтам_сводятся_в_строку():
    """У проекта на 240 поддоменов одно письмо приходит 240 раз."""
    ws = _лист([_письмо(msg_id='1', subject='Сайт недоступен - msk.a.ru'),
                _письмо(msg_id='2', subject='Сайт недоступен - spb.a.ru')])
    строки = _строки(ws)
    assert len(строки) == 1
    assert строки[0]['Кол-во'] == 2
    assert 'msk.a.ru' in строки[0]['Сайты'] and 'spb.a.ru' in строки[0]['Сайты']


def test_критичное_сверху():
    ws = _лист([_письмо(msg_id='1', priority='info', subject='Мелочь'),
                _письмо(msg_id='2', priority='critical', subject='Пожар')])
    уровни = [r['Уровень'] for r in _строки(ws)]
    assert уровни[0] == 'Критично'


def test_шкалы_писем_и_api_сведены_к_общей():
    """У писем priority, у API severity - в колонке должно быть одно слово."""
    ws = _лист([_письмо(priority='important')],
               [ServiceIssue(project_id='p', service='webmaster', host='a.ru',
                             severity='possible', code='X', title='Т')])
    assert {r['Уровень'] for r in _строки(ws)} == {'Важно'}


def test_отзыв_2гис_сохраняет_оценку_и_ссылку():
    ws = _лист([_письмо(source='twogis', priority='info', rating=2,
                        subject='Новый отзыв',
                        review_url='https://2gis.ru/otziv/1')])
    r = _строки(ws)[0]
    assert '★' in str(r['Подробности'])
    # ссылка кликабельна
    ячейка = next(c for row in ws.iter_rows() for c in row
                  if c.value and 'отзыв' in str(c.value).lower() and c.hyperlink)
    assert ячейка.hyperlink


def test_автокликер_идёт_под_таблицей():
    ws = _лист([_письмо()], автокликер={'services': [
        {'service': 'Вебмастер', 'problems': 12, 'clicked': 9,
         'checking': 2, 'no_button': 1, 'errors': 0}]})
    значения = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    assert any('Автокликер' in v for v in значения)
    последняя_фильтра = int(ws.auto_filter.ref.split(':')[1][1:])
    строка_блока = next(c.row for row in ws.iter_rows() for c in row
                        if c.value and 'Автокликер' in str(c.value))
    assert строка_блока > последняя_фильтра


def test_пустой_лист_объясняет_причину():
    ws = _лист()
    значения = ' '.join(str(c.value) for row in ws.iter_rows()
                        for c in row if c.value)
    assert 'событий не найдено' in значения
    assert 'пароли приложений' in значения


def test_лист_отзывов_переименован():
    """«Я.Бизнес и GMB» → «Отзывы»: на листе давно только докупка отзывов."""
    wb = Workbook()
    wb.remove(wb.active)
    reporter._build_yabusiness_sheet(wb, {
        'available': True, 'total_branches': 1, 'low_rating_count': 0,
        'branches': [{'city': 'Пермь', 'yandex': {'rating': 4.9, 'count': 10},
                      'twogis': {}, 'rating': 4.9, 'low_rating': False,
                      'order': 2}]})
    assert 'Отзывы' in wb.sheetnames
    assert 'Я.Бизнес и GMB' not in wb.sheetnames
