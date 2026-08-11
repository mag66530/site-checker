# -*- coding: utf-8 -*-
"""Лист «Трафик и траст»: раскладка блоков и ИКС.

Два живых случая:
  • блоки стояли в намертво заданных колонках (M/Q/X). Когда трафика не было,
    лист открывался с пустой левой третью, а таблицы уезжали за край экрана;
  • ИКС читался из корня хоста, где поля sqi нет вовсе - прогон отчитывался
    «хостов 242», а в отчёте стояли прочерки."""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import reporter

ТРАСТ = {'available': True, 'has_dr': False,
         'hosts': [{'host': 'a.ru', 'sqi': 220, 'dr': None}]}
ССЫЛКИ = {'available': True, 'hosts': [
    {'host': 'a.ru', 'links': 21, 'donors': 7, 'panel_url': 'https://x',
     'issues': [], 'recent': 0, 'spam': 0}]}
ТРАФИК = {'rows': [{'d1': '2026-07-01', 'd2': '2026-07-31', 'period': 'Месяц',
                    'kind': 'текущий', 'country': 'Россия', 'visits': 76,
                    'direct': 64, 'yandex': 5, 'google': 7, 'leads': 1,
                    'conv': 1.3, 'bounce': 20}]}


def _колонки(traffic=None, trust=None, link_profile=None, gsc_pages=None):
    """{название блока: буква колонки, где он начинается}."""
    wb = Workbook()
    wb.remove(wb.active)
    reporter._build_traffic_overview_sheet(wb, traffic, trust, link_profile,
                                           gsc_pages)
    if 'Трафик и траст' not in wb.sheetnames:
        return {}
    ws = wb['Трафик и траст']
    места = {}
    for row in ws.iter_rows(min_row=3, max_row=6):
        for c in row:
            v = str(c.value or '')
            for имя, кусок in (('Траст', 'Траст проекта'),
                               ('Ссылки', 'Ссылочный профиль'),
                               ('Трафик', 'Страна')):
                if кусок in v and имя not in места:
                    места[имя] = get_column_letter(c.column)
    return места


def test_единственный_блок_встаёт_слева():
    """Раньше ссылочный профиль начинался с Q - две трети листа пустовали."""
    assert _колонки(link_profile=ССЫЛКИ) == {'Ссылки': 'B'}


def test_без_трафика_блоки_сдвигаются_влево():
    места = _колонки(trust=ТРАСТ, link_profile=ССЫЛКИ)
    assert места['Траст'] == 'B'
    assert места['Ссылки'] < 'M'          # раньше было Q


def test_блоки_не_наезжают_друг_на_друга():
    места = _колонки(trust=ТРАСТ, link_profile=ССЫЛКИ)
    assert места['Траст'] != места['Ссылки']


def test_с_трафиком_порядок_прежний():
    места = _колонки(traffic=ТРАФИК, trust=ТРАСТ, link_profile=ССЫЛКИ)
    assert места['Трафик'] == 'B'
    assert места['Траст'] > 'B' and места['Ссылки'] > места['Траст']


def test_пустой_траст_не_резервирует_место():
    """Условие раскладки должно совпадать с условием отрисовки блока."""
    места = _колонки(trust={'available': False, 'note': 'нет токена'},
                     link_profile=ССЫЛКИ)
    assert места == {'Ссылки': 'B'}


# ── ИКС: правильный источник данных ──────────────────────────────────

def test_икс_читается_из_summary(monkeypatch):
    """В корне хоста поля sqi нет - только в /summary/."""
    import trust_check
    import webmaster_api

    запрошено = []

    def _get(token, path, proxy_url=None, params=None):
        запрошено.append(path)
        if path == '/user/':
            return {'user_id': 1}
        if path.endswith('/hosts/'):
            return {'hosts': [{'host_id': 'https:a.ru:443',
                               'ascii_host_url': 'https://a.ru/'}]}
        if path.endswith('/summary/'):
            return {'sqi': 220, 'searchable_pages_count': 10}
        return {'ascii_host_url': 'https://a.ru/', 'verified': True}

    # fetch_sqi импортирует обе функции ИЗ webmaster_api - патчим там.
    monkeypatch.setattr(webmaster_api, '_get', _get)
    monkeypatch.setattr(webmaster_api, '_project_hosts', lambda pid: set())
    hosts = trust_check.fetch_sqi('imp', 'токен')
    assert hosts and hosts[0]['sqi'] == 220
    assert any(p.endswith('/summary/') for p in запрошено), запрошено
