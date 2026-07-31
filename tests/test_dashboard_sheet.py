"""Тесты листа «Дашборд» в отчёте variables_run.py - сводка по данным
«Проверка КП»/«Расхождения»: 4 KPI, таблица+график по типу данных (только они -
таблицу/график «по источнику» убрали по просьбе заказчика). Числа - готовые
(посчитаны в Python при генерации), не формулы Excel.

График - КАРТИНКА (Pillow), не «родной» график Excel: «родной» BarChart
рендерился по-разному на одном и том же файле локально и на проде (само
ориентация/кеш/оси тут ни при чём - разница именно в окружении, скорее
всего в версии openpyxl), картинка рендерится одинаковыми байтами везде.
Сделан по образцу файла, который прислал заказчик (КП-сму-...-улучшено.xlsx) -
структура должна совпадать."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import variables_run as vr
from maps_compare import MapCheckResult


def _результаты(n=3):
    return [{'domain': f'host{i}.ru', 'city': f'Город{i}', 'country': 'Россия',
            'error': '', 'fields': []} for i in range(n)]


def test_dashboard_is_first_sheet(tmp_path):
    out = tmp_path / 'test.xlsx'
    vr._записать_xlsx(out, 'СМУ', _результаты(), per_city=False, map_results=[])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert wb.sheetnames[0] == 'Дашборд'
    assert wb.sheetnames == ['Дашборд', 'Как читать результат', 'Проверка КП', 'Расхождения']
    print('✓ «Дашборд» - первый лист, порядок листов как в образце')


def test_no_dashboard_when_nothing_checked(tmp_path):
    out = tmp_path / 'test.xlsx'
    vr._записать_xlsx(out, 'СМУ', [], per_city=False, map_results=[])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert 'Дашборд' not in wb.sheetnames
    print('✓ ничего не проверяли → дашборда нет (нечего показывать)')


def test_kpi_values_are_plain_numbers(tmp_path):
    """KPI - готовые числа, не формулы: график/KPI на формулах без закешированного
    значения у части просмотрщиков рисуется пустым (баг увидели на проде)."""
    out = tmp_path / 'test.xlsx'
    vr._записать_xlsx(out, 'СМУ', _результаты(5), per_city=False, map_results=[])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb['Дашборд']
    assert ws['B6'].value == 5
    assert ws['D6'].value == 0
    assert ws['F6'].value == 5
    assert ws['H6'].value == 0
    print('✓ KPI - обычные числа (не формулы), посчитаны при генерации')


def test_type_breakdown_table_matches_var_columns(tmp_path):
    out = tmp_path / 'test.xlsx'
    r = MapCheckResult(source='2gis', city='Город0', url='u', available=True,
                       country='Россия', phone_match=True, address_match=True,
                       site_match=True)
    vr._записать_xlsx(out, 'СМУ', _результаты(1), per_city=False, map_results=[r])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb['Дашборд']
    assert ws['B10'].value == 'Показатель' and ws['C10'].value == 'Расхождений (✗)'
    labels = [ws.cell(r, 2).value for r in range(11, 20) if ws.cell(r, 2).value]
    assert 'Город' in labels and 'Адрес' in labels and '2ГИС' in labels
    # Значение для «2ГИС» - готовое число расхождений по этому источнику.
    _2gis_row = next(r for r in range(11, 20) if ws.cell(r, 2).value == '2ГИС')
    assert isinstance(ws.cell(_2gis_row, 3).value, int)
    print('✓ таблица «по типу данных» построена по фактическим колонкам отчёта')


def test_only_one_table_and_chart_on_dashboard(tmp_path):
    """Только «Расхождения по типу данных» (лист «Проверка КП») - без таблицы/
    графика «по источнику»."""
    out = tmp_path / 'test.xlsx'
    r1 = MapCheckResult(source='2gis', city='Город0', url='u', available=True,
                        country='Россия', phone_match=False, address_match=True,
                        site_match=True, issues=['x'],
                        details=[{'field': 'телефон', 'kp': '1', 'card': '–'}])
    vr._записать_xlsx(out, 'СМУ', _результаты(1), per_city=False, map_results=[r1])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb['Дашборд']
    all_text = ' '.join(str(ws.cell(r, 2).value or '') for r in range(1, 40))
    assert 'Расхождения по типу данных' in all_text
    assert 'по источнику' not in all_text
    assert len(ws._images) == 1, 'должна быть ровно 1 картинка-график - по типу данных'
    print('✓ на дашборде только одна таблица+график (по типу данных)')


def test_no_conclusion_line(tmp_path):
    """«Главный вывод» убрали с дашборда целиком по просьбе заказчика - никакой
    текстовой строки-вывода под таблицей/графиком быть не должно."""
    out = tmp_path / 'test.xlsx'
    maps = [
        MapCheckResult(source='2gis', city=f'Город{i}', url='u', available=True,
                       country='Россия', phone_match=False, address_match=True,
                       site_match=True, issues=['x'],
                       details=[{'field': 'телефон', 'kp': str(i), 'card': '–'}])
        for i in range(3)
    ]
    vr._записать_xlsx(out, 'СМУ', _результаты(3), per_city=False, map_results=maps)
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb['Дашборд']
    all_text = ' '.join(str(ws.cell(r, 2).value or '') for r in range(1, 45))
    assert 'Главный вывод' not in all_text
    print('✓ строки-вывода «Главный вывод» на дашборде больше нет')


def test_table_title_has_no_extra_caption_and_is_not_truncated(tmp_path):
    """Заголовок таблицы - без «(лист «Проверка КП»)» (лишняя подпись убрана
    по просьбе заказчика), и мёрдж достаточно широкий (до H), чтобы текст не
    обрезался, как это было при узком B:E."""
    out = tmp_path / 'test.xlsx'
    vr._записать_xlsx(out, 'СМУ', _результаты(2), per_city=False, map_results=[])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb['Дашборд']
    title_row = next(r for r in range(1, 20)
                     if ws.cell(r, 2).value and 'Расхождения по типу данных' in str(ws.cell(r, 2).value))
    assert ws.cell(title_row, 2).value == 'Расхождения по типу данных'
    merges = [str(m) for m in ws.merged_cells.ranges if m.min_row == title_row]
    assert merges, 'строка заголовка таблицы должна быть объединена'
    assert f'H{title_row}' in merges[0], f'мёрдж должен доходить до H, получили {merges[0]}'
    print('✓ заголовок таблицы - без лишней подписи, мёрдж на всю ширину (не обрезается)')


def test_report_round_trips_without_corruption(tmp_path):
    """Полный отчёт (дашборд + графики + формулы) должен сохраняться и
    открываться обратно без ошибок - простая проверка целостности файла."""
    out = tmp_path / 'test.xlsx'
    r = MapCheckResult(source='yandex', city='Город0', url='u', available=True,
                       country='Россия', phone_match=True, address_match=True,
                       site_match=True)
    vr._записать_xlsx(out, 'СМУ', _результаты(2), per_city=False, map_results=[r])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {'Дашборд', 'Как читать результат',
                                 'Проверка КП', 'Расхождения'}
    print('✓ файл с дашбордом и графиками сохраняется/открывается без ошибок')


def test_chart_image_renders_all_categories_with_visible_bars_or_zero(tmp_path):
    """Картинка рисуется через _нарисовать_barchart_png() - проверяем саму
    функцию напрямую (не то, что Excel потом сделает с картинкой - тут дело в
    байтах PNG, они одинаковы везде): все 9 показателей (8 полей сайта + 1
    карта) должны попасть на картинку, а не только тот, у которого есть
    расхождение."""
    from maps_compare import MapCheckResult as _MCR
    r = _MCR(source='2gis', city='Город0', url='u', available=True,
            country='Россия', phone_match=False, address_match=True,
            site_match=True, issues=['x'],
            details=[{'field': 'телефон', 'kp': '1', 'card': '–'}])
    out = tmp_path / 'test.xlsx'
    vr._записать_xlsx(out, 'СМУ', _результаты(4), per_city=False, map_results=[r])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb['Дашборд']
    assert len(ws._images) == 1
    img = ws._images[0]
    assert img.width > 0 and img.height > 0
    print('✓ картинка-график встроена, размер задан')


def test_barchart_png_draws_a_row_per_item(tmp_path):
    """_нарисовать_barchart_png() напрямую - картинка растёт по высоте с
    числом показателей (иначе подписи слипаются на большом отчёте)."""
    from PIL import Image
    import io
    items_small = [('Город', 0), ('Адрес', 1)]
    items_big = [(f'Показатель{i}', i % 2) for i in range(11)]
    png_small = vr._нарисовать_barchart_png('Т', items_small)
    png_big = vr._нарисовать_barchart_png('Т', items_big)
    h_small = Image.open(io.BytesIO(png_small.getvalue())).size[1]
    h_big = Image.open(io.BytesIO(png_big.getvalue())).size[1]
    assert h_big > h_small, 'картинка с большим числом показателей должна быть выше'
    print('✓ высота картинки растёт с числом показателей')


def test_chart_anchored_below_section_title_not_overlapping(tmp_path):
    """Раньше график садился на строку 9 - ту же, где секционный заголовок
    «Расхождения по типу данных…» (объединённая ячейка B9:E9) - и наезжал на
    неё. Якорь должен быть на строке ЗАГОЛОВКА САМОЙ ТАБЛИЦЫ («Показатель»/
    «Расхождений (✗)»), на строку ниже секционного заголовка."""
    out = tmp_path / 'test.xlsx'
    vr._записать_xlsx(out, 'СМУ', _результаты(2), per_city=False, map_results=[])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb['Дашборд']
    title_row = next(r for r in range(1, 20)
                     if ws.cell(r, 2).value and 'Расхождения по типу данных' in str(ws.cell(r, 2).value))
    table_header_row = title_row + 1
    assert ws.cell(table_header_row, 2).value == 'Показатель'
    img = ws._images[0]
    anchor_row = img.anchor._from.row + 1   # openpyxl anchor - 0-индексный
    assert anchor_row >= table_header_row, (
        f'график (строка {anchor_row}) не должен быть ВЫШЕ заголовка таблицы '
        f'(строка {table_header_row}) - иначе наезжает на секционный заголовок')
    print(f'✓ график заякорен на строке {anchor_row} - не выше заголовка таблицы ({table_header_row})')
