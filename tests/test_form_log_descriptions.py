"""FORM_LOG_DESCRIPTIONS - пояснения «что и КАК проверяется» к части колонок
log_forms.xlsx (комментарий к заголовку, наведите мышь), по образцу
BLOCK_DESCRIPTIONS в content_checker.py (см. test_every_block_has_description
в tests/test_content_checker.py). В отличие от content_checker, тут НЕ
требуем описание для всех колонок - только защищаем уже написанные от
опечаток и проверяем, что init_excel_log реально их проставляет."""
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "forms_tester"))

t = pytest.importorskip("test_all")


def test_все_ключи_описаний_существуют_в_шапке():
    for header in t.FORM_LOG_DESCRIPTIONS:
        assert header in t.LOG_HEADERS, \
            f"«{header}» есть в FORM_LOG_DESCRIPTIONS, но нет в LOG_HEADERS (опечатка?)"


def test_описания_непустые():
    for header, desc in t.FORM_LOG_DESCRIPTIONS.items():
        assert isinstance(desc, str) and desc.strip(), f"Пустое описание для «{header}»"


def test_описания_обязательны_для_пунктов_этой_доработки():
    # Явный список - если запись из FORM_LOG_DESCRIPTIONS случайно потеряется,
    # тест упадёт (в отличие от простого "если есть - непустое" выше).
    ожидаем = (
        "CSRF-защита",
        "Серверная валидация",
        "Защита от спама (пассивно)",
        "Защита от спама (активно)",
        "Модалка открывается",
        "Модалка закрывается",
        "Подсказки полей",
        "Ошибки валидации",
    )
    for header in ожидаем:
        assert header in t.FORM_LOG_DESCRIPTIONS, f"Нет описания для «{header}»"


def test_init_excel_log_ставит_комментарии_к_заголовкам(tmp_path):
    path = str(tmp_path / "log_forms_test.xlsx")
    t.init_excel_log(path, очистить=True)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["Логи"]
    col_by_header = {ws.cell(row=1, column=c).value: c
                      for c in range(1, ws.max_column + 1)}

    for header, desc in t.FORM_LOG_DESCRIPTIONS.items():
        cell = ws.cell(row=1, column=col_by_header[header])
        assert cell.comment is not None, f"Нет комментария у заголовка «{header}»"
        assert cell.comment.text == desc

    # Контраст: столбец БЕЗ записи в словаре не должен получать комментарий
    # (иначе значит init_excel_log что-то себе выдумывает).
    no_desc = [h for h in t.LOG_HEADERS if h not in t.FORM_LOG_DESCRIPTIONS]
    assert no_desc, "ожидали хотя бы один столбец без описания - для контраста"
    assert ws.cell(row=1, column=col_by_header[no_desc[0]]).comment is None


if __name__ == "__main__":
    import traceback
    fns = [v for k, v in sorted(globals().items())
           if k.startswith("test_") and callable(v)]
    ok = 0
    for fn in fns:
        try:
            if "tmp_path" in fn.__code__.co_varnames[:fn.__code__.co_argcount]:
                import tempfile
                with tempfile.TemporaryDirectory() as d:
                    fn(Path(d))
            else:
                fn()
            print(f"✓ {fn.__name__}")
            ok += 1
        except Exception:
            print(f"✗ {fn.__name__}")
            traceback.print_exc()
    print(f"\n{ok}/{len(fns)} прошло")
    sys.exit(0 if ok == len(fns) else 1)
