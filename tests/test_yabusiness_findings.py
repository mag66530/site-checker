# -*- coding: utf-8 -*-
"""Я.Бизнес и отзывы попадают в «Проблемы».

Раньше эти находки жили ТОЛЬКО на листе «Я.Бизнес и GMB»: ни в «Проблемах»,
ни в «Плане работ» их не было, поэтому в общий список задач они не попадали."""
from report_priorities import (
    classify, group_into_tasks, review_priority_findings, yabusiness_findings,
)


def _тексты(находки):
    return ' | '.join(f.problem for f in находки)


def _yb(**kw):
    основа = {'available': True, 'total_subdomains': 2, 'active_orgs': 1,
              'missing': [], 'matched': [], 'orphan_orgs': []}
    основа.update(kw)
    return основа


# ── Поддомены без карточки ────────────────────────────────────────────

def test_поддомен_без_карточки_это_ошибка():
    находки = yabusiness_findings(_yb(missing=[
        {'url': 'https://ekb.a.ru/', 'city': 'Екатеринбург', 'country': 'Россия'}]))
    assert len(находки) == 1
    f = находки[0]
    assert f.level == 'Ошибка' and f.section == 'Я.Бизнес'
    assert f.city == 'Екатеринбург' and f.url == 'https://ekb.a.ru/'
    assert classify(f)['priority'] == 1


def test_поддомены_с_карточкой_не_находка():
    """✅-список - справка, в «Проблемы» не идёт."""
    assert yabusiness_findings(_yb(matched=[
        {'url': 'https://msk.a.ru/', 'city': 'Москва',
         'org': {'permalink': '1', 'region': 'Москва'}}])) == []


def test_проверка_не_выполнялась_молчим():
    assert yabusiness_findings({'available': False, 'note': 'нет сессии'}) == []
    assert yabusiness_findings(None) == []


# ── Карточки без поддомена ────────────────────────────────────────────

def test_карточка_без_поддомена_это_предупреждение():
    находки = yabusiness_findings(_yb(orphan_orgs=[
        {'permalink': '12345', 'city': 'Сочи', 'region': 'Краснодарский край',
         'addr': 'ул. Ленина, 1'}]))
    f = находки[0]
    assert f.level == 'Предупреждение'
    assert '12345' in f.url                      # ссылка на карточку в картах
    assert 'Краснодарский край' in f.detail and 'ул. Ленина, 1' in f.detail


# ── Сеть ──────────────────────────────────────────────────────────────

def test_филиалы_не_в_сети_одна_строка_на_сайт():
    находки = yabusiness_findings(_yb(chain_check={
        'united': False, 'chains': 1, 'chain_members': 8,
        'standalone_companies': 3,
        'standalone_list': [{'permalink': '1', 'city': 'Тверь'},
                            {'permalink': '2', 'city': 'Тула'}]}))
    assert len(находки) == 1
    f = находки[0]
    assert f.level == 'Ошибка'
    assert 'вне сети: 3' in f.detail and 'Тверь' in f.detail
    assert 'филиалов в них: 8' in f.fix_note


def test_филиалы_в_сети_не_находка():
    assert yabusiness_findings(_yb(chain_check={
        'united': True, 'chains': 1, 'chain_members': 11,
        'standalone_companies': 0, 'standalone_list': []})) == []


# ── Профиль: строка на каждое незаполненное поле ───────────────────────

def test_профиль_строка_на_каждое_поле():
    находки = yabusiness_findings(_yb(profile_check={'all_full': False, 'orgs': [
        {'permalink': '77', 'city': 'Казань', 'filled': 6, 'total': 8,
         'missing': ['фото', 'особенности']}]}))
    assert len(находки) == 2                       # два поля - две строки
    т = _тексты(находки)
    assert 'не заполнено: фото' in т and 'не заполнено: особенности' in т
    assert all(f.city == 'Казань' for f in находки)
    assert 'заполнено 6 из 8' in находки[0].fix_note


def test_полный_профиль_не_находка():
    assert yabusiness_findings(_yb(profile_check={'all_full': True, 'orgs': [
        {'permalink': '77', 'city': 'Казань', 'filled': 8, 'total': 8,
         'missing': []}]})) == []


# ── Отзывы ────────────────────────────────────────────────────────────

def _филиал(**kw):
    основа = {'city': 'Пермь', 'yandex': {'rating': 4.2, 'count': 30},
              'twogis': {'rating': 4.9, 'count': 12}, 'rating': 4.2,
              'low_rating': True, 'negative': False, 'order': 2,
              'yandex_url': 'https://yandex.ru/maps/org/1'}
    основа.update(kw)
    return основа


def test_низкий_рейтинг_предупреждение_с_обоими_источниками():
    находки = review_priority_findings(
        {'available': True, 'branches': [_филиал()]})
    f = находки[0]
    assert f.level == 'Предупреждение' and f.section == 'Отзывы'
    assert '4.2' in f.problem
    assert 'Яндекс 4.2 (30 отз.)' in f.detail and '2ГИС 4.9 (12 отз.)' in f.detail
    assert 'Докупить отзывы: 2' in f.fix_note


def test_нормальный_рейтинг_не_находка():
    assert review_priority_findings({'available': True, 'branches': [
        _филиал(rating=4.9, low_rating=False)]}) == []


def test_критично_низкий_рейтинг_помечен_первым_в_работу():
    f = review_priority_findings({'available': True, 'branches': [
        _филиал(rating=3.4, negative=True, order=3)]})[0]
    assert 'первым' in f.fix_note and 'Докупить отзывы: 3' in f.fix_note


def test_рейтинг_не_получен_отдельная_находка_а_не_низкий_рейтинг():
    """rating=None помечается low_rating, но это «не проверили», а не «плохо»."""
    находки = review_priority_findings({'available': True, 'branches': [
        _филиал(rating=None, unknown=True,
                yandex={'rating': None, 'count': None},
                twogis={'rating': None, 'count': None})]})
    f = находки[0]
    assert 'не удалось получить рейтинг' in f.problem
    assert 'ниже 4.7' not in f.problem
    assert 'yandex_map_url' in f.fix_note           # подсказка, где чинить ссылки
    assert classify(f)['title'] == 'Починить сбор рейтингов филиалов'


def test_филиал_без_единого_отзыва_не_путается_с_низким_рейтингом():
    """Карточка ответила, отзывов 0: рейтинга нет, но это не сбой чтения."""
    f = review_priority_findings({'available': True, 'branches': [
        _филиал(rating=None, no_reviews=True,
                yandex={'rating': None, 'count': 0},
                twogis={'rating': None, 'count': 0})]})[0]
    assert 'нет ни одного отзыва' in f.problem
    assert classify(f)['title'] == 'Набрать первые отзывы филиалам без отзывов'


def test_отзывы_есть_а_рейтинга_нет():
    """Отзывы оставлены без звёзд - рейтинг не считается."""
    f = review_priority_findings({'available': True, 'branches': [
        _филиал(rating=None, no_rating_yet=True,
                yandex={'rating': None, 'count': 0},
                twogis={'rating': None, 'count': 3})]})[0]
    assert 'рейтинг не выставлен' in f.problem
    assert classify(f)['title'] == 'Добрать отзывы с оценкой'


def test_отзывы_не_выполнялись_молчим():
    assert review_priority_findings({'available': False, 'note': 'нет конфига'}) == []


# ── Лист «Я.Бизнес и GMB»: только таблица докупки ─────────────────────

def _лист_отзывов(rp):
    from openpyxl import Workbook

    import reporter
    wb = Workbook()
    wb.remove(wb.active)
    reporter._build_yabusiness_sheet(wb, rp)
    return wb


def test_лист_содержит_только_таблицу_докупки():
    wb = _лист_отзывов({'available': True, 'total_branches': 1,
                        'low_rating_count': 1, 'branches': [_филиал()]})
    ws = wb['Отзывы']
    значения = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    текст = ' '.join(значения)
    assert 'Отзывы: приоритет докупки' in текст
    assert 'Пермь' in текст and 'Докупить' in текст
    # блоков Я.Бизнеса на листе больше нет - они в «Проблемах»
    for ушло in ('Поддомены без организации', 'Поддомены с организацией',
                 'Организации без поддомена', 'объединены в Сеть',
                 'Заполненность профиля'):
        assert ушло not in текст, f'блок «{ушло}» вернулся на лист'


def test_лист_фильтруется_по_городам():
    wb = _лист_отзывов({'available': True, 'total_branches': 1,
                        'low_rating_count': 0, 'branches': [
                            _филиал(rating=4.9, low_rating=False)]})
    assert wb['Отзывы'].auto_filter.ref                # таблица рабочая


def test_листа_нет_без_проверки_отзывов():
    """Данные Я.Бизнеса без отзывов лист не создают - всё уехало в «Проблемы»."""
    from openpyxl import Workbook

    import reporter
    wb = Workbook()
    wb.remove(wb.active)
    reporter._build_yabusiness_sheet(wb, None)
    assert 'Отзывы' not in wb.sheetnames


def test_отзывы_не_выполнялись_лист_объясняет_причину():
    wb = _лист_отзывов({'available': False, 'note': 'нет конфига reviews-sm.csv'})
    значения = [str(c.value) for row in wb['Отзывы'].iter_rows()
                for c in row if c.value]
    assert any('reviews-sm.csv' in v for v in значения)


# ── Задачи «Плана работ» ──────────────────────────────────────────────

def test_находки_группируются_в_задачи_с_понятными_названиями():
    находки = (yabusiness_findings(_yb(
        missing=[{'url': 'https://ekb.a.ru/', 'city': 'Екатеринбург'}],
        profile_check={'all_full': False, 'orgs': [
            {'permalink': '77', 'city': 'Казань', 'filled': 7, 'total': 8,
             'missing': ['фото']}]}))
        + review_priority_findings({'available': True, 'branches': [_филиал()]}))
    названия = {t.title for t in group_into_tasks(находки)}
    assert 'Завести карточки организаций на города поддоменов' in названия
    assert 'Дозаполнить профили организаций' in названия
    assert 'Докупить отзывы филиалам с низким рейтингом' in названия
    # ни одна находка не свалилась в безопасный дефолт
    assert not any(t.title.startswith('Проверить: ') for t in group_into_tasks(находки))
