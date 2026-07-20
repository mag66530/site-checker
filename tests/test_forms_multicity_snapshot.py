"""Страховка мультигорода в forms_run: даже если движок в мультигородном прогоне
теряет строки всех городов кроме последнего (баг в браузерном пути), снимок строк
КАЖДОГО города сразу после его прогона + пересборка лога восстанавливают формы
всех городов. Тест воспроизводит худший случай (файл в конце содержит только
последний город) и проверяет, что снимки его чинят.
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import forms_run


def _make_log(path, hdr, city_rows):
    """Создать log_forms.xlsx с листом «Логи»: hdr + строки одного города."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Логи'
    ws.append(hdr)
    for r in city_rows:
        ws.append(r)
    wb.save(path)


def test_snapshot_recovers_all_cities(tmp_path):
    path = str(tmp_path / 'log_forms.xlsx')
    hdr = ['Дата', 'Город', 'Страница', 'Название', 'Статус']
    cities = ['Москва', 'Алматы', 'Steelgroup']

    # Воспроизводим цикл forms_run С БАГОМ: каждый «прогон города» ПЕРЕЗАТИРАЕТ
    # файл только своими строками (как будто движок теряет прошлые города).
    снимки = []
    шапка = None
    for city in cities:
        _make_log(path, hdr,
                  [['20.07', city, 'Главная', 'Заказ звонка', 'Успешно']])  # только этот город
        h, rows = forms_run._read_logi(path)
        шапка = h
        мои = [r for r in rows if forms_run._row_city(h, r).lower() == city.lower()]
        снимки.append((city, мои))

    # Без пересборки в файле остался бы только последний город:
    _, before = forms_run._read_logi(path)
    assert {forms_run._row_city(hdr, r) for r in before} == {'Steelgroup'}

    # Пересборка из снимков - как в forms_run после цикла.
    forms_run._rewrite_logi(path, шапка, [r for _c, rs in снимки for r in rs])

    _, after = forms_run._read_logi(path)
    городами = {forms_run._row_city(hdr, r) for r in after}
    assert городами == {'Москва', 'Алматы', 'Steelgroup'}   # все три вернулись
    print('✓ снимки по городам восстановили формы всех городов (было: только последний)')


def test_snapshot_no_dup_when_engine_accumulates(tmp_path):
    # Если движок НЕ теряет города (накапливает), снимок каждого города берёт
    # только его строки → в пересборке нет дублей.
    path = str(tmp_path / 'log_forms.xlsx')
    hdr = ['Дата', 'Город', 'Страница', 'Название', 'Статус']
    from openpyxl import Workbook, load_workbook
    cities = ['Москва', 'Алматы']
    снимки = []
    шапка = None
    acc = []
    for city in cities:
        acc.append(['20.07', city, 'Главная', 'Заказ звонка', 'Успешно'])
        wb = Workbook(); ws = wb.active; ws.title = 'Логи'; ws.append(hdr)
        for r in acc:                       # накопление (нормальное поведение)
            ws.append(r)
        wb.save(path)
        h, rows = forms_run._read_logi(path)
        шапка = h
        снимки.append((city, [r for r in rows
                              if forms_run._row_city(h, r).lower() == city.lower()]))
    forms_run._rewrite_logi(path, шапка, [r for _c, rs in снимки for r in rs])
    _, after = forms_run._read_logi(path)
    assert [forms_run._row_city(hdr, r) for r in after] == ['Москва', 'Алматы']  # по 1 на город
    print('✓ при нормальном накоплении дублей нет - по одной строке на город')


if __name__ == '__main__':
    import pytest
    raise SystemExit(pytest.main([__file__, '-v', '-s']))
