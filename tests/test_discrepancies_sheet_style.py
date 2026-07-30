"""Тесты оформления листа «Расхождения»: колонка «Где проверяли» (без «(площадка)»
в заголовке - просьба заказчика) подсвечена своим цветом под каждую площадку,
по образцу присланного файла (там это было условное форматирование по
значению ячейки; мы просто красим сразу при генерации - площадка каждой
строки уже известна)."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import variables_run as vr


def _wb_with_style():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Side
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="EEF3FB")
    thin = Side(style="thin", color="C9CFDB")
    return wb, hdr_fill, thin


def _d(площадка, **kw):
    base = {"domain": "host.ru", "city": "Город", "площадка": площадка,
           "поле": "телефон", "кп": "1", "факт": "2", "примечание": "",
           "ссылка": "https://host.ru/"}
    base.update(kw)
    return base


def test_column_header_renamed_without_suffix():
    wb, hdr_fill, thin = _wb_with_style()
    vr._написать_расхождения_лист(wb, hdr_fill, thin, [_d("Сайт")])
    ws = wb["Расхождения"]
    assert ws.cell(1, 3).value == "Где проверяли"
    print('✓ заголовок колонки - просто «Где проверяли», без «(площадка)»')


def test_площадка_values_are_bold():
    wb, hdr_fill, thin = _wb_with_style()
    vr._написать_расхождения_лист(wb, hdr_fill, thin, [_d("2ГИС"), _d("Что-то новое")])
    ws = wb["Расхождения"]
    assert ws.cell(2, 3).font.bold is True
    assert ws.cell(3, 3).font.bold is True, 'жирным - даже для незнакомой площадки'
    print('✓ значения в «Где проверяли» - жирным, известная площадка и нет')


def test_each_площадка_has_its_own_color():
    wb, hdr_fill, thin = _wb_with_style()
    расхождения = [_d("2ГИС"), _d("Яндекс.Карты"), _d("Google"), _d("Сайт")]
    vr._написать_расхождения_лист(wb, hdr_fill, thin, расхождения)
    ws = wb["Расхождения"]
    colors = {}
    for i, d in enumerate(расхождения, 2):
        cell = ws.cell(i, 3)
        colors[d["площадка"]] = (cell.fill.fgColor.rgb, cell.font.color.rgb)
    # Все 4 цвета должны быть заданы и РАЗНЫМИ (иначе площадки визуально не различить).
    assert all(fg is not None for fg, _ in colors.values())
    assert len({fg for fg, _ in colors.values()}) == 4, "цвета площадок должны отличаться"
    # Значения должны совпадать с образцом (условное форматирование в присланном файле).
    assert colors["2ГИС"] == ("00D9E8FB", "001155CC") or colors["2ГИС"][0].endswith("D9E8FB")
    print('✓ каждая площадка подсвечена своим (разным) цветом')


def test_unknown_площадка_has_no_special_color():
    """На всякий случай - незнакомое значение площадки не должно падать и не
    должно красить ячейку в случайный цвет."""
    wb, hdr_fill, thin = _wb_with_style()
    vr._написать_расхождения_лист(wb, hdr_fill, thin, [_d("Что-то новое")])
    ws = wb["Расхождения"]
    cell = ws.cell(2, 3)
    assert cell.value == "Что-то новое"
    assert cell.fill.fgColor.rgb in (None, "00000000")
    print('✓ незнакомая площадка - без цвета, не падает')
