"""Тесты классификации ячеек матрицы проверок (_матрица_классифицировать) и
её сборки (построить_матрицу_проверок).

Главное, что фиксируем — исправление бага: у колонок «Модалка открывается/
закрывается» не было правил, поэтому «Да» (пройденная проверка!) падало в
дефолт ⚠ «значение не распознано» и выглядело как ручная проверка. Теперь
«Да» → ✓, «Нет» → ✗, «Проверить» → ⚠.

Реальные дефекты (УЯЗВИМА, «не защищена», CSRF «Нет», «ложный успех») остаются ✗.
Новых символов в матрице нет — только ✓/✗/⚠/–.

Без браузера — чистая функция и рендер xlsx (openpyxl)."""
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "forms_tester"))

t = pytest.importorskip("test_all")


def _sym(col, val):
    return t._матрица_классифицировать(col, val)[0]


# ── Регресс бага: у «Модалки» не было правил → «Да» уходило в дефолт ⚠ ──
def test_модалка_закрывается_да_галочка_а_не_предупреждение():
    assert _sym("Модалка закрывается", "Да") == "✓"


def test_модалка_открывается_да_галочка():
    assert _sym("Модалка открывается", "Да") == "✓"


def test_модалка_не_закрылась_ошибка():
    assert _sym("Модалка закрывается", "Нет") == "✗"


def test_модалка_проверить_остаётся_ручной():
    assert _sym("Модалка закрывается", "Проверить") == "⚠"


# ── Реальные дефекты не «размываются» в ✓/⚠ ──
def test_серверная_валидация_уязвима_ошибка():
    assert _sym("Серверная валидация", "УЯЗВИМА") == "✗"


def test_серверная_валидация_защищена_галочка():
    assert _sym("Серверная валидация", "Защищена") == "✓"


def test_двойная_отправка_не_защищена_ошибка():
    assert _sym("Двойная отправка", "не защищена") == "✗"


def test_csrf_нет_ошибка():
    assert _sym("CSRF-защита", "Нет") == "✗"


# ── Кнопка «Отправить»: всегда-активная - знак внимания ⚠ (по требованию) ──
def test_кнопка_всегда_активная_жёлтый_флажок():
    assert _sym("Кнопка по заполнению", "не блокируется") == "⚠"   # активна на пустой
    assert _sym("Кнопка по заполнению", "корректно") == "✓"
    assert _sym("Кнопка по заполнению", "не разблокируется") == "✗"  # реальный баг
    assert _sym("Кнопка по заполнению", "не найдено") == "–"


# ── Обработка ошибок на многошаговом заказе (пункт «Далее») ──
def test_детектор_ошибки_оформления_ловит_незаполненные_поля():
    # Форма пропускает по «Далее», но заказ не оформить: детектор ловит
    # незаполненные обязательные поля на чекауте (не только старые 4 фразы).
    assert t.response_indicates_form_error("Заполните обязательные поля")
    assert t.response_indicates_form_error("Необходимо заполнить обязательные поля")
    assert t.response_indicates_form_error("Заказ не оформлен")
    # Успешная страница ошибкой не считается (нет ложных срабатываний):
    assert t.response_indicates_form_error("Заказ сформирован. Спасибо!") == ""


def test_ошибка_на_шаге_далее_крест():
    assert _sym("Обработка ошибок", "ошибка на шаге «Далее»") == "✗"
    assert _sym("Обработка ошибок", "корректно") == "✓"


def test_csrf_есть_галочка():
    assert _sym("CSRF-защита", "Есть") == "✓"


def test_обработка_ошибок_ложный_успех_ошибка():
    assert _sym("Обработка ошибок", "ложный успех") == "✗"


def test_обработка_ошибок_молчит_больше_не_ошибка():
    # «молчит» на искусственно оборванный запрос больше НЕ помечаем ✗ - это был
    # ложный дефект на рабочих формах (правило убрано, значение не производится).
    assert _sym("Обработка ошибок", "молчит") != "✗"


def test_защита_спама_блокирует_повтор_внимание():
    # Кнопка «Отправлено» после отправки - от случайного двойного клика защищает,
    # но НЕ доказывает серверную защиту (на части форм кнопка косметическая,
    # заявки летят). Поэтому ⚠ «проверьте вручную», а не ложный ✓.
    assert _sym("Защита от спама (активно)", "Блокирует повтор") == "⚠"


def test_пустое_значение_прочерк():
    assert _sym("Серверная валидация", "") == "–"
    assert _sym("Серверная валидация", None) == "–"


# ── «Данные дошли до сервера»: «Да» (пройдено) должно быть ✓, а не ⚠ ──
def test_данные_дошли_да_галочка():
    assert _sym("Данные дошли до сервера", "Да") == "✓"


def test_данные_дошли_нет_ошибка():
    assert _sym("Данные дошли до сервера", "Нет") == "✗"


def test_данные_дошли_не_долетело_настоящий_вердикт_ошибка():
    # РЕАЛЬНАЯ строка из данные_формы_вердикт - «Не долетело» (не «Нет»).
    # В «не долетело» нет подстроки «нет», раньше это уходило в дефолт ⚠
    # (ложная ручная проверка на настоящей потере данных). Теперь → ✗.
    assert t.данные_формы_вердикт(True, "name=X", {"phone": "79161234567"})[0] == "Не долетело"
    assert _sym("Данные дошли до сервера", "Не долетело") == "✗"
    assert _sym("Данные дошли до сервера", "не долетело") == "✗"


def test_данные_дошли_проверить_ручная():
    assert _sym("Данные дошли до сервера", "Проверить") == "⚠"


# ── Умный дефолт: новая колонка без правил не пугает ⚠ на «да/есть» ──
def test_новая_колонка_да_не_предупреждение():
    assert _sym("Совсем новая проверка", "Да") == "✓"
    assert _sym("Совсем новая проверка", "Есть защита") == "✓"


def test_новая_колонка_непонятное_значение_ручная():
    assert _sym("Совсем новая проверка", "абракадабра") == "⚠"


# ── Честный статус: «нет подтверждения» и «сервер отклонил» → ✗, не «успешно» ──
def test_статус_нет_подтверждения_ошибка():
    assert _sym("Статус", "НЕТ ПОДТВЕРЖДЕНИЯ (форма не показала успех)") == "✗"


def test_статус_сервер_отклонил_ошибка():
    assert _sym("Статус", "ОШИБКА (сервер отклонил отправку)") == "✗"


def test_статус_успешно_галочка():
    assert _sym("Статус", "УСПЕШНО (Playwright - как ручная отправка)") == "✓"


def test_нет_подтверждения_распознаётся_как_форма():
    # Статус должен начинаться с одного из submit-префиксов, иначе строка формы
    # выпадет из матрицы. «нет подтверждения» добавлен в _МАТРИЦА_SUBMIT_ST.
    st = "нет подтверждения (форма не показала успех)"
    assert any(st.startswith(s) for s in t._МАТРИЦА_SUBMIT_ST)


# ── В матрице нет новых символов (только ✓/✗/⚠/–) ──
def test_нет_нового_символа_в_правилах():
    все_символы = set()
    for правила in t._МАТРИЦА_ПРАВИЛА.values():
        for _pat, sym, _cm in правила:
            все_символы.add(sym)
    assert все_символы <= {"✓", "✗", "⚠", "–"}, f"лишние символы: {все_символы}"


# ── Сборка матрицы: «Да» модалки реально становится ✓, дефект остаётся ✗ ──
def _минимальный_лог(path):
    from openpyxl import Workbook
    wb = Workbook()
    sv = wb.active
    sv.title = "Сводка"
    ws = wb.create_sheet("Логи")
    ws.append(list(t.LOG_HEADERS))
    row = {h: "" for h in t.LOG_HEADERS}
    row["Дата"] = "14.07.2026"
    row["Город"] = "Москва"
    row["Страница"] = "Товар"
    row["URL"] = "https://example.ru/tovar/"
    row["Название"] = "Купить в один клик"
    row["Где находится"] = "Товар"
    row["Статус"] = "УСПЕШНО (Playwright - как ручная отправка)"
    row["Модалка закрывается"] = "Да"
    row["CSRF-защита"] = "Есть"
    row["Двойная отправка"] = "не защищена"
    row["Защита от спама (активно)"] = "Блокирует повтор"
    # Справочные колонки - должны исчезнуть из матрицы (остаться в «Логи»).
    row["Поля очищены"] = "не очищены"
    row["Защита от спама (пассивно)"] = "Не обнаружено"
    ws.append([row[h] for h in t.LOG_HEADERS])
    wb.save(path)


def test_матрица_рендерит_модалку_галочкой(tmp_path):
    from openpyxl import load_workbook
    path = tmp_path / "log_forms.xlsx"
    _минимальный_лог(path)
    t.построить_матрицу_проверок(str(path))

    wb = load_workbook(path)
    assert "Москва" in wb.sheetnames
    ws = wb["Москва"]
    got = {str(ws.cell(i, 1).value): str(ws.cell(i, 2).value)
           for i in range(2, ws.max_row + 1)}

    assert got.get("Модалка закрывается") == "✓"   # был баг: ⚠
    assert got.get("CSRF-защита") == "✓"           # SameSite-защита
    assert got.get("Двойная отправка") == "✗"       # реальный дефект
    assert got.get("Защита от спама (активно)") == "⚠"  # «блокирует повтор» = кнопка косметическая → проверить

    # Справочные колонки убраны из матрицы (шум ⚠ на каждой форме).
    assert "Поля очищены" not in got
    assert "Защита от спама (пассивно)" not in got

    # В легенде нет посторонних символов.
    sv = wb["Сводка"]
    legend_syms = {str(sv.cell(r, 7).value) for r in range(2, sv.max_row + 1)}
    legend_syms.discard("None")
    assert legend_syms <= {"✓", "✗", "⚠", "–"}


def test_пустые_строки_прочерки_скрыты(tmp_path):
    # Если у ВСЕХ форм листа проверка «–» (нет данных - напр. на сайте нигде нет
    # выпадающих списков/полей загрузки), строка НЕ выводится в матрицу (данные
    # при этом остаются в листе «Логи»). А строки с реальным вердиктом остаются.
    from openpyxl import load_workbook
    path = tmp_path / "log_forms.xlsx"
    _минимальный_лог(path)          # одна форма: без списков/файлов/автозаполнения
    t.построить_матрицу_проверок(str(path))
    wb = load_workbook(str(path))
    ws = wb["Москва"]
    labels = {str(ws.cell(i, 1).value) for i in range(2, ws.max_row + 1)}
    assert "Выпадающие списки" not in labels
    assert "Типы файлов формы" not in labels
    assert "Автозаполнение полей" not in labels
    # реальные вердикты остаются:
    assert "Статус" in labels
    assert "CSRF-защита" in labels
    assert "Двойная отправка" in labels


def test_сшивка_заказа_берёт_вердикт_а_не_заполнено(tmp_path):
    # После сшивки шаг-форма «данные покупателя» (ЗАПОЛНЕНО без отправки) и
    # шаг-проверка «Оформление заказа» (вердикт) идут под ОДНИМ именем. Обе
    # строки «голые» - консолидация должна взять ВЕРДИКТ заказа, а не «ЗАПОЛНЕНО».
    from openpyxl import Workbook, load_workbook
    path = tmp_path / "log_forms.xlsx"
    wb = Workbook(); sv = wb.active; sv.title = "Сводка"
    ws = wb.create_sheet("Логи"); ws.append(list(t.LOG_HEADERS))

    def _row(status, коммент=""):
        r = {h: "" for h in t.LOG_HEADERS}
        r.update({"Город": "Москва", "Страница": "Товар",
                  "Название": "Оформление заказа", "Статус": status,
                  "Комментарий": коммент})
        return [r[h] for h in t.LOG_HEADERS]

    ws.append(_row("ЗАПОЛНЕНО (без отправки)"))              # шаг-форма (первый)
    ws.append(_row("ОШИБКА (не выбран тип плательщика)",
                   "Оформление: не выбран тип плательщика"))  # вердикт заказа
    wb.save(str(path))
    t.консолидировать_форм_строки(str(path))

    ws2 = load_workbook(str(path))["Логи"]
    assert ws2.max_row == 2, "две строки должны слиться в одну"
    hdr = [c.value for c in ws2[1]]
    si = hdr.index("Статус")
    assert str(ws2.cell(2, si + 1).value).startswith("ОШИБКА")


def test_заметки_не_ужимаются_до_ячейки(tmp_path):
    # openpyxl пишет <x:SizeWithCells/>, из-за чего Excel обрезает длинный
    # комментарий до узкой ячейки. Матрица должна убрать этот флаг (оставив
    # MoveWithCells), иначе комментарии «не влезают».
    import zipfile
    from openpyxl import load_workbook
    path = tmp_path / "log_forms.xlsx"
    _минимальный_лог(path)
    t.построить_матрицу_проверок(str(path))

    with zipfile.ZipFile(str(path)) as z:
        vml = [n for n in z.namelist() if n.lower().endswith(".vml")]
        assert vml, "в файле нет VML-заметок"
        swc = sum(z.read(n).count(b"SizeWithCells") for n in vml)
        mwc = sum(z.read(n).count(b"MoveWithCells") for n in vml)
    assert swc == 0, "SizeWithCells не убран - Excel будет обрезать комментарии"
    assert mwc > 0, "MoveWithCells должен остаться"
    # файл всё ещё открывается и комментарий на месте
    load_workbook(str(path))


def test_матрица_заметки_щедрые_по_размеру(tmp_path):
    """Окна заметок должны быть достаточно БОЛЬШИМИ, чтобы текст влезал целиком -
    раньше в отчёте они были дефолтными (144x79) и текст обрезался. Проверяем по
    VML: ширина фигуры-заметки >= 300px, высота >= 110px."""
    import re
    import zipfile
    path = tmp_path / "log_forms.xlsx"
    _минимальный_лог(path)
    t.построить_матрицу_проверок(str(path))
    widths, heights = [], []
    with zipfile.ZipFile(str(path)) as z:
        for n in z.namelist():
            if n.lower().endswith(".vml"):
                vml = z.read(n).decode("utf-8", "replace")
                widths += [int(m) for m in re.findall(r"width:(\d+)px", vml)]
                heights += [int(m) for m in re.findall(r"height:(\d+)px", vml)]
    assert widths, "в матрице нет заметок с размером в VML"
    assert max(widths) >= 300, f"окна заметок узкие: {sorted(set(widths))}"
    assert max(heights) >= 110, f"окна заметок низкие: {sorted(set(heights))}"


def _vml_stats(path):
    import re
    import zipfile
    widths = []
    sw = 0
    with zipfile.ZipFile(str(path)) as z:
        for n in z.namelist():
            if n.lower().endswith(".vml"):
                d = z.read(n).decode("utf-8", "replace")
                widths += [int(m) for m in re.findall(r"width:(\d+)px", d)]
                sw += d.count("SizeWithCells")
    return widths, sw


def test_порядок_как_в_forms_run_заметки_целые(tmp_path):
    """Регресс бага «заметки снова 144x79»: в forms_run ПОСЛЕ матрицы шло ещё
    одно load+save (подгонка ширины «Комментарий»), а openpyxl при перечитывании
    сбрасывает размер окон заметок на дефолт и возвращает SizeWithCells. Теперь
    матрица строится ПОСЛЕДНЕЙ. Повторяем этот порядок и проверяем, что окна
    заметок остались крупными."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    path = tmp_path / "log_forms.xlsx"
    _минимальный_лог(path)
    t.консолидировать_форм_строки(str(path))
    # подгонка ширины колонки «Комментарий» - ДО матрицы (как в forms_run)
    wb = load_workbook(str(path))
    ws = wb["Логи"]
    hh = [str(c.value or "").lower() for c in ws[1]]
    if "комментарий" in hh:
        ws.column_dimensions[get_column_letter(hh.index("комментарий") + 1)].width = 120
        wb.save(str(path))
    t.построить_матрицу_проверок(str(path))        # ПОСЛЕДНЕЙ
    widths, sw = _vml_stats(path)
    assert widths and max(widths) >= 300, f"заметки сбились: {sorted(set(widths))}"
    assert sw == 0, "SizeWithCells вернулся - заметки будут ужиматься"


def test_load_save_после_матрицы_ломает_заметки(tmp_path):
    """Документируем ПРИЧИНУ: любой load+save ПОСЛЕ матрицы возвращает дефолтный
    размер заметок и SizeWithCells - поэтому матрица должна быть последней."""
    from openpyxl import load_workbook
    path = tmp_path / "log_forms.xlsx"
    _минимальный_лог(path)
    t.построить_матрицу_проверок(str(path))
    w1, sw1 = _vml_stats(path)
    assert max(w1) >= 300 and sw1 == 0            # сразу после матрицы - крупные
    load_workbook(str(path)).save(str(path))       # повторное сохранение
    w2, sw2 = _vml_stats(path)
    assert max(w2) < 200 and sw2 > 0               # сбилось - вот почему матрица последняя


# ── Пер-ячеечные КОНКРЕТНЫЕ причины (двойная отправка / серверная валидация) ──
def _лог_с_формой(path, поля):
    """Мини-лог с ОДНОЙ формой (Москва/Товар/«Купить в один клик») + заданные
    значения проверок. Для тестов пер-ячеечных пояснений и инфо-подсказок."""
    from openpyxl import Workbook
    wb = Workbook(); sv = wb.active; sv.title = "Сводка"
    ws = wb.create_sheet("Логи"); ws.append(list(t.LOG_HEADERS))
    row = {h: "" for h in t.LOG_HEADERS}
    row.update({"Дата": "21.07.2026", "Город": "Москва", "Страница": "Товар",
                "URL": "https://example.ru/tovar/",
                "Название": "Купить в один клик", "Где находится": "Товар",
                "Статус": "УСПЕШНО (Playwright - как ручная отправка)"})
    row.update(поля)
    ws.append([row[h] for h in t.LOG_HEADERS])
    wb.save(str(path))


def _коммент_ячейки(ws, label, col=2):
    """Текст всплывающей заметки первой формы (кол.2) в строке-проверке label."""
    for i in range(2, ws.max_row + 1):
        if str(ws.cell(i, 1).value) == label:
            c = ws.cell(i, col).comment
            return c.text if c else None
    return None


def test_ключ_детали_стабилен_и_различает():
    k = t._матрица_ключ_детали
    # обрезка пробелов - запись и чтение дают один ключ
    assert k("Москва", "Товар", "Форма", "Двойная отправка") == \
           k(" Москва ", "Товар", "Форма ", "Двойная отправка")
    # разные колонки / формы - разные ключи
    assert k("Москва", "Товар", "Форма", "Двойная отправка") != \
           k("Москва", "Товар", "Форма", "Серверная валидация")
    assert k("Москва", "Товар", "A", "X") != k("Москва", "Товар", "B", "X")


def test_запись_чтение_детали_роундтрип(tmp_path):
    old = t.MATRIX_DETAILS_FILE
    t.MATRIX_DETAILS_FILE = str(tmp_path / "matrix_details.json")
    try:
        t.record_matrix_detail("Москва", "Товар", "Форма",
                               "Двойная отправка", "кнопка перекрыта")
        t.record_matrix_detail("Москва", "Товар", "Форма",
                               "Серверная валидация", "без почты (поле «email»)")
        d = t.load_matrix_details()
        assert d[t._матрица_ключ_детали("Москва", "Товар", "Форма",
                                        "Двойная отправка")] == "кнопка перекрыта"
        assert d[t._матрица_ключ_детали("Москва", "Товар", "Форма",
                 "Серверная валидация")] == "без почты (поле «email»)"
        # пустую деталь НЕ пишем и НЕ затираем прежнюю
        t.record_matrix_detail("Москва", "Товар", "Форма", "Двойная отправка", "   ")
        d2 = t.load_matrix_details()
        assert d2[t._матрица_ключ_детали("Москва", "Товар", "Форма",
                                         "Двойная отправка")] == "кнопка перекрыта"
    finally:
        t.MATRIX_DETAILS_FILE = old


def test_матрица_показывает_конкретную_причину_двойной_отправки(tmp_path):
    # «под вопросом» ⚠: в ячейке должна всплывать ПОЧЕМУ именно (перекрыта кнопка),
    # а не общая фраза «проверьте вручную».
    from openpyxl import load_workbook
    old = t.MATRIX_DETAILS_FILE
    t.MATRIX_DETAILS_FILE = str(tmp_path / "matrix_details.json")
    try:
        path = tmp_path / "log_forms.xlsx"
        _лог_с_формой(path, {"Двойная отправка": "под вопросом"})
        причина = ("Кнопка перекрыта другим элементом - чистую двойную отправку "
                   "автоматически не проверить.")
        t.record_matrix_detail("Москва", "Товар", "Купить в один клик",
                               "Двойная отправка", причина)
        t.построить_матрицу_проверок(str(path))
        ws = load_workbook(str(path))["Москва"]
        assert str([ws.cell(i, 2).value for i in range(2, ws.max_row + 1)
                    if str(ws.cell(i, 1).value) == "Двойная отправка"][0]) == "⚠"
        got = _коммент_ячейки(ws, "Двойная отправка")
        assert got and "перекрыта" in got                       # КОНКРЕТНАЯ причина
        assert "Не удалось однозначно проверить" not in got     # общая фраза заменена
    finally:
        t.MATRIX_DETAILS_FILE = old


def test_матрица_показывает_без_чего_серверная_валидация(tmp_path):
    # УЯЗВИМА ✗: в ячейке - БЕЗ ЧЕГО именно ушло (без почты, с именем поля),
    # а не общая «принял заведомо невалидные данные».
    from openpyxl import load_workbook
    old = t.MATRIX_DETAILS_FILE
    t.MATRIX_DETAILS_FILE = str(tmp_path / "matrix_details.json")
    try:
        path = tmp_path / "log_forms.xlsx"
        _лог_с_формой(path, {"Серверная валидация": "УЯЗВИМА"})
        деталь = ("сервер ПРИНЯЛ невалидную отправку без почты (поле «email») - "
                  "это и есть уязвимость")
        t.record_matrix_detail("Москва", "Товар", "Купить в один клик",
                               "Серверная валидация", деталь)
        t.построить_матрицу_проверок(str(path))
        ws = load_workbook(str(path))["Москва"]
        assert str([ws.cell(i, 2).value for i in range(2, ws.max_row + 1)
                    if str(ws.cell(i, 1).value) == "Серверная валидация"][0]) == "✗"
        got = _коммент_ячейки(ws, "Серверная валидация")
        assert got and "без почты" in got and "поле «email»" in got
        assert "заведомо невалидные данные" not in got          # общая фраза заменена
    finally:
        t.MATRIX_DETAILS_FILE = old


def test_нет_детали_остаётся_общая_фраза_правила(tmp_path):
    # Если проба деталь не записала (файла нет) - показываем общий текст правила,
    # а не пусто. Это гарантирует обратную совместимость со старыми прогонами.
    from openpyxl import load_workbook
    old = t.MATRIX_DETAILS_FILE
    t.MATRIX_DETAILS_FILE = str(tmp_path / "нет_такого_файла.json")
    try:
        path = tmp_path / "log_forms.xlsx"
        _лог_с_формой(path, {"Двойная отправка": "под вопросом"})
        t.построить_матрицу_проверок(str(path))
        ws = load_workbook(str(path))["Москва"]
        got = _коммент_ячейки(ws, "Двойная отправка")
        assert got and "Не удалось однозначно проверить" in got  # общая фраза правила
    finally:
        t.MATRIX_DETAILS_FILE = old


def test_матрица_инфо_подсказка_на_названии_проверки_а_не_на_галочке(tmp_path):
    # R3 (уточнение): инфо-заметка теперь на НАЗВАНИИ проверки (первый столбец),
    # а НЕ на каждой зелёной галочке. Символ ✓ остаётся, но заметки на ✓-ячейке нет.
    from openpyxl import load_workbook
    old = t.MATRIX_DETAILS_FILE
    t.MATRIX_DETAILS_FILE = str(tmp_path / "matrix_details.json")   # деталей нет
    try:
        path = tmp_path / "log_forms.xlsx"
        _лог_с_формой(path, {"Уведомление пользователю": "Да", "CSRF-защита": "Есть"})
        t.построить_матрицу_проверок(str(path))
        ws = load_workbook(str(path))["Москва"]
        # символ - зелёная галочка
        assert str([ws.cell(i, 2).value for i in range(2, ws.max_row + 1)
                    if str(ws.cell(i, 1).value) == "Уведомление пользователю"][0]) == "✓"
        # инфо-заметка на НАЗВАНИИ проверки (первый столбец, col=1)
        info_label = _коммент_ячейки(ws, "Уведомление пользователю", col=1)
        assert info_label and "подтверждение" in info_label
        # на ✓-ЯЧЕЙКЕ (col=2) заметки БОЛЬШЕ НЕТ
        assert _коммент_ячейки(ws, "Уведомление пользователю", col=2) is None
        # у CSRF - инфо тоже на названии
        assert (_коммент_ячейки(ws, "CSRF-защита", col=1) or "").strip()
        assert _коммент_ячейки(ws, "CSRF-защита", col=2) is None
    finally:
        t.MATRIX_DETAILS_FILE = old


def test_инфо_подсказки_покрывают_основные_проверки():
    # Ключи инфо-словаря должны быть реальными колонками правил (не опечатка).
    for col in t._МАТРИЦА_ИНФО:
        assert col in t._МАТРИЦА_IDENT or col in t._МАТРИЦА_ПРАВИЛА \
            or col in t.LOG_HEADERS, f"инфо для неизвестной колонки: {col}"
    # флагманский пример пользователя присутствует
    assert "Уведомление пользователю" in t._МАТРИЦА_ИНФО


# ── Согласие: 0 чекбоксов не должно выглядеть «всё ок» ──
def test_наличие_чекбоксов_0_конкретный_коммент():
    # «0» → ✗ с КОНКРЕТНЫМ пояснением, что галочек согласия нет ВООБЩЕ.
    sym, cm = t._матрица_классифицировать("Наличие чек боксов согласия", "0 (нужно ≥2)")
    assert sym == "✗"
    assert "НЕТ ни одного" in cm and "персональных данных" in cm
    # «1» → ✗, но текст другой (одна есть, нужно ≥2)
    sym1, cm1 = t._матрица_классифицировать("Наличие чек боксов согласия", "1 (нужно ≥2)")
    assert sym1 == "✗" and "только один" in cm1
    # «2» → ✓
    assert t._матрица_классифицировать("Наличие чек боксов согласия", "2 (нужно ≥2)")[0] == "✓"


def test_предустановлены_пусто_даёт_прочерк():
    # При 0 чекбоксах источник пишет "" в «не предустановлены» → в матрице «–»
    # (не ✓): проверять предустановку не на чем, зелёная галочка вводила в
    # заблуждение.
    assert t._матрица_классифицировать("Чек боксы согласия не предустановлены", "")[0] == "–"
    # обычные значения по-прежнему работают
    assert t._матрица_классифицировать("Чек боксы согласия не предустановлены", "да")[0] == "✓"
    assert t._матрица_классифицировать("Чек боксы согласия не предустановлены", "НЕТ - стоят по умолчанию")[0] == "✗"


def test_без_согласия_0_чекбоксов_пишет_что_галочки_нет(tmp_path):
    # «Без согласия не отправить» ✗ + на форме 0 чекбоксов → в ячейке КОНКРЕТНО:
    # галочки согласия нет вообще, отметить нечего (а не «забыли галочку»).
    from openpyxl import load_workbook
    old = t.MATRIX_DETAILS_FILE
    t.MATRIX_DETAILS_FILE = str(tmp_path / "matrix_details.json")
    try:
        path = tmp_path / "log_forms.xlsx"
        _лог_с_формой(path, {"Без согласия не отправить": "нет",
                             "Наличие чек боксов согласия": "0 (нужно ≥2)"})
        t.построить_матрицу_проверок(str(path))
        ws = load_workbook(str(path))["Москва"]
        assert str([ws.cell(i, 2).value for i in range(2, ws.max_row + 1)
                    if str(ws.cell(i, 1).value) == "Без согласия не отправить"][0]) == "✗"
        got = _коммент_ячейки(ws, "Без согласия не отправить")
        assert got and "нет галочки согласия" in got.lower()
        assert "отметить нечего" in got
    finally:
        t.MATRIX_DETAILS_FILE = old


if __name__ == "__main__":
    import traceback
    import inspect
    import tempfile
    fns = [v for k, v in sorted(globals().items())
           if k.startswith("test_") and callable(v)]
    ok = 0
    for fn in fns:
        try:
            if "tmp_path" in inspect.signature(fn).parameters:
                with tempfile.TemporaryDirectory() as d:
                    fn(Path(d))
            else:
                fn()
            print(f"✓ {fn.__name__}"); ok += 1
        except Exception:
            print(f"✗ {fn.__name__}"); traceback.print_exc()
    print(f"\n{ok}/{len(fns)} прошло")
    sys.exit(0 if ok == len(fns) else 1)
