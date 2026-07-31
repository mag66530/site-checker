"""Тест: лист «Проверка КП» (+ «Расхождения») не создаётся, если ВООБЩЕ
нечего показать (ни сайт, ни карты не проверяли) - раньше пустая вкладка с
одной шапкой всё равно появлялась. Если проверяли ХОТЯ БЫ карты (даже без
сайта, --no-check-site) - лист всё равно нужен, чтобы показать карты."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import variables_run as vr
from maps_compare import MapCheckResult


def test_no_sheet_when_nothing_checked(tmp_path):
    out = tmp_path / 'test.xlsx'
    vr._записать_xlsx(out, 'СМУ', [], per_city=False, map_results=[])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert 'Проверка КП' not in wb.sheetnames
    assert 'Расхождения' not in wb.sheetnames
    print('✓ ни сайт, ни карты не проверяли → ни один из листов не создаётся')


def test_sheet_created_when_results_present(tmp_path):
    out = tmp_path / 'test.xlsx'
    результаты = [{'domain': 'stalmetural.ru', 'city': 'Москва', 'country': 'Россия',
                  'error': '', 'fields': []}]
    vr._записать_xlsx(out, 'СМУ', результаты, per_city=False, map_results=[])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert 'Проверка КП' in wb.sheetnames
    assert 'Расхождения' in wb.sheetnames
    print('✓ результаты непустые → «Проверка КП» и «Расхождения» создаются как раньше')


def test_sheet_created_from_maps_only_when_site_not_checked(tmp_path):
    """--no-check-site: результаты=[], но карты проверяли - лист всё равно
    нужен, чтобы показать карты (просьба заказчика: карты - в том же листе)."""
    out = tmp_path / 'test.xlsx'
    r = MapCheckResult(source='yandex', city='Казань', url='u', available=True,
                       country='Россия', phone_match=True, address_match=True,
                       site_match=True)
    vr._записать_xlsx(out, 'СМУ', [], per_city=False, map_results=[r])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert 'Проверка КП' in wb.sheetnames
    assert 'Расхождения' in wb.sheetnames
    print('✓ только карты (сайт не проверяли) → лист всё равно создаётся')
