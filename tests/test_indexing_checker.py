"""Тесты indexing_checker - «Соблюдение директив вживую» (пункт 1.7):
чистые функции без сети. Живой запрос (_fetch_for_directive_check) юнит-
тестом не покрывается - только смоук-тестом на реальном HTTP (см.
scratchpad прошлых сессий)."""
from indexing_checker import _sample_disallowed_by_rule, _directive_compliance_verdict


# ── _sample_disallowed_by_rule ──────────────────────────────────────────

def test_несколько_путей_с_одним_правилом_один_в_выборке():
    disallowed = [
        {'path': '/catalog/a/?PAGEN_1=2', 'rule': '*PAGEN_1*', 'agent': '*'},
        {'path': '/catalog/b/?PAGEN_1=3', 'rule': '*PAGEN_1*', 'agent': '*'},
        {'path': '/catalog/c/?PAGEN_1=4', 'rule': '*PAGEN_1*', 'agent': '*'},
    ]
    sample = _sample_disallowed_by_rule(disallowed)
    assert len(sample) == 1
    assert sample[0]['path'] == '/catalog/a/?PAGEN_1=2'


def test_разные_правила_все_представлены():
    disallowed = [
        {'path': '/a/', 'rule': '/a/', 'agent': '*'},
        {'path': '/b/', 'rule': '/b/', 'agent': '*'},
        {'path': '/c/', 'rule': '/c/', 'agent': 'yandex'},
    ]
    sample = _sample_disallowed_by_rule(disallowed)
    assert {s['rule'] for s in sample} == {'/a/', '/b/', '/c/'}


def test_обрезка_по_лимиту():
    disallowed = [{'path': f'/{i}/', 'rule': f'/{i}/', 'agent': '*'} for i in range(50)]
    sample = _sample_disallowed_by_rule(disallowed, limit=5)
    assert len(sample) == 5


def test_пустой_список_пустая_выборка():
    assert _sample_disallowed_by_rule([]) == []
    assert _sample_disallowed_by_rule(None) == []


# ── _directive_compliance_verdict ───────────────────────────────────────

def test_недоступна_напрямую_ok():
    assert _directive_compliance_verdict(404, False) == 'ok'
    assert _directive_compliance_verdict(301, False) == 'ok'
    assert _directive_compliance_verdict(500, False) == 'ok'
    assert _directive_compliance_verdict(None, False) == 'ok'


def test_отвечает_200_с_noindex_protected():
    assert _directive_compliance_verdict(200, True) == 'protected'


def test_отвечает_200_без_noindex_robots_only():
    # Главный кейс находки: страница реально доступна и держится ТОЛЬКО на
    # честном слове robots.txt, без собственной подстраховки noindex'ом.
    assert _directive_compliance_verdict(200, False) == 'robots_only'


# ── Секция 3а отчёта (reporter.py) не падает на новой структуре данных ──

def test_секция_отчёта_рендерится_с_находками_и_без():
    from openpyxl import Workbook
    from reporter import _build_indexing_sheet

    summary_с_находкой = {
        'host': 'example.ru', 'disallowed': [],
        'directive_check': {'checked': 3, 'findings': [
            {'rule': '/starye-tovary/', 'path': '/starye-tovary/', 'status': 200},
        ]},
    }
    wb = Workbook()
    _build_indexing_sheet(wb, [], summary_с_находкой)
    ws = wb['Индексация']
    text = ' '.join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert 'Заблокированные страницы' in text
    assert '/starye-tovary/' in text

    summary_без_находок = {
        'host': 'example.ru', 'disallowed': [],
        'directive_check': {'checked': 3, 'findings': []},
    }
    wb2 = Workbook()
    _build_indexing_sheet(wb2, [], summary_без_находок)
    ws2 = wb2['Индексация']
    text2 = ' '.join(str(c.value) for row in ws2.iter_rows() for c in row if c.value)
    assert 'Заблокированные страницы' in text2
    assert 'либо недоступны напрямую' in text2


if __name__ == "__main__":
    import traceback
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    ok = 0
    for fn in fns:
        try:
            fn(); print(f"✓ {fn.__name__}"); ok += 1
        except Exception:
            print(f"✗ {fn.__name__}"); traceback.print_exc()
    print(f"\n{ok}/{len(fns)} прошло")
    import sys
    sys.exit(0 if ok == len(fns) else 1)
