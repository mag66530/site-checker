"""Тесты карт в отчёте variables_run.py: отдельного листа «Карты» больше нет
(вернули по просьбе заказчика) - карты идут ДОПОЛНИТЕЛЬНЫМИ колонками на
листе «Проверка КП», той же строкой, что и сайт, в том же оформлении (символ
✓/✗/⚠/– цветным шрифтом, заливка+комментарий только на несовпадении). Лист
«Расхождения» тоже вернули - расхождения по картам туда тоже попадают, плюс
служебная колонка «Ошибок (✗)» (формула COUNTIF по строке) - дашборд на ней
считает «строк без единого расхождения» / «строк с 3+ расхождениями»."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import variables_run as vr
from maps_compare import MapCheckResult


def _wb_with_style():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Side
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="EEF3FB")
    thin = Side(style="thin", color="C9CFDB")
    return wb, hdr_fill, thin


def _site_res(city='Москва', domain='stalmetural.ru', country='Россия'):
    return {'domain': domain, 'city': city, 'country': country, 'error': '', 'fields': []}


def test_no_separate_maps_sheet():
    wb, hdr_fill, thin = _wb_with_style()
    r = MapCheckResult(source='yandex', city='Москва', url='u', available=True,
                       country='Россия', name='X', phone_match=True,
                       address_match=True, site_match=True)
    vr._записать_проверку_кп_лист(wb, hdr_fill, thin, [_site_res()], False, [r])
    assert 'Карты' not in wb.sheetnames
    assert 'Проверка КП' in wb.sheetnames
    print('✓ отдельного листа «Карты» больше нет - всё на «Проверке КП»')


def test_map_columns_appended_after_site_columns():
    wb, hdr_fill, thin = _wb_with_style()
    results = [
        MapCheckResult(source='yandex', city='Москва', url='u', available=True,
                       country='Россия', phone_match=True, address_match=True,
                       site_match=True),
        MapCheckResult(source='2gis', city='Москва', url='u', available=True,
                       country='Россия', phone_match=True, address_match=True,
                       site_match=True),
    ]
    vr._записать_проверку_кп_лист(wb, hdr_fill, thin, [_site_res()], False, results)
    ws = wb['Проверка КП']
    header = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    # 2 ID + 8 сайт-переменных (не per_city) + 2 карты (Яндекс/2ГИС, Google не
    # проверяли) + служебная «Ошибок (✗)» в самом конце.
    assert header[-3:-1] == ['Яндекс.Карты', '2ГИС']
    assert header[-1] == 'Ошибок (✗)'
    assert len(header) == 13
    print('✓ карты - дополнительные колонки после сайта, «Ошибок (✗)» - последней')


def test_map_mismatch_same_row_as_site_and_goes_to_discrepancies():
    wb, hdr_fill, thin = _wb_with_style()
    r = MapCheckResult(source='yandex', city='Москва', url='https://yandex.ru/x',
                       available=True, country='Россия', name='Стальметурал',
                       phone_match=False, address_match=True, site_match=True,
                       issues=['телефон на карте (+7 000) не совпал с КП'],
                       details=[{'field': 'телефон', 'kp': '4991303669', 'card': '+7 000'}])
    расхождения, _meta = vr._записать_проверку_кп_лист(
        wb, hdr_fill, thin, [_site_res()], False, [r])
    ws = wb['Проверка КП']
    row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    assert row[0] == 'Россия' and row[1] == 'Москва'
    assert row[-2] == '✗'   # предпоследняя колонка - Яндекс.Карты (последняя - «Ошибок»)
    assert len(расхождения) == 1
    d = расхождения[0]
    assert d['city'] == 'Москва'
    assert d['площадка'] == 'Яндекс.Карты' and d['поле'] == 'телефон'
    assert d['кп'] == '4991303669' and d['факт'] == '+7 000'
    assert d['ссылка'] == 'https://yandex.ru/x'
    print('✓ карта того же города - та же строка, расхождение ушло в общий список')


def test_map_only_city_without_site_check_gets_own_row():
    """Сайт не проверяли (--no-check-site) - результаты=[], но карты есть.
    Город из карт всё равно должен появиться строкой (колонки сайта - «–»)."""
    wb, hdr_fill, thin = _wb_with_style()
    r = MapCheckResult(source='yandex', city='Казань', url='u', available=True,
                       country='Россия', phone_match=True, address_match=True,
                       site_match=True)
    vr._записать_проверку_кп_лист(wb, hdr_fill, thin, [], False, [r])
    ws = wb['Проверка КП']
    row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    assert row[1] == 'Казань'
    assert row[-2] == '✓'
    print('✓ город без сайт-проверки, но с картой - своя строка на листе')


def test_ok_map_cell_has_no_comment():
    wb, hdr_fill, thin = _wb_with_style()
    r = MapCheckResult(source='yandex', city='Москва', url='u', available=True,
                       country='Россия', phone_match=True, address_match=True,
                       site_match=True)
    vr._записать_проверку_кп_лист(wb, hdr_fill, thin, [_site_res()], False, [r])
    ws = wb['Проверка КП']
    cell = list(ws.iter_rows(min_row=3, max_row=3))[0][-2]
    assert cell.value == '✓'
    assert cell.comment is None
    print('✓ карта совпала: без комментария, как и у сайта')


def test_unavailable_card_is_warning_not_in_discrepancies():
    wb, hdr_fill, thin = _wb_with_style()
    r = MapCheckResult(source='2gis', city='Уфа', url='u', available=False,
                       country='Россия', error='таймаут')
    расхождения, _meta = vr._записать_проверку_кп_лист(
        wb, hdr_fill, thin, [_site_res(city='Уфа')], False, [r])
    ws = wb['Проверка КП']
    cell = list(ws.iter_rows(min_row=3, max_row=3))[0][-2]
    assert cell.value == '⚠'
    assert 'таймаут' in cell.comment.text
    assert расхождения == [], 'недоступность карточки - не расхождение по значению'
    print('✓ недоступная карточка: ⚠ + причина, но не попадает в «Расхождения»')


def test_no_link_shown_as_dash_without_fill_or_comment():
    wb, hdr_fill, thin = _wb_with_style()
    r = MapCheckResult(source='2gis', city='Брянск', url='', available=False,
                       country='Россия', no_link=True)
    vr._записать_проверку_кп_лист(wb, hdr_fill, thin, [_site_res(city='Брянск')], False, [r])
    ws = wb['Проверка КП']
    cell = list(ws.iter_rows(min_row=3, max_row=3))[0][-2]
    assert cell.value == '–'
    assert cell.comment is None
    print('✓ нет ссылки на карту → «–» без заливки и комментария')


def test_source_not_run_is_blank_not_dash():
    """2ГИС не включали в этот прогон вообще (в map_results для этого города
    вообще нет записи с source='2gis') - раз колонки «2ГИС» тогда и не будет,
    проверяем на примере: два источника были в ДРУГОМ городе, а у этого только
    Яндекс - колонка 2ГИС для него должна быть пустой, не «–»."""
    wb, hdr_fill, thin = _wb_with_style()
    results = [
        MapCheckResult(source='yandex', city='Москва', url='u', available=True,
                       country='Россия', phone_match=True, address_match=True,
                       site_match=True),
        MapCheckResult(source='2gis', city='Казань', url='u', available=True,
                       country='Россия', phone_match=True, address_match=True,
                       site_match=True),
    ]
    vr._записать_проверку_кп_лист(
        wb, hdr_fill, thin, [_site_res(), _site_res(city='Казань')], False, results)
    ws = wb['Проверка КП']
    rows = list(ws.iter_rows(min_row=3, max_row=4, values_only=True))
    moscow = next(row for row in rows if row[1] == 'Москва')
    assert moscow[-2] is None, '2ГИС для Москвы не проверяли - пусто, не «–»'
    print('✓ источник, которого нет у этого города в map_results - пустая ячейка')


def test_all_three_map_sources_columns_in_order():
    """Яндекс.Карты + Google + 2ГИС включены разом - колонки в порядке
    _MAP_SOURCES (Яндекс, Google, 2ГИС), не в порядке появления в результатах."""
    wb, hdr_fill, thin = _wb_with_style()
    results = [
        MapCheckResult(source='2gis', city='Москва', url='u', available=True,
                       country='Россия', phone_match=True, address_match=True,
                       site_match=True),
        MapCheckResult(source='google', city='Москва', url='u', available=True,
                       country='Россия', phone_match=True, address_match=True,
                       site_match=True),
        MapCheckResult(source='yandex', city='Москва', url='u', available=True,
                       country='Россия', phone_match=True, address_match=True,
                       site_match=True),
    ]
    vr._записать_проверку_кп_лист(wb, hdr_fill, thin, [_site_res()], False, results)
    ws = wb['Проверка КП']
    header = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert header[-4:-1] == ['Яндекс.Карты', 'Google', '2ГИС']
    print('✓ все три источника карт - колонки в фиксированном порядке')


def test_no_maps_no_extra_columns():
    """Без карт (map_results=[]) - сетка как раньше + «Ошибок (✗)» в конце."""
    wb, hdr_fill, thin = _wb_with_style()
    vr._записать_проверку_кп_лист(wb, hdr_fill, thin, [_site_res()], False, [])
    ws = wb['Проверка КП']
    header = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert len(header) == 11   # 2 ID + 8 сайт-переменных + «Ошибок (✗)»
    assert header[-1] == 'Ошибок (✗)'
    print('✓ без карт - колонок как в исходном отчёте + служебная «Ошибок (✗)»')


def test_error_count_formula_present():
    """«Ошибок (✗)» - формула COUNTIF по всей строке данных (не число из
    Python) - должна пересчитываться в Excel, если строку поправят руками."""
    wb, hdr_fill, thin = _wb_with_style()
    vr._записать_проверку_кп_лист(wb, hdr_fill, thin, [_site_res()], False, [])
    ws = wb['Проверка КП']
    cell = list(ws.iter_rows(min_row=3, max_row=3))[0][-1]
    assert isinstance(cell.value, str) and cell.value.startswith('=COUNTIF(')
    assert '"✗"' in cell.value
    print('✓ «Ошибок (✗)» - формула COUNTIF, не статичное число')


def test_meta_reports_var_cols_and_sources():
    wb, hdr_fill, thin = _wb_with_style()
    r = MapCheckResult(source='2gis', city='Москва', url='u', available=True,
                       country='Россия', phone_match=True, address_match=True,
                       site_match=True)
    _расхождения, meta = vr._записать_проверку_кп_лист(
        wb, hdr_fill, thin, [_site_res()], False, [r])
    assert meta['last_row'] == 3
    assert ('2ГИС', meta['err_col_letter']) not in meta['var_cols']  # err_col - не var_col
    assert any(label == '2ГИС' for label, _col in meta['var_cols'])
    assert meta['sources'] == ['Сайт', '2ГИС']
    print('✓ метаданные для дашборда - строки данных, последняя var-колонка, источники')


# ── лист «Расхождения»: единообразие сайт/карты + осмысленное примечание ───


def test_site_discrepancy_has_separate_площадка_and_поле():
    """«Площадка»/«Поле» - раздельные колонки (не склеены через «: »), как в
    листе «Расхождения» - «Сайт» + «Общий Город», а не «Сайт: Общий Город»."""
    res = {'domain': 'stalmetural.ru', 'city': 'Москва', 'country': 'Россия',
          'error': '', 'fields': [
              {'field': 'Тел. Общий Город', 'expected': '+7 495 111-11-11',
               'found': '+7 495 222-22-22', 'status': 'bug', 'note': ''}]}
    расхождения, _meta = vr._записать_проверку_кп_лист(*_wb_with_style(), [res], False, [])
    assert len(расхождения) == 1
    d = расхождения[0]
    assert d['площадка'] == 'Сайт'
    assert d['поле'] == 'Общий Город'
    assert d['ссылка'] == 'https://stalmetural.ru/'
    print('✓ расхождение по сайту: площадка «Сайт» + поле отдельно, есть ссылка на сайт')


def test_map_discrepancy_note_explains_missing_value():
    """Карточка ничего не нашла (found='–') - примечание должно объяснять
    ПОЧЕМУ, а не быть пустым (раньше note всегда была '')."""
    wb, hdr_fill, thin = _wb_with_style()
    r = MapCheckResult(source='2gis', city='Казань', url='https://2gis.ru/x',
                       available=True, country='Россия', phone_match=False,
                       address_match=True, site_match=True,
                       issues=['телефон на карте не совпал с КП'],
                       details=[{'field': 'телефон', 'kp': '8432538626', 'card': '–'}])
    расхождения, _meta = vr._записать_проверку_кп_лист(
        wb, hdr_fill, thin, [_site_res(city='Казань')], False, [r])
    assert len(расхождения) == 1
    d = расхождения[0]
    assert 'не найден' in d['примечание'] and '2ГИС' in d['примечание']
    assert d['площадка'] == '2ГИС' and d['поле'] == 'телефон'
    assert d['ссылка'] == 'https://2gis.ru/x'
    print('✓ пустая карточка → примечание объясняет «не найден на карточке», не пусто')


def test_map_discrepancy_note_explains_value_mismatch():
    """Карточка нашла ДРУГОЕ значение (не пусто) - другая формулировка
    примечания, не путать с «не найден»."""
    wb, hdr_fill, thin = _wb_with_style()
    r = MapCheckResult(source='yandex', city='Москва', url='u', available=True,
                       country='Россия', phone_match=False, address_match=True,
                       site_match=True,
                       issues=['телефон на карте не совпал с КП'],
                       details=[{'field': 'телефон', 'kp': '4991303669',
                                'card': '+7 000 000-00-00'}])
    расхождения, _meta = vr._записать_проверку_кп_лист(
        wb, hdr_fill, thin, [_site_res()], False, [r])
    note = расхождения[0]['примечание']
    assert 'не совпадает' in note and 'не найден' not in note
    print('✓ карточка нашла другое значение → «не совпадает с КП», не «не найден»')
