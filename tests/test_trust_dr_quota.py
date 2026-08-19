"""Квота Open PageRank: состояние из заголовков, сообщение и плашка в отчёте.

Квота считается в ДОМЕНАХ за месяц - у ИМП это 242 хоста за один прогон.
Когда она кончается, часть хостов остаётся без DR, и это должно быть
объяснено: строкой в логе и плашкой на листе «Трафик и траст». Во всех
остальных случаях плашки быть НЕ должно.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import trust_check as tc
from tests.test_trust_dr import _Ответ, _подменить


def test_код_429_помечает_квоту(monkeypatch):
    _подменить(monkeypatch, _Ответ(код=429, текст='quota',
                                   headers={'X-Domains-Limit': '1000'}))
    состояние, строки = {}, []

    tc.fetch_dr(['inmetprom.ru'], 'opr_live_x', log=строки.append,
                state=состояние)

    assert состояние['quota_out'] is True
    assert состояние['limit'] == 1000
    assert any('429' in s for s in строки), строки
    print('✓ HTTP 429 → quota_out + лимит из заголовка')


def test_остаток_ноль_прекращает_обход(monkeypatch):
    """Ответ пришёл, но остаток доменов кончился - следующие куски не шлём."""
    ответ = _Ответ(данные={'results': [
        {'domain': 'inmetprom.ru', 'found': True, 'open_page_rank': 1.17},
    ]}, headers={'X-Domains-Remaining': '0', 'X-Domains-Limit': '100'})
    з = _подменить(monkeypatch, ответ)
    состояние, строки = {}, []

    out = tc.fetch_dr([f'{i}.inmetprom.ru' for i in range(250)],
                      'opr_live_x', log=строки.append, state=состояние)

    assert состояние['quota_out'] is True
    assert out == {'inmetprom.ru': 1.17}
    assert len(з.вызовы) == 1, 'после нулевого остатка ходить некуда'
    print('✓ остаток 0 → обход прекращён, а не 250 бесполезных доменов')


def test_запас_квоты_не_поднимает_флаг(monkeypatch):
    ответ = _Ответ(данные={'results': [
        {'domain': 'inmetprom.ru', 'found': True, 'open_page_rank': 1.17},
    ]}, headers={'X-Domains-Remaining': '758', 'X-Domains-Limit': '1000'})
    _подменить(monkeypatch, ответ)
    состояние = {}

    tc.fetch_dr(['inmetprom.ru'], 'opr_live_x', state=состояние)

    assert not состояние.get('quota_out')
    assert состояние['remaining'] == 758
    print('✓ пока квота есть - флага нет')


def test_сообщение_считает_хосты():
    хосты = [{'host': 'inmetprom.ru', 'dr': 1.17},
             {'host': 'spb.inmetprom.ru', 'dr': None},
             {'host': 'kazan.inmetprom.ru', 'dr': None}]

    текст = tc._quota_note({'limit': 1000}, хосты)

    assert 'у 1 из 3 хостов' in текст
    assert '1000' in текст
    # Плашка живёт в узкой колонке - длинный текст раздувает строку.
    assert len(текст) < 130, текст
    print('✓ в сообщении честные «посчитано N из M», и оно короткое')


def test_плашка_в_отчёте_только_при_исчерпанной_квоте(tmp_path):
    """Лист «Трафик и траст» с плашкой и без неё."""
    from openpyxl import load_workbook
    import reporter

    хосты = [{'host': 'inmetprom.ru', 'sqi': 220, 'dr': 1.17},
             {'host': 'spb.inmetprom.ru', 'sqi': 90, 'dr': None}]

    def _лист(примечание):
        траст = {'available': True, 'hosts': хосты,
                 'dr_quota_note': примечание}
        wb = reporter.Workbook()
        reporter._build_traffic_overview_sheet(wb, None, trust=траст)
        путь = tmp_path / f'{"с" if примечание else "без"}.xlsx'
        wb.save(путь)
        return load_workbook(путь)['Трафик и траст']

    # Текст берём тот же, что уйдёт в реальный отчёт, а не выдуманный:
    # иначе тест продолжает проходить после смены формулировки.
    текст = tc._quota_note({'limit': 1000}, хосты)

    def _найти(лист):
        """Плашка - единственная ячейка со знаком ⚠ (шапка колонки «DR (Open
        PageRank)» под поиск по названию сервиса попадала бы тоже)."""
        return [c.value for row in лист.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith('⚠')]

    с_плашкой = _найти(_лист(текст))
    без_плашки = _найти(_лист(None))

    assert len(с_плашкой) == 1 and 'Open PageRank' in с_плашкой[0]
    assert без_плашки == []
    print('✓ плашка есть только когда квота кончилась')
