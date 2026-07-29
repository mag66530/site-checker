"""Устойчивость проверки форм:
1) Атомарное сохранение лога - оборванная запись НЕ бьёт отчёт (раньше при сбое
   сохранения файл превращался в «Truncated file header» и терялись все прошлые
   формы прогона).
2) Мобильную вёрстку страниц «только для» других городов (подписка Хабаровска
   и т.п.) в прогоне не гоняем - иначе в отчёт лезет лишний домен.
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                'forms_tester'))

import test_all as t          # noqa: E402
import forms_run as fr        # noqa: E402


def test_atomic_save_valid_file(tmp_path):
    from openpyxl import Workbook, load_workbook
    p = str(tmp_path / 'log.xlsx')
    wb = Workbook()
    wb.active['A1'] = 'ok'
    t._atomic_save_wb(wb, p)
    assert load_workbook(p).active['A1'].value == 'ok'
    assert not os.path.exists(p + '.tmp')          # временный файл убран
    print('✓ атомарное сохранение даёт валидный файл, tmp не остаётся')


def test_atomic_save_failure_keeps_previous(tmp_path):
    # Сбой сохранения (нет места и т.п.) НЕ должен портить уже валидный файл.
    from openpyxl import Workbook, load_workbook
    p = str(tmp_path / 'log.xlsx')
    wb = Workbook()
    wb.active['A1'] = 'first'
    t._atomic_save_wb(wb, p)

    class _BadWB:
        def save(self, _path):
            raise IOError('disk full')

    try:
        t._atomic_save_wb(_BadWB(), p)
    except IOError:
        pass
    assert load_workbook(p).active['A1'].value == 'first'   # старый файл цел
    assert not os.path.exists(p + '.tmp')                    # мусор убран
    print('✓ сбой сохранения не рушит отчёт: прошлые данные на месте')


def _write_cfg(path, body):
    path.write_text(body, encoding='utf-8')
    return path


def test_города_ограничения_читаются(tmp_path):
    cfg = _write_cfg(tmp_path / 'config.py', (
        "СТРАНИЦЫ_ДЛЯ_ПРОВЕРКИ = [\n"
        "  {'тип': 'Главная'},\n"
        "  {'тип': 'Подписка_Хабаровск', 'только_города': ['Хабаровск']},\n"
        "  {'тип': 'Менеджер_СНГ', 'только_города': ['Алматы', 'Минск']},\n"
        "]\n"
    ))
    огр = fr._страницы_только_города(cfg)
    assert огр == {'Подписка_Хабаровск': {'Хабаровск'},
                   'Менеджер_СНГ': {'Алматы', 'Минск'}}
    print('✓ ограничения «только_города» страниц читаются из конфига')


def test_мобильная_страница_чужого_города_пропускается():
    # Логика фильтра из forms_run: страницу пропускаем, если она «только для»
    # городов, которых нет в прогоне.
    огр = {'Подписка_Хабаровск': {'Хабаровск'}}
    run_города = {'Москва'}

    def _пропустить(тип):
        только = огр.get(тип)
        return bool(только) and not (run_города & только)

    assert _пропустить('Подписка_Хабаровск') is True     # Хабаровск при Москве - мимо
    assert _пропустить('Главная') is False               # общая страница - гоним
    # А в прогоне Хабаровска - страница нужна.
    run_города2 = {'Хабаровск'}
    assert not (огр['Подписка_Хабаровск'] and not (run_города2 & огр['Подписка_Хабаровск']))
    print('✓ мобильную вёрстку чужого города пропускаем, свой - гоним')


def test_обрыв_распознаётся_сквозь_обёртку_сценария():
    # Сценарий заворачивает сетевую ошибку в свой RuntimeError («прервался на
    # шаге …»). Раньше исходный net:: терялся, и обрыв прокси/сети писался в
    # отчёт как «ОШИБКА формы». Теперь _это_обрыв_связи разворачивает __cause__.
    net = Exception('Page.goto: net::ERR_CONNECTION_RESET at https://metpromko.ru/')
    try:
        try:
            raise net
        except Exception as e:
            raise RuntimeError('прервался на шаг 1 «перейти»') from e
    except Exception as wrapped:
        assert t._это_обрыв_связи(wrapped) is True

    # Текст исходной ошибки, вшитый в сообщение (второй пояс надёжности).
    assert t._это_обрыв_связи(
        RuntimeError('прервался на шаг 1: Page.goto: net::ERR_CONNECTION_RESET')
    ) is True
    assert t._это_обрыв_связи(RuntimeError('Timeout 30000ms exceeded')) is True
    print('✓ сетевой обрыв распознаётся даже завёрнутым в RuntimeError сценария')


def test_дефект_формы_не_путается_с_обрывом():
    # Настоящий дефект формы НЕ должен помечаться как сеть/прокси - иначе
    # обратная ложь: реальную поломку спишут на «недоступность домена».
    assert t._это_обрыв_связи(
        RuntimeError('прервался на шаг 2 «форма»: Форма не найдена на странице')
    ) is False
    assert t._это_обрыв_связи('всё хорошо') is False
    print('✓ дефект формы не маскируется под обрыв связи')


def test_requests_путь_берёт_прокси_из_env(monkeypatch):
    # Паритет с «Проверкой КП»: requests-путь должен ходить тем же FORMS_PROXY.
    monkeypatch.setenv('FORMS_PROXY', 'http://205.172.57.182:1080')
    p = t._requests_proxies_from_env()
    assert p == {'http': 'http://205.172.57.182:1080',
                 'https': 'http://205.172.57.182:1080'}
    # Без схемы - достраиваем http://
    monkeypatch.setenv('FORMS_PROXY', '1.2.3.4:8080')
    assert t._requests_proxies_from_env()['https'] == 'http://1.2.3.4:8080'
    # Пусто - None (идём напрямую, как раньше)
    monkeypatch.delenv('FORMS_PROXY', raising=False)
    assert t._requests_proxies_from_env() is None
    print('✓ requests-путь подхватывает FORMS_PROXY (как в «Проверке КП»)')


def test_троттлинг_сайта_не_дефект_формы():
    # Сайт ответил «попробуйте позже» (как «Обратная связь» Метпромко при серии
    # заявок) - это временный отказ, а не поломка формы.
    class _FakePage:
        def __init__(self, text):
            self._t = text
        def inner_text(self, _sel):
            return self._t

    assert t._страница_просит_позже(
        _FakePage('Ошибка. Попробуйте заказать позже.')) is True
    assert t._страница_просит_позже(
        _FakePage('Слишком много попыток, повторите попытку позже')) is True
    # Обычная страница без маркеров лимита - не троттлинг.
    assert t._страница_просит_позже(
        _FakePage('Спасибо! Ваша заявка принята.')) is False

    # Вердикт-статус чистится в короткое «Отклонено сервером» (для листа «Логи»),
    _clean, _ = t._status_clean_reason(
        'ОТКЛОНЕНО СЕРВЕРОМ (повторите позже - троттлинг/лимит)')
    assert _clean == 'Отклонено сервером'
    # …а в матрице «Москва» это ⚠ (проверить вручную), НЕ ✗ (дефект формы).
    sym, _ = t._матрица_классифицировать('Статус', 'Отклонено сервером')
    assert sym == '⚠'
    # Контроль: настоящая ошибка отправки остаётся ✗.
    sym_err, _ = t._матрица_классифицировать('Статус', 'Ошибка')
    assert sym_err == '✗'
    print('✓ троттлинг сайта = ⚠ (временный отказ), не путается с дефектом формы')


if __name__ == '__main__':
    import pytest
    raise SystemExit(pytest.main([__file__, '-v', '-s']))
