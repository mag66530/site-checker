"""
uniqueness_run.py - фоновый прогон проверки уникальности контента через text.ru.

Запускается страницей приложения отдельным процессом (как pagespeed_run.py).

Что делает:
  1) собирает НЕБОЛЬШУЮ выборку страниц ТОЛЬКО главного домена проекта
     (главная + каталог + N категорий + N товаров) - города-поддомены не трогаем,
     это дубли по дизайну и лишний расход символов text.ru;
  2) для каждой страницы достаёт «основной текст» и шлёт в text.ru, исключая
     собственные домены проекта (exceptdomain) - чтобы уникальность считалась
     относительно ЧУЖИХ сайтов;
  3) пишет last_run.json (для страницы) и Excel-отчёт.

Ключ text.ru берётся из переменной окружения TEXTRU_KEY (её проставляет страница
из секретов). В аргументах ключ не передаём - чтобы не светился.

Прогресс печатается в stdout - страница тайлит лог.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import traceback
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

import sources as S
import uniqueness_checker as UC

CACHE_DIR = ROOT / "cache" / "uniqueness"


def _out_dir(pid: str) -> Path:
    d = CACHE_DIR / pid
    d.mkdir(parents=True, exist_ok=True)
    return d


def _log(msg: str) -> None:
    print(msg, flush=True)


def _load_products(pid: str, project: dict, sources) -> list[str]:
    """Товарные pathname'ы: сперва локальная база листингов, затем sitemap."""
    try:
        from product_links import load_product_links
        base = load_product_links(pid)
        if base and base.get("pathnames"):
            return list(base["pathnames"])
    except Exception as e:  # noqa: BLE001
        _log(f"⚠ База листингов недоступна ({e}).")
    try:
        import asyncio
        import sitemap
        sm = asyncio.run(sitemap.load_product_pathnames(
            project, sources.categories, sources.filters,
            log=lambda lvl, msg: _log(msg)))
        return list((sm or {}).get("pathnames") or [])
    except Exception as e:  # noqa: BLE001
        _log(f"⚠ Товары из sitemap недоступны ({e}).")
        return []


def _sample_urls(pid, project, sources, cats, prods):
    """Выборка главного домена: главная + каталог + N категорий + N товаров."""
    if prods > 0:
        sources.products = _load_products(pid, project, sources)
    plan = S.build_plan(
        sources,
        random_subdomains_count=0,                 # только главный домен
        mandatory_city=project.get("mandatory_city", "Москва"),
        mandatory_hosts=[],
        categories_per_subdomain=cats,
        filters_per_subdomain=0,                    # фильтры пропускаем (тонкие/дубли)
        products_per_subdomain=prods,
        check_products=prods > 0 and bool(sources.products),
    )
    return [(t.url, t.type_code) for t in plan.tasks]


def _all_project_hosts(sources) -> list[str]:
    """Все домены/поддомены проекта - для exceptdomain (исключить свои сайты)."""
    return [s.url for s in sources.subdomains]


def run(args) -> int:
    pid = args.project
    try:
        project = S.load_project_config(pid)
    except Exception as e:  # noqa: BLE001
        _log(f"✗ Не удалось загрузить конфиг проекта: {e}")
        return 2
    project_name = project.get("name", pid)
    sources = S.load_sources(project)

    _log(f"Проект: {project_name} · уникальность контента (text.ru)")

    typed = _sample_urls(pid, project, sources, args.categories, args.products)
    if not typed:
        _log("✗ Нет URL для проверки.")
        return 2

    exceptdomain = UC._project_domains(_all_project_hosts(sources))
    _log(f"К проверке: {len(typed)} страниц главного домена. "
         f"Исключаем свои домены: {exceptdomain[:120]}…")

    key = os.environ.get("TEXTRU_KEY", "").strip()
    if not key:
        _log("✗ TEXTRU_KEY не задан (ключ text.ru). Задайте секрет textru_key.")
        return 2
    client = UC.TextRuClient(key)

    results = UC.run_batch(
        typed, client, exceptdomain=exceptdomain, log=_log,
        submit_pause=args.submit_pause, poll_interval=args.poll_interval,
        max_wait=args.max_wait,
    )
    summary = UC.summarize(results, threshold=args.threshold)
    _log(f"Готово: проверено {summary['checked']}/{summary['total']}, "
         f"ниже {args.threshold}% - {summary['below']}, ошибок {summary['errors']}.")

    out = _out_dir(pid)
    date_disp = datetime.now().strftime("%d.%m.%Y")
    xlsx_name = f"{pid.upper()}-уникальность-{date_disp}.xlsx"
    _save_xlsx(out / xlsx_name, project_name, results, summary, args.threshold)
    _log(f"Отчёт: {out / xlsx_name}")

    payload = {
        "project": pid, "project_name": project_name,
        "run_at": datetime.now().isoformat(timespec="seconds"),
        "threshold": args.threshold, "summary": summary,
        "exceptdomain": exceptdomain,
        "rows": [{
            "url": r.url, "type": r.type_code, "unique": r.unique,
            "chars": r.chars, "error": r.error,
            "sources": r.sources[:10],
        } for r in results],
        "xlsx_name": xlsx_name,
    }
    (out / "last_run.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    _log("✅ ГОТОВО")
    return 0


def _save_xlsx(path, project_name, results, summary, threshold):
    """Простой Excel: по странице - URL, тип, % уникальности, чужие источники."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception as e:  # noqa: BLE001
        _log(f"⚠ openpyxl недоступен ({e}) - Excel не сохранён (есть last_run.json).")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Уникальность"
    ws.sheet_view.showGridLines = False
    bold = Font(bold=True)
    red = Font(color="B91C1C", bold=True)
    green = Font(color="15803D", bold=True)
    hdr_fill = PatternFill("solid", fgColor="EEF1F5")
    top = Alignment(vertical="top", wrap_text=True)

    ws["A1"] = f"Уникальность контента - {project_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"Проверено {summary['checked']}/{summary['total']} страниц · "
                f"средняя уникальность {summary['avg_unique']}% · "
                f"ниже {threshold}%: {summary['below']} · порог {threshold}%")
    ws["A2"].font = Font(italic=True, color="6B7280")

    heads = ["Страница", "Тип", "Уникальность", "Символов",
             "С каким сайтом пересекается (совпадение %)"]
    for j, h in enumerate(heads, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = bold
        c.fill = hdr_fill
    widths = [58, 12, 14, 10, 70]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + j)].width = w

    row = 5
    for r in results:
        ws.cell(row=row, column=1, value=r.url).alignment = top
        ws.cell(row=row, column=2, value=r.type_code).alignment = top
        if r.error:
            cu = ws.cell(row=row, column=3, value=f"— ({r.error})")
            cu.font = Font(color="6B7280", italic=True)
        elif r.unique is not None:
            cu = ws.cell(row=row, column=3, value=f"{r.unique:.1f}%")
            cu.font = red if r.unique < threshold else green
        ws.cell(row=row, column=3).alignment = top
        ws.cell(row=row, column=4, value=r.chars or None).alignment = top
        src = "\n".join(
            f"{s['url']} ({s['plagiat']:.1f}%)" if s.get("plagiat") is not None
            else s["url"] for s in (r.sources or []))
        ws.cell(row=row, column=5, value=src or ("—" if not r.error else "")).alignment = top
        row += 1
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description="Проверка уникальности контента (text.ru)")
    ap.add_argument("--project", required=True, help="id проекта (smu/imp/mpe/…)")
    ap.add_argument("--categories", type=int, default=3, help="сколько категорий проверить")
    ap.add_argument("--products", type=int, default=3, help="сколько товаров проверить")
    ap.add_argument("--threshold", type=float, default=95.0, help="порог уникальности, %%")
    ap.add_argument("--submit-pause", type=float, default=1.5)
    ap.add_argument("--poll-interval", type=float, default=15.0)
    ap.add_argument("--max-wait", type=float, default=480.0)
    args = ap.parse_args()
    try:
        sys.exit(run(args))
    except Exception as e:  # noqa: BLE001
        _log(f"✗ ОШИБКА: {e}")
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
