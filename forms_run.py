"""
forms_run.py - один фоновый процесс проверки форм для проекта.

Запускается страницей «Проверка форм» в фоне (как autoclick_run.py для
кликеров). Готовит рабочую папку cache/forms/<project>/, кладёт туда
config.py выбранного проекта и log_forms.xlsx, и гоняет движок форм-тестера
(forms_tester/test_all.py → run_test). Весь вывод идёт в stdout, который
вызывающая сторона перенаправляет в лог-файл.

Поддомены (города): если у проекта есть справочник forms_tester/projects/
<id>/cities.csv (город;url;почта), можно прогнать формы по выбранным городам.
Для каждого города подменяется поддомен в URL, а в отчёт пишутся колонки
«Город» и «Почта получателя» (куда должна прийти заявка).

Запуск:
    python forms_run.py --project smu
    python forms_run.py --project smu --cities "Москва,Санкт-Петербург,Казань"
"""
import argparse
import csv
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

ROOT = Path(__file__).parent
ENGINE = ROOT / 'forms_tester'                  # пакет с движком (test_all, name_format, form_tester)
PROJECTS_ROOT = ENGINE / 'projects'             # forms_tester/projects/<id>/config.py
WORK_ROOT = ROOT / 'cache' / 'forms'            # рабочие папки прогонов (в .gitignore)

PROJECT_NAMES = {
    'smu': 'СМУ - Стальметурал',
    'imp': 'ИМП - Инметпром',
    'mpe': 'МПЭ - Мепэн',
    'mpe_cart': 'МПЭ - Корзина',
    'avia': 'АПС - Авиапромсталь',
    'metpromko': 'Метпромко',
}


def _имена_заказов(src_config: Path) -> list[str]:
    """Названия сценариев/форм/шагов из блоков «тип: Оформление*» конфига проекта -
    это и есть сквозной заказ (корзина → оформление). Нужны, чтобы «Проверка целей»
    прогоняла ТОЛЬКО заказ (через ТОЛЬКО_ФОРМЫ), а не все формы. Конфиг импортируем
    напрямую - в нём только словари/константы, движок не тянется."""
    import importlib.util
    try:
        spec = importlib.util.spec_from_file_location('cfg_orders', src_config)
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
    except Exception:
        return []
    имена: set[str] = set()
    for блок in getattr(m, 'СТРАНИЦЫ_ДЛЯ_ПРОВЕРКИ', []) or []:
        if not str(блок.get('тип', '')).startswith('Оформление'):
            continue
        for ключ in ('формы', 'модалки', 'сценарии'):
            for it in блок.get(ключ, []) or []:
                if it.get('название'):
                    имена.add(str(it['название']).strip())
                for шаг in it.get('шаги', []) or []:
                    if шаг.get('название'):
                        имена.add(str(шаг['название']).strip())
    return sorted(имена)


def _страницы_форм(src_config: Path) -> list[tuple[str, str]]:
    """Список (тип, url) страниц форм из СТРАНИЦЫ конфига - для проверки мобильной
    вёрстки. Дедуп по url (у карточки товара и оформления часто один адрес)."""
    import importlib.util
    try:
        spec = importlib.util.spec_from_file_location('cfg_pages', src_config)
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
    except Exception:
        return []
    out: list[tuple[str, str]] = []
    seen: set[str] = set()
    for тип, url in (getattr(m, 'СТРАНИЦЫ', {}) or {}).items():
        u = str(url or '').strip()
        if u and u not in seen:
            seen.add(u)
            out.append((str(тип), u))
    return out


def _страницы_только_города(src_config: Path) -> dict:
    """{тип_страницы: set(городов)} для блоков СТРАНИЦЫ_ДЛЯ_ПРОВЕРКИ с
    «только_города» - такие страницы существуют лишь в этих городах (напр.
    подписка Хабаровска). Нужно, чтобы НЕ гонять их мобильную вёрстку в других
    городах, иначе в отчёт всплывает лишний домен (habarovsk.stalmetural.ru)."""
    import importlib.util
    try:
        spec = importlib.util.spec_from_file_location('cfg_only_cities', src_config)
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
    except Exception:
        return {}
    out: dict[str, set] = {}
    for блок in getattr(m, 'СТРАНИЦЫ_ДЛЯ_ПРОВЕРКИ', []) or []:
        тип = str(блок.get('тип', '')).strip()
        только = блок.get('только_города')
        if тип and только:
            out[тип] = {str(g).strip() for g in только}
    return out

# Проекты-варианты со своим config.py, но БЕЗ своего cities.csv - берут
# справочник городов у «родителя». Так «МПЭ - Корзина» гоняет те же города,
# что и Мепэн, без дублирования файла на 160 строк.
CITIES_FROM = {
    'mpe_cart': 'mpe',
}


def _stamp(msg):
    print(f'[{datetime.now().strftime("%H:%M:%S")}] {msg}', flush=True)


def _load_admin_zones(src_config: Path):
    """Читает АДМИН_ЗОНЫ из ИСХОДНОГО config.py проекта (не из рабочей копии, где
    домены уже подменены под город). Возвращает список зон {домен, города} или
    None, если у проекта админ-зоны не настроены (тогда проверку админки не
    делаем - например, ИМП/МПЭ, где админка устроена иначе)."""
    try:
        import importlib.util
        spec = importlib.util.spec_from_file_location('_orig_cfg_zones', str(src_config))
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        зоны = getattr(mod, 'АДМИН_ЗОНЫ', None)
        if зоны:
            return [dict(z) for z in зоны]
    except Exception as e:  # noqa: BLE001
        _stamp(f'⚠️ Не удалось прочитать АДМИН_ЗОНЫ из конфига: {e}')
    return None


def _load_cities(project: str):
    """Справочник городов проекта: [(город, url, почта), ...]. Пусто, если файла нет.
    Для проектов-вариантов (CITIES_FROM) берём справочник у родителя."""
    project = CITIES_FROM.get(project, project)
    f = PROJECTS_ROOT / project / 'cities.csv'
    if not f.is_file():
        return []
    out = []
    with open(f, encoding='utf-8', newline='') as fh:
        for row in csv.DictReader(fh):
            city = (row.get('город') or '').strip()
            url = (row.get('url') or '').strip().rstrip('/')
            mail = (row.get('почта') or '').strip()
            if city and url:
                out.append((city, url, mail))
    return out


def _read_logi(path: str):
    """(шапка, [строки]) листа «Логи», или (None, []) если листа/файла нет."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        if 'Логи' not in wb.sheetnames:
            wb.close()
            return None, []
        ws = wb['Логи']
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return (rows[0], list(rows[1:])) if rows else (None, [])
    except Exception:  # noqa: BLE001
        return None, []


def _row_city(hdr, row) -> str:
    try:
        return str(row[list(hdr).index('Город')] or '').strip()
    except Exception:  # noqa: BLE001
        return ''


def _rewrite_logi(path: str, hdr, rows) -> None:
    """Заменить строки листа «Логи» на переданные (шапку сохранить); другие листы
    («Сводка»/«Цели») не трогаем."""
    from openpyxl import load_workbook
    wb = load_workbook(path)
    if 'Логи' not in wb.sheetnames:
        wb.close()
        return
    ws = wb['Логи']
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    for r in rows:
        ws.append(r)
    tmp = f'{path}.tmp'                       # атомарно: temp + rename (как в движке)
    wb.save(tmp)
    os.replace(tmp, path)


def main() -> int:
    ap = argparse.ArgumentParser(description='Прогон проверки форм для проекта.')
    ap.add_argument('--project', required=True, choices=list(PROJECT_NAMES),
                    help='Идентификатор проекта: smu / imp / mpe')
    ap.add_argument('--no-clear-excel', action='store_true',
                    help='Не очищать log_forms.xlsx перед прогоном')
    ap.add_argument('--show-browser', action='store_true',
                    help='Показывать окно браузера (по умолчанию скрыто, headless)')
    ap.add_argument('--cities', default='',
                    help='Список городов через запятую (из cities.csv). Пусто = основной сайт.')
    ap.add_argument('--forms-file', default='',
                    help='Путь к JSON-файлу со списком выбранных форм (имена). '
                         'Пусто = проверять все формы проекта.')
    ap.add_argument('--no-admin', action='store_true',
                    help='Не проверять админку (Уровень 1) после прогона.')
    ap.add_argument('--file-probe', action='store_true',
                    help='Проба серверной фильтрации загрузки: грузит безвредные '
                         'файлы 31 типа (13 опасных .php/.svg/.exe/… + 18 обычных) '
                         'в формы с полем загрузки и отправляет - что примет '
                         'сервер. Создаёт тест-заявки на боевых формах.')
    ap.add_argument('--only-orders', action='store_true',
                    help='Прогнать ТОЛЬКО сквозной заказ (блоки «Оформление») - '
                         'для подтверждения заказ-целей из «Проверки целей».')
    ap.add_argument('--check-goals', action='store_true',
                    help='Ловить цели Яндекс.Метрики (для «Проверки целей»). В '
                         'обычной «Проверке форм» НЕ ставится - формы без целей.')
    ap.add_argument('--xss-probe', action='store_true',
                    help='Проба защиты от XSS: в поле имени шлёт безвредный '
                         'payload-маркер и смотрит, исполнился ли он на ответе '
                         '(колонка «Защита от XSS»). Создаёт тест-заявку с маркером.')
    ap.add_argument('--server-validation-probe', action='store_true',
                    help='Проба серверной валидации: снимает клиентские '
                         'ограничения (required/pattern/maxlength) через JS и '
                         'реально отправляет невалидные данные (пустое '
                         'обязательное поле / битый e-mail / превышение длины) - '
                         'как правка в DevTools. Пропускает формы заказа. Может '
                         'создать до 3 тест-заявок на форму, если сервер их примет.')
    ap.add_argument('--rate-limit-probe', action='store_true',
                    help='Активная проверка лимита запросов: до 3 быстрых '
                         'повторных ВАЛИДНЫХ отправок подряд - смотрит, '
                         'блокирует ли сайт позднюю попытку. Пассивная проверка '
                         '(капча/honeypot) идёт всегда и этим флагом не '
                         'управляется. Пропускает формы заказа. Создаёт до 3 '
                         'тест-заявок на форму, если лимита нет.')
    a = ap.parse_args()

    # Видимый браузер (show-browser) физически невозможен без дисплея: на сервере/
    # в облаке (Streamlit Cloud - нет $DISPLAY) headed-запуск падает с «launched a
    # headed browser without XServer» на КАЖДОЙ форме (весь прогон = «Ошибка»). Если
    # дисплея нет - принудительно гоним скрыто (headless), отчёт формируется как
    # обычно. Локально (Windows/Mac/Linux с X-сервером) галочка работает как раньше.
    show_browser = a.show_browser
    if show_browser and sys.platform.startswith('linux') and not os.environ.get('DISPLAY'):
        _stamp('⚠️ «Показывать окно браузера» недоступно без дисплея (сервер/облако) '
               '- гоню скрыто (headless). На отчёт это не влияет.')
        show_browser = False

    name = PROJECT_NAMES[a.project]
    src_config = PROJECTS_ROOT / a.project / 'config.py'
    if not src_config.is_file():
        _stamp(f'✗ Нет файла конфигурации: {src_config}')
        return 2

    # Справочник городов и какие из них гнать
    cities_all = _load_cities(a.project)
    by_name = {c[0]: c for c in cities_all}
    main_host = urlparse(cities_all[0][1]).netloc if cities_all else ''   # домен основного сайта

    wanted = [c.strip() for c in a.cities.split(',') if c.strip()]
    if wanted and cities_all:
        run_cities = [by_name[c] for c in wanted if c in by_name]
    elif cities_all:
        run_cities = [cities_all[0]]                       # по умолчанию - основной город
    else:
        run_cities = [('', '', '')]                        # нет справочника - обычный прогон

    _stamp(f'ПРОВЕРКА ФОРМ СТАРТ - проект {name}'
           + (f' - городов: {len(run_cities)}' if run_cities and run_cities[0][0] else ''))
    if a.xss_probe:
        _stamp('🛡️ Проба защиты от XSS ВКЛючена: в поле имени уходит безвредный '
               'payload-маркер. Создаётся тест-заявка с маркером - после прогона '
               'удалите её в админке (компьютеру/данным маркер не вредит).')
    if a.server_validation_probe:
        _stamp('🛡️ Проба серверной валидации ВКЛючена: отправляю невалидные '
               'данные в обход клиентской проверки (формы заказа пропускаются). '
               'Если сервер их примет - в админке появится тест-заявка с меткой '
               '«ТЕСТ-ВАЛИДАЦИЯ» или пустым/битым полем - удалите её после прогона.')
    if a.rate_limit_probe:
        _stamp('🚦 Активная проверка лимита запросов ВКЛючена: до 3 быстрых '
               'повторных отправок на форму (формы заказа пропускаются). Если '
               'лимита нет - в админке появятся до 3 тест-заявок - удалите их '
               'после прогона.')

    work = WORK_ROOT / a.project
    work.mkdir(parents=True, exist_ok=True)
    base_config = src_config.read_text(encoding='utf-8')

    # Выбор форм из интерфейса: дописываем в конфиг список ТОЛЬКО_ФОРМЫ (движок
    # прогонит только формы с этими именами). Пустой/отсутствующий файл - гоним всё.
    forms_filter = []
    if a.forms_file:
        try:
            import json
            forms_filter = json.loads(Path(a.forms_file).read_text(encoding='utf-8')) or []
        except Exception as e:  # noqa: BLE001
            _stamp(f'⚠️ Не удалось прочитать список форм ({a.forms_file}): {e}')
            forms_filter = []
    # Только заказ (для «Проверки целей»): фильтр = имена сценариев оформления.
    if a.only_orders and not forms_filter:
        forms_filter = _имена_заказов(src_config)
        if forms_filter:
            _stamp(f'Режим «только заказ»: прогоняю сценарий оформления '
                   f'({len(forms_filter)} шаг(ов)).')
        else:
            _stamp('⚠️ Режим «только заказ»: в конфиге нет блока «Оформление» - '
                   'прогоняю все формы.')
    if forms_filter:
        base_config = base_config.rstrip() + '\n\nТОЛЬКО_ФОРМЫ = ' + repr(list(forms_filter)) + '\n'
        _stamp(f'Выбрано форм: {len(forms_filter)} (остальные пропускаем).')

    # Базовый домен для подмены берём НЕ из первой строки cities.csv, а из самого
    # конфига: тот город-домен, что реально встречается в URL-ах СТРАНИЦ. Иначе если
    # первый город в справочнике - поддомен (напр. abakan.mepen.ru), а конфиг на
    # mepen.ru, замена не находила совпадения и все города шли на основной сайт (МПЭ).
    for _c in cities_all:
        _h = urlparse(_c[1]).netloc
        if _h and f'//{_h}' in base_config:
            main_host = _h
            break

    sys.path.insert(0, str(ENGINE))
    sys.path.insert(0, str(work))
    prev = os.getcwd()
    try:
        os.chdir(work)
    except OSError as e:
        _stamp(f'✗ Не удалось перейти в {work}: {e}')
        return 2

    rc = 0
    try:
        from form_tester.runner import run_test
        from form_tester.stop_signal import make_stop_check
        stop = make_stop_check()

        # Страховка мультигорода: сразу после прогона КАЖДОГО города снимаем его
        # строки из «Логи» (пока следующий город их не затронул) и в конце
        # пересобираем лог из всех снимков. Так формы всех городов гарантированно
        # попадают в отчёт, даже если движок в мультигороде теряет часть строк.
        _мультигород = len([c for c in run_cities if c and c[0]]) > 1
        _снимки: list[tuple[str, list]] = []
        _логи_шапка = None

        for i, (city, city_url, city_mail) in enumerate(run_cities):
            if stop():
                _stamp('⛔ Остановлено')
                break
            # Подменяем домен в конфиге под город (для Москвы/основного - без изменений)
            cfg = base_config
            if city and main_host:
                target = urlparse(city_url).netloc
                if target and target != main_host:
                    cfg = cfg.replace(f'//{main_host}', f'//{target}')
            (work / 'config.py').write_text(cfg, encoding='utf-8')

            if city:
                _stamp(f'── Город: {city}  ({city_url})  → заявка должна прийти на {city_mail or "?"} ──')

            run_test(
                ОЧИСТИТЬ_EXCEL=(not a.no_clear_excel and i == 0),   # чистим лог только перед первым
                stop_flag=stop,
                headless=not show_browser,
                город=city,
                почта_получателя=city_mail,
                проба_файлов=a.file_probe,
                xss_проба=a.xss_probe,
                валидация_проба=a.server_validation_probe,
                лимит_проба=a.rate_limit_probe,
                проверять_цели=a.check_goals,
            )

            # Снимок строк этого города сразу после его прогона.
            if _мультигород and city:
                _h, _rows = _read_logi('log_forms.xlsx')
                if _h is not None:
                    _логи_шапка = _h
                    _мои = [r for r in _rows
                            if _row_city(_h, r).lower() == city.strip().lower()]
                    _снимки.append((city, _мои))

        # Пересобираем «Логи» из снимков всех городов (формы каждого - со своим
        # городом), пока не начались общие блоки (cookie/вёрстка/админка).
        if _мультигород and _логи_шапка is not None:
            try:
                _все = [r for _c, rs in _снимки for r in rs]
                _rewrite_logi('log_forms.xlsx', _логи_шапка, _все)
                _городов = len([1 for _c, rs in _снимки if rs])
                _stamp(f'🧩 Формы всех городов собраны в отчёт: {len(_все)} строк '
                       f'по {_городов} город(ам).')
            except Exception as e:  # noqa: BLE001
                _stamp(f'⚠️ Не удалось пересобрать лог по городам: {e}')

        # ── Пункт 2.12: cookie-уведомление + ссылка на политику + живочат ──
        # ВАЖНО: пишем строки 2.12 в лог ПЕРВЫМИ (до проверок админки) - они
        # добавляют строки через append_log_row по фиксированному порядку колонок,
        # а проверки админки ниже ВСТАВЛЯЮТ доп. колонки; если 2.12 писать после,
        # значения разъедутся по колонкам. Открываем главную каждого города свежим
        # контекстом (без cookie = «новый пользователь»). Тихо пропускается без городов.
        if not (stop and stop()):
            try:
                import privacy_check
                _города_212 = [(c, u) for (c, u, _m) in run_cities if u]
                if _города_212:
                    privacy_check.выполнить_проверку(
                        _города_212, excel_path='log_forms.xlsx',
                        show=show_browser, log=_stamp)
            except Exception as e:  # noqa: BLE001
                _stamp(f'⚠️ Проверка 2.12 (cookie/чат) не выполнена: {e}')

        # ── Мобильная вёрстка форм (горизонтальный скролл + тач-размеры) ──
        # Тоже через append_log_row, поэтому ДО проверок админки (иначе колонки
        # разъедутся). Меряем на домене первого выбранного города - шаблон вёрстки
        # у поддоменов один, гонять по всем городам смысла нет.
        if not (stop and stop()):
            try:
                import mobile_check
                _pages = _страницы_форм(src_config)
                _огр_города = _страницы_только_города(src_config)
                _run_города = {c[0] for c in run_cities if c and c[0]}
                if _pages and run_cities:
                    _c0_url = run_cities[0][1]
                    _c0_host = urlparse(_c0_url).netloc
                    _моб = []
                    for _тип, _purl in _pages:
                        # Страницы «только для» других городов (подписка Хабаровска
                        # и т.п.) в этом прогоне пропускаем - иначе в отчёт лезет
                        # чужой домен, которого мы не выбирали.
                        _только = _огр_города.get(_тип)
                        if _только and not (_run_города & _только):
                            continue
                        _u = _purl
                        if main_host and _c0_host and _c0_host != main_host:
                            _u = _u.replace(f'//{main_host}', f'//{_c0_host}')
                        _моб.append((_тип, _u))
                    mobile_check.выполнить_проверку(
                        _моб, excel_path='log_forms.xlsx',
                        show=show_browser, log=_stamp)
            except Exception as e:  # noqa: BLE001
                _stamp(f'⚠️ Проверка мобильной вёрстки не выполнена: {e}')

        # ── Уровень 1: проверка админки (если заданы креды admin.local.json) ──
        # У СМУ разные админки для РФ / СНГ / Steelgroup (АДМИН_ЗОНЫ в конфиге), но
        # логин/пароль общие. Логинимся в каждую нужную зону, читаем «Уведомления
        # с форм» за сегодня и сверяем с отправками. Пропускается без admin.local.json.
        if not a.no_admin and not (stop and stop()):
            try:
                import admin_check
                проект_дир = PROJECTS_ROOT / a.project
                зоны = _load_admin_zones(src_config)   # None → у проекта нет админ-зон
                if зоны:
                    # Логин/пароль приходят из окружения (введены на странице) или
                    # из admin.local.json; если нигде нет - выполнить_проверку тихо
                    # пропустит проверку.
                    admin_check.выполнить_проверку(
                        str(проект_дир), зоны,
                        excel_path='log_forms.xlsx',
                        submitted_path='submitted_forms.json',
                        show=show_browser, log=_stamp,
                        xss_проба=a.xss_probe,
                    )
                    # ── Пункт 2.11: заказы из корзины → список «Заказы» админки ──
                    # Тот же логин и зоны, но другой раздел (sale_order.php).
                    # Тихо пропустится без оформленных заказов/кредов.
                    try:
                        import order_admin_check
                        order_admin_check.выполнить_проверку(
                            str(проект_дир), зоны,
                            orders_path='placed_orders.json',
                            excel_path='log_forms.xlsx',
                            show=show_browser, log=_stamp,
                        )
                    except Exception as e:  # noqa: BLE001
                        _stamp(f'⚠️ Проверка заказов в админке не выполнена: {e}')
            except Exception as e:  # noqa: BLE001
                _stamp(f'⚠️ Проверка админки не выполнена: {e}')

        # ── Пункт 2.9: письмо об оформлении заказа приходит покупателю ──
        # Заказ(ы) оформлены на почту покупателя (её задаёт страница через
        # ORDER_BUYER_EMAIL). Если заданы креды тестового ящика (ORDER_MAIL_*) -
        # заходим по IMAP и подтверждаем письмо; иначе (своя почта) - напоминаем
        # проверить вручную. Без почты покупателя блок тихо пропускается.
        if not (stop and stop()):
            try:
                import order_mail_check
                order_mail_check.выполнить_проверку(
                    orders_path='placed_orders.json',
                    excel_path='log_forms.xlsx',
                    log=_stamp,
                )
            except Exception as e:  # noqa: BLE001
                _stamp(f'⚠️ Проверка письма о заказе не выполнена: {e}')

        # После пересборки «Логи» по городам «Сводку» тоже освежаем из полного
        # лога (иначе она осталась бы от последнего города).
        if _мультигород:
            try:
                from test_all import write_summary_sheet
                write_summary_sheet('log_forms.xlsx')
            except Exception as e:  # noqa: BLE001
                _stamp(f'⚠️ Сводку по городам обновить не удалось: {e}')

        # ── Читаемость: сводим строки одной формы в одну (1 форма = 1 строка) ──
        # Все проверки формы (Состав/Стилизация/Списки/Чекбоксы/Enter/Двойная/
        # Ошибки/Согласие/Вёрстка…) раньше шли отдельными строками (~11 на форму).
        # Делаем в самом конце, когда все колонки (в т.ч. «Статус в админке»)
        # уже на месте. Отключается переменной FORMS_NO_CONSOLIDATE=1.
        if not os.environ.get('FORMS_NO_CONSOLIDATE'):
            try:
                from test_all import консолидировать_форм_строки
                консолидировать_форм_строки('log_forms.xlsx')
            except Exception as e:  # noqa: BLE001
                _stamp(f'⚠️ Консолидация отчёта не выполнена: {e}')

        # ── Наглядность: матрица проверок по доменам (лист на город, проверки -
        # строки, формы - столбцы, ✓/✗/⚠/–) + легенда на «Сводке». Строится
        # ПОСЛЕ консолидации (нужны уже сведённые строки). Отключается
        # переменной FORMS_NO_MATRIX=1.
        if not os.environ.get('FORMS_NO_MATRIX'):
            try:
                from test_all import построить_матрицу_проверок
                построить_матрицу_проверок('log_forms.xlsx')
            except Exception as e:  # noqa: BLE001
                _stamp(f'⚠️ Матрица проверок не построена: {e}')

        # Финал: колонку «Комментарий» тянем шире (её ширину сбивают вставки
        # колонок проверок админки/письма). Делаем в самом конце, когда все
        # колонки уже на месте.
        try:
            from openpyxl import load_workbook as _lw
            from openpyxl.utils import get_column_letter as _gcl
            _wb = _lw('log_forms.xlsx')
            if 'Логи' in _wb.sheetnames:
                _ws = _wb['Логи']
                _h = [str(c.value or '').strip().lower() for c in _ws[1]]
                if 'комментарий' in _h:
                    _ws.column_dimensions[_gcl(_h.index('комментарий') + 1)].width = 120
                    _wb.save('log_forms.xlsx')
        except Exception:  # noqa: BLE001
            pass
    except SystemExit as e:
        rc = int(e.code) if isinstance(e.code, int) else 1
    except Exception as e:
        import traceback
        _stamp(f'✗ Ошибка прогона: {e}')
        traceback.print_exc()
        rc = 1
    finally:
        try:
            os.chdir(prev)
        except OSError:
            pass

    if rc == 0:
        _stamp(f'Лог сохранён: {work / "log_forms.xlsx"}')
        # Telegram: отчёт по формам получателям проекта (креды - в окружении, их
        # проставляет страница из секретов). НЕ шлём, когда формы гоняются ВНУТРИ
        # «Проверки целей» (--check-goals): там свой сводный отчёт по целям.
        if not a.check_goals:
            try:
                import telegram_notify as tn
                from telegram_notify import escape_html
                _отчёт = work / 'log_forms.xlsx'
                _города = [c for c, _u, _m in run_cities if c]
                import datetime as _dt
                _дата = _dt.datetime.now(_dt.timezone(_dt.timedelta(hours=5))).strftime('%d.%m.%Y')
                _имя = f'Form-{a.project}-{_дата}.xlsx'
                _бренд = name.split(' - ')[0].strip()
                _части = [f'Проверка форм {escape_html(_бренд)}']
                if _города:
                    _части.append(f'Города: {escape_html(", ".join(_города))}')
                _части.append('📎 Полный отчёт - в прикреплённом xlsx-файле')
                _текст = '\n\n'.join(_части)
                _res = tn.send_report_from_env(
                    project_name=name, summary_text=_текст,
                    report_file=_отчёт if _отчёт.is_file() else None,
                    report_filename=_имя,
                    log=lambda lvl, msg: _stamp(msg))
                if not _res.get('skipped'):
                    _stamp(f'✓ Telegram: отправлено {_res.get("sent", 0)}, '
                           f'не доставлено {_res.get("failed", 0)}')
            except Exception as e:  # noqa: BLE001
                _stamp(f'⚠ Telegram-отправка не удалась ({e}) - отчёт всё равно готов.')
        _stamp('✅ ВСЁ ГОТОВО')
    return rc


if __name__ == '__main__':
    raise SystemExit(main())
