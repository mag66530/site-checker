"""
goals_tester.py - проверка ВСЕХ целей Яндекс.Метрики проекта (страница
«Проверка целей» в панели).

Эталон - каталог целей, выгруженный из Метрики («Конверсии»):
catalogs/goals-<проект>.json (номер, название, условие, тип). Движок открывает
страницы сайта в браузере, выполняет безопасные действия (клики по телефонам,
почте, соцсетям/мессенджерам, кнопкам открытия форм - БЕЗ отправки заявок) и
слушает запросы к Метрике: каждая сработавшая JS-цель шлёт hit вида
goal://<хост>/<идентификатор>.

Вердикты по типам целей:
  js         - Сработала / НЕ сработала (если ждали от кликов) / Нет автодействия;
               формные цели (отправка заявки) не дублируем заявкой - статус
               «Прогоном форм» + подтягиваем результат из последнего отчёта форм.
  url/url_re - открываем страницу: 200 + счётчик отправил визит → «Сработает».
  auto       - автоцель Метрики (клики tel/mailto, формы, файлы…) - фиксируется
               Метрикой автоматически, отдельная проверка не нужна (информируем).
  jivo       - события чата Jivo запускает оператор/посетитель - только вручную.
  composite  - составная, Метрика считает её из шагов - смотрим цели-шаги.
"""
from __future__ import annotations

import json
import os
import re
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import unquote, urljoin

ROOT = Path(__file__).parent
CATALOGS = ROOT / 'catalogs'

# Метрика: hit сработавшей JS-цели
_RE_GOAL = re.compile(r"goal://[^/]+/([^&\s\"?#]+)")

# Безопасные «общие» клики на каждой странице (без отправки форм):
# телефоны, почта, мессенджеры и соцсети (клик фиксируется, переход гасим).
GENERIC_CLICK_SELECTORS = [
    "a[href^='tel:']",
    "a[href^='mailto:']",
    "a[href*='wa.me'], a[href*='api.whatsapp'], a[href*='whatsapp:']",
    "a[href*='t.me'], a[href*='tg://']",
    "a[href*='vk.com'], a[href*='vk.me']",
    "a[href*='ok.ru']",
    "a[href*='dzen.ru'], a[href*='zen.yandex']",
    "a[href*='rutube.ru']",
    "a[href*='max.ru'], a[href*='web.max']",
    "a[href*='yandex.ru/profile'], a[href*='yandex.ru/maps/org']",
]
MAX_PER_SELECTOR = 5     # кликаем не больше N элементов на селектор (шапка+подвал+
                         # блоки «в корзину»: похожие/ранее просмотренные/акции)

# План действий по проектам: какие страницы открыть и какие кнопки нажать
# (кнопки открытия модалок дают цели «*click»; сами формы НЕ отправляем).
# «ожидаемые» - идентификаторы js-целей, которые ДОЛЖНЫ сработать от этих
# действий: если такая не сработала - это красное «НЕ сработала». Цели без
# автодействия получают серый статус «Нет автодействия» (не шумим ложным красным).
ACTIONS = {
    'smu': {
        'страницы': [
            ('Главная',   'https://stalmetural.ru/',
             [# СРОЧНО (до автозакрытия): модалка «Ваш город - Москва?».
              # На боевом сайте кнопка = «Все верно» (button.city-popup__btn--yes),
              # в части версий - «Да (N)». Жмём «оставить город» = закрыть модалку,
              # иначе её оверлей блокирует ВСЕ клики страницы (→ куча «Не проверено»).
              '!button.city-popup__btn--yes, .city-popup__btn--yes, '
              'button:has-text("Все верно"), button:has-text("Да"), '
              '.city-confirm button:has-text("Да")',
              # «Заказать звонок» в шапке = #call-back-form → фиксирует callorderclick
              # (устаревшей цели call_ordering сайт больше не шлёт).
              '#call-back-form', '#txt-back-form', '#txt-back-form-footer',
              '#call-back-form-main, [class*="manager-connect"], a:has-text("Связаться с менеджером")',
              'a:has-text("Показать больше категорий")',                   # morecatalog
              'a:has-text("Перейти ко всем категориям каталога")',         # gotomorecatalog
              'a:has-text("категориям производства"), a:has-text("ко всем категориям произ")',
              'button:has-text("Показать больше"), a:has-text("Показать больше")',
              'a:has-text("ко всем категориям"), a:has-text("всем категориям")']),
            ('Контакты',  'https://stalmetural.ru/contacts/',
             [# смена города: «изменить» в шапке → выбрать другой город (izmenit_gorod)
              {'цепочка': ['a:has-text("изменить"), text=изменить',
                           'a:has-text("Санкт-Петербург"), a:has-text("Казань")']}]),
            # breadcrumbphone - клик по номеру в «хлебных крошках» каталога
            ('Каталог',   'https://stalmetural.ru/catalog/',
             ['.breadcrumbs a[href^="tel:"], [class*="breadcrumb"] a[href^="tel:"]',
              'button:has-text("Показать больше"), a:has-text("Показать больше")',
              'a:has-text("ко всем категориям"), a:has-text("всем категориям")']),
            # Листинг труб (реальные карточки с кнопкой «в корзину»): tocart/addocart,
            # «Расчет стоимости» (.cost-calc) → в модалке «В корзину»
            # (raschetst/raschetaddtocart), «Скачать прайс-лист» ловится onclick-генериком.
            ('Листинг',   'https://stalmetural.ru/catalog/truba-profilnaya/',
             [# кнопка «В корзину» на СМУ - это div.btn.btn-catalog (не <button>/<a>)
              '.add-to-cart, div.btn:has-text("В корзину"), .btn-catalog:has-text("В корзину"), button:has-text("В корзину"), a:has-text("В корзину"), text=Добавить в корзину',
              # «Расчет стоимости» (div.cost-calc) открывает попап → в нём «В корзину»
              # (div.btn) даёт raschetst + raschetaddtocart. Кнопка - div, не <button>.
              {'цепочка': ['div.cost-calc >> visible=true',
                           '[class*="popup"] div.btn:has-text("В корзину"), .modal div.btn:has-text("В корзину"), [class*="popup"] button:has-text("В корзину"), div.btn:has-text("В корзину")']},
              '.one-click-to-buy, text=Купить в один клик',
              'text=Срочный заказ, text=Быстрый заказ',
              'text=Не нашли что искали, text=Не нашли']),
            # «Скачать прайс-лист» (a.btn, onclick reachGoal) → price_download_category
            ('Листинг (прайс)', 'https://stalmetural.ru/catalog/list-goryachekatanyy/',
             ['a.btn:has-text("Скачать прайс"), a:has-text("Скачать прайс-лист"), text=Скачать прайс']),
            # ВАЖНО: товар кладём в корзину ДО страницы «Корзина», иначе корзина пуста
            # и поля купона нет. «Добавить в корзину» на карточке реально добавляет
            # позицию (как в «Проверке форм»). Заодно tocart/addocart.
            ('Товар (труба)', 'https://stalmetural.ru/catalog/truba-profilnaya/2972110-truba-profilnaya-100kh10-mm-gost-8639-82-kvadratnaya/',
             ['div.btn:has-text("Добавить в корзину"), .add-to-cart, text=Добавить в корзину, text=В корзину',
              '.one-click-to-buy']),
            # Корзина (в ней уже есть товар): клик по полю купона + ввод буквы → coupon.
            ('Корзина',   'https://stalmetural.ru/basket/',
             [{'цепочка': [{'ввод': '.basket-coupon-block-field input, input[id*="coupon" i], '
                                    'input[name*="coupon" i], input[placeholder*="упон" i], '
                                    'input[placeholder*="ромокод" i], .coupon input, [class*="coupon"] input',
                            'текст': 'а'}]}]),
            ('Товар',     'https://stalmetural.ru/catalog/izgotovlenie-pruzhin/1285453-izgotovlenie-pruzhin-rastyazheniya/',
             ['.one-click-to-buy', '#call-back-form-product',
              '.copy-btn:has(.an-ico-link-price)',
              '[class*="favorite"], [class*="favourite"], [class*="to-fav"], button:has(.an-ico-heart)',
              '[class*="share"], .an-ico-share, button:has-text("Поделиться")',
              '.add-to-cart, text=Добавить в корзину, text=В корзину']),
            ('Доставка',  'https://stalmetural.ru/delivery/', ['#call-back-form-delivery']),
            ('Вакансии',  'https://stalmetural.ru/vacancy/',
             ['text=Откликнуться, text=Отклик, button:has-text("Откликнуться")']),
            # Избранное/поделиться настроены ТОЛЬКО на хабаровском поддомене -
            # ловим click_favorites/click_share именно там.
            ('Товар (Хабаровск)', 'https://habarovsk.stalmetural.ru/catalog/truba-profilnaya/2972110-truba-profilnaya-100kh10-mm-gost-8639-82-kvadratnaya/',
             ['[class*="favorite"], [class*="favourite"], [class*="to-fav"], button:has(.an-ico-heart)',
              '[class*="share"], .an-ico-share, button:has-text("Поделиться")']),
            ('Избранное (Хабаровск)', 'https://habarovsk.stalmetural.ru/favorites/', []),
            # Поиск: заходим на страницу результатов - url-цель «/search/» засчитывается.
            ('Поиск', 'https://stalmetural.ru/search/?q=труба', []),
            # 404: несуществующий адрес - должна сработать цель 404error
            ('Страница 404', 'https://stalmetural.ru/nesuschestvuyushaya-404-xyz/', []),
        ],
        'ожидаемые': [
            'tel', 'email', 'clickwapp', 'clicktg', 'clickvk', 'clickmax',
            'click_vk_podval', 'click_ok_podval', 'click_tg_podval',
            'click_dzen_podval', 'click_rutube_podval', 'click_max_podval',
            'breadcrumbphone',
            'callorderclick', 'zayavkaclick', 'svyazclick', 'oneclickbuy',
            'morecatalog', 'gotomorecatalog', 'moreuslugi',
            'moreproizvodstvo', 'click_favorites', 'click_share', 'addocart',
            'tocart', '404error', 'click_yes_confirm',
            'izmenit_gorod', 'raschetst', 'raschetaddtocart',
            'price_download_category',
        ],
        # На РФ-сайте такой кнопки/формы нет (статус «Не найдена на сайте»).
        # Значение = точное пояснение, почему цель не проверяется на РФ.
        'нет_на_сайте': {
            'phone_header': 'клика по этой кнопке в шапке на сайте нет '
                            '(старый идентификатор Метрики)',
            'phone_footer': 'клика по этой кнопке в подвале на сайте нет '
                            '(старый идентификатор Метрики)',
            'zvonok_text_category': 'такой формы «Заказать звонок» на сайте нет',
            'managerclick': 'кнопка «Связаться с менеджером» есть только на сайтах СНГ, '
                            'на РФ её нет - проверяется в отчётах стран СНГ',
            'managerform': 'форма «Связаться с менеджером» есть только на сайтах СНГ, '
                           'на РФ её нет - проверяется в отчётах стран СНГ',
            'subscribeform': 'форма подписки на рассылку есть только на Хабаровске - '
                             'проверяется «Проверкой форм» для города Хабаровск',
            'click_favorites': 'кнопка «Добавить в избранное» настроена только на '
                               'Хабаровске (если Хабаровск в прогоне - цель зелёная)',
            'click_share': 'кнопка «Поделиться» настроена только на Хабаровске '
                           '(если Хабаровск в прогоне - цель зелёная)',
            'click_yandexorg_podval': 'это виджет Яндекс-карты (iframe) в подвале - '
                                      'клик внутрь чужой карты автотест не '
                                      'воспроизводит, проверьте вручную',
        },
        # СМЕЖНЫЕ ЦЕЛИ: на кнопке/форме РЕАЛЬНО срабатывает один идентификатор, а в
        # Метрике цель заведена под другим (устаревшим/переименованным). Если
        # «фактический» ID сработал в прогоне - цель считаем достигнутой и пишем в
        # пояснении: «ловится вот такой, а в Метрике вот такой».
        #   {ожидаемый_id: (фактический_id, пояснение)}
        'смежные': {
            'call_ordering': ('callorderclick',
                              'кнопка «Заказать звонок» в шапке'),
            'raschet_stoimosti_dostavki': ('callorderform',
                              'форма «Расчёт стоимости доставки»'),
        },
    },
    'imp': {
        'страницы': [
            ('Главная',   'https://inmetprom.ru/',
             [# ШАПКА: заявка/просчёт, каталог (klik-na-katalog-v-shapke), акции, прайс
              '[data-my-modal="#modal-callback"], .banner-fast-order__application',
              'header a:has-text("Каталог"), .header a:has-text("Каталог")',
              'header a:has-text("Акции"), a:has-text("Акции")',
              'header a:has-text("Прайс"), a:has-text("Прайс-лист")',
              # СМЕНА ГОРОДА: открыть модалку города → кликнуть другой город (izmenit_gorod)
              {'цепочка': ['[data-my-modal="#city-modal-id"], .choose-city',
                           '.city-modal__city:not(.active), #city-modal-id a:has-text("Санкт-Петербург")']},
              # ПОДВАЛ: звонок менеджеру, написать нам, городской телефон
              '.footer-manager-call, footer a:has-text("Звонок менеджеру")',
              '.footer-email, footer a:has-text("Написать нам")',
              '.footer-phone, footer .telephone-utf']),
            ('Контакты',  'https://inmetprom.ru/contacts/',
             ['.footer-manager-call', '.footer-email']),
            ('Листинг',   'https://inmetprom.ru/catalog/reshetchatyj-nastil/',
             ['a:has-text("Скачать прайс-лист"), button:has-text("Скачать прайс-лист"), a:has-text("Скачать прайс")',
              '.add-to-cart-btn',
              'text=Быстрый заказ, [class*="fast-order"], [class*="bystryy"]',
              '.tags a, [class*="tag"] a, [class*="tags"] a',
              'button:has-text("Показать ещё"), a:has-text("Показать ещё")']),
            # Спецпредложения/акции: «в корзину» и «быстрый заказ» карточек акций.
            ('Акции',     'https://inmetprom.ru/specials/',
             ['.add-to-cart-btn, button:has-text("В корзину")',
              'text=Быстрый заказ, [class*="fast-order"], [class*="bystryy"]']),
            ('Товар',     'https://inmetprom.ru/list-gesti-0-2-mm-klass-1-gost-13345-85/',
             ['.add-to-cart-btn',                          # «в корзину» ВСЕХ блоков (похожие/ранее/с этим товаром/акции)
              'text=Быстрый заказ, [class*="fast-order"], [class*="bystryy"]',
              'text=Оставить отзыв, text=Написать отзыв, a:has-text("Отзыв")',
              'text=Нужна консультация, [class*="konsultac"]']),
            ('Вакансии',  'https://inmetprom.ru/vakansii/',
             ['text=Узнать подробнее, text=Откликнуться, text=Интересует вакансия']),
            ('Страница 404', 'https://inmetprom.ru/nesuschestvuyushaya-404-xyz/', []),
        ],
        'ожидаемые': [
            'tel-shapka-gorod', 'klik-v-shapke-po-elektronnoy-pochte',
            'klik-v-shapke-na-whatsapp', 'klik-na-tg-v-shapke',
            'klik-na-prays-list-iz-shapki', 'klik-na-katalog-v-shapke',
            'click-proschet', 'klik-na-aktsii', 'click-podval-gorod-nomer',
            'klik-v-podvale-zvonok-menedzheru', 'click-podval-email',
            'klik-v-podvale-po-napisat-nam', 'klik-na-tg-v-mobilke',
            'klik-na-whatsapp-v-mobilke', 'click-telephone-utf-gorod',
            'izmenit_gorod', 'click_yes_confirm', 'bystryy-zakaz-katalog',
            'bystryy-zakaz-listing', 'v-cart-listing', 'v-cart-listing-img',
            'bistrii-zakaz-listing', 'bistrii-zakaz-listing-img',
            'listing-skachat-prays-list', 'listing-klik-po-tegam', 'v-cart-kartochka',
            'bistrii-zakaz-cartochka', 'v-cart-kartochka-ranee-prosmotrennye',
            'bistrii-zakaz-cartochka-ranee-prosmotrennye',
            'v-korzinu-kartochka-pokhozhiye-tovary',
            'bystryy-zakaz-kartochka-pokhozhiye-tovary',
            'v-korzinu-kartochka-s-etim-tovarom-pokupayut',
            'bystryy-zakaz-kartochka-s-etim-tovarom-pokupayut',
            'v-korzinu-kartochka-tovara-aktsii', 'bystryy-zakaz-kartochka-tovara-aktsii',
            'tovar_konsultaciya', 'klik-tel', '404error',
        ],
    },
    'mpe': {
        'страницы': [
            # Цели МПЭ прошиты в onclick="ym(...,'reachGoal','X')" - генерик-кликер
            # движка снимает их сам (login/citys/about/catalog/smotretvse/raschet/
            # rekvizity_podval). Ниже - только то, что требует особых шагов.
            ('Главная',   'https://mepen.ru/',
             # rasschitatzakaz и raschet_stoimosti_dostavki - скрытые input.d-none с
             # onclick reachGoal; их снимает dispatch-проход по input (см. движок).
             ['header.header-kostyl .bottom-header-right button.popup_form',
              'a.link_more',                               # «Узнать больше» → about
              'a.footer-link[href="/catalog/"]']),         # «Смотреть все» в подвале → smotretvse
            ('Контакты',  'https://mepen.ru/contacts/',
             ['[onclick*="rekvizity_contacts"], a:has-text("Реквизиты компании")']),
            # Реквизиты: «Скачать реквизиты» → skachat_rekvizity
            ('Реквизиты', 'https://mepen.ru/rekvizity/',
             ['[onclick*="skachat_rekvizity"], a:has-text("Скачать реквизиты")']),
            # Листинг с корзиной (болты): «В корзину» → tocart, клик по карточке →
            # klik_kartochka_tovara.
            ('Листинг (болты)', 'https://mepen.ru/catalog/zheleznodorozhnaya-avtomatika/zheleznodorozhnyy-krepezh/bolt/',
             ['[onclick*="tocart"], button:has-text("В корзину"), text=В корзину',
              '[onclick*="klik_kartochka_tovara"]']),
            # Товар кладём в корзину ДО «Корзины» (иначе поля купона нет).
            ('Товар',     'https://mepen.ru/catalog/tovar/telezhka-tip-b-gcl/',
             ['[onclick*="tovar_v_korzinu"], text=в корзину, text=В корзину',
              'text=Нужна консультация', 'text=Нашли дешевле']),
            # Корзина (товар уже добавлен): клик по полю купона → skidochnyy_kupon.
            ('Корзина',   'https://mepen.ru/personal/basket/',
             ['[onclick*="skidochnyy_kupon"], text=Введите код купона, '
              'input[class*="coupon"], [class*="coupon"] input, text=Купон']),
            # Авторизация: иконка аккаунта (a.lk_link.personal_popups → i.fa-user)
            # открывает модалку «Авторизация», в ней кнопка «Авторизоваться»
            # (button type=submit, onclick reachGoal klik_avtorizovatsya). Форма
            # пустая - вход не произойдёт, но onclick срабатывает до отправки.
            ('Авторизация', 'https://mepen.ru/',
             [{'цепочка': ['a.lk_link.personal_popups, a[href="/personal/auth.php"], '
                           'a[data-idform="auth"], .top-header-right a:has(.fa-user)',
                           {'dispatch': 'button[onclick*="klik_avtorizovatsya"], '
                                        '#auth_form button[type="submit"], '
                                        '.modal button:has-text("Авторизоваться")'}]}]),
            # Поиск: заходим на страницу результатов - url-цель «/search/» засчитывается.
            ('Поиск',     'https://mepen.ru/search/?q=болт', []),
            ('Страница 404', 'https://mepen.ru/nesuschestvuyushaya-404-xyz/', []),
        ],
        'ожидаемые': [
            'tel', 'email', 'catalog', 'raschet', 'rasschitatzakaz', 'smotretvse',
            'telegram', 'clickwapp', 'tocart', 'tovar_v_korzinu',
            'klik_kartochka_tovara', 'tovar_konsultaciya', 'klik_nashli_deshevle',
            'citys', 'about', 'rekvizity_podval', 'rekvizity_contacts',
            'skachat_rekvizity', 'skidochnyy_kupon', 'klik_avtorizovatsya',
        ],
    },
}


def загрузить_каталог(pid: str) -> dict | None:
    f = CATALOGS / f'goals-{pid}.json'
    if not f.is_file():
        return None
    return json.loads(f.read_text(encoding='utf-8'))


# Цели, которые настроены ТОЛЬКО на хабаровском поддомене СМУ (РФ). На всех
# остальных сайтах СМУ (Москва и страны СНГ) их нет - помечаем «только Хабаровск»,
# чтобы не считать ошибкой. На РФ-прогоне мы заходим на Хабаровск и ловим их вживую.
_ХАБАРОВСК_ТОЛЬКО = {
    'click_favorites': 'кнопка «Добавить в избранное» настроена только на '
                       'хабаровском поддомене - на этом сайте её нет',
    'click_share': 'кнопка «Поделиться» настроена только на хабаровском '
                   'поддомене - на этом сайте её нет',
    'subscribeform': 'форма подписки на рассылку есть только на хабаровском '
                     'поддомене - на этом сайте её нет',
}

# Формы/кнопки, которых нет НИ НА ОДНОМ сайте СМУ (устаревшие идентификаторы в
# Метрике) - помечаем «Не найдена на сайте» и на РФ, и на странах.
_НЕТ_ВЕЗДЕ = {
    'zvonok_text_category': 'такой формы «Заказать звонок» (шаблонный текст '
                            'категории) на сайте нет - старый идентификатор Метрики',
}


def _план_для_домена(домен: str) -> dict:
    """Универсальный план прогона для сайта на той же платформе, что СМУ РФ
    (СНГ-домены stalmetural.*, smg.az и т.п.): те же кнопки и ожидаемые цели,
    только другой домен. Товар/доставку не трогаем - их адреса у стран разные."""
    d = (домен or '').rstrip('/')
    if not d:
        return {'страницы': []}
    return {
        'страницы': [
            ('Главная', d + '/',
             ['#call-back-form', '#txt-back-form', '#txt-back-form-footer',
              '#call-back-form-main, [class*="manager-connect"], a:has-text("Связаться с менеджером")',
              'button:has-text("Показать больше"), a:has-text("Показать больше")',
              'a:has-text("ко всем категориям"), a:has-text("всем категориям")']),
            ('Контакты', d + '/contacts/', []),
            ('Каталог', d + '/catalog/',
             ['.breadcrumbs a[href^="tel:"], [class*="breadcrumb"] a[href^="tel:"]',
              'button:has-text("Показать больше"), a:has-text("Показать больше")',
              'a:has-text("ко всем категориям"), a:has-text("всем категориям")']),
            ('Поиск', d + '/search/?q=труба', []),   # url-цель «/search/»
            ('Страница 404', d + '/nesuschestvuyushaya-404-xyz/', []),
        ],
        'ожидаемые': ACTIONS['smu']['ожидаемые'],
        # На страны Хабаровск-цели не распространяются + устаревшие формы (zvonok…).
        'нет_на_сайте': {**_ХАБАРОВСК_ТОЛЬКО, **_НЕТ_ВЕЗДЕ},
    }


def _план_мпэ_для_домена(домен: str) -> dict:
    """План прогона МПЭ для страны СНГ: те же СТРУКТУРНЫЕ страницы, что у РФ
    (реквизиты, корзина, авторизация, поиск), с подменённым доменом. Цели МПЭ
    прошиты в общем JS-бандле, поэтому заходить надо на те же разделы, что и на
    РФ - иначе сканер reachGoal их не увидит (страны раньше гоняли СМУ-шаблон!)."""
    d = (домен or '').rstrip('/')
    if not d:
        return {'страницы': []}
    _auth = {'цепочка': ['a.lk_link.personal_popups, a[href="/personal/auth.php"], '
                         'a[data-idform="auth"], .top-header-right a:has(.fa-user)',
                         {'dispatch': 'button[onclick*="klik_avtorizovatsya"], '
                                      '#auth_form button[type="submit"], '
                                      '.modal button:has-text("Авторизоваться")'}]}
    return {
        'страницы': [
            ('Главная',   d + '/',
             ['a.link_more', 'a.footer-link[href="/catalog/"]']),
            ('Контакты',  d + '/contacts/',
             ['[onclick*="rekvizity_contacts"], a:has-text("Реквизиты компании")']),
            ('Реквизиты', d + '/rekvizity/',
             ['[onclick*="skachat_rekvizity"], a:has-text("Скачать реквизиты")']),
            ('Каталог',   d + '/catalog/', []),
            ('Корзина',   d + '/personal/basket/',
             ['[onclick*="skidochnyy_kupon"], [class*="coupon"] input']),
            ('Авторизация', d + '/', [_auth]),
            ('Поиск',     d + '/search/?q=bolt', []),
            ('Страница 404', d + '/nesuschestvuyushaya-404-xyz/', []),
        ],
        'ожидаемые': ACTIONS['mpe']['ожидаемые'],
    }


def _план_страна(pid: str, домен: str) -> dict:
    """Универсальный план для суб-проекта страны - по базовому проекту."""
    if _базовый(pid) == 'mpe':
        return _план_мпэ_для_домена(домен)
    return _план_для_домена(домен)


def _базовый(pid: str) -> str:
    """Базовый проект для суб-проекта страны: smu-uz → smu (формы и их конфиг
    едины для всех стран проекта, лежат под базовым кодом)."""
    return (pid or '').split('-')[0]


def _формные_цели(pid: str) -> set[str]:
    """Идентификаторы целей, привязанных к ОТПРАВКЕ форм в конфиге форм-тестера
    (их не триггерим здесь, чтобы не слать заявки)."""
    p = ROOT / 'forms_tester' / 'projects' / _базовый(pid) / 'config.py'
    if not p.is_file():
        return set()
    txt = p.read_text(encoding='utf-8')
    return set(re.findall(r'"цель"\s*:\s*"([\w\-.]+)"', txt))


def _формные_url(pid: str) -> list[str]:
    """URL, до которых дошёл прогон форм (в т.ч. страница после оформления заказа).
    По ним подтверждаем url-цели, на которые обычный прогон целей не попадает."""
    base = _базовый(pid)
    f = ROOT / 'cache' / 'forms' / base / 'fired_urls.json'
    if f.is_file():
        try:
            return [str(u) for u in json.loads(f.read_text(encoding='utf-8')) if u]
        except Exception:
            return []
    return []


def _результаты_форм(pid: str) -> dict[str, str]:
    """Статусы целей из последнего прогона форм: идентификатор → статус.
    Берём из ДВУХ источников:
      1) fired_goals.json - ВСЕ цели, сработавшие при формах (ловятся из вывода
         форм-прогона, включая те, что движок пишет только в лог);
      2) лист «Цели» отчёта форм (цели спец-сценариев)."""
    base = _базовый(pid)
    out: dict[str, str] = {}
    # 1) Все пойманные при формах цели.
    fj = ROOT / 'cache' / 'forms' / base / 'fired_goals.json'
    if fj.is_file():
        try:
            for gid in json.loads(fj.read_text(encoding='utf-8')):
                gid = str(gid).strip()
                if gid:
                    out[gid] = 'Сработала'
        except Exception:
            pass
    # 2) Лист «Цели» отчёта форм (статусы спец-сценариев дополняют/уточняют).
    f = ROOT / 'cache' / 'forms' / base / 'log_forms.xlsx'
    if not f.is_file():
        return out
    try:
        from openpyxl import load_workbook
        wb = load_workbook(f, data_only=True)
        if 'Цели' not in wb.sheetnames:
            return out
        ws = wb['Цели']
        hdr = [str(c.value or '').strip() for c in ws[1]]
        i_id = hdr.index('Цель (идентификатор)') if 'Цель (идентификатор)' in hdr else -1
        i_st = hdr.index('Статус') if 'Статус' in hdr else -1
        if i_id < 0 or i_st < 0:
            return out
        for row in ws.iter_rows(min_row=2, values_only=True):
            gid = str(row[i_id] or '').strip()
            st = str(row[i_st] or '').strip()
            if gid:
                out[gid] = st          # последняя строка по цели побеждает
    except Exception:
        pass
    return out


def _совпало(цель: dict, fired: set[str]) -> str | None:
    """Какой из сработавших идентификаторов закрывает цель (учитывая «содержит»)."""
    ids = цель.get('идентификаторы') or []
    for gid in ids:
        if цель.get('содержит'):
            for f in fired:
                if gid.lower() in f.lower():
                    return f
        elif gid in fired:
            return gid
    return None


def выполнить_прогон(pid: str, headless: bool = True, log=print, stop=None) -> dict:
    """Открывает страницы, кликает, слушает Метрику. Возвращает
    {'fired': set(id), 'страницы': [{'название','url','код','счётчик','визит'}]}."""
    каталог = загрузить_каталог(pid) or {}
    # Явный план проекта; для суб-проектов (страны СМУ) - универсальный по домену.
    план = ACTIONS.get(pid) or _план_страна(pid, каталог.get('домен', ''))
    counter = str(каталог.get('счётчик') or '')
    fired: set[str] = set()
    привязки: set[str] = set()       # reachGoal-идентификаторы, найденные явно
    визиты: dict[str, bool] = {}     # url → watch-hit отправлен
    страницы_инфо = []
    _re_reach = re.compile(r"reachGoal\W{1,4}([\w\-]+)")
    _seen_js: set[str] = set()
    _код_части: list[str] = []        # весь код страниц + их JS (для поиска целей)

    def _собрать_привязки(html: str, base_url: str):
        """Копит reachGoal-идентификаторы И ВЕСЬ код (HTML + все JS того же хоста),
        чтобы потом надёжно проверить, упоминается ли цель в коде сайта вообще
        (как строковый литерал), а не только сразу после reachGoal(."""
        привязки.update(_re_reach.findall(html))
        _код_части.append(html.lower())
        try:
            import requests as _rq
            from concurrent.futures import ThreadPoolExecutor
            host = re.sub(r'^https?://', '', base_url).split('/')[0].split(':')[0]
            base_host = '.'.join(host.split('.')[-2:])   # stalmetural.ru
            srcs = re.findall(r'<script[^>]+src=["\']([^"\']+)["\']', html)
            _to_fetch = []
            for src in srcs[:40]:
                u = src if src.startswith('http') else urljoin(base_url, src)
                # берём JS с того же домена/поддоменов (там и живут reachGoal),
                # пропускаем очевидные библиотеки (в них целей нет)
                if base_host not in u or u in _seen_js:
                    continue
                if re.search(r'(jquery|bootstrap|swiper|slick|popper|fancybox|'
                             r'owl\.carousel|lazyload|polyfill)', u, re.I):
                    continue
                _seen_js.add(u)
                _to_fetch.append(u)

            def _fetch_js(u):
                try:
                    return _rq.get(u, timeout=6,
                                   headers={'User-Agent': 'Mozilla/5.0'},
                                   verify=os.environ.get('REQUESTS_CA_BUNDLE', True)).text
                except Exception:
                    return ''
            # ПАРАЛЛЕЛЬНО (раньше 40 файлов по 15с последовательно = до 2 мин на
            # страницу; теперь пул потоков + таймаут 6с = секунды).
            if _to_fetch:
                with ThreadPoolExecutor(max_workers=8) as _ex:
                    for js in _ex.map(_fetch_js, _to_fetch):
                        if js:
                            привязки.update(_re_reach.findall(js))
                            _код_части.append(js.lower())
        except Exception:
            pass

    from playwright.sync_api import sync_playwright
    # Облачная среда (агентский прокси режет TLS браузера): гоняем трафик страницы
    # через сетевой стек драйвера (route.fetch). Локально флага нет - напрямую.
    _via_driver = bool(os.environ.get('CCR_AGENT_PROXY_ENABLED'))
    with sync_playwright() as pw:
        b = pw.chromium.launch(headless=headless,
                               args=["--disable-blink-features=AutomationControlled",
                                     "--no-sandbox"])
        # Прикидываемся обычным Chrome: часть сайтов (напр. inmetprom.ru) отдаёт
        # 403 «голому» headless-браузеру. Реальный User-Agent + заголовки часто
        # снимают такую блокировку.
        _UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
               "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36")
        ctx = b.new_context(locale='ru-RU', viewport={'width': 1440, 'height': 900},
                            ignore_https_errors=_via_driver, user_agent=_UA,
                            extra_http_headers={
                                'Accept-Language': 'ru-RU,ru;q=0.9,en;q=0.8',
                                'Accept': ('text/html,application/xhtml+xml,'
                                           'application/xml;q=0.9,image/avif,'
                                           'image/webp,*/*;q=0.8'),
                            })
        ctx.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
        try:
            ctx.clear_cookies()   # чтобы вылезла модалка выбора города (кука не стоит)
        except Exception:
            pass
        page = ctx.new_page()

        текущий_url = {'u': ''}

        def _на_запрос(req):
            u = req.url
            if 'mc.yandex' in u or 'mc.webvisor' in u:
                m = _RE_GOAL.search(unquote(u))
                if m:
                    gid = m.group(1)
                    if gid not in fired:
                        fired.add(gid)
                        log(f"   🎯 цель: {gid}")
                if f'/watch/{counter}' in u:
                    визиты[текущий_url['u']] = True

        ctx.on('request', _на_запрос)
        # новые вкладки (клики по соцсетям с target=_blank) сразу закрываем
        ctx.on('page', lambda p: p != page and p.close())
        if _via_driver:
            def _route(route, request):
                try:
                    route.fulfill(response=route.fetch(timeout=40000))
                except Exception:
                    try:
                        route.abort()
                    except Exception:
                        pass
            ctx.route('**/*', _route)

        _всего = len(план['страницы'])
        _модалка_нет_сделана = False   # «Нет» в модалке города жмём один раз за прогон
        for _idx, (название, url, клики) in enumerate(план['страницы'], 1):
            if stop and stop():
                log('⛔ Остановлено')
                break
            log(f"ПРОГРЕСС {_idx}/{_всего}")
            log(f"- Страница: {название}  {url}")
            текущий_url['u'] = url
            код = 0
            try:
                resp = page.goto(url, wait_until='domcontentloaded', timeout=45000)
                код = resp.status if resp else 0
                # 5xx часто транзиентны (сервер моргнул) - пробуем ещё пару раз,
                # прежде чем метить сайт «недоступен».
                for _try in range(2):
                    if код < 500:
                        break
                    log(f"   ↻ код {код}, повтор через 3с…")
                    page.wait_for_timeout(3000)
                    resp = page.goto(url, wait_until='domcontentloaded', timeout=45000)
                    код = resp.status if resp else код
                page.wait_for_timeout(1000)
            except Exception as e:
                log(f"   ⚠️ не открылась: {e}")
                страницы_инфо.append({'название': название, 'url': url, 'код': код,
                                      'счётчик': False, 'визит': False})
                continue
            # ОДИН РАЗ за прогон: жмём «Нет»/«Выбрать город» в модалке города, чтобы
            # поймать цель click_no_confirm (её иначе не проверить - мы всегда
            # оставляем город через «Да»). После этого перезагружаем страницу -
            # модалка появится снова, и ниже её закроет «Да» как обычно.
            # ОБЕ цели модалки города (один раз за прогон): dispatch по «Нет» и «Да».
            # dispatch (untrusted) вызывает reachGoal сразу, но НЕ открывает пикер и
            # НЕ закрывает модалку по-настоящему - поэтому ловим и click_no_confirm, и
            # click_yes_confirm, а реальное закрытие ниже делает «Да».
            if not _модалка_нет_сделана:
                for _btn in ('.city-confirm-popup__btn--no', '.city-popup__btn--no',
                             '[class*="city"] button:has-text("Выбрать город")',
                             '.city-confirm-popup__btn--yes', '.city-popup__btn--yes',
                             '[class*="city"] button:has-text("Да")',
                             'button:has-text("Все верно")'):
                    try:
                        el = page.locator(_btn).first
                        if el.count():
                            el.dispatch_event('click')
                            page.wait_for_timeout(250)
                    except Exception:
                        continue
                _модалка_нет_сделана = True
            # УНИВЕРСАЛЬНОЕ закрытие модалки выбора города (СРАЗУ после загрузки).
            # Разные проекты - разные кнопки «оставить текущий город»:
            #   СМУ: «Все верно» (button.city-popup__btn--yes) или «Да (N)»;
            #   МПЭ: «Вы находитесь в г. Москва?» → «Да»;
            #   страны СНГ: свои варианты.
            # Если её не закрыть - прозрачный оверлей перехватывает ВСЕ клики
            # страницы и почти все цели выходят «Не проверено».
            for _cm in ('button.city-popup__btn--yes', '.city-popup__btn--yes',
                        '.city-confirm-popup__btn--yes',
                        '[class*="city"] button:has-text("Да")',
                        '.city-confirm button:has-text("Да")',
                        'button:has-text("Все верно")',
                        '[class*="popup"] button:has-text("Да")'):
                try:
                    el = page.locator(_cm).first
                    if el.count() and el.is_visible():
                        el.click(timeout=2500, force=True)
                        page.wait_for_timeout(500)
                        break
                except Exception:
                    continue
            # СРОЧНЫЕ клики (префикс '!'): выполняются сразу после загрузки -
            # до прокрутки и генериков. Нужны для модалки города «Да/Нет»,
            # которая автозакрывается через несколько секунд.
            for sel in [s[1:] for s in клики
                        if isinstance(s, str) and s.startswith('!')]:
                try:
                    el = page.locator(sel).first
                    el.click(timeout=3500)
                    page.wait_for_timeout(500)
                except Exception:
                    pass

            html = page.content()
            есть_счётчик = counter in html if counter else False
            _собрать_привязки(html, url)

            # прокрутка вниз (ленивые блоки + подвал с соцсетями)
            try:
                page.mouse.wheel(0, 20000)
                page.wait_for_timeout(700)
            except Exception:
                pass

            # общие безопасные «клики» (тел/почта/соцсети) - через dispatch_event:
            # событие click всплывает, Метрика ловит внешнюю ссылку своим слушателем,
            # но БЕЗ перехода по href и без новой вкладки (быстро, без go_back).
            for sel in GENERIC_CLICK_SELECTORS:
                try:
                    els = page.locator(sel)
                    n = min(els.count(), MAX_PER_SELECTOR)
                    for i in range(n):
                        try:
                            els.nth(i).dispatch_event('click')
                            page.wait_for_timeout(110)
                        except Exception:
                            continue
                except Exception:
                    continue

            # ЭЛЕМЕНТЫ С ЦЕЛЬЮ В onclick: на СМУ/МПЭ цели прошиты прямо в
            # onclick="ym(...,'reachGoal','X')" (в т.ч. скрытые input и submit-кнопки).
            # Скорость: ВСЕ строки onclick забираем ОДНИМ page.evaluate (раньше был
            # get_attribute на каждый из ~40 «городов» = десятки round-trip'ов),
            # дедупим по id в Python, а дальше жмём проверенным Playwright
            # dispatch_event (вызывает onclick/reachGoal БЕЗ перехода по ссылке и БЕЗ
            # реального сабмита - пустая заявка не уходит).
            _RG_SEL = ('a[onclick*="reachGoal"], div[onclick*="reachGoal"], '
                       'span[onclick*="reachGoal"], button[onclick*="reachGoal"], '
                       'input[onclick*="reachGoal"]')
            try:
                _oncs = page.evaluate(
                    "(s) => [...document.querySelectorAll(s)].map(e => "
                    "e.getAttribute('onclick') || '')", _RG_SEL)
            except Exception:
                _oncs = []
            try:
                _all_rg = page.locator(_RG_SEL)
                _видели_цель: set[str] = set()
                _idx: list[int] = []
                for _i, _oc in enumerate(_oncs):
                    _m = re.search(r"reachGoal[^)]*['\"]([\w\-.]+)['\"]", _oc)
                    _gid = _m.group(1) if _m else f'#{_i}'
                    if _gid in _видели_цель:
                        continue
                    _видели_цель.add(_gid)
                    _idx.append(_i)
                    if len(_idx) >= 60:
                        break
                for _i in _idx:
                    try:
                        _all_rg.nth(_i).dispatch_event('click')
                    except Exception:
                        continue
                page.wait_for_timeout(700)   # даём beacon'ам уйти
                page.keyboard.press('Escape')
                page.wait_for_timeout(120)
            except Exception:
                pass

            # подвал: соцсети/мессенджеры - dispatch_event (событие всплывает,
            # Метрика ловит внешнюю ссылку, но без новой вкладки и без ухода).
            _soc = ("vk.com", "vk.me", "ok.ru", "t.me", "dzen.ru", "rutube.ru",
                    "max.ru", "wa.me", "whatsapp", "yandex.ru/maps",
                    "yandex.ru/profile")
            try:
                foot = page.locator("footer a[href], .footer a[href]")
                for i in range(min(foot.count(), 25)):
                    try:
                        el = foot.nth(i)
                        href = (el.get_attribute("href") or "").lower()
                        if not any(s in href for s in _soc):
                            continue
                        el.dispatch_event('click')
                        page.wait_for_timeout(90)
                    except Exception:
                        continue
            except Exception:
                pass

            # клики проекта (кнопки модалок форм, «в корзину» разных блоков и т.п.).
            # Кликаем НЕСКОЛЬКО элементов на селектор: на карточке товара кнопки
            # «в корзину»/«быстрый заказ» повторяются в блоках «похожие», «ранее
            # просмотренные», «с этим товаром покупают», «акции» - у каждой СВОЯ
            # цель. Escape ДО клика снимает модалку прошлого (перекрытие - причина
            # «через раз»), Escape ПОСЛЕ закрывает открытую.
            for sel in клики:
                if isinstance(sel, str) and sel.startswith('!'):
                    continue        # срочный - уже выполнен сразу после загрузки
                # ЦЕПОЧКА: {'цепочка': [шаг1, шаг2, ...]} - действия ПОДРЯД без
                # Escape между ними (модалка остаётся открытой). Шаг - селектор
                # (клик) или {'ввод': селектор, 'текст': 'а'} (клик + ввод текста,
                # напр. поле купона в корзине). Нужно, когда одна цель ведёт к
                # другой: «Расчёт стоимости» → в модалке «В корзину».
                if isinstance(sel, dict) and sel.get('цепочка'):
                    try:
                        page.keyboard.press('Escape')
                        page.wait_for_timeout(150)
                        for step in sel['цепочка']:
                            if isinstance(step, dict) and step.get('ввод'):
                                el = page.locator(step['ввод']).first
                                if el.count() == 0:
                                    break
                                el.scroll_into_view_if_needed(timeout=1500)
                                el.click(timeout=2500)
                                el.type(step.get('текст', 'а'), delay=120)
                                page.wait_for_timeout(900)
                                continue
                            if isinstance(step, dict) and step.get('dispatch'):
                                # dispatch_event: вызвать onclick (reachGoal) БЕЗ
                                # реального сабмита (для submit-кнопок вроде
                                # «Авторизоваться» - иначе форма уходит и цель теряется).
                                el = page.locator(step['dispatch']).first
                                if el.count() == 0:
                                    break
                                try:
                                    el.dispatch_event('click')
                                except Exception:
                                    pass
                                page.wait_for_timeout(700)
                                continue
                            el = page.locator(step).first
                            if el.count() == 0:
                                break
                            el.scroll_into_view_if_needed(timeout=1500)
                            try:
                                el.click(timeout=2500)
                            except Exception:
                                el.click(timeout=2500, force=True)
                            page.wait_for_timeout(900)
                        page.keyboard.press('Escape')
                        page.wait_for_timeout(200)
                        if page.url != url:
                            try:
                                page.go_back(wait_until='domcontentloaded', timeout=15000)
                            except Exception:
                                page.goto(url, wait_until='domcontentloaded', timeout=30000)
                            page.wait_for_timeout(600)
                    except Exception:
                        pass
                    continue
                try:
                    total = page.locator(sel).count()
                except Exception:
                    continue
                for i in range(min(total, MAX_PER_SELECTOR)):
                    try:
                        page.keyboard.press('Escape')
                        page.wait_for_timeout(100)
                        el = page.locator(sel).nth(i)
                        if el.count() == 0:
                            break
                        el.scroll_into_view_if_needed(timeout=1500)
                        try:
                            el.click(timeout=2000)
                        except Exception:
                            el.click(timeout=2000, force=True)
                        # держим паузу для beacon корзины (tocart/addocart уходят ajax-ом)
                        page.wait_for_timeout(650)
                        # клик мог быть по ссылке-переходу (Каталог/Акции/Прайс в
                        # шапке): цель сработала, но нужно вернуться и кликать дальше.
                        if page.url != url:
                            try:
                                page.go_back(wait_until='domcontentloaded', timeout=15000)
                            except Exception:
                                page.goto(url, wait_until='domcontentloaded', timeout=30000)
                            page.wait_for_timeout(500)
                            break   # после ухода нумерация сбилась - к следующему селектору
                        page.keyboard.press('Escape')
                        page.wait_for_timeout(100)
                    except Exception:
                        continue
            page.wait_for_timeout(200)

            страницы_инфо.append({'название': название, 'url': url, 'код': код,
                                  'счётчик': есть_счётчик,
                                  'визит': визиты.get(url, False)})
        b.close()

    return {'fired': fired, 'страницы': страницы_инфо, 'привязки': привязки,
            'код': ''.join(_код_части)}


def _url_цели_проверка(каталог: dict, страницы_инфо: list,
                       формные_url: list | None = None) -> dict[str, dict]:
    """Для url-целей: найти открытую страницу, чей адрес содержит «url_часть».
    Дополнительно проверяем URL, до которых дошёл прогон форм (formные_url):
    оформленный заказ / «спасибо» - на них обычный прогон целей не попадает, а
    формы реально доходят. Совпадение по ним даёт статус «Сработала (формы)»."""
    формные_url = формные_url or []
    out = {}
    for g in каталог.get('цели', []):
        if g['тип'] not in ('url', 'url_re'):
            continue
        часть = g.get('url_часть') or ''

        def _match(u: str) -> bool:
            try:
                return bool(re.search(часть, u) if g['тип'] == 'url_re'
                            else часть.lower() in (u or '').lower())
            except re.error:
                return False

        hit = None
        for s in страницы_инфо:
            if _match(s['url']):
                hit = s
                break
        через_формы = False
        if hit is None:                     # не нашли в обычном прогоне - ищем в формах
            for u in формные_url:
                if _match(u):
                    hit = {'название': 'Прогон форм', 'url': u, 'код': 200,
                           'счётчик': True, 'визит': True}
                    через_формы = True
                    break
        out[g['номер']] = {'страница': hit, 'через_формы': через_формы}
    return out


# Цели, которым нужно ОСОБОЕ действие (его автотест намеренно не делает: это не
# «нет в коде», а «нужен ручной шаг / отдельная настройка автотеста»). Ключи -
# подстроки в названии/условии/идентификаторе цели → человеческое описание.
_СПЕЦ_ДЕЙСТВИЯ = [
    (('вход', 'авториз', 'логин', 'login', 'акка', 'регистрац', 'registr'),
     'вход или регистрация в личном кабинете'),
    (('избранн', 'favorit', 'wishlist', 'сравн', 'sravn'),
     'добавление товара в избранное/сравнение'),
    (('купон', 'kupon', 'скидочн', 'промокод', 'promo'),
     'применение купона или промокода'),
    (('скачив', 'скачать', 'download', 'реквизит', 'rekvizit'),
     'скачивание файла или реквизитов'),
    (('смотреть', 'smotretvse', 'показать', 'показат', 'ещё', 'eshe', 'load', 'pagina'),
     'просмотр/подгрузка каталога (пагинация, «смотреть всё»)'),
    (('оплат', 'оформ', 'oformit', 'checkout'),
     'оформление или оплата заказа'),
]


def _нужно_спец_действие(g: dict) -> str | None:
    """Если цель завязана на особое действие (вход, оплата, купон, избранное,
    скачивание, пагинация) - вернуть человеческое описание, иначе None."""
    text = ((g.get('название', '') + ' ' + g.get('условие', '') + ' '
             + ' '.join(g.get('идентификаторы') or [])).lower())
    for keys, label in _СПЕЦ_ДЕЙСТВИЯ:
        if any(k in text for k in keys):
            return label
    return None


# Цели-клики по соцсетям/мессенджерам/телефону/почте. Такие ссылки Метрика часто
# считает сама («внешняя ссылка»/«мессенджер»/«клик по телефону») - reachGoal в
# коде может не быть, но САМ КЛИК мы выполняем (GENERIC_CLICK_SELECTORS на каждой
# странице), значит действие посетителя совершено и Метрика цель зачтёт.
_СОЦ_КОНТАКТ_КЛЮЧИ = (
    'vk', 'вконтакте', 'ok_podval', 'одноклассник', 'dzen', 'дзен', 'rutube',
    'рутуб', '_tg', 'tg_', 'clicktg', 'телеграм', 'telegram', 'whatsapp', 'wapp',
    'ватсап', 'max_podval', 'clickmax', 'yandexorg', 'яндекс-организ',
    'соц.сет', 'соцсет',
)


def _клик_соцсети(g: dict) -> bool:
    text = (' '.join(g.get('идентификаторы') or []).lower() + ' '
            + (g.get('название', '') or '').lower())
    return any(k in text for k in _СОЦ_КОНТАКТ_КЛЮЧИ)


def _вход_в_аккаунт(g: dict) -> bool:
    """Цель «вход/авторизация в личном кабинете». На сайтах входа как такового
    нет (или он не найден) - помечаем «Не найдено», а не «нужно спец-действие»."""
    ids = ' '.join(g.get('идентификаторы') or []).lower()
    name = (g.get('название', '') or '').lower()
    return ('login' in ids or 'avtoriz' in ids
            or 'вход в аккаунт' in name or 'авторизов' in name)


def _лид_цель(g: dict) -> bool:
    """Составная «лид»-цель («Основная цель на лиды», «Весь сайт», d_Goal/Lid_Goal):
    Метрика сама агрегирует её из под-целей - отдельного goal-сигнала нет, считается
    в кабинете. Помечаем как автоцель Метрики."""
    ids = ' '.join(g.get('идентификаторы') or []).lower()
    name = (g.get('название', '') or '').lower()
    return ('d_goal' in ids or 'lid_goal' in ids or 'lidgoal' in ids
            or 'основная цель на лиды' in name or 'весь сайт' in name
            or 'лиды не основные' in name)


# Цвета статусов и порядок вывода - на уровне модуля (используются в
# классификации и в рисовании листов).
_GREEN, _RED, _GREY, _BLUE = '1E8E3E', 'C62828', '757575', '1565C0'
_ФОН = {_GREEN: 'E6F4EA', _RED: 'FCE8E6', _BLUE: 'E8F0FE', _GREY: 'F1F3F4'}
_ПОРЯДОК = {'Сработала': 0, 'Сработала (формы)': 1, 'Сработала (другой ID)': 1,
            'Сработала (в коде)': 1, 'Действие выполнено': 2,
            'Сработает': 3, 'Нет в коде сайта': 4, 'НЕ сработала': 5, 'Проблема': 6,
            'Нужно спец-действие': 7, 'Не найдено на сайте': 8,
            'Не найдена на сайте': 8, 'Не проверено': 9,
            'Нет автопроверки': 10, 'Проверяется формами': 11,
            'Форма не в этом прогоне': 11, 'Сайт недоступен': 9,
            'Автоцель (Метрика сама)': 12, 'Только в Метрике': 12, 'Составная': 13,
            'Вручную': 14}


def _классифицировать(pid: str, каталог: dict, прогон: dict) -> dict:
    """Сводит каталог целей с результатами прогона: список строк со статусом/
    цветом/пояснением + счётчики категорий. Без рисования (его делают отдельно)."""
    import re as _re2
    fired = прогон['fired']
    страницы = прогон['страницы']
    привязки = {i.lower() for i in прогон.get('привязки', set())}
    код = прогон.get('код', '')       # весь код страниц + JS (нижний регистр)
    формные = _формные_цели(pid)
    форм_статусы = _результаты_форм(pid)
    # Прогонялись ли формы в этом запуске (есть результат «Проверки форм»). Если да,
    # но цель формы всё равно не поймалась - это не «запустите формы», а «форма
    # отправлена, но цель при ней не зафиксировалась» (проблема настройки цели/формы).
    _формы_прогнали = bool(форм_статусы) or bool(_формные_url(pid))
    url_map = _url_цели_проверка(каталог, страницы, _формные_url(pid))
    _план = ACTIONS.get(pid) or _план_страна(pid, каталог.get('домен', ''))
    ожидаемые = {i.lower() for i in _план.get('ожидаемые', [])}
    # нет_на_сайте: список ИЛИ словарь {id: пояснение}. Приводим к словарю.
    _ннс_raw = _план.get('нет_на_сайте', [])
    if isinstance(_ннс_raw, dict):
        нет_на_сайте = {k.lower(): v for k, v in _ннс_raw.items()}
    else:
        нет_на_сайте = {i.lower(): '' for i in _ннс_raw}
    # смежные: {ожидаемый_id: (фактический_id, пояснение)} - на кнопке/форме
    # реально срабатывает фактический, а в Метрике заведён ожидаемый.
    смежные = {k.lower(): v for k, v in (_план.get('смежные') or {}).items()}
    GREEN, RED, GREY, BLUE = _GREEN, _RED, _GREY, _BLUE
    _код_кэш: dict[str, bool] = {}

    def _id_в_коде(gid: str) -> bool:
        """Цель упоминается в коде сайта: есть в reachGoal-списке ИЛИ встречается
        как строковый литерал ('id' / \"id\") где-либо в HTML/JS проверенных
        страниц. Литерал в кавычках защищает от ложных совпадений (tel в hotel)."""
        gid = (gid or '').lower()
        if not gid:
            return False
        if gid in привязки or any(gid in b for b in привязки):
            return True
        if gid in _код_кэш:
            return _код_кэш[gid]
        found = bool(_re2.search(r'["\']' + _re2.escape(gid) + r'["\']', код)) if код else False
        _код_кэш[gid] = found
        return found

    def _привязана(g) -> str:
        ids = g.get('идентификаторы') or []
        return 'есть' if any(_id_в_коде(i) for i in ids) else 'не найдена'

    def _форма_поймала(g) -> str | None:
        """Идентификатор цели, которую ПОЙМАЛА «Проверка форм» (лист «Цели» её
        отчёта). Формы отправляются по-настоящему, поэтому цели на onsubmit там
        реально фиксируются - даже если статический скан кода их не нашёл. Это
        снимает ложное «нет в коде сайта» с целей отправки форм."""
        for gid in (g.get('идентификаторы') or []):
            st = (форм_статусы.get(gid) or '').lower()
            if st.startswith('сработал') or 'зафиксир' in st:
                return gid
        return None

    # По ИДЕНТИФИКАТОРУ (надёжно: это цель ОТПРАВКИ формы).
    _ФОРМ_ID = ('form', 'forma', 'formu', 'otpravlen', 'zayavka', 'subscribe',
                'podpisk', 'findtome', 'raschet_stoimosti')
    # По НАЗВАНИЮ - только явные фразы отправки (НЕ «клик ... из формы ...»:
    # это клик по кнопке внутри формы, а не отправка - иначе ловили бы raschetaddtocart).
    _ФОРМ_ИМЯ = ('отправка форм', 'отправить форм', 'отправленная форма',
                 'отправленной форм', 'заполнение форм', 'форма подписки',
                 'подписк на рассыл')

    def _похоже_на_форму(g) -> bool:
        """Цель ОТПРАВКИ формы (fires on onsubmit): её reachGoal почти никогда не
        виден статически в коде, поэтому «нет в коде» - ложь. Определяем по
        идентификатору (надёжно) или явной фразе отправки в названии."""
        ids = ' '.join(g.get('идентификаторы') or []).lower()
        name = (g.get('название') or '').lower()
        return (any(w in ids for w in _ФОРМ_ID)
                or any(w in name for w in _ФОРМ_ИМЯ))

    def _смежная(g):
        """Если цель заведена в Метрике под одним id, а на кнопке/форме реально
        срабатывает другой (смежный), и этот другой ПОЙМАН в прогоне - вернуть
        (фактический_id, пояснение). Иначе None."""
        for gid in (g.get('идентификаторы') or []):
            пара = смежные.get(gid.lower())
            if not пара:
                continue
            факт, поясн = пара
            _поймана = (факт in fired
                        or (форм_статусы.get(факт, '').lower().startswith('сработал'))
                        or 'зафиксир' in форм_статусы.get(факт, '').lower())
            if _поймана:
                return (факт, поясн)
        return None

    def _авто_действие_сделано(условие: str) -> bool:
        """Автоцель Метрики: выполнил ли автотест соответствующее действие.
        Клики по телефону/почте/соцсетям/мессенджерам мы делаем; формы/файлы/
        поиск - нет (формы закрывает «Проверка форм»)."""
        c = (условие or '').lower()
        делаем = ('телефон', 'номер', 'email', 'почт', 'соц', 'мессенджер',
                  'whatsapp', 'telegram', 'вконтакте')
        не_делаем = ('форм', 'файл', 'скачив', 'поиск', 'чат', 'контактные данные',
                     'оформлени')
        if any(k in c for k in не_делаем):
            return False
        return any(k in c for k in делаем)

    счёт = {'ok': 0, 'ok_forms': 0, 'bad': 0, 'no_code': 0, 'forms': 0,
            'special': 0, 'manual': 0, 'info': 0}
    # Прозрачен ли код сайта для статического анализа. Если reachGoal почти не
    # нашли (сайт грузит цели через GTM/минифицированный бандл - как ИМП), то
    # вывод «нет в коде» НЕЛЬЗЯ делать - это была бы наша слепота, а не баг сайта.
    _код_надёжен = len(привязки) >= 3
    _строки: list[dict] = []
    for g in каталог.get('цели', []):
        t = g['тип']
        способ = статус = детали = ''
        цвет = GREY
        if t == 'js':
            hit = _совпало(g, fired)
            форма_id = _форма_поймала(g)
            в_формах = (any(gid in формные for gid in (g.get('идентификаторы') or []))
                        or _похоже_на_форму(g))
            _особое = _нужно_спец_действие(g)
            if _лид_цель(g):
                # Составная лид-цель: Метрика агрегирует сама на сервере - в отчёте
                # всегда «автоцель», а не зелёное срабатывание под-цели.
                способ, статус, цвет = 'автоцель Метрики', 'Автоцель (Метрика сама)', BLUE
                детали = ('составная «лид»-цель (Основная цель на лиды/Весь сайт): '
                          'Метрика собирает её из под-целей сама на сервере - '
                          'отдельного goal-сигнала в трафике нет, смотрите в кабинете')
                счёт['info'] += 1
            elif hit:
                способ, статус, цвет = 'клики автотеста', 'Сработала', GREEN
                детали = f'зафиксирован идентификатор «{hit}»'
                счёт['ok'] += 1
            elif _смежная(g):
                # На кнопке/форме реально срабатывает СМЕЖНЫЙ идентификатор (он
                # пойман), а в Метрике цель заведена под другим. По факту действие
                # выполнено и цель достигается - зелёный + понятное пояснение.
                _факт, _поясн = _смежная(g)
                способ, статус, цвет = 'смежная цель', 'Сработала (другой ID)', GREEN
                _этот = (g.get('идентификаторы') or ['?'])[0]
                детали = (f'по факту срабатывает «{_факт}» ({_поясн}), а в Метрике эта '
                          f'цель заведена под идентификатором «{_этот}». Действие '
                          'выполнено и засчитывается - но сверьте идентификатор в '
                          'кабинете (возможно, старый/переименованный)')
                счёт['ok'] += 1
            elif форма_id:
                # Цель реально сработала при ОТПРАВКЕ формы («Проверка форм») -
                # даже если статический скан reachGoal не нашёл. Это НЕ «нет в коде».
                способ, статус, цвет = 'через формы', 'Сработала (формы)', GREEN
                детали = ('зафиксирована при отправке формы на странице «Проверка '
                          f'форм» (идентификатор «{форма_id}»)')
                счёт['ok_forms'] += 1
            elif any(gid.lower() in нет_на_сайте
                     for gid in (g.get('идентификаторы') or [])):
                # Подтверждено вручную: такой кнопки/формы на сайте нет.
                _пояс = next((нет_на_сайте[gid.lower()]
                              for gid in (g.get('идентификаторы') or [])
                              if нет_на_сайте.get(gid.lower())), '')
                способ, статус, цвет = 'вручную', 'Не найдена на сайте', GREY
                детали = _пояс or ('кнопки/формы под эту цель на сайте нет '
                                   '(проверено вручную)')
                счёт['manual'] += 1
            elif _код_надёжен and _привязана(g) == 'есть':
                # reachGoal этой цели ЕСТЬ в коде сайта (в onclick или JS-бандле) -
                # значит цель ПРИВЯЗАНА и сработает при действии посетителя. Проверяем
                # ПЕРЕД входом и формами: если reachGoal есть в коде, цель рабочая,
                # даже если goal-хит в прогоне не пойман (findtome, вход в аккаунт,
                # managerform на СНГ). Сканируем весь код страниц разом - надёжно.
                способ, статус, цвет = 'reachGoal в коде', 'Сработала (в коде)', GREEN
                _прив = ', '.join(i for i in (g.get('идентификаторы') or []) if _id_в_коде(i))
                детали = (f'reachGoal цели ({_прив}) найден в коде сайта - цель '
                          'привязана и срабатывает при действии посетителя '
                          '(клик по кнопке/ссылке/форме)')
                счёт['ok'] += 1
            elif _вход_в_аккаунт(g) and not any(
                    gid.lower() in ожидаемые for gid in (g.get('идентификаторы') or [])):
                # Вход в личный кабинет, а reachGoal в коде не нашли: кнопку/страницу
                # входа найти не удалось - проверяется вручную.
                способ, статус, цвет = 'вручную', 'Не найдено на сайте', GREY
                детали = ('кнопку/страницу входа в личный кабинет на сайте найти '
                          'не удалось - проверьте вручную')
                счёт['manual'] += 1
            elif в_формах and not _формы_прогнали:
                # Цель привязана к отправке формы, а результата форм ещё нет:
                # подскажем запустить «Проверку форм» (результат подтянется сам).
                способ, статус, цвет = 'через формы', 'Проверяется формами', BLUE
                детали = ('цель срабатывает при отправке формы - запустите «Проверку '
                          'форм» (её результат автоматически подтянется в этот отчёт)')
                счёт['forms'] += 1
            elif в_формах:
                # reachGoal формы в коде НЕ найден (форма на этом сайте отсутствует
                # или её нет в Москва-прогоне - напр. «Связаться с менеджером» только
                # для СНГ, а формы гоняются по Москве). Не ошибка сайта.
                способ, статус, цвет = 'через формы', 'Форма не в этом прогоне', GREY
                детали = ('цель срабатывает при отправке формы, но эта форма не '
                          'проверена в текущем прогоне (формы гоняются по Москве; '
                          'часть форм есть только на отдельных городах/странах СНГ '
                          'или подгружается динамически) - проверьте вручную/в кабинете')
                счёт['manual'] += 1
            elif _клик_соцсети(g):
                # Клик по соцсети/мессенджеру мы ВЫПОЛНЯЕМ на каждой странице
                # (GENERIC_CLICK_SELECTORS). Такие цели Метрика чаще считает сама
                # как «внешнюю ссылку»/«мессенджер» (reachGoal в коде может не быть,
                # goal-хита в трафике нет), но действие посетителя совершено - в
                # Метрике цель зачтётся.
                способ, статус, цвет = 'клик по соцсети', 'Действие выполнено', GREEN
                детали = ('клик по ссылке соцсети/мессенджера выполнен. Метрика '
                          'считает такие цели сама («внешняя ссылка»/«мессенджер») - '
                          'отдельного goal-сигнала в трафике нет, но действие '
                          'совершено, и в Метрике цель зачтётся')
                счёт['ok'] += 1
            elif _особое:
                # Цель срабатывает только на особое действие (купон, оплата, избранное,
                # скачивание и т.п.), которого автотест не делает. Ставим ПЕРЕД «нет в
                # коде»: reachGoal купона грузится в ленивом чанке корзины, которого мы
                # не касаемся - это не «нет в коде», а «нужно спец-действие».
                способ, статус, цвет = 'спец-действие', 'Нужно спец-действие', GREY
                _прив = f"; reachGoal в коде: {_привязана(g)}" if _код_надёжен else ''
                детали = (f'нужно {_особое} - автотест этот шаг не выполняет '
                          f'(можно добавить отдельным сценарием){_прив}')
                счёт['special'] += 1
            elif _код_надёжен and _привязана(g) == 'не найдена':
                # Код сайта прозрачен, а reachGoal этой цели в нём не встречается -
                # это не наша слепота, а отсутствие отправки на стороне сайта.
                способ, статус, цвет = 'проверка кода', 'Нет в коде сайта', RED
                детали = ('reachGoal этой цели НЕ найден в коде сайта - цель создана '
                          'в Метрике, но сайт её не отправляет')
                счёт['no_code'] += 1
            else:
                # Код непрозрачен (GTM/бандл) или действие не входило в прогон.
                способ, статус, цвет = 'вручную', 'Не проверено', GREY
                _подск = f"; привязка reachGoal в коде: {_привязана(g)}" if _код_надёжен else ''
                детали = ('цель грузится через GTM/бандл или требует действия, '
                          'которого не было в прогоне - проверяется вручную/в '
                          'Метрике' + _подск)
                счёт['manual'] += 1
        elif t in ('url', 'url_re'):
            способ = 'визит страницы'
            _um = url_map.get(g['номер']) or {}
            s = _um.get('страница')
            if s and _um.get('через_формы'):
                # url-цель подтверждена прогоном форм (оформленный заказ / «спасибо»).
                способ = 'через формы'
                статус, цвет = 'Сработала (формы)', GREEN
                детали = ('цель = визит на страницу; «Проверка форм» реально дошла до '
                          f'{s["url"]} (напр. оформила заказ) - значит визит на эту '
                          'страницу засчитан и цель фиксируется')
                счёт['ok_forms'] += 1
            elif s and s['код'] == 200 and s['счётчик']:
                статус, цвет = 'Сработала', GREEN
                детали = (f"цель = визит на страницу; мы открыли {s['url']} "
                          "(ответ 200, счётчик Метрики на странице стоит) - "
                          "значит визит засчитан и цель фиксируется"
                          + (', визит подтверждён в трафике' if s['визит'] else ''))
                счёт['ok'] += 1
            elif s and (s['код'] or 0) >= 500:
                # Сервер вернул 5xx (транзиентная ошибка сайта, не наша и не цели) -
                # не красный «баг», а серое «сайт временно недоступен».
                способ, статус, цвет = 'визит страницы', 'Сайт недоступен', GREY
                детали = (f"сайт вернул код {s['код']} на {s['url']} - похоже, "
                          "временная ошибка сервера (при вас сайт работает) - "
                          "перепроверьте прогоном позже")
                счёт['manual'] += 1
            elif s:
                статус, цвет = 'Проблема', RED
                детали = f"страница {s['url']}: код {s['код']}, счётчик {'есть' if s['счётчик'] else 'НЕ найден'}"
                счёт['bad'] += 1
            elif ('favorites' in (g.get('url_часть', '') or '').lower()
                  or 'избранн' in (g.get('название', '') or '').lower()) \
                    and pid != _базовый(pid):
                # «Посещение страницы Избранного» - Избранное есть только на
                # хабаровском поддомене (РФ). На странах его нет - не ошибка.
                способ, статус, цвет = 'вручную', 'Не найдена на сайте', GREY
                детали = ('страница «Избранное» есть только на хабаровском поддомене '
                          '(РФ) - на этом сайте её нет')
                счёт['manual'] += 1
            else:
                статус, цвет = 'Нет автопроверки', GREY
                детали = (f"цель = визит на страницу «{g.get('url_часть','')}». Обычный "
                          "прогон целей на неё не попадает (заказ/оплата/«спасибо»). "
                          "Включите галочку «Сначала прогнать формы» - форма оформляет "
                          "заказ по-настоящему, и если её финальный адрес совпадёт с "
                          "этой целью, статус станет «Сработала (формы)»")
                счёт['manual'] += 1
        elif t == 'auto':
            сделано = _авто_действие_сделано(g.get('условие', ''))
            способ = 'автоцель Метрики'
            if сделано:
                статус, цвет = 'Действие выполнено', GREEN
                детали = ('автоцель Метрики (клик по телефону/почте/соцсети). Такие '
                          'цели Метрика считает САМА на своём сервере - отдельного '
                          'goal-сигнала в трафике нет, увидеть факт срабатывания извне '
                          'нельзя. Но нужное действие автотест выполнил - значит в '
                          'Метрике цель зачтётся')
                счёт['ok'] += 1
            else:
                статус, цвет = 'Автоцель (Метрика сама)', BLUE
                детали = ('это автоцель Метрики: Яндекс считает её сам на сервере '
                          '(отправка формы/файл/поиск/контакты) - отдельного '
                          'goal-сигнала в трафике нет, извне факт срабатывания не '
                          'виден. Действие закрывает «Проверка форм» (для форм) либо '
                          'оно совершается реальным посетителем - смотрите цифры в '
                          'самой Метрике')
                счёт['info'] += 1
        elif t == 'jivo':
            способ, статус, цвет = 'вручную', 'Вручную', GREY
            детали = 'события чата Jivo зависят от посетителя/оператора'
            счёт['manual'] += 1
        else:  # composite
            способ, статус, цвет = 'по шагам', 'Составная', BLUE
            детали = 'Метрика вычисляет из шагов - смотри цели-шаги выше/ниже'
            счёт['info'] += 1

        _строки.append({'номер': g['номер'], 'название': g['название'],
                        'условие': g['условие'], 'статус': статус,
                        'детали': детали, 'цвет': цвет})

    # Сортировка по статусу для читаемости листа.
    _строки.sort(key=lambda x: (_ПОРЯДОК.get(x['статус'].split(':')[0].strip(), 20),
                                x['название']))
    return {'строки': _строки, 'счёт': счёт, 'страницы': страницы,
            'код_надёжен': _код_надёжен, 'проект': каталог.get('проект', ''),
            'счётчик': каталог.get('счётчик', ''),
            'всего': len(каталог.get('цели', []))}


def _рисовать_цели(ws, строки):
    """Рисует лист целей: шапка + цветные плашки статусов + сетка + автофильтр."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    ws.sheet_view.showGridLines = False
    headers = ['№ цели', 'Название', 'Статус', 'Что это значит', 'Условие (из Метрики)']
    for c, (h, w) in enumerate(zip(headers, (12, 42, 20, 62, 44)), 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='EEF3FB')
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'
    _tside = Side(style='thin', color='D9DCE1')
    _tbord = Border(left=_tside, right=_tside, top=_tside, bottom=_tside)
    r = 2
    for s in строки:
        vals = [s['номер'], s['название'], s['статус'], s['детали'], s['условие']]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = Alignment(wrap_text=(c in (2, 3, 4)), vertical='top')
            cell.border = _tbord
        st = ws.cell(r, 3)
        st.font = Font(color=s['цвет'], bold=True)
        st.fill = PatternFill('solid', fgColor=_ФОН.get(s['цвет'], 'FFFFFF'))
        r += 1
    for c in range(1, 6):
        ws.cell(1, c).border = _tbord
    ws.auto_filter.ref = f"A1:E{max(2, r - 1)}"


def _рисовать_сводку(sm, данные):
    """Лист «Сводка» одной страны: итоги по категориям + страницы прогона."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    GREEN, RED, GREY, BLUE = _GREEN, _RED, _GREY, _BLUE
    счёт = данные['счёт']
    страницы = данные['страницы']
    _код_надёжен = данные['код_надёжен']
    каталог = {'проект': данные['проект'], 'счётчик': данные['счётчик'],
               'цели': [None] * данные['всего']}
    sm.sheet_view.showGridLines = False
    sm.column_dimensions['A'].width = 4
    sm.column_dimensions['B'].width = 32
    sm.column_dimensions['C'].width = 9
    sm.column_dimensions['D'].width = 74
    _thin = Side(style='thin', color='D9DCE1')
    _bord = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    sm.merge_cells('A1:D1')
    sm['A1'] = f"Проверка целей Метрики - {каталог.get('проект','')}"
    sm['A1'].font = Font(bold=True, size=15)
    sm.merge_cells('A2:D2')
    sm['A2'] = (f"Счётчик {каталог.get('счётчик','')} · целей в каталоге: "
                f"{len(каталог.get('цели', []))} · прогон "
                f"{datetime.now().strftime('%d.%m.%Y %H:%M')}")
    sm['A2'].font = Font(italic=True, color='5F6368')

    _подтв = счёт['ok'] + счёт['ok_forms']
    _пробл = счёт['no_code'] + счёт['bad']
    sm.merge_cells('A3:D3')
    sm['A3'] = (f"Подтверждено: {_подтв}   ·   Проблемы: {_пробл}   ·   "
                f"Требует действия/вручную: {счёт['special'] + счёт['manual']}   ·   "
                f"Формы и авто: {счёт['forms'] + счёт['info']}")
    sm['A3'].font = Font(bold=True, color='3C4043')

    # Категории сгруппированы по смыслу; между группами - пустая строка-разделитель.
    _группы = [
        ('ПОДТВЕРЖДЕНО', [
            ('✅ Сработали при кликах', счёт['ok'], GREEN,
             'цель реально зафиксирована во время кликов автотеста'),
            ('✅ Сработали через формы', счёт['ok_forms'], GREEN,
             'цель поймана при отправке формы («Проверка форм») - reachGoal рабочий'),
        ]),
        ('ПРОБЛЕМЫ (к разработчикам)', [
            ('❌ Нет в коде сайта', счёт['no_code'], RED,
             'reachGoal этой цели в коде сайта не найден - цель в Метрике есть, но '
             'сайт её не отправляет'),
            ('❌ НЕ сработала (reachGoal есть)', счёт['bad'], RED,
             'reachGoal в коде есть, действие выполняли, но цель не поймалась - '
             'проверить её настройку/кнопку'),
        ]),
        ('ТРЕБУЕТ ДЕЙСТВИЯ / ВРУЧНУЮ', [
            ('🟡 Нужно спец-действие', счёт['special'], GREY,
             'цель на вход/оплату/купон/избранное/скачивание - автотест этот шаг '
             'пока не делает (можно добавить сценарием)'),
            ('🖐 Не проверено', счёт['manual'], GREY,
             'цель грузится через GTM/бандл или её страницы не было в прогоне - '
             'смотрится вручную/в Метрике'),
        ]),
        ('ФОРМЫ И АВТО-ЦЕЛИ', [
            ('📝 Проверяется формами', счёт['forms'], BLUE,
             'цель отправки формы - запустите «Проверку форм», результат подтянется сюда'),
            ('ℹ️ Авто / составные', счёт['info'], BLUE,
             'Метрика считает сама на сервере (goal-сигнала в трафике нет - увидеть '
             'извне нельзя)'),
        ]),
    ]

    hr = 5
    for c, h in enumerate(['', 'Категория', 'Кол-во', 'Что это значит'], 1):
        cell = sm.cell(hr, c, h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='5B6470')
        cell.border = _bord
        cell.alignment = Alignment(horizontal='center' if c == 3 else 'left')
    rr = hr + 1
    for заг, строки in _группы:
        sm.cell(rr, 2, заг).font = Font(bold=True, size=9, color='80868B')
        rr += 1
        for назв, кол, цв, пояс in строки:
            sm.cell(rr, 2, назв).font = Font(bold=True, color=цв)
            cc = sm.cell(rr, 3, кол)
            cc.font = Font(bold=True, color=цв)
            cc.alignment = Alignment(horizontal='center')
            cc.fill = PatternFill('solid', fgColor=_ФОН.get(цв, 'FFFFFF'))
            пc = sm.cell(rr, 4, пояс)
            пc.alignment = Alignment(wrap_text=True, vertical='top')
            for c in (2, 3, 4):
                sm.cell(rr, c).border = _bord
            sm.row_dimensions[rr].height = 30
            rr += 1
        rr += 1   # пустая строка-разделитель между группами

    if not _код_надёжен:
        sm.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=4)
        sm.cell(rr, 2, '⚠️ Код сайта грузится через GTM/бандл - reachGoal статически '
                       'не виден, поэтому «нет в коде» здесь не выносим (чтобы не '
                       'обвинять сайт зря).').font = Font(italic=True, color='B06000')
        rr += 2

    sm.cell(rr, 2, 'Страницы прогона').font = Font(bold=True)
    rr += 1
    for s in страницы:
        sm.cell(rr, 2, s['название']).font = Font(bold=True)
        sm.cell(rr, 4, f"{s['url']} - код {s['код']}, "
                       f"счётчик {'✓' if s['счётчик'] else '✗ НЕ найден'}")
        rr += 1


def построить_отчёт(pid: str, каталог: dict, прогон: dict,
                    out_path: str | Path) -> Path:
    """Отчёт по одному сайту: лист «Сводка» + лист «Цели Метрики»."""
    from openpyxl import Workbook
    данные = _классифицировать(pid, каталог, прогон)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Цели Метрики'
    _рисовать_цели(ws, данные['строки'])
    _рисовать_сводку(wb.create_sheet('Сводка', 0), данные)
    wb.active = 0
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def построить_сводный_отчёт(результаты: list, out_path: str | Path) -> Path:
    """Один файл на несколько сайтов: лист «Сводка» (строка на сайт) + по листу
    целей на каждый сайт. результаты = [(pid, каталог, прогон, метка), ...]."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    сайты = []
    for pid, каталог, прогон, метка in результаты:
        d = _классифицировать(pid, каталог, прогон)
        d['метка'] = метка or (каталог.get('проект') or pid)
        сайты.append(d)

    wb = Workbook()
    sm = wb.active
    sm.title = 'Сводка'
    sm.sheet_view.showGridLines = False
    _thin = Side(style='thin', color='D9DCE1')
    _bord = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    for c, w in enumerate([26, 8, 12, 12, 20, 14], 1):
        sm.column_dimensions[get_column_letter(c)].width = w
    sm.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    sm.cell(1, 1, 'Проверка целей Метрики - сводка по сайтам · '
                  f"{datetime.now().strftime('%d.%m.%Y %H:%M')}").font = Font(bold=True, size=14)
    hdr = ['Сайт', 'Целей', '✅ Подтв.', '❌ Проблемы', '🟡 Действие/вручную', '📝 Формы/авто']
    for c, h in enumerate(hdr, 1):
        cell = sm.cell(3, c, h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='5B6470')
        cell.border = _bord
        cell.alignment = Alignment(horizontal='left' if c == 1 else 'center',
                                   wrap_text=True, vertical='center')
    r = 4
    for d in сайты:
        s = d['счёт']
        подтв = s['ok'] + s['ok_forms']
        пробл = s['no_code'] + s['bad']
        vals = [d['метка'], d['всего'], подтв, пробл,
                s['special'] + s['manual'], s['forms'] + s['info']]
        for c, v in enumerate(vals, 1):
            cell = sm.cell(r, c, v)
            cell.border = _bord
            cell.alignment = Alignment(horizontal='left' if c == 1 else 'center')
            if c == 1:
                cell.font = Font(bold=True)
            elif c == 3 and подтв:
                cell.font = Font(bold=True, color=_GREEN)
                cell.fill = PatternFill('solid', fgColor=_ФОН[_GREEN])
            elif c == 4 and пробл:
                cell.font = Font(bold=True, color=_RED)
                cell.fill = PatternFill('solid', fgColor=_ФОН[_RED])
        r += 1

    # По листу целей на каждый сайт (имя листа = метка, уникальное, <=31 символ).
    занятые = set()
    for d in сайты:
        имя = (d['метка'] or 'Сайт').replace('/', '-').replace('\\', '-')[:28] or 'Сайт'
        база, i = имя, 2
        while имя in занятые:
            имя = f"{база[:26]} {i}"
            i += 1
        занятые.add(имя)
        _рисовать_цели(wb.create_sheet(имя), d['строки'])

    wb.active = 0
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
