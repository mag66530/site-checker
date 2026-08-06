"""
reporter.py - формирование xlsx-отчёта.

Структура:
  • Лист «Обзор» - метрики, сводка, параметры прогона
  • Лист «План работ» / «Проблемы» - приоритезированные задачи и находки
  • Лист «Страницы» - каждая проверенная страница отдельной строкой (код,
    статус, скорость, отдел, битые переменные, откуда перешли, найдено
    проблем) - без деталей находок, те на «Проблемы»
"""
import re
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, range_boundaries

from report_priorities import (
    PRIORITY_LABEL, collect_findings, group_into_tasks, extra_site_tasks,
    classify, indexing_site_findings, metadata_site_findings,
    home_dupes_findings, arsenkin_findings, page404_findings,
    stress_check_findings, ps_filters_findings,
    service_issues_findings, w3c_findings,
    url_format_findings, robots_hygiene_findings, content_sections_findings,
    static_delivery_findings, ux_interactive_findings, interlinking_note,
    yabusiness_findings, review_priority_findings,
)


# ── Стили (цвета как в Node.js версии) ──────────────────────────────


class C:
    text = '09090B'
    text_soft = '3F3F46'
    text_muted = '71717A'
    # Раньше border_light был 'E4E4E7' - настолько светлый, что в Excel
    # границы данных не было видно («почему в отчёте нет границ»).
    # Делаем оба варианта заметнее.
    border = 'A8B2BD'
    border_light = 'C7D0DA'
    surface = 'FAFAFA'
    bg_elev = 'FFFFFF'
    accent = '0052CC'
    accent_soft = 'EEF3FB'
    header_navy = '1F3864'
    ok = '15803D'
    ok_soft = 'F0FDF4'
    warn = 'B45309'
    warn_soft = 'FFFBEB'
    err = 'B91C1C'
    err_soft = 'FEF2F2'


# Метки статусов на русском
STATUS_LABEL = {
    'ok': 'Работает',
    'redirect': 'Перенаправление',
    'redirect_loop': 'Циклический редирект',
    'not_found': 'Страница не найдена',
    'client_error': 'Ошибка на сайте',
    'server_error': 'Сервер не отвечает',
    'timeout': 'Нет ответа',
    'network_error': 'Нет соединения',
}

SPEED_LABEL = {
    'fast': 'ОК',
    'normal': 'ОК',
    'slow': 'Медленно',
    'very_slow': 'Долгий ответ сервера',
}

SPEED_COLOR = {
    'fast': C.ok,
    'normal': C.ok,
    'slow': C.warn,
    'very_slow': C.err,
}

SPEED_FILL = {
    'fast': C.ok_soft,
    'normal': C.ok_soft,
    'slow': C.warn_soft,
    'very_slow': C.err_soft,
}

_NOTIF_CAT_DEPT = {
    'server':    ['разработка'],
    'speed':     ['разработка'],
    'security':  ['разработка'],
    'indexing':  ['SEO'],
    'coverage':  ['SEO'],
    'structure': ['SEO'],
    'other':     ['SEO'],
}


def _dept_result(r) -> str:
    """Отдел для колонки «Кому чинить» листа «Страницы».

    Тег ставим ТОЛЬКО при проблеме со статусом или скоростью.
    Если статус «Работает» и скорость «ОК» - поле пустое, всё в порядке.
    (Битые переменные и контент-баги показаны в своих колонках/листе,
    здесь их не дублируем - иначе тег появлялся бы у рабочих страниц.)

    Карта:
      • сервер не отвечает / таймаут / нет соединения (5xx) → разработка
      • прочие ошибки на сайте (4xx, кроме 404)              → разработка
      • долгий ответ сервера (медленно)                      → разработка
      • 404 / страница не найдена                            → SEO
      • редиректы (предупреждение)                           → SEO
    """
    tags: list[str] = []
    if r.is_error:
        if r.status == 'not_found':
            tags.append('SEO')
        else:  # server_error, timeout, network_error, client_error
            tags.append('разработка')
    elif r.is_warning:
        tags.append('SEO')
    if r.speed_rating in ('slow', 'very_slow') and 'разработка' not in tags:
        tags.append('разработка')
    return ', '.join(dict.fromkeys(tags))


def _dept_notif(n) -> str:
    return ', '.join(_NOTIF_CAT_DEPT.get(n.category, ['SEO']))


def _argb(color: str) -> str:
    """6-значный RGB ('71717A') -> 8-значный ARGB с непрозрачной альфой
    ('FF71717A'). Без этого openpyxl сам подставляет альфу '00' (прозрачно) -
    Font(color='71717A').color.rgb == '0071717A', а не 'FF71717A' - из-за
    этого текст/заливка/границы могут отрисовываться невидимыми в клиентах,
    которые честно учитывают альфа-канал (не все версии Excel его игнорируют)."""
    return color if len(color) == 8 else f'FF{color}'


def _font(size=10, bold=False, italic=False, underline=None, color=C.text, name='Arial'):
    return Font(
        name=name, size=size, bold=bold, italic=italic,
        underline=underline, color=_argb(color),
    )


def _border(color=C.border):
    side = Side(style='thin', color=_argb(color))
    return Border(top=side, left=side, bottom=side, right=side)


def _fill(color):
    argb = _argb(color)
    return PatternFill(start_color=argb, end_color=argb, fill_type='solid')


def _wrap_line_count(text: str, width_units: float) -> int:
    """Сколько визуальных строк займёт text при переносе на ширину
    width_units (сумма Excel-ширин колонок) - грубая, но честная оценка без
    доступа к реальному измерению текста в Excel."""
    chars_per_line = max(10, int(width_units) - 2)
    return sum(max(1, -(-len(line) // chars_per_line))
              for line in str(text or '').split('\n'))


def _row_height_for(text: str, width_units: float, *, line_px: int = 15,
                    min_px: int = 20) -> int:
    return max(min_px, _wrap_line_count(text, width_units) * line_px + 8)


def _align(horizontal='left', vertical='center', wrap=False, indent=1):
    return Alignment(
        horizontal=horizontal, vertical=vertical,
        wrap_text=wrap, indent=indent,
    )


# ── Описание пути для 404 ──────────────────────────────────────────


def _build_path_description(result) -> str:
    """Колонка «Откуда перешли»: пусто / прямая ссылка / цепочка редиректов."""
    chain = result.redirect_chain or []
    if not chain:
        if not result.is_ok:
            return 'Прямая ссылка из каталога (без переходов)'
        return ''

    # Цепочка редиректов: 301: from → to → to2
    steps = []
    for i, hop in enumerate(chain):
        if i == 0:
            steps.append(f"{hop['code']}: {hop['from']}")
        steps.append(f"→ {hop['to']}")
    return '  '.join(steps)


# ── Лист «Структура страниц» ───────────────────────────────────────

# Порядок и подписи групп страниц. Категории/теги делятся по факту наполнения:
# страница с товарами → «Листинг», страница-витрина/пустая → «Разделы каталога».
def _grp_listing(r):
    return (r.type_code in ('category', 'filter')
            and getattr(r.content, 'page_kind', '') == 'listing')


def _grp_section(r):
    return (r.type_code in ('category', 'filter')
            and getattr(r.content, 'page_kind', '') in ('section', 'empty'))


_STRUCT_GROUPS = [
    ('Главная',           lambda r: r.type_code == 'main'),
    ('Каталог',           lambda r: r.type_code == 'catalog'),
    ('Листинг',           _grp_listing),
    ('Разделы каталога',  _grp_section),
    ('Карточки товаров',  lambda r: r.type_code == 'product'),
    ('Прочие страницы',   lambda r: r.type_code == 'custom'),
]


# «Схлопнутые» столбцы грида: 3 столбца цены и 3 столбца кнопок сводим в один
# смысловой каждый - так таблица читается, а тип цены/кнопки виден в ячейке.

def _price_cell(bk):
    # Одна галочка: есть цена в любом виде (₽ ИЛИ «по запросу») → ✓; нет ни того
    # ни другого или скрыто стилями → БАГ. Без «₽ + запрос» - это лишний шум.
    price = bk.get('price')
    if price and price.required and not price.present:
        return ('БАГ', 'bug')
    if price and price.present:
        return ('✓', 'ok')
    return ('-', 'absent')


def _btn_cell(bk):
    order = bk.get('btn_order'); cart = bk.get('btn_cart'); one = bk.get('btn_oneclick')
    if order and order.required and not order.present:
        return ('БАГ', 'bug')
    has_cart = bool(cart and cart.present); has_one = bool(one and one.present)
    if has_cart and has_one:
        return ('в корзину + 1 клик', 'okinfo')
    if has_cart:
        return ('в корзину', 'okinfo')
    if has_one:
        return ('1 клик', 'okinfo')
    if order and order.present:
        return ('✓', 'ok')
    return ('-', 'absent')


_COLLAPSE = [
    {'trigger': 'price', 'label': 'Цена',
     'desc': 'Цена на карточках: «₽» - рублёвая, «по запросу» - цена по запросу. '
             '«БАГ» - цены нет вовсе.',
     'keys': {'price', 'price_real', 'price_request'}, 'fn': _price_cell},
    {'trigger': 'btn_order', 'label': 'Кнопка заказа',
     'desc': 'Кнопка заказа: «в корзину» (товар с ценой) или «1 клик» (по запросу). '
             '«БАГ» - нет ни одной.',
     'keys': {'btn_order', 'btn_cart', 'btn_oneclick'}, 'fn': _btn_cell},
]


def _grid_columns(blocks):
    """Столбцы грида: реальные блоки + схлопнутые «Цена»/«Кнопка заказа»."""
    by_trigger = {c['trigger']: c for c in _COLLAPSE}
    consumed = set().union(*(c['keys'] for c in _COLLAPSE))
    cols = []
    for b in blocks:
        if b.key in by_trigger:
            c = by_trigger[b.key]
            cols.append({'kind': 'virtual', 'label': c['label'],
                         'desc': c['desc'], 'fn': c['fn']})
        elif b.key in consumed:
            continue                       # под-блок схлопнут - пропускаем
        else:
            cols.append({'kind': 'block', 'key': b.key, 'label': b.label,
                         'desc': getattr(b, 'description', '')})
    return cols


def _cell_state(col, by_key):
    """(значение, состояние) для ячейки грида."""
    if col['kind'] == 'virtual':
        return col['fn'](by_key)
    b = by_key.get(col['key'])
    if b is None:
        return ('', 'absent')
    # Жёлтое предупреждение (не красный баг), напр. «Фото товаров»: стоит заглушка.
    if getattr(b, 'warn', False):
        return (f'Заглушка ({b.count})' if b.count else 'Заглушка', 'warn')
    if b.required and not b.present:
        if b.count:
            return (f'БАГ ({b.count})', 'bug')
        return ('БАГ', 'bug')
    if b.present:
        if b.count is not None:
            return (b.count, 'count')
        return ('✓', 'ok')
    return ('-', 'absent')


def _style_cell(cell, value, state):
    cell.value = value
    if state == 'bug':
        cell.font = _font(size=10, bold=True, color=C.err); cell.fill = _fill(C.err_soft)
    elif state == 'warn':          # жёлтое предупреждение (заглушка фото и т.п.)
        cell.font = _font(size=10, bold=True, color=C.warn); cell.fill = _fill(C.warn_soft)
    elif state == 'ok':
        cell.font = _font(size=10, bold=True, color=C.ok); cell.fill = _fill(C.ok_soft)
    elif state == 'okinfo':       # значение-текст (по запросу / в корзину…)
        cell.font = _font(size=9, color=C.ok)
    elif state == 'count':
        cell.font = _font(size=10, color=C.text_soft)
    else:                          # absent
        cell.value = '-'
        cell.font = _font(size=10, color=C.text_muted)


def _plural_pages(n):
    n10, n100 = n % 10, n % 100
    if n10 == 1 and n100 != 11:
        return 'страница'
    if 2 <= n10 <= 4 and not 12 <= n100 <= 14:
        return 'страницы'
    return 'страниц'


_KIND_LABEL = {'listing': 'Листинг', 'section': 'Раздел каталога',
               'empty': 'Пустой раздел'}


def _contacts_problem_text(r):
    """Текст расхождений контактов с КП (адреса всех городов / телефон страницы)."""
    parts = []
    ca = getattr(r, 'contacts_addr', None)
    if ca and ca.get('mismatched'):
        mm = ca['mismatched']
        ex = '; '.join(f'{m["city"]}: сайт «{m["site"]}» / КП «{m["kp"]}»' for m in mm[:5])
        parts.append('адреса не совпадают с КП - ' + ex
                     + (f' и ещё {len(mm) - 5}' if len(mm) > 5 else ''))
    pp = getattr(r, 'page_phone', None)
    if pp and pp.get('status') in ('bug', 'critical'):
        parts.append(f'телефон: {pp.get("comment", "не совпадает с КП")}')
    return '; '.join(parts)


def _broken_links_text(r):
    """Битые ссылки (404/410) в контенте страницы - краткий текст для отчёта."""
    bl = getattr(r, 'broken_links', None)
    if not bl or not bl.get('broken'):
        return ''
    items = bl['broken']
    ex = '; '.join(f'{b["code"]} {b["url"]}' for b in items[:3])
    more = f' и ещё {len(items) - 3}' if len(items) > 3 else ''
    return f'битые ссылки ({len(items)}): ' + ex + more


# Человеческие формулировки багов для «Что чинить» / «Что не так». Иначе из
# машинного названия столбца получалось коряво: «нет: Цена (есть)».
_BUG_PHRASES = {
    'price': 'нет цены',
    'price_real': 'нет цены суммой',
    'btn_order': 'нет кнопки заказа',
    'product_cards': 'нет карточек товаров',
    'photos': 'нет фото у части товаров',
    'h1': 'нет заголовка H1',
    'breadcrumbs': 'нет хлебных крошек',
    'img_alt': 'картинки без alt',
    'content_text': 'нет текста на странице',
    'rec_price': 'нет цен в нижних блоках',
    'form_nf': 'нет формы «Не нашли что искали»',
    'tech_map': 'нет карты',
    'tech_images': 'нет картинок',
    'tech_search': 'нет строки поиска',
    'hdr_phone': 'нет телефона в шапке',
    'hdr_callback': 'нет «Заказать звонок» в шапке',
    'hdr_request': 'нет «Оставить заявку» в шапке',
    'hdr_city': 'нет выбора города в шапке',
    'ftr_phone': 'нет телефона в подвале',
    'ftr_email': 'нет e-mail в подвале',
    'ftr_writeus': 'нет «Написать нам» в подвале',
    'ftr_address': 'нет адреса в подвале',
}


def _problem_text(r):
    """Понятная формулировка проблемы страницы для списка «Что чинить»."""
    parts = []
    _ct = _contacts_problem_text(r)
    if _ct:
        parts.append(_ct)
    content = getattr(r, 'content', None)
    if content is not None:
        if getattr(content, 'is_soft_404', False):
            parts.append('страница отдаёт 404 (не найдена) - проверить ссылку или убрать из каталога')
        elif getattr(content, 'page_kind', '') == 'empty':
            parts.append('раздел пуст - нет ни товаров, ни подразделов')
        else:
            # Человеческая фраза по каждому багу (+ число для фото, + пояснение,
            # напр. «в коде есть, но покупатель не видит»).
            bugs = []
            for b in content.bugs:
                phrase = _BUG_PHRASES.get(b.key, b.label)
                if b.key in ('photos', 'img_alt') and getattr(b, 'count', None):
                    phrase += f' ({b.count})'
                if getattr(b, 'note', ''):
                    # У картинок без alt пояснение - список адресов: через «:»
                    if b.key == 'img_alt':
                        phrase += f': {b.note}'
                    else:
                        phrase += f' ({b.note})'
                bugs.append(phrase)
            if bugs:
                parts.append(', '.join(bugs))
    _bl = _broken_links_text(r)
    if _bl:
        parts.append(_bl)
    return '; '.join(parts) if parts else 'проблема'


def _build_structure_sheet(wb, results):
    """Лист структурной проверки - дашборд, что чинить, сводка и детали."""
    # Тех. страницы выносим отдельной секцией (у них нет структуры - только
    # доступность), чтобы они не искажали статистику структурной проверки.
    pages = [r for r in results if getattr(r, 'content', None) is not None
             and getattr(r, 'type_code', '') != 'tech']
    if not pages:
        return

    ws = wb.create_sheet('Структура страниц')
    ws.sheet_view.showGridLines = False

    total_pages = len(pages)
    pages_with_bugs = sum(1 for r in pages if r.content_bugs > 0)
    ok_pages = total_pages - pages_with_bugs
    total_bugs = sum(r.content_bugs for r in pages)
    ws.sheet_properties.tabColor = C.err if total_bugs else C.ok

    # ── Ширины ──
    ws.column_dimensions['A'].width = 2.5
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 11
    max_block_cols = max((len(_grid_columns(r.content.blocks)) for r in pages), default=12)
    for col_idx in range(5, 5 + max(max_block_cols, 9) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13
    last_col = 13                       # карточки дашборда занимают B..M
    LASTL = get_column_letter(last_col)

    def fill_block(r1, c1, r2, c2, bg, bc=C.border_light):
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                cell = ws.cell(row=rr, column=cc)
                if bg:
                    cell.fill = _fill(bg)
                cell.border = _border(color=bc)

    # ── Заголовок ──
    ws.merge_cells(f'B2:{LASTL}2')
    c = ws['B2']
    c.value = 'Структура страниц'
    c.font = _font(size=20, bold=True, color=C.text)
    ws.row_dimensions[2].height = 30

    ws.merge_cells(f'B3:{LASTL}3')
    c = ws['B3']
    c.value = ('Что должно быть на каждой странице для продаж - и чего не хватает. '
               'Красное нужно чинить, серый прочерк - этого просто нет (норма).')
    c.font = _font(size=11, color=C.text_soft)
    c.alignment = _align(wrap=True, vertical='center')
    ws.row_dimensions[3].height = 18

    # ── Дашборд: 3 карточки на всю ширину (B-E, F-I, J-M) ──
    cards = [
        (total_pages, 'ПРОВЕРЕНО СТРАНИЦ', C.accent, C.accent_soft),
        (ok_pages, 'БЕЗ ПРОБЛЕМ', C.ok, C.ok_soft),
        (pages_with_bugs, 'НУЖНО ПОЧИНИТЬ',
         C.err if pages_with_bugs else C.ok, C.err_soft if pages_with_bugs else C.ok_soft),
    ]
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[6].height = 16
    for i, (value, label, color, bg) in enumerate(cards):
        c1 = 2 + i * 4
        c2 = c1 + 3
        fill_block(5, c1, 6, c2, bg)
        ws.merge_cells(start_row=5, start_column=c1, end_row=5, end_column=c2)
        v = ws.cell(row=5, column=c1, value=value)
        v.font = _font(size=26, bold=True, color=color)
        v.alignment = _align(horizontal='center', vertical='center')
        ws.merge_cells(start_row=6, start_column=c1, end_row=6, end_column=c2)
        l = ws.cell(row=6, column=c1, value=label)
        l.font = _font(size=9, bold=True, color=C.text_muted)
        l.alignment = _align(horizontal='center')

    # ── «Что чинить» - главный блок ──
    bug_pages = [r for r in pages if r.content_bugs > 0]
    # Тех. страницы с расхождением контактов с КП (адреса городов / телефон) -
    # тоже выводим наверх как ошибку.
    for r in results:
        if getattr(r, 'type_code', '') == 'tech' and (
                (getattr(r, 'content_bugs', 0) or 0) > 0
                or _contacts_problem_text(r) or _broken_links_text(r)):
            bug_pages.append(r)
    bug_pages = sorted(bug_pages, key=lambda r: -(getattr(r, 'content_bugs', 0) or 0))
    row = 8
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
    hc = ws.cell(row=row, column=2)
    if bug_pages:
        hc.value = f'  Что чинить - {len(bug_pages)} {_plural_pages(len(bug_pages))}'
        hc.font = _font(size=14, bold=True, color=C.err)
        hc.fill = _fill(C.err_soft)
    else:
        hc.value = '  ✓ Всё в порядке - структурных проблем не найдено'
        hc.font = _font(size=14, bold=True, color=C.ok)
        hc.fill = _fill(C.ok_soft)
    hc.alignment = _align(indent=1, vertical='center')
    for cc in range(2, last_col + 1):
        ws.cell(row=row, column=cc).fill = _fill(C.err_soft if bug_pages else C.ok_soft)
    ws.row_dimensions[row].height = 26
    row += 1

    if bug_pages:
        # Шапка списка
        for ci, h in [(2, 'Город'), (3, 'Тип страницы'), (4, 'Открыть'),
                      (5, 'Что не так')]:
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = _font(size=9, bold=True, color=C.text_muted)
            cell.fill = _fill(C.surface)
            cell.alignment = _align(indent=1)
            cell.border = _border()
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=last_col)
        for cc in range(5, last_col + 1):
            ws.cell(row=row, column=cc).fill = _fill(C.surface)
            ws.cell(row=row, column=cc).border = _border()
        row += 1
        for idx, r in enumerate(bug_pages[:50]):
            band = C.surface if idx % 2 else C.bg_elev
            kind = _KIND_LABEL.get(getattr(r.content, 'page_kind', ''), r.type_label)
            cc = ws.cell(row=row, column=2, value=r.city)
            cc.font = _font(size=10, bold=True); cc.fill = _fill(band)
            cc.alignment = _align(indent=1); cc.border = _border(color=C.border_light)
            kc = ws.cell(row=row, column=3, value=kind)
            kc.font = _font(size=10, color=C.text_soft); kc.fill = _fill(band)
            kc.alignment = _align(indent=1); kc.border = _border(color=C.border_light)
            uc = ws.cell(row=row, column=4, value='открыть')
            uc.hyperlink = r.url
            uc.font = _font(size=10, color=C.accent, underline='single')
            uc.fill = _fill(band)
            uc.alignment = _align(horizontal='center'); uc.border = _border(color=C.border_light)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=last_col)
            _ptext = _problem_text(r)
            mc = ws.cell(row=row, column=5, value=_ptext)
            mc.font = _font(size=10, color=C.err)
            # Одна строка фиксированной высоты: длинный текст (списки адресов)
            # визуально обрезается, НЕ раздувая таблицу. Полный текст - в
            # тултипе (навести курсор), в строке формул (клик по ячейке) или
            # растянув строку вручную. Данные не меняем - только отображение.
            mc.alignment = _align(indent=1, wrap=True)
            ws.row_dimensions[row].height = 20
            if len(_ptext) > 100:
                mc.comment = Comment(_ptext, 'Site Checker', height=260, width=420)
            for cc2 in range(5, last_col + 1):
                ws.cell(row=row, column=cc2).fill = _fill(band)
                ws.cell(row=row, column=cc2).border = _border(color=C.border_light)
            row += 1
        if len(bug_pages) > 50:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
            ws.cell(row=row, column=2,
                    value=f'… и ещё {len(bug_pages) - 50} - см. таблицы ниже').font = \
                _font(size=10, italic=True, color=C.text_muted)
            row += 1

    # ── Подробные таблицы по типам ──
    row += 2
    ws.cell(row=row, column=2, value='Подробно по типам страниц').font = \
        _font(size=13, bold=True, color=C.text)
    ws.cell(row=row + 1, column=2,
            value='✓ есть · БАГ обязательного нет · «-» необязательного нет (норма) · '
                  'число = сколько найдено. Наведите курсор на заголовок столбца - пояснение.').font = \
        _font(size=9, italic=True, color=C.text_muted)
    ws.merge_cells(start_row=row + 1, start_column=2, end_row=row + 1, end_column=last_col)
    row += 3

    for group_label, predicate in _STRUCT_GROUPS:
        group_pages = [r for r in pages if predicate(r)]
        if not group_pages:
            continue
        columns = _grid_columns(group_pages[0].content.blocks)
        n_cols = len(columns)
        g_bugs = sum(r.content_bugs for r in group_pages)

        # Заголовок секции
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4 + n_cols)
        gc = ws.cell(row=row, column=2)
        gc.value = (f'  {group_label} - {len(group_pages)} стр.'
                    + (f'  ·  проблем: {g_bugs}' if g_bugs else '  ·  все в порядке'))
        gc.font = _font(size=11, bold=True, color=C.err if g_bugs else C.ok)
        gc.fill = _fill(C.accent_soft)
        gc.alignment = _align(indent=1, vertical='center')
        for cc in range(2, 5 + n_cols):
            ws.cell(row=row, column=cc).fill = _fill(C.accent_soft)
        ws.row_dimensions[row].height = 22
        row += 1

        # Шапка таблицы
        headers = ([('Город', ''), ('Открыть', ''), ('Проблем', '')]
                   + [(c['label'], c['desc']) for c in columns])
        hdr_row = row
        for ci, (h, desc) in enumerate(headers, start=2):
            cell = ws.cell(row=hdr_row, column=ci)
            cell.value = h
            cell.font = _font(size=9, bold=True, color=C.bg_elev)
            cell.fill = _fill(C.header_navy)
            cell.alignment = _align(horizontal='center', wrap=True, indent=0)
            cell.border = _border()
            if desc:
                cell.comment = Comment(desc, 'Site Checker', height=120, width=260)
        ws.row_dimensions[hdr_row].height = 54
        row += 1

        for idx, r in enumerate(group_pages):
            by_key = {b.key: b for b in r.content.blocks}
            band = C.surface if idx % 2 else C.bg_elev

            cc = ws.cell(row=row, column=2, value=r.city)
            cc.font = _font(size=10); cc.fill = _fill(band)
            cc.alignment = _align(indent=1); cc.border = _border(color=C.border_light)

            uc = ws.cell(row=row, column=3, value='открыть')
            uc.hyperlink = r.url
            uc.font = _font(size=10, color=C.accent, underline='single')
            uc.fill = _fill(band)
            uc.alignment = _align(horizontal='center', indent=0)
            uc.border = _border(color=C.border_light)

            pc = ws.cell(row=row, column=4)
            pc.value = r.content_bugs if r.content_bugs else ''
            pc.font = _font(size=11, bold=True, color=C.err)
            pc.alignment = _align(horizontal='center', indent=0)
            pc.fill = _fill(C.err_soft) if r.content_bugs else _fill(band)
            pc.border = _border(color=C.border_light)

            if getattr(r.content, 'is_soft_404', False) and n_cols:
                ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=4 + n_cols)
                cell = ws.cell(row=row, column=5, value='Страница отдаёт 404 (не найдена)')
                cell.font = _font(size=10, bold=True, color=C.err)
                cell.alignment = _align(indent=1)
                for k in range(n_cols):
                    cm = ws.cell(row=row, column=5 + k)
                    cm.fill = _fill(C.err_soft); cm.border = _border(color=C.border_light)
                row += 1
                continue

            for ci, col in enumerate(columns):
                cell = ws.cell(row=row, column=5 + ci)
                cell.alignment = _align(horizontal='center', indent=0)
                cell.border = _border(color=C.border_light)
                value, state = _cell_state(col, by_key)
                _style_cell(cell, value, state)
                if state in ('absent', 'count', 'okinfo'):
                    cell.fill = _fill(band)
                # У заглушки фото - всплывающая подсказка с названиями товаров.
                if state == 'warn' and col.get('kind') == 'block':
                    _b = by_key.get(col.get('key'))
                    _nm = getattr(_b, 'note', '') if _b else ''
                    if _nm:
                        cell.comment = Comment('Стоит заглушка «нет фото» у товаров: '
                                               + _nm, 'Site Checker', height=120, width=300)
            row += 1
        row += 2  # пробел между секциями

    # ── Технические страницы (оплата, доставка, контакты, реквизиты, политики,
    # карта сайта) ── Проверяем их «как все»: доступность (открывается / 404 /
    # ошибка) + структуру (H1, хлебные крошки) + битые переменные. H1 обязателен;
    # крошки справочно (их отсутствие на служебной странице багом не считаем).
    from urllib.parse import urlparse as _urlparse
    from sources import tech_page_label as _tech_label
    tech = [r for r in results if getattr(r, 'type_code', '') == 'tech']
    if tech:
        def _tech_bad(r):
            if not r.is_ok:
                return True
            if getattr(r.content, 'is_soft_404', False):
                return True
            return bool(r.content_bugs or r.has_text_issues
                        or _broken_links_text(r))
        _bad = sum(1 for r in tech if _tech_bad(r))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        gc = ws.cell(row=row, column=2)
        gc.value = (f'  Технические страницы - {len(tech)} стр.'
                    + (f'  ·  проблем: {_bad}' if _bad else '  ·  все в порядке'))
        gc.font = _font(size=11, bold=True, color=C.err if _bad else C.ok)
        gc.fill = _fill(C.accent_soft)
        gc.alignment = _align(indent=1, vertical='center')
        for cc in range(2, 11):
            ws.cell(row=row, column=cc).fill = _fill(C.accent_soft)
        ws.row_dimensions[row].height = 22
        row += 1
        _tech_headers = [
            (2, 'Страница', 'Название страницы - кликабельная ссылка, ведёт на страницу.'),
            (3, 'Статус', 'Открывается ли страница: «Работает» / код ответа (404 и т.п.) / «404-заглушка» (отдаёт 200, но контент «страница не найдена»).'),
            (4, 'Проблем', 'Сколько проблем на странице: структурные баги, битые переменные, расхождения контактов с КП и битые ссылки (404).'),
            (5, 'H1', 'Заголовок H1. Обязателен - у нормальной страницы он есть.'),
            (6, 'Крошки', 'Хлебные крошки. Справочно: показываем есть/нет, отсутствие на служебной странице не баг.'),
            (7, 'Текст', 'Есть ли на странице собственный текст (помимо сквозных шапки и подвала). Обязателен.'),
            (8, 'Битые перем.', 'Битые шаблонные переменные ({{…}}, %name% и т.п.). Число = сколько найдено.'),
            (9, 'Элементы страницы', 'Спец-проверки в зависимости от страницы: картинки, ссылка на каталог, карта, форма обратной связи, строка поиска (✓ есть / - нет / БАГ - обязательного нет). Обязательны: карта на «Контактах», картинки на «О компании», строка поиска на странице поиска. Если включена проверка ссылок - тут же «Ссылки: N ✓» или «N битых» (404/410).'),
            (10, 'Что не так', 'Подробно: структурные баги (нет карты/картинок/строки поиска и т.п.) и расхождения контактов с КП (адреса городов / телефон страницы).'),
        ]
        hdr_row = row
        for ci, h, desc in _tech_headers:
            cell = ws.cell(row=hdr_row, column=ci, value=h)
            cell.font = _font(size=9, bold=True, color=C.text_muted)
            cell.fill = _fill(C.surface)
            cell.alignment = _align(horizontal='center', wrap=True, indent=0)
            cell.border = _border()
            if desc:
                cell.comment = Comment(desc, 'Site Checker', height=120, width=260)
        ws.row_dimensions[hdr_row].height = 40
        row += 1
        for idx, r in enumerate(tech):
            band = C.surface if idx % 2 else C.bg_elev

            # Страница - человеческое название (Оплата, Доставка…) как ссылка.
            try:
                _path = _urlparse(r.url).path or r.url
            except Exception:
                _path = r.url
            pgc = ws.cell(row=row, column=2, value=_tech_label(_path))
            pgc.hyperlink = r.url
            pgc.font = _font(size=10, color=C.accent, underline='single')
            pgc.fill = _fill(band)
            pgc.alignment = _align(indent=1)
            pgc.border = _border(color=C.border_light)

            _soft = getattr(r.content, 'is_soft_404', False)
            if not r.is_ok:
                _status = str(r.http_code) if r.http_code else 'не открылась'
            elif _soft:
                _status = '404-заглушка'
            else:
                _status = 'Работает'
            _status_ok = r.is_ok and not _soft
            sc = ws.cell(row=row, column=3, value=_status)
            sc.font = _font(size=10, bold=not _status_ok, color=C.ok if _status_ok else C.err)
            sc.fill = _fill(band if _status_ok else C.err_soft)
            sc.alignment = _align(horizontal='center', indent=0)
            sc.border = _border(color=C.border_light)

            _probs = (r.content_bugs or 0) + len(r.text_issues or [])
            _ca = getattr(r, 'contacts_addr', None)
            if _ca:
                _probs += len(_ca.get('mismatched') or [])
            _pp = getattr(r, 'page_phone', None)
            if _pp and _pp.get('status') in ('bug', 'critical'):
                _probs += 1
            _blk = getattr(r, 'broken_links', None)
            _broken_n = len(_blk['broken']) if (_blk and _blk.get('broken')) else 0
            _probs += _broken_n
            pc = ws.cell(row=row, column=4)
            pc.value = _probs if _probs else ''
            pc.font = _font(size=11, bold=True, color=C.err)
            pc.alignment = _align(horizontal='center', indent=0)
            pc.fill = _fill(C.err_soft) if _probs else _fill(band)
            pc.border = _border(color=C.border_light)

            # H1 / Крошки: если страница не открылась или это 404-заглушка -
            # структуры нет, ставим «-». Иначе берём из блоков контента.
            by_key = {b.key: b for b in r.content.blocks} if (r.is_ok and r.content) else {}
            for ci, key in ((5, 'h1'), (6, 'breadcrumbs'), (7, 'content_text')):
                cell = ws.cell(row=row, column=ci)
                cell.alignment = _align(horizontal='center', indent=0)
                cell.border = _border(color=C.border_light)
                if not by_key or _soft:
                    cell.value = '-'; cell.font = _font(size=10, color=C.text_muted)
                    cell.fill = _fill(band)
                else:
                    value, state = _cell_state({'kind': 'block', 'key': key}, by_key)
                    _style_cell(cell, value, state)
                    if state in ('absent', 'count', 'okinfo'):
                        cell.fill = _fill(band)

            # Битые переменные - число найденных.
            _ti = len(r.text_issues or []) if r.is_ok else 0
            vc = ws.cell(row=row, column=8)
            vc.alignment = _align(horizontal='center', indent=0)
            vc.border = _border(color=C.border_light)
            if _ti:
                vc.value = _ti; vc.font = _font(size=10, bold=True, color=C.err)
                vc.fill = _fill(C.err_soft)
            else:
                vc.value = '-'; vc.font = _font(size=10, color=C.text_muted)
                vc.fill = _fill(band)

            # Элементы страницы - спец-проверки (картинки/каталог-ссылка/карта/форма)
            # + краткий итог сверки адресов/телефона с КП.
            _spec = [b for b in (r.content.blocks if (r.is_ok and r.content) else [])
                     if b.key.startswith('tech_')]
            _addr_bad = False
            _parts = []
            for b in _spec:
                if b.required and not b.present:
                    _parts.append(f'{b.label}: БАГ')   # обязательный элемент, а его нет
                    _addr_bad = True
                else:
                    _parts.append(f'{b.label} {"✓" if b.present else "-"}')
            if _ca:
                _mm = _ca.get('mismatched') or []
                _txt = f'Адреса городов {_ca.get("matched", 0)}/{_ca.get("on_page", 0)}'
                if _mm:
                    _txt += f' · расхождений {len(_mm)}'
                    _addr_bad = True
                _parts.append(_txt)
            if _pp:
                _ps = _pp.get('status')
                _parts.append('Телефон ' + {'ok': '✓', 'info': 'инфо'}.get(_ps, 'расхождение'))
                if _ps in ('bug', 'critical'):
                    _addr_bad = True
            if _broken_n:
                _parts.append(f'Ссылки: {_broken_n} битых')
                _addr_bad = True
            elif _blk:                       # проверяли - все ссылки открылись
                _parts.append(f'Ссылки: {_blk.get("checked", 0)} ✓')
            ec = ws.cell(row=row, column=9)
            ec.alignment = _align(indent=1)
            ec.border = _border(color=C.border_light)
            ec.fill = _fill(C.err_soft if _addr_bad else band)
            ec.value = ' · '.join(_parts) if _parts else '-'
            ec.font = _font(size=9, color=C.err if _addr_bad else
                            (C.text_soft if _parts else C.text_muted))

            # Что не так - подробно: структурные баги (нет карты/картинок/строки
            # поиска и т.п.) и расхождения контактов с КП. Пусто, если проблем нет.
            _has_problem = ((r.content_bugs or 0) > 0
                            or bool(_contacts_problem_text(r)) or _broken_n > 0)
            _wn = _problem_text(r) if _has_problem else ''
            wn = ws.cell(row=row, column=10, value=_wn or '-')
            # Одна строка: длинные списки не растягивают таблицу; полный текст -
            # в тултипе / строке формул / при растяжении строки вручную.
            wn.alignment = _align(indent=1, wrap=True)
            ws.row_dimensions[row].height = 20
            if len(_wn) > 100:
                wn.comment = Comment(_wn, 'Site Checker', height=260, width=420)
            wn.border = _border(color=C.border_light)
            wn.fill = _fill(C.err_soft if _wn else band)
            wn.font = _font(size=9, color=C.err if _wn else C.text_muted)
            row += 1
        row += 2


# ── Лист «Страницы» - короткая сводка всех проверенных страниц ──────


def _build_pages_overview_sheet(wb, results, findings, w3c_check=None):
    """Лист «Страницы»: код/статус/скорость/битые переменные/откуда перешли
    + сколько находок на странице (из «Проблемы») по каждой проверенной
    странице - без деталей самих находок (те - на «Проблемы»). Плюс, если
    была валидация W3C (выборка страниц) - ошибок валидатора и время
    загрузки ресурсов (детали по типам ресурсов - на «Валидация и скорость»)."""
    if not results:
        return
    from collections import Counter
    counts = Counter(f.url for f in findings if f.url)
    w3c_by_url = {p.get('url'): p for p in (w3c_check or {}).get('pages') or []}

    ws = wb.create_sheet('Страницы')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = (
        C.err if any(r.is_error for r in results) else
        C.warn if any(r.is_warning for r in results) else C.ok)

    # «Источник» - откуда взялся адрес: обычная выборка по каталогу проекта,
    # случайная проверка карт сайта, свой список или тех. страницы. Без него в
    # таблице не отличить страницу каталога от случайной из sitemap.
    headers = [('Город', 16), ('Поддомен', 24), ('Тип', 13), ('Источник', 16),
              ('Адрес страницы', 55), ('Код', 8), ('Статус', 18),
              ('Скорость, с', 11), ('Оценка скорости', 16),
              ('Кому чинить', 18), ('Найдено проблем', 14),
              ('Битые переменные', 16), ('Откуда перешли', 46)]
    if w3c_by_url:
        headers += [('W3C ошибок', 12), ('Загрузка, мс', 12)]
    last_col = get_column_letter(1 + len(headers))
    ws.column_dimensions['A'].width = 3
    for i, (_title, w) in enumerate(headers, 2):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(f'B2:{last_col}2')
    c = ws['B2']
    c.value = f'Все проверенные страницы - {len(results)}'
    c.font = _font(size=16, bold=True)
    ws.row_dimensions[2].height = 26

    ws.merge_cells(f'B3:{last_col}3')
    c = ws['B3']
    c.value = ('Быстрый обзор: код ответа, статус, скорость и сколько находок '
              'на странице (сами находки - на листе «Проблемы»).'
              + (' W3C/загрузка - только для страниц из выборки валидации '
                 '(детали - «Валидация и скорость»).' if w3c_by_url else ''))
    c.font = _font(size=10, italic=True, color=C.text_soft)
    ws.row_dimensions[3].height = 18

    hdr_row = 5
    for i, (title, _w) in enumerate(headers, 2):
        cell = ws.cell(row=hdr_row, column=i)
        cell.value = title
        cell.font = _font(size=10, bold=True, color=C.bg_elev)
        cell.fill = _fill(C.header_navy)
        cell.border = _border()
        cell.alignment = _align(indent=1)
    ws.row_dimensions[hdr_row].height = 20

    def sort_key(r):
        score = 0 if r.is_error else 1 if r.is_warning else 2 if r.has_text_issues else 3
        return (score, r.city or '')
    sorted_results = sorted(results, key=sort_key)

    row = hdr_row + 1
    for r in sorted_results:
        speed_sec = ''
        if r.elapsed_ms is not None:
            speed_sec = f'{r.elapsed_ms / 1000:.2f}'.replace('.', ',')
        speed_label = SPEED_LABEL.get(r.speed_rating, '') if r.speed_rating else ''
        n_prob = counts.get(r.url, 0)
        text_issue_text = ''
        if r.has_text_issues:
            n = len(r.text_issues)
            text_issue_text = f'{n} {"находка" if n == 1 else "находок"}'
        vals = [r.city, r.subdomain, r.type_label,
               getattr(r, 'source', '') or 'Каталог проекта', r.url,
               r.http_code if r.http_code else '-',
               STATUS_LABEL.get(r.status, r.status), speed_sec, speed_label,
               _dept_result(r), n_prob or None, text_issue_text,
               _build_path_description(r)]
        if w3c_by_url:
            wp = w3c_by_url.get(r.url)
            w3c_errors = w3c_total_ms = None
            if wp and not wp.get('error'):
                _h, _cs = wp.get('html') or {}, wp.get('css') or {}
                if not _h.get('error') and not _cs.get('error'):
                    w3c_errors = (_h.get('errors', 0) or 0) + (_cs.get('errors', 0) or 0)
                w3c_total_ms = (wp.get('timings') or {}).get('total_ms')
            vals += [w3c_errors, w3c_total_ms]
        for ci, v in enumerate(vals, 2):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = _font(size=9, color=C.text_soft)
            cell.alignment = _align(indent=1)
            cell.border = _border(color=C.border_light)
        ws.cell(row=row, column=3).font = _font(name='Consolas', size=9,
                                                color=C.text_muted)
        # Источник: адрес не из каталога проекта выделяем - именно он объясняет,
        # почему в прогоне вдруг оказалась незнакомая страница.
        _src_cell = ws.cell(row=row, column=5)
        if (getattr(r, 'source', '') or 'Каталог проекта') != 'Каталог проекта':
            _src_cell.font = _font(size=9, bold=True, color=C.accent)
            _src_cell.fill = _fill(C.accent_soft)
        url_cell = ws.cell(row=row, column=6)
        url_cell.hyperlink = r.url
        url_cell.font = _font(name='Consolas', size=9, color=C.accent,
                              underline='single')
        status_cell = ws.cell(row=row, column=8)
        status_color = C.ok if r.is_ok else C.warn if r.is_warning else C.err
        status_fill = C.ok_soft if r.is_ok else C.warn_soft if r.is_warning else C.err_soft
        status_cell.font = _font(size=9, bold=True, color=status_color)
        status_cell.fill = _fill(status_fill)
        if r.speed_rating:
            speed_cell = ws.cell(row=row, column=10)
            speed_cell.font = _font(size=9, bold=True, color=SPEED_COLOR[r.speed_rating])
            speed_cell.fill = _fill(SPEED_FILL[r.speed_rating])
        if n_prob:
            prob_cell = ws.cell(row=row, column=12)
            prob_cell.font = _font(size=9, bold=True, color=C.err)
            prob_cell.fill = _fill(C.err_soft)
        if r.has_text_issues:
            issue_cell = ws.cell(row=row, column=13)
            issue_cell.font = _font(size=9, bold=True, color=C.warn)
            issue_cell.fill = _fill(C.warn_soft)
        path_cell = ws.cell(row=row, column=14)
        if r.redirect_chain:
            path_cell.font = _font(name='Consolas', size=9, color=C.text_soft)
        elif not r.is_ok:
            path_cell.font = _font(size=9, italic=True, color=C.text_muted)
        if w3c_by_url and w3c_errors:
            err_cell = ws.cell(row=row, column=15)
            err_cell.font = _font(size=9, bold=True, color=C.warn)
            err_cell.fill = _fill(C.warn_soft)
        ws.row_dimensions[row].height = 16
        row += 1

    last = row - 1
    ws.auto_filter.ref = f'B{hdr_row}:{last_col}{last}'
    ws.freeze_panes = f'B{hdr_row + 1}'


# ── Лист «Хосты и аномалии» - проблемы уровня сайта/хоста целиком ───


def _build_hosts_anomalies_sheet(wb, service_issues, wm_metrics, link_profile,
                                 anomalies=None):
    """Лист «Хосты и аномалии»: фатальные проблемы из сервисов (Вебмастер/
    GSC/Метрика) + аномалии обхода/ссылок «от себя-прошлого» + всплески
    мусорных/иноязычных запросов (ГСК) и переходов со спам-сайтов
    (Метрика). Не привязаны к одной странице - раньше были погребены
    внутри группового листа «Аналитика», здесь вынесены наверх (обычно
    самые срочные)."""
    fatal = [i for i in (service_issues or [])
             if getattr(i, 'severity', None) == 'fatal']
    anomaly_rows = (_collect_anomaly_rows(wm_metrics, link_profile)
                    if wm_metrics and wm_metrics.get('available') else [])
    _a = anomalies or {}
    gsc = _a.get('gsc') or {}
    mtr = _a.get('metrika') or {}
    _gsc_bad = gsc.get('available') and (gsc.get('spiked') or gsc.get('spam_queries_count'))
    _mtr_bad = mtr.get('available') and (mtr.get('spiked') or mtr.get('spam_domains_count'))
    if not fatal and not anomaly_rows and not _gsc_bad and not _mtr_bad:
        return

    n_red = sum(1 for a in anomaly_rows if a.get('severity') in ('fatal', 'critical'))
    ws = wb.create_sheet('Хосты и аномалии')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C.err if (fatal or n_red) else C.warn

    ws.column_dimensions['A'].width = 3
    for col, w in (('B', 26), ('C', 30), ('D', 16), ('E', 10), ('F', 90)):
        ws.column_dimensions[col].width = w

    ws.merge_cells('B2:F2')
    c = ws['B2']
    c.value = 'Проблемы уровня сайта: фатальные проблемы и аномалии обхода'
    c.font = _font(size=16, bold=True)
    ws.row_dimensions[2].height = 26

    ws.merge_cells('B3:F3')
    c = ws['B3']
    c.value = ('Находки не по страницам, а по хостам целиком - обычно самые '
              'срочные (сайт может частично или полностью выпасть из индекса).')
    c.font = _font(size=10, italic=True, color=C.text_soft)
    ws.row_dimensions[3].height = 18

    row = 5

    def _hdr(cols, r):
        for col, title in cols:
            cell = ws[f'{col}{r}']
            cell.value = title
            cell.font = _font(size=9, bold=True, color=C.bg_elev)
            cell.fill = _fill(C.header_navy)
            cell.border = _border()
            cell.alignment = _align(indent=1, wrap=True)
        ws.row_dimensions[r].height = 20

    if fatal:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        c = ws.cell(row=row, column=2,
                    value=f'Фатальные проблемы в сервисах  ({len(fatal)})')
        c.font = _font(size=13, bold=True, color=C.err)
        c.fill = _fill(C.accent_soft)
        c.alignment = _align(indent=1)
        ws.row_dimensions[row].height = 24
        row += 1
        _hdr((('B', 'Хост'), ('C', 'Что нашли'), ('D', 'Дата')), row)
        row += 1
        for i in fatal:
            for col, v in ((2, i.host), (3, i.title), (4, i.date)):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = _font(size=10, color=C.text)
                cell.fill = _fill(C.err_soft)
                cell.border = _border(color=C.border_light)
                cell.alignment = _align(indent=1, wrap=True)
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1

    if anomaly_rows:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        c = ws.cell(row=row, column=2,
                    value=f'Аномалии обхода/ссылок («от себя-прошлого»)  '
                          f'({len(anomaly_rows)})')
        c.font = _font(size=13, bold=True, color=C.err if n_red else C.warn)
        c.fill = _fill(C.accent_soft)
        c.alignment = _align(indent=1)
        ws.row_dimensions[row].height = 24
        row += 1
        _hdr((('B', 'Хост'), ('C', 'Метрика'), ('D', 'Было → сейчас'),
             ('E', 'Уровень'), ('F', 'Что это значит')), row)
        row += 1
        _F_WIDTH = 90
        for a in anomaly_rows:
            fatal_like = a.get('severity') in ('fatal', 'critical')
            text = a.get('text', '')
            vals = [(2, a.get('host', '')), (3, a.get('metric', '')),
                   (4, _fmt_ba(a.get('before'), a.get('after'))),
                   (5, _ANOM_SEV.get(a.get('severity'), (9, ''))[1]),
                   (6, text)]
            for col, v in vals:
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = _font(size=10, color=C.text)
                cell.border = _border(color=C.border_light)
                cell.alignment = _align(indent=1, wrap=(col == 6), vertical='top')
                if fatal_like:
                    cell.fill = _fill(C.err_soft)
            if a.get('panel_url'):
                hc = ws.cell(row=row, column=2)
                hc.hyperlink = a['panel_url']
                hc.font = _font(size=10, color=C.accent, underline='single')
            ws.row_dimensions[row].height = _row_height_for(text, _F_WIDTH, min_px=18)
            row += 1

    # ── Всплески мусорных/иноязычных запросов (ГСК) и переходов со
    # спам-сайтов (Метрика) - проектная сводка, не по одному хосту. ──
    if _gsc_bad or _mtr_bad:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        c = ws.cell(row=row, column=2, value='Всплески мусорных запросов и переходов')
        c.font = _font(size=13, bold=True, color=C.err)
        c.fill = _fill(C.accent_soft)
        c.alignment = _align(indent=1)
        ws.row_dimensions[row].height = 24
        row += 1

        def _q_line(text, color=C.text, bold=False):
            nonlocal row
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            cc = ws.cell(row=row, column=2, value=text)
            cc.font = _font(size=10, color=color, bold=bold)
            cc.alignment = _align(indent=1, wrap=True)
            ws.row_dimensions[row].height = 18
            row += 1

        if _gsc_bad:
            _q_line(
                f'Google Search Console: показы сейчас/было '
                f'{gsc.get("total_impr_cur")}/{gsc.get("total_impr_prev")}'
                + _ratio_word(gsc.get('impr_spike'), gsc.get('spiked'))
                + f' · мусорных/иноязычных запросов: '
                f'{gsc.get("spam_queries_count", 0)} (показов '
                f'{gsc.get("spam_impr_cur", 0)}, было {gsc.get("spam_impr_prev", 0)})',
                color=C.err, bold=True)
            for q in (gsc.get('spam_queries') or [])[:12]:
                _q_line(f'  «{q.get("query", "")}» - {q.get("impressions")} показов',
                       color=C.text_soft)
        if _mtr_bad:
            _q_line(
                f'Метрика: переходы-рефералы сейчас/было '
                f'{mtr.get("total_cur")}/{mtr.get("total_prev")}'
                + _ratio_word(mtr.get('referral_spike'), mtr.get('spiked'))
                + f' · спам-доменов-рефереров: {mtr.get("spam_domains_count", 0)} '
                f'(переходов {mtr.get("spam_cur", 0)}, было {mtr.get("spam_prev", 0)})',
                color=C.err, bold=True)
            for d in (mtr.get('spam_domains') or [])[:15]:
                _q_line(f'  {d.get("domain", "")} - {d.get("visits")} переходов',
                       color=C.text_soft)


# ── Лист уведомлений ──────────────────────────────────────────────

_NOTIF_PRIORITY_ORDER = ['critical', 'important', 'recommendation', 'info']
_NOTIF_PRIORITY_LABEL = {
    'critical':       '🔴 Критические',
    'important':      '🟠 Важные',
    'recommendation': '🟡 Рекомендации',
    'info':           '⚪ Инфо',
}
_NOTIF_PRIORITY_COLOR = {
    'critical':       C.err,
    'important':      C.warn,
    'recommendation': 'CA8A04',
    'info':           C.text_muted,
}
_NOTIF_PRIORITY_BG = {
    'critical':       C.err_soft,
    'important':      C.warn_soft,
    'recommendation': 'FEFCE8',
    'info':           C.surface,
}
_NOTIF_CATEGORY_LABEL = {
    'server':     'Сервер',
    'indexing':   'Индексирование',
    'speed':      'Скорость',
    'security':   'Безопасность',
    'structure':  'Структура',
    'coverage':   'Покрытие',
    'other':      'Прочее',
}

# Группировка уведомлений по теме (один и тот же текст письма приходит по
# каждому домену отдельно - схлопываем в одну строку, домены в список).
_DOMAIN_TLDS = (
    # рф/ru/su + СНГ/региональные зоны (.kz/.kg/.uz/.ua и т.д.) - чтобы один
    # бренд в разных зонах не дробил тему на отдельные строки + gTLD.
    'ru|рф|su|by|kz|kg|uz|ua|am|az|ge|md|tj|tm|ee|lv|lt|'
    'com|net|org|info|biz|pro|online|store|site|shop|me|cc|io'
)
# URL c путём целиком (group1 = host+path) - для извлечения режем по '/'.
_URL_RE = re.compile(r'https?://([^\s,;()<>"\']+)', re.IGNORECASE)
_HOST_RE = re.compile(
    r'\b((?:[a-zа-я0-9](?:[a-zа-я0-9-]*[a-zа-я0-9])?\.)+(?:' + _DOMAIN_TLDS + r'))\b',
    re.IGNORECASE,
)


def _extract_domains(text: str) -> list:
    """Вытащить хосты/домены из текста темы письма (URL и «голые» хосты)."""
    if not text:
        return []
    raw = [m.group(1).split('/')[0] for m in _URL_RE.finditer(text)]
    raw += [m.group(1) for m in _HOST_RE.finditer(text)]
    seen, out = set(), []
    for d in raw:
        d = d.strip('.').lower()
        if d.startswith('www.'):
            d = d[4:]
        if d and d not in seen:
            seen.add(d)
            out.append(d)
    return out


def _canon_theme(subject: str) -> str:
    """Тема без конкретного домена/URL - ключ группировки и текст для отчёта."""
    s = subject or ''
    s = _URL_RE.sub('', s)
    s = _HOST_RE.sub('', s)
    s = re.sub(r'\s+', ' ', s).strip(' .,:;/---«»"\'')
    return s or (subject or '').strip()


def _group_notifs_by_theme(items: list) -> list:
    """Схлопнуть письма с одинаковой темой в группы.

    Возвращает список dict: theme, date (минимальная), domains (список),
    first (репрезентативное письмо), count. Порядок - первое появление темы.
    """
    from collections import OrderedDict
    groups = OrderedDict()
    for n in items:
        key = _canon_theme(n.subject)
        g = groups.get(key)
        if g is None:
            g = {'theme': key, 'date': n.date, 'first': n,
                 'domains': [], 'count': 0}
            groups[key] = g
        g['count'] += 1
        if n.date and (not g['date'] or n.date < g['date']):
            g['date'] = n.date
        for d in _extract_domains(n.subject):
            if d not in g['domains']:
                g['domains'].append(d)
    return list(groups.values())


def _notif_row_height(domains_str: str, preview: str) -> float:
    """Высота строки под перенос длинного списка доменов / превью."""
    import math
    dom_lines = max(1, math.ceil(len(domains_str or '') / 50))
    prev_lines = max(1, math.ceil(len((preview or '')[:400]) / 70))
    lines = max(dom_lines, prev_lines)
    return min(300, max(44, lines * 14))


# Оценка отзыва 2ГИС → текст звёзд + ярлык качества + цвет.
def _review_rating_cell(rating):
    """(текст, цвет) для колонки «Оценка». rating: 1..5 или None."""
    if not rating:
        return '-', C.text_muted
    stars = '★' * int(rating)
    if rating >= 4:
        return f'{stars} Хороший', C.ok
    if rating == 3:
        return f'{stars} Средний', C.warn
    return f'{stars} Плохой', C.err


# Отдел для ошибки сервиса (Вебмастер-API): серверное → разработка, иначе SEO.
def _dept_service_issue(i) -> str:
    code = (getattr(i, 'code', '') or '').upper()
    if any(k in code for k in ('SERVER', 'DNS', 'SLOW', 'RESPONSE', 'THREAT',
                               'SITE_NOT_LOADED', 'SITE_ERROR', '5XX')):
        return 'разработка'
    return 'SEO'


_SEV2PRIO = {'fatal': 'critical', 'critical': 'critical', 'possible': 'important',
             'recommendation': 'recommendation', 'info': 'info'}


def _group_service_issues(items: list) -> list:
    """Схлопнуть ошибки сервиса по одной проблеме: один и тот же тип проблемы
    приходит по каждому сайту отдельно - собираем сайты в список.
    Возвращает dict: title, code, hosts (список), date (мин), count, first."""
    from collections import OrderedDict, Counter
    groups = OrderedDict()
    for i in items:
        title = getattr(i, 'title', '') or getattr(i, 'code', '')
        key = (title, getattr(i, 'code', ''))
        g = groups.get(key)
        if g is None:
            g = {'title': title, 'code': getattr(i, 'code', ''),
                 'hosts': [], 'date': getattr(i, 'date', ''),
                 'count': 0, 'first': i, 'states': Counter()}
            groups[key] = g
        g['count'] += 1
        host = getattr(i, 'host', '')
        if host and host not in g['hosts']:
            g['hosts'].append(host)
        st = getattr(i, 'state', '') or '-'
        g['states'][st] += 1
        d = getattr(i, 'date', '')
        if d and (not g['date'] or d < g['date']):
            g['date'] = d
    return list(groups.values())


# Коды состояния проблемы Вебмастера → человекочитаемо.
_WM_STATE_LABELS = {
    'IN_PROGRESS': 'на проверке',
    'CHECKING': 'на проверке',
    'UNDEFINED': 'на проверке',   # состояние не определено = идёт перепроверка
    'PROBLEM_ACTUAL': 'проблема актуальна',
    'PRESENT': 'проблема актуальна',
    'ACTUAL': 'проблема актуальна',
    'NEW': 'новая',
}


def _state_human(code: str):
    """Код состояния → текст. Пусто/«-» → None (не выводим).
    Старый кеш с уже-человеческим текстом - отдаём как есть."""
    s = (code or '').strip()
    if not s or s == '-':
        return None
    up = s.upper()
    if up in _WM_STATE_LABELS:
        return _WM_STATE_LABELS[up]
    return s.lower()


def _format_states(states) -> str:
    """Counter кодов состояния → «16 - на проверке. 45 - проблема актуальна».
    Коды агрегируются по человекочитаемой метке."""
    from collections import Counter
    agg = Counter()
    for code, n in states.items():
        h = _state_human(code)
        if h:
            agg[h] += n
    return '\n'.join(f'{n} - {label}' for label, n in agg.most_common())


# Секции в порядке убывания релевантности:
# (source_key, title, has_priority)
_NOTIF_SECTIONS = [
    ('yandex_webmaster', 'Вебмастер. Почта',        True),
    ('gsc',              'Google Search Console',   True),
    ('ya_business',      'Я.Бизнес',                False),
    ('twogis',           '2ГИС',                    False),
    ('google_accounts',  'Google',                  False),
]


def _build_notifications_sheet(wb, notifications):
    """Лист «Уведомления» - письма по источникам (Вебмастер/GSC/Я.Бизнес/
    2ГИС/Google) за период проверки. Структурирован секциями. Ошибки прямо
    из сервисов (не из почты) - отдельным листом «Ошибки сервисов», чтобы
    не дублировать одни и те же данные дважды. Добавляется всегда: при
    пустых данных показывает заглушку."""
    notifications = notifications or []
    ws = wb.create_sheet('Уведомления')
    ws.sheet_view.showGridLines = False

    has_critical = any(n.priority == 'critical' for n in notifications)
    ws.sheet_properties.tabColor = C.err if has_critical else C.accent

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 14   # Дата
    ws.column_dimensions['C'].width = 18   # Приоритет / пусто
    ws.column_dimensions['D'].width = 18   # Категория / пусто
    ws.column_dimensions['E'].width = 50   # Тема
    ws.column_dimensions['F'].width = 42   # Домены / Сайты
    ws.column_dimensions['G'].width = 60   # Превью / Состояние (Вебмастер-API)
    ws.column_dimensions['H'].width = 22   # Отдел / Кол-во
    ws.column_dimensions['I'].width = 22   # Отдел (секция Вебмастер-API)

    # ── Заголовок листа ──
    ws.merge_cells('B2:H2')
    c = ws['B2']
    c.value = 'Уведомления'
    c.font = _font(size=16, bold=True)
    ws.row_dimensions[2].height = 26

    ws.merge_cells('B3:H3')
    c = ws['B3']
    c.value = (
        'Письма от Яндекс.Вебмастера, GSC, Я.Бизнеса, 2ГИС и Google '
        'за период проверки. Красная вкладка = есть критические уведомления.'
    )
    c.font = _font(size=10, italic=True, color=C.text_soft)
    c.alignment = _align(wrap=True, vertical='top')
    ws.row_dimensions[3].height = 24

    # Нет писем - показываем заглушку и выходим
    if not notifications:
        ws.merge_cells('B5:H5')
        c = ws['B5']
        c.value = ('За период проверки писем не найдено. '
                   'Если ждёте уведомления - проверьте секреты почты и пароли приложений '
                   '(Gmail требует App Password), затем запустите прогон с галкой '
                   '«Собрать уведомления из почты».')
        c.font = _font(size=11, color=C.text_soft)
        c.alignment = _align(wrap=True, vertical='top')
        ws.row_dimensions[5].height = 60
        return

    # Разбиваем по источникам
    from collections import defaultdict
    by_source = defaultdict(list)
    for n in notifications:
        by_source[n.source].append(n)

    row = 5

    for source_key, section_title, has_priority in _NOTIF_SECTIONS:
        items = by_source.get(source_key, [])
        if not items:
            continue

        # ── Заголовок секции ──
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        sc = ws.cell(row=row, column=2)
        sc.value = f'{section_title}  ({len(items)})'
        sc.font = _font(size=13, bold=True, color=C.accent)
        sc.fill = _fill(C.accent_soft)
        sc.alignment = _align(indent=1)
        ws.row_dimensions[row].height = 24
        row += 1

        if has_priority:
            # ── Источник с классификацией: группируем по приоритету ──
            p_groups = defaultdict(list)
            for n in items:
                p_groups[n.priority].append(n)

            for priority in _NOTIF_PRIORITY_ORDER:
                p_items = p_groups.get(priority, [])
                if not p_items:
                    continue

                p_color = _NOTIF_PRIORITY_COLOR[priority]
                p_bg = _NOTIF_PRIORITY_BG[priority]
                p_label = _NOTIF_PRIORITY_LABEL[priority]

                # Подзаголовок приоритета
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
                pc = ws.cell(row=row, column=2)
                pc.value = f'  {p_label}  ({len(p_items)})'
                pc.font = _font(size=10, bold=True, color=p_color)
                pc.fill = _fill(p_bg)
                pc.alignment = _align(indent=2)
                ws.row_dimensions[row].height = 20
                row += 1

                # Шапка: одна строка на тему, домены списком + их количество
                for ci, h in enumerate(['Дата', 'Серьёзность', 'Категория', 'Тема',
                                        'Сайты', 'Кол-во', 'Отдел'], 2):
                    cell = ws.cell(row=row, column=ci)
                    cell.value = h
                    cell.font = _font(size=9, bold=True, color=C.text_muted)
                    cell.fill = _fill(C.surface)
                    cell.alignment = _align()
                    cell.border = _border()
                ws.row_dimensions[row].height = 20
                row += 1

                # Строки - одна на уникальную тему (без учёта доменной зоны),
                # все домены в колонке «Сайты», их число - в «Кол-во».
                groups = _group_notifs_by_theme(p_items)
                for g in sorted(groups, key=lambda x: len(x['domains']), reverse=True):
                    n0 = g['first']
                    domains_str = ', '.join(g['domains'])
                    ws.row_dimensions[row].height = _notif_row_height(domains_str, '')

                    for ci, (val, kw) in enumerate([
                        (g['date'], {'color': C.text_soft}),
                        (_NOTIF_PRIORITY_LABEL[priority], {'bold': priority == 'critical', 'color': p_color}),
                        (_NOTIF_CATEGORY_LABEL.get(n0.category, n0.category), {'color': C.text_soft}),
                        (g['theme'], {'bold': priority == 'critical', 'color': p_color}),
                        (domains_str, {'size': 9, 'color': C.text_soft}),
                        (len(g['domains']), {'size': 10, 'bold': True, 'color': C.text_soft}),
                        (_dept_notif(n0), {'size': 9, 'color': C.text_soft}),
                    ], 2):
                        cell = ws.cell(row=row, column=ci)
                        cell.value = val
                        cell.font = _font(**kw)
                        cell.alignment = _align(
                            wrap=True, vertical='top',
                            horizontal='center' if ci == 7 else 'general')
                        cell.border = _border(color=C.border_light)
                        if priority == 'critical' and ci in (5, 6, 7):
                            cell.fill = _fill(p_bg)

                    row += 1

                row += 1  # пробел между приоритетами

        elif source_key == 'twogis':
            # ── 2ГИС: одна строка на отзыв (без группировки), колонка
            # «Оценка» (★ + качество), превью = только ссылка «Читать». ──
            for ci, h in enumerate(['Дата', 'Оценка', '', 'Тема', '', 'Ссылка', 'Отдел'], 2):
                cell = ws.cell(row=row, column=ci)
                cell.value = h
                cell.font = _font(size=9, bold=True, color=C.text_muted)
                cell.fill = _fill(C.surface)
                cell.alignment = _align()
                cell.border = _border()
            ws.row_dimensions[row].height = 20
            row += 1

            for n in sorted(items, key=lambda x: x.date or '', reverse=True):
                ws.row_dimensions[row].height = 30
                rating_txt, rating_color = _review_rating_cell(getattr(n, 'rating', None))
                review_url = getattr(n, 'review_url', None)

                for ci, (val, kw) in enumerate([
                    (n.date, {'color': C.text_soft}),
                    (rating_txt, {'bold': True, 'color': rating_color}),
                    ('', {}),
                    (n.subject, {'color': C.text}),
                    ('', {}),
                    ('', {}),   # ссылка проставляется ниже
                    (_dept_notif(n), {'size': 9, 'color': C.text_soft}),
                ], 2):
                    cell = ws.cell(row=row, column=ci)
                    cell.value = val
                    cell.font = _font(**kw)
                    cell.alignment = _align(wrap=True, vertical='top')
                    cell.border = _border(color=C.border_light)

                # Колонка «Ссылка» (G = 7): кликабельная «Читать полностью»
                link_cell = ws.cell(row=row, column=7)
                if review_url:
                    link_cell.value = 'Читать полностью'
                    link_cell.hyperlink = review_url
                    link_cell.font = _font(size=9, color=C.accent, underline='single')
                else:
                    link_cell.value = '-'
                    link_cell.font = _font(size=9, color=C.text_muted)
                link_cell.alignment = _align(vertical='top')
                link_cell.border = _border(color=C.border_light)

                row += 1

        else:
            # ── Источник без классификации: плоский список ──
            # Шапка
            for ci, h in enumerate(['Дата', '', '', 'Тема', 'Домены', 'Превью', 'Отдел'], 2):
                cell = ws.cell(row=row, column=ci)
                cell.value = h
                cell.font = _font(size=9, bold=True, color=C.text_muted)
                cell.fill = _fill(C.surface)
                cell.alignment = _align()
                cell.border = _border()
            ws.row_dimensions[row].height = 20
            row += 1

            groups = _group_notifs_by_theme(items)
            for g in sorted(groups, key=lambda x: x['date'] or '', reverse=True):
                n0 = g['first']
                domains_str = ', '.join(g['domains'])
                theme = g['theme']
                if g['count'] > 1:
                    theme = f'{theme}  ×{g["count"]}'
                ws.row_dimensions[row].height = _notif_row_height(
                    domains_str, n0.body_preview)

                for ci, (val, kw) in enumerate([
                    (g['date'], {'color': C.text_soft}),
                    ('', {}),
                    ('', {}),
                    (theme, {'bold': False, 'color': C.text}),
                    (domains_str, {'size': 9, 'color': C.text_soft}),
                    ((n0.body_preview or '')[:400], {'size': 9, 'color': C.text_soft}),
                    (_dept_notif(n0), {'size': 9, 'color': C.text_soft}),
                ], 2):
                    cell = ws.cell(row=row, column=ci)
                    cell.value = val
                    cell.font = _font(**kw)
                    cell.alignment = _align(wrap=True, vertical='top')
                    cell.border = _border(color=C.border_light)

                row += 1

        row += 2  # пробел между секциями


# ── Лист «Ошибки сервисов» (Вебмастер/GSC/Метрика - из API) ─────────
# Ошибки прямо из сервисов (не из почты) - находки, детали для «Плана
# работ» (Task.where = «Лист «Ошибки сервисов»»).

_SVC_SECTION = [
    ('webmaster', 'Яндекс.Вебмастер'),
    ('gsc',       'Google Search Console'),
    ('metrika',   'Яндекс.Метрика'),
]
_SVC_SEV_LABEL = {
    'fatal': '🔴 Фатальная', 'critical': '🔴 Критическая',
    'possible': '🟠 Возможная', 'recommendation': '🟡 Рекомендация',
    'info': '⚪ Инфо',
}
_SVC_SEV_COLOR = {
    'fatal': C.err, 'critical': C.err, 'possible': C.warn,
    'recommendation': 'CA8A04', 'info': C.text_muted,
}
_SVC_SEV_ORDER = {'fatal': 0, 'critical': 1, 'possible': 2,
                  'recommendation': 3, 'info': 4}


def _build_service_issues_sheet(wb, service_issues):
    """Лист «Ошибки сервисов» - проблемы сайтов прямо из сервисов (не из почты).
    Добавляется только если есть данные."""
    issues = service_issues or []
    if not issues:
        return

    ws = wb.create_sheet('Ошибки сервисов')
    ws.sheet_view.showGridLines = False
    has_crit = any(getattr(i, 'severity', '') in ('fatal', 'critical') for i in issues)
    ws.sheet_properties.tabColor = C.err if has_crit else C.accent

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 28   # Сайт
    ws.column_dimensions['C'].width = 18   # Серьёзность
    ws.column_dimensions['D'].width = 50   # Проблема
    ws.column_dimensions['E'].width = 13   # Дата
    ws.column_dimensions['F'].width = 10   # Открыть

    ws.merge_cells('B2:F2')
    c = ws['B2']
    c.value = 'Ошибки сайтов из сервисов'
    c.font = _font(size=16, bold=True)
    ws.row_dimensions[2].height = 26

    ws.merge_cells('B3:F3')
    c = ws['B3']
    c.value = ('Проблемы напрямую из Яндекс.Вебмастера / GSC / Метрики (диагностика: '
               'сайтмапы, дубли, мусорные ссылки, ошибки сервера и индексации). '
               'Не из почты - из самих сервисов по API.')
    c.font = _font(size=10, italic=True, color=C.text_soft)
    c.alignment = _align(wrap=True, vertical='top')
    ws.row_dimensions[3].height = 30

    from collections import defaultdict
    by_service = defaultdict(list)
    for i in issues:
        by_service[getattr(i, 'service', 'webmaster')].append(i)

    row = 5
    for svc_key, svc_title in _SVC_SECTION:
        svc_items = by_service.get(svc_key, [])
        if not svc_items:
            continue

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        sc = ws.cell(row=row, column=2)
        sc.value = f'{svc_title}  ({len(svc_items)})'
        sc.font = _font(size=13, bold=True, color=C.accent)
        sc.fill = _fill(C.accent_soft)
        sc.alignment = _align(indent=1)
        ws.row_dimensions[row].height = 24
        row += 1

        for ci, h in enumerate(['Сайт', 'Серьёзность', 'Проблема', 'Дата', 'Открыть'], 2):
            cell = ws.cell(row=row, column=ci)
            cell.value = h
            cell.font = _font(size=9, bold=True, color=C.text_muted)
            cell.fill = _fill(C.surface)
            cell.alignment = _align()
            cell.border = _border()
        ws.row_dimensions[row].height = 20
        row += 1

        for i in sorted(svc_items, key=lambda x: (_SVC_SEV_ORDER.get(
                getattr(x, 'severity', 'info'), 9), getattr(x, 'host', ''))):
            sev = getattr(i, 'severity', 'info')
            sev_color = _SVC_SEV_COLOR.get(sev, C.text_muted)
            ws.row_dimensions[row].height = 30

            for ci, (val, kw) in enumerate([
                (getattr(i, 'host', ''), {'size': 10, 'color': C.text}),
                (_SVC_SEV_LABEL.get(sev, sev),
                 {'size': 9, 'bold': sev in ('fatal', 'critical'), 'color': sev_color}),
                (getattr(i, 'title', '') or getattr(i, 'code', ''),
                 {'size': 10, 'color': C.text_soft}),
                (getattr(i, 'date', ''), {'size': 9, 'color': C.text_muted}),
            ], 2):
                cell = ws.cell(row=row, column=ci)
                cell.value = val
                cell.font = _font(**kw)
                cell.alignment = _align(wrap=True, vertical='top')
                cell.border = _border(color=C.border_light)

            # «Открыть» - ссылка в панель сервиса
            link_cell = ws.cell(row=row, column=6)
            _u = _wm_alive_url(getattr(i, 'url', ''))
            if _u:
                link_cell.value = 'открыть'
                link_cell.hyperlink = _u
                link_cell.font = _font(size=9, color=C.accent, underline='single')
            else:
                link_cell.value = '-'
                link_cell.font = _font(size=9, color=C.text_muted)
            link_cell.alignment = _align(horizontal='center')
            link_cell.border = _border(color=C.border_light)

            row += 1

        row += 2


# ── Лист «Индексация» (п.1.7: robots.txt / noindex / canonical) ─────


def _idx_signals_text(ix):
    """Краткая сводка сигналов индексации страницы для колонки «Сигналы»."""
    parts = []
    if ix.get('robots_disallowed'):
        parts.append(f'robots.txt: Disallow {ix.get("robots_rule")}')
    if ix.get('meta_noindex'):
        parts.append(f'meta: {ix.get("meta_robots")}')
    if ix.get('x_robots_noindex'):
        parts.append(f'X-Robots-Tag: {ix.get("x_robots")}')
    if ix.get('canonical_disallowed'):
        parts.append(f'canonical → закрытый URL: {ix.get("canonical")}')
    elif ix.get('canonical_self') is False:
        parts.append(f'canonical → {ix.get("canonical")}')
    return '; '.join(parts)


# Группировка «одна проблема - одна строка + список URL» (_issue_groups /
# _render_issue_groups) удалена вместе с последними листами, которые её
# использовали: теперь каждая находка - отдельная строка «Проблем», где
# группировку делает автофильтр, а не отчёт.


# ── Лист «Валидация и скорость» (п.1.16: W3C HTML/CSS + время ресурсов) ─


# Лист «Страницы в ГСК» удалён - метрика (индексировано/просканировано
# + дельта) теперь секцией на «Трафик и траст».


# ── Лист «Страница 404» (п.1.18) ────────────────────────────────────


def _wm_alive_url(url, section='optimization/checklist/'):
    """Живая ссылка в панель Вебмастера. Старые кеши хранят мёртвые пути
    (/diagnostics/ и /links/external/ отдают 404 - панель переехала) -
    подменяем хвост на актуальный раздел."""
    if not url:
        return None
    for dead in ('diagnostics/', 'links/external/'):
        if url.endswith(dead):
            return url[:-len(dead)] + section
    return url

# Лист «404 в индексе» удалён - те же находки (dead/errors) уже полностью
# в «Проблемы» через report_priorities._index_404_findings(), отдельная
# таблица дублировала одно и то же.


# ── Лист «Фильтры ПС» (п.1.19: санкции поисковых систем) ───────────


# ── Лист «Нагрузка и парсинг» (ошибки сервера: парсинг/нагрузка/дубли) ──


# ── Ссылочный профиль (lite-проверка беклинков, Вебмастер) - секция на
# листе «Трафик и траст» ────────────────────────────────────────────


def _lp_rank(h):
    """Сортировка хостов: сначала самые проблемные. Группы: 0 - обвал массы,
    1 - спам-доноры/всплеск, 2 - прочие предупреждения, 3 - профиля нет,
    4 - норма. Внутри группы - по глубине просадки, затем по числу спама."""
    hist = h.get('history') or {}
    if hist.get('dropped'):
        grp = 0
    elif h.get('spam_count') or hist.get('spiked'):
        grp = 1
    elif h.get('warnings'):
        grp = 2
    elif h.get('infos'):
        grp = 3
    else:
        grp = 4
    return (grp, -(hist.get('drop_pct') or 0),
            -(h.get('recent_spam_count') or 0),
            -(h.get('spam_count') or 0), h.get('host') or '')


# Лист «Ссылочный профиль» удалён - таблица по хостам теперь секцией на
# «Трафик и траст» (компактнее: без колонки «Панель», без ручной сверки
# Google - это справочная строка, не критично для сводки).


# ── Секция «Аномалии» (низ листа «Аналитика») ─────────────────────
# Сводит в одном месте резкие отклонения: аномалии Вебмастера (обход,
# проблемы, страницы/ИКС - Блок B) + внезапные мусорные доноры и скачки
# ссылочной массы (Блок A, детали - на листе «Трафик и траст»).

_ANOM_SEV = {'fatal': (0, '🔴 фатально'), 'critical': (1, '🔴 критично'),
             'possible': (2, '⚠ возможно'), 'info': (3, 'инфо')}


def _fmt_ba(before, after):
    """«было → сейчас» для колонки динамики."""
    b = '–' if before is None else str(before)
    a = '–' if after is None else str(after)
    return f'{b} → {a}' if before is not None else a


def _collect_anomaly_rows(wm_metrics, link_profile):
    """Плоский список аномалий из Вебмастера (wm_metrics) и ссылочного
    профиля (link_profile): [{host, metric, before, after, delta_pct,
    severity, text}]."""
    rows = []
    for h in (wm_metrics or {}).get('hosts') or []:
        for a in h.get('anomalies') or []:
            rows.append({**a, 'host': h.get('host', ''),
                         'panel_url': h.get('panel_url')})
    # Ссылочный профиль → аномалии (детали на листе «Трафик и траст»).
    for h in (link_profile or {}).get('hosts') or []:
        host = h.get('host', '')
        purl = h.get('panel_url')
        if h.get('recent_spam_count'):
            rows.append({
                'host': host, 'metric': 'Внезапные мусорные доноры',
                'before': None, 'after': h['recent_spam_count'], 'delta_pct': None,
                'severity': 'critical', 'panel_url': purl,
                'text': f'{h["recent_spam_count"]} новых спам-доноров за ~30 дн. '
                        f'- негативное SEO? (детали - лист «Трафик и траст»)'})
        # Обвал ссылочной массы - это про ПОТЕРЮ доноров, не про мусор; ему
        # место на листе «Трафик и траст», в аномалии не тащим (иначе
        # десятки строк на каждом прогоне). Всплеск (возможный спам) - тащим.
        hist = h.get('history') or {}
        if hist.get('spiked'):
            rows.append({
                'host': host, 'metric': 'Рост ссылок',
                'before': hist.get('first'), 'after': hist.get('latest'),
                'delta_pct': None, 'severity': 'possible', 'panel_url': purl,
                'text': f'резкий рост ×{hist.get("spike_factor")} - проверить на спам/накрутку'})
    rows.sort(key=lambda r: (_ANOM_SEV.get(r.get('severity'), (9,))[0],
                             r.get('host', '')))
    return rows


# Лист «Аномалии» удалён: часть A (Вебмастер/ссылочный профиль) и часть B
# (ГСК-запросы/Метрика-рефералы) обе перенесены на «Хосты и аномалии»
# (_build_hosts_anomalies_sheet) - «только с проблемами», без «всё чисто».


# ── Лист «Настройки в админке» (доп. чек-лист: функции настройки) ──


# ── Лист «Я.Бизнес/GMB» ─────────────────────────────────────────────


def _build_yabusiness_sheet(wb, yabusiness, review_priority=None):
    """Лист «Я.Бизнес/GMB»: каждый поддомен зарегистрирован под свой регион
    (Яндекс.Бизнес). Внизу - единая проверка отзывов (приоритет докупки,
    Яндекс+2ГИС). Данные из кабинета Справочника на сессии. Добавляется,
    только если проверка выполнялась."""
    if not yabusiness:
        return
    missing = yabusiness.get('missing') or []
    matched = yabusiness.get('matched') or []
    orphans = yabusiness.get('orphan_orgs') or []
    has_problem = bool(missing) or not yabusiness.get('available')

    ws = wb.create_sheet('Я.Бизнес и GMB')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = (C.err if missing else C.ok
                                    if yabusiness.get('available') else C.warn)
    for col, w in (('A', 3), ('B', 26), ('C', 30), ('D', 42),
                   ('E', 12), ('F', 12), ('G', 13), ('H', 11), ('I', 3)):
        ws.column_dimensions[col].width = w

    ws.merge_cells('B2:D2')
    c = ws['B2']
    c.value = 'Я.Бизнес / GMB'
    c.font = _font(size=16, bold=True)
    ws.row_dimensions[2].height = 26

    ws.merge_cells('B3:D3')
    c = ws['B3']
    c.value = ('Каждый поддомен (город) должен быть зарегистрирован в '
               'Яндекс.Бизнесе под своим регионом. Берём организации '
               'аккаунта из кабинета Справочника (город/регион карточки) и '
               'сверяем с городами поддоменов. «Сети» без единого города '
               'пропускаем (это группы). Данные - на сессии Яндекса (как '
               'автокликеры); при партнёрском доступе перейдём на API.')
    c.font = _font(size=10, italic=True, color=C.text_soft)
    c.alignment = _align(wrap=True, vertical='top')
    ws.row_dimensions[3].height = 56

    row = 5
    if not yabusiness.get('available'):
        ws.merge_cells(f'B{row}:D{row}')
        cc = ws[f'B{row}']
        cc.value = f'⚪ {yabusiness.get("note", "Проверка не выполнена.")}'
        cc.font = _font(size=10, color=C.text_muted)
        cc.alignment = _align(indent=1, wrap=True)
        return

    n_sub = yabusiness.get('total_subdomains', 0)
    ws.merge_cells(f'B{row}:D{row}')
    cc = ws[f'B{row}']
    cc.value = (f'Поддоменов: {n_sub}  ·  с орг под свой город: '
                f'{len(matched)}  ·  БЕЗ орг: {len(missing)}  ·  активных '
                f'карточек в аккаунте: {yabusiness.get("active_orgs", 0)} '
                f'(сетей/пустых: {yabusiness.get("chains_or_empty", 0)})')
    cc.font = _font(size=11, bold=True, color=C.err if missing else C.ok)
    cc.fill = _fill(C.surface)
    cc.alignment = _align(indent=1, wrap=True)
    ws.row_dimensions[row].height = 22
    row += 2

    def _hdr(text):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        h = ws.cell(row=row, column=2, value=text)
        h.font = _font(size=11, bold=True, color=C.text)
        h.alignment = _align(indent=1)
        ws.row_dimensions[row].height = 20
        row += 1

    # Поддомены без орг - главная находка.
    if missing:
        _hdr(f'❌ Поддомены без организации под их город ({len(missing)})')
        for m in missing:
            ws.cell(row=row, column=2, value=m.get('city') or '').font = _font(
                size=10, color=C.err)
            uc = ws.cell(row=row, column=3, value=m.get('url') or '')
            uc.font = _font(size=9, color=C.accent, underline='single')
            if m.get('url'):
                uc.hyperlink = m['url']
            ws.cell(row=row, column=4,
                    value='нет карточки в Я.Бизнесе под этот город').font = \
                _font(size=9, color=C.text_soft)
            ws.row_dimensions[row].height = 15
            row += 1
        row += 1

    # Поддомены с орг.
    if matched:
        _hdr(f'✅ Поддомены с организацией ({len(matched)})')
        for m in matched:
            o = m.get('org') or {}
            ws.cell(row=row, column=2, value=m.get('city') or '').font = _font(
                size=10, color=C.ok)
            ws.cell(row=row, column=3,
                    value=f'орг {o.get("permalink","")} · регион '
                    f'{o.get("region","")}').font = _font(
                size=9, color=C.text_soft)
            ws.cell(row=row, column=4, value=o.get('addr') or '').font = _font(
                size=9, color=C.text_muted)
            ws.row_dimensions[row].height = 15
            row += 1
        row += 1

    # Организации без поддомена (лишние/чужие города).
    if orphans:
        _hdr(f'⚠ Организации без поддомена ({len(orphans)})')
        for o in orphans:
            ws.cell(row=row, column=2, value=o.get('city') or '').font = _font(
                size=10, color=C.warn)
            ws.cell(row=row, column=3,
                    value=f'орг {o.get("permalink","")} · регион '
                    f'{o.get("region","")}').font = _font(size=9, color=C.text_soft)
            ws.cell(row=row, column=4, value=o.get('addr') or '').font = _font(
                size=9, color=C.text_muted)
            ws.row_dimensions[row].height = 15
            row += 1
        row += 1

    # ── Пункт: все филиалы объединены в Сеть ──
    cch = yabusiness.get('chain_check') or {}
    if cch:
        row += 1
        united = cch.get('united')
        _hdr(('✅ ' if united else '❌ ') + 'Все филиалы объединены в Сеть')
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        cc = ws.cell(row=row, column=2)
        if united:
            cc.value = (f'все филиалы объединены в сети (в сетях '
                        f'{cch.get("chain_members", 0)} филиалов, отдельных '
                        f'компаний нет)')
            cc.font = _font(size=10, color=C.ok)
        else:
            cc.value = (f'НЕ объединены: {cch.get("standalone_companies", 0)} '
                        f'отдельных компаний (карточек) вне сети; в сетях '
                        f'{cch.get("chain_members", 0)} филиалов, сетей '
                        f'{cch.get("chains", 0)} - отдельные свести в Сеть')
            cc.font = _font(size=10, color=C.err)
        cc.alignment = _align(indent=1, wrap=True)
        ws.row_dimensions[row].height = 30
        row += 2

    # ── Пункт: максимально заполнен профиль ──
    pch = yabusiness.get('profile_check') or {}
    porgs = pch.get('orgs') or []
    if porgs:
        row += 1
        _hdr(('✅ ' if pch.get('all_full') else '⚠ ')
             + f'Заполненность профиля организаций ({len(porgs)})')
        for o in porgs:
            miss = o.get('missing') or []
            ws.cell(row=row, column=2, value=o.get('city') or '').font = _font(
                size=10, color=C.ok if not miss else C.warn)
            ws.cell(row=row, column=3,
                    value=f'заполнено {o.get("filled",0)}/{o.get("total",0)}'
                    ).font = _font(size=9, color=C.text_soft)
            ws.cell(row=row, column=4,
                    value=('всё заполнено' if not miss
                           else 'не заполнено: ' + ', '.join(miss))).font = \
                _font(size=9, color=C.text_muted if not miss else C.warn)
            ws.row_dimensions[row].height = 15
            row += 1

    # ── Отзывы: приоритет докупки (единая проверка отзывов, Яндекс + 2ГИС) ──
    rp = review_priority
    if rp:
        row += 1
        if not rp.get('available'):
            _hdr('Отзывы: приоритет докупки')
            ws.merge_cells(start_row=row, start_column=2, end_row=row,
                           end_column=8)
            cc = ws.cell(row=row, column=2,
                         value='⚪ ' + (rp.get('note') or 'не выполнялось'))
            cc.font = _font(size=10, color=C.text_muted)
            cc.alignment = _align(indent=1, wrap=True)
            row += 1
        else:
            low = rp.get('low_rating_count', 0)
            tot = rp.get('total_branches', 0)
            _hdr(('❌ ' if low else '✅ ')
                 + f'Отзывы: приоритет докупки - с рейтингом < 4.7: {low} '
                 + f'из {tot}')
            ws.merge_cells(start_row=row, start_column=2, end_row=row,
                           end_column=8)
            nt = ws.cell(row=row, column=2, value=(
                'Рейтинг филиала = худший из Яндекс/2ГИС. Докупаем по 2 отзыва '
                '(3 при низком рейтинге). Порядок: рейтинг < 4.7 выше, затем '
                'города от миллионников к меньшим.'))
            nt.font = _font(size=9, italic=True, color=C.text_soft)
            nt.alignment = _align(indent=1, wrap=True)
            ws.row_dimensions[row].height = 26
            row += 1
            cols = [('Город', 2), ('Рейтинг Я', 3), ('Отз. Я', 4),
                    ('Рейтинг 2ГИС', 5), ('Отз. 2ГИС', 6),
                    ('Рейтинг (мин)', 7), ('Докупить', 8)]
            for name, ci in cols:
                h = ws.cell(row=row, column=ci, value=name)
                h.font = _font(size=9, bold=True, color=C.text)
                h.fill = _fill(C.surface)
                h.alignment = _align(indent=1)
                h.border = _border()
            ws.row_dimensions[row].height = 16
            row += 1

            def _rv(v):
                return '' if v is None else v

            for b in (rp.get('branches') or []):
                y = b.get('yandex') or {}
                g = b.get('twogis') or {}
                low_b = b.get('low_rating')
                vals = [b.get('city') or '', _rv(y.get('rating')),
                        _rv(y.get('count')), _rv(g.get('rating')),
                        _rv(g.get('count')), _rv(b.get('rating')),
                        b.get('order')]
                for j, (_name, ci) in enumerate(cols):
                    cell = ws.cell(row=row, column=ci, value=vals[j])
                    clr = C.err if (ci == 7 and low_b) else C.text
                    cell.font = _font(size=9, color=clr,
                                      bold=(ci == 7 and low_b))
                    cell.alignment = _align(indent=1)
                    cell.border = _border()
                ws.row_dimensions[row].height = 15
                row += 1


_TRAFFIC_COLS = [
    ('Год', 8), ('Срез', 11),
    ('Итого по каналам', 15), ('Прямые заходы', 13), ('Яндекс', 10),
    ('Google', 10), ('Лиды', 8), ('Конверсия, %', 12), ('Отказы, %', 10),
    ('Глубина', 9), ('Время на сайте', 13),
    ('Главная', 10), ('Категория', 11), ('Услуга', 9), ('Товар', 9),
    ('Фильтр', 9), ('Тег', 8), ('Информационная', 14), ('Техническая', 12),
]


_MONTHS_NOM = ['', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
               'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']


def _fmt_duration(sec):
    """Секунды → «м:сс» (0 → «0:00»)."""
    sec = int(sec or 0)
    return f'{sec // 60}:{sec % 60:02d}'


def _traffic_srez(period, r):
    y, m, _d = r['d1'].split('-')
    if period == 'Месяц':
        return f'{_MONTHS_NOM[int(m)]} {y}'
    if period == 'Год':
        return y
    return '.'.join(reversed(r['d1'].split('-')))


def _traffic_nums(r):
    return [r.get('visits', 0), r.get('direct', 0), r.get('yandex', 0),
           r.get('google', 0), r.get('leads', 0), r.get('conv', 0),
           r.get('bounce', 0)]


# Индексы (0-based среди _traffic_nums) с инверсией: рост - плохо, красный.
# Прямые заходы (1) - непропорциональный рост обычно значит потерю UTM/
# referrer-разметки, а не реальный органический рост; Отказы (6) - как обычно.
_TRAFFIC_INVERT_IDX = {1, 6}


def _build_traffic_overview_sheet(wb, traffic, trust=None, link_profile=None,
                                  gsc_pages=None):
    """Лист «Трафик и траст»: компактная сводка трафика по странам/периодам
    (визиты, каналы, лиды, конверсия, отказы). Плюс
    траст проекта (ИКС/DR), ссылочный профиль (lite, Вебмастер) и
    страницы в ГСК (индексировано/просканировано + Δ) - все метрики, не
    находки, поэтому не в «Проблемы», а здесь, рядом с трафиком (раньше
    жили отдельными листами). Каждый блок трафика (текущий/прошлый/Δ)
    отделён жирной рамкой сверху и снизу, чтобы блоки не сливались."""
    groups = (traffic or {}).get('groups')
    if not groups:
        rows = (traffic or {}).get('rows') or []
        if rows:
            groups = [{'country': 'Все домены', 'counters': traffic.get('counters', 0),
                      'rows': rows}]
    has_traffic = bool(groups and any(g.get('rows') for g in groups))
    if not has_traffic and not trust and not link_profile and not gsc_pages:
        return

    ws = wb.create_sheet('Трафик и траст')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C.accent

    # 4 независимые зоны колонок, каждая - свой курсор строк: Трафик (B:K),
    # Траст (M:O), Ссылочный профиль (Q:V), ГСК (X:Z), между зонами узкие
    # колонки-разделители (L/P/W). Раньше Траст и Ссылочный профиль стояли
    # ОДИН ПОД ДРУГИМ в одних и тех же колонках (у обоих «Хост» начинался
    # с одной точки) - узкие числовые колонки траста (ИКС/DR) навязывали
    # свою ширину колонкам ссылочного профиля (Ссылок/Доноров) и наоборот,
    # то же самое, что раньше было с трафиком. Теперь у каждой таблицы
    # СВОИ колонки - ширина одной никак не зависит от другой.
    widths = [16, 9, 13, 10, 13, 10, 10, 8, 12, 10]
    ws.column_dimensions['A'].width = 3
    for i, w in enumerate(widths, 2):
        ws.column_dimensions[get_column_letter(i)].width = w
    for col in ('L', 'P', 'W'):
        ws.column_dimensions[col].width = 3
    for col, w in zip('MNO', (18, 11, 11)):              # Траст
        ws.column_dimensions[col].width = w
    for col, w in zip('QRSTUV', (16, 9, 9, 13, 14, 60)):  # Ссылочный профиль
        ws.column_dimensions[col].width = w
    for col, w in zip('XYZ', (46, 12, 14)):               # ГСК (длинные подписи)
        ws.column_dimensions[col].width = w
    TRUST_COL, LP_COL, GSC_COL = 11, 15, 22  # смещения от B(2): M/Q/X
    # Шапка таблицы трафика (при наличии трафика) стоит на row=5, данные - с
    # row=6 (title@2, подзаголовок@3, зазор@4). Боковые блоки (Траст/Ссылочный
    # профиль/ГСК) равняют свою шапку и данные на те же номера строк, чтобы
    # по горизонтали не оказались рядом «шапка» одного блока и «данные»
    # другого.
    _SIDE_HDR_ROW, _SIDE_DATA_ROW = 5, 6

    ws.merge_cells('B2:Z2')
    c = ws['B2']
    c.value = 'Трафик и траст проекта'
    c.font = _font(size=16, bold=True)
    ws.row_dimensions[2].height = 26

    row = 3
    if has_traffic:
        cols = ['Страна', 'Период', 'Срез', 'Визиты', 'Прямые заходы', 'Яндекс',
               'Google', 'Лиды', 'Конверсия, %', 'Отказы, %']
        ws.merge_cells(f'B{row}:K{row}')
        c = ws.cell(row=row, column=2)
        c.value = ('Текущий период и прошлый рядом, Δ - разница в %. Рост «Прямых '
                  'заходов» или «Отказов» - красный: обычно значит потерю UTM/'
                  'referrer-разметки или ухудшение, а не хороший знак.')
        c.font = _font(size=10, italic=True, color=C.text_soft)
        c.alignment = _align(wrap=True)
        ws.row_dimensions[row].height = 28
        row += 2

        hdr_row = row
        for i, title in enumerate(cols, 2):
            cell = ws.cell(row=hdr_row, column=i)
            cell.value = title
            cell.font = _font(size=9, bold=True, color=C.bg_elev)
            cell.fill = _fill(C.header_navy)
            cell.border = _border()
            cell.alignment = _align(horizontal='center', wrap=True)
        ws.row_dimensions[hdr_row].height = 24

        THICK = Side(style='medium', color='FF404040')
        THIN = Side(style='thin', color=f'FF{C.border_light}')

        def _row_border(top, bottom):
            return Border(top=THICK if top else THIN, bottom=THICK if bottom else THIN,
                          left=THIN, right=THIN)

        def _delta_txt(cur, prev, invert):
            if not prev:
                return '–', C.text_muted, None
            pct = round((cur - prev) / prev * 100, 1)
            if pct == 0:
                return '0%', C.text_muted, None
            up_bad = invert
            color = (C.err if up_bad else C.ok) if pct > 0 else (C.ok if up_bad else C.err)
            bg = C.err_soft if color == C.err else C.ok_soft
            sign = '+' if pct > 0 else ''
            return f'{sign}{pct}%', color, bg

        order = ['День', 'Месяц', 'Год']
        row = hdr_row + 1
        for g in groups:
            grows = g.get('rows') or []
            if not grows:
                continue
            seen_periods = sorted({r['period'] for r in grows},
                                  key=lambda p: order.index(p) if p in order else 9)
            for period in seen_periods:
                prs = [r for r in grows if r['period'] == period]
                cur = next((r for r in prs if r['kind'] == 'текущий'), None)
                prev = next((r for r in prs if r['kind'] == 'прошлый'), None)
                block = [r for r in (cur, prev) if r]
                for i_row, r in enumerate(block):
                    is_cur = r is cur
                    vals = [g.get('country', ''), period, _traffic_srez(period, r)] \
                        + _traffic_nums(r)
                    for ci, v in enumerate(vals, 2):
                        cell = ws.cell(row=row, column=ci, value=v)
                        cell.font = _font(size=10, bold=(ci == 2 and is_cur))
                        cell.alignment = _align(
                            horizontal='center' if ci > 4 else 'left', indent=1)
                        cell.border = _row_border(top=(i_row == 0), bottom=False)
                    ws.row_dimensions[row].height = 16
                    row += 1
                if cur and prev:
                    cn, pn = _traffic_nums(cur), _traffic_nums(prev)
                    dvals = [g.get('country', ''), period, 'Δ, %']
                    for ci_v, (c_val, p_val) in enumerate(zip(cn, pn)):
                        txt, color, bg = _delta_txt(
                            c_val, p_val, invert=ci_v in _TRAFFIC_INVERT_IDX)
                        dvals.append((txt, color, bg))
                    for ci, v in enumerate(dvals, 2):
                        cell = ws.cell(row=row, column=ci)
                        if isinstance(v, tuple):
                            txt, color, bg = v
                            cell.value = txt
                            cell.font = _font(size=10, bold=True, color=color)
                            if bg:
                                cell.fill = _fill(bg)
                        else:
                            cell.value = v
                            cell.font = _font(size=10, bold=True, color=C.text)
                        cell.alignment = _align(
                            horizontal='center' if ci > 4 else 'left', indent=1)
                        cell.border = _row_border(top=False, bottom=True)
                    ws.row_dimensions[row].height = 16
                    row += 1
                elif block:
                    # Единственная строка блока (нет пары текущий/прошлый) - и
                    # верхняя, и нижняя рамка жирные (блок из одной строки).
                    for ci in range(2, 12):
                        ws.cell(row=row - 1, column=ci).border = _row_border(True, True)

        last = row - 1
        ws.freeze_panes = f'B{hdr_row + 1}'
        if last >= hdr_row + 1:
            ws.auto_filter.ref = f'B{hdr_row}:K{last}'
        row += 1

    # ── Траст проекта (ИКС + DR) - метрика, не находка. Своя зона колонок
    # (TRUST_COL), свой курсор строк - высота этого блока (число хостов)
    # никак не связана с Ссылочным профилем или ГСК. Заголовок и данные
    # выровнены на те же номера строк, что у таблицы трафика (row=3 -
    # заголовок блока, row=5 - шапка таблицы, row=6 - первая строка данных),
    # чтобы по горизонтали не съезжались «шапка одного блока» и «данные
    # другого» на одной строке.
    if trust and trust.get('available') and trust.get('hosts'):
        trust_row = 3
        ws.merge_cells(start_row=trust_row, start_column=2 + TRUST_COL,
                       end_row=trust_row, end_column=4 + TRUST_COL)
        c = ws.cell(row=trust_row, column=2 + TRUST_COL, value='Траст проекта (ИКС + DR)')
        c.font = _font(size=13, bold=True, color=C.text)
        c.fill = _fill(C.accent_soft)
        c.alignment = _align(indent=1)
        ws.row_dimensions[trust_row].height = 24
        trust_row = _SIDE_HDR_ROW
        for col, title in ((2, 'Хост'), (3, 'ИКС (Яндекс)'), (4, 'DR (Open PageRank)')):
            h = ws.cell(row=trust_row, column=col + TRUST_COL, value=title)
            h.font = _font(size=9, bold=True, color=C.bg_elev)
            h.fill = _fill(C.header_navy)
            h.border = _border()
            h.alignment = _align(indent=1)
        ws.row_dimensions[trust_row].height = 20
        trust_row += 1
        for hh in trust['hosts']:
            sqi, dr = hh.get('sqi'), hh.get('dr')
            vals = [(2, hh.get('host', ''), C.text, False),
                   (3, '–' if sqi is None else sqi,
                    C.err if (sqi is not None and sqi < 10) else C.text,
                    sqi is not None and sqi < 10),
                   (4, '–' if dr is None else
                    (int(dr) if isinstance(dr, (int, float)) and float(dr).is_integer()
                     else dr), C.text, False)]
            for col, v, color, bold in vals:
                cell = ws.cell(row=trust_row, column=col + TRUST_COL, value=v)
                cell.font = _font(size=10, color=color, bold=bold)
                cell.border = _border(color=C.border_light)
                cell.alignment = _align(indent=1)
            ws.row_dimensions[trust_row].height = 16
            trust_row += 1

    # ── Ссылочный профиль (lite, Вебмастер) - метрика, не находка. Своя
    # зона колонок (LP_COL) - «Что не так» не зависит от ширины колонок
    # ни трафика, ни траста.
    lp_hosts = sorted((link_profile or {}).get('hosts') or [], key=_lp_rank) \
        if link_profile and link_profile.get('available') else []
    if lp_hosts:
        lp_row = 3
        ws.merge_cells(start_row=lp_row, start_column=2 + LP_COL,
                       end_row=lp_row, end_column=7 + LP_COL)
        c = ws.cell(row=lp_row, column=2 + LP_COL,
                    value='Ссылочный профиль (lite, Вебмастер)')
        c.font = _font(size=13, bold=True, color=C.text)
        c.fill = _fill(C.accent_soft)
        c.alignment = _align(indent=1)
        ws.row_dimensions[lp_row].height = 24
        lp_row = _SIDE_HDR_ROW
        _lp_headers = ('Хост', 'Ссылок', 'Доноров', 'Динамика', 'Статус', 'Что не так')
        for i, title in enumerate(_lp_headers, 2):
            h = ws.cell(row=lp_row, column=i + LP_COL, value=title)
            h.font = _font(size=9, bold=True, color=C.bg_elev)
            h.fill = _fill(C.header_navy)
            h.border = _border()
            h.alignment = _align(indent=1)
        ws.row_dimensions[lp_row].height = 20
        lp_row += 1
        _LP_STATUS = {0: ('❌ обвал', C.err), 1: ('⚠ спам/всплеск', C.err),
                      2: ('⚠ внимание', C.warn), 3: ('· нет профиля', C.text_muted),
                      4: ('✅ норма', C.ok)}
        for h in lp_hosts:
            hist = h.get('history') or {}
            grp = _lp_rank(h)[0]
            label, color = _LP_STATUS[grp]
            problems = list(h.get('warnings') or [])
            if h.get('spam_hosts'):
                problems.append('спам-доноры: ' + ', '.join(h['spam_hosts']))
            problems_text = '; '.join(problems) if problems else '–'
            dyn = (f'{hist.get("first")} → {hist.get("latest")}'
                  if hist.get('points') else '-')
            vals = [(2, h.get('host', ''), C.text), (3, h.get('total', 0), C.text_soft),
                   (4, h.get('distinct_hosts', 0), C.text_soft), (5, dyn, C.text_soft),
                   (6, label, color), (7, problems_text,
                                      color if problems else C.text_muted)]
            for col, v, color_ in vals:
                cell = ws.cell(row=lp_row, column=col + LP_COL, value=v)
                cell.font = _font(size=9, color=color_, bold=(col == 6 and grp <= 1))
                cell.border = _border(color=C.border_light)
                cell.alignment = _align(indent=1, wrap=(col == 7))
            ws.row_dimensions[lp_row].height = 16
            lp_row += 1

    # ── Страницы в ГСК (индексировано/просканировано) - метрика, не
    # находка. Своя зона колонок (GSC_COL) - «Показатель» тут длинные
    # текстовые подписи, им тоже не ужиться в чужих узких колонках.
    if gsc_pages and gsc_pages.get('available'):
        gsc_row = 3
        ws.merge_cells(start_row=gsc_row, start_column=2 + GSC_COL,
                       end_row=gsc_row, end_column=4 + GSC_COL)
        c = ws.cell(row=gsc_row, column=2 + GSC_COL,
                    value='Страницы в ГСК (Google Search Console)')
        c.font = _font(size=13, bold=True, color=C.text)
        c.fill = _fill(C.accent_soft)
        c.alignment = _align(indent=1)
        ws.row_dimensions[gsc_row].height = 24
        gsc_row = _SIDE_HDR_ROW
        for col, title in ((2, 'Показатель'), (3, 'Значение'), (4, 'Δ к прошлому')):
            h = ws.cell(row=gsc_row, column=col + GSC_COL, value=title)
            h.font = _font(size=9, bold=True, color=C.bg_elev)
            h.fill = _fill(C.header_navy)
            h.border = _border()
            h.alignment = _align(indent=1)
        ws.row_dimensions[gsc_row].height = 20
        gsc_row += 1
        deltas = gsc_pages.get('deltas') or {}
        for label, key in (('Проиндексировано', 'indexed'),
                           ('Просканировано, но пока не проиндексировано', 'crawled_not_indexed'),
                           ('Сумма', 'total')):
            val = gsc_pages.get(key)
            dv = deltas.get(key)
            if dv is None:
                dv_text, dv_color = '–', C.text_muted
            elif dv > 0:
                dv_text, dv_color = f'▲ +{dv:g}', C.ok
            elif dv < 0:
                dv_text, dv_color = f'▼ {dv:g}', C.err
            else:
                dv_text, dv_color = '= 0', C.text_muted
            for col, v, color, bold in ((2, label, C.text, False),
                                        (3, val if val is not None else '–', C.text, True),
                                        (4, dv_text, dv_color, True)):
                cell = ws.cell(row=gsc_row, column=col + GSC_COL, value=v)
                cell.font = _font(size=10, color=color, bold=bold)
                cell.border = _border(color=C.border_light)
                cell.alignment = _align(indent=1, horizontal='center' if col > 2 else 'left')
            ws.row_dimensions[gsc_row].height = 18
            gsc_row += 1
        note = 'Числа из отчёта GSC «Индексирование → Страницы».'
        if gsc_pages.get('manual'):
            note += ' Введены вручную.'
        ws.merge_cells(start_row=gsc_row, start_column=2 + GSC_COL,
                       end_row=gsc_row, end_column=4 + GSC_COL)
        c = ws.cell(row=gsc_row, column=2 + GSC_COL, value=note)
        c.font = _font(size=9, italic=True, color=C.text_muted)
        c.alignment = _align(indent=1)
        ws.row_dimensions[gsc_row].height = 16
        gsc_row += 1

    # ── Ни один блок не наполнился: лист остаётся с одним заголовком, и по
    # нему не понять, сломалось что-то или просто не настроено. Пишем прямо,
    # чего не хватает - как на листе «Я.Бизнес и GMB».
    _пусто = []
    if not has_traffic:
        _пусто.append(('Трафик',
                       (traffic or {}).get('note')
                       or 'нет данных Яндекс.Метрики. Нужны OAuth-токен Метрики '
                          'и номер счётчика в настройках проекта '
                          '(metrika_oauth, metrika_counter).'))
    if not (trust and trust.get('available') and trust.get('hosts')):
        _пусто.append(('Траст проекта',
                       (trust or {}).get('note')
                       or 'ИКС не получен. Нужен OAuth-токен Вебмастера '
                          '(webmaster_oauth) в настройках проекта; DR требует '
                          'ключ Open PageRank.'))
    if not lp_hosts:
        _пусто.append(('Ссылочный профиль',
                       (link_profile or {}).get('note')
                       or 'данные Вебмастера не получены. Нужен OAuth-токен '
                          'Вебмастера со scope webmaster:hostinfo.'))
    if not (gsc_pages and gsc_pages.get('available')):
        _пусто.append(('Страницы в ГСК',
                       (gsc_pages or {}).get('note')
                       or 'нет доступа к Google Search Console. Нужен '
                          'сервис-аккаунт GSC (gsc_service_account) с правами '
                          'на ресурс сайта.'))
    if len(_пусто) == 4:            # не наполнилось НИЧЕГО
        _r = 4
        for _имя, _почему in _пусто:
            ws.merge_cells(start_row=_r, start_column=2, end_row=_r, end_column=12)
            c = ws.cell(row=_r, column=2, value=f'⚪ {_имя}: {_почему}')
            c.font = _font(size=10, color=C.text_muted)
            c.alignment = _align(wrap=True, indent=1)
            ws.row_dimensions[_r].height = 30
            _r += 1


# Лист «Динамика трафика» удалён по прямому указанию - трафик по странам/
# периодам уже есть на «Трафик и траст»; разбивку по типам страниц (была
# только здесь) сознательно не переносим никуда.


def _ratio_word(spike, spiked):
    """Пояснение к коэффициенту ×: во сколько раз изменилось."""
    if not spike:
        return ''
    try:
        z = float(spike)
    except (TypeError, ValueError):
        return f'  ·  ×{spike}'
    if spiked:
        return f'  ·  ×{spike} – ВСПЛЕСК'
    if z <= 0.9:
        return f'  ·  ×{spike} – ниже, чем было'
    if z >= 1.1:
        return f'  ·  ×{spike} – чуть выше'
    return f'  ·  ×{spike} – без изменений'


# Лист «Траст проекта» удалён - ИКС/DR теперь секцией на «Трафик и траст».


def _build_admin_settings_sheet(wb, admin_settings):
    """Лист «Настройки в админке»: работают ли функции настройки поддоменов/
    категорий/товаров/тех.страниц (браузерная проверка + round-trip
    сохранения). Добавляется, только если проверка выполнялась."""
    if not admin_settings:
        return
    checks = admin_settings.get('checks') or []
    verdict = admin_settings.get('verdict') or 'ok'

    ws = wb.create_sheet('Настройки в админке')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = (C.err if verdict == 'fail'
                                    else C.warn if verdict == 'warn' else C.ok)
    for col, w in (('A', 3), ('B', 22), ('C', 78), ('D', 40), ('E', 3)):
        ws.column_dimensions[col].width = w

    ws.merge_cells('B2:D2')
    c = ws['B2']
    c.value = 'Настройки в админке'
    c.font = _font(size=16, bold=True)
    ws.row_dimensions[2].height = 26

    ws.merge_cells('B3:D3')
    c = ws['B3']
    c.value = ('Браузер заходит в админку Bitrix и проверяет, что функции '
               'настройки работают. Поддомены: создание (симуляция-dry-run), '
               'массовая загрузка, правка, удаление, скрытие. Категории: '
               'полный CRUD на временном скрытом разделе «[ТЕСТ ЧЕКЕРА]» '
               '(создание → правка → скрытие → удаление, удаляется в конце) + '
               'массовая загрузка. Товары (опционально): CRUD + сортировка + '
               'вывод в разные категории на временном скрытом товаре. '
               'Тех.страницы - редактор файлов. Ниже - аудит каждой операции '
               '«было → стало». Боевые данные не меняются: тест-раздел и '
               'тест-товар удаляются, поддомены реально не создаются.')
    c.font = _font(size=10, italic=True, color=C.text_soft)
    c.alignment = _align(wrap=True, vertical='top')
    ws.row_dimensions[3].height = 62

    row = 5
    if not admin_settings.get('available'):
        ws.merge_cells(f'B{row}:D{row}')
        c = ws[f'B{row}']
        c.value = f'⚪ {admin_settings.get("note", "Проверка не выполнена.")}'
        c.font = _font(size=10, color=C.text_muted)
        c.alignment = _align(indent=1, wrap=True)
        return

    ws.merge_cells(f'B{row}:D{row}')
    c = ws[f'B{row}']
    _dom = admin_settings.get('domain') or ''
    c.value = f'Админка: {_dom}'
    c.font = _font(size=10, bold=True, color=C.text)
    c.alignment = _align(indent=1)
    row += 2

    # Пояснение режимов операций (появляется, если есть хоть один аудит).
    _has_ops = any(ch.get('operations') for ch in checks)
    if _has_ops:
        ws.merge_cells(f'B{row}:D{row}')
        c = ws[f'B{row}']
        c.value = ('Режимы операций: «выполнено» - реально сделано и '
                   'откатано (запись в БД проверена); «симуляция» - dry-run '
                   'мастера, на сайте ничего не создаётся; «функция» - '
                   'проверено только наличие (реально не трогаем - боевые '
                   'данные).')
        c.font = _font(size=9, italic=True, color=C.text_muted)
        c.alignment = _align(wrap=True, indent=1)
        ws.row_dimensions[row].height = 30
        row += 2

    _MODE_LABEL = {'executed': 'выполнено', 'simulated': 'симуляция',
                   'ui': 'функция'}
    _RES = {'ok': ('✓', C.ok), 'fail': ('✗', C.err),
            'skip': ('–', C.text_muted)}

    for ch in checks:
        ws.row_dimensions[row].height = 18
        b = ws[f'B{row}']
        b.value = ('✅ ' if ch.get('ok') else '❌ ') + (ch.get('title') or '')
        b.font = _font(size=11, bold=True,
                       color=C.ok if ch.get('ok') else C.err)
        b.alignment = _align(indent=1, vertical='top')
        ws.merge_cells(f'C{row}:D{row}')
        d = ws[f'C{row}']
        d.value = ch.get('detail') or ''
        d.font = _font(size=10,
                       color=C.text_soft if ch.get('ok') else C.err)
        d.alignment = _align(wrap=True, vertical='top')
        row += 1
        for w in ch.get('warnings') or []:
            ws.merge_cells(f'C{row}:D{row}')
            wc = ws[f'C{row}']
            wc.value = f'⚠ {w}'
            wc.font = _font(size=9, color=C.warn)
            wc.alignment = _align(wrap=True, indent=1)
            ws.row_dimensions[row].height = 16
            row += 1

        # Таблица операций (аудит было→стало) - если есть.
        ops = ch.get('operations') or []
        if ops:
            # Шапка мини-таблицы
            for col, title in (('B', 'Операция'), ('C', 'Что менялось (было → стало)'),
                               ('D', 'Режим')):
                hc = ws[f'{col}{row}']
                hc.value = title
                hc.font = _font(size=9, bold=True, color=C.text_muted)
                hc.fill = _fill(C.surface)
                hc.alignment = _align(indent=1)
                hc.border = _border(color=C.border_light)
            ws.row_dimensions[row].height = 16
            row += 1
            for o in ops:
                mark, mcolor = _RES.get(o.get('result'), ('•', C.text_muted))
                oc = ws[f'B{row}']
                oc.value = f'{mark} {o.get("label", "")}'
                oc.font = _font(size=10, color=mcolor)
                oc.alignment = _align(indent=1, vertical='top', wrap=True)
                oc.border = _border(color=C.border_light)
                # было → стало (+ примечание)
                before, after = o.get('before', ''), o.get('after', '')
                if before and after:
                    txt = f'{before}  →  {after}'
                else:
                    txt = after or before or '-'
                if o.get('note'):
                    txt += f'\n({o["note"]})'
                cc = ws[f'C{row}']
                cc.value = txt
                cc.font = _font(size=9, color=C.text_soft)
                cc.alignment = _align(wrap=True, vertical='top', indent=1)
                cc.border = _border(color=C.border_light)
                mc = ws[f'D{row}']
                mc.value = _MODE_LABEL.get(o.get('mode'), o.get('mode', ''))
                mc.font = _font(size=9, color=C.text_muted)
                mc.alignment = _align(vertical='top', horizontal='center')
                mc.border = _border(color=C.border_light)
                ws.row_dimensions[row].height = 30 if o.get('note') else 18
                row += 1
        row += 1


# ── Лист «Ошибки JavaScript» (п.1.14: консоль браузера) ────────────


# ── Лист «Метаданные» (п.1.8: title/description/H1, дубли, URL) ─────

# ── Лист «Автокликер» ──────────────────────────────────────────────


# ── Лист «Фильтрация» (доп. чек-лист: фильтры товаров работают) ────

# Вердикт → (метка «работает/не работает» + причина, цвет, это баг?)
_FILTER_VERDICT = {
    'ok':            ('✅ работает',                                   'ok',  False),
    'empty':         ('❌ не работает – после фильтра пусто (ничего не найдено)', 'err', True),
    'not_narrowed':  ('❌ не работает – фильтр применился, но выдача не изменилась', 'err', True),
    'apply_failed':  ('⚠ не проверено – не удалось применить фильтр автоматически (проверить вручную)', 'warn', False),
    'http_error':    ('❌ не работает – ошибка загрузки страницы',     'err', True),
    'no_cards':      ('⚠ не проверено – карточки не распознаны (селектор card)', 'warn', False),
    'filter_absent': ('⚠ не проверено – фильтр не найден на странице (селектор filter)', 'warn', False),
    'config_error':  ('⚠ не проверено – ошибка конфига кейса',        'warn', False),
}


def _build_autoclick_sheet(wb, autoclick):
    """Итоги автокликера (перекликивание ошибок в Вебмастере/ГСК) - сводка
    по сайтам. Добавляется только если автокликер запускался."""
    if not autoclick:
        return
    ws = wb.create_sheet('Автокликер')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C.accent

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 40   # Сайт
    ws.column_dimensions['C'].width = 16   # Сервис
    ws.column_dimensions['D'].width = 12   # Проблем
    ws.column_dimensions['E'].width = 14   # Прокликано
    ws.column_dimensions['F'].width = 16   # Проверяются
    ws.column_dimensions['G'].width = 14   # Без кнопки
    ws.column_dimensions['H'].width = 12   # Ошибки

    ws.merge_cells('B2:H2')
    c = ws['B2']
    c.value = 'Автокликер - перекликивание ошибок'
    c.font = _font(size=16, bold=True)
    ws.row_dimensions[2].height = 26

    # Недоступен (нет браузера / облако)
    if not autoclick.get('available'):
        ws.merge_cells('B4:H4')
        c = ws['B4']
        c.value = autoclick.get('note') or (
            'Автокликер не запускался: нужен локальный залогиненный Chrome '
            '(CDP 9222). На облаке недоступен.')
        c.font = _font(size=11, color=C.text_soft)
        c.alignment = _align(wrap=True, vertical='top')
        ws.row_dimensions[4].height = 44
        return

    sites = autoclick.get('sites') or []
    _t_prob = sum(s.get('problems', 0) for s in sites)
    _t_click = sum(s.get('clicked', 0) for s in sites)
    _t_check = sum(s.get('checking', 0) for s in sites)
    _t_skip = sum(s.get('no_button', 0) for s in sites)

    ws.merge_cells('B3:H3')
    c = ws['B3']
    c.value = (f'Сайтов обработано: {len(sites)}.  Проблем: {_t_prob}.  '
               f'Прокликано: {_t_click}.  Уже проверяются: {_t_check}.  '
               f'Без кнопки: {_t_skip}.')
    c.font = _font(size=11, color=C.text_soft)
    ws.row_dimensions[3].height = 22

    hdr_row = 5
    headers = ['Сайт', 'Сервис', 'Проблем', 'Прокликано',
               'Проверяются', 'Без кнопки', 'Ошибки']
    for ci, h in enumerate(headers, 2):
        cell = ws.cell(row=hdr_row, column=ci, value=h)
        cell.font = _font(size=9, bold=True, color=C.text_muted)
        cell.fill = _fill(C.surface)
        cell.alignment = _align(horizontal='center' if ci > 3 else 'left')
        cell.border = _border()
    ws.row_dimensions[hdr_row].height = 22
    ws.freeze_panes = f'B{hdr_row + 1}'

    row = hdr_row + 1
    for s in sorted(sites, key=lambda x: x.get('clicked', 0), reverse=True):
        ws.row_dimensions[row].height = 20
        vals = [
            (s.get('site', ''), 'left', C.text),
            (s.get('service', ''), 'center', C.text_soft),
            (s.get('problems', 0), 'center', C.text_soft),
            (s.get('clicked', 0), 'center', C.ok if s.get('clicked') else C.text_muted),
            (s.get('checking', 0), 'center', C.warn if s.get('checking') else C.text_muted),
            (s.get('no_button', 0), 'center', C.text_muted),
            (s.get('errors', 0), 'center', C.err if s.get('errors') else C.text_muted),
        ]
        for ci, (val, halign, color) in enumerate(vals, 2):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = _font(size=10, color=color,
                              bold=(ci == 5 and bool(s.get('clicked'))))
            cell.alignment = _align(horizontal=halign, indent=1 if halign == 'left' else 0)
            cell.border = _border(color=C.border_light)
        row += 1


# Листы «Регион и СНГ» и «Заголовки и мета» удалены (04.08.2026): находки
# (чужой город/телефон/почта, СНГ-чистота, технический регион по гео-тегам,
# единственность title/description/H1) полностью и читаемо попадают в
# «Проблемы» через report_priorities.py (_region_findings/_cis_findings/
# _geo_findings/_meta_unique_findings).


# ── Пересборка листов в тематические группы ─────────────────────────
# Каждый детальный лист строится как раньше (временный), затем переносится
# СЕКЦИЕЙ в один из 7 групповых листов. Так весь рендер сохраняется без
# переписывания, а отчёт группируется по темам.

# Группа → упорядоченный список исходных листов (что в неё сливается).
_SHEET_GROUPS = [
    # «Структура страниц» - НЕ в группе: остаётся отдельным листом сразу
    # после «Обзора» (как было до пересборки).
    ('Техничка', []),               # находки - в «Проблемы», листы удалены
    ('Верстка', []),                # находки - в «Проблемы», лист удалён
    ('Безопасность', []),          # находки - в «Проблемы», лист удалён
    ('КП', []),                     # находки - в «Проблемы», листы удалены
    ('Формы', []),                 # детальный отчёт форм - отдельный файл
    ('Админка', ['Настройки в админке']),
    ('Аналитика', [
        'Уведомления', 'Ошибки сервисов', 'Автокликер',
    ]),
    ('Контент', ['Уникальность']),
]

# Групповые листы, которым добавляем поясняющую секцию, даже если исходных
# листов нет (чтобы структура из 7 листов существовала и была понятной).
_GROUP_NOTES = {
    'Верстка': ('Вёрстка и адаптивность (viewport, CSS, меню шапки, mixed '
               'content, favicon), поиск по сайту (находит категории/теги) '
               'и живой тест фильтрации товаров. Находки – на листе '
               '«Проблемы» (раздел «Вёрстка»), приоритет – на «План работ».'),
    'КП': ('Сверка контактов с картой присутствия (телефон/почта/адрес), '
          'верные переменные города, чистота СНГ-доменов от РФ и '
          'технический регион поддоменов (гео-теги). Находки – на листе '
          '«Проблемы» (раздел «Регион и город» / «СНГ-домены» / «Контакты '
          'по городам»), приоритет – на «План работ».'),
    'Безопасность': ('Заголовки безопасности ответа сервера (HSTS, CSP, '
                     'X-Content-Type-Options, защита от кликджекинга) по '
                     'выборке страниц. Нет заголовка – предупреждение, '
                     'битое значение – баг. Находки – на листе «Проблемы» '
                     '(раздел «Безопасность»), приоритет – на «План работ».'),
    'Формы': ('Детальная проверка форм – в отдельном отчёте форм-тестера '
              '(свой файл). Здесь, в основном отчёте, форма заявки/телефон '
              'проверяется как часть страниц (см. лист «Структура страниц», '
              'а находки теста 404-страницы – на листе «Проблемы», раздел '
              '«Страница 404»).'),
    'Контент': ('Изображения (alt, современные форматы webp/avif, вес, '
                'lazy, уникальность картинок категорий, фото товаров не '
                'дублируются между категориями) – находки на листе '
                '«Проблемы» (раздел «Изображения»). SEO-текст частотных '
                'категорий (нейроответы) – там же, раздел «Метаданные»; '
                'блоки товара (похожие/отзывы/сортировка/цены) – на листе '
                '«Структура страниц».'),
}

_GROUP_TAB_RANK = {C.err: 0, C.warn: 1}   # для агрегированного цвета вкладки

_RE_WHERE_SHEET = re.compile(r'^Лист «([^»]+)»$')


def _old_to_group_map() -> dict:
    """{старое имя листа: имя группового листа}, куда его сливает
    _regroup_into_groups (Индексация -> Техничка, Вёрстка -> Верстка и т.п.)."""
    return {old: grp for grp, olds in _SHEET_GROUPS for old in olds}


# Листы, которых больше нет: их находки целиком собраны в «Проблемы» и там
# различаются колонкой «Раздел». Ссылки на такой лист («см. лист «Индексация»»)
# без подмены вели бы на несуществующую вкладку.
_MOVED_TO_PROBLEMS = {
    'Индексация', 'Метаданные', 'Ошибки JavaScript', 'Валидация и скорость',
    'Страница 404', 'Дубли главной', 'Индексация (Арсенкин)', 'Фильтры ПС',
    'Нагрузка и парсинг', 'Вёрстка', 'Безопасность', 'Изображения', 'Разметка',
}


def _sheet_ref(name: str) -> str:
    """«X» -> ссылка на существующую вкладку для текста «см. лист {ref}»:
    лист, который _regroup_into_groups сливает в групповой (Ошибки сервисов ->
    Аналитика) - «Группа», раздел «X»; удалённый лист, находки которого уехали
    в «Проблемы» (Индексация/Метаданные/…) - «Проблемы», раздел «X». Листы вне
    группировки (Обзор/Проблемы/Структура страниц/Страницы/Хосты и аномалии/…)
    возвращаются как есть - «X»."""
    grp = _old_to_group_map().get(name)
    if grp and grp != name:
        return f'«{grp}», раздел «{name}»'
    if name in _MOVED_TO_PROBLEMS:
        return f'«Проблемы», раздел «{name}»'
    return f'«{name}»'


def _fix_where_refs(tasks):
    """Task.where вида «Лист «X»» может называть лист, который
    _regroup_into_groups сливает в групповой (Индексация/Фильтры ПС/…
    -> «Техничка», Ошибки сервисов -> «Аналитика» и т.п.) - без исправления
    ссылка в «Плане работ» ведёт на несуществующую вкладку. Переписываем на
    «Лист «Группа», раздел «X»» ДО того, как этот текст попадёт на лист
    (регруппировка выполняется позже, tasks строятся раньше)."""
    for t in tasks:
        m = _RE_WHERE_SHEET.match(t.where or '')
        if not m:
            continue
        t.where = f'Лист {_sheet_ref(m.group(1))}'
    return tasks


def _append_sheet_as_section(dst, src, start_row, title, gap=2):
    """Скопировать содержимое листа src в dst начиная со start_row (значения,
    стили, слияния, ширины колонок, высоты строк, гиперссылки, комментарии).
    Перед секцией – цветная полоса-разделитель с title. Возвращает следующую
    свободную строку."""
    # Полоса-разделитель секции (навигационный якорь).
    dst.merge_cells(start_row=start_row, start_column=2,
                    end_row=start_row, end_column=8)
    band = dst.cell(row=start_row, column=2, value='▸ ' + title)
    band.font = _font(size=12, bold=True, color='FFFFFF')
    band.fill = _fill(C.text_soft)
    band.alignment = _align(indent=1)
    dst.row_dimensions[start_row].height = 22
    row0 = start_row + 1
    offset = row0 - 1                          # src-строка r → dst-строка r+offset

    for col, dim in src.column_dimensions.items():
        if dim.width:
            cur = dst.column_dimensions[col].width or 0
            dst.column_dimensions[col].width = max(cur, dim.width)

    max_row, max_col = src.max_row, src.max_column
    for r in range(1, max_row + 1):
        for cc in range(1, max_col + 1):
            s = src.cell(row=r, column=cc)
            if s.value is None and not s.has_style:
                continue
            d = dst.cell(row=r + offset, column=cc)
            d.value = s.value
            if s.has_style:
                d.font = copy(s.font)
                d.fill = copy(s.fill)
                d.border = copy(s.border)
                d.alignment = copy(s.alignment)
                d.number_format = s.number_format
            if s.hyperlink:
                d.hyperlink = s.hyperlink.target
            if s.comment:
                d.comment = Comment(s.comment.text,
                                    s.comment.author or 'Site Checker')
        rd = src.row_dimensions.get(r)
        if rd is not None and rd.height:
            dst.row_dimensions[r + offset].height = rd.height

    for mr in list(src.merged_cells.ranges):
        c1, r1, c2, r2 = range_boundaries(str(mr))
        dst.merge_cells(start_row=r1 + offset, start_column=c1,
                        end_row=r2 + offset, end_column=c2)

    return max_row + offset + gap


def _regroup_into_groups(wb):
    """Собрать детальные листы в 7 тематических групповых листов.
    Обзор остаётся первым (следом - План работ и Проблемы)."""
    for group_name, members in _SHEET_GROUPS:
        present = [m for m in members if m in wb.sheetnames]
        note = _GROUP_NOTES.get(group_name)
        if not present:
            # Пустая группа (все детальные листы удалены/не запускались в
            # этом прогоне) - вкладку не создаём вообще, даже с одной
            # пояснительной заметкой: реальные данные всё равно на
            # «Проблемы»/«План работ», а лишний почти пустой лист - просто
            # балласт (навигация на «Обзоре» его тоже не перечисляет).
            continue
        grp = wb.create_sheet(group_name)
        grp.sheet_view.showGridLines = False
        grp.column_dimensions['A'].width = 3
        # Цвет вкладки - худший среди секций.
        _rank = 9
        for m in present:
            _tc = getattr(wb[m].sheet_properties, 'tabColor', None)
            _tcv = getattr(_tc, 'rgb', None) or _tc
            if isinstance(_tcv, str):
                _rank = min(_rank, _GROUP_TAB_RANK.get(_tcv[-6:].upper(), 9))
        grp.sheet_properties.tabColor = (
            C.err if _rank == 0 else C.warn if _rank == 1 else C.ok)

        row = 2
        if note:
            grp.merge_cells(start_row=row, start_column=2,
                            end_row=row, end_column=8)
            c = grp.cell(row=row, column=2, value=note)
            c.font = _font(size=10, italic=True, color=C.text_soft)
            c.alignment = _align(wrap=True, vertical='top', indent=1)
            grp.row_dimensions[row].height = 60
            row += 2
        for m in present:
            row = _append_sheet_as_section(grp, wb[m], row, m)
        # Удаляем исходные листы после переноса.
        for m in present:
            del wb[m]
        # Если имя группы совпало с именем её листа (напр. «Безопасность»),
        # create_sheet выше добавил суффикс «1» - возвращаем имя после удаления
        # исходного листа (место освободилось).
        if grp.title != group_name:
            grp.title = group_name

    # Порядок: Обзор → План работ → Проблемы → Структура страниц → Страницы →
    # Хосты и аномалии → Трафик и траст → 7 групп → Я.Бизнес/GMB.
    order = (['Обзор', 'План работ', 'Проблемы', 'Структура страниц',
             'Страницы', 'Хосты и аномалии', 'Трафик и траст']
             + [g for g, _ in _SHEET_GROUPS if g in wb.sheetnames]
             + ['Я.Бизнес и GMB'])
    ordered = [wb[n] for n in order if n in wb.sheetnames]
    ordered += [ws for ws in wb.worksheets if ws not in ordered]
    wb._sheets = ordered


_UNIQ_TYPE_RU = {'main': 'Главная', 'catalog': 'Каталог', 'category': 'Категория',
                 'product': 'Товар', 'filter': 'Фильтр', 'custom': 'URL'}


def _build_uniqueness_sheet(wb, uniqueness):
    """Лист «Уникальность»: по каждой странице - % уникальности по text.ru и с
    какими ЧУЖИМИ сайтами пересекается контент. Источники пересечения показываем
    ВСЕГДА, если они есть (даже когда уникальность выше порога). Добавляется
    только если проверка выполнялась."""
    if not uniqueness or not uniqueness.get('rows'):
        return
    rows = uniqueness['rows']
    thr = float(uniqueness.get('threshold', 95) or 95)
    summ = uniqueness.get('summary', {}) or {}
    _below = [r for r in rows if r.get('unique') is not None and r['unique'] < thr]
    has_bugs = bool(_below)

    ws = wb.create_sheet('Уникальность')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C.err if has_bugs else C.ok

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 58   # URL страницы
    ws.column_dimensions['C'].width = 13   # Тип
    ws.column_dimensions['D'].width = 14   # Уникальность
    ws.column_dimensions['E'].width = 70   # Источники пересечения
    ws.column_dimensions['F'].width = 3

    ws.merge_cells('B2:E2')
    c = ws['B2']
    c.value = 'Уникальность контента (text.ru)'
    c.font = _font(size=16, bold=True)
    ws.row_dimensions[2].height = 26

    ws.merge_cells('B3:E3')
    c = ws['B3']
    c.value = ('Проверяем SEO-текст страниц главного домена через text.ru: процент '
               'уникальности и с какими ЧУЖИМИ сайтами пересекается контент (свои '
               'домены/поддомены исключены). Красным - страницы ниже порога '
               f'({thr:.0f}%). Сайты-источники пересечения показаны ВСЕГДА, если '
               'они есть, - даже при высокой уникальности.')
    c.font = _font(size=10, italic=True, color=C.text_soft)
    c.alignment = _align(wrap=True, vertical='top')
    ws.row_dimensions[3].height = 40

    row = 5
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    c = ws.cell(row=row, column=2)
    _avg = summ.get('avg_unique')
    c.value = (f'Проверено страниц: {summ.get("checked", 0)}/{summ.get("total", 0)} · '
               f'средняя уникальность: {_avg if _avg is not None else "—"}% · '
               f'ниже {thr:.0f}%: {summ.get("below", 0)} · ошибок: {summ.get("errors", 0)}')
    c.font = _font(size=11, bold=True)
    row += 2

    # ── Конкуренты: с кем пересекаемся и на скольких страницах ──
    _comps = summ.get('competitors') or []
    _checked = summ.get('checked', 0) or 0
    if _comps:
        c = ws.cell(row=row, column=2)
        c.value = 'Конкуренты (пересечение по каталогу)'
        c.font = _font(size=12, bold=True)
        row += 1
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c = ws.cell(row=row, column=2)
        c.value = ('С какими ЧУЖИМИ доменами пересекается наш контент и на скольких '
                   f'из {_checked} проверенных страниц. Если домен пересекается с '
                   'большинством страниц - вероятно, у нас скопировали каталог '
                   'целиком (а не одну страницу).')
        c.font = _font(size=10, italic=True, color=C.text_soft)
        c.alignment = _align(wrap=True, vertical='top')
        ws.row_dimensions[row].height = 30
        row += 1
        for comp in _comps[:15]:
            _n = comp.get('pages', 0)
            frac = (_n / _checked) if _checked else 0
            ws.cell(row=row, column=2).value = comp.get('domain', '')
            ws.cell(row=row, column=2).alignment = _align(vertical='top')
            cc = ws.cell(row=row, column=3)
            cc.value = f'{_n} из {_checked}'
            cc.font = _font(bold=True, color=(C.err if frac >= 0.5 else C.text))
            cc.alignment = _align(vertical='top')
            if frac >= 0.5:
                cf = ws.cell(row=row, column=4)
                cf.value = 'возможно скопирован каталог'
                cf.font = _font(italic=True, color=C.err)
                cf.alignment = _align(wrap=True, vertical='top')
            row += 1
        row += 1

    # ── Структура каталога у топ-конкурента ──
    _struct = uniqueness.get('structure') or {}
    if _struct.get('competitor'):
        c = ws.cell(row=row, column=2)
        c.value = f'Структура каталога у конкурента {_struct["competitor"]}'
        c.font = _font(size=12, bold=True)
        row += 1
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c = ws.cell(row=row, column=2)
        if _struct.get('error'):
            c.value = f'Не удалось сверить структуру: {_struct["error"]}'
            c.font = _font(size=10, italic=True, color=C.text_muted)
        else:
            _op = _struct.get('overlap_pct', 0)
            c.value = (f'{_struct.get("matched_count", 0)} из {_struct.get("our_count", 0)} '
                       f'наших категорий (по слагу) есть у него - совпадение структуры '
                       f'{_op:.0f}%. У конкурента всего {_struct.get("their_count", 0)} '
                       'адресов в sitemap. Высокая доля = вероятно, скопировали каталог.')
            c.font = _font(size=11, bold=True, color=(C.err if _op >= 50 else C.text))
        c.alignment = _align(wrap=True, vertical='top')
        ws.row_dimensions[row].height = 30
        row += 1
        _mt = _struct.get('matched') or []
        if _mt:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            c = ws.cell(row=row, column=2)
            c.value = 'Совпавшие категории: ' + ', '.join(_mt[:60])
            c.font = _font(size=10, color=C.text_soft)
            c.alignment = _align(wrap=True, vertical='top')
            ws.row_dimensions[row].height = 44
            row += 1
        row += 1

    heads = ['Страница', 'Тип', 'Уникальность', 'С каким сайтом пересекается (совпадение %)']
    for j, h in enumerate(heads, start=2):
        c = ws.cell(row=row, column=j)
        c.value = h
        c.font = _font(bold=True, color=C.text)
        c.fill = _fill(C.surface)
        c.alignment = _align(wrap=True, vertical='top')
    row += 1

    for r in rows:
        u = r.get('unique')
        ws.cell(row=row, column=2).value = r.get('url', '')
        ws.cell(row=row, column=2).alignment = _align(wrap=True, vertical='top')
        ws.cell(row=row, column=3).value = _UNIQ_TYPE_RU.get(r.get('type'), r.get('type', ''))
        ws.cell(row=row, column=3).alignment = _align(vertical='top')
        cu = ws.cell(row=row, column=4)
        if r.get('error'):
            cu.value = f'— ({r["error"]})'
            cu.font = _font(italic=True, color=C.text_muted)
        elif u is not None:
            cu.value = f'{u:.1f}%'
            cu.font = _font(bold=True, color=(C.err if u < thr else C.ok))
        cu.alignment = _align(vertical='top')
        # Источники пересечения - ВСЕГДА, если есть.
        src = '\n'.join(
            (f'{s.get("url", "")}  ({s["plagiat"]:.1f}%)'
             if s.get('plagiat') is not None else s.get('url', ''))
            for s in (r.get('sources') or []) if s.get('url'))
        # Авто-углубление: своя бесплатная сверка с топ-конкурентом.
        _dp = r.get('deep')
        if _dp and _dp.get('a_in_b') is not None:
            _dl = (f'🔎 наша сверка: {_dp["a_in_b"]:.0f}% нашего текста на '
                   f'{_dp.get("competitor", "")}')
            if _dp.get('samples'):
                _dl += ' · фразы: ' + '; '.join(f'«{x}»' for x in _dp['samples'][:2])
            src = (src + '\n' + _dl) if src else _dl
        cs = ws.cell(row=row, column=5)
        cs.value = src or ('—' if not r.get('error') else '')
        cs.alignment = _align(wrap=True, vertical='top')
        row += 1


# ── Лист «Проблемы» - плоский список всех находок (report_priorities) ──


_LEVEL_COLOR = {'Ошибка': 'err', 'Предупреждение': 'warn'}
_LEVEL_FILL = {'Ошибка': 'err_soft', 'Предупреждение': 'warn_soft'}


def _build_problems_sheet(wb, findings, source_by_url=None, interlinking=None):
    """Лист «Проблемы»: одна строка = одна находка на одной странице (все
    проверки чек-листа - report_priorities.collect_findings). Фильтруется
    по любой колонке (автофильтр).

    source_by_url - {адрес: откуда взят}, чтобы к странице из карты сайта или
    своего списка приписать источник: иначе непонятно, откуда в отчёте взялся
    незнакомый URL, которого нет в каталоге проекта.

    interlinking - вывод о перелинковке по САЙТУ (report_priorities.
    interlinking_note): это не дефект конкретной страницы, поэтому идёт
    отдельным блоком ПОД таблицей и только когда есть проблема."""
    source_by_url = source_by_url or {}
    _MAX_ROWS = 3000
    ws = wb.create_sheet('Проблемы')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C.err if any(
        f.level == 'Ошибка' for f in findings) else (C.warn if findings else C.ok)

    for col, w in (('A', 3), ('B', 4), ('C', 14), ('D', 20), ('E', 46),
                  ('F', 12), ('G', 14), ('H', 55), ('I', 40), ('J', 46)):
        ws.column_dimensions[col].width = w

    ws.merge_cells('B2:J2')
    c = ws['B2']
    c.value = 'Все находки по страницам - одна строка = одна проблема на одной странице'
    c.font = _font(size=14, bold=True)
    ws.row_dimensions[2].height = 24

    ws.merge_cells('B3:J3')
    c = ws['B3']
    c.value = ('Фильтруйте столбцы: «Уровень» - что критично, «Раздел» - чья '
              'зона ответственности, «Проблема» - конкретная задача. '
              'Ошибка = чинить, Предупреждение = улучшение. «Как исправить» - '
              'готовая рекомендация с пояснением, почему это важно.')
    c.font = _font(size=10, italic=True, color=C.text_soft)
    c.alignment = _align(wrap=True, vertical='top')
    ws.row_dimensions[3].height = 30

    hdr = 5
    headers = (('B', '№'), ('C', 'Уровень'), ('D', 'Раздел'),
              ('E', 'Проблема'), ('F', 'Город'), ('G', 'Тип страницы'),
              ('H', 'Адрес страницы'), ('I', 'Подробности'),
              ('J', 'Как исправить'))
    for col, title in headers:
        cell = ws[f'{col}{hdr}']
        cell.value = title
        cell.font = _font(size=10, bold=True, color=C.bg_elev)
        cell.fill = _fill(C.header_navy)
        cell.border = _border()
        cell.alignment = _align(indent=1)
    ws.row_dimensions[hdr].height = 20

    ordered = sorted(findings, key=lambda f: (f.level != 'Ошибка', f.section,
                                              f.problem, f.url))
    row = hdr + 1
    for i, f in enumerate(ordered[:_MAX_ROWS], 1):
        color = getattr(C, _LEVEL_COLOR.get(f.level, 'text_soft'))
        fill_name = _LEVEL_FILL.get(f.level)
        meta = classify(f)
        fix = meta['title']
        if meta.get('why'):
            fix = f'{fix} - {meta["why"]}'
        # Справочные цифры находки (замеры скорости, объёмы, «проверено N
        # путей») - раньше они стояли строкой на детальном листе; листов
        # больше нет, поэтому дописываем их сюда.
        if getattr(f, 'fix_note', ''):
            fix = f'{fix} {f.fix_note}'
        _тип = f.page_type
        _ист = source_by_url.get(f.url)
        if _ист and _ист != 'Каталог проекта':
            _тип = f'{_тип} · {_ист}' if _тип else _ист
        vals = (i, f.level, f.section, f.problem, f.city, _тип,
               f.url, f.detail, fix)
        for col_i, (col, _) in enumerate(headers):
            cell = ws[f'{col}{row}']
            cell.value = vals[col_i]
            cell.font = _font(size=9, color=color if col == 'C' else C.text_soft,
                              bold=(col == 'C'))
            cell.alignment = _align(wrap=(col in ('E', 'H', 'I', 'J')),
                                    vertical='top', indent=1)
            cell.border = _border(color=C.border_light)
            if col == 'C' and fill_name:
                cell.fill = _fill(getattr(C, fill_name))
        if f.url:
            u = ws[f'H{row}']
            u.hyperlink = f.url
            u.font = _font(size=9, color=C.accent, underline='single')
        ws.row_dimensions[row].height = 15
        row += 1

    last = row - 1
    if last >= hdr + 1:
        ws.auto_filter.ref = f'B{hdr}:J{last}'
        ws.freeze_panes = f'B{hdr + 1}'
    if len(ordered) > _MAX_ROWS:
        row += 1
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        c = ws.cell(row=row, column=2)
        c.value = f'…показаны первые {_MAX_ROWS} из {len(ordered)} находок.'
        c.font = _font(size=10, italic=True, color=C.text_muted)

    # Перелинковка - вывод по всему сайту, а не по странице: в таблицу с
    # автофильтром такую строку класть некуда (нет ни адреса, ни типа
    # страницы), поэтому показываем блоком под таблицей.
    if interlinking:
        row += 2
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        c = ws.cell(row=row, column=2, value='Перелинковка (внутренний вес)')
        c.font = _font(size=12, bold=True, color=C.warn)
        c.fill = _fill(C.warn_soft)
        c.alignment = _align(indent=1)
        ws.row_dimensions[row].height = 22
        row += 1
        for текст, стиль in ((interlinking.get('text', ''), C.text),
                             (interlinking.get('detail', ''), C.text_muted)):
            if not текст:
                continue
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
            c = ws.cell(row=row, column=2, value=текст)
            c.font = _font(size=10, color=стиль)
            c.alignment = _align(wrap=True, vertical='top', indent=1)
            ws.row_dimensions[row].height = _row_height_for(текст, 230)
            row += 1


# ── Лист «План работ» - приоритезированные задачи (report_priorities) ──


def _build_work_plan_sheet(wb, tasks):
    """Лист «План работ»: находки, сгруппированные в задачи и
    отсортированные по приоритету (report_priorities.group_into_tasks +
    extra_site_tasks)."""
    ws = wb.create_sheet('План работ')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = (
        C.err if any(t.priority == 1 for t in tasks)
        else C.warn if tasks else C.ok)

    for col, w in (('A', 3), ('B', 14), ('C', 34), ('D', 55), ('E', 10),
                  ('F', 50), ('G', 20), ('H', 40)):
        ws.column_dimensions[col].width = w

    ws.merge_cells('B2:H2')
    c = ws['B2']
    c.value = 'План работ - что чинить и в каком порядке'
    c.font = _font(size=14, bold=True)
    ws.row_dimensions[2].height = 24

    ws.merge_cells('B3:H3')
    c = ws['B3']
    c.value = ('Сверху вниз: сначала «Критично» - то, что прямо сейчас теряет '
              'заказы и позиции. Столбец «Объём» показывает, сколько страниц '
              'или объектов затронуто.')
    c.font = _font(size=10, italic=True, color=C.text_soft)
    c.alignment = _align(wrap=True, vertical='top')
    ws.row_dimensions[3].height = 30

    hdr = 5
    headers = (('B', 'Приоритет'), ('C', 'Задача'), ('D', 'Что именно не так'),
              ('E', 'Объём'), ('F', 'Почему это важно'), ('G', 'Кому'),
              ('H', 'Где смотреть детали'))
    for col, title in headers:
        cell = ws[f'{col}{hdr}']
        cell.value = title
        cell.font = _font(size=10, bold=True, color=C.bg_elev)
        cell.fill = _fill(C.header_navy)
        cell.border = _border()
        cell.alignment = _align(indent=1)
    ws.row_dimensions[hdr].height = 20

    _PRIO_COLOR = {1: C.err, 2: C.warn, 3: C.text_muted}
    _PRIO_FILL = {1: C.err_soft, 2: C.warn_soft, 3: C.surface}
    row = hdr + 1
    for t in tasks:
        color = _PRIO_COLOR.get(t.priority, C.text_muted)
        fill_color = _PRIO_FILL.get(t.priority)
        vals = (PRIORITY_LABEL.get(t.priority, str(t.priority)), t.title,
               t.what, t.volume, t.why, t.owner, t.where)
        for col_i, (col, _) in enumerate(headers):
            cell = ws[f'{col}{row}']
            cell.value = vals[col_i]
            cell.font = _font(size=10, bold=(col == 'B'),
                              color=color if col == 'B' else C.text_soft)
            cell.alignment = _align(wrap=(col in ('C', 'D', 'F', 'H')),
                                    vertical='top', indent=1,
                                    horizontal='center' if col == 'E' else 'left')
            cell.border = _border(color=C.border_light)
            if col == 'B' and fill_color:
                cell.fill = _fill(fill_color)
        ws.row_dimensions[row].height = 32
        row += 1

    last = row - 1
    if last >= hdr + 1:
        ws.auto_filter.ref = f'B{hdr}:H{last}'
        ws.freeze_panes = f'B{hdr + 1}'


# ── Главная функция ────────────────────────────────────────────────


def build_report(
    *,
    project_name: str,
    started_at_ms: int,
    finished_at_ms: int,
    selected_subdomains: list,    # список Subdomain
    results: list,                 # список CheckResult
    output_path: Path | str,
    metrika_reports: list = None,  # список Report404 - находки в «Проблемы», раздел «404 в индексе»
    metrika_data_date: str = None, # дата отчёта Метрики (YYYY-MM-DD) - сейчас не используется
    metrika_is_stale: bool = False,# True если данные не за вчера - сейчас не используется
    metrika_404_goal: dict = None, # has_404_goal() - сейчас не используется
    notifications: list = None,    # список WebmasterNotification - добавит лист «Уведомления»
    service_issues: list = None,   # список ServiceIssue - добавит лист «Ошибки сервисов»
    autoclick: dict = None,        # итоги автокликера - добавит лист «Автокликер»
    indexing_summary: dict = None, # sitemap↔robots (п.1.7) - в лист «Индексация»
    meta_summary: dict = None,     # дубли мета/URL (п.1.8) - в лист «Метаданные»
    filters_test: dict = None,     # итоги фильтр-теста - секция на листе «Вёрстка»
    console_check: dict = None,    # ошибки JS в консоли (п.1.14) - лист «Ошибки JavaScript»
    calltracking_check: dict = None,  # браузерная проверка замены рекл. номера - находки в «Проблемы»
    w3c_check: dict = None,        # валидация W3C + скорость (п.1.16) - лист «Валидация и скорость»
    p404_check: dict = None,       # страница 404 (п.1.18) - лист «Страница 404»
    ps_filters: dict = None,       # фильтры ПС (п.1.19) - лист «Фильтры ПС»
    search_check: dict = None,     # поиск находит категории - секция «Вёрстки»
    index_404_check: dict = None,  # 404 среди страниц в индексе - находки в «Проблемы»
    stress_check: dict = None,     # ошибки сервера: парсинг/нагрузка/дубли - лист «Нагрузка и парсинг»
    link_profile: dict = None,     # lite-профиль ссылок (Вебмастер) - секция на «Трафик и траст»
    wm_metrics: dict = None,       # аномалии Вебмастера - лист «Хосты и аномалии»
    admin_settings: dict = None,   # функции настройки в админке (п.1.21) - лист «Настройки в админке»
    yabusiness: dict = None,       # Я.Бизнес/GMB (поддомен под свой регион) - лист «Я.Бизнес и GMB»
    gsc_pages: dict = None,        # количество страниц в ГСК (индекс/не-индекс/сумма) - секция на «Трафик и траст»
    home_dupes: dict = None,       # дубли главной страницы - лист «Дубли главной»
    traffic: dict = None,          # сравнение трафика день/месяц/год - лист «Трафик и траст»
    arsenkin: dict = None,         # индексация URL через Арсенкин - лист «Индексация (Арсенкин)»
    review_priority: dict = None,  # приоритет докупки отзывов - лист «Отзывы (докупка)»
    anomalies: dict = None,        # аномалии ГСК/Метрика - лист «Хосты и аномалии»
    trust: dict = None,            # ИКС + DR - секция на «Трафик и траст»
    uniqueness: dict = None,       # уникальность контента (text.ru) - лист «Уникальность»
) -> Path:
    """Сформировать xlsx-отчёт и сохранить в output_path."""
    wb = Workbook()
    # Удаляем дефолтный пустой лист
    wb.remove(wb.active)

    # ── Подсчёт метрик ─────────────────────────────────────────────
    total = len(results)
    ok_count = sum(1 for r in results if r.is_ok)
    warn_count = sum(1 for r in results if r.is_warning)
    err_count = total - ok_count - warn_count
    duration_sec = (finished_at_ms - started_at_ms) // 1000

    pages_with_issues = [r for r in results if r.has_text_issues]
    total_text_issues = sum(len(r.text_issues) for r in pages_with_issues)

    # Структурные проблемы (баги контента: нет цены, кнопок, H1 и т.п.)
    pages_with_content = [r for r in results if getattr(r, 'content', None) is not None]
    pages_with_content_bugs = [r for r in pages_with_content if r.content_bugs > 0]
    total_content_bugs = sum(r.content_bugs for r in pages_with_content)

    # Индексация (п.1.7): страницы выборки, закрытые от индексации
    indexing_bad_pages = [r for r in results if getattr(r, 'has_indexing_issues', False)]
    indexing_sitemap_conflicts = len((indexing_summary or {}).get('disallowed') or [])

    # Метаданные (п.1.8): проблемы title/description/H1 + дубли + единственность
    meta_bad_pages = [r for r in results
                      if getattr(r, 'has_meta_issues', False)
                      or getattr(r, 'has_meta_unique_issues', False)]

    # Вёрстка (п.1.11): нет viewport / битые CSS
    layout_bad_pages = [r for r in results if getattr(r, 'has_layout_issues', False)]

    # Разметка (п.1.12): OG/Schema.org
    markup_bad_pages = [r for r in results if getattr(r, 'has_markup_issues', False)]

    # Заголовки безопасности (доп. 1.8): битые значения HSTS/CSP/X-Frame
    security_bad_pages = [r for r in results
                          if getattr(r, 'has_security_issues', False)]

    # Изображения (п.1.15): картинки без alt
    images_bad_pages = [r for r in results
                        if getattr(r, 'has_image_issues', False)]
    _mdups = (meta_summary or {}).get('duplicates') or {}
    meta_dup_groups = (len(_mdups.get('same_city') or [])
                       + len(_mdups.get('cross_city') or [])
                       + sum(1 for d in ((meta_summary or {})
                                         .get('url_duplicates') or [])
                             if d.get('problem') != 'not_301'))

    # Находки со всех проверок (лист «Проблемы») + приоритезированный план
    # работ (лист «План работ») - report_priorities.py. wm_metrics - задачи
    # уровня хоста, в «Проблемы» не попадают (те же данные видны на листе
    # «Хосты и аномалии», отдельный дубль не нужен).
    _page_findings = (collect_findings(results, console_check=console_check,
                                       index_404_check=index_404_check,
                                       metrika_reports=metrika_reports,
                                       calltracking_check=calltracking_check,
                                       search_check=search_check,
                                       filters_test=filters_test)
                      + metadata_site_findings(meta_summary)
                      + home_dupes_findings(home_dupes)
                      + arsenkin_findings(arsenkin)
                      + page404_findings(p404_check)
                      + stress_check_findings(stress_check)
                      # Бывшие секции листа «Техничка»: интерактив (слайдер/
                      # меню/cookie/модалка) и доставка статики (сжатие, кеш,
                      # общее время загрузки). Листа больше нет - находки
                      # живут только здесь.
                      + ux_interactive_findings(console_check)
                      + static_delivery_findings(w3c_check)
                      # Я.Бизнес и отзывы: раньше жили только на листе
                      # «Я.Бизнес и GMB» и не попадали ни в «Проблемы», ни в
                      # «План работ». Лист пока оставлен ради таблицы
                      # приоритета докупки - в колонки «Проблем» она не
                      # ложится (там строка = страница сайта).
                      + yabusiness_findings(yabusiness)
                      + review_priority_findings(review_priority))
    # Сайт-уровневые находки индексации (пути/файлы) и санкции ПС - только
    # в «Проблемы» (список), «План работ» их агрегирует extra_site_tasks()
    # ниже - через group_into_tasks они бы задвоились. Остальные (дубли
    # метаданных, дубли главной, Арсенкин, 404-тест, нагрузка/парсинг)
    # своей агрегации нигде больше не имеют - идут через group_into_tasks()
    # как обычные находки. Ошибки сервисов и W3C - тоже только в «Проблемы»
    # (агрегация service_issues в «План работ» - по хосту, не по issue;
    # W3C вообще без своей задачи, только числа-колонки на «Страницы»).
    _findings = (_page_findings + indexing_site_findings(indexing_summary)
                + ps_filters_findings(ps_filters)
                + service_issues_findings(service_issues)
                + w3c_findings(w3c_check)
                # Тоже бывшие секции «Технички»: ЧПУ-адреса (строка на адрес),
                # гигиена robots.txt и разделы «Отгрузки»/новости. Как и
                # прочие сайт-уровневые - только в «Проблемы», в «План работ»
                # они попадают через group_into_tasks ниже не идут.
                + url_format_findings(indexing_summary)
                + robots_hygiene_findings(indexing_summary)
                + content_sections_findings(indexing_summary))
    _tasks = (group_into_tasks(_page_findings)
             + extra_site_tasks(indexing_summary=indexing_summary,
                                wm_metrics=wm_metrics,
                                service_issues=service_issues,
                                ps_filters=ps_filters))
    _fix_where_refs(_tasks)
    _tasks.sort(key=lambda t: (t.priority, -t.volume))
    _critical_tasks = sum(1 for t in _tasks if t.priority == 1)

    # ═══════════════════════════════════════════════════════════════
    # ЛИСТ 1: Обзор
    # ═══════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet('Обзор')
    ws1.sheet_view.showGridLines = False

    # Ширины колонок - B..H под 7 карточек метрик, дальше секции идут тем же
    # диапазоном (описания смёрджены на всю ширину B:H).
    ws1.column_dimensions['A'].width = 3
    for col, w in (('B', 20), ('C', 17), ('D', 13), ('E', 13),
                  ('F', 13), ('G', 13), ('H', 20)):
        ws1.column_dimensions[col].width = w
    ws1.column_dimensions['I'].width = 3

    # Заголовок
    ws1.merge_cells('B2:H2')
    c = ws1['B2']
    c.value = 'Отчёт по проверке сайта'
    c.font = _font(size=20, bold=True)
    ws1.row_dimensions[2].height = 30

    ws1.merge_cells('B3:H3')
    started_dt = datetime.fromtimestamp(started_at_ms / 1000)
    c = ws1['B3']
    c.value = f'{project_name} · {started_dt.strftime("%d.%m.%Y, %H:%M:%S")}'
    c.font = _font(size=11, color=C.text_muted)
    ws1.row_dimensions[3].height = 20

    # ─── 7 карточек метрик ───────────────────────────────────────────
    card_row = 6
    ws1.row_dimensions[card_row].height = 22
    ws1.row_dimensions[card_row + 1].height = 38

    _total_findings = len(_findings)
    _err_findings = sum(1 for f in _findings if f.level == 'Ошибка')
    _warn_findings = sum(1 for f in _findings if f.level == 'Предупреждение')

    def _card_fill(kind: str, value: int) -> str:
        # 'good' - позитивная метрика (всегда зелёная), 'bad' - красная, пока
        # значение > 0 (иначе зелёная - нечего чинить), 'warn' - жёлтая при
        # значении > 0, 'neutral' - нейтральная заливка (просто счётчик).
        if kind == 'good':
            return C.ok_soft
        if kind == 'bad':
            return C.err_soft if value else C.ok_soft
        if kind == 'warn':
            return C.warn_soft if value else C.ok_soft
        return C.surface

    metrics = [
        ('B', 'СТРАНИЦ ПРОВЕРЕНО', total, C.text, 'neutral'),
        ('C', 'ОТКРЫВАЮТСЯ', ok_count, C.ok, 'good'),
        ('D', 'НЕ ОТКРЫВАЮТСЯ', err_count, C.err, 'bad'),
        ('E', 'ВСЕГО НАХОДОК', _total_findings, C.text, 'neutral'),
        ('F', 'ИЗ НИХ ОШИБОК', _err_findings, C.err, 'bad'),
        ('G', 'ПРЕДУПРЕЖДЕНИЙ', _warn_findings, C.warn, 'warn'),
        ('H', 'КРИТИЧНЫХ ЗАДАЧ', _critical_tasks, C.err, 'bad'),
    ]
    for col, label, value, color, kind in metrics:
        top = ws1[f'{col}{card_row}']
        top.value = label
        top.font = _font(size=9, bold=True, color=C.text_muted)
        top.alignment = _align()
        top.fill = _fill(C.surface)
        top.border = _border()

        bot = ws1[f'{col}{card_row + 1}']
        bot.value = value
        bot.font = _font(size=22, bold=True, color=color)
        bot.alignment = _align()
        bot.fill = _fill(_card_fill(kind, value))
        bot.border = _border()

    # ─── Коротко о главном (сводка) ──────────────────────────────────
    sum_row = card_row + 3
    ws1.row_dimensions[sum_row].height = 26
    ws1.merge_cells(f'B{sum_row}:H{sum_row}')
    c = ws1[f'B{sum_row}']
    c.value = 'Коротко о главном'
    c.font = _font(size=12, bold=True)
    c.alignment = _align()
    c.fill = _fill(C.surface)
    c.border = _border()

    sum_body_row = sum_row + 1
    ws1.merge_cells(f'B{sum_body_row}:H{sum_body_row}')
    c = ws1[f'B{sum_body_row}']
    summary_text = (
        f'Из {total} проверенных страниц: '
        f'{ok_count} работают, {warn_count} с перенаправлениями, {err_count} не открываются.'
    )
    if total_text_issues > 0:
        summary_text += (
            f'\nДополнительно: на {len(pages_with_issues)} страницах найдено '
            f'{total_text_issues} битых переменных в текстах - см. «Проблемы».'
        )
    if total_content_bugs > 0:
        summary_text += (
            f'\nВ контенте {total_content_bugs} проблем на {len(pages_with_content_bugs)} страницах '
            f'(нет цены, кнопок заказа или заголовка) - см. лист «Структура страниц».'
        )
    _idx_blanket = (indexing_summary or {}).get('blanket_disallow') or []
    _idx_assets = (indexing_summary or {}).get('assets_closed') or []
    _idx_mc = (((indexing_summary or {}).get('sitemap_audit') or {})
               .get('missing_catalog') or {})
    _idx_missing = ((_idx_mc.get('categories') or [])
                    + (_idx_mc.get('filters') or [])
                    + (_idx_mc.get('services') or []))
    _idx_hm_junk = (((indexing_summary or {}).get('html_sitemap') or {})
                    .get('junk_links') or [])
    if (indexing_bad_pages or indexing_sitemap_conflicts
            or _idx_blanket or _idx_assets or _idx_missing or _idx_hm_junk):
        _idx_bits = []
        if indexing_bad_pages:
            _idx_bits.append(f'расхождения с robots.txt на {len(indexing_bad_pages)} '
                             f'{_plural_pages(len(indexing_bad_pages))}')
        if indexing_sitemap_conflicts:
            _idx_bits.append(f'{indexing_sitemap_conflicts} путей каталога под Disallow '
                             f'в robots.txt')
        if _idx_blanket:
            _idx_bits.append('в robots.txt есть «Disallow: /» - сайт закрыт целиком')
        if _idx_assets:
            _idx_bits.append(f'{len(_idx_assets)} файлов .css/.js закрыты в robots.txt')
        if _idx_missing:
            _idx_bits.append(f'{len(_idx_missing)} важных ссылок '
                             f'(категории/фильтры/услуги) нет в sitemap')
        if _idx_hm_junk:
            _idx_bits.append(f'{len(_idx_hm_junk)} служебных ссылок в HTML-карте')
        summary_text += ('\nИндексация: ' + ', '.join(_idx_bits)
                         + ' - см. «Проблемы» и «План работ».')
    if meta_bad_pages or meta_dup_groups:
        _mb = []
        _mb_ref = '«Проблемы»'
        if meta_bad_pages:
            _mb.append(f'проблемы на {len(meta_bad_pages)} '
                       f'{_plural_pages(len(meta_bad_pages))}')
        if meta_dup_groups:
            _mb.append(f'{meta_dup_groups} групп дублей (title/описания/URL)')
            # Дубли (в отличие от находок по странице) не попадают в
            # «Проблемы» - только на сам лист метаданных.
            _mb_ref = f'«Проблемы» (страницы) и лист {_sheet_ref("Метаданные")} (дубли)'
        summary_text += ('\nМетаданные: ' + ', '.join(_mb) + ' - см. ' + _mb_ref + '.')
    if layout_bad_pages:
        summary_text += (f'\nВёрстка: проблемы (viewport/CSS) на '
                         f'{len(layout_bad_pages)} '
                         f'{_plural_pages(len(layout_bad_pages))} - '
                         f'см. «Проблемы».')
    if markup_bad_pages:
        summary_text += (f'\nРазметка: проблемы (OG/Schema.org) на '
                         f'{len(markup_bad_pages)} '
                         f'{_plural_pages(len(markup_bad_pages))} - '
                         f'см. «Проблемы».')
    if security_bad_pages:
        summary_text += (f'\nБезопасность: ошибки заголовков на '
                         f'{len(security_bad_pages)} '
                         f'{_plural_pages(len(security_bad_pages))} - '
                         f'см. «Проблемы».')
    if images_bad_pages:
        summary_text += (f'\nИзображения: картинки без alt на '
                         f'{len(images_bad_pages)} '
                         f'{_plural_pages(len(images_bad_pages))} - '
                         f'см. «Проблемы».')
    _filters_cases = (filters_test or {}).get('cases') or []
    _filters_bad = sum(1 for c in _filters_cases
                       if _FILTER_VERDICT.get(c.get('verdict'),
                                              (None, None, False))[2])
    if _filters_bad:
        summary_text += (f'\nФильтрация: {_filters_bad} '
                         f'{"фильтр" if _filters_bad == 1 else "фильтров"} '
                         f'работают некорректно - см. «Проблемы».')
    _console_bad = sum(1 for p in ((console_check or {}).get('pages') or [])
                       if p.get('errors'))
    if _console_bad:
        summary_text += (f'\nОшибки JavaScript: на {_console_bad} '
                         f'{_plural_pages(_console_bad)} есть ошибки в консоли '
                         f'- см. «Проблемы».')
    if stress_check and stress_check.get('available'):
        _sp = stress_check.get('parsing') or {}
        _sl = stress_check.get('load') or {}
        _sd = stress_check.get('duplicates') or {}
        _s5 = (len(_sp.get('server_errors') or [])
               + sum(p.get('server_5xx', 0) for p in (_sl.get('pages') or []))
               + len(_sd.get('server_errors') or []))
        _stress_ref = _sheet_ref('Нагрузка и парсинг')
        if _sp.get('banned'):
            summary_text += ('\nНагрузка и парсинг: сайт закрыл доступ '
                             f'(принял бота за парсера) - см. лист {_stress_ref}.')
        elif _s5:
            summary_text += (f'\nНагрузка и парсинг: ошибок сервера (5xx) '
                             f'{_s5} - см. лист {_stress_ref}.')
    if link_profile and link_profile.get('available'):
        _lp_w = sum(len(h.get('warnings') or [])
                    for h in (link_profile.get('hosts') or []))
        if _lp_w:
            summary_text += (f'\nСсылочный профиль: замечаний {_lp_w} '
                             f'(обвал/всплеск/спам) - см. лист '
                             f'«Трафик и траст».')
    if admin_settings and admin_settings.get('available'):
        _adm_bad = [c.get('title') for c in (admin_settings.get('checks') or [])
                    if not c.get('ok')]
        if _adm_bad:
            summary_text += ('\nНастройки в админке: не работают - '
                             + ', '.join(_adm_bad)
                             + f' (см. лист {_sheet_ref("Настройки в админке")}).')
    summary_text += '\nПодробности по страницам - на листе «Страницы» (фильтр по колонке «Статус»).'
    # Ссылки на старые листы → на группу-лист (блок внутри группы), т.к.
    # детальные листы теперь секции в 7 тематических листах.
    _sheet_to_group = {m: g for g, ms in _SHEET_GROUPS for m in ms}
    for _old, _grp in _sheet_to_group.items():
        summary_text = summary_text.replace(
            f'лист «{_old}»', f'лист «{_grp}» (блок «{_old}»)')
    c.value = summary_text
    c.font = _font(size=11, color=C.text_soft)
    c.alignment = _align(wrap=True)
    c.fill = _fill(C.bg_elev)
    c.border = _border()

    # Высота строки - по факту переносов при реальной ширине смёрженных
    # колонок (грубая, но честная оценка - без Excel не измерить пиксели).
    _obzor_width = sum(ws1.column_dimensions[c_].width or 0 for c_ in 'BCDEFGH')
    ws1.row_dimensions[sum_body_row].height = _row_height_for(
        summary_text, _obzor_width, min_px=30)

    # ─── Параметры прогона ─────────────────────────────────────────
    param_row = sum_body_row + 2
    ws1.row_dimensions[param_row].height = 22
    ws1.merge_cells(f'B{param_row}:H{param_row}')
    c = ws1[f'B{param_row}']
    c.value = 'Параметры прогона'
    c.font = _font(size=10, bold=True, color=C.text_muted)
    c.alignment = _align()

    params = [('Длительность', f'{duration_sec} сек')]
    if selected_subdomains:
        cities = ', '.join(s.city for s in selected_subdomains)
        params.append(('Поддоменов', f'{len(selected_subdomains)} ({cities})'))

    for i, (key, value) in enumerate(params):
        r = param_row + 1 + i
        ws1.row_dimensions[r].height = 22
        k = ws1[f'B{r}']
        k.value = key
        k.font = _font(size=10, color=C.text_muted)
        k.alignment = Alignment(horizontal='left', vertical='top', indent=1)

        ws1.merge_cells(f'C{r}:H{r}')
        v = ws1[f'C{r}']
        v.value = value
        v.font = _font(size=10, color=C.text_soft)
        v.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # ─── Навигация по отчёту (для тех, кто открыл впервые) ──────────
    nav_row = param_row + len(params) + 2
    ws1.row_dimensions[nav_row].height = 22
    ws1.merge_cells(f'B{nav_row}:H{nav_row}')
    c = ws1[f'B{nav_row}']
    c.value = 'Из чего состоит отчёт'
    c.font = _font(size=10, bold=True, color=C.text_muted)
    c.alignment = _align()

    # Отчёт собран в 7 тематических листов (каждый - несколько блоков-секций).
    nav_items = [
        ('Обзор', 'эта страница: сколько проверено, сколько работает и сколько сломано.'),
        ('План работ', 'все задачи по приоритету: что чинить, почему это важно и кому передать.'),
        ('Проблемы', 'каждая находка отдельной строкой: уровень, раздел, адрес страницы. Фильтруется как угодно. Здесь же вся SEO-техничка: индексация (robots/sitemap/canonical/ЧПУ), метаданные, микроразметка, ошибки JavaScript, валидность W3C, скорость и доставка статики, тест страницы 404, санкции ПС, нагрузка/парсинг, 404 в индексе.'),
        ('Структура страниц', 'что чинить в контенте по типам страниц (главная/каталог/листинг/разделы/карточки товаров/технические) - где нет цены, кнопок заказа, заголовка. Красное = баг.'),
        ('Страницы', 'каждая проверенная страница: адрес, код ответа, статус, скорость, битые переменные в тексте, откуда перешли и сколько находок (детали - в «Проблемах»).'),
        ('Хосты и аномалии', 'проблемы уровня сайта/хоста целиком (не одной страницы): фатальные проблемы из сервисов и аномалии обхода/ссылок «от себя-прошлого» - обычно самое срочное.'),
        ('Трафик и траст', 'краткая сводка трафика по странам/периодам (визиты, каналы, лиды, конверсия, отказы) + траст проекта (ИКС/DR), lite-профиль беклинков и страницы в ГСК.'),
        ('Админка', 'работа функций настройки в админке: поддомены/категории/товары/тех.страницы + CRUD (создание/правка/скрытие/удаление) с аудитом «было → стало».'),
        ('Аналитика', 'письма Вебмастера/GSC/Я.Бизнеса/2ГИС/Google, ошибки сервисов (сайтмапы/дубли/мусорные ссылки), прокликивание исправлений.'),
        ('Контент', 'если есть лист - уникальность контента (text.ru); изображения, вёрстка, КП и контакты - все на листе «Проблемы».'),
        ('Я.Бизнес и GMB', 'если есть лист - каждый поддомен (город) зарегистрирован в Яндекс.Бизнесе под своим регионом; поддомены без организации.'),
    ]
    nav_hdr = nav_row + 1
    for col, title in (('B', 'Лист'), ('C', 'Что там')):
        cell = ws1[f'{col}{nav_hdr}']
        cell.value = title
        cell.font = _font(size=9, bold=True, color=C.bg_elev)
        cell.fill = _fill(C.accent)
        cell.border = _border()
        cell.alignment = _align(indent=1)
    ws1.merge_cells(f'C{nav_hdr}:H{nav_hdr}')
    ws1.row_dimensions[nav_hdr].height = 20
    _nav_desc_w = sum(ws1.column_dimensions[c_].width or 0 for c_ in 'CDEFGH')
    for i, (sheet_name, desc) in enumerate(nav_items):
        r = nav_hdr + 1 + i
        ws1.row_dimensions[r].height = _row_height_for(desc, _nav_desc_w)
        k = ws1[f'B{r}']
        k.value = sheet_name
        k.font = _font(size=10, bold=True, color=C.accent, underline='single')
        k.alignment = Alignment(horizontal='left', vertical='top', indent=1)
        k.border = _border(color=C.border_light)
        # Ссылка на лист ставится безусловно: на момент сборки «Обзора»
        # остальные листы (План работ/Проблемы/группы) ещё не созданы -
        # появятся ниже по коду, но к моменту сохранения файла будут
        # существовать все, внутренняя ссылка на них верна.
        k.hyperlink = f"#'{sheet_name}'!A1"
        ws1.merge_cells(f'C{r}:H{r}')
        v = ws1[f'C{r}']
        v.value = desc
        v.font = _font(size=10, color=C.text_soft)
        v.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        v.border = _border(color=C.border_light)

    # ─── С чего начать - первые 6 задач (report_priorities) ─────────
    top_row = nav_hdr + len(nav_items) + 1
    ws1.row_dimensions[top_row].height = 22
    ws1.merge_cells(f'B{top_row}:H{top_row}')
    c = ws1[f'B{top_row}']
    c.value = (f'С чего начать - первые задачи (критичных задач всего: '
              f'{_critical_tasks})' if _tasks else
              'С чего начать - критичных задач не найдено')
    c.font = _font(size=10, bold=True, color=C.text_muted)
    c.alignment = _align()

    _top6 = _tasks[:6]
    _TOP_FILL = {1: C.err_soft, 2: C.warn_soft, 3: C.surface}
    if _top6:
        top_hdr = top_row + 1
        for col, title in (('B', '№'), ('C', 'Задача'), ('D', 'Почему сейчас'),
                           ('G', 'Кому')):
            cell = ws1[f'{col}{top_hdr}']
            cell.value = title
            cell.font = _font(size=9, bold=True, color=C.bg_elev)
            cell.fill = _fill(C.accent)
            cell.border = _border()
            cell.alignment = _align()
        ws1.merge_cells(f'D{top_hdr}:F{top_hdr}')
        ws1.merge_cells(f'G{top_hdr}:H{top_hdr}')
        ws1.row_dimensions[top_hdr].height = 20
        _c_w = ws1.column_dimensions['C'].width or 17
        _dg_w = sum(ws1.column_dimensions[c_].width or 0 for c_ in 'DEF')
        _gh_w = sum(ws1.column_dimensions[c_].width or 0 for c_ in 'GH')
        for i, t in enumerate(_top6, 1):
            r = top_hdr + i
            ws1.row_dimensions[r].height = max(
                _row_height_for(t.title, _c_w),
                _row_height_for(t.what, _dg_w),
                _row_height_for(t.owner, _gh_w))
            num = ws1[f'B{r}']
            num.value = i
            num.font = _font(size=10, bold=True)
            num.alignment = _align()
            num.fill = _fill(_TOP_FILL.get(t.priority, C.surface))
            num.border = _border(color=C.border_light)

            task_c = ws1[f'C{r}']
            task_c.value = t.title
            task_c.font = _font(size=10, bold=True)
            task_c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            task_c.border = _border(color=C.border_light)

            ws1.merge_cells(f'D{r}:F{r}')
            why_c = ws1[f'D{r}']
            why_c.value = t.what
            why_c.font = _font(size=10, color=C.text_soft)
            why_c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            why_c.border = _border(color=C.border_light)

            ws1.merge_cells(f'G{r}:H{r}')
            owner_c = ws1[f'G{r}']
            owner_c.value = t.owner
            owner_c.font = _font(size=10, color=C.text_soft)
            owner_c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            owner_c.border = _border(color=C.border_light)

    # ─── Листы «План работ» и «Проблемы» (report_priorities) ────────
    _build_work_plan_sheet(wb, _tasks)
    # Откуда взят каждый адрес - чтобы в «Проблемах» было видно страницы,
    # попавшие в прогон из карты сайта / своего списка, а не из каталога.
    _src_by_url = {r.url: getattr(r, 'source', '') for r in results
                   if getattr(r, 'source', '')}
    _build_problems_sheet(wb, _findings, _src_by_url,
                          interlinking=interlinking_note(results))

    # ─── Лист структурной проверки (идёт сразу после «Обзора») ──────
    _build_structure_sheet(wb, results)

    # ─── Лист «Страницы» - краткая сводка всех проверенных страниц ──
    _build_pages_overview_sheet(wb, results, _findings, w3c_check)

    # ─── Лист «Хосты и аномалии» - проблемы уровня сайта/хоста ──────
    _build_hosts_anomalies_sheet(wb, service_issues, wm_metrics, link_profile,
                                 anomalies)

    # ─── Лист «Трафик и траст» - краткая сводка + ИКС/DR + линкпрофиль ──
    _build_traffic_overview_sheet(wb, traffic, trust, link_profile, gsc_pages)

    # Листы «Заголовки и мета» и «Регион и СНГ» удалены - их находки
    # (единственность title/description/H1, чужой город/телефон/почта,
    # СНГ-чистота, технический регион по гео-тегам) полностью попадают в
    # «Проблемы» через report_priorities.py (_meta_unique_findings/
    # _region_findings/_cis_findings/_geo_findings).

    # Лист «Вёрстка» удалён - находки (viewport/CSS/меню/mixed content/
    # favicon, поиск по сайту, фильтрация товаров) полностью в «Проблемы».

    # Лист «Разметка» удалён - находки полностью в «Проблемы».

    # Листы «Безопасность» и «Изображения» удалены - находки (заголовки
    # безопасности; конкретная картинка без alt/битая, дубли картинок
    # категорий/товаров) полностью попадают в «Проблемы» через
    # report_priorities.py (generic-обработчик issues/warnings и
    # _images_findings соответственно).

    # Групповой лист «Техничка» и все его секции (Индексация, Метаданные,
    # Ошибки JavaScript, Валидация и скорость, Страница 404, Дубли главной,
    # Индексация (Арсенкин), Фильтры ПС, Нагрузка и парсинг) удалены: их
    # находки целиком собраны в «Проблемы» - там таблица с автофильтром, где
    # видно и уровень, и раздел, и адрес. Всё, что было НЕ находкой (✅-строки,
    # счётчики, замеры), либо ушло хвостом в колонку «Как исправить», либо
    # убрано как справочный шум. Сборщики: indexing_site_findings,
    # url_format_findings, robots_hygiene_findings, content_sections_findings,
    # metadata_site_findings, _console_findings, ux_interactive_findings,
    # w3c_findings, static_delivery_findings, page404_findings,
    # home_dupes_findings, arsenkin_findings, ps_filters_findings,
    # stress_check_findings. Перелинковка (вывод по сайту, а не по странице) -
    # блоком под таблицей «Проблемы» через interlinking_note.

    # Лист «Страницы в ГСК» удалён - секция на «Трафик и траст».

    # Лист «404 в индексе» удалён - находки уже в «Проблемы».

    # Листы «Ссылочный профиль» и «Траст проекта» удалены - обе метрики
    # (не находки) теперь секциями на листе «Трафик и траст».
    _build_uniqueness_sheet(wb, uniqueness)

    # ─── Лист «Настройки в админке» - если проверка выполнялась ────────
    _build_admin_settings_sheet(wb, admin_settings)

    # ─── Лист «Я.Бизнес/GMB» + единая проверка отзывов (приоритет докупки) ──
    _build_yabusiness_sheet(wb, yabusiness, review_priority)

    # Лист «Динамика трафика» удалён - сводка трафика уже на «Трафик и траст».

    # Листы «Контакты по городам» и «Замена рекл. номера» удалены: находки
    # (конкретное поле/город/что не так) полностью попадают в «Проблемы»
    # через report_priorities.py (_kp_findings/_calltracking_findings) -
    # полная таблица по каждому городу (даже без проблем) не нужна.

    # Лист «Все детали» удалён - «Битые переменные» и «Откуда перешли»
    # теперь колонки на «Страницы» (там же «Найдено проблем»), а не
    # отдельный дублирующий лист со списком тех же страниц.

    # ═══════════════════════════════════════════════════════════════
    # Лист «Битые тексты» удалён - находки полностью в «Проблемы» через
    # report_priorities._text_issue_findings().

    # Лист «404 из Метрики» удалён - находки (с пометкой, подтверждено ли
    # обходом сайта или только по Метрике) теперь в «Проблемы», раздел
    # «404 в индексе» (см. _metrika_404_findings в report_priorities.py).


    # ═══════════════════════════════════════════════════════════════
    # ЛИСТ 5: Уведомления (письма Вебмастера/GSC/Я.Бизнеса/2ГИС/Google)
    # ═══════════════════════════════════════════════════════════════
    # notifications=None - сбор уведомлений был ВЫКЛЮЧЕН, листа нет.
    # notifications=[] - сбор включён, писем нет: лист с заглушкой
    # («проверено, писем нет» - это результат, а не отсутствие проверки).
    if notifications is not None:
        _build_notifications_sheet(wb, notifications)

    # ЛИСТ: «Ошибки сервисов» - находки прямо из сервисов (не из почты),
    # детали к агрегированным задачам «Разобрать проблемы в сервисах» в
    # «Плане работ».
    _build_service_issues_sheet(wb, service_issues)

    # ЛИСТ: «Автокликер» - итоги перекликивания ошибок (если запускался).
    _build_autoclick_sheet(wb, autoclick)

    # Фильтрация товаров - теперь секцией на листе «Вёрстка» (см. выше).

    # ── Пересборка детальных листов в 7 тематических групп ──────────
    # (Техничка / Верстка / КП / Формы / Админка / Аналитика / Контент)
    _regroup_into_groups(wb)

    # ── Сохраняем ──────────────────────────────────────────────────
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# ── Утилита для имени файла ─────────────────────────────────────────


def make_report_filename(project_id: str, started_at_ms: int, reports_dir: Path) -> str:
    """
    Имя файла: smu-21.05.2026.xlsx
    Если уже есть - smu-21.05.2026_2.xlsx, _3 и т.д.
    """
    d = datetime.fromtimestamp(started_at_ms / 1000)
    date_part = d.strftime('%d.%m.%Y')
    prefix = f'{project_id}-{date_part}'

    base_name = f'{prefix}.xlsx'
    if not (reports_dir / base_name).exists():
        return base_name

    n = 2
    while (reports_dir / f'{prefix}_{n}.xlsx').exists():
        n += 1
    return f'{prefix}_{n}.xlsx'
