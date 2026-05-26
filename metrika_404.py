"""
metrika_404.py — загрузка и парсинг 404-отчётов из почты Яндекс.Метрики.

Письма приходят от Яндекс.Метрики (devnull@yandex.ru) в специальную папку.
Каждый день — 7-8 писем (по одному на страну: АЗ, РБ, УЗ, АМ, КЗ, РФ, КГ + АЗ-перевод).
В каждом письме 2 xlsx: «таблица» (URL → визиты) и «график» (распределение по часам).
Нам интересна только «таблица».

Структура xlsx «таблица»:
  Строка 1: метаданные периода
  Строка 2: фильтры
  Строка 3: атрибуция
  Строка 4: пусто
  Строка 5: заголовки — обычно «Заголовок страницы | Просмотры | Посетители»
  Строка 6+: данные

Хранилище: cache/metrika-404/{project_id}/{country}/{YYYY-MM-DD}.json
"""
import base64
import email
import imaplib
import json
import re
import socket
import ssl
from dataclasses import dataclass, asdict
from datetime import datetime
from email.header import decode_header
from io import BytesIO
from pathlib import Path
from typing import Optional, Callable
from urllib.parse import urlparse

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).parent
CACHE_DIR = PROJECT_ROOT / 'cache' / 'metrika-404'


# ── Конфиг почтовых ящиков по проектам ───────────────────────────────


# IMAP-сервер Яндекса (один для всех ящиков @yandex.ru)
YANDEX_IMAP_HOST = 'imap.yandex.ru'
YANDEX_IMAP_PORT = 993

# По каждому проекту: с какого ящика и из какой папки забирать письма.
# Логин и пароль приложения хранятся в Streamlit Secrets:
#   metrika_smu_email = "stalmetural19@yandex.ru"
#   metrika_smu_password = "пароль приложения"
# Имена папок в Яндекс-почте — на русском, важно сохранять кириллицу как есть.
MAILBOX_CONFIG = {
    'smu': {
        'folder': 'Я.Метрика 404 и др',
        'secret_email': 'metrika_smu_email',
        'secret_password': 'metrika_smu_password',
    },
    # Когда добавим ИМП и МПЭ — просто допишем сюда новые ключи
}


# Какие страны/регионы ожидаются в письмах (по проектам).
# Используется чтобы определить страну из темы письма.
# Темы вида: «Отчёт «АЗ 404 отчет» за 25.05.2026»
# Ключ — обозначение в теме, значение — человекочитаемое имя.
COUNTRY_LABELS = {
    'РФ': 'Россия',
    'КЗ': 'Казахстан',
    'РБ': 'Беларусь',
    'УЗ': 'Узбекистан',
    'АЗ': 'Азербайджан',
    'АМ': 'Армения',
    'КГ': 'Кыргызстан',
    'АЗ (перевод)': 'Азербайджан (перевод)',
}


# ── Структуры данных ────────────────────────────────────────────────


@dataclass
class Page404:
    """Одна 404-страница из отчёта Метрики."""
    page_title: str              # «Страница не найдена | Стальметурал»
    page_url: Optional[str]      # URL — из колонки «Адрес страницы» или из заголовка
    views: int                   # просмотры
    visitors: int                # уникальные посетители
    referer: Optional[str] = None  # откуда пришли (колонка «Реферер»)


@dataclass
class Report404:
    """Один отчёт за день по одной стране."""
    project_id: str          # smu / imp / mpe
    country_code: str        # РФ / КЗ / ...
    country_name: str        # Россия / Казахстан / ...
    report_date: str         # YYYY-MM-DD
    received_at: str         # ISO datetime когда забрали из почты
    pages: list[Page404]
    total_views: int
    total_pages: int

    def to_dict(self) -> dict:
        d = asdict(self)
        # dataclass-список → dict-список вручную
        d['pages'] = [asdict(p) for p in self.pages]
        return d


# ── Регулярки для парсинга темы письма ──────────────────────────────


# «Отчёт «АЗ 404 отчет» за 25.05.2026» → ('АЗ', '25.05.2026')
# Также «Отчёт «АЗ (перевод) 404 отчет» за 25.05.2026»
SUBJECT_RE = re.compile(
    r'Отч[её]т\s*[«"]([\wА-Яа-я\s()]+?)\s+404\s+отч[её]т[»"]\s+за\s+(\d{2}\.\d{2}\.\d{4})',
    re.IGNORECASE,
)


def parse_subject(subject: str) -> Optional[dict]:
    """Из темы письма достать страну и дату.
    Возвращает {'country': 'АЗ', 'date': '2026-05-25'} или None.
    """
    if not subject:
        return None

    # Декодируем MIME-заголовок (может прийти как =?utf-8?B?...?=)
    decoded_parts = decode_header(subject)
    decoded = ''
    for part, enc in decoded_parts:
        if isinstance(part, bytes):
            decoded += part.decode(enc or 'utf-8', errors='replace')
        else:
            decoded += part

    m = SUBJECT_RE.search(decoded)
    if not m:
        return None

    country = m.group(1).strip()
    date_str = m.group(2)  # 25.05.2026
    # Переводим в YYYY-MM-DD для удобства сортировки
    dd, mm, yyyy = date_str.split('.')
    iso_date = f'{yyyy}-{mm}-{dd}'
    return {'country': country, 'date': iso_date}


# ── Парсинг xlsx-вложения «таблица» ─────────────────────────────────


# Сигнатура: «таблица» в имени файла. У Метрики имя файла всегда содержит «таблица»
# (а второй файл — «график»).
def is_table_attachment(filename: str) -> bool:
    if not filename:
        return False
    fn_lower = filename.lower()
    return 'таблица' in fn_lower and fn_lower.endswith('.xlsx')


def parse_table_xlsx(xlsx_bytes: bytes) -> list[Page404]:
    """Распарсить xlsx-«таблицу» отчёта Метрики в список Page404."""
    try:
        wb = load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f'Не удалось открыть xlsx: {e}')

    # Берём первый лист (у Метрики всегда один лист «Отчет»)
    ws = wb[wb.sheetnames[0]]

    # Ищем строку с заголовками — она должна точно соответствовать одному из вариантов:
    # «Заголовок страницы», «URL», «Адрес страницы», и т.п. В соседних колонках —
    # «Просмотры»/«Посетители»/«Визиты». То есть строка с НЕСКОЛЬКИМИ короткими ячейками.
    header_row_idx = None
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 20:
            break
        cells = [str(c).strip() if c else '' for c in row]
        # Считаем «заголовочной» только если:
        # 1) Есть колонка точно похожая на название страницы
        # 2) И ещё есть колонка похожая на метрику (просмотры/посетители/визиты)
        has_title_col = any(
            c.lower() in ('заголовок страницы', 'url', 'адрес страницы', 'страница')
            or c.lower().startswith('заголовок страниц')
            for c in cells
        )
        has_metric_col = any(
            'просмотр' in c.lower() or 'посетит' in c.lower() or 'визит' in c.lower()
            for c in cells
        )
        if has_title_col and has_metric_col:
            header_row_idx = i
            headers = cells
            break

    if header_row_idx is None:
        return []  # Нет заголовков — нет и данных

    # Определяем индексы колонок гибко
    title_idx = None
    url_idx = None
    views_idx = None
    visitors_idx = None
    referer_idx = None
    for idx, h in enumerate(headers):
        h_lower = h.lower()
        if h_lower in ('адрес страницы', 'url'):
            if url_idx is None:
                url_idx = idx
        elif 'заголовок' in h_lower or h_lower == 'страница':
            if title_idx is None:
                title_idx = idx
        elif 'просмотр' in h_lower:
            views_idx = idx
        elif 'посетит' in h_lower:
            visitors_idx = idx
        elif 'реферер' in h_lower or 'переход' in h_lower or 'источник' in h_lower:
            referer_idx = idx

    # Минимум — должны быть либо title, либо url
    if title_idx is None and url_idx is None:
        return []

    # Читаем данные начиная со следующей строки после заголовков
    pages = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), 1):
        if all(c is None or str(c).strip() == '' for c in row):
            continue

        # Если первая ячейка — «Итого и средние», это сводная строка, пропускаем
        first_cell = str(row[0]).strip().lower() if row[0] else ''
        if 'итого' in first_cell:
            continue

        title = str(row[title_idx]).strip() if title_idx is not None and title_idx < len(row) and row[title_idx] else ''
        url_from_col = str(row[url_idx]).strip() if url_idx is not None and url_idx < len(row) and row[url_idx] else ''
        if not title and not url_from_col:
            continue

        views = _to_int(row[views_idx]) if views_idx is not None and views_idx < len(row) else 0
        visitors = _to_int(row[visitors_idx]) if visitors_idx is not None and visitors_idx < len(row) else 0
        referer = None
        if referer_idx is not None and referer_idx < len(row) and row[referer_idx]:
            ref_str = str(row[referer_idx]).strip()
            # «Не определен» — это пусто
            if ref_str and ref_str.lower() != 'не определен':
                referer = ref_str

        # URL: сначала из явной колонки, потом из заголовка
        url = url_from_col or _extract_url(title)

        pages.append(Page404(
            page_title=title or (url or ''),
            page_url=url or None,
            views=views,
            visitors=visitors,
            referer=referer,
        ))

    return pages


def _to_int(v) -> int:
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace(' ', '').replace(',', '.')
    try:
        return int(float(s))
    except ValueError:
        return 0


URL_IN_TEXT_RE = re.compile(r'https?://[^\s|]+')


def _extract_url(text: str) -> Optional[str]:
    """Найти URL в строке (если он там есть)."""
    if not text:
        return None
    m = URL_IN_TEXT_RE.search(text)
    return m.group(0) if m else None


# ── IMAP-клиент ─────────────────────────────────────────────────────


def _connect_via_http_proxy(
    proxy_url: str,
    target_host: str,
    target_port: int,
    timeout: int = 30,
) -> socket.socket:
    """
    Установить TCP-соединение к target_host:target_port через HTTP-прокси
    используя метод CONNECT. Возвращает голый сокет, поверх которого
    можно навернуть SSL.
    
    proxy_url: http://user:pass@host:port
    """
    p = urlparse(proxy_url)
    proxy_host = p.hostname
    proxy_port = p.port or 8080
    
    # Готовим CONNECT-запрос
    connect_line = f'CONNECT {target_host}:{target_port} HTTP/1.1'
    headers = [
        connect_line,
        f'Host: {target_host}:{target_port}',
        'User-Agent: Mozilla/5.0 (site-checker)',
        'Proxy-Connection: keep-alive',
    ]
    if p.username and p.password:
        creds = f'{p.username}:{p.password}'
        token = base64.b64encode(creds.encode()).decode()
        headers.append(f'Proxy-Authorization: Basic {token}')
    request = '\r\n'.join(headers) + '\r\n\r\n'

    sock = socket.create_connection((proxy_host, proxy_port), timeout=timeout)
    sock.sendall(request.encode())

    # Читаем ответ — должен быть "HTTP/1.1 200 ..."
    response = b''
    while b'\r\n\r\n' not in response:
        chunk = sock.recv(4096)
        if not chunk:
            sock.close()
            raise ConnectionError(f'Прокси {proxy_host}:{proxy_port} закрыл соединение без ответа')
        response += chunk
        if len(response) > 16384:
            sock.close()
            raise ConnectionError('Слишком длинный ответ от прокси')

    status_line = response.split(b'\r\n', 1)[0].decode('utf-8', errors='replace')
    if ' 200 ' not in status_line:
        sock.close()
        raise ConnectionError(
            f'Прокси отказал в CONNECT к {target_host}:{target_port}. '
            f'Ответ: {status_line}. Возможно, прокси не разрешает '
            f'порт {target_port} (только 443/HTTPS).'
        )
    return sock


# ── IMAP UTF-7 (modified) кодек ──────────────────────────────────────
# Имена папок на IMAP-серверах кодируются специальной кодировкой
# IMAP UTF-7 modified (RFC 3501 §5.1.3). Это не обычный UTF-7:
#   • Символ "&" заменяется на "&-"
#   • Не-ASCII символы кодируются Base64 (без padding, с '+' и '/' заменёнными на '+' и ',')
#     и обрамляются "&" ... "-"


def _imap_utf7_encode(s: str) -> bytes:
    """Кодировать строку в IMAP UTF-7 modified для использования в IMAP-командах."""
    if not s:
        return b''
    res = []
    i = 0
    while i < len(s):
        ch = s[i]
        if ch == '&':
            res.append(b'&-')
            i += 1
        elif 0x20 <= ord(ch) <= 0x7E:
            res.append(ch.encode('ascii'))
            i += 1
        else:
            # Накапливаем все подряд не-ASCII символы
            j = i
            while j < len(s) and (ord(s[j]) < 0x20 or ord(s[j]) > 0x7E) and s[j] != '&':
                j += 1
            chunk = s[i:j]
            # UTF-16BE → Base64 → заменить '/' на ','
            b16 = chunk.encode('utf-16-be')
            b64 = base64.b64encode(b16).rstrip(b'=').replace(b'/', b',')
            res.append(b'&' + b64 + b'-')
            i = j
    return b''.join(res)


def _imap_utf7_decode(b: bytes) -> str:
    """Декодировать IMAP UTF-7 modified обратно в обычную строку."""
    if isinstance(b, str):
        b = b.encode('ascii', errors='replace')
    res = []
    i = 0
    while i < len(b):
        ch = b[i:i+1]
        if ch == b'&':
            # Ищем закрывающий '-'
            end = b.find(b'-', i + 1)
            if end < 0:
                # Битая последовательность — оставим как есть
                res.append(b[i:].decode('ascii', errors='replace'))
                break
            seq = b[i+1:end]
            if not seq:
                # "&-" → литерал '&'
                res.append('&')
            else:
                # Base64: '/' было заменено на ',', возвращаем обратно, паддинг добавляем
                b64 = seq.replace(b',', b'/')
                padding = b'=' * (-len(b64) % 4)
                try:
                    raw = base64.b64decode(b64 + padding)
                    res.append(raw.decode('utf-16-be'))
                except Exception:
                    res.append(seq.decode('ascii', errors='replace'))
            i = end + 1
        else:
            res.append(ch.decode('ascii', errors='replace'))
            i += 1
    return ''.join(res)


class IMAP4_SSL_via_Proxy(imaplib.IMAP4_SSL):
    """
    IMAP4_SSL который сначала идёт через HTTP-прокси (CONNECT),
    потом наворачивает SSL на этот сокет.

    Переопределяем только _create_socket — он, согласно imaplib, должен
    вернуть готовый SSL-сокет. Метод open() родительского класса дальше
    сам обернёт его в makefile() как полагается. Не трогаем self.file —
    в Python 3.12+ это property без сеттера.
    """
    def __init__(self, host, port, proxy_url, ssl_context=None, timeout=30):
        self._proxy_url = proxy_url
        self._connect_timeout = timeout
        self._custom_ssl_context = ssl_context or ssl.create_default_context()
        super().__init__(host, port, ssl_context=ssl_context, timeout=timeout)

    def _create_socket(self, timeout=None):
        # 1. Поднимаем голый TCP-сокет к Яндексу через CONNECT на прокси
        raw_sock = _connect_via_http_proxy(
            self._proxy_url, self.host, self.port,
            timeout=self._connect_timeout,
        )
        # 2. Заворачиваем в SSL — родителю отдаём готовый ssl-сокет
        ssl_sock = self._custom_ssl_context.wrap_socket(
            raw_sock, server_hostname=self.host,
        )
        return ssl_sock



def fetch_metrika_emails(
    project_id: str,
    email_addr: str,
    password: str,
    folder: str,
    *,
    since_days: int = 30,
    log: Optional[Callable] = None,
    progress: Optional[Callable] = None,
    proxy_url: Optional[str] = None,
) -> list[Report404]:
    """
    Подключиться по IMAP, скачать новые письма из указанной папки,
    распарсить вложения и вернуть список отчётов Report404.

    since_days — забираем письма не старше N дней (по умолчанию 30).
    proxy_url — если задан, IMAP-соединение пойдёт через HTTP CONNECT-прокси.
                Нужно когда основной хостинг (Streamlit Cloud в США) блокируется
                Яндексом. Прокси должен разрешать CONNECT на порт 993.
    """
    reports = []

    if log:
        log('info', f'Подключаюсь к {YANDEX_IMAP_HOST}:{YANDEX_IMAP_PORT} как {email_addr}…')
        if proxy_url:
            p = urlparse(proxy_url)
            log('info', f'IMAP через прокси {p.hostname}:{p.port}')

    ssl_ctx = ssl.create_default_context()

    # Создаём IMAP-клиент: через прокси или напрямую
    if proxy_url:
        M = IMAP4_SSL_via_Proxy(
            YANDEX_IMAP_HOST, YANDEX_IMAP_PORT,
            proxy_url=proxy_url,
            ssl_context=ssl_ctx,
            timeout=60,
        )
    else:
        M = imaplib.IMAP4_SSL(YANDEX_IMAP_HOST, YANDEX_IMAP_PORT, ssl_context=ssl_ctx, timeout=60)

    try:
        try:
            M.login(email_addr, password)
        except imaplib.IMAP4.error as e:
            raise PermissionError(
                f'Не удалось войти в почту: {e}. '
                f'Проверьте: 1) email и пароль приложения в Streamlit Secrets; '
                f'2) что IMAP включён в Яндекс-почте '
                f'(mail.yandex.ru → Все настройки → Почтовые программы).'
            )

        if log:
            log('info', 'Логин успешен. Запрашиваю список папок…')

        # Папки на Яндексе с русскими именами требуют кодировки IMAP UTF-7,
        # но самый надёжный способ — заключить имя в кавычки.
        # Сначала найдём папку среди списка
        try:
            status, folders = M.list()
        except Exception as e:
            raise RuntimeError(f'M.list() упало: {type(e).__name__}: {e}')

        if log:
            log('info', f'Получено папок: {len(folders) if folders else 0}')

        target_folder_encoded = None
        if status == 'OK' and folders:
            for f in folders:
                # В ответе строка типа: b'(\\HasNoChildren) "|" "&BB8-.&BBwAVA..."'
                # Имя папки уже в IMAP UTF-7 (если содержит кириллицу)
                try:
                    line = f.decode('ascii', errors='replace') if isinstance(f, bytes) else f
                except Exception:
                    continue
                # Имя в конце строки — берём из последних кавычек
                m = re.search(r'"([^"]+)"\s*$', line)
                if not m:
                    continue
                name_encoded = m.group(1)
                # Декодируем из IMAP UTF-7 чтобы сравнить с тем что ищем
                try:
                    name_decoded = _imap_utf7_decode(name_encoded.encode('ascii'))
                except Exception:
                    name_decoded = name_encoded
                if name_decoded == folder:
                    target_folder_encoded = name_encoded
                    if log:
                        log('info', f'Папка найдена в листинге: {name_encoded}')
                    break

        if target_folder_encoded is None:
            # Папка не нашлась в листинге — кодируем имя сами
            target_folder_encoded = _imap_utf7_encode(folder).decode('ascii')
            if log:
                log('warn', f'Папка не нашлась в листинге. Использую закодированное имя: {target_folder_encoded}')

        if log:
            log('info', f'Открываю папку «{folder}» (IMAP-имя: {target_folder_encoded})…')

        # Кодируем имя папки В BYTES сразу — imaplib попытается сделать bytes(arg, 'ascii')
        # и упадёт на любом не-ASCII. Передавая bytes, обходим эту проверку.
        # target_folder_encoded уже содержит только ASCII (это IMAP UTF-7), но
        # обернём для надёжности.
        folder_bytes = target_folder_encoded.encode('ascii', errors='replace')
        select_arg = b'"' + folder_bytes + b'"'

        try:
            status, _ = M.select(select_arg, readonly=True)
        except Exception as e:
            raise RuntimeError(f'M.select({select_arg!r}) упало: {type(e).__name__}: {e}')

        if status != 'OK':
            raise FileNotFoundError(f'Папка «{folder}» не найдена в почте')

        if log:
            log('info', 'Папка открыта. Ищу письма…')

        # Ищем все письма за последние N дней от Яндекс.Метрики
        from datetime import datetime, timedelta
        since_date = (datetime.now() - timedelta(days=since_days)).strftime('%d-%b-%Y')
        # IMAP-команда для поиска (только ASCII символы — даты и yandex.ru)
        criteria_str = f'(SINCE "{since_date}" FROM "yandex.ru")'
        criteria_bytes = criteria_str.encode('ascii')

        try:
            status, data = M.search(None, criteria_bytes)
        except Exception as e:
            raise RuntimeError(f'M.search() упало: {type(e).__name__}: {e}')

        if status != 'OK':
            if log:
                log('warn', 'IMAP search вернул ошибку')
            return reports

        msg_ids = data[0].split()
        if log:
            log('info', f'Найдено писем за последние {since_days} дней: {len(msg_ids)}')

        for i, msg_id in enumerate(msg_ids):
            if progress:
                progress(i + 1, len(msg_ids))

            status, msg_data = M.fetch(msg_id, '(RFC822)')
            if status != 'OK':
                continue

            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)

            # Парсим тему: страна + дата
            subj_info = parse_subject(msg.get('Subject', ''))
            if not subj_info:
                continue  # Не наше письмо

            # Достаём xlsx-«таблицу»
            table_pages = None
            for part in msg.walk():
                if part.get_content_maintype() == 'multipart':
                    continue
                fname = part.get_filename()
                if fname:
                    # Декодируем имя файла из MIME
                    decoded = decode_header(fname)
                    fname_decoded = ''.join(
                        p.decode(enc or 'utf-8', errors='replace') if isinstance(p, bytes) else p
                        for p, enc in decoded
                    )
                    if is_table_attachment(fname_decoded):
                        payload = part.get_payload(decode=True)
                        if payload:
                            try:
                                table_pages = parse_table_xlsx(payload)
                            except Exception as e:
                                if log:
                                    log('warn', f'Не удалось распарсить {fname_decoded}: {e}')
                            break

            if table_pages is None:
                continue  # Не нашли «таблицу» в этом письме

            country_code = subj_info['country']
            country_name = COUNTRY_LABELS.get(country_code, country_code)

            report = Report404(
                project_id=project_id,
                country_code=country_code,
                country_name=country_name,
                report_date=subj_info['date'],
                received_at=datetime.now().isoformat(),
                pages=table_pages,
                total_views=sum(p.views for p in table_pages),
                total_pages=len(table_pages),
            )
            reports.append(report)

        if log:
            log('info', f'Успешно распарсено отчётов: {len(reports)}')

    finally:
        try:
            M.logout()
        except Exception:
            pass

    return reports


# ── Хранилище ──────────────────────────────────────────────────────


def report_storage_path(project_id: str, country_code: str, report_date: str) -> Path:
    """Где лежит JSON одного отчёта."""
    safe_country = re.sub(r'[^\w-]', '_', country_code)
    return CACHE_DIR / project_id / safe_country / f'{report_date}.json'


def save_report(report: Report404) -> Path:
    """Сохранить отчёт в JSON."""
    path = report_storage_path(report.project_id, report.country_code, report.report_date)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(report.to_dict(), f, ensure_ascii=False, indent=2)
    return path


def load_report(project_id: str, country_code: str, report_date: str) -> Optional[Report404]:
    """Прочитать отчёт из JSON. Возвращает None если файла нет."""
    path = report_storage_path(project_id, country_code, report_date)
    if not path.exists():
        return None
    with open(path, 'r', encoding='utf-8') as f:
        d = json.load(f)
    pages = [Page404(**p) for p in d.pop('pages')]
    return Report404(**d, pages=pages)


def list_stored_reports(project_id: str) -> list[dict]:
    """
    Список всех сохранённых отчётов для проекта.
    Возвращает [{'country': 'РФ', 'date': '2026-05-25', 'total_pages': 12, 'total_views': 45}, ...].
    """
    project_dir = CACHE_DIR / project_id
    if not project_dir.exists():
        return []
    result = []
    for country_dir in sorted(project_dir.iterdir()):
        if not country_dir.is_dir():
            continue
        country = country_dir.name
        for file in sorted(country_dir.glob('*.json'), reverse=True):
            try:
                with open(file, 'r', encoding='utf-8') as f:
                    d = json.load(f)
                result.append({
                    'country_code': d['country_code'],
                    'country_name': d.get('country_name', d['country_code']),
                    'date': d['report_date'],
                    'total_pages': d.get('total_pages', 0),
                    'total_views': d.get('total_views', 0),
                    'file': str(file),
                })
            except Exception:
                continue
    # Сортируем: сначала свежие даты, в пределах даты — по алфавиту страны
    result.sort(key=lambda r: (r['date'], r['country_code']), reverse=True)
    return result


def save_reports_batch(reports: list[Report404]) -> int:
    """Сохранить пачку отчётов, возвращает число НОВЫХ (не перезаписанных)."""
    new_count = 0
    for r in reports:
        path = report_storage_path(r.project_id, r.country_code, r.report_date)
        is_new = not path.exists()
        save_report(r)
        if is_new:
            new_count += 1
    return new_count


# ── Высокоуровневые операции ────────────────────────────────────────


def get_stored_dates(project_id: str) -> set[str]:
    """Множество дат (YYYY-MM-DD), за которые в кеше есть хотя бы один отчёт."""
    return {r['date'] for r in list_stored_reports(project_id)}


def fetch_incremental(
    project_id: str,
    email_addr: str,
    password: str,
    folder: str,
    *,
    proxy_url: Optional[str] = None,
    lookback_days: int = 3,
    log: Optional[Callable] = None,
    progress: Optional[Callable] = None,
) -> dict:
    """
    Инкрементальная загрузка: идём в IMAP только за последние lookback_days дней,
    парсим только письма за даты, которых ещё нет в кеше, сразу сохраняем.

    Возвращает {'fetched': N_новых_отчётов, 'skipped': N_уже_было, 'errors': N_ошибок}.
    """
    existing_dates = get_stored_dates(project_id)

    reports = fetch_metrika_emails(
        project_id=project_id,
        email_addr=email_addr,
        password=password,
        folder=folder,
        since_days=lookback_days,
        log=log,
        progress=progress,
        proxy_url=proxy_url,
    )

    fetched = 0
    skipped = 0
    for r in reports:
        # Проверяем: нет ли уже такого отчёта (по дате + стране)
        path = report_storage_path(r.project_id, r.country_code, r.report_date)
        if path.exists():
            skipped += 1
        else:
            save_report(r)
            fetched += 1

    return {'fetched': fetched, 'skipped': skipped, 'total_in_letters': len(reports)}


def get_latest_available_date(project_id: str) -> Optional[str]:
    """Самая свежая дата отчёта, которая есть в кеше. None если кеш пустой."""
    dates = get_stored_dates(project_id)
    if not dates:
        return None
    return max(dates)


def load_reports_for_date(project_id: str, target_date: str) -> list[Report404]:
    """Все отчёты (по странам) за одну конкретную дату."""
    reports = []
    listing = list_stored_reports(project_id)
    for r in listing:
        if r['date'] != target_date:
            continue
        rep = load_report(project_id, r['country_code'], target_date)
        if rep:
            reports.append(rep)
    return reports


def load_reports_for_period(
    project_id: str,
    days: int,
) -> list[Report404]:
    """Все отчёты за последние N дней (включая сегодня)."""
    from datetime import datetime as _dt, timedelta as _td
    today = _dt.now().date()
    cutoff = today - _td(days=days - 1)

    reports = []
    listing = list_stored_reports(project_id)
    for r in listing:
        try:
            d = _dt.strptime(r['date'], '%Y-%m-%d').date()
        except ValueError:
            continue
        if d >= cutoff:
            rep = load_report(project_id, r['country_code'], r['date'])
            if rep:
                reports.append(rep)
    # Сортируем: свежие сверху, потом по стране
    reports.sort(key=lambda r: (r.report_date, r.country_code), reverse=True)
    return reports

