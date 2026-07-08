"""
Проверка админки Bitrix «Заказы» (sale_order.php) - пункт чек-листа 2.11
(в части «оформления с корзины приходят в админку»).

Идея та же, что у admin_check для форм: после оформления тест-заказа через
корзину тест заходит в админку, открывает список «Заказы» и проверяет, что НАШ
заказ реально там появился. Формы проверяет admin_check («Уведомления с форм»),
а заказы - здесь («Заказы»); это разные разделы админки, но один логин/пароль.

Сопоставление - по ДАТЕ-ВРЕМЕНИ СОЗДАНИЯ заказа (в списке видны только «Дата
создания» до секунды и «№ заказа», ФИО/почты покупателя там нет). Момент
оформления мы знаем из placed_orders.json (ts), заказ создаётся ровно тогда же -
берём заказ с ближайшим временем в узком окне. Логин/пароль и вход
переиспользуем из admin_check (тот же Bitrix).
"""
import re
from datetime import datetime
from pathlib import Path

# Переиспользуем вход/креды/детектор логина у форм-проверки (тот же Bitrix).
import admin_check as _ac


def построить_url_списка(домен: str, размер: int = 100) -> str:
    """URL списка «Заказы» (свежие сверху по умолчанию). SIZEN_1 побольше, чтобы
    точно захватить наш только что оформленный заказ на первой странице."""
    домен = домен.rstrip("/")
    return (f"{домен}/bitrix/admin/sale_order.php?lang=ru"
            f"&PAGEN_1=1&SIZEN_1={int(размер)}")


_DT_RE = re.compile(r"(\d{2}\.\d{2}\.\d{4})\s+(\d{2}:\d{2}:\d{2})")
_ID_RE = re.compile(r"№\s*(\d+)")


def _текст(html: str) -> str:
    html = (html or "").replace("&nbsp;", " ")
    html = re.sub(r"<[^>]+>", " ", html)
    return re.sub(r"\s+", " ", html).strip()


def разобрать_заказы(html: str) -> list:
    """Разбирает таблицу «Заказы» по строкам adm-list-table-row.
    Возвращает [{id, дата, время, дата_время}]. Устойчиво к вложенным таблицам:
    дату-время и № берём регулярками по всей строке (а не по позиции ячейки)."""
    заказы = []
    строки = re.findall(
        r'<tr[^>]*class="[^"]*adm-list-table-row[^"]*"[^>]*>(.*?)</tr>',
        html, re.S | re.I)
    for tr in строки:
        t = _текст(tr)
        mdt = _DT_RE.search(t)
        mid = _ID_RE.search(t)
        if not mdt:
            continue
        заказы.append({
            "id": mid.group(1) if mid else "",
            "дата": mdt.group(1),
            "время": mdt.group(2),
            "дата_время": f"{mdt.group(1)} {mdt.group(2)}",
        })
    return заказы


def _parse_iso(ts: str):
    try:
        return datetime.strptime((ts or "")[:19], "%Y-%m-%dT%H:%M:%S")
    except Exception:
        return None


def _parse_admin_dt(s: str):
    try:
        return datetime.strptime((s or "")[:19], "%d.%m.%Y %H:%M:%S")
    except Exception:
        return None


def сопоставить(заказы: list, оформленные: list, окно_назад_с: int = 300,
                окно_вперёд_с: int = 900) -> tuple:
    """Сверяет наши оформленные заказы (placed_orders) с заказами из админки по
    времени создания. Для каждого нашего заказа берём заказ из админки в окне
    [ts-5мин, ts+15мин], ближайший по времени. Каждый заказ засчитывается один раз.
    Возвращает (результаты, свободные_заказы)."""
    использованные = set()
    результаты = []
    for o in оформленные:
        ts = _parse_iso(o.get("ts", ""))
        база = {"город": o.get("город", ""), "название": o.get("название", ""),
                "ts": o.get("ts", "")}
        кандидаты = []
        for i, z in enumerate(заказы):
            if i in использованные:
                continue
            zt = _parse_admin_dt(z.get("дата_время", ""))
            if ts is None or zt is None:
                continue
            delta = (zt - ts).total_seconds()
            if -окно_назад_с <= delta <= окно_вперёд_с:
                кандидаты.append((abs(delta), i, z))
        if not кандидаты:
            результаты.append({**база, "статус": "НЕ найдено", "заказ": None,
                               "примечание": "заказ в списке «Заказы» админки не найден "
                                             "за окно после оформления"})
            continue
        кандидаты.sort(key=lambda p: p[0])
        _, i, z = кандидаты[0]
        использованные.add(i)
        результаты.append({**база, "статус": "Есть в админке", "заказ": z,
                           "примечание": ""})
    свободные = [z for i, z in enumerate(заказы) if i not in использованные]
    return результаты, свободные


def записать_в_логи(excel_path: str, результаты: list) -> None:
    """Дописывает колонку «Заказ в админке» в лист «Логи» у строк оформления
    заказа (сопоставление по городу + названию сценария). По образцу
    admin_check.записать_в_логи."""
    if not результаты:
        return
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_path)
    if "Логи" not in wb.sheetnames:
        return
    ws = wb["Логи"]
    headers = [str(c.value or "").strip() for c in ws[1]]

    def col(name):
        for i, h in enumerate(headers):
            if h.lower() == name.lower():
                return i + 1
        return -1

    i_city, i_name = col("Город"), col("Название")
    i_comment = col("Комментарий")
    if i_city == -1 or i_name == -1:
        return

    i_ord = col("Заказ в админке")
    if i_ord == -1:
        # ставим после «Письмо покупателю» / «Статус в админке» / «Статус»
        anchor = col("Письмо покупателю")
        if anchor == -1:
            anchor = col("Статус в админке")
        if anchor == -1:
            anchor = col("Статус")
        i_ord = (anchor + 1) if anchor != -1 else len(headers) + 1
        ws.insert_cols(i_ord)
        hc = ws.cell(1, i_ord, "Заказ в админке")
        hc.font = Font(bold=True)
        hc.fill = PatternFill("solid", fgColor="EEF3FB")
        ws.column_dimensions[get_column_letter(i_ord)].width = 18
        headers = [str(c.value or "").strip() for c in ws[1]]
        i_comment = col("Комментарий")

    def norm(v):
        return _ac._norm(str(v or ""))

    строки = {}
    for r in range(2, ws.max_row + 1):
        key = (norm(ws.cell(r, i_city).value), norm(ws.cell(r, i_name).value))
        строки.setdefault(key, []).append(r)
    занято = set()

    for res in результаты:
        key = (norm(res.get("город", "")), norm(res.get("название", "")))
        rows = [r for r in строки.get(key, []) if r not in занято]
        if not rows:
            continue
        r = rows[0]
        занято.add(r)
        z = res.get("заказ")
        есть = str(res.get("статус", "")).startswith("Есть")
        cell = ws.cell(r, i_ord, "Есть в админке" if есть else "НЕ найдено")
        cell.font = Font(color="1E8E3E" if есть else "C62828", bold=True)
        if i_comment != -1:
            деталь = (f"заказ №{z['id']} в админке ({z.get('время', '')})"
                      if (есть and z) else (res.get("примечание", "") or "заказ не найден"))
            if деталь:
                prev = str(ws.cell(r, i_comment).value or "").strip()
                ws.cell(r, i_comment, (prev + "; " if prev else "") + деталь)

    wb.save(excel_path)


def выполнить_проверку(проект_дир, зоны, orders_path: str = "placed_orders.json",
                       excel_path: str = "log_forms.xlsx", show: bool = False,
                       log=print) -> bool:
    """По каждой АДМИН-ЗОНЕ логинится, читает «Заказы» и сверяет с нашими
    оформленными заказами (placed_orders.json). Пишет колонку «Заказ в админке».
    Тихо пропускается без кредов/заказов. По образцу admin_check.выполнить_проверку."""
    import json

    creds = _ac.загрузить_креды(проект_дир)
    if not creds:
        log("ℹ️ Проверка заказов в админке пропущена: не заданы логин/пароль.")
        return False

    p = Path(orders_path)
    оформленные = []
    if p.is_file():
        try:
            оформленные = [o for o in (json.loads(p.read_text(encoding="utf-8")) or []) if o]
        except Exception:
            оформленные = []
    if not оформленные:
        log("ℹ️ Проверка заказов в админке: оформленных заказов за прогон нет.")
        return False

    if not зоны:
        зоны = [{"домен": "", "города": []}]

    # Раскладываем заказы по зонам (по нашему городу) - как admin_check.
    по_зонам = {}
    for o in оформленные:
        z = _ac._зона_отправки(зоны, o.get("город", ""))
        if z is None:
            continue
        по_зонам.setdefault(id(z), (z, []))[1].append(o)

    from playwright.sync_api import sync_playwright
    все_результаты = []
    итог_есть = итог_нет = 0

    with sync_playwright() as pw:
        b = pw.chromium.launch(headless=not show,
                               args=["--disable-blink-features=AutomationControlled"])
        ctx = b.new_context(locale="ru-RU")
        try:
            for z, з_заказы in по_зонам.values():
                домен = z.get("домен") or ""
                кратко = _ac._домен_кратко(домен) or "(основной)"
                log(f"🔎 Заказы в админке [{кратко}]: вход и чтение "
                    f"({len(з_заказы)} заказ.) …")
                page = ctx.new_page()
                html = ""
                заказы = []
                try:
                    _ac.войти(page, домен, creds["login"], creds["password"])
                    page.goto(построить_url_списка(домен),
                              wait_until="domcontentloaded", timeout=30000)
                    page.wait_for_timeout(1500)
                    html = page.content()
                    заказы = разобрать_заказы(html)
                finally:
                    try:
                        page.close()
                    except Exception:
                        pass

                if not заказы and _ac._это_форма_входа(html):
                    log(f"⚠️ Заказы [{кратко}]: не удалось войти - проверьте логин/пароль.")
                    for o in з_заказы:
                        все_результаты.append({
                            "город": o.get("город", ""), "название": o.get("название", ""),
                            "ts": o.get("ts", ""), "статус": "НЕ найдено", "заказ": None,
                            "примечание": "не удалось войти в админку (логин/пароль)"})
                        итог_нет += 1
                    continue

                if not заказы:
                    dbg = f"order_admin_debug_{кратко}.html"
                    try:
                        with open(dbg, "w", encoding="utf-8") as fh:
                            fh.write(html)
                        log(f"   ⚠️ [{кратко}] заказов не распознано - сохранил {dbg}.")
                    except Exception:
                        pass

                результаты, свободные = сопоставить(заказы, з_заказы)
                все_результаты.extend(результаты)
                е = sum(1 for r in результаты if r["статус"].startswith("Есть"))
                итог_есть += е
                итог_нет += len(результаты) - е
                log(f"   [{кратко}] заказов в админке: {len(заказы)}; "
                    f"найдено {е} из {len(результаты)}.")
        finally:
            b.close()

    try:
        записать_в_логи(excel_path, все_результаты)
    except Exception as e:  # noqa: BLE001
        log(f"⚠️ Не удалось записать результат заказов в Excel: {e}")
    log(f"✅ Проверка заказов в админке: найдено {итог_есть}, НЕ найдено {итог_нет}. "
        f"Смотри колонку «Заказ в админке» на листе «Логи».")
    return True
