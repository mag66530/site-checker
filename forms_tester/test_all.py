"""
Проверка форм и модалок на сайте: requests или Playwright.
Точка входа: run_test(). Конфиг подхватывается с диска через importlib.reload.
"""
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, unquote
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
import json
import re
from playwright.sync_api import sync_playwright

from name_format import build_test_name, cfg_enabled

# Не используем «from config import *»: run_test() перезагружает config с диска (importlib.reload).


def normalize_phone_for_submit(phone: str) -> str:
    """
    Телефон для отправки в форму: только цифры, без скобок и пробелов.
    Многие бэкенды не принимают маску «+7 (916) …» и заявка не попадает в админку.
    """
    if not phone:
        return ""
    digits = "".join(c for c in phone if c.isdigit())
    if len(digits) == 11 and digits.startswith("8"):
        return "7" + digits[1:]
    if len(digits) == 10:
        return "7" + digits
    return digits


def response_indicates_captcha_block(text: str) -> bool:
    """
    Текст ответа явно указывает на ошибку/блокировку по капче.
    Не использует подстроку «captcha» по всему HTML - в разметке часто есть recaptcha/hcaptcha в скриптах без ошибки отправки.
    """
    t = (text or "").lower()
    if "капч" in t and any(
        w in t
        for w in (
            "не пройден",
            "не прошл",
            "не пройдена",
            "введите",
            "укажите",
            "ошибк",
            "пройдите",
            "неверн",
            "error",
            "wrong",
            "invalid",
        )
    ):
        return True
    return any(
        p in t
        for p in (
            "please complete the captcha",
            "wrong captcha",
            "invalid captcha",
            "captcha is incorrect",
            "verification failed",
            "captcha_error",
        )
    )


def response_indicates_form_error(text: str) -> str:
    """Если страница ПОСЛЕ отправки/оформления показывает явную ошибку - возвращает
    её краткое описание, иначе пустую строку. Фразы специфичны (не подсказки полей),
    чтобы не было ложных срабатываний.
    """
    low = (text or "").lower()
    маркеры = [
        ("доступ запрещ", "Форма защищена reCAPTCHA (доступ запрещен)"),
        ("access denied", "Форма защищена reCAPTCHA (доступ запрещен)"),
        ("при расчете заказа произошла ошибка", "Оформление: при расчёте заказа произошла ошибка"),
        ("при расчёте заказа произошла ошибка", "Оформление: при расчёте заказа произошла ошибка"),
        ("не выбран тип плательщик", "Оформление: не выбран тип плательщика"),
        ("нет платежных систем", "Оформление: нет платёжных систем для оплаты"),
        ("нет платёжных систем", "Оформление: нет платёжных систем для оплаты"),
        # Незаполненные обязательные поля на чекауте (пункт «Обработка ошибок»
        # на многошаговом заказе: форма пропускает по «Далее», но заказ не
        # оформить). Фразы - именно сообщения об ОШИБКЕ, не подписи полей.
        ("заполните обязательные поля", "Оформление: не заполнены обязательные поля"),
        ("заполните все обязательные", "Оформление: не заполнены обязательные поля"),
        ("необходимо заполнить обязательные", "Оформление: не заполнены обязательные поля"),
        ("поле обязательно для заполнения", "Оформление: не заполнено обязательное поле"),
        ("обязательные поля не заполнены", "Оформление: не заполнены обязательные поля"),
        ("заказ не оформлен", "Оформление: заказ не оформлен"),
        ("заказ не был оформлен", "Оформление: заказ не оформлен"),
        ("ошибка оформления заказа", "Оформление: ошибка оформления заказа"),
    ]
    for needle, reason in маркеры:
        if needle in low:
            return reason
    return ""


# Пункт 2.7: уведомление пользователю о заявке после отправки формы (попап/
# картинка «спасибо», сообщение об успехе или смена текста кнопки на подтверждение).
_МАРКЕРЫ_УВЕДОМЛЕНИЯ = (
    "спасибо", "заявка принят", "заявка отправлен", "благодар", "мы свяжемся",
    "ваша заявка", "заявка успешно", "успешно отправл", "отправлено", "принято в обработ",
    "заявка получен", "мы получили", "будем на связи", "заявка зарегистрирован",
)


def _текст_подтверждает_отправку(text: str) -> bool:
    """True, если в тексте есть слова-маркеры подтверждения заявки пользователю.
    Чистая функция (без браузера) - легко тестируется."""
    t = (text or "").lower().replace("ё", "е")
    return any(m in t for m in _МАРКЕРЫ_УВЕДОМЛЕНИЯ)


def _текст_кнопки(loc, timeout_ms: int = 1000) -> str:
    """Видимый текст кнопки: inner_text ИЛИ атрибут value (у <input type=submit>
    текст лежит в value, а inner_text ПУСТОЙ) ИЛИ aria-label. Без этого смену
    «Отправить» → «Отправлено» на input-кнопках тул не видел - и ложно писал
    «нет подтверждения», хотя форма отправилась. Любая ошибка → пустая строка."""
    try:
        t = (loc.inner_text(timeout=timeout_ms) or "").strip()
    except Exception:  # noqa: BLE001
        t = ""
    if t:
        return t
    for attr in ("value", "aria-label"):
        try:
            v = (loc.get_attribute(attr) or "").strip()
            if v:
                return v
        except Exception:  # noqa: BLE001
            pass
    return ""


def детект_уведомления_пользователю(page, текст_кнопки_до: str = "",
                                    текст_кнопки_после: str = "",
                                    кнопка=None, таймаут_мс: int = 6000) -> str:
    """Проверяет (пункт 2.7), увидел ли пользователь подтверждение отправки заявки:
    всплывающий попап/картинка «спасибо», текст успеха на видимой странице или
    смена текста кнопки на подтверждение. Возвращает короткую пометку для отчёта:
    «Да (попап)» / «Да (кнопка)» / «Да (текст)» / «Нет». Ошибки браузера гасим.

    ОПРОС В ОКНЕ ВРЕМЕНИ: подтверждение прилетает ajax-ом и может появиться не
    сразу (кнопка «Отправить» → «Отправлено» через 1-4 с). Один снимок ловил
    только быстрые формы, а медленные помечал «Нет», хотя пользователь уведомление
    видит. Поэтому проверяем несколько раз в течение таймаут_мс и возвращаем первое
    сработавшее «Да». Текст кнопки читаем ЖИВЫМ локатором (если передан) - чтобы
    поймать позднюю смену, а не только снимок «после»."""
    до = (текст_кнопки_до or "").strip()

    def _однажды() -> str:
        # 1) Видимый попап/модалка с текстом успеха (всплывающая «картинка»).
        for sel in ("[class*='popup']", "[class*='modal']", "[role='dialog']",
                    "[class*='thank']", "[class*='success']", "[class*='spasibo']",
                    "[class*='thanks']"):
            try:
                loc = page.locator(sel)
                for i in range(min(loc.count(), 6)):
                    el = loc.nth(i)
                    if el.is_visible() and _текст_подтверждает_отправку(
                            el.inner_text(timeout=1000)):
                        return "Да (попап)"
            except Exception:  # noqa: BLE001
                pass
        # 2) Кнопка сменила текст на подтверждение («Отправить» → «Отправлено»).
        после = (текст_кнопки_после or "").strip()
        if кнопка is not None:
            try:
                живой = _текст_кнопки(кнопка, 1000)
                if живой:
                    после = живой
            except Exception:  # noqa: BLE001
                pass
        if после and после != до and _текст_подтверждает_отправку(после):
            return "Да (кнопка)"
        # 3) Текст успеха на видимой части страницы.
        try:
            if _текст_подтверждает_отправку(
                    page.locator("body").inner_text(timeout=2000)):
                return "Да (текст)"
        except Exception:  # noqa: BLE001
            pass
        return ""

    шаг_мс = 500
    попыток = max(1, int(таймаут_мс) // шаг_мс)
    for _ in range(попыток):
        итог = _однажды()
        if итог:
            return итог
        try:
            page.wait_for_timeout(шаг_мс)
        except Exception:  # noqa: BLE001
            break
    return "Нет"


# ── Защита формы от XSS (ввод <script> не исполняется) ──────────────────
# Отправляем форму с БЕЗВРЕДНЫМ маркером-payload в поле имени и смотрим, исполнился
# ли он / попал ли сырым HTML на ответ (Фаза 1) и в админку (Фаза 2, только СМУ).
# Payload безвреден: ставит переменную window.__xssFired и БОЛЬШЕ НИЧЕГО (не крадёт
# куки, не шлёт запросы, исчезает с закрытием вкладки). Если он сработал -
# экранирования нет, и настоящий (вредоносный) ввод тоже сработал бы: уязвимо.
# Голый <script> не используем: через reflection он не исполняется (ложное «ок»);
# <svg onload> срабатывает БЕЗ сетевого запроса. Маркер уникальный - чтобы не
# спутать наш ввод с разметкой сайта.
_XSS_MARK = "xSs0PrObe7"
_XSS_PAYLOAD = f'<svg onload="window.__xssFired=1" data-xssprobe="{_XSS_MARK}"></svg>'


def _xss_reset(page) -> None:
    """Сбрасывает флаг срабатывания ПЕРЕД отправкой (чтобы поймать именно наш)."""
    try:
        page.evaluate("window.__xssFired = 0")
    except Exception:  # noqa: BLE001
        pass


def _xss_наблюдение(page, html: str = "") -> tuple:
    """(исполнился, отражён_сырым) после отправки / на странице админки.
    исполнился - payload выполнил JS (window.__xssFired===1); отражён_сырым -
    payload попал в DOM как ЖИВОЙ элемент (data-xssprobe=маркер), а не как
    экранированный текст (&lt;svg…). Ошибки браузера гасим."""
    исполнился = False
    try:
        исполнился = bool(page.evaluate("window.__xssFired === 1"))
    except Exception:  # noqa: BLE001
        pass
    сырой = False
    try:
        сырой = page.locator(f"[data-xssprobe='{_XSS_MARK}']").count() > 0
    except Exception:  # noqa: BLE001
        h = (html or "").lower()
        сырой = (f'data-xssprobe="{_XSS_MARK}"'.lower() in h) and ("&lt;svg" not in h)
    return исполнился, сырой


def xss_вердикт(отправлена: bool, исполнился: bool, сырой: bool) -> tuple:
    """Сигналы XSS-пробы → (статус, деталь). ЧИСТАЯ функция (юнит-тест без браузера).
    Статус: «Защищена» / «УЯЗВИМА» / «Проверить»."""
    if not отправлена:
        return "Проверить", "форма не отправилась (капча/ошибка) - XSS не проверен"
    if исполнился:
        return "УЯЗВИМА", "введённый скрипт ИСПОЛНИЛСЯ - экранирования нет"
    if сырой:
        return "УЯЗВИМА", "ввод попал в страницу сырым HTML (не экранирован)"
    return "Защищена", "ввод экранирован / не отражён на видимой странице"


# Пункт 2.13: согласие на обработку персональных данных + ссылка на политику.
_ПОЛИТИКА_МАРКЕРЫ_213 = (
    "politik", "policy", "privacy", "personal", "confiden", "soglas", "agree",
    "полит", "конфиденциальн", "персональн", "обработк", "согласи", "152",
)


def ссылка_ведёт_на_политику(href: str, text: str) -> bool:
    """True, если ссылка (по href или подписи) ведёт на политику/согласие. Чистая."""
    h = (href or "").lower().replace("ё", "е")
    t = (text or "").lower().replace("ё", "е")
    return any(m in h for m in _ПОЛИТИКА_МАРКЕРЫ_213) or \
        any(m in t for m in _ПОЛИТИКА_МАРКЕРЫ_213)


def проверка_согласия_2_13(scope, page) -> dict:
    """Пункт 2.13 для одной формы: считает видимые чекбоксы согласия, проверяет
    предустановлены ли они, есть ли ссылка на политику, и обязательно ли согласие
    (без него форму не отправить). Ничего не отправляет: обязательность проверяем
    по атрибуту required и браузерной проверке формы checkValidity().
    Возвращает {чекбоксов, предустановлены, ссылка, валидация}."""
    res = {"чекбоксов": 0, "предустановлены": False, "ссылка": False, "валидация": False}
    # Зона поиска галочек: часто чекбоксы согласия лежат ВНЕ <form> - в теле
    # модалки/попапа (у МПЭ так почти везде). Если в самой форме видимых галочек
    # нет, расширяем поиск до ближайшего контейнера модалки - иначе визуально
    # видимые 2 галочки давали «0». Валидность (checkValidity) считаем по ФОРМЕ.
    контейнер = scope
    try:
        _в_форме = 0
        _cbf = scope.locator("input[type='checkbox']")
        for i in range(min(_cbf.count(), 12)):
            try:
                if _cbf.nth(i).is_visible():
                    _в_форме += 1
            except Exception:  # noqa: BLE001
                continue
        if not _в_форме:
            anc = scope.locator(
                "xpath=ancestor::*[contains(@class,'modal') or contains(@class,'popup') "
                "or contains(@class,'fancybox') or contains(@class,'my-modal') "
                "or @role='dialog'][1]")
            if anc.count():
                контейнер = anc.first
    except Exception:  # noqa: BLE001
        pass
    try:
        cb = контейнер.locator("input[type='checkbox']")
        видимых = 0
        for i in range(min(cb.count(), 12)):
            el = cb.nth(i)
            try:
                if not el.is_visible():
                    continue
            except Exception:  # noqa: BLE001
                continue
            видимых += 1
            try:
                if el.is_checked():
                    res["предустановлены"] = True
            except Exception:  # noqa: BLE001
                pass
            try:
                if el.evaluate("e => !!(e.required || "
                               "e.getAttribute('aria-required')==='true')"):
                    res["валидация"] = True
            except Exception:  # noqa: BLE001
                pass
        res["чекбоксов"] = видимых
    except Exception:  # noqa: BLE001
        pass
    # Ссылка на политику среди ссылок формы/модалки.
    try:
        links = контейнер.locator("a")
        for j in range(min(links.count(), 20)):
            a = links.nth(j)
            if ссылка_ведёт_на_политику(a.get_attribute("href") or "",
                                        a.inner_text(timeout=400) or ""):
                res["ссылка"] = True
                break
    except Exception:  # noqa: BLE001
        pass
    # Если по атрибуту required не увидели - спросим браузер: без согласия форма
    # невалидна? (снимаем видимые галочки согласия и вызываем checkValidity()).
    if not res["валидация"] and res["чекбоксов"]:
        try:
            cb = scope.locator("input[type='checkbox']")
            for i in range(min(cb.count(), 12)):
                el = cb.nth(i)
                try:
                    if el.is_visible() and el.is_checked():
                        el.uncheck(force=True)
                except Exception:  # noqa: BLE001
                    pass
            ok = scope.evaluate("f => { try { return f.checkValidity ? "
                                "f.checkValidity() : true } catch(e){ return true } }")
            if ok is False:
                res["валидация"] = True
        except Exception:  # noqa: BLE001
            pass
    return res


def _parse_accept_types(accept: str):
    """accept-атрибут <input type=file> → (список типов, принимает_любые?).
    Пустой accept или '*'/'*/*' = загрузчик берёт ЛЮБЫЕ файлы. Чистая
    функция - тестируется без браузера."""
    a = (accept or "").strip()
    if not a or a in ("*", "*/*", "*.*"):
        return [], True
    parts = [p.strip() for p in a.split(",") if p.strip()]
    if any(p in ("*", "*/*", "*.*") for p in parts):
        return parts, True
    return parts, bool(not parts)


def проверка_полей_форм(scope, page) -> dict:
    """Аудит полей формы БЕЗ отправки (доп. чек-лист):
      • Маска телефона: ограничено ли поле цифрами/длиной. Смотрим атрибуты
        (type=tel / pattern / maxlength / inputmode / JS-маска) И поведение -
        вводим буквы и лишние цифры, проверяем, что осталось; корректное
        значение восстанавливаем, отправку не трогаем.
      • Загрузка файлов: находим <input type=file>, читаем accept и выводим
        разрешённые типы (только вывод, без вердикта - тема безопасности:
        через это поле грузили вредоносные файлы, важно видеть, что можно).
    Возвращает {телефон_ограничен(bool|None), телефон_детали, файл_есть,
    файл_типы, файл_любые}."""
    res = {"телефон_ограничен": None, "телефон_детали": "поле телефона не найдено",
           "почта_ок": None, "почта_детали": "поле почты не найдено",
           "дата_ок": None, "дата_детали": "поле даты не найдено",
           "обязательность_ок": None, "обязательность_детали": "не определено",
           "длина_ок": None, "длина_детали": "поля с ограничением длины не найдены",
           "файл_есть": False, "файл_типы": [], "файл_любые": False}

    # ── Телефон ──
    phone_sel = (
        "input[type='tel'], input[name*='phone' i], input[name*='tel' i], "
        "input[placeholder*='телефон' i], input[placeholder*='phone' i], "
        "input[autocomplete='tel']")
    try:
        pl = scope.locator(phone_sel).first
        if pl.count():
            attrs = pl.evaluate(
                "el => ({type:(el.type||'').toLowerCase(),"
                " pattern: el.getAttribute('pattern'),"
                " maxlength: el.maxLength,"
                " inputmode:(el.getAttribute('inputmode')||'').toLowerCase(),"
                " mask: el.getAttribute('data-mask')||el.getAttribute('data-phone-mask')"
                "||el.getAttribute('data-tel')||'',"
                " cls: el.className||''})")
            детали = []
            if attrs.get("type") == "tel":
                детали.append("type=tel")
            if attrs.get("pattern"):
                детали.append("pattern")
            _ml = attrs.get("maxlength")
            if isinstance(_ml, int) and 0 < _ml <= 25:
                детали.append(f"maxlength={_ml}")
            if attrs.get("inputmode") in ("numeric", "tel"):
                детали.append(f"inputmode={attrs['inputmode']}")
            _has_mask = bool(attrs.get("mask")) or bool(
                re.search(r"(mask|imask|inputmask|js-tel|js-phone)",
                          attrs.get("cls", ""), re.I))
            if _has_mask:
                детали.append("JS-маска")
            attr_ok = bool(детали)

            # Поведенческая проба: буквы + много цифр → что останется в поле.
            beh_ok = None
            try:
                saved = pl.input_value(timeout=1500)
                pl.fill("", timeout=1500, force=True)
                pl.type("ab1cd2345678901234567890", timeout=2500)
                got = pl.input_value(timeout=1500) or ""
                letters = sum(c.isalpha() for c in got)
                beh_ok = (letters == 0 and len(got) <= 18)
                pl.fill(saved, timeout=1500, force=True)   # вернуть телефон
                if beh_ok and not attr_ok:
                    детали.append("ввод фильтруется")
            except Exception:  # noqa: BLE001
                beh_ok = None

            res["телефон_ограничен"] = bool(attr_ok or beh_ok)
            res["телефон_детали"] = (
                ", ".join(детали) if детали else
                "нет ограничения (не tel, без pattern/maxlength/inputmode/"
                "маски; ввод не фильтруется)")
    except Exception:  # noqa: BLE001
        pass

    # ── Почта: валидируется ли формат (пункт 2.14) ──
    # Одним evaluate: сохранить значение, проверить type=email, иначе ввести
    # «abcdef» (без @) и «a@b.ru» и сравнить checkValidity - восстановить.
    # checkValidity НЕ отправляет форму, только спрашивает браузер о валидности.
    try:
        _em = scope.evaluate(
            "f => {"
            " const e = f.querySelector(\"input[type='email'], input[name*='mail' i],"
            " input[name*='email' i], input[placeholder*='mail' i],"
            " input[placeholder*='почт' i], input[autocomplete='email']\");"
            " if(!e) return {found:false};"
            " const s=e.value; const t=(e.type||'').toLowerCase();"
            " if(t==='email'){ return {found:true, typeEmail:true}; }"
            " let v1=true,v2=true; try{ e.value='abcdef'; v1=e.checkValidity();"
            " e.value='a@b.ru'; v2=e.checkValidity(); }catch(_){}"
            " e.value=s; return {found:true, typeEmail:false, valid:(v1===false && v2===true)};"
            "}")
        if _em and _em.get("found"):
            if _em.get("typeEmail"):
                res["почта_ок"] = True
                res["почта_детали"] = "type=email (браузер проверяет формат)"
            elif _em.get("valid"):
                res["почта_ок"] = True
                res["почта_детали"] = "проверяет формат (pattern/валидация)"
            else:
                res["почта_ок"] = False
                res["почта_детали"] = "НЕ проверяет формат (примет любой текст без @)"
    except Exception:  # noqa: BLE001
        pass

    # ── Дата: валидируется ли формат (пункт 2.14) ──
    # Находим поле даты (type=date или текстовое «дата/дд.мм/дата рождения»).
    # type=date - формат гарантирует браузер; readonly = датапикер (руками
    # некорректное не ввести); pattern - РЕАЛЬНО проверяем, что мусорную дату
    # он отклоняет (checkValidity, форму не отправляем); маска/inputmode - тоже
    # ограничение. Иначе поле принимает любой текст = формат не проверяется.
    try:
        _dt = scope.evaluate(
            "f => {"
            " const e = f.querySelector(\"input[type='date'], input[name*='date' i],"
            " input[name*='дата' i], input[placeholder*='дата' i],"
            " input[placeholder*='дд.мм' i], input[placeholder*='dd.mm' i],"
            " input[placeholder*='дд/мм' i], input[name*='birth' i],"
            " input[autocomplete='bday']\");"
            " if(!e) return {found:false};"
            " const t=(e.type||'').toLowerCase();"
            " if(t==='date') return {found:true, typeDate:true};"
            " const pattern=e.getAttribute('pattern');"
            " let vinv=null;"
            " if(pattern){ const s=e.value; try{ e.value='зз.зз.зззз';"
            " vinv=e.checkValidity(); }catch(_){} e.value=s; }"
            " return {found:true, typeDate:false, ro:!!e.readOnly, pattern:!!pattern,"
            " vinv:vinv, ml:e.maxLength, im:(e.getAttribute('inputmode')||'').toLowerCase(),"
            " mask:e.getAttribute('data-mask')||e.getAttribute('data-date-mask')||'',"
            " cls:e.className||''};"
            "}")
        if _dt and _dt.get("found"):
            _ml_d = _dt.get("ml")
            _mask_d = bool(_dt.get("mask")) or bool(
                re.search(r"(mask|datepicker|flatpickr|calendar|air-?datepicker)",
                          _dt.get("cls", ""), re.I))
            if _dt.get("typeDate"):
                res["дата_ок"] = True
                res["дата_детали"] = "type=date (браузер проверяет формат)"
            elif _dt.get("ro"):
                res["дата_ок"] = True
                res["дата_детали"] = "датапикер (дата выбирается из календаря, руками не ввести)"
            elif _dt.get("pattern") and _dt.get("vinv") is False:
                res["дата_ок"] = True
                res["дата_детали"] = "проверяет формат (pattern отклоняет некорректную дату)"
            elif _mask_d or (isinstance(_ml_d, int) and 0 < _ml_d <= 10
                             and _dt.get("im") in ("numeric", "tel")):
                res["дата_ок"] = True
                res["дата_детали"] = "маска даты (ввод ограничен по формату)"
            else:
                res["дата_ок"] = False
                res["дата_детали"] = "НЕ проверяет формат (примет любой текст)"
    except Exception:  # noqa: BLE001
        pass

    # ── Обязательность полей / уведомления о заполнении (пункт 2.14) ──
    # Одним evaluate: сохранить значения всех полей, очистить видимые текстовые,
    # спросить form.checkValidity() (пустая форма невалидна = браузер покажет
    # «заполните поле»), восстановить значения. Форма НЕ отправляется.
    try:
        _rq = scope.evaluate(
            "f => {"
            " const inp=[...f.querySelectorAll('input,textarea,select')];"
            " const saved=inp.map(e=>e.value);"
            " const isText=e=>['text','tel','email','search','url','number','',"
            "'textarea'].includes((e.type||e.tagName||'').toLowerCase());"
            " const vis=e=>e.offsetParent!==null;"
            " const req=inp.filter(e=>vis(e)&&isText(e)&&(e.required||"
            "e.getAttribute('aria-required')==='true')).length;"
            " inp.forEach(e=>{ if(isText(e)) e.value=''; });"
            " let valid=true; try{ valid=f.checkValidity(); }catch(_){}"
            " inp.forEach((e,i)=>{ e.value=saved[i]; });"
            " return {req, emptyInvalid: valid===false};"
            "}")
        if _rq is not None:
            _req = int(_rq.get("req") or 0)
            _empty_invalid = bool(_rq.get("emptyInvalid"))
            res["обязательность_ок"] = bool(_req > 0 or _empty_invalid)
            res["обязательность_детали"] = (
                f"обязательных полей: {_req}; пустую форму не отправить: "
                f"{'да' if _empty_invalid else 'нет'}")
    except Exception:  # noqa: BLE001
        pass

    # ── Ограничение длины: РЕАЛЬНО пробуем ввести больше лимита (пункт 2.14) ──
    # Не верим атрибуту на слово: в каждое поле с maxlength вводим (лимит+5)
    # символов настоящими нажатиями (type) и проверяем, что браузер обрезал по
    # лимиту, как у живого пользователя. Значение сохраняем и возвращаем; форму
    # не трогаем. Телефон пропускаем (его покрывает проба маски + маска добавляет
    # формат-символы → ложные срабатывания). type=number тоже пропускаем: браузеры
    # к нему maxlength не применяют - иначе ложное «не держит».
    try:
        проверено, не_держат, примеры = 0, [], []
        поля = scope.locator(
            "input:not([type='checkbox']):not([type='radio']):not([type='file'])"
            ":not([type='hidden']):not([type='submit']):not([type='button'])"
            ":not([type='date']):not([type='time']):not([type='color'])"
            ":not([type='range']):not([type='number']), textarea")
        for i in range(min(поля.count(), 30)):
            if проверено >= 8:
                break
            el = поля.nth(i)
            try:
                info = el.evaluate(
                    "e => ({ml: e.maxLength, vis: e.offsetParent!==null,"
                    " off: !!(e.disabled||e.readOnly),"
                    " tel: ((e.type||'').toLowerCase()==='tel')"
                    " || /phone|tel/i.test(e.name||'')"
                    " || /телефон|phone/i.test(e.getAttribute('placeholder')||''),"
                    " label:(e.name||e.getAttribute('placeholder')||e.id||'поле')})")
            except Exception:  # noqa: BLE001
                continue
            ml = info.get("ml")
            if not isinstance(ml, int) or not (0 < ml <= 40):
                continue                        # нет реального лимита / слишком большой
            if not info.get("vis") or info.get("off") or info.get("tel"):
                continue                        # скрыто/недоступно/это телефон
            try:
                saved = el.input_value(timeout=1500)
            except Exception:  # noqa: BLE001
                continue
            got = None
            try:
                el.fill("", timeout=1500, force=True)
                el.type(("1234567890" * 5)[:ml + 5], timeout=3000)   # лимит+5 цифр
                got = el.input_value(timeout=1500) or ""
            except Exception:  # noqa: BLE001
                got = None
            try:
                el.fill(saved, timeout=1500, force=True)             # вернуть как было
            except Exception:  # noqa: BLE001
                pass
            if got is None:
                continue
            проверено += 1
            if len(got) > ml:
                не_держат.append(f"«{info['label']}»: лимит {ml}, ввелось {len(got)}")
            else:
                примеры.append(f"{info['label']}={ml}")
        if проверено == 0:
            res["длина_ок"] = None
            res["длина_детали"] = "поля с ограничением длины не найдены"
        elif не_держат:
            res["длина_ок"] = False
            res["длина_детали"] = "ограничение НЕ работает: " + "; ".join(не_держат[:4])
        else:
            res["длина_ок"] = True
            _пр = ", ".join(примеры[:4])
            res["длина_детали"] = (f"полей с лимитом: {проверено}, все не дают ввести "
                                   f"больше" + (f" ({_пр})" if _пр else ""))
    except Exception:  # noqa: BLE001
        pass

    # ── Загрузка файлов ──
    try:
        fi = scope.locator("input[type='file']")
        n = fi.count()
        if n:
            res["файл_есть"] = True
            типы, любые = set(), False
            for i in range(min(n, 6)):
                acc = fi.nth(i).get_attribute("accept") or ""
                parts, unrestricted = _parse_accept_types(acc)
                типы.update(parts)
                if unrestricted:
                    любые = True
            res["файл_типы"] = sorted(типы)
            res["файл_любые"] = любые or not типы
    except Exception:  # noqa: BLE001
        pass
    return res


def состав_формы(scope) -> dict:
    """Пункт «Все элементы формы (поля/кнопки/чекбоксы/радио/списки) присутствуют».
    Перепись ВИДИМЫХ элементов формы ОДНИМ evaluate - чистое чтение DOM, без ввода
    и без отправки. Вердикт «Проверить» - только если форма совсем без элементов
    ввода. «Соответствие дизайну» - визуально по макету (авто не проверяем).
    Возвращает {ок(bool|None), детали, поля, чекбоксы, радио, радиоГрупп, списки,
    кнопки, отправка, файл}."""
    c = {"поля": 0, "чекбоксы": 0, "радио": 0, "радиоГрупп": 0, "списки": 0,
         "кнопки": 0, "отправка": False, "файл": 0}
    try:
        r = scope.evaluate(
            "el => {"
            " const vis = e => !!(e.offsetWidth || e.offsetHeight || e.getClientRects().length);"
            " const inputs = [...el.querySelectorAll('input')].filter(vis);"
            " const tt = ['text','tel','email','search','url','number','password',"
            "'date','time','datetime-local','month','week',''];"
            " const ty = e => (e.type||'text').toLowerCase();"
            " const поля = inputs.filter(e => tt.includes(ty(e))).length"
            "   + [...el.querySelectorAll('textarea')].filter(vis).length;"
            " const чекбоксы = inputs.filter(e => ty(e)==='checkbox').length;"
            " const радио = inputs.filter(e => ty(e)==='radio');"
            " const nm = радио.map(e => e.name||'');"
            " const радиоГрупп = new Set(nm.filter(n=>n)).size + nm.filter(n=>!n).length;"
            " const файл = inputs.filter(e => ty(e)==='file').length;"
            " const списки = [...el.querySelectorAll('select')].filter(vis).length;"
            " const btns = [...el.querySelectorAll("
            "\"button, input[type='submit'], input[type='button'], input[type='reset']\")].filter(vis);"
            " const отправка = btns.some(b => { const g=b.tagName.toLowerCase();"
            "   const t=(b.getAttribute('type')||'').toLowerCase();"
            "   return g==='input' ? t==='submit' : (t===''||t==='submit'); });"
            " return {поля, чекбоксы, радио: радио.length, радиоГрупп, списки,"
            "   кнопки: btns.length, отправка, файл};"
            "}")
        if isinstance(r, dict):
            for k in c:
                if k in r and r[k] is not None:
                    c[k] = r[k]
    except Exception:  # noqa: BLE001
        return {"ок": None, "детали": "состав формы определить не удалось", **c}

    есть_ввод = bool(c["поля"] or c["чекбоксы"] or c["списки"] or c["радио"])
    радио_txt = f"{c['радио']}" + (f" ({c['радиоГрупп']} групп)" if c["радио"] else "")
    if c["кнопки"]:
        кноп_txt = f"{c['кнопки']}" + (" (есть отправка)" if c["отправка"]
                                       else " (кнопки отправки нет)")
    else:
        кноп_txt = "0"
    детали = (f"поля ввода: {c['поля']}, чекбоксы: {c['чекбоксы']}, "
              f"радио: {радио_txt}, списки: {c['списки']}, кнопки: {кноп_txt}, "
              f"файл: {c['файл']}. Соответствие дизайну — по макету вручную.")
    if not есть_ввод:
        детали = "элементы ввода не найдены (та ли форма/скоуп?). " + детали
    elif not c["кнопки"]:
        детали += " Кнопка отправки не найдена — возможно, отправка через JS/ссылку."
    return {"ок": есть_ввод, "детали": детали, **c}


def _rgb_в_hex(s: str) -> str:
    """'rgb(61, 72, 88)' / 'rgba(...)' → '#3D4858'. Прозрачное → 'прозрачный'.
    Прочее (named-цвет и т.п.) возвращаем как есть."""
    try:
        m = re.findall(r"[\d.]+", s or "")
        if not m:
            return s or ""
        r, g, b = int(float(m[0])), int(float(m[1])), int(float(m[2]))
        a = float(m[3]) if len(m) > 3 else 1.0
        if a == 0:
            return "прозрачный"
        return f"#{r:02X}{g:02X}{b:02X}"
    except Exception:  # noqa: BLE001
        return s or ""


def стиль_формы(scope) -> dict:
    """Пункт «Форма стилизована по макету (цвета, шрифты, отступы)».

    Чистое чтение вычисленных стилей (getComputedStyle) видимых текст-полей и
    кнопки отправки - без ввода и без отправки, как `состав_формы`. Абсолютное
    «совпало с макетом дизайнера» без Figma-файла проекта проверить нельзя (у
    каждого проекта свой макет), поэтому проверяем ОБЪЕКТИВНОЕ:
      • факт стилизации - поля не «голый» браузерный дефолт (border-style:
        inset/outset - UA-дефолт, который вручную никто не верстает);
      • консистентность - все поля одинаковы по шрифту/размеру/рамке/радиусу
        (разнобой = реальный баг вёрстки).
    В детали всегда кладём снятый «слепок» стилей - QA сверяет с макетом глазами.

    → {состояние: 'не найдено'|'консистентно'|'разнобой'|'дефолт', ок, детали}.
    ок=None - нечего оценивать (нет полей); True - консистентно и застилизовано;
    False - разнобой или дефолт (в отчёте станет «Проверить»)."""
    try:
        r = scope.evaluate(
            "el => {"
            " const vis = e => !!(e.offsetWidth || e.offsetHeight || e.getClientRects().length);"
            " const g = (e,p) => getComputedStyle(e)[p];"
            " const snap = e => ({"
            "   font: (g(e,'fontFamily')||'').split(',')[0].replace(/[\"']/g,'').trim(),"
            "   size: g(e,'fontSize'), color: g(e,'color'), bg: g(e,'backgroundColor'),"
            "   bw: g(e,'borderTopWidth'), bs: g(e,'borderTopStyle'), bc: g(e,'borderTopColor'),"
            "   radius: g(e,'borderTopLeftRadius'),"
            "   pad: g(e,'paddingTop')+' '+g(e,'paddingLeft') });"
            " const tt = ['text','tel','email','search','url','number','password',"
            "'date','time','datetime-local','month','week',''];"
            " const ty = e => (e.type||'text').toLowerCase();"
            " const поля = [...el.querySelectorAll('input')].filter(vis)"
            "   .filter(e => tt.includes(ty(e)))"
            "   .concat([...el.querySelectorAll('textarea')].filter(vis)).slice(0,8);"
            " const btns = [...el.querySelectorAll("
            "\"button, input[type='submit'], a.btn, [class*='btn'], [class*='submit']\")].filter(vis);"
            " const typed = btns.filter(e => (e.getAttribute('type')||'').toLowerCase()==='submit');"
            " const btn = typed[0] || btns[0] || null;"
            " return {поля: поля.map(snap), кнопка: btn ? snap(btn) : null};"
            "}")
    except Exception:  # noqa: BLE001
        return {"ок": None, "состояние": "не найдено",
                "детали": "стили формы прочитать не удалось"}

    поля = (r or {}).get("поля") or []
    if not поля:
        return {"ок": None, "состояние": "не найдено",
                "детали": "видимых текст-полей нет - стилизацию оценивать нечего"}

    # «Слепок» первого поля - человекочитаемо, для сверки с макетом глазами.
    p0 = поля[0]
    рамка = f"{p0['bw']} {p0['bs']} {_rgb_в_hex(p0['bc'])}"
    слепок = (f"шрифт {p0['font']} {p0['size']}, цвет {_rgb_в_hex(p0['color'])}, "
              f"фон {_rgb_в_hex(p0['bg'])}, рамка {рамка}, радиус {p0['radius']}, "
              f"паддинг {p0['pad']}")

    # Кнопка: брендированная (свой фон) или дефолтная серая/прозрачная.
    b = (r or {}).get("кнопка")
    if b:
        _bg = (b["bg"] or "").replace(" ", "")
        деф_кноп = _bg in ("rgb(239,239,239)", "rgba(0,0,0,0)", "") or "inset" in (b["bs"] or "")
        кноп_txt = ("кнопка дефолтная (не брендирована)" if деф_кноп
                    else f"кнопка брендирована (фон {_rgb_в_hex(b['bg'])})")
    else:
        кноп_txt = "кнопка отправки не найдена"

    # 1) Факт стилизации: border-style inset/outset - «голый» браузерный дефолт.
    дефолтные = [i + 1 for i, f in enumerate(поля)
                 if (f["bs"] or "").lower() in ("inset", "outset")]
    if дефолтные:
        return {"ок": False, "состояние": "дефолт",
                "детали": (f"поля не застилизованы - браузерный дефолт "
                           f"(border-style: {p0['bs']}), макет не применён. "
                           f"Поля: {len(поля)}. {слепок}. {кноп_txt}")}

    # 2) Консистентность: подпись = шрифт|размер|рамка|радиус (цвет/фон не берём -
    # placeholder vs заполненное поле легально отличаются по цвету текста).
    def подпись(f):
        return "|".join([f["font"], f["size"], f["bw"], f["bs"],
                         f["bc"], f["radius"]])
    подписи = {подпись(f) for f in поля}
    if len(подписи) > 1:
        # Находим, чем выделяется «белая ворона» (по частоте подписи).
        from collections import Counter
        cnt = Counter(подпись(f) for f in поля)
        норм = cnt.most_common(1)[0][0]
        выбив = [i + 1 for i, f in enumerate(поля) if подпись(f) != норм]
        return {"ок": False, "состояние": "разнобой",
                "детали": (f"поля различаются по стилю (шрифт/размер/рамка/радиус): "
                           f"выбиваются поля {выбив} из {len(поля)}. "
                           f"Эталон (поле 1): {слепок}. {кноп_txt}")}

    return {"ок": True, "состояние": "консистентно",
            "детали": (f"поля стилизованы единообразно ({len(поля)} шт.). "
                       f"{слепок}. {кноп_txt}")}


# JS: очистить ВСЕ поля (сохранив значения в window.__valSaved) и спросить браузер
# о нативной валидации - БЕЗ отправки (checkValidity не шлёт). Также считаем
# «похожие на обязательные» поля: native required / класс-маркер / core-поле
# (имя/телефон/почта) - чтобы покрыть формы с кастомной JS-валидацией без атрибута.
_JS_VAL_NATIVE = r"""
f => {
  const form = f.tagName==='FORM' ? f : (f.querySelector('form') || f);
  const skip = ['hidden','submit','button','reset','image'];
  const vis = e => !!(e.offsetWidth || e.offsetHeight || e.getClientRects().length);
  const ctrls = [...form.querySelectorAll('input,textarea,select')]
    .filter(e => !skip.includes((e.type||'').toLowerCase()));
  const looksReq = e => {
    if (e.required || e.getAttribute('aria-required')==='true') return true;
    const cls = (e.className||'') + ' ' + (e.parentElement ? e.parentElement.className : '');
    if (/require|mandat|обязат/i.test(cls)) return true;
    const nm = ((e.name||'') + ' ' + (e.placeholder||'') + ' '
                + (e.getAttribute('autocomplete')||'')).toLowerCase();
    if (/phone|tel|тел|mail|почт|name|имя|fio|фио/.test(nm)) return true;
    return false;
  };
  const fillable = ctrls.filter(e => {
    const t=(e.type||'text').toLowerCase();
    return vis(e) && (e.tagName==='TEXTAREA' || e.tagName==='SELECT'
      || ['text','tel','email','search','url','number','password',''].includes(t)); });
  const reqLike = ctrls.filter(looksReq);
  window.__valSaved = ctrls.map(e => ({v: e.value, c: e.checked}));
  ctrls.forEach(e => { const t=(e.type||'').toLowerCase();
    if (t==='checkbox'||t==='radio') e.checked=false; else e.value=''; });
  let formValid = true, msgs = [];
  try { formValid = form.checkValidity(); } catch(_){}
  for (const e of ctrls) { try {
    if (!e.checkValidity() && e.validationMessage) msgs.push(e.validationMessage);
  } catch(_){} }
  return {fillable: fillable.length, reqLike: reqLike.length, formValid,
          msgs: [...new Set(msgs)].slice(0,3)};
}
"""

# JS: вернуть поля к сохранённым значениям (после пробы).
_JS_VAL_RESTORE = r"""
f => {
  const form = f.tagName==='FORM' ? f : (f.querySelector('form') || f);
  const skip = ['hidden','submit','button','reset','image'];
  const ctrls = [...form.querySelectorAll('input,textarea,select')]
    .filter(e => !skip.includes((e.type||'').toLowerCase()));
  const s = window.__valSaved || [];
  ctrls.forEach((e,i) => { if (s[i]) { const t=(e.type||'').toLowerCase();
    if (t==='checkbox'||t==='radio') e.checked=s[i].c; else e.value=s[i].v; } });
  return true;
}
"""

# JS: перехватчик submit с preventDefault - ловим ФАКТ попытки отправки и ОТМЕНЯЕМ её.
_JS_VAL_ARM = r"""
f => {
  const g = f.tagName==='FORM' ? f : f.querySelector('form');
  if (!g) return 'noform';
  window.__valSub = false;
  window.__valH = function(e){ window.__valSub=true; e.preventDefault(); e.stopPropagation(); };
  g.addEventListener('submit', window.__valH, true);
  return 'ok';
}
"""

_JS_VAL_DISARM = r"""
f => { const g = f.tagName==='FORM' ? f : f.querySelector('form');
  if (g && window.__valH) g.removeEventListener('submit', window.__valH, true); return true; }
"""

# JS: ищем ВИДИМУЮ ошибку валидации в форме - сообщение (текст+цвет) или подсветку поля.
_JS_VAL_SCAN = r"""
f => {
  const root = f.tagName==='FORM' ? f : (f.querySelector('form') || f);
  const vis = el => { const r=el.getBoundingClientRect(); const s=getComputedStyle(el);
    return r.width>0 && r.height>0 && s.visibility!=='hidden'
        && s.display!=='none' && s.opacity!=='0'; };
  const red = c => { const m=(c||'').match(/[\d.]+/g); if(!m) return false;
    const r=+m[0],g=+m[1],b=+m[2]; return r>=120 && r-g>=40 && r-b>=40; };
  const sels = "[class*=error i],[class*=invalid i],[aria-invalid=true],.errortext,"
    + ".form-error,.field-error,.help-block,.invalid-feedback,[role=alert]";
  for (const el of root.querySelectorAll(sels)) {
    if (!vis(el)) continue; const t=(el.innerText||'').trim();
    if (!t || t.length>160) continue;
    const c=getComputedStyle(el).color;
    return {found:true, text:t.slice(0,120), color:c, red:red(c), kind:'msg'};
  }
  for (const e of root.querySelectorAll('input,textarea,select')) {
    if (!vis(e)) continue; let inv=false;
    try { inv = e.matches(':invalid') || e.getAttribute('aria-invalid')==='true'; } catch(_){}
    if (inv) { const bc=getComputedStyle(e).borderTopColor;
      return {found:true, text:'(поле подсвечено)', color:bc, red:red(bc), kind:'field'}; }
  }
  return {found:false};
}
"""


def проверка_отображения_ошибок(scope, page, sub, is_order) -> dict:
    """Пункт «Ошибки валидации отображаются корректно (цвета, текст ошибок)».

    Проверяет НЕ наличие правил валидации (это делает 2.14), а ПОКАЗ ошибки
    пользователю: при пустом/невалидном вводе появляется ли видимое сообщение,
    красное ли оно, есть ли внятный текст.

    Безопасность - заявка НЕ уходит ни в каком случае:
      • нативную валидацию ловим `checkValidity()` без отправки;
      • кастомную (JS Bitrix) - контролируемым пустым сабмитом под ДВОЙНОЙ
        защитой: перехватчик submit с preventDefault + abort POST-запроса.
    Поля восстанавливаем из window.__valSaved в конце.

    → {состояние: 'есть'|'нет'|'без реакции'|'не найдено'|'проверить вручную', детали}."""
    def _restore():
        try:
            scope.evaluate(_JS_VAL_RESTORE)
        except Exception:  # noqa: BLE001
            pass

    try:
        r = scope.evaluate(_JS_VAL_NATIVE)
    except Exception:  # noqa: BLE001
        return {"состояние": "не найдено", "детали": "валидацию формы прочитать не удалось"}
    if not isinstance(r, dict):
        _restore()
        return {"состояние": "не найдено", "детали": "форму прочитать не удалось"}

    # Нет заполняемых полей - показ ошибок проверять нечего.
    if not r.get("fillable"):
        _restore()
        return {"состояние": "не найдено",
                "детали": "заполняемых полей нет - показ ошибок проверять нечего"}

    # Нативная валидация ловит пустые поля → браузер сам покажет ошибку (гарантированно).
    if r.get("formValid") is False:
        _restore()
        msgs = r.get("msgs") or []
        текст = ("; ".join(msgs))[:150] if msgs else "браузерная проверка активна"
        return {"состояние": "есть",
                "детали": f"нативная валидация - браузер покажет ошибку ({текст})"}

    # Нативной валидации нет и «обязательных» полей не видно (нет имя/тел/почта,
    # нет маркеров) - пустая отправка легальна, проверять нечего.
    if not r.get("reqLike"):
        _restore()
        return {"состояние": "не найдено",
                "детали": "нативной валидации нет, обязательных полей не видно - проверять нечего"}

    # Есть похожие на обязательные поля, но нативной валидации нет → форма
    # полагается на кастомный JS. На форме заказа авто-пробу не делаем
    # (перестраховка на чекауте) - помечаем ручной проверкой.
    if is_order:
        _restore()
        return {"состояние": "проверить вручную",
                "детали": ("нет нативной валидации, форма заказа - авто-проба пропущена; "
                           "проверьте показ ошибки вручную")}

    # Кастомный проход: пустой сабмит под двойной защитой (preventDefault + abort POST).
    posted = {"n": 0}
    armed = listener = False
    err = {"found": False}
    attempted = False

    def _val_route(route):
        try:
            req = route.request
            if (req.method or "").upper() == "POST" and not _ds_это_трекер(req.url):
                posted["n"] += 1
                route.abort("failed")
            else:
                route.continue_()
        except Exception:  # noqa: BLE001
            try:
                route.continue_()
            except Exception:  # noqa: BLE001
                pass

    try:
        try:
            listener = (scope.evaluate(_JS_VAL_ARM) == "ok")
        except Exception:  # noqa: BLE001
            listener = False
        page.route("**/*", _val_route)
        armed = True
        try:
            sub.click(timeout=5000)
        except Exception:  # noqa: BLE001
            try:
                sub.click(timeout=5000, force=True)
            except Exception:  # noqa: BLE001
                pass
        page.wait_for_timeout(1300)
        attempted = bool(posted["n"])
        try:
            attempted = attempted or bool(page.evaluate("() => !!window.__valSub"))
        except Exception:  # noqa: BLE001
            pass
        try:
            err = scope.evaluate(_JS_VAL_SCAN) or {"found": False}
        except Exception:  # noqa: BLE001
            err = {"found": False}
    finally:
        if armed:
            try:
                page.unroute("**/*", _val_route)
            except Exception:  # noqa: BLE001
                try:
                    page.unroute("**/*")
                except Exception:  # noqa: BLE001
                    pass
        if listener:
            try:
                scope.evaluate(_JS_VAL_DISARM)
            except Exception:  # noqa: BLE001
                pass
        try:
            sub.evaluate("b => { try { b.disabled=false; b.removeAttribute('disabled'); } catch(_){} }")
        except Exception:  # noqa: BLE001
            pass
        _restore()
        page.wait_for_timeout(200)

    # Если форма ПОПЫТАЛАСЬ отправить пустые данные - клиентской валидации нет
    # (реальная валидация блокирует ДО отправки). Любая «ошибка» тут - это уже
    # реакция на наш перехват, а не валидация. Заявка перехвачена, не ушла.
    if attempted:
        return {"состояние": "нет",
                "детали": ("пустая форма попыталась отправиться без клиентской валидации - "
                           "ошибок не показано (заявка перехвачена, не ушла)")}
    if err.get("found"):
        цвет = _rgb_в_hex(err.get("color") or "")
        крас = "красная" if err.get("red") else "НЕ красная - проверить цвет"
        поле = " (подсветка поля)" if err.get("kind") == "field" else ""
        txt = (err.get("text") or "").strip()
        return {"состояние": "есть",
                "детали": f"показана ошибка{поле}: «{txt}», цвет {цвет} - {крас}"}
    return {"состояние": "без реакции",
            "детали": ("на пустую отправку форма не показала ошибку и не отправилась - "
                       "проверьте, видима ли реакция валидации")}


# ── Серверная валидация: нельзя отправить неверные данные в обход клиентской
# проверки (как правка в DevTools) ────────────────────────────────────────
# В отличие от проверка_отображения_ошибок (там перехватчик + route.abort
# специально ГАСЯТ запрос, чтобы ничего не ушло) - здесь наоборот: снимаем
# HTML5-констрейнт и подставляем невалидное значение НАПРЯМУЮ через JS (минуя
# маску/кастомный JS-валидатор, ровно как правка в DevTools), а запросу
# ДАЁМ реально уйти - и смотрим, что ответил сервер. Три вида нарушения,
# каждый - отдельная (по-настоящему невалидная) отправка; пропускаем те,
# для которых на форме нет подходящего поля.
_SRVVAL_ВИДЫ = (
    ("empty", "обязательное поле пусто"),
    ("empty_name", "без имени"),
    ("empty_phone", "без телефона"),
    ("bad_email", "некорректный e-mail"),
    ("too_long", "превышение длины поля"),
)

# JS: находит подходящее поле под вид нарушения, снимает required/pattern/
# maxlength, подставляет невалидное значение (с узнаваемым маркером в тексте,
# где это возможно - если сервер всё же примет, заявку легко узнать в
# админке) и диспатчит input/change - чтобы сработали и кастомные JS-валидаторы,
# слушающие события, а не только нативный констрейнт. Метит поле временным
# data-атрибутом для точного восстановления после пробы.
_JS_SRVVAL_TAMPER = r"""
(f, тип) => {
  const vis = e => !!(e.offsetWidth || e.offsetHeight || e.getClientRects().length);
  const fire = e => { e.dispatchEvent(new Event('input', {bubbles:true}));
                       e.dispatchEvent(new Event('change', {bubbles:true})); };
  const looksReq = e => {
    if (e.required || e.getAttribute('aria-required')==='true') return true;
    const cls = (e.className||'') + ' ' + (e.parentElement ? e.parentElement.className : '');
    if (/require|mandat|обязат/i.test(cls)) return true;
    const nm = ((e.name||'') + ' ' + (e.placeholder||'') + ' '
                + (e.getAttribute('autocomplete')||'')).toLowerCase();
    if (/phone|tel|тел|mail|почт|name|имя|fio|фио/.test(nm)) return true;
    return false;
  };
  const skip = ['hidden','submit','button','reset','image','checkbox','radio','file'];
  const ctrls = [...f.querySelectorAll('input,textarea,select')]
    .filter(e => vis(e) && !skip.includes((e.type||'').toLowerCase()));

  let target = null;
  if (тип === 'empty') {
    // Обязательное, но НЕ похожее на «имя» - чтобы маркер в имени клиента,
    // если сервер всё же примет заявку, остался читаемым в админке.
    target = ctrls.find(e => looksReq(e) && !/^(name|fio|имя)$/i.test(e.name||''))
      || ctrls.find(looksReq);
  } else if (тип === 'bad_email') {
    target = f.querySelector("input[type='email'], input[name*='mail' i],"
      + " input[name*='email' i], input[placeholder*='mail' i],"
      + " input[placeholder*='почт' i], input[autocomplete='email']");
  } else if (тип === 'too_long') {
    // Textarea/текстовое поле, НЕ телефон/почта; избегаем «имени», пока есть
    // другой кандидат - иначе 4000 симв. затрут узнаваемое имя в заявке.
    const notPhoneMail = e => !/phone|tel|тел|mail|почт/i.test((e.name||'') + (e.placeholder||''));
    const isText = e => e.tagName === 'TEXTAREA'
      || ['text','search',''].includes((e.type||'').toLowerCase());
    target = ctrls.find(e => e.tagName === 'TEXTAREA' && notPhoneMail(e))
      || ctrls.find(e => isText(e) && notPhoneMail(e) && !/^(name|fio|имя)$/i.test(e.name||''))
      || ctrls.find(e => isText(e) && notPhoneMail(e));
  } else if (тип === 'empty_name') {
    // Поле ИМЕНИ - очистим именно его (проба «прошло без имени»). Остальные
    // поля (телефон/почта) остаются с тест-данными - заявку видно в админке.
    target = ctrls.find(e => /name|fio|имя|фио|фамил/i.test(
      (e.name||'') + ' ' + (e.placeholder||'') + ' ' + (e.getAttribute('autocomplete')||'')));
  } else if (тип === 'empty_phone') {
    // Поле ТЕЛЕФОНА - очистим именно его (проба «прошло без телефона»).
    target = ctrls.find(e => (e.type||'').toLowerCase()==='tel'
      || /phone|tel|тел/i.test(
        (e.name||'') + ' ' + (e.placeholder||'') + ' ' + (e.getAttribute('autocomplete')||'')));
  }
  if (!target) return {done: false};

  target.setAttribute('data-srvval-tmp', '1');
  const old = target.value;
  target.removeAttribute('required');
  target.removeAttribute('pattern');
  target.removeAttribute('maxlength');
  if ((target.type || '').toLowerCase() === 'email') target.type = 'text';

  if (тип === 'empty' || тип === 'empty_name' || тип === 'empty_phone') target.value = '';
  else if (тип === 'bad_email') target.value = 'test-validation-probe-not-an-email';
  else if (тип === 'too_long') target.value = 'ТЕСТ-ВАЛИДАЦИЯ ' + 'A'.repeat(4000);

  fire(target);
  return {done: true, поле: (target.name || target.id || target.placeholder || 'поле'), old: old};
}
"""

_JS_SRVVAL_RESTORE = r"""
(f, old) => {
  const t = f.querySelector('[data-srvval-tmp="1"]');
  if (!t) return false;
  t.value = old || '';
  t.dispatchEvent(new Event('input', {bubbles:true}));
  t.dispatchEvent(new Event('change', {bubbles:true}));
  t.removeAttribute('data-srvval-tmp');
  return true;
}
"""


# ── Общий детектор «эта КОНКРЕТНАЯ попытка отправки успешна» ─────────────
# Используется и в проба_серверной_валидации, и в активная_проба_лимита - обе
# кликают ПОСЛЕ того, как форма уже была легитимно отправлена (и ещё раз -
# тестом двойной отправки) выше по потоку. К этому моменту страница часто уже
# показывает «Спасибо!» от ТОЙ, первой отправки - обычный детект_уведомления_
# пользователю() смотрит «видно ли подтверждение СЕЙЧАС» без сравнения с
# состоянием ДО и на таких сайтах засчитывает успех на КАЖДОЙ попытке подряд,
# даже если сервер её на самом деле отверг (ложный «УЯЗВИМА»/«не сработала»).
#
# Поэтому здесь два независимых сигнала:
#   1) Реальный HTTP-ответ формы (page.expect_response) - самый надёжный,
#      не зависит от того, что уже нарисовано на странице от прошлых попыток.
#   2) Если ответ поймать не удалось (переход страницы / не-AJAX форма) -
#      визуальный фолбэк, но ТОЛЬКО по переходу «подтверждения не было видно
#      ДО этого клика → появилось ПОСЛЕ» (не «видно ли оно сейчас»).
def _подтверждение_видно(page) -> bool:
    """Снимок «видно ли ПРЯМО СЕЙЧАС подтверждение отправки» - попап ИЛИ текст
    на странице (тот же вокабуляр, что и в детект_уведомления_пользователю,
    но без опроса во времени - чистый снимок для сравнения «было/стало»)."""
    for sel in ("[class*='popup']", "[class*='modal']", "[role='dialog']",
                "[class*='thank']", "[class*='success']", "[class*='spasibo']",
                "[class*='thanks']"):
        try:
            loc = page.locator(sel)
            for i in range(min(loc.count(), 6)):
                el = loc.nth(i)
                if el.is_visible() and _текст_подтверждает_отправку(el.inner_text(timeout=500)):
                    return True
        except Exception:  # noqa: BLE001
            pass
    try:
        if _текст_подтверждает_отправку(page.locator("body").inner_text(timeout=1500)):
            return True
    except Exception:  # noqa: BLE001
        pass
    return False


def _ответ_формы_вердикт(текст: str, статус) -> str:
    """«успешно»/«ошибка»/«неизвестно» по перехваченному HTTP-ответу формы.
    ЧИСТАЯ функция (юнит-тест без браузера). Тот же вокабуляр, что и
    _interpret_response_status, адаптирован под (текст, статус) вместо
    requests.Response - переиспользует response_indicates_captcha_block."""
    if статус is None:
        return "неизвестно"
    t = (текст or "").lower()
    try:
        j = json.loads(текст) if текст else None
    except (ValueError, TypeError):
        j = None
    if isinstance(j, dict):
        if j.get("success") is True or j.get("ok") is True:
            return "успешно"
        if j.get("success") is False or j.get("ok") is False or j.get("error") or j.get("errors"):
            return "ошибка"
    if "csrf" in t and "invalid" in t:
        return "ошибка"
    if response_indicates_captcha_block(текст or ""):
        return "ошибка"
    if "ошибк" in t and any(x in t for x in ("не удалось", "не отправлен", "отклонен", "invalid", "error")):
        return "ошибка"
    if "спасибо" in t or "успешно" in t or "принят" in t:
        return "успешно"
    if статус == 200:
        return "успешно"
    return "ошибка"


def _клик_поймать_ответ(page, sub, timeout_ms: int = 5000):
    """Кликает по кнопке ОДИН раз, пытаясь поймать HTTP-ответ реальной
    отправки формы (не трекер аналитики - _ds_это_трекер). Возвращает
    (текст, статус) пойманного ответа, или (None, None), если поймать не
    удалось (переход страницы / не-AJAX форма / ответ не успел) - клик при
    этом уже произошёл, повторно НЕ кликает."""
    кликнуто = {"done": False}

    def _click():
        try:
            sub.click(timeout=5000)
        except Exception:  # noqa: BLE001
            sub.click(timeout=5000, force=True)
        кликнуто["done"] = True

    try:
        with page.expect_response(
                lambda r: (r.request.method or "").upper() == "POST"
                          and not _ds_это_трекер(r.url),
                timeout=timeout_ms) as _ri:
            _click()
        resp = _ri.value
        try:
            текст = resp.text()
        except Exception:  # noqa: BLE001
            текст = ""
        return текст, resp.status
    except Exception:  # noqa: BLE001
        if not кликнуто["done"]:
            try:
                sub.click(timeout=5000)
            except Exception:  # noqa: BLE001
                try:
                    sub.click(timeout=5000, force=True)
                except Exception:  # noqa: BLE001
                    pass
        return None, None


def _клик_и_вердикт_отправки(page, sub) -> tuple:
    """Кликает по кнопке и решает, была ли ИМЕННО ЭТА попытка успешной.
    Возвращает (успех: bool, источник: 'ответ'|'визуально')."""
    _подтв_до = _подтверждение_видно(page)
    try:
        _btn_до = _текст_кнопки(sub, 1000)
    except Exception:  # noqa: BLE001
        _btn_до = ""

    _текст_отв, _статус_отв = _клик_поймать_ответ(page, sub)
    if _статус_отв is not None:
        return (_ответ_формы_вердикт(_текст_отв, _статус_отв) == "успешно"), "ответ"

    page.wait_for_timeout(1200)
    try:
        _btn_после = _текст_кнопки(sub, 1000)
    except Exception:  # noqa: BLE001
        _btn_после = ""
    _подтв_после = _подтверждение_видно(page)
    успех = (
        (_подтв_после and not _подтв_до)
        or bool(_btn_после and _btn_после != _btn_до
                and _текст_подтверждает_отправку(_btn_после))
    )
    return успех, "визуально"


def проба_серверной_валидации(scope, page, sub, is_order: bool,
                              снимок=None) -> dict:
    """Пункт «нельзя отправить неверные данные через DevTools» (серверная
    валидация). Вызывается ПОСЛЕ обычной легитимной отправки формы (чтобы не
    исказить остальные колонки её строки отчёта), только под галочкой.

    Форму заказа (is_order) пропускаем - та же перестраховка на чекауте, что
    и в проверка_отображения_ошибок (не делаем на нём авто-пробу).

    `снимок` - валидные значения полей, снятые ДО легитимной отправки (пока
    форма была заполнена). После отправки форма часто ОЧИЩЕНА и прикрыта
    оверлеем «Спасибо»: без восстановления проба тамперит пустую форму, сервер
    отклоняет её из-за ВСЕХ пустых полей (а не из-за нашей подмены), клик не
    даёт нового POST - и вердикт уходил в «не удалось определить». Поэтому перед
    каждой пробой возвращаем валидные значения и портим РОВНО одно поле.

    Возвращает {попытки: {вид: 'принято'|'отклонено'|'неприменимо'|
    'не удалось определить'}, детали}."""
    if is_order:
        return {"попытки": {}, "детали": "форма заказа - проба пропущена (перестраховка на чекауте)"}

    попытки = {}
    поля = {}          # вид → имя поля, которое портили (для конкретного коммента)
    for вид, _описание in _SRVVAL_ВИДЫ:
        # Вернуть валидные значения перед каждой пробой (см. docstring).
        if снимок:
            try:
                scope.evaluate(_JS_RATELIMIT_RESTORE, снимок)
            except Exception:  # noqa: BLE001
                pass
        try:
            info = scope.evaluate(_JS_SRVVAL_TAMPER, вид)
        except Exception:  # noqa: BLE001
            info = None
        if not info or not info.get("done"):
            попытки[вид] = "неприменимо"
            continue
        поля[вид] = info.get("поле")
        try:
            успех, источник = _клик_и_вердикт_отправки(page, sub)
            if успех:
                попытки[вид] = "принято"
            elif источник == "ответ" or _видна_ошибка_отправки(page):
                попытки[вид] = "отклонено"
            else:
                попытки[вид] = "не удалось определить"
        except Exception as e:  # noqa: BLE001
            попытки[вид] = "не удалось определить"
            print(f"      ⚠️ Проба серверной валидации ({вид}): {e}")
        finally:
            try:
                scope.evaluate(_JS_SRVVAL_RESTORE, info.get("old", ""))
            except Exception:  # noqa: BLE001
                pass

    детали = серверная_валидация_детали(попытки, поля)
    return {"попытки": попытки, "детали": детали, "поля": поля}


def серверная_валидация_детали(попытки: dict, поля: dict | None = None) -> str:
    """Человекочитаемая расшифровка пробы серверной валидации С УКАЗАНИЕМ ПОЛЯ -
    идёт в «Комментарий» отчёта (не путать с кратким статусом-вердиктом из
    валидация_сервера_вердикт). Для каждого проверенного вида пишем, ПРОШЛА ли
    невалидная отправка и в каком именно поле. ЧИСТАЯ функция (юнит-тест без
    браузера).

    Пример: «сервер ПРИНЯЛ невалидную отправку с некорректным e-mail (поле
    «email») - это и есть уязвимость. сервер отклонил: без имени (поле «fio»)»."""
    поля = поля or {}

    def _поле_txt(вид):
        имя = str(поля.get(вид) or "").strip()
        return f" (поле «{имя}»)" if имя else ""

    def _рус(вид):
        # Дружеское русское имя по имени поля - для «прошло без телефона».
        имя = str(поля.get(вид) or "").lower()
        if re.search(r"phone|tel|тел", имя):
            return "телефона"
        if re.search(r"mail|почт", имя):
            return "почты"
        if re.search(r"name|имя|fio|фио|фамил", имя):
            return "имени"
        return ""

    def _что(вид, описание):
        # Краткое ЧТО именно проверяли (без исхода) - исход добавит группировка.
        if вид == "empty_name":
            return f"без имени{_поле_txt(вид)}"
        if вид == "empty_phone":
            return f"без телефона{_поле_txt(вид)}"
        if вид == "empty":
            return f"без {_рус(вид) or 'обязательного поля'}{_поле_txt(вид)}"
        if вид == "bad_email":
            return f"с некорректным e-mail{_поле_txt(вид)}"
        if вид == "too_long":
            return f"со слишком длинным текстом{_поле_txt(вид)}"
        return f"{описание}{_поле_txt(вид)}"

    # Группируем по ИСХОДУ и ведём с уязвимости (что сервер ПРИНЯЛ - в этом суть
    # «УЯЗВИМА»), затем отклонённые, затем неоднозначные. Так понятнее, чем
    # вперемешку по видам.
    прошли, отклонены, неясно = [], [], []
    for вид, описание in _SRVVAL_ВИДЫ:
        рез = попытки.get(вид)
        if рез in (None, "неприменимо"):
            continue
        ч = _что(вид, описание)
        if рез == "принято":
            прошли.append(ч)
        elif рез == "отклонено":
            отклонены.append(ч)
        else:
            неясно.append(ч)
    куски = []
    if прошли:
        куски.append("сервер ПРИНЯЛ невалидную отправку " + "; ".join(прошли)
                     + " - это и есть уязвимость")
    if отклонены:
        куски.append("сервер отклонил: " + ", ".join(отклонены))
    if неясно:
        куски.append("неоднозначно (сервер не ответил чётко): "
                     + ", ".join(неясно))
    return ". ".join(куски)


def валидация_сервера_вердикт(попытки: dict) -> tuple:
    """(статус, деталь) для строки «Серверная валидация». ЧИСТАЯ функция
    (юнит-тест без браузера). Статус: «Защищена» / «УЯЗВИМА» / «Проверить»."""
    применимые = {в: р for в, р in (попытки or {}).items() if р != "неприменимо"}
    принятые = [в for в, р in применимые.items() if р == "принято"]
    if принятые:
        return "УЯЗВИМА", f"сервер принял невалидные данные: {', '.join(принятые)}"
    if применимые and all(р == "отклонено" for р in применимые.values()):
        return "Защищена", f"сервер отклонил все проверенные варианты ({len(применимые)})"
    if not применимые:
        return ("Проверить",
                "не нашли подходящих полей для пробы (нет обязательных/email/текстовых)")
    return ("Проверить",
            "нет однозначного ответа сервера на повторную отправку - форма часто "
            "блокируется после первой отправки; проверьте вручную (мусор в e-mail / "
            "пустое обязательное поле в обход подсказки)")


# ── Лимит запросов: защита от спама/ботов ────────────────────────────────
# Слой A (пассивно, всегда включён, НИ ОДНОГО лишнего запроса): признаки
# антибот-защиты в уже загруженной странице - капча-виджет по HTML и
# honeypot-поле по живому DOM. Слой B (активный залп, отдельная галочка,
# выключена по умолчанию) - ниже.
_КАПЧА_МАРКЕРЫ = (
    ("g-recaptcha", "reCAPTCHA"), ("grecaptcha", "reCAPTCHA"), ("recaptcha", "reCAPTCHA"),
    ("h-captcha", "hCaptcha"), ("hcaptcha", "hCaptcha"),
    ("cf-turnstile", "Cloudflare Turnstile"), ("turnstile", "Cloudflare Turnstile"),
    ("smartcaptcha", "Яндекс SmartCaptcha"), ("smart-captcha", "Яндекс SmartCaptcha"),
)


def защита_от_спама_из_html(html: str) -> dict:
    """{капча, какая} - есть ли в уже загруженной странице маркер капча-виджета
    (script/class/iframe узнаваемых сервисов). ЧИСТАЯ функция (юнит-тест без
    браузера). Не проверяет, что капча реально требуется при отправке - только
    что она технически присутствует."""
    low = (html or "").lower()
    for маркер, имя in _КАПЧА_МАРКЕРЫ:
        if маркер in low:
            return {"капча": True, "какая": имя}
    return {"капча": False, "какая": ""}


# JS: honeypot-поле - скрытое (type=hidden) поле с типовым «ловушечным» именем
# (как hideit у Метпромко, forms_tester/projects/metpromko/config.py). Именно
# ТИПОВЫЕ имена, не любое hidden - иначе ложно сработает на csrf/hash/sessid.
_JS_HONEYPOT = r"""
f => {
  const known = /^(hideit|honeypot|honey[-_]?pot|trap|bot[-_]?field|homepage)$/i;
  for (const e of f.querySelectorAll("input[type='hidden']")) {
    const nm = (e.name || e.id || '').trim();
    if (nm && known.test(nm)) return {найдено: true, имя: nm};
  }
  return {найдено: false, имя: ''};
}
"""


def лимит_пассивно_вердикт(обнаружено: dict) -> tuple:
    """(статус, деталь) для колонки «Защита от спама (пассивно)». ЧИСТАЯ
    функция (юнит-тест без браузера). Статус: «Есть защита» / «Не обнаружено»."""
    о = обнаружено or {}
    признаки = []
    if о.get("капча"):
        признаки.append(о.get("капча_какая") or "капча")
    if о.get("honeypot"):
        признаки.append(f"honeypot-поле «{о.get('honeypot_имя')}»")
    if признаки:
        return "Есть защита", "обнаружено: " + ", ".join(признаки)
    return ("Не обнаружено",
            "видимых признаков защиты нет (капча/honeypot) - это НЕ доказывает "
            "отсутствие лимита на сервере, только что снаружи не видно")


# Слой B: активный залп - до 3 БЫСТРЫХ повторных отправок ВАЛИДНЫМИ (не
# невалидными - здесь проверяем throttling, а не валидацию) данными. Число
# попыток сознательно не настраивается через UI (чтобы не провоцировать
# «покрутить побольше» - каждая лишняя попытка на боевом сайте лишняя).
_RATELIMIT_ПОПЫТОК = 3
_ЛИМИТ_МАРКЕРЫ = (
    "слишком часто", "слишком много попыток", "много запросов",
    "попробуйте позже", "повторите попытку позже", "too many requests",
    "too many attempts", "try again later", "rate limit",
)

# JS: снимок/восстановление значений полей между попытками - AJAX-форма часто
# ОЧИЩАЕТ поля после успешной отправки (см. «Поля очищаются» выше), без
# восстановления 2-я/3-я попытка ушла бы пустой не из-за лимита, а из-за
# нехватки данных. По образцу window.__valSaved у проверка_отображения_ошибок.
_JS_RATELIMIT_SNAPSHOT = r"""
f => {
  const skip = ['hidden','submit','button','reset','image'];
  const ctrls = [...f.querySelectorAll('input,textarea,select')]
    .filter(e => !skip.includes((e.type||'').toLowerCase()));
  return ctrls.map(e => ({v: e.value, c: e.checked}));
}
"""
_JS_RATELIMIT_RESTORE = r"""
(f, saved) => {
  const skip = ['hidden','submit','button','reset','image'];
  const ctrls = [...f.querySelectorAll('input,textarea,select')]
    .filter(e => !skip.includes((e.type||'').toLowerCase()));
  ctrls.forEach((e, i) => { if (saved[i]) { const t = (e.type||'').toLowerCase();
    if (t==='checkbox'||t==='radio') e.checked = saved[i].c; else e.value = saved[i].v;
    e.dispatchEvent(new Event('input', {bubbles:true}));
    e.dispatchEvent(new Event('change', {bubbles:true})); } });
  return true;
}
"""


def _текст_похож_на_блок_лимита(text: str) -> bool:
    """True, если видимый текст страницы похож на капча-блок (переиспользуем
    response_indicates_captcha_block) или явную фразу про лимит запросов.
    ЧИСТАЯ функция (юнит-тест без браузера)."""
    if response_indicates_captcha_block(text or ""):
        return True
    t = (text or "").lower()
    return any(m in t for m in _ЛИМИТ_МАРКЕРЫ)


def активная_проба_лимита(scope, page, sub, is_order: bool,
                          снимок=None) -> dict:
    """До 3 быстрых повторных отправок ВАЛИДНЫМИ данными - проверяем, блокирует
    ли сайт позднюю попытку (рабочая защита от спама/ботов, а не просто «на
    вид есть капча»). Форма заказа пропускается - перестраховка на чекауте
    (та же, что и у server_validation/показа ошибок валидации).

    `снимок` - валидные значения полей, снятые ДО легитимной отправки. Раньше
    снимок брался ЗДЕСЬ, но форма к этому моменту УЖЕ очищена отправкой, поэтому
    2-я/3-я попытки уходили пустыми (не из-за лимита, а из-за нехватки данных) -
    вердикт получался «неоднозначно». Теперь восстанавливаем валидные значения
    перед КАЖДОЙ попыткой, снятые ещё до очистки.

    Возвращает {попытки: [{'n', 'успех', 'блок'}, ...], детали}."""
    if is_order:
        return {"попытки": [], "детали": "форма заказа - проба пропущена (перестраховка на чекауте)"}

    saved = снимок
    if not saved:
        # Фолбэк: снимок не передан - пробуем снять сейчас (форма может быть
        # уже пустой, тогда 2-3 попытки менее надёжны - как раньше).
        try:
            saved = scope.evaluate(_JS_RATELIMIT_SNAPSHOT)
        except Exception:  # noqa: BLE001
            saved = None

    попытки = []
    for i in range(1, _RATELIMIT_ПОПЫТОК + 1):
        try:
            # Восстанавливаем валидные значения перед КАЖДОЙ попыткой (в т.ч.
            # первой: легитимная отправка выше уже очистила форму).
            if saved:
                try:
                    scope.evaluate(_JS_RATELIMIT_RESTORE, saved)
                except Exception:  # noqa: BLE001
                    pass
            успех, _источник = _клик_и_вердикт_отправки(page, sub)
            _текст = ""
            try:
                _текст = page.locator("body").inner_text(timeout=1500) or ""
            except Exception:  # noqa: BLE001
                pass
            попытки.append({
                "n": i,
                "успех": успех,
                "блок": _текст_похож_на_блок_лимита(_текст),
            })
        except Exception as e:  # noqa: BLE001
            попытки.append({"n": i, "успех": False, "блок": False, "ошибка": str(e)})

    детали = "; ".join(
        f"попытка {p['n']}: " + ("блок" if p.get("блок") else
                                  ("принято" if p.get("успех") else "неясно"))
        for p in попытки)
    return {"попытки": попытки, "детали": детали}


def лимит_активно_вердикт(попытки: list) -> tuple:
    """(статус, деталь) для строки «Защита от спама (активно)». ЧИСТАЯ функция
    (юнит-тест без браузера). Статус: «Сработала защита» / «Не сработала за N
    попытки» / «Проверить»."""
    if not попытки:
        return "Проверить", "проба не выполнялась (форма заказа или сбой)"
    заблокированные = [p for p in попытки if p.get("блок")]
    if заблокированные:
        n = заблокированные[0]["n"]
        return "Сработала защита", f"сайт заблокировал попытку №{n} из {len(попытки)}"
    if all(p.get("успех") for p in попытки):
        return (f"Не сработала за {len(попытки)} попытки",
                f"все {len(попытки)} быстрых попытки прошли одинаково успешно - "
                "лимит не сработал на этом масштабе (не значит, что его нет вообще)")
    # Первая отправка прошла, а дальше не проходят (сервер при этом не блокировал) -
    # форма САМА блокирует повтор (кнопка «Отправлено»). Это защита от быстрого/
    # случайного спама → помечаем как сработавшую (✓): серверный лимит так не
    # проверить, но с точки зрения пользователя форма защищена.
    if попытки[0].get("успех"):
        return ("Блокирует повтор",
                "кнопка становится «Отправлено» после отправки - от СЛУЧАЙНОГО "
                "двойного клика защищает, но это НЕ доказывает серверную защиту "
                "(на части форм кнопка косметическая, повторы всё равно уходят - "
                "в админке несколько заявок с одного теста). Проверьте вручную: "
                "перезагрузите страницу и отправьте 3-4 раза подряд")
    # Первая отправка не прошла - результат неоднозначный, надёжнее проверить руками.
    return ("Проверить",
            "быстрые повторы дали неоднозначный результат - проверьте вручную: "
            "перезагрузите страницу и отправьте форму 3-4 раза подряд")


def проверка_списков(scope) -> dict:
    """Пункт «Выпадающие списки открываются и корректно отображают варианты».
    Читает ВИДИМЫЕ <select> формы, значение НЕ меняет (иначе на форме заказа
    select региона/доставки триггерит пересчёт/AJAX). «Открытие» нативного select
    гарантирует сам браузер - проверяем суть: есть ли в списке варианты.
    → {состояние: 'не найдено' | 'корректно' | 'ошибка', коммент}.
    Ошибка = пустой <select> (совсем без <option>) - это однозначно битый список.
    Списки с плейсхолдером/вариантами не трогаем (могут догружаться по AJAX)."""
    try:
        sels = scope.evaluate(
            "el => [...el.querySelectorAll('select')]"
            ".filter(s => !!(s.offsetWidth||s.offsetHeight||s.getClientRects().length))"
            ".map(s => ({name:(s.name||s.id||'список'), total:s.options.length,"
            " real:[...s.options].filter(o=>(o.value||'').trim()"
            " && (o.textContent||'').trim()).length}))")
    except Exception:  # noqa: BLE001
        return {"состояние": "не найдено", "коммент": ""}
    if not sels:
        return {"состояние": "не найдено", "коммент": ""}
    битые = [s for s in sels if int(s.get("total") or 0) == 0]
    if битые:
        имена = ", ".join(f"«{s.get('name', 'список')}»" for s in битые[:4])
        return {"состояние": "ошибка",
                "коммент": (f"Выпадающий список {имена} пустой — в нём нет ни одного "
                            f"варианта для выбора.")}
    return {"состояние": "корректно", "коммент": ""}


def проверка_чекбоксов(scope) -> dict:
    """Пункт «Чекбоксы и радиокнопки работают (отмечаются/снимаются)». Проверяем
    ПО СМЫСЛУ: настоящим кликом переключаем элемент и смотрим, изменилось ли его
    состояние; ЗАТЕМ принудительно возвращаем всё как было (чтобы не сломать
    согласие/выбор перед отправкой формы).
    → {состояние: 'не найдено' | 'корректно' | 'ошибка', коммент}.
    Ошибка = клик НЕ меняет состояние (чекбокс не ставится/не снимается,
    радиокнопка не выбирается). Уже выбранную радиокнопку глубоко не трогаем -
    снять её кликом нельзя, это штатное поведение, а не баг."""
    sel = "input[type='checkbox'], input[type='radio']"
    try:
        loc = scope.locator(sel)
        n = loc.count()
    except Exception:  # noqa: BLE001
        return {"состояние": "не найдено", "коммент": ""}
    if n == 0:
        return {"состояние": "не найдено", "коммент": ""}
    # Снимок исходных состояний ВСЕХ чекбоксов/радио - вернём в конце
    # принудительно, что бы ни натворили клики (иначе можно случайно сбросить
    # согласие на обработку данных или выбранный способ оплаты).
    _q = "input[type='checkbox'],input[type='radio']"
    try:
        snapshot = scope.evaluate(
            f"el => [...el.querySelectorAll(\"{_q}\")].map(e => e.checked)")
    except Exception:  # noqa: BLE001
        snapshot = None
    проверено, сломаны = 0, []
    for i in range(min(n, 15)):
        if проверено >= 8:
            break
        el = loc.nth(i)
        try:
            info = el.evaluate(
                "e => ({t:(e.type||'').toLowerCase(),"
                " vis:!!(e.offsetWidth||e.offsetHeight||e.getClientRects().length),"
                " off:!!(e.disabled||e.readOnly), checked:!!e.checked,"
                " label:(e.name||e.id||e.getAttribute('aria-label')||'элемент')})")
        except Exception:  # noqa: BLE001
            continue
        if not info["vis"] or info["off"]:
            continue
        typ, was, label = info["t"], info["checked"], info["label"]
        try:
            if typ == "checkbox":
                el.click(timeout=2500)
                after = bool(el.evaluate("e => e.checked"))
                проверено += 1
                if after == was:
                    сломаны.append(f"чекбокс «{label}» не переключается кликом")
            else:  # radio
                if was:
                    continue                       # уже выбран - переключать нечем
                el.click(timeout=2500)
                after = bool(el.evaluate("e => e.checked"))
                проверено += 1
                if not after:
                    сломаны.append(f"радиокнопка «{label}» не выбирается кликом")
        except Exception:  # noqa: BLE001
            continue
    # Принудительно возвращаем все чекбоксы/радио в исходное состояние.
    if snapshot is not None:
        try:
            scope.evaluate(
                "(el, states) => { const nodes = el.querySelectorAll("
                f"\"{_q}\");"
                " states.forEach((v, i) => { if (nodes[i]) nodes[i].checked = v; }); }",
                snapshot)
        except Exception:  # noqa: BLE001
            pass
    if проверено == 0:
        return {"состояние": "не найдено", "коммент": ""}
    if сломаны:
        return {"состояние": "ошибка",
                "коммент": "Не реагируют на клик: " + "; ".join(сломаны[:4]) + "."}
    return {"состояние": "корректно", "коммент": ""}


def проверка_автозаполнения(scope) -> dict:
    """Пункт «Поля с автозаполнением работают корректно». Настоящее браузерное
    автозаполнение (из сохранённого профиля) в headless честно не воспроизвести,
    поэтому проверяем ОБЪЕКТИВНЫЕ признаки, от которых оно зависит - на личных
    полях (имя/почта/телефон):
      • не ОТКЛЮЧЕНО ли автозаполнение (autocomplete=off/false/nope) - тогда
        браузер не подставит сохранённые данные, пользователь вводит вручную;
      • ПРИНИМАЕТ ли поле «автозаполненное» значение - ставим значение разом
        (как автозаполнение, через .value + события input/change) и смотрим, что
        поле его не очистило. Значение возвращаем как было (отправку не ломаем).
    → {состояние: 'не найдено'|'корректно'|'отключено'|'искажается', коммент}.
    Всё в одном evaluate: атомарно, с восстановлением исходных значений."""
    try:
        r = scope.evaluate(
            "f => {"
            " const vis=e=>{const b=e.getBoundingClientRect();const s=getComputedStyle(e);"
            "   return b.width>0&&b.height>0&&s.visibility!=='hidden'&&s.display!=='none';};"
            " const off=new Set(['off','false','nope','none','disabled']);"
            " const kind=e=>{const t=(e.type||'').toLowerCase();"
            "   const s=((e.name||'')+' '+(e.getAttribute('autocomplete')||'')+' '"
            "     +(e.placeholder||'')+' '+(e.id||'')).toLowerCase();"
            "   if(t==='email'||/mail|почт/.test(s)) return 'email';"
            "   if(t==='tel'||/phone|tel|телеф/.test(s)) return 'phone';"
            "   if(/name|имя|фио|\\bfio|фамил/.test(s)) return 'name'; return null;};"
            " const vals={name:'Иван Петров',email:'test.autofill@example.ru',phone:'+79991234567'};"
            " const els=[...f.querySelectorAll('input,textarea')].filter(e=>{"
            "   const t=(e.type||'').toLowerCase();"
            "   return !['hidden','submit','button','checkbox','radio','file','password'].includes(t)"
            "     && vis(e) && !e.disabled && !e.readOnly;});"
            " const отключены=[], искажены=[]; let проверено=0;"
            " for(const e of els){ const k=kind(e); if(!k) continue; проверено++;"
            "   const label=(e.name||e.placeholder||e.id||k);"
            "   const ac=(e.getAttribute('autocomplete')||'').trim().toLowerCase();"
            "   if(off.has(ac)) отключены.push(label);"
            "   const orig=e.value;"
            "   try{ e.focus(); e.value=vals[k];"
            "     e.dispatchEvent(new Event('input',{bubbles:true}));"
            "     e.dispatchEvent(new Event('change',{bubbles:true}));"
            "     if(((e.value||'').trim())==='') искажены.push(label);"
            "   }catch(_){}"
            "   finally{ try{ e.value=orig;"
            "     e.dispatchEvent(new Event('input',{bubbles:true})); }catch(_){} } }"
            " return {проверено, отключены:[...new Set(отключены)].slice(0,4),"
            "   искажены:[...new Set(искажены)].slice(0,4)};"
            "}")
    except Exception:  # noqa: BLE001
        return {"состояние": "не найдено", "коммент": ""}
    if not r or int(r.get("проверено") or 0) == 0:
        return {"состояние": "не найдено", "коммент": ""}
    if r.get("отключены"):
        return {"состояние": "отключено",
                "коммент": ("Автозаполнение выключено (autocomplete=off) на полях: "
                            + ", ".join(f"«{x}»" for x in r["отключены"])
                            + " - браузер не подставит сохранённые имя/почту/телефон, "
                            "пользователь вводит вручную.")}
    if r.get("искажены"):
        return {"состояние": "искажается",
                "коммент": ("Поле очищает автозаполненное значение (не принимает ввод "
                            "разом): " + ", ".join(f"«{x}»" for x in r["искажены"]) + ".")}
    return {"состояние": "корректно", "коммент": ""}


def проверка_подсказок(scope) -> dict:
    """Пункт «Подсказки в полях (placeholder, help text) соответствуют назначению».
    «Хорошо ли сформулировано» - оценивает человек; мы проверяем ОБЪЕКТИВНОЕ:
    подсказка НЕ ПРОТИВОРЕЧИТ назначению поля (ловим copy-paste: у «Телефона»
    placeholder про e-mail и т.п.). Назначение поля (P) берём из type/autocomplete/
    name/label - НЕ из placeholder; на что похожа подсказка (Q) - из placeholder+
    help-text. Флажим только пары из «списка противоречий» и только при уверенных
    P и Q. Только чтение DOM (ничего не меняем).
    → {состояние: 'не найдено'|'корректно'|'несоответствие'|'нет подсказок', коммент}."""
    try:
        r = scope.evaluate(
            "f => {"
            " const vis=e=>{const b=e.getBoundingClientRect();const s=getComputedStyle(e);"
            "   return b.width>0&&b.height>0&&s.visibility!=='hidden'&&s.display!=='none';};"
            " const norm=s=>(s||'').toLowerCase().replace(/ё/g,'е').trim();"
            " const cat=txt=>{ const s=norm(txt); const raw=(txt||'').trim(); if(!s) return null;"
            "   if(s.includes('@')||/mail|почт|e-?mail|электрон/.test(s)) return 'EMAIL';"
            "   if(/телеф|phone|\\bтел\\b|\\bмоб/.test(s)||/^[\\s+\\d()\\-_]{6,}$/.test(raw)) return 'PHONE';"
            "   if(/фамил|\\bимя\\b|\\bфио\\b|как вас зов|ваше имя|контактн.{0,3}лиц/.test(s)) return 'NAME';"
            "   if(/сообщ|вопрос|коммент|что вас интерес|message|напишите/.test(s)) return 'COMMENT';"
            "   if(/адрес|улиц|street/.test(s)) return 'ADDRESS';"
            "   if(/\\bгород\\b|\\bcity\\b/.test(s)) return 'CITY';"
            "   if(/дд[.\\/]мм|\\bдата\\b|\\bdate\\b/.test(s)) return 'DATE'; return null; };"
            " const labelText=e=>{ let t=''; if(e.id){ const l=document.querySelector('label[for=\"'"
            "   +CSS.escape(e.id)+'\"]'); if(l) t=l.innerText; } if(!t){ const p=e.closest('label'); if(p) t=p.innerText; } return t; };"
            " const helpText=e=>{ let t=''; const db=e.getAttribute('aria-describedby');"
            "   if(db){ for(const id of db.split(/\\s+/)){ const h=document.getElementById(id); if(h) t+=' '+h.innerText; } }"
            "   const c=e.closest('.form-group,.field,.form-field,.input-wrap,.form-row')||e.parentElement;"
            "   if(c){ const h=c.querySelector('.hint,.help,.help-text,.form-text,.field-desc,.field-hint,small'); if(h&&h!==e) t+=' '+h.innerText; } return t; };"
            " const field=e=>{ const t=(e.type||'').toLowerCase();"
            "   if(t==='email') return 'EMAIL'; if(t==='tel') return 'PHONE'; if(t==='date') return 'DATE';"
            "   const ac=norm(e.getAttribute('autocomplete'));"
            "   if(/email/.test(ac)) return 'EMAIL'; if(/tel/.test(ac)) return 'PHONE';"
            "   if(/name|given-name|family-name/.test(ac)) return 'NAME';"
            "   if(/street-address|address-line/.test(ac)) return 'ADDRESS';"
            "   const nm=norm((e.name||'')+' '+(e.id||''));"
            "   if(/mail/.test(nm)) return 'EMAIL'; if(/phone|tel|телеф/.test(nm)) return 'PHONE';"
            "   if(/comment|сообщ|коммент|message|вопрос/.test(nm)) return 'COMMENT';"
            "   if(/name|имя|фио|\\bfio|фамил/.test(nm)) return 'NAME';"
            "   if(/\\bгород|\\bcity/.test(nm)) return 'CITY'; if(/адрес|address/.test(nm)) return 'ADDRESS';"
            "   if(/компан|company|организ/.test(nm)) return 'COMPANY';"
            "   return cat(labelText(e)); };"
            " const bad={PHONE:['EMAIL','NAME'],EMAIL:['PHONE','NAME'],NAME:['EMAIL','PHONE'],"
            "   COMMENT:['EMAIL','PHONE'],ADDRESS:['EMAIL','PHONE'],CITY:['EMAIL','PHONE']};"
            " const RUS={EMAIL:'почта',PHONE:'телефон',NAME:'имя',COMMENT:'сообщение',ADDRESS:'адрес',CITY:'город',COMPANY:'компания',DATE:'дата'};"
            " const els=[...f.querySelectorAll('input,textarea')].filter(e=>{ const t=(e.type||'').toLowerCase();"
            "   return !['hidden','submit','button','checkbox','radio','file','password','image','reset'].includes(t) && vis(e); });"
            " const несоотв=[]; let проверено=0, без_n=0; const без=[];"
            " for(const e of els){ if(проверено>=20) break; проверено++;"
            "   const P=field(e); const ph=e.placeholder||''; const hint=(ph+' '+helpText(e)).trim();"
            "   const label=(e.name||e.id||ph||'поле');"
            "   if(!hint){ без_n++; if(без.length<6) без.push(label); continue; }"
            "   const Q=cat(hint);"
            "   if(P&&Q&&P!==Q&&(bad[P]||[]).includes(Q)){"
            "     несоотв.push(`«${label}» (поле для «${RUS[P]}», а подсказка про «${RUS[Q]}»: ${ph.slice(0,28)})`); } }"
            " return {проверено, несоотв:[...new Set(несоотв)].slice(0,4), без, без_n}; }")
    except Exception:  # noqa: BLE001
        return {"состояние": "не найдено", "коммент": ""}
    if not r or int(r.get("проверено") or 0) == 0:
        return {"состояние": "не найдено", "коммент": ""}
    if r.get("несоотв"):
        return {"состояние": "несоответствие",
                "коммент": ("Подсказка не соответствует назначению поля: "
                            + "; ".join(r["несоотв"]) + ".")}
    if int(r.get("без_n") or 0) >= int(r.get("проверено") or 0):
        return {"состояние": "нет подсказок",
                "коммент": ("Ни у одного поля формы нет placeholder/подсказки"
                            + (" (" + ", ".join(f"«{x}»" for x in r["без"]) + ")" if r.get("без") else "")
                            + " - поля опираются только на подписи, если они есть.")}
    return {"состояние": "корректно", "коммент": ""}


# JS: снимок + очистка видимых текстовых полей (реф. и значения - на window).
_REQ_CLEAR_JS = (
    "f => { const vis=e=>{const b=e.getBoundingClientRect();const s=getComputedStyle(e);"
    " return b.width>0&&b.height>0&&s.visibility!=='hidden'&&s.display!=='none';};"
    " const els=[...f.querySelectorAll('input,textarea')].filter(e=>{const t=(e.type||'').toLowerCase();"
    " return !['hidden','submit','button','checkbox','radio','file'].includes(t)&&vis(e);});"
    " window.__reqEls=els; window.__reqSaved=els.map(e=>e.value);"
    " els.forEach(e=>{ e.value=''; e.dispatchEvent(new Event('input',{bubbles:true}));"
    "   e.dispatchEvent(new Event('change',{bubbles:true}));"
    "   e.dispatchEvent(new Event('keyup',{bubbles:true})); }); return els.length; }")
# JS: восстановить значения из снимка (идемпотентно).
_REQ_RESTORE_JS = (
    "f => { const els=window.__reqEls||[]; const s=window.__reqSaved||[];"
    " els.forEach((e,i)=>{ e.value=s[i]; e.dispatchEvent(new Event('input',{bubbles:true}));"
    "   e.dispatchEvent(new Event('change',{bubbles:true}));"
    "   e.dispatchEvent(new Event('keyup',{bubbles:true})); }); }")


def проверка_кнопки_обязательные(scope, page, кнопка_css: str = "") -> dict:
    """Пункт «Кнопка отправки активна только после заполнения обязательных полей».
    Проверяем ПО СМЫСЛУ: очищаем поля → смотрим кнопку; возвращаем значения →
    смотрим кнопку. Значения ВОССТАНАВЛИВАЕМ гарантированно (finally), с паузами -
    чтобы поймать и синхронные, и слегка отложенные валидаторы.
    → {состояние, коммент}:
      корректно        - пусто: кнопка заблокирована, заполнено: активна;
      не блокируется   - кнопка активна и при пустой форме (валидация по клику);
      не разблокируется- кнопка заблокирована и ПОСЛЕ заполнения (баг - не отправить)."""
    sel = (кнопка_css or "button[type='submit'], input[type='submit'], button.btn").strip()
    try:
        btn = scope.locator(sel).first
        if btn.count() == 0:
            return {"состояние": "не найдено", "коммент": ""}
    except Exception:  # noqa: BLE001
        return {"состояние": "не найдено", "коммент": ""}

    def _disabled():
        try:
            return bool(btn.evaluate(
                "b => !!(b.disabled || b.getAttribute('aria-disabled')==='true'"
                " || /disabl|inactive|not-?allowed/i.test(b.className||'')"
                " || getComputedStyle(b).pointerEvents==='none')"))
        except Exception:  # noqa: BLE001
            return False

    empty_dis = filled_dis = None
    _cleared = False
    try:
        n = scope.evaluate(_REQ_CLEAR_JS)
        if not n:
            return {"состояние": "не найдено", "коммент": ""}
        _cleared = True
        page.wait_for_timeout(450)
        empty_dis = _disabled()
        scope.evaluate(_REQ_RESTORE_JS)
        page.wait_for_timeout(450)
        filled_dis = _disabled()
    except Exception:  # noqa: BLE001
        empty_dis = None
    finally:
        if _cleared:
            try:
                scope.evaluate(_REQ_RESTORE_JS)     # идемпотентно - точно вернём как было
            except Exception:  # noqa: BLE001
                pass
            try:
                page.evaluate("() => { delete window.__reqEls; delete window.__reqSaved; }")
            except Exception:  # noqa: BLE001
                pass

    if empty_dis is None or filled_dis is None:
        return {"состояние": "не найдено", "коммент": ""}
    if empty_dis and not filled_dis:
        return {"состояние": "корректно", "коммент": ""}
    if empty_dis and filled_dis:
        return {"состояние": "не разблокируется",
                "коммент": ("Кнопка отправки заблокирована ДАЖЕ после заполнения "
                            "обязательных полей - форму нельзя отправить (баг).")}
    return {"состояние": "не блокируется",
            "коммент": ("Кнопка «Отправить» активна ДАЖЕ на пустой форме. По "
                        "требованию она должна быть НЕактивной, пока не заполнены "
                        "обязательные поля. Не критично (форма может валидировать по "
                        "клику и показывать подсказки), но это знак внимания.")}


# ── Двойная отправка (двойной клик по кнопке) ────────────────────────
_DS_ТРЕКЕРЫ = ("mc.yandex", "metri", "google-analytics", "googletagmanager",
               "doubleclick", "top-fwz1", "vk.com", "facebook.", "criteo",
               "/collect", "stat.", "counter", "analytics")


def _ds_это_трекер(url: str) -> bool:
    """POST к аналитике/счётчикам (Метрика, GA и т.п.) - это НЕ отправка формы,
    в подсчёте заявок его игнорируем."""
    u = (url or "").lower()
    return any(t in u for t in _DS_ТРЕКЕРЫ)


def _ds_похоже_на_заказ(url: str, action: str) -> bool:
    """Похоже ли на форму ОФОРМЛЕНИЯ ЗАКАЗА по URL/action (тогда двойную отправку
    проверяем БЕЗОПАСНО - один клик, чтобы не создать второй заказ). Основной
    признак - тип блока «Оформление…» (передаётся отдельно); это запасная сетка.
    Слово «заказ» из названия НЕ берём: его содержит и «Заказать звонок»."""
    s = ((url or "") + " " + (action or "")).lower()
    return any(k in s for k in ("onepagecheckout", "/checkout", "checkout/",
                                "saveorder", "opc_submit", "/order", "/cart",
                                "korzin"))


def _видна_ошибка_отправки(page) -> bool:
    """Видит ли пользователь СООБЩЕНИЕ ОБ ОШИБКЕ на странице (после упавшей
    отправки): видимый элемент с error-классом / role=alert или короткий видимый
    текст с маркером ошибки. Чистая проверка DOM (ничего не меняет)."""
    try:
        return bool(page.evaluate(
            "() => {"
            " const vis = el => { const r=el.getBoundingClientRect();"
            "   const s=getComputedStyle(el);"
            "   return r.width>0 && r.height>0 && s.visibility!=='hidden'"
            "   && s.display!=='none' && s.opacity!=='0'; };"
            " for (const el of document.querySelectorAll("
            "   '[class*=error i],[class*=fail i],[role=alert],.alert-danger,"
            "    .text-danger,.form-error,.has-error,.invalid-feedback')) {"
            "   if (vis(el) && (el.innerText||'').trim()) return true; }"
            " const words=['ошибк','не удалось','не отправ','попробуйте','повторите',"
            "   'проверьте','что-то пошло','failed','error'];"
            " for (const el of document.querySelectorAll('div,span,p,li,label,strong,small,b')) {"
            "   if (!vis(el)) continue; const t=(el.innerText||'').trim().toLowerCase();"
            "   if (t && t.length<140 && words.some(w=>t.includes(w))) return true; }"
            " return false; }"))
    except Exception:  # noqa: BLE001
        return False


# Типы для пробы серверной фильтрации загрузки: (расширение, опасное?).
# Контент БЕЗВРЕДНЫЙ (не эксплойт) - меняем только расширение. «Опасные» -
# исполняемые/скриптовые/веб (могут стать вектором атаки): их приём сервером
# = находка. Обычные (документы/картинки/архивы) приём = норма, просто
# показываем в списке.
_ПРОБА_ТИПЫ = [
    # опасные: скрипты/исполняемые/веб
    (".php", True), (".phtml", True), (".html", True), (".htm", True),
    (".js", True), (".svg", True),   # SVG может содержать JS - частый вектор
    (".exe", True), (".sh", True), (".bat", True), (".py", True),
    (".jar", True), (".dll", True), (".msi", True),
    # обычные: документы
    (".csv", False), (".doc", False), (".docx", False), (".xlsx", False), (".pptx", False),
    (".pdf", False), (".rtf", False), (".txt", False), (".md", False),
    (".json", False), (".yaml", False),
    # обычные: архивы
    (".zip", False), (".tar", False),
    # обычные: картинки
    (".jpg", False), (".jpeg", False), (".png", False), (".gif", False),
    (".webp", False), (".bmp", False),
]
# Признаки, что сервер ОТКЛОНИЛ файл по типу (валидация есть - это хорошо).
_ФАЙЛ_ОТКЛОНЁН = (
    "недопустим", "разрешены только", "разрешён", "разрешен", "неверн",
    "запрещен", "запрещён", "invalid", "not allowed", "формат файла",
    "тип файла", "только изображени", "unsupported", "wrong file",
    "недопустимый формат", "нельзя загружать",
)
# Признаки, что сервер отклонил файл по РАЗМЕРУ (лимит есть - это хорошо).
_ФАЙЛ_ВЕЛИК = (
    "слишком большой", "превышает", "превышен размер", "размер файла",
    "не более", "макс. размер", "максимальный размер", "too large",
    "file too big", "maximum size", "exceeds", "20 кб", "20кб", "20 kb",
)
# Пункт «разрешены только PDF/DOC/DOCX и размер до 20 КБ»: ЧТО сервер должен
# принимать. Всё остальное (другой формат ИЛИ больше лимита) должно
# отклоняться. Бизнес-правило вынесено в константы - легко поменять.
_ФАЙЛ_РАЗРЕШЁННЫЕ = (".pdf", ".doc", ".docx")
_ФАЙЛ_МАКС_КБ = 20


def _безвредный_файл(ext: str, размер_байт: int = 0):
    """Создаёт временный БЕЗВРЕДНЫЙ файл с расширением ext (кэшируется).
    Возвращает путь. SVG - без скриптов; остальные - обычный текст. Если
    размер_байт>0 - дополняет файл до этого размера (проба лимита размера);
    кэш отдельный по (ext, размер)."""
    import tempfile
    d = getattr(_безвредный_файл, "_dir", None)
    if d is None:
        d = tempfile.mkdtemp(prefix="fileprobe_")
        _безвредный_файл._dir = d
    имя = "test" + ext if not размер_байт else f"test_{размер_байт}{ext}"
    path = os.path.join(d, имя)
    if not os.path.exists(path):
        if ext == ".svg":
            data = ('<svg xmlns="http://www.w3.org/2000/svg" width="1" '
                    'height="1"></svg>')
        else:
            data = ("SECURITY TEST (benign, no payload) - "
                    "checking server-side file-type filtering.")
        # Добить до нужного размера безвредным паддингом (1 байт = 1 символ).
        не_хватает = размер_байт - len(data.encode("utf-8"))
        if не_хватает > 0:
            data += "A" * не_хватает
        with open(path, "w", encoding="utf-8") as fh:
            fh.write(data)
    return path


def проба_загрузки_файла(scope, page) -> dict:
    """Проба серверной фильтрации: по каждому типу из _ПРОБА_ТИПЫ грузим
    БЕЗВРЕДНЫЙ файл и отправляем форму. Форма подтвердила отправку без
    ошибки о типе = сервер ПРИНЯЛ этот тип. Опасный принят = находка.
    Идём, пока поле загрузки остаётся на странице (AJAX-формы не уходят -
    можно проверить все типы; форма с переходом закроется - проверим,
    сколько успеем).

    ВНИМАНИЕ: каждый принятый тип = отдельная тест-заявка в админке (спам
    на боевых формах) - потому проба за отдельной галочкой.

    Возвращает {было_поле, принятые_опасные, принятые_обычные, отклонённые,
    проверено, детали}."""
    res = {"было_поле": False, "принятые_опасные": [], "принятые_обычные": [],
           "отклонённые": [], "проверено": 0, "детали": "",
           "большой_принят": None}
    if scope.locator("input[type='file']").count() == 0:
        return res
    res["было_поле"] = True
    for ext, опасный in _ПРОБА_ТИПЫ:
        fi = scope.locator("input[type='file']")
        if fi.count() == 0:            # форма закрылась после отправки
            break
        try:
            fi.first.set_input_files(_безвредный_файл(ext), timeout=5000)
        except Exception:  # noqa: BLE001
            continue
        res["проверено"] += 1
        try:
            _click_modal_submit(scope)
        except Exception:  # noqa: BLE001
            pass
        page.wait_for_timeout(3000)
        txt = ""
        try:
            txt = (page.inner_text("body", timeout=1500) or "").lower()
        except Exception:  # noqa: BLE001
            pass
        отклонён = any(m in txt for m in _ФАЙЛ_ОТКЛОНЁН)
        успех = _текст_подтверждает_отправку(txt)
        if успех and not отклонён:
            (res["принятые_опасные"] if опасный
             else res["принятые_обычные"]).append(ext)
        elif отклонён:
            res["отклонённые"].append(ext)
        # неясный исход тоже не прерывает - пробуем следующий тип, пока
        # поле на месте (в начале цикла проверяем fi.count()).

    # Проба ЛИМИТА размера: грузим РАЗРЕШЁННЫЙ тип (pdf), но БОЛЬШЕ лимита -
    # сервер должен отклонить. Принял → лимит не проверяется. Best-effort:
    # если поля уже нет (форма ушла) - оставляем None (не проверено).
    try:
        fi = scope.locator("input[type='file']")
        if fi.count():
            fi.first.set_input_files(
                _безвредный_файл(".pdf", (_ФАЙЛ_МАКС_КБ + 5) * 1024), timeout=5000)
            try:
                _click_modal_submit(scope)
            except Exception:  # noqa: BLE001
                pass
            page.wait_for_timeout(3000)
            _txt = ""
            try:
                _txt = (page.inner_text("body", timeout=1500) or "").lower()
            except Exception:  # noqa: BLE001
                pass
            _велик = (any(m in _txt for m in _ФАЙЛ_ОТКЛОНЁН)
                      or any(m in _txt for m in _ФАЙЛ_ВЕЛИК))
            if _текст_подтверждает_отправку(_txt) and not _велик:
                res["большой_принят"] = True      # приняли большой - лимита нет
            elif _велик:
                res["большой_принят"] = False     # отклонили - лимит есть
    except Exception:  # noqa: BLE001
        pass

    _оп = res["принятые_опасные"]
    _об = res["принятые_обычные"]
    _всего = len(_ПРОБА_ТИПЫ)
    _n = res["проверено"]
    # Форма ушла/сбросилась после отправки - проверить успели не все типы.
    _частично = (_n < _всего)
    _хвост = (f" (проверено {_n} из {_всего} - форма закрылась после "
              f"отправки, остальные типы не проверены)" if _частично else "")
    if _оп:
        res["детали"] = ("сервер ПРИНЯЛ опасные типы: " + ", ".join(_оп)
                         + " - фильтрации нет"
                         + ("; также принял: " + ", ".join(_об) if _об else ""))
    elif _об:
        res["детали"] = ("опасные типы отклонены; принимает обычные: "
                         + ", ".join(_об) + _хвост)
    elif res["отклонённые"]:
        res["детали"] = ("из проверенных сервер отклонил все ("
                         + ", ".join(res["отклонённые"]) + ")" + _хвост)
    else:
        res["детали"] = "результат неясен (форма не подтвердила отправку)" + _хвост
    return res


def фильтр_файлов_вердикт(res: dict) -> tuple:
    """(значение_колонки, детали) по правилу «принимать только PDF/DOC/DOCX и
    размер до 20 КБ». ЧИСТАЯ функция (юнит-тест без браузера) - на вход результат
    проба_загрузки_файла. Значение колонки подобрано под правила матрицы: ✗ на
    нарушении, ✓ на «корректно», ⚠ на «проверить»."""
    оп = [str(e) for e in (res.get("принятые_опасные") or [])]
    об = [str(e) for e in (res.get("принятые_обычные") or [])]
    вне = [e for e in об if e.lower() not in _ФАЙЛ_РАЗРЕШЁННЫЕ]
    разрешённые_принятые = [e for e in об if e.lower() in _ФАЙЛ_РАЗРЕШЁННЫЕ]
    большой = res.get("большой_принят")

    проблемы = []
    if оп:
        проблемы.append("приняты ОПАСНЫЕ типы: " + ", ".join(оп))
    if вне:
        проблемы.append("приняты посторонние типы: " + ", ".join(вне)
                        + " (по правилу разрешены только PDF/DOC/DOCX)")
    if большой is True:
        проблемы.append(f"принят файл больше {_ФАЙЛ_МАКС_КБ} КБ - лимит размера "
                        "не проверяется")
    if проблемы:
        return "✗ " + "; ".join(проблемы), "; ".join(проблемы)

    # Нарушений формата/размера не найдено.
    if not об and not оп:
        return ("проверить: ни один тип не подтверждён",
                "форма не подтвердила отправку ни одного файла - проверьте "
                f"вручную: PDF/DOC/DOCX до {_ФАЙЛ_МАКС_КБ} КБ должен пройти, "
                f"больше {_ФАЙЛ_МАКС_КБ} КБ и посторонний формат - отклониться.")
    подтв = ("принимает " + ", ".join(разрешённые_принятые)
             if разрешённые_принятые else "посторонние форматы отклонены")
    if большой is False:
        return (f"корректно: только PDF/DOC/DOCX ≤{_ФАЙЛ_МАКС_КБ} КБ",
                f"правило соблюдено: {подтв}; файл больше {_ФАЙЛ_МАКС_КБ} КБ "
                "отклонён.")
    # Формат ок, но размер автоматически проверить не удалось (форма ушла).
    return ("проверить размер: формат корректен (PDF/DOC/DOCX)",
            f"по формату правило соблюдено ({подтв}), но лимит {_ФАЙЛ_МАКС_КБ} КБ "
            f"автоматически не проверен - проверьте вручную: приложите PDF больше "
            f"{_ФАЙЛ_МАКС_КБ} КБ, он должен отклониться.")


def _извлечь_цели_из_запроса(url: str, body: str = "") -> list:
    """Имена целей Метрики из запроса reachGoal. Цель уходит на mc.yandex.* как
    page-url=goal://<домен>/<цель>. Это бывает и в URL (обычный GET), и в ТЕЛЕ
    запроса (POST / navigator.sendBeacon) - раньше смотрели только URL, из-за чего
    цели, уходящие POST-ом (напр. findtome при отправке формы), не ловились.
    Проверяем оба места. Чистая функция - легко тестируется."""
    if "mc.yandex" not in (url or "") and "mc.webvisor" not in (url or ""):
        return []
    hay = unquote(url or "") + " " + unquote(body or "")
    return re.findall(r"goal://[^/]+/([^&\s\"?#]+)", hay)


def _form_field_map_from_config(форма_config: dict) -> dict | None:
    """Явная карта полей: HTML name → токен источника (ПОЧТА, ТЕЛЕФОН, …) или буквальная строка."""
    raw = форма_config.get("поля") or форма_config.get("fields")
    if not isinstance(raw, dict) or not raw:
        return None
    out = {}
    for k, v in raw.items():
        nk = str(k).strip()
        if nk:
            out[nk] = v
    return out or None


def _normalize_field_token_key(s: str) -> str:
    return str(s).strip().upper().replace(" ", "_")


def _resolve_form_field_token(
    token,
    *,
    имя_теста: str,
    телефон: str,
    почта: str,
    имя: str,
    комментарий: str,
    город: str,
) -> str:
    """
    Значение токена из «поля» формы.
    Известные ключи: ПОЧТА, ТЕЛЕФОН, ИМЯ, КОММЕНТАРИЙ, ГОРОД, ИМЯ_ТЕСТА (и email/phone/…).
    Любая другая непустая строка - буквальное значение для поля.
    """
    if token is None:
        return ""
    s0 = str(token).strip()
    if not s0:
        return ""
    aliases = {
        "EMAIL": "ПОЧТА",
        "PHONE": "ТЕЛЕФОН",
        "NAME": "ИМЯ",
        "COMMENT": "КОММЕНТАРИЙ",
        "CITY": "ГОРОД",
        "TEST_NAME": "ИМЯ_ТЕСТА",
    }
    key = _normalize_field_token_key(s0)
    if key in aliases:
        key = aliases[key]
    mapping = {
        "ПОЧТА": (почта or "").strip(),
        "ТЕЛЕФОН": (телефон or "").strip(),
        "ИМЯ": (имя or "").strip(),
        "КОММЕНТАРИЙ": (комментарий or "").strip(),
        "ГОРОД": (город or "").strip(),
        "ИМЯ_ТЕСТА": (имя_теста or "").strip(),
    }
    if key in mapping:
        return mapping[key]
    return s0


def _apply_container_expand(scope, форма_config: dict):
    """
    расширить_контейнер / expand_container: row | form | none (пусто - без расширения).
    Например, css попал на .bx-soa-customer, а textarea - в соседней колонке: укажите row или css: div.row.
    """
    mode = (
        форма_config.get("расширить_контейнер") or форма_config.get("expand_container") or ""
    ).strip().lower()
    if mode in ("", "none", "нет", "no"):
        return scope
    if mode in ("row", "строка"):
        row = scope.locator(
            "xpath=ancestor::div[contains(concat(' ', normalize-space(@class), ' '), ' row ')][1]"
        )
        if row.count() > 0:
            return row.first
        return scope
    if mode in ("form", "форма"):
        f = scope.locator("xpath=ancestor::form[1]")
        if f.count() > 0:
            return f.first
        return scope
    return scope


def _pw_fill_named_field(scope, name_attr: str, value: str) -> bool:
    """Заполняет input/textarea/select по атрибуту name."""
    if not (value or "").strip():
        return False
    v = value.strip()
    esc = name_attr.replace("\\", "\\\\").replace('"', '\\"')
    loc = scope.locator(f'[name="{esc}"]')
    n = loc.count()
    if n == 0:
        return False
    # bx-soa (и др.) держат СКРЫТЫЕ копии полей с тем же name (шаблоны). Берём
    # ВИДИМУЮ копию, иначе заполняли скрытую - на экране поле оставалось пустым.
    el = None
    for _j in range(min(n, 12)):
        cand = loc.nth(_j)
        try:
            if cand.is_visible():
                el = cand
                break
        except Exception:
            continue
    if el is None:
        el = loc.first
    try:
        el.wait_for(state="visible", timeout=12000)
    except Exception:
        pass
    try:
        el.scroll_into_view_if_needed()
    except Exception:
        pass
    try:
        tag = el.evaluate("e => e.tagName.toLowerCase()")
    except Exception:
        tag = "input"
    try:
        if tag == "select":
            try:
                el.select_option(value=v)
            except Exception:
                el.select_option(label=v)
        else:
            el.fill(v, force=True)
        return True
    except Exception as e:
        print(f"      ⚠️ поле name={name_attr!r}: {e}")
        return False


def _browser_headers(url: str) -> dict:
    return {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        "Referer": url,
    }


def _ajax_post_headers(url: str, submit_url: str) -> dict:
    """Заголовки как у браузерного AJAX (Bitrix /local/ajax/form.php и т.п.)."""
    h = _browser_headers(url)
    if "/ajax/" in submit_url or "form.php" in submit_url:
        h["X-Requested-With"] = "XMLHttpRequest"
        h["Accept"] = "application/json, text/javascript, */*; q=0.01"
    return h


def extract_form_security_from_html(html: str) -> dict:
    """
    Вытаскивает hash/sessid из HTML и скриптов (часто пусто в разметке, задаётся JS).
    """
    out = {}
    for key, patterns in (
        (
            "hash",
            (
                r'name=["\']hash["\'][^>]*value=["\']([^"\']*)["\']',
                r'["\']hash["\']\s*:\s*["\']([^"\']+)["\']',
                r"#hash['\"]\s*\)\s*\.val\(['\"]([^'\"]+)['\"]",
            ),
        ),
        (
            "sessid",
            (
                r'bitrix_sessid["\']?\s*[:=]\s*["\']([^"\']+)',
                r'["\']sessid["\']\s*:\s*["\']([^"\']+)["\']',
            ),
        ),
    ):
        for pat in patterns:
            m = re.search(pat, html, re.I | re.DOTALL)
            if m and m.group(1).strip():
                out[key] = m.group(1).strip()
                break
    return out


# ── CSRF: наличие токена/поля защиты сессии (если требуется) ────────────
# Ищем СКРЫТОЕ поле формы с именем, характерным для CSRF/session-токена
# (не только Bitrix sessid - АПС/ИМП на нём явно не подтверждены, поэтому
# вокабуляр общий). Читаем ЖИВОЙ DOM формы (form.evaluate), а не статический
# HTML: на Bitrix-сайтах значение часто проставляется JS уже после загрузки
# страницы (см. docstring extract_form_security_from_html выше) - к моменту
# вызова (после заполнения полей, перед кликом submit) оно уже должно стоять.
_CSRF_ПОЛЕ_JS = r"""
f => {
  const re = /csrf|sessid|_token|authenticity_token|requestverificationtoken|xsrf|nonce/i;
  for (const e of f.querySelectorAll("input[type='hidden']")) {
    const nm = e.name || e.id || "";
    if (re.test(nm)) {
      return {найдено: true, заполнено: !!(e.value && e.value.trim()), имя: nm};
    }
  }
  return {найдено: false, заполнено: false, имя: ""};
}
"""


def _найти_csrf_поле(form) -> dict:
    """{найдено, заполнено, имя, ошибка} для скрытого CSRF/sessid-подобного
    поля формы. ошибка=True - DOM прочитать не удалось (это НЕ значит «токена
    нет»: для отчёта такой случай уходит в «Проверить», а не в «Нет»)."""
    try:
        r = form.evaluate(_CSRF_ПОЛЕ_JS)
        if isinstance(r, dict):
            r.setdefault("ошибка", False)
            return r
    except Exception:  # noqa: BLE001
        pass
    return {"найдено": False, "заполнено": False, "имя": "", "ошибка": True}


# Имена cookie, которые несут авторизацию/сессию (именно их подделывает CSRF).
# Аналитические/технические cookie (ga/ym/…) для CSRF не важны - их не считаем.
_CSRF_СЕССИОННАЯ_RE = re.compile(
    r"sessid|session|phpsessid|bitrix_sm|bx_|auth|token|_identity|remember|login|user",
    re.I)


def _samesite_защищает(cookie: dict) -> bool:
    """True, если cookie НЕ отправится при межсайтовом POST (значит, защищает от
    CSRF). Браузеры с 2020 трактуют отсутствие атрибута как Lax, поэтому
    «незащищённой» считаем ТОЛЬКО явную SameSite=None. Lax/Strict/пусто → защищает."""
    ss = str((cookie or {}).get("sameSite") or "").strip().lower()
    return ss != "none"


def csrf_куки_инфо(cookies: list) -> dict:
    """Классифицирует cookie сайта для оценки CSRF. ЧИСТАЯ функция (юнит-тест
    без браузера). → {сессионные_есть, все_защищены_samesite, незащищённые,
    детали}. «Защищённая» = SameSite Lax/Strict/по умолчанию (см. выше)."""
    сессионные = [c for c in (cookies or [])
                  if _CSRF_СЕССИОННАЯ_RE.search(str((c or {}).get("name") or ""))]
    if not сессионные:
        return {"сессионные_есть": False, "все_защищены_samesite": False,
                "незащищённые": [], "детали": "сессионных cookie нет"}
    незащ = [str(c.get("name")) for c in сессионные if not _samesite_защищает(c)]
    имена = ", ".join(sorted({str(c.get("name")) for c in сессионные})[:4])
    return {"сессионные_есть": True, "все_защищены_samesite": not незащ,
            "незащищённые": незащ, "детали": f"session-cookie: {имена}"}


def _прочитать_куки_csrf(page) -> dict | None:
    """Живые cookie контекста для текущего URL → csrf_куки_инфо. None, если
    прочитать не удалось (тогда вердикт не опирается на SameSite)."""
    try:
        cookies = page.context.cookies(page.url)
    except Exception:  # noqa: BLE001
        try:
            cookies = page.context.cookies()
        except Exception:  # noqa: BLE001
            return None
    return csrf_куки_инфо(cookies)


def csrf_вердикт(найдено: bool, заполнено: bool, ошибка: bool = False,
                 куки: dict | None = None) -> tuple:
    """(статус, деталь) для колонки «CSRF-защита». ЧИСТАЯ функция (юнит-тест
    без браузера). Статус: «Есть» / «Нет» / «Проверить».

    CSRF защищают ДВА независимых механизма, и достаточно любого:
      1) CSRF-токен - скрытое поле, которое сервер сверяет (найдено/заполнено);
      2) SameSite-cookie - браузер не пошлёт session-cookie при межсайтовом
         POST (куки: см. csrf_куки_инфо). Раньше проверяли только (1), поэтому
         на сайтах с SameSite-защитой была ложная «Нет»."""
    if ошибка:
        return "Проверить", "не удалось прочитать форму - CSRF не проверен"
    if найдено and заполнено:
        return "Есть", "скрытое поле токена сессии найдено и заполнено"
    # Токена в форме нет (или он пустой) - смотрим SameSite-защиту cookie.
    if куки is not None:
        if куки.get("сессионные_есть"):
            if куки.get("все_защищены_samesite"):
                return ("Есть",
                        "CSRF-токена в форме нет, но session-cookie помечены "
                        f"SameSite=Lax/Strict ({куки.get('детали','')}) - "
                        "межсайтовый POST не пошлёт их, подделка не пройдёт")
            return ("Нет",
                    "session-cookie с SameSite=None и без CSRF-токена "
                    f"({', '.join(куки.get('незащищённые') or [])}) - "
                    "браузер отправит их при межсайтовом POST, форма уязвима к CSRF")
        # Сессионных cookie нет вовсе: подделывать нечего (публичная форма без
        # авторизации) - CSRF к ней неприменим.
        if найдено:
            return ("Проверить", "поле токена есть, но пустое - "
                    "возможно, заполняется иначе или не успело проставиться JS")
        return ("Есть",
                "форма не использует сессионные cookie - подделать "
                "авторизованный запрос нельзя, CSRF неприменим")
    # Куки не переданы (старое поведение / чтение не удалось).
    if найдено:
        return ("Проверить",
                "поле токена есть, но пустое - возможно, заполняется иначе "
                "(не hidden-полем) или не успело проставиться JS")
    return ("Нет",
            "скрытое поле CSRF/sessid не найдено - если на сайте нет другой "
            "защиты (например SameSite-cookie), форма может быть уязвима к CSRF")


def _куки_из_ответа_requests(response) -> list:
    """Список {name, sameSite} из заголовков Set-Cookie ответа requests - тот же
    вход, что page.context.cookies() в браузере, чтобы оценить CSRF по SameSite и
    для code-пути (requests). SameSite в cookiejar не хранится, поэтому читаем
    СЫРЫЕ заголовки Set-Cookie. Любая ошибка → пустой список (тогда вердикт CSRF
    опирается только на наличие токена)."""
    setc = []
    try:
        raw = response.raw.headers
        if hasattr(raw, "getlist"):
            setc = raw.getlist("Set-Cookie")
    except Exception:  # noqa: BLE001
        setc = []
    if not setc:
        try:
            one = response.headers.get("Set-Cookie") or ""
        except Exception:  # noqa: BLE001
            one = ""
        setc = [one] if one else []
    out = []
    for line in setc:
        try:
            name = str(line).split("=", 1)[0].strip()
            if not name:
                continue
            m = re.search(r"samesite\s*=\s*(\w+)", str(line), re.I)
            out.append({"name": name, "sameSite": (m.group(1) if m else "")})
        except Exception:  # noqa: BLE001
            continue
    return out


# Скрытое поле токена/сессии - тот же вокабуляр, что и _CSRF_ПОЛЕ_JS (браузер).
_CSRF_ИМЯ_RE = re.compile(
    r"csrf|sessid|_token|authenticity_token|requestverificationtoken|xsrf|nonce",
    re.I)


def _html_структурные_проверки(form, html: str = "", куки: dict | None = None) -> dict:
    """Структурные проверки формы ПО СТАТИЧЕСКОМУ HTML (без браузера) - те же
    колонки отчёта, что заполняет браузерный путь, но на разобранной
    BeautifulSoup-форме (code-путь «по коду», requests). Раньше у форм,
    проверенных по коду, эти колонки были пустыми и в матрице показывались
    прочерками; теперь заполняются реальными вердиктами там, где ответ виден
    прямо в разметке. Чистая функция (юнит-тест без сети).

    Возвращает словарь log-ключ → значение В ТОМ ЖЕ ВОКАБУЛЯРЕ, что и браузерный
    путь (чтобы правила матрицы и консолидация работали без изменений)."""
    out: dict = {}
    if form is None:
        return out

    def _find_all(tag, **attrs):
        try:
            return form.find_all(tag, attrs=attrs) if attrs else form.find_all(tag)
        except Exception:  # noqa: BLE001
            return []

    def _attr(el, name):
        try:
            return el.get(name)
        except Exception:  # noqa: BLE001
            return None

    def _has(el, name):
        try:
            return el.has_attr(name)
        except Exception:  # noqa: BLE001
            return False

    # ── CSRF: скрытое поле токена/сессии + SameSite-cookie (как csrf_вердикт). ──
    найдено = заполнено = False
    for h in _find_all("input", type="hidden"):
        nm = str(_attr(h, "name") or _attr(h, "id") or "")
        if _CSRF_ИМЯ_RE.search(nm):
            найдено = True
            if str(_attr(h, "value") or "").strip():
                заполнено = True
            break
    # Значение hash/sessid на Bitrix часто проставляет JS - подхватим из скриптов
    # (как extract_form_security_from_html), тогда «Есть», а не «Проверить».
    if найдено and not заполнено and html:
        try:
            sec = extract_form_security_from_html(html)
            if (sec.get("sessid") or "").strip() or (sec.get("hash") or "").strip():
                заполнено = True
        except Exception:  # noqa: BLE001
            pass
    _cs, _ = csrf_вердикт(найдено, заполнено, ошибка=False, куки=куки)
    out["csrf_защита"] = _cs

    # ── Согласие 2.13: чекбоксы, предустановка, ссылка на политику, обязательность.
    видимые_cb = []
    for cb in _find_all("input", type="checkbox"):
        st = str(_attr(cb, "style") or "").replace(" ", "").lower()
        if "display:none" in st or "visibility:hidden" in st:
            continue
        видимые_cb.append(cb)
    предустановлены = any(_has(cb, "checked") for cb in видимые_cb)
    обязательно = any(
        _has(cb, "required")
        or str(_attr(cb, "aria-required") or "").lower() == "true"
        for cb in видимые_cb)
    ссылка = False
    for a in _find_all("a"):
        txt = ""
        try:
            txt = a.get_text() or ""
        except Exception:  # noqa: BLE001
            txt = ""
        if ссылка_ведёт_на_политику(str(_attr(a, "href") or ""), txt):
            ссылка = True
            break
    out["согласие_чекбоксы"] = f"{len(видимые_cb)} (нужно ≥2)"
    out["согласие_предустановка"] = ('да' if not предустановлены
                                     else 'НЕТ - стоят по умолчанию')
    out["согласие_ссылка"] = 'да' if ссылка else 'нет'
    out["согласие_обязательно"] = 'да' if обязательно else 'нет'

    # ── Выпадающие списки: пустой <select> - битый; со списком - корректно. ──
    selects = _find_all("select")
    if not selects:
        out["выпадающие_списки"] = "не найдено"
    elif any(len(_find_all_options(s)) == 0 for s in selects):
        out["выпадающие_списки"] = "ошибка"
    else:
        out["выпадающие_списки"] = "корректно"

    # ── Типы файлов: загрузчик без accept принимает ЛЮБЫЕ файлы. ──
    files = _find_all("input", type="file")
    if not files:
        out["типы_файлов"] = ""
    elif any(not str(_attr(f, "accept") or "").strip() for f in files):
        out["типы_файлов"] = "⚠ ЛЮБЫЕ типы (accept не задан)"
    else:
        типы = []
        for f in files:
            типы += [t.strip() for t in str(_attr(f, "accept") or "").split(",")
                     if t.strip()]
        out["типы_файлов"] = ", ".join(dict.fromkeys(типы)) or "заданы"

    # ── Подсказки (placeholder) у текстовых полей: есть - корректно; совсем нет -
    # мягкая заметка (⚠). Флажим только ПОЛНОЕ отсутствие, как браузерный путь. ──
    текстовые = []
    for el in _find_all("input") + _find_all("textarea"):
        try:
            nm = el.name
        except Exception:  # noqa: BLE001
            nm = ""
        if nm == "textarea":
            текстовые.append(el)
            continue
        t = str(_attr(el, "type") or "text").lower()
        if t in ("text", "email", "tel", "search", "url", "number", "password"):
            текстовые.append(el)
    if not текстовые:
        out["подсказки"] = ""
    elif any(str(_attr(el, "placeholder") or "").strip() for el in текстовые):
        out["подсказки"] = "корректно"
    else:
        out["подсказки"] = "нет подсказок"

    return out


def _find_all_options(select_el) -> list:
    try:
        return select_el.find_all("option")
    except Exception:  # noqa: BLE001
        return []


# ── Данные формы дошли до сервера (видно в DevTools → Network) ────────────
# Проверяется ВСЕГДА, без единого лишнего запроса - вешается на ту же самую
# единственную легитимную отправку формы, что и так уже происходит (для ЛЮБОЙ
# формы, включая заказы - тут нет второго клика, только пассивное наблюдение
# за первым и единственным). Ловит баг «сайт нарисовал Спасибо, но реально
# отправил не то (или не всё), что было в полях» - направление, ОБРАТНОЕ
# «Серверной валидации» (та проверяет, что БИТЫЕ данные сервер отклонит; этот
# пункт - что ХОРОШИЕ данные не потеряются по дороге).
_JS_СНИМОК_ПОЛЕЙ = r"""
f => {
  const skip = ['hidden','submit','button','reset','image','file','checkbox','radio'];
  const vis = e => !!(e.offsetWidth || e.offsetHeight || e.getClientRects().length);
  const out = {};
  for (const e of f.querySelectorAll('input,textarea,select')) {
    const t = (e.type || '').toLowerCase();
    if (skip.includes(t) || !vis(e)) continue;
    const v = (e.value || '').trim();
    if (!v) continue;
    out[e.name || e.id || ('поле_' + Object.keys(out).length)] = v;
  }
  return out;
}
"""


def _снять_поля_формы(form) -> dict:
    """{поля: {имя: значение}, ошибка: bool} - снимок ВИДИМЫХ непустых полей
    формы. Снимать нужно максимально поздно (прямо перед кликом submit) - к
    этому моменту чужие пробы (проверка_кнопки_обязательные,
    проверка_отображения_ошибок) уже закончили свои временные очистки/
    восстановления полей выше по потоку."""
    try:
        r = form.evaluate(_JS_СНИМОК_ПОЛЕЙ)
        if isinstance(r, dict):
            return {"поля": r, "ошибка": False}
    except Exception:  # noqa: BLE001
        pass
    return {"поля": {}, "ошибка": True}


def _плоские_значения(x) -> list:
    """Разворачивает произвольно вложенный dict/list (тело JSON-запроса) в
    плоский список строк-значений - структуру, в которую сайт завернул поля
    формы перед fetch(), заранее не знаем."""
    if isinstance(x, dict):
        out = []
        for v in x.values():
            out.extend(_плоские_значения(v))
        return out
    if isinstance(x, list):
        out = []
        for v in x:
            out.extend(_плоские_значения(v))
        return out
    if x is None:
        return []
    return [str(x)]


def _тело_запроса_для_поиска(content_type: str, raw: str) -> str:
    """Текст для поиска значений полей: СЫРОЕ тело ПЛЮС (если удалось
    разобрать) те же значения через штатный парсер формата - раздельное
    percent/plus-кодирование (urlencoded) и JSON-escaping (\\n, \\") делают
    сырой текст плохим для точного поиска многострочных/спецсимвольных
    значений, поэтому разбор ДОБАВЛЯЕТСЯ к сырому тексту, а не заменяет его
    (multipart не разбираем вовсе - там значения в частях буквальные, сырого
    текста достаточно). Ошибка разбора не фатальна - тихий откат на сырой текст."""
    raw = raw or ""
    ct = (content_type or "").lower()
    try:
        if "urlencoded" in ct:
            from urllib.parse import parse_qsl
            return raw + " " + " ".join(в for _, в in parse_qsl(raw))
        if "json" in ct:
            return raw + " " + " ".join(_плоские_значения(json.loads(raw)))
    except Exception:  # noqa: BLE001
        pass
    return raw


def _значение_долетело(значение: str, тело: str, тело_цифры: str) -> bool:
    """Найдено ли конкретное заполненное значение в теле запроса. Порядок
    попыток - от строгой к снисходительной: точная подстрока → без учёта
    регистра → похоже на телефон (10+ цифр) - по последним 10 цифрам (сайт
    мог переформатировать маску +7/8/пробелы/скобки - это не потеря данных)
    → длинное значение (20+ символов) - по первым ~20 символам (сайт мог
    обрезать по своему maxlength - тоже не потеря данных). Совсем короткие
    значения (короче 2 символов) шума ради не сверяем - считаются найденными."""
    v = (значение or "").strip()
    if len(v) < 2:
        return True
    if v in тело:
        return True
    v_low, тело_low = v.lower(), тело.lower()
    if v_low in тело_low:
        return True
    digits = "".join(c for c in v if c.isdigit())
    if len(digits) >= 10 and digits[-10:] in тело_цифры:
        return True
    if len(v) >= 20 and v_low[:20] in тело_low:
        return True
    return False


def данные_формы_вердикт(пойман_хоть_один_post: bool, тело: str, ожидаемые: dict) -> tuple:
    """(статус, деталь) для «Данные дошли до сервера». ЧИСТАЯ функция
    (юнит-тест без браузера). ожидаемые - {поле: значение}: снимок видимых
    непустых полей формы прямо перед кликом отправки. Статус: «Да» /
    «Не долетело» / «Проверить» (нет заполненных полей ИЛИ не поймали ни
    одного POST - тот же класс неопределённости, что и «под вопросом» у
    двойной отправки при несчитанных отправках)."""
    ожидаемые = {п: (з or "").strip() for п, з in (ожидаемые or {}).items() if (з or "").strip()}
    if not ожидаемые:
        return "Проверить", "нет заполненных полей для сверки"
    if not пойман_хоть_один_post:
        return "Проверить", "не удалось поймать сетевой POST-запрос отправки формы"
    тело = тело or ""
    тело_цифры = "".join(c for c in тело if c.isdigit())
    пропало = [п for п, з in ожидаемые.items()
               if not _значение_долетело(з, тело, тело_цифры)]
    if пропало:
        return ("Не долетело",
                "в теле запроса на сервер не нашли значение полей: " + ", ".join(пропало))
    return "Да", f"все заполненные поля ({len(ожидаемые)}) нашли в теле запроса"


def format_form_selector_type(форма_config: dict | None) -> str:
    """
    Тип селектора формы для лога Excel: id, class, data-source, css, name, text.
    """
    if not форма_config:
        return ""
    for key in ("id", "class", "data-source", "css", "name", "text"):
        if key in форма_config:
            return key
    return ""


def format_form_config_for_log(форма_config: dict) -> str:
    """
    Колонка Excel «Значение типа»: только значение (id, class, data-source, css, name или text),
    без префиксов вида «class=» и без названия теста.
    Если задан нестандартный «индекс» (не 0), добавляется суффикс [n] для различия форм.
    """
    if not форма_config:
        return ""
    try:
        idx = int(форма_config.get("индекс", 0))
    except (TypeError, ValueError):
        idx = 0
    for key in ("id", "class", "data-source", "css", "name", "text"):
        if key in форма_config:
            val = str(форма_config[key]).strip()
            if not val:
                continue
            pref = f"{key}:"
            if val.lower().startswith(pref):
                val = val[len(pref) :].strip()
            if idx:
                return f"{val} [{idx}]"
            return val
    if "индекс" in форма_config:
        return str(форма_config["индекс"])
    return ""


_HTML_TAG_NAMES = frozenset(
    {
        "a",
        "abbr",
        "article",
        "aside",
        "b",
        "body",
        "br",
        "button",
        "canvas",
        "dd",
        "div",
        "dl",
        "dt",
        "em",
        "footer",
        "form",
        "h1",
        "h2",
        "h3",
        "h4",
        "h5",
        "h6",
        "header",
        "hr",
        "html",
        "i",
        "img",
        "input",
        "label",
        "li",
        "main",
        "nav",
        "ol",
        "option",
        "p",
        "pre",
        "section",
        "select",
        "small",
        "span",
        "strong",
        "sub",
        "sup",
        "svg",
        "table",
        "tbody",
        "td",
        "textarea",
        "tfoot",
        "th",
        "thead",
        "title",
        "tr",
        "ul",
        "video",
        "path",
        "g",
        "defs",
    }
)


def _normalize_scenario_click_css_selector(s: str) -> str:
    """
    Типичная ошибка в шаге «клик»: несколько классов через пробел без точек.
    В CSS это классы на одном элементе: «btn btn-transparent-blue» → «.btn.btn-transparent-blue».
    Не трогаем уже корректные селекторы (#id, [attr], >>, text=, теги button/div/…).
    """
    s = (s or "").strip()
    if not s:
        return s
    low = s.lower()
    if low.startswith((".", "#", "[", "/", "*", "(", "text=", ">>", "role=", "xpath=")):
        return s
    if ">>" in s:
        return s
    parts = s.split()

    def _simple_token(t: str) -> bool:
        return bool(t) and all(c.isalnum() or c in "_-" for c in t)

    if len(parts) >= 2:
        if all(_simple_token(p) for p in parts) and not any(
            "." in p or "#" in p for p in parts
        ):
            return "." + ".".join(parts)
    if len(parts) == 1:
        p = parts[0]
        if "." in p or "#" in p or "[" in p:
            return s
        if p.lower() in _HTML_TAG_NAMES:
            return s
        if _simple_token(p):
            return "." + p
    return s


def _playwright_form_css_selector(форма_config: dict) -> str:
    """CSS-селектор формы для page.locator (или пусто, если тип «text»).
    Один токен class: form.X матчит любой токен в class (в т.ч. an-row calculation-order).
    Для «короткой» формы в конфиге пишут class: calculation-order - тогда :not(.an-row),
    чтобы не сливалась со второй формой an-row calculation-order.
    Один токен (кроме calculation-order): form.token - как CSS-класс (подходит и при class="row col-md-12").
    Несколько слов: form.a.b.c как раньше.
    Ключ «css» - произвольный CSS-селектор (должен находить форму или её корень).
    Ключ «name» - атрибут name (CSS [name="..."]), часто поле Bitrix; для контейнера заказа
    лучше «css» (например .bx-soa-customer), если по name попадает один input.
    Ключ «text» - не CSS; возвращает пусто (см. отправить_форму_через_playwright).
    """
    if "text" in форма_config and str(форма_config.get("text", "")).strip():
        return ""
    if "css" in форма_config:
        raw = str(форма_config["css"]).strip()
        # Как у шагов «клик»: «row» → «.row», иначе locator ищет несуществующий тег <row>.
        return _normalize_scenario_click_css_selector(raw)
    if "id" in форма_config:
        fid = форма_config["id"].replace('"', '\\"')
        return f'form#{fid}'
    if "class" in форма_config:
        cls = str(форма_config["class"]).strip()
        parts = cls.split()
        if len(parts) == 1:
            token = parts[0].replace("\\", "\\\\").replace('"', '\\"')
            # В конфиге одна строка «calculation-order», вторая «an-row calculation-order»
            if token == "calculation-order":
                return "form.calculation-order:not(.an-row)"
            # form.row, а не form[class="row"]: иначе не матчится при class="row другие-классы"
            return f"form.{token}"
        return "form." + ".".join(parts)
    if "data-source" in форма_config:
        ds = форма_config["data-source"].replace("'", "\\'")
        return f"form[data-source='{ds}']"
    if "name" in форма_config:
        nm = str(форма_config["name"]).strip().replace("\\", "\\\\").replace('"', '\\"')
        return f'[name="{nm}"]'
    return ""


def _alternate_form_root_selector(sel: str) -> str | None:
    """
    Многие сайты без <form>: блоки в div / section. Тогда form.* не матчится - пробуем div.*.
    form[attr] (например data-source) - то же: div[attr].
    """
    if sel.startswith("form.") or sel.startswith("form#"):
        return "div" + sel[4:]
    if sel.startswith("form["):
        return "div" + sel[4:]
    return None


def _expand_form_selector_fallbacks(sel: str) -> list[str]:
    """
    Цепочка CSS для page.locator / soup.select: основной селектор и более общие варианты.
    """
    s = (sel or "").strip()
    if not s:
        return []
    out: list[str] = []
    seen: set[str] = set()

    def add(x: str) -> None:
        x = (x or "").strip()
        if x and x not in seen:
            seen.add(x)
            out.append(x)

    add(s)
    alt = _alternate_form_root_selector(s)
    if alt:
        add(alt)
    # form.foo.bar → .foo.bar (часто нет тега <form>, нужен просто класс)
    if s.startswith("form.") and not s.startswith("form["):
        tail = s[4:]
        if tail.startswith("."):
            add(tail)
    # div#id → #id (часто id на обёртке не div)
    for v in list(out):
        if v.startswith("div#") and "[" not in v:
            add("#" + v[4:])
    # div.a.b.c → .a.b.c (класс на любом теге)
    for v in list(out):
        if v.startswith("div.") and "[" not in v[4:]:
            add("." + v[4:])
    # data-source: без тега (ищем любой элемент с атрибутом)
    m = re.search(r"data-source\s*=\s*(['\"])([^'\"]+)\1", s)
    if m:
        ds_esc = m.group(2).replace("'", "\\'")
        add(f"[data-source='{ds_esc}']")
    return out


_MODAL_TRIGGER_PREFIXES = frozenset(
    {"text", "css", "id", "class", "data-source", "ds", "name"}
)


def _parse_modal_trigger(raw: str) -> tuple[str, str]:
    """
    Триггер модалки в конфиге:
    - без префикса или data-source: - значение в атрибуте data-source (как раньше);
    - text: - подпись кнопки/ссылки (роль button/link);
    - css: - произвольный CSS-селектор Playwright;
    - id: - id элемента;
    - class: - класс(ы) на button/a;
    - name: - атрибут name (например name:ORDER_PROP_3).
    Префикс срабатывает только если он из известного списка (строка «foo:bar» без префикса остаётся целиком для data-source).
    """
    s = (raw or "").strip()
    if not s:
        return ("data-source", "")
    if ":" in s:
        prefix, rest = s.split(":", 1)
        p = prefix.strip().lower()
        if p in _MODAL_TRIGGER_PREFIXES:
            if p == "ds":
                p = "data-source"
            return (p, rest.strip())
    return ("data-source", s)


def format_modal_selector_type(raw: str | None) -> str:
    """Тип триггера модалки: id, class, css, text, data-source, name и т.д."""
    return _parse_modal_trigger(raw or "")[0]


def format_modal_value_for_log(raw: str | None) -> str:
    """Колонка «Значение типа»: только значение без префикса «class:» и т.п. (тип - в колонке «Тип»)."""
    return _parse_modal_trigger(raw or "")[1]


def encode_modal_trigger(kind: str, value: str) -> str:
    """Собирает строку триггера для config (согласовано с _parse_modal_trigger)."""
    v = (value or "").strip()
    k = (kind or "data-source").strip().lower()
    if k == "ds":
        k = "data-source"
    if k == "data-source":
        return v
    if not v:
        return ""
    return f"{k}:{v}"


def _find_modal_opener(page, kind: str, val: str):
    """Элемент, по которому кликают, чтобы открыть модалку."""
    val = (val or "").strip()
    if not val:
        return None

    def _fallback_by_id_or_hash():
        """Если data-source не сработал - часто в конфиге указан id кнопки/ссылки (без префикса id:)."""
        id_esc = val.replace('"', '\\"')
        for sel in (f'[id="{id_esc}"]', f"#{val}"):
            try:
                loc = page.locator(sel).first
                if loc.count():
                    return loc
            except Exception:
                continue
        return None
    if kind == "css":
        loc = page.locator(val).first
        return loc if loc.count() else None
    if kind == "text":
        for role in ("button", "link"):
            loc = page.get_by_role(role, name=val, exact=True)
            if loc.count():
                return loc.first
        loc = page.get_by_role("button", name=val)
        if loc.count():
            return loc.first
        combo = page.locator("button, a, [role='button']").filter(has_text=val)
        if combo.count():
            return combo.first
        return None
    if kind == "id":
        id_esc = val.replace('"', '\\"')
        for sel in (f'[id="{id_esc}"]', f"#{val}"):
            try:
                loc = page.locator(sel).first
                if loc.count():
                    return loc
            except Exception:
                continue
        return None
    if kind == "class":
        parts = val.split()
        if not parts:
            return None
        chain = ".".join(parts)
        loc = page.locator(
            f"button.{chain}, a.{chain}, [role='button'].{chain}"
        ).first
        if loc.count():
            return loc
        return None
    if kind == "name":
        nm_esc = val.replace("\\", "\\\\").replace('"', '\\"')
        loc = page.locator(f'[name="{nm_esc}"]').first
        return loc if loc.count() else None
    # data-source (по умолчанию)
    esc = val.replace("'", "\\'")
    for sel in (
        f"button[data-source='{esc}']",
        f"a[data-source='{esc}']",
        f"[data-source='{esc}']",
    ):
        loc = page.locator(sel).first
        if loc.count():
            return loc
    return _fallback_by_id_or_hash()


def _find_modal_root(page):
    """Видимый контейнер модалки или формы после клика по триггеру."""
    try:
        page.wait_for_selector(
            '[role="dialog"] input, [role="dialog"] textarea, '
            ".modal input, .modal textarea, .popup input, "
            '[class*="modal"] input',
            state="visible",
            timeout=9000,
        )
    except Exception:
        pass
    page.wait_for_timeout(350)

    try:
        dlg = page.locator('[role="dialog"], [role="alertdialog"]').filter(
            has=page.locator(
                'input:not([type="hidden"]), textarea, '
                'input[name="name"], input[type="tel"], input[type="email"]'
            )
        ).last
        dlg.wait_for(state="visible", timeout=4000)
        if dlg.count() and dlg.locator("input, textarea").count() > 0:
            return dlg
    except Exception:
        pass

    for sel in (
        ".modal:visible",
        '[class*="popup"]:visible',
        '[class*="modal"]:visible',
        ".popup:visible",
    ):
        loc = page.locator(sel).last
        try:
            if loc.count() and loc.locator("input:not([type='hidden'])").count() > 0:
                loc.wait_for(state="visible", timeout=2000)
                return loc
        except Exception:
            continue

    try:
        form = page.locator("form:has(input)").last
        if form.count():
            form.wait_for(state="visible", timeout=4000)
            return form
    except Exception:
        pass

    for sel in ("#form-callback", "#callback-form", ".modal"):
        loc = page.locator(sel).first
        if loc.count():
            return loc
    return page.locator('[role="dialog"]').last


def _modal_scope_for_fill(page, modal):
    """Область, в которой реально есть поля (часто modal - пустой #form-callback)."""
    try:
        if modal.count():
            n = modal.locator(
                "input:not([type='hidden']):not([type='checkbox']):not([type='radio']), textarea"
            ).count()
            if n > 0:
                return modal
    except Exception:
        pass
    for sel in ('[role="dialog"]', '[role="alertdialog"]'):
        loc = page.locator(sel).last
        if loc.count() and loc.locator("input, textarea").count() > 0:
            return loc
    for sel in (".modal:visible", ".popup:visible", '[class*="modal"]:visible'):
        loc = page.locator(sel).last
        if loc.count() and loc.locator("input").count() > 0:
            return loc
    return modal


def _fill_modal_fields(
    modal,
    имя_теста: str,
    телефон: str,
    почта: str,
    комментарий: str,
    page=None,
):
    """Заполняет типовые поля в модалке (имя, телефон, почта, комментарий)."""
    scope = _modal_scope_for_fill(page, modal) if page is not None else modal

    def _fill_first(locator_str: str, value: str) -> bool:
        if value is None or str(value).strip() == "":
            return False
        try:
            el = scope.locator(locator_str).first
            if el.count():
                el.scroll_into_view_if_needed()
                el.fill(value, timeout=8000, force=True)
                return True
        except Exception:
            pass
        return False

    # Имя - типовые name + Bitrix / произвольные
    name_sels = tuple(
        f'input[name="{nm}"]'
        for nm in (
            "name",
            "NAME",
            "fio",
            "FIO",
            "client_name",
            "user_name",
            "contact_name",
            "form_text_1",
            "form_text_2",
            "PROPERTY_NAME",
        )
    )
    filled_name = False
    for sel in name_sels:
        if _fill_first(sel, имя_теста):
            filled_name = True
            break

    if not filled_name and имя_теста:
        for sub in ("Имя", "ФИО", "имя", "Ваше имя", "Name", "Фио"):
            try:
                el = scope.locator(f'input[placeholder*="{sub}"]')
                if el.count():
                    el.first.scroll_into_view_if_needed()
                    el.first.fill(имя_теста, timeout=5000, force=True)
                    filled_name = True
                    break
            except Exception:
                pass

    phone_sels = (
        "input[name='telephone']",
        "input[name='phone']",
        "input[name='PHONE']",
        "input[name='PHONE_MOBILE']",
        "input[name='form_phone']",
        "input[name='form_text_3']",
        "input[name='form_text_2']",
        "input[name='tel']",
        "input[name='TEL']",
        "input[name='PROPERTY_PHONE']",
        "input[type='tel']",
        "input[name*='phone']",
        "input[name*='PHONE']",
        "input[name*='tel']",
    )
    filled_phone = False
    for sel in phone_sels:
        if _fill_first(sel, телефон):
            filled_phone = True
            break

    if not filled_phone and телефон:
        for sub in ("Телефон", "телефон", "Phone", "Мобильный", "Tel"):
            try:
                el = scope.locator(f'input[placeholder*="{sub}"]')
                if el.count():
                    el.first.scroll_into_view_if_needed()
                    el.first.fill(телефон, timeout=5000, force=True)
                    filled_phone = True
                    break
            except Exception:
                pass

    # Bitrix: подряд form_text_N
    if (not filled_name or not filled_phone) and (имя_теста or телефон):
        try:
            bits = scope.locator('input[name^="form_text_"]')
            n = bits.count()
            if n >= 1 and not filled_name and имя_теста:
                bits.first.scroll_into_view_if_needed()
                bits.first.fill(имя_теста, timeout=5000, force=True)
                filled_name = True
            if n >= 2 and not filled_phone and телефон:
                bits.nth(1).scroll_into_view_if_needed()
                bits.nth(1).fill(телефон, timeout=5000, force=True)
                filled_phone = True
        except Exception:
            pass

    # Последний шанс: первые два видимых текстовых поля в форме модалки
    if (not filled_name or not filled_phone) and (имя_теста or телефон):
        try:
            form = scope.locator("form").last
            if form.count():
                inputs = form.locator(
                    "input[type='text'], input[type='tel'], input:not([type])"
                )
                cnt = inputs.count()
                for i in range(min(cnt, 6)):
                    inp = inputs.nth(i)
                    try:
                        if not inp.is_visible():
                            continue
                        tp = (inp.get_attribute("type") or "text").lower()
                        if tp in ("hidden", "checkbox", "radio", "submit", "button"):
                            continue
                        if not filled_name and имя_теста:
                            inp.scroll_into_view_if_needed()
                            inp.fill(имя_теста, force=True, timeout=4000)
                            filled_name = True
                            continue
                        if filled_name and not filled_phone and телефон:
                            inp.scroll_into_view_if_needed()
                            inp.fill(телефон, force=True, timeout=4000)
                            filled_phone = True
                            break
                    except Exception:
                        continue
        except Exception:
            pass

    if почта:
        if not _fill_first('input[type="email"]', почта):
            _fill_first('input[name="email"]', почта)

    if комментарий:
        try:
            scope.locator(
                "textarea, input[name='comment'], input[name='message'], input[name='COMMENT']"
            ).first.fill(комментарий, timeout=5000, force=True)
        except Exception:
            pass

    if имя_теста and not filled_name:
        print(
            "      ⚠️ Модалка: не удалось заполнить имя - откройте DevTools и проверьте name/placeholder полей."
        )
    if телефон and not filled_phone:
        print(
            "      ⚠️ Модалка: не удалось заполнить телефон - проверьте name/placeholder или укажите ТЕЛЕФОН в конфиге."
        )


def _модалка_видна(modal) -> bool:
    """Виден ли контейнер модалки СЕЙЧАС - для проверки «закрывается». Не
    требует полей внутри (после отправки форма могла смениться на «Спасибо»
    с одной кнопкой «ОК», без input/textarea)."""
    try:
        return bool(modal.count()) and modal.is_visible()
    except Exception:  # noqa: BLE001
        return False


def _модалка_открылась(modal) -> bool:
    """Открылась ли модалка ПО-НАСТОЯЩЕМУ (для проверки «открывается»): видна
    И внутри есть хотя бы одно поле - иначе _find_modal_root() мог тихо
    съехать на пустой fallback вроде [role="dialog"] с нулевым count()."""
    if not _модалка_видна(modal):
        return False
    try:
        return modal.locator("input, textarea, select").count() > 0
    except Exception:  # noqa: BLE001
        return False


# Селекторы типовой кнопки закрытия модалки - крестик/«Закрыть» в разных
# вариантах разметки (data-атрибуты Bootstrap/micromodal, классы, aria-label).
_MODAL_CLOSE_SELECTORS = (
    "[data-dismiss='modal']", "[data-micromodal-close]", "[data-close]",
    "[aria-label*='закрыть' i]", "[aria-label*='close' i]",
    ".modal-close", ".popup-close", ".close-btn", ".btn-close",
    "[class*='close' i]",
)


def _закрыть_модалку_способ(page, modal, способ: str) -> bool:
    """Пробует ОДИН способ закрыть модалку. Возвращает True, только если
    модалка реально пропала (или потеряла видимость) ПОСЛЕ попытки."""
    try:
        if способ == "крестик":
            найдена = False
            for sel in _MODAL_CLOSE_SELECTORS:
                try:
                    btn = modal.locator(sel).first
                    if btn.count() and btn.is_visible():
                        btn.click(timeout=3000, force=True)
                        найдена = True
                        break
                except Exception:  # noqa: BLE001
                    continue
            if not найдена:
                return False
        elif способ == "esc":
            page.keyboard.press("Escape")
        elif способ == "клик_вне":
            try:
                box = modal.bounding_box()
                vp = page.viewport_size
            except Exception:  # noqa: BLE001
                box, vp = None, None
            if not box or not vp:
                return False
            # Точка заведомо ВНЕ прямоугольника модалки, но внутри вьюпорта -
            # угол экрана обычно свободен от интерактивных элементов сайта.
            кандидаты = [(5, 5), (vp["width"] - 5, 5), (5, vp["height"] - 5)]
            pt = next((p for p in кандидаты
                       if not (box["x"] <= p[0] <= box["x"] + box["width"]
                               and box["y"] <= p[1] <= box["y"] + box["height"])), None)
            if not pt:
                return False
            page.mouse.click(*pt)
        else:
            return False
        page.wait_for_timeout(500)
        return not _модалка_видна(modal)
    except Exception:  # noqa: BLE001
        return False


def _проба_закрытия_модалки(page, modal) -> tuple:
    """Пробует закрыть модалку тремя способами по очереди (крестик/кнопка →
    Esc → клик вне модалки), пока один не сработает. Жёсткий вердикт «Да»/
    «Нет» - без «проверить вручную» (по просьбе: тул должен сам дойти до
    ответа, а не перекладывать на человека)."""
    if not _модалка_видна(modal):
        return "Нет", "модалка уже не видна к моменту проверки"
    for способ, имя in (("крестик", "крестик/кнопка закрытия"),
                        ("esc", "клавиша Esc"),
                        ("клик_вне", "клик вне модалки")):
        if _закрыть_модалку_способ(page, modal, способ):
            return "Да", имя
    return "Нет", "не закрылась ни крестиком, ни Esc, ни кликом вне модалки"


# Для форм, открытых через сценарий «клик → форма» (не через выделенный блок
# «модалки») - определяем, лежит ли найденная форма внутри модалки/попапа.
# XPath ancestor:: от САМОЙ формы (не .filter(has=form)!) - .filter(has=…)
# переоценивает селектор `form`-локатора ЗАНОВО внутри каждого кандидата, и
# если у формы селектор сам заякорен на id (напр. "#modal1 form" - обычное
# дело, когда форма ищется по id/css из конфига), повторный скоуп внутри
# кандидата #modal1 ищет #modal1 ВНУТРИ #modal1 - ничего не находит. XPath
# ancestor:: работает от уже разрешённого узла формы - этой проблемы нет.
# ancestor-OR-SELF: локатор формы иногда указывает на САМ popup-контейнер
# (напр. Bitrix «Оставить заявку»: селектор формы разрешается в div#txt-back с
# class="popup", а не во внутренний <form>). Тогда «модалка вокруг формы» = сам
# элемент, и ancestor:: её не находил → модалка шла прочерком. self учитывает и
# этот случай, и обычный (форма ВНУТРИ .popup).
_MODAL_ANCESTOR_XPATH = (
    "xpath=(ancestor-or-self::*[contains(@class,'modal') or contains(@class,'popup') "
    "or contains(@class,'fancybox') or contains(@class,'mfp') "
    "or contains(@class,'lightbox') "
    "or @role='dialog' or @role='alertdialog' "
    "or @aria-modal='true'])[last()]"
)


def _найти_модалку_вокруг(form):
    """Локатор модалки/попапа вокруг уже найденной формы, или None, если форма
    открыта не в модалке (обычная форма на странице) - тогда пробу закрытия
    для неё не запускаем, это не относится к пункту."""
    try:
        anc = form.locator(_MODAL_ANCESTOR_XPATH)
        if anc.count():
            return anc.first
    except Exception:  # noqa: BLE001
        pass
    return None


def _click_modal_submit(modal):
    """Отправка формы в модалке - несколько типовых вариантов кнопки."""
    for sel in (
        "button.form__submit",
        "button[type='submit']",
        "input[type='submit']",
        "button.btn[type='submit']",
        "input[name='web_form_submit']",
        "button.send",
        "button.js-form-submit",
    ):
        try:
            btn = modal.locator(sel).first
            if btn.count() and btn.is_visible():
                btn.click(timeout=8000, force=True)
                return
        except Exception:
            continue
    try:
        modal.get_by_role(
            "button", name=re.compile(r"отправ|заказ|оформ|отправить|send", re.I)
        ).first.click(timeout=8000, force=True)
        return
    except Exception:
        pass
    try:
        modal.locator("button[type='submit']").first.click(timeout=8000, force=True)
    except Exception:
        try:
            modal.locator("form").locator("button").last.click(timeout=5000, force=True)
        except Exception:
            pass


def _label_has_policy_link(locator) -> bool:
    """В label часто ссылка на политику - клик по label открывает страницу, а не чекбокс."""
    try:
        return locator.locator("a[href]").count() > 0
    except Exception:
        return False


def _click_checkbox_via_label_or_js(box, page):
    """Ставит галочку на input; не кликает по label, если внутри ссылка (иначе откроется политика)."""
    try:
        box.scroll_into_view_if_needed()
        box.check(force=True, timeout=8000)
        return
    except Exception:
        pass
    try:
        box.click(force=True, timeout=8000)
        return
    except Exception:
        pass
    try:
        box.evaluate(
            """el => {
            el.checked = true;
            el.dispatchEvent(new Event('input', {bubbles: true}));
            el.dispatchEvent(new Event('change', {bubbles: true}));
        }"""
        )
        return
    except Exception:
        pass
    try:
        bid = box.get_attribute("id")
        if bid:
            bid_esc = bid.replace("\\", "\\\\").replace('"', '\\"')
            lab = page.locator(f'label[for="{bid_esc}"]')
            if lab.count() > 0 and not _label_has_policy_link(lab.first):
                lab.first.scroll_into_view_if_needed()
                lab.first.click(timeout=8000)
                return
    except Exception:
        pass
    try:
        if box.evaluate(
            """el => {
            const lab = el.closest('label');
            if (!lab) return false;
            if (lab.querySelector('a[href]')) return false;
            lab.click();
            return true;
        }"""
        ):
            return
    except Exception:
        pass
    try:
        box.evaluate(
            """el => {
            el.checked = true;
            el.dispatchEvent(new Event('input', {bubbles: true}));
            el.dispatchEvent(new Event('change', {bubbles: true}));
        }"""
        )
    except Exception:
        pass


def _playwright_check_consent_checkboxes(form_locator, page):
    """
    Отмечает обязательные чекбоксы (152-ФЗ и т.д.).
    Bitrix/кастомные темы часто скрывают input - кликаем по label, иначе force/check + JS.
    """
    boxes = form_locator.locator("input[type='checkbox'][required]")
    n = boxes.count()
    for i in range(n):
        box = boxes.nth(i)
        _click_checkbox_via_label_or_js(box, page)


def _ensure_modal_consent(scope, page):
    """
    Согласие в модалке: required, типичные name/id (Bitrix), клик по подписи на русском,
    затем оставшиеся видимые чекбоксы в форме (кроме рассылки).
    """
    if not scope.count():
        return

    _playwright_check_consent_checkboxes(scope, page)

    name_fragments = (
        "agree",
        "privacy",
        "personal",
        "consent",
        "policy",
        "accept",
        "licence",
        "SOGLASIE",
        "pd",
        "cbpd",
        "fzs",
        "152",
        "PERSONAL",
        "form_checkbox",
        "licenses",
    )
    for frag in name_fragments:
        for attr in ("name", "id"):
            try:
                loc = scope.locator(
                    f'input[type="checkbox"][{attr}*="{frag}"]'
                )
                for j in range(loc.count()):
                    box = loc.nth(j)
                    try:
                        if box.is_checked():
                            continue
                    except Exception:
                        pass
                    _click_checkbox_via_label_or_js(box, page)
            except Exception:
                pass

    # Явные имена из типовых сайтов
    for nm in (
        "privacy_agree",
        "terms_agree",
        "licence_popup",
        "PERSONAL_DATA",
        "SOGLASIE",
        "agreement",
    ):
        try:
            loc = scope.locator(f'input[type="checkbox"][name="{nm}"]')
            if loc.count():
                box = loc.first
                if not box.is_checked():
                    _click_checkbox_via_label_or_js(box, page)
        except Exception:
            pass

    # Не кликаем по произвольным label с текстом «политик…» - там часто <a href>, открывается страница политики.
    # Только чекбоксы внутри label с подходящим текстом (без клика по ссылке).
    for pattern in (r"соглас", r"персональн", r"обработк"):
        try:
            labs = scope.locator("label:has(input[type='checkbox'])").filter(
                has_text=re.compile(pattern, re.I)
            )
            for j in range(min(labs.count(), 4)):
                cb = labs.nth(j).locator("input[type='checkbox']").first
                if cb.count():
                    try:
                        if cb.is_checked():
                            continue
                    except Exception:
                        pass
                    _click_checkbox_via_label_or_js(cb, page)
        except Exception:
            pass

    # Последний проход: неотмеченные чекбоксы в форме модалки (часто один - согласие)
    try:
        form = scope.locator("form").last
        if not form.count():
            form = scope
        boxes = form.locator("input[type='checkbox']")
        n = min(boxes.count(), 8)
        for i in range(n):
            box = boxes.nth(i)
            try:
                if not box.is_visible():
                    continue
                if box.is_checked():
                    continue
                nm = (
                    (box.get_attribute("name") or "")
                    + " "
                    + (box.get_attribute("id") or "")
                ).lower()
                if any(
                    x in nm
                    for x in ("newsletter", "subscribe", "mailing", "рассыл", "уведомл")
                ):
                    continue
                _click_checkbox_via_label_or_js(box, page)
            except Exception:
                continue
    except Exception:
        pass

    page.wait_for_timeout(400)


def _interpret_response_status(result: requests.Response) -> str:
    """Пытается понять, принята ли заявка, а не просто HTTP 200."""
    text = result.text.lower()
    try:
        j = result.json()
        if isinstance(j, dict):
            if j.get("success") is True or j.get("ok") is True:
                return "УСПЕШНО (JSON)"
            if j.get("success") is False or j.get("ok") is False:
                return f"ОШИБКА API: {j}"
            if j.get("error") or j.get("errors"):
                return f"ОШИБКА API: {j}"
    except (json.JSONDecodeError, ValueError):
        pass

    if "csrf" in text and "invalid" in text:
        return "ОШИБКА: НЕВЕРНЫЙ CSRF ТОКЕН"
    if response_indicates_captcha_block(result.text):
        return "ОШИБКА: ТРЕБУЕТСЯ КАПЧА"
    if "ошибк" in text and any(
        x in text for x in ("не удалось", "не отправлен", "отклонен", "invalid", "error")
    ):
        return "ОШИБКА: ОТВЕТ СЕРВЕРА (см. лог)"

    if "спасибо" in text or "успешно" in text or "принят" in text:
        return "УСПЕШНО"
    if result.status_code == 200:
        return "УСПЕШНО (статус 200) - проверьте админку (возможна ложная успешность)"
    return f"ОШИБКА (статус {result.status_code})"


def scenario_blocks_from_page(страница: dict) -> list:
    """
    Список сценариев на странице: [{'название', 'включено', 'шаги'}, ...].
    Поддержка: ключ «сценарии»; иначе legacy - плоский список шагов в «шаги».
    """
    if "сценарии" in страница and страница["сценарии"] is not None:
        cs = страница["сценарии"] or []
        return [c for c in cs if isinstance(c, dict)]
    legacy = страница.get("шаги") or []
    if not legacy:
        return []
    first = legacy[0]
    if not isinstance(first, dict):
        return []
    if "действие" in first:
        title = str(страница.get("название_сценария") or "").strip()
        return [
            {"название": title or "Сценарий", "включено": True, "шаги": list(legacy)}
        ]
    if "шаги" in first and "действие" not in first:
        return [c for c in legacy if isinstance(c, dict)]
    return []


# Единый словарь алиасов действий шага (рус + англ). Используется и в сценариях, и в «подготовке».
_STEP_ACTION_ALIASES = {
    "wait": "пауза",
    "pause": "пауза",
    "sleep": "пауза",
    "click": "клик",
    "hover": "наведение",
    "modal": "модалка",
    "form": "форма",
    "goto": "перейти",
    "navigate": "перейти",
    "url": "перейти",
}


def normalize_step_action(raw) -> str:
    """Нормализованное имя действия шага (англ. алиасы → рус.)."""
    a = (raw or "").strip().lower()
    return _STEP_ACTION_ALIASES.get(a, a)


def prep_steps_from_page(страница: dict) -> list:
    """
    Шаги «подготовки» страницы (ключ «подготовка», алиас «prep»).

    Это best-effort действия сразу после загрузки страницы: снять оверлей
    (например подтверждение города), навести на карточку и т.п. - то, что
    раньше было зашито в коде под конкретные сайты. Возвращает только
    включённые шаги с нормализованным действием (пауза / клик / наведение).
    """
    if not isinstance(страница, dict):
        return []
    raw = страница.get("подготовка")
    if raw is None:
        raw = страница.get("prep")
    if not raw:
        return []
    out = []
    for s in raw:
        if not isinstance(s, dict):
            continue
        if not cfg_enabled(s.get("включено", True)):
            continue
        step = dict(s)
        step["действие"] = normalize_step_action(
            s.get("действие") or s.get("action") or ""
        )
        out.append(step)
    return out


# --- Лог Excel: единый формат колонок (создание файла и запись строк) ---
# Технические колонки («Режим», «Тип», «Значение типа», «Код») убраны из отчёта.
# Статус приводим к короткому слову (Успешно / Заполнено / Ошибка), а причину
# неудачи кладём в колонку «Комментарий» (см. _status_clean_reason).
LOG_HEADERS = [
    "Дата", "Время", "Город", "Страница", "URL",
    "Название", "Где находится", "Имя", "Телефон", "Почта", "Почта получателя",
    "Статус", "Уведомление пользователю", "Типы файлов формы",
    "Выпадающие списки", "Чекбоксы/радио", "Двойная отправка",
    # Данные формы дошли до сервера (видно в DevTools → Network) - снимок
    # значений полей перед кликом сверяется с реально пойманным телом POST
    # той же единственной легитимной отправки. НЕ путать с «Серверная
    # валидация» ниже - то проверяет обратное направление.
    "Данные дошли до сервера",
    "Enter отправляет", "Поля очищены",
    # Пункт «Форма стилизована по макету»: консистентность полей + факт
    # стилизации (не браузерный дефолт). «Консистентно / Разнобой / Дефолт».
    "Стилизация полей",
    # Пункт 2.13 (согласие на обработку ПД + политика) - разбит на отдельные
    # колонки, как в чек-листах (раньше всё лежало одной строкой в «Комментарий»).
    "Наличие чек боксов согласия", "Чек боксы согласия не предустановлены",
    "Ссылка на политику", "Без согласия не отправить",
    # Защита от XSS (проба под галочкой) - «Защищена / УЯЗВИМА / Проверить».
    "Защита от XSS",
    # CSRF: наличие токена/поля защиты сессии, если требуется - «Есть /
    # Нет / Проверить». Проверяется всегда (пассивно, без лишних запросов).
    "CSRF-защита",
    # Серверная валидация (проба под галочкой, отдельная строка после обычной
    # отправки) - «Защищена / УЯЗВИМА / Проверить».
    "Серверная валидация",
    # Лимит запросов / защита от спама-ботов: пассивный слой всегда («Есть
    # защита» / «Не обнаружено»), активный залп - отдельная строка под галочкой.
    "Защита от спама (пассивно)", "Защита от спама (активно)",
    # Модальные окна работают корректно (если есть) - «открывается» только
    # у выделенного блока «модалки» (Да/Нет); «закрывается» - у любой формы,
    # найденной внутри модалки/попапа, отдельной строкой после отправки.
    "Модалка открывается", "Модалка закрывается",
    "Обработка ошибок",
    "Автозаполнение полей",
    "Подсказки полей",
    # Пункт «Ошибки валидации отображаются корректно (цвета, текст ошибок)»:
    # показывает ли форма видимую ошибку при пустом/невалидном вводе.
    # «есть / нет / без реакции / не найдено / проверить вручную».
    "Ошибки валидации",
    "Кнопка по заполнению",
    "Комментарий",
]

# Ключи строки-словаря в порядке колонок LOG_HEADERS.
LOG_KEYS_ORDER = [
    "дата", "время", "город", "страница", "url",
    "название", "где", "имя", "телефон", "почта", "почта_получателя",
    "статус", "уведомление", "типы_файлов", "выпадающие_списки",
    "чекбоксы_радио", "двойная_отправка", "данные_дошли", "enter_отправляет",
    "поля_очищены",
    "стилизация",
    "согласие_чекбоксы", "согласие_предустановка", "согласие_ссылка",
    "согласие_обязательно",
    "защита_от_xss",
    "csrf_защита",
    "серверная_валидация",
    "защита_от_спама_пассивно", "защита_от_спама_активно",
    "модалка_открылась", "модалка_закрывается",
    "обработка_ошибок",
    "автозаполнение",
    "подсказки",
    "ошибки_валидации",
    "кнопка_обязательные",
    "комментарий",
]

# Пояснения «что и КАК именно проверяется» к части колонок «Логов» - тем же
# способом, что и BLOCK_DESCRIPTIONS в content_checker.py: ложатся в комментарий
# к заголовку столбца в log_forms.xlsx (наведите мышь), чтобы человек, который
# впервые открыл отчёт, понимал механику проверки, а не только название.
# Заполнено не для всех 30 столбцов - только там, где описание уже написано;
# остальные заголовки просто идут без комментария (см. init_excel_log).
FORM_LOG_DESCRIPTIONS = {
    "CSRF-защита":
        'Проверяется ВСЕГДА, пассивно (без лишних запросов) - смотрим уже '
        'заполненную форму. Ищем скрытое поле <input type="hidden"> с именем '
        'вроде csrf/sessid/_token/authenticity_token/requestverificationtoken/'
        'xsrf/nonce (вокабуляр общий, не только под Bitrix). «Есть» - поле '
        'найдено и в нём уже стоит значение. «Проверить» - поле есть, но '
        'пустое (может заполняться иначе), либо форму не удалось прочитать. '
        '«Нет» - такого поля не нашли вовсе. Активно обойти защиту (отправить '
        'без токена или с чужим значением) тул не пытается - это отдельная, '
        'пока не реализованная проверка.',
    "Данные дошли до сервера":
        'Проверяется ВСЕГДА, без единого лишнего запроса - на той же самой '
        'единственной легитимной отправке формы (включая заказы). Тул '
        'снимает значения видимых полей ПРЯМО ПЕРЕД кликом отправки, слушает '
        'реальный сетевой POST-запрос той же отправки и сверяет: каждое '
        'заполненное значение должно найтись в теле запроса (точно, без '
        'учёта регистра, либо по последним 10 цифрам для телефона - если '
        'сайт переформатировал маску, либо по началу строки для длинного '
        'текста - если сайт обрезал по своему maxlength; это не считается '
        'потерей данных). «Да» - все значения нашли. «Не долетело» - '
        'какое-то значение пропало (указаны поля) - ловит баг «сайт '
        'нарисовал Спасибо, но реально отправил не то, что было в полях». '
        '«Проверить» - не поймали ни одного POST (например обычная '
        'навигация без AJAX) или заполненных полей не было. НЕ путать с '
        '«Серверная валидация» ниже - там проверяется обратное направление: '
        'что БИТЫЕ данные сервер отклонит, а не что ХОРОШИЕ не потеряются.',
    "Серверная валидация":
        'Только под галочкой «Проба серверной валидации», после обычной '
        'отправки формы (чтобы не испортить остальные колонки её строки). '
        'Тул через page.evaluate() снимает html5-ограничения у поля '
        '(required/pattern/maxlength/type) и жмёт настоящую кнопку отправки - '
        'как если бы человек в DevTools обошёл проверку на странице и отправил '
        'её руками. Пробует до 3 вариантов, пропуская неприменимые: пустое '
        'обязательное поле, некорректный e-mail, слишком длинный текст. '
        '«Защищена» - сервер отклонил все проверенные варианты. «УЯЗВИМА» - '
        'хотя бы один вариант сервер принял как валидный (заявка реально ушла '
        'в админку клиента с тестовой меткой в имени - после проверки её '
        'нужно удалить вручную). «Проверить» - не нашлось подходящих полей.',
    "Защита от спама (пассивно)":
        'Проверяется ВСЕГДА, без единого лишнего запроса - только по уже '
        'загруженной странице. Ищем два признака: 1) капча-виджет по классу/'
        'скрипту (reCAPTCHA, hCaptcha, Cloudflare Turnstile, Яндекс '
        'SmartCaptcha); 2) honeypot-поле - скрытый input с типовым '
        '«ловушечным» именем (hideit/honeypot/trap/bot-field и т.п. - боты '
        'его слепо заполняют, реальные люди не видят). «Есть защита» - нашли '
        'хотя бы один признак. «Не обнаружено» - видимых признаков нет; это '
        'НЕ доказывает отсутствие лимита на сервере, только что снаружи по '
        'разметке его не видно (правку см. в колонке «активно»).',
    "Защита от спама (активно)":
        'Только под отдельной галочкой «Активная проверка лимита запросов» '
        '(по умолчанию выключена). Тул реально отправляет ту же форму 3 раза '
        'подряд за несколько секунд, валидными данными с тестовой меткой - '
        'как если бы кто-то вручную быстро закинул несколько заявок - и после '
        'каждой попытки смотрит на реакцию сайта (текст про блок/лимит/капчу '
        'на странице). «Сработала защита» - одна из попыток была '
        'заблокирована (указан номер попытки). «Не сработала за 3 попытки» - '
        'все 3 прошли одинаково успешно (не значит, что лимита нет вообще - '
        'только что 3 быстрые попытки его не выявили). Число попыток жёстко '
        'зашито (3) и не растёт от повторных запусков галочки - но каждый '
        'запуск и правда кладёт до 3 тестовых заявок в админку, если защиты нет.',
    "Модалка открывается":
        'У формы, размещённой в модальном окне/попапе, тул кликает по '
        'кнопке-триггеру и проверяет не просто видимость контейнера, а что '
        'внутри него реально появилось хотя бы одно поле ввода - иначе это '
        'мог быть пустой fallback-селектор, а не настоящая открывшаяся '
        'модалка. «Да»/«Нет» по факту.',
    "Модалка закрывается":
        'После работы с формой внутри модалки тул по очереди пробует три '
        'штатных способа закрытия и смотрит, исчез ли контейнер: 1) клик по '
        'кнопке-крестику (типовые селекторы вроде [data-dismiss=modal], '
        '.modal-close, .btn-close, aria-label «закрыть»); 2) клавиша Esc; '
        '3) клик по фону вне окна модалки. «Да» - и каким именно способом '
        'из трёх получилось. «Нет» - ни один из трёх способов не закрыл окно '
        '(жёсткий вердикт без «проверить вручную» - тул сам доходит до ответа).',
    "Подсказки полей":
        'Placeholder/подсказка поля не должна ПРОТИВОРЕЧИТЬ его назначению '
        '(ловим copy-paste-баги вроде поля «Телефон» с подсказкой про '
        'e-mail). Тул определяет назначение поля из type/autocomplete/name/'
        'подписи (НЕ из самого placeholder), затем отдельно смотрит, на что '
        'похож текст подсказки - и флагает только явные несовпадения из '
        'проверенного списка пар (например подсказка «похожа на почту» у '
        'поля, определённого как телефон). Хорошо ли ЗВУЧИТ подсказка - '
        'решает человек, тул ловит только объективное противоречие.',
    "Ошибки валидации":
        'Проверяет не сами правила валидации, а их ПОКАЗ пользователю: '
        'появляется ли видимое сообщение об ошибке при пустом/невалидном '
        'вводе. Нативную HTML5-валидацию тул ловит безопасно через '
        'checkValidity() без реальной отправки; кастомную (свой JS сайта) - '
        'контролируемым пустым сабмитом под двойной защитой (перехват submit '
        '+ обрыв запроса) - заявка никогда не уходит ни в каком случае. '
        '«Есть» - браузер или сайт показывают ошибку. «Нет»/«без реакции» - '
        'обязательное поле можно отправить пустым без всякой видимой реакции. '
        'На форме заказа автопробу тул не делает (перестраховка на чекауте) - '
        'сразу помечает «проверить вручную».',
}

# Отдельный лист «Цели» (Яндекс.Метрика): свои колонки - форма/кнопка + идентификатор цели.
GOAL_HEADERS = [
    "Дата", "Время", "Город", "Страница", "Форма / кнопка", "Где находится",
    "Цель (идентификатор)", "URL", "Статус", "Комментарий",
]
GOAL_KEYS_ORDER = [
    "дата", "время", "город", "страница", "название", "где",
    "ид", "url", "статус", "комментарий",
]


def _строка_это_цель(row: dict) -> bool:
    """Строка относится к целям Метрики (идёт на отдельный лист «Цели»)."""
    return (str(row.get("тип", "")).upper().startswith("ЦЕЛЬ")
            or str(row.get("код", "")).startswith("ym"))


def _status_clean_reason(raw: str):
    """Из «сырого» статуса делает (короткое_слово, причина).

    Короткое слово - Успешно / Заполнено / Ошибка (без скобок и деталей).
    Причина заполняется только когда что-то не сработало (идёт в «Комментарий»),
    при успехе - пустая строка.
    """
    s = (raw or "").strip()
    up = s.upper()
    detail = ""
    m = re.search(r"[(（](.+?)[)）]", s)
    if m:
        detail = m.group(1).strip()
    elif ":" in s:
        detail = s.split(":", 1)[1].strip()
    d = detail.upper()

    if up.startswith("УСПЕШНО"):
        return "Успешно", ""
    if up.startswith("ЗАПОЛНЕНО"):
        return "Заполнено", ""
    if up.startswith("СРАБОТАЛА"):
        return "Сработала", ""
    if up.startswith("НЕ СРАБОТАЛА"):
        return "Не сработала", detail
    if up.startswith("ФОРМА НЕ НАЙДЕНА"):
        return "Ошибка", "Форма не найдена на странице - изменился селектор или она не загрузилась"
    if up.startswith("НЕТ СЕЛЕКТОРА ФОРМЫ"):
        return "Ошибка", "Не задан селектор формы в настройках"
    if up.startswith("ОШИБКА"):
        if "КАПЧА" in up:
            return "Ошибка", "Форма защищена капчей - заявка не отправилась"
        if "CSRF" in up:
            return "Ошибка", "Сайт отклонил отправку (неверный токен сессии)"
        if "СООБЩЕНИЕ НА СТРАНИЦЕ" in d or "ТЕКСТ НА СТРАНИЦЕ" in d:
            return "Ошибка", "Сайт показал сообщение об ошибке"
        if "НЕТ ПРИЗНАКА УСПЕХА" in d:
            return "Ошибка", "Нет подтверждения отправки на странице"
        if "СЦЕНАРИЙ ПРЕРВАН" in d:
            return "Ошибка", "Сценарий прервался на одном из шагов"
        if "ФОРМА ПРЕРВАНА" in d:
            return "Ошибка", "Не удалось дойти до формы"
        if d.startswith("СТАТУС"):
            return "Ошибка", f"Сервер вернул {detail.lower()}"
        if "ОТВЕТ СЕРВЕРА" in d:
            return "Ошибка", "Сервер ответил ошибкой"
        if up.startswith("ОШИБКА API"):
            return "Ошибка", "Сервер вернул ошибку (API)"
        # деталь - просто селектор (нет пробелов, есть . # или это form/div):
        # значит форму не нашли (часто из-за того, что страница не успела загрузиться)
        if detail and " " not in detail and re.fullmatch(r"[\w .#>:\[\]='\"\-]+", detail) \
                and any(ch in detail for ch in ".#[") :
            return "Ошибка", f"Форма не найдена или страница не успела загрузиться: {detail}"
        if "НЕ НАЙДЕН" in d or "НЕТ ЭЛЕМЕНТОВ" in d:
            return "Ошибка", "Форма не найдена или страница не успела загрузиться"
        return "Ошибка", detail or "Не удалось отправить - подробности в подробном логе"
    # запасной случай: короткое слово без скобок/деталей
    base = re.split(r"[(:（]", s, 1)[0].strip()
    return (base.capitalize() if base else s), ""


def _atomic_save_wb(wb, path: str) -> None:
    """Атомарное сохранение книги: пишем во ВРЕМЕННЫЙ файл рядом и переименовываем
    (rename атомарен на одном диске). Так log_forms.xlsx НИКОГДА не остаётся полу-
    записанным: сбой сохранения (нет места на диске, обрыв) оставляет ПРЕДЫДУЩИЙ
    валидный файл целым - теряется максимум одна текущая строка, а не весь отчёт.
    Раньше запись шла прямо в файл: оборванное сохранение било его, и все
    последующие чтения падали («Truncated file header») - терялись все прошлые
    формы прогона."""
    import os as _os
    tmp = f"{path}.tmp"
    try:
        wb.save(tmp)
        _os.replace(tmp, path)
    except Exception:
        try:
            if _os.path.exists(tmp):
                _os.remove(tmp)
        except Exception:  # noqa: BLE001
            pass
        raise


def init_excel_log(path: str, очистить: bool = True) -> None:
    """Готовит файл лога: при «очистить» удаляет старый, создаёт новый с шапкой LOG_HEADERS.
    Столбцам с описанием в FORM_LOG_DESCRIPTIONS ставится комментарий к заголовку
    (наведите мышь) - что именно и КАК проверяется, по образцу BLOCK_DESCRIPTIONS
    из content_checker.py."""
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
    if очистить and os.path.exists(path):
        try:
            os.remove(path)
            print(f"✅ Старый файл удален: {path}")
        except Exception:
            print(f"⚠️ Не удалось удалить {path} (возможно, открыт в Excel)")
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Логи"
        for col, val in enumerate(LOG_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=val)
            # стартовая ширина по заголовку (потом подрастёт под содержимое)
            ws.column_dimensions[get_column_letter(col)].width = len(str(val)) + 3
            desc = FORM_LOG_DESCRIPTIONS.get(val)
            if desc:
                cell.comment = Comment(desc, 'Site Checker', height=200, width=340)
        _atomic_save_wb(wb, path)
        print(f"✅ Создан новый Excel файл: {path}")


def append_log_row(path: str, row: dict) -> None:
    """Добавляет строку в конец файла. Строки форм/сценариев идут на лист «Логи»,
    строки целей Метрики - на отдельный лист «Цели». Колонку «Статус» красит:
    зелёный - Успешно/Заполнено/Зафиксирована, красный - Ошибка."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    wb = load_workbook(path)

    # Подсказка «где находится форма» - из общего словаря по названию, если
    # строка её ещё не несёт (заполняем и для форм, и для целей-кнопок).
    if not row.get("где"):
        try:
            from form_locations import where as _where_form
            row = {**row, "где": _where_form(row.get("название", ""))}
        except Exception:
            pass

    цель_строка = _строка_это_цель(row)
    if цель_строка:
        # лист «Цели» создаём при первой цели
        if "Цели" in wb.sheetnames:
            ws = wb["Цели"]
        else:
            ws = wb.create_sheet("Цели")
            for col, val in enumerate(GOAL_HEADERS, 1):
                ws.cell(1, col, val)
                ws.column_dimensions[get_column_letter(col)].width = len(str(val)) + 3
        keys, headers = GOAL_KEYS_ORDER, GOAL_HEADERS
    else:
        # Пишем строго в лист «Логи» (а не в активный): рядом есть «Сводка»/«Цели».
        ws = wb["Логи"] if "Логи" in wb.sheetnames else wb.active
        keys, headers = LOG_KEYS_ORDER, LOG_HEADERS

    r = ws.max_row + 1
    for col, key in enumerate(keys, 1):
        val = row.get(key, "")
        ws.cell(r, col, val)
        # авто-ширина: растим колонку под содержимое (с разумным потолком).
        # «Комментарий» тянем шире (там длинные пояснения) - до 120.
        letter = get_column_letter(col)
        cur = ws.column_dimensions[letter].width or (len(headers[col - 1]) + 3)
        _cap = 120 if str(headers[col - 1]).strip().lower() == "комментарий" else 70
        ws.column_dimensions[letter].width = min(max(cur, len(str(val)) + 3), _cap)
    try:
        si = keys.index("статус") + 1
        sval = str(row.get("статус", "")).strip().lower()
        if sval in ("успешно", "заполнено", "зафиксирована", "сработала"):
            ws.cell(r, si).font = Font(color="1E8E3E", bold=True)   # зелёный
        elif sval.startswith("ошибк") or sval.startswith("не сработала"):
            ws.cell(r, si).font = Font(color="C62828", bold=True)   # красный
    except Exception:
        pass
    # Пункт 2.7: колонку «Уведомление пользователю» подсветим (Да - зелёный,
    # Нет - оранжевый, чтобы обратить внимание). Пустую не трогаем.
    try:
        ui = keys.index("уведомление") + 1
        uval = str(row.get("уведомление", "")).strip()
        if uval.startswith("Да"):
            ws.cell(r, ui).font = Font(color="1E8E3E", bold=True)
        elif uval.startswith("Нет"):
            ws.cell(r, ui).font = Font(color="B26A00", bold=True)
    except Exception:
        pass
    _atomic_save_wb(wb, path)


def консолидировать_форм_строки(path: str) -> None:
    """Пост-обработка отчёта: сводит все строки ОДНОЙ формы в ОДНУ строку.

    Раньше каждая проверка формы (Состав, Стилизация, Списки, Чекбоксы, Enter,
    Двойная отправка, Обработка ошибок, Ошибки валидации, Согласие 2.13, Поля
    2.14, Вёрстка …) писалась отдельной строкой - на форму выходило ~11 строк,
    отчёт был нечитаем. Здесь группируем строки листа «Логи» по (Город, Страница,
    имя формы) и склеиваем: каждая колонка-проверка берёт своё значение, «Статус»
    - от самой отправки формы, «Комментарий» - объединение пояснений.

    Имя формы: у строк-проверок оно идёт после «префикс: » (напр. «Стилизация
    формы (…): Заказ звонка»), у самой формы - без префикса. Ни одно имя формы
    двоеточия с пробелом не содержит, поэтому делим по первому «: ».

    Колонко-независимо (читаем реальную шапку) - переживает вставку «Статуса в
    админке» и любые новые колонки. Идемпотентно. При ошибке файл не трогаем.
    Лист «Цели» и «Сводка» не затрагиваются."""
    from openpyxl.styles import Font
    from collections import OrderedDict
    try:
        wb = load_workbook(path)
    except Exception:  # noqa: BLE001
        return
    if "Логи" not in wb.sheetnames:
        return
    ws = wb["Логи"]
    if ws.max_row < 2:
        return
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) if h is not None else "" for h in rows[0]]
    data = rows[1:]
    idx = {h: i for i, h in enumerate(hdr)}
    # Нужны эти колонки; если шапка нестандартная - тихо выходим (не рискуем).
    for нужн in ("Город", "Страница", "Название", "Статус", "Комментарий"):
        if нужн not in idx:
            return
    GI, PI, NI, SI, CI = (idx["Город"], idx["Страница"], idx["Название"],
                          idx["Статус"], idx["Комментарий"])

    def _base(n):
        n = str(n or "")
        return n.split(": ", 1)[1] if ": " in n else n

    def _empty(v):
        return v is None or (isinstance(v, str) and v.strip() in ("", "-"))

    groups = OrderedDict()
    for r in data:
        key = (r[GI], r[PI], _base(r[NI]))
        groups.setdefault(key, []).append(r)

    # Уже сведено (нет строк-проверок с «: ») - второй раз не трогаем.
    if len(groups) == len(data):
        return

    merged = []
    for (_g, _p, bn), grp in groups.items():
        row = [None] * len(hdr)
        for col in range(len(hdr)):
            for r in grp:
                if not _empty(r[col]):
                    row[col] = r[col]
                    break
        row[NI] = bn
        # «Статус» - от строки самой формы (без префикса); иначе худший из группы.
        bare = [r for r in grp if ": " not in str(r[NI] or "")]
        if bare:
            # Обычно у формы одна «голая» строка. После сшивки оформления заказа
            # (шаг-форма «данные покупателя» + шаг-проверка «Оформление заказа»
            # под одним именем) их ДВЕ: «ЗАПОЛНЕНО (без отправки)» и настоящий
            # вердикт заказа. Берём НАСТОЯЩИЙ вердикт, а не «заполнено».
            _bare_st = [str(r[SI]) for r in bare if r[SI] not in (None, "")]
            row[SI] = next((s for s in _bare_st
                            if not s.lower().startswith("заполнено")),
                           _bare_st[0] if _bare_st else bare[0][SI])
        else:
            статусы = [str(r[SI]) for r in grp if r[SI] not in (None, "")]
            row[SI] = next((s for s in статусы
                            if s.lower().startswith(("ошибк", "проверить"))),
                           статусы[0] if статусы else "")
        # «Комментарий» - объединяем непустые пояснения проверок (без повторов).
        коммы = []
        for r in grp:
            c = str(r[CI] or "").strip()
            if c and c not in коммы:
                коммы.append(c)
        row[CI] = " | ".join(коммы)
        merged.append(row)

    # Переписываем данные листа: чистим строки со 2-й, пишем сведённые, красим.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    for r_i, row in enumerate(merged, start=2):
        for c_i, val in enumerate(row, start=1):
            ws.cell(r_i, c_i, val)
        try:
            sval = str(row[SI] or "").strip().lower()
            if sval in ("успешно", "заполнено", "зафиксирована", "сработала"):
                ws.cell(r_i, SI + 1).font = Font(color="1E8E3E", bold=True)
            elif sval.startswith("ошибк") or sval.startswith("не сработала"):
                ws.cell(r_i, SI + 1).font = Font(color="C62828", bold=True)
        except Exception:  # noqa: BLE001
            pass
        # «Уведомление пользователю» - подсветка Да/Нет (как в append_log_row).
        try:
            ui = idx.get("Уведомление пользователю")
            if ui is not None:
                uval = str(row[ui] or "").strip()
                if uval.startswith("Да"):
                    ws.cell(r_i, ui + 1).font = Font(color="1E8E3E", bold=True)
                elif uval.startswith("Нет"):
                    ws.cell(r_i, ui + 1).font = Font(color="B26A00", bold=True)
        except Exception:  # noqa: BLE001
            pass
    try:
        _atomic_save_wb(wb, path)
        print(f"   🧹 Отчёт сведён: {len(data)} строк → {len(merged)} (1 форма = 1 строка)")
    except Exception as e:  # noqa: BLE001
        print(f"   ⚠️ Консолидация отчёта не сохранена: {e}")


# Идент. колонки (не проверки) - не попадают в строки «Проверка» матрицы.
_МАТРИЦА_IDENT = {"Дата", "Время", "Город", "Страница", "URL", "Название",
                  "Где находится", "Имя", "Телефон", "Почта", "Почта получателя",
                  "Комментарий"}
# Справочные колонки: НЕ дефекты (не «прошло/не прошло»), а информация. Раньше
# они давали ⚠ на каждой форме (шум), хотя ошибкой не являются. Убираем из
# матрицы «прошло/не прошло», но ОСТАВЛЯЕМ в листе «Логи» (данные не теряются):
#   • «Поля очищены» - код сам помечает статусом OK, «не критично»;
#   • «Защита от спама (пассивно)» - «снаружи не видно» не доказывает отсутствие
#     защиты; реальную защиту показывает строка «активно» (поведенческая проба).
_МАТРИЦА_СКРЫТЬ = {"Поля очищены", "Защита от спама (пассивно)"}
# Служебные строки (не формы) - не становятся столбцами матрицы.
_МАТРИЦА_SKIP_NAME = ("согласие и политика", "cookie-уведомление",
                      "ссылка на политику", "живочат", "мобильная вёрстка",
                      "вёрстка на устройствах")
# submit-статусы - только они превращают строку «Логи» в столбец-форму матрицы
# (у строк мобильной вёрстки статус «OK», у части служебных - города нет вовсе).
_МАТРИЦА_SUBMIT_ST = ("успешно", "ошибка", "заполнено", "нет на сайте",
                      "проверить", "сработала", "не сработала",
                      "нет подтверждения")

# (колонка) -> [(паттерн_в_значении, символ, пояснение_на_✗_или_⚠|None), …]
# Пояснение - ГОТОВАЯ фраза, НЕ повторяет сырое значение колонки. На ✓/– коммент
# не нужен - и так ясно. Первое совпадение по порядку побеждает.
_МАТРИЦА_ПРАВИЛА = {
    "Статус": [
        ("ошибк", "✗", "Форма не отправилась – подробности в листе «Логи»."),
        ("не сработ", "✗", "Форма не отправилась – подробности в листе «Логи»."),
        ("нет подтверждения", "✗", "После отправки НЕТ подтверждения пользователю и ответа сервера об успехе – для посетителя форма выглядит нерабочей (заявка могла не уйти). Проверьте форму вручную и приход заявки."),
        ("нет на сайте", "–", None),
        ("успешно", "✓", None), ("заполнено", "✓", None), ("сработ", "✓", None),
        ("ok", "✓", None), ("ок", "✓", None),
    ],
    "Статус в админке": [
        ("не найд", "✗", "Заявка не найдена в админке – уведомление с формы не пришло или не сохранилось."),
        ("есть в админке", "✓", None), ("найд", "✓", None),
    ],
    "Уведомление пользователю": [
        ("нет", "✗", "После отправки пользователь не видит подтверждения, что заявка принята."),
        ("да", "✓", None),
    ],
    "Типы файлов формы": [
        ("опасн", "✗", "Поле загрузки приняло файл с опасным расширением – сервер не фильтрует."),
        ("посторонн", "✗", "Сервер принял посторонний формат – по правилу разрешены только PDF/DOC/DOCX."),
        ("больше", "✗", "Сервер принял файл больше 20 КБ – лимит размера не проверяется."),
        ("любые", "✗", "Поле загрузки принимает любые файлы – тип не ограничен (accept не задан)."),
        ("проверить", "⚠", "Фильтр файлов проверен не полностью – проверьте вручную (см. комментарий в листе «Логи»)."),
    ],
    "Выпадающие списки": [
        ("ошибк", "✗", "Список пустой – вариантов для выбора нет."),
        ("не найд", "–", None), ("корректно", "✓", None),
    ],
    "Чекбоксы/радио": [
        ("ошибк", "✗", "Чекбокс/радио не переключается кликом."),
        ("не найд", "–", None), ("корректно", "✓", None),
    ],
    "Двойная отправка": [
        ("не защищ", "✗", "Кнопка не блокируется после клика – форму можно отправить дважды подряд."),
        ("под вопрос", "⚠", "Не удалось однозначно проверить защиту от двойной отправки – проверьте вручную."),
        ("защищ", "✓", None),
    ],
    "Данные дошли до сервера": [
        # Реальный вердикт функции - «Не долетело» (не «Нет»): раньше правило
        # ловило только «нет», а «не долетело» не содержит подстроки «нет» и
        # проваливалось в дефолт ⚠ «проверьте вручную» - ложная ручная проверка
        # на НАСТОЯЩЕЙ потере данных. Ловим и «не долет…», и «не дошл…», и «нет».
        ("не долет", "✗", "Заполненные поля НЕ найдены в теле POST-запроса – часть данных до сервера не дошла."),
        ("не дошл", "✗", "Заполненные поля НЕ найдены в теле POST-запроса – часть данных до сервера не дошла."),
        ("нет", "✗", "Заполненные поля НЕ найдены в теле POST-запроса – часть данных до сервера не дошла."),
        ("проверить", "⚠", "Не удалось поймать сетевой POST отправки (форма без явного POST/переход) – проверьте вручную."),
        ("да", "✓", None),
    ],
    "Enter отправляет": [
        ("нет", "⚠", "Форму нельзя отправить нажатием Enter, только кнопкой – проверьте, ожидается ли это."),
        ("да", "✓", None),
    ],
    "Поля очищены": [
        ("не очищ", "⚠", "Поля после отправки не очищаются."),
        ("очищ", "✓", None),
    ],
    "Стилизация полей": [
        ("разнобой", "✗", "Поля формы визуально различаются между собой (шрифт/рамка/радиус) – похоже на баг вёрстки."),
        ("дефолт", "✗", "Поля не оформлены – браузерный вид по умолчанию, вёрстка не применена."),
        ("не найд", "–", None), ("консистентно", "✓", None),
    ],
    "Чек боксы согласия не предустановлены": [
        ("нет", "✗", "Чекбокс согласия отмечен галочкой заранее – должен быть пустым."),
        ("да", "✓", None),
    ],
    "Ссылка на политику": [
        ("нет", "✗", "На форме нет ссылки на политику обработки персональных данных."),
        ("да", "✓", None),
    ],
    "Без согласия не отправить": [
        ("нет", "✗", "Форму можно отправить, не поставив галочку согласия."),
        ("да", "✓", None),
    ],
    "Защита от XSS": [
        ("уязв", "✗", "Внедрённый код выполнился на странице – форма уязвима к XSS."),
        ("проверить", "⚠", "XSS не проверен – форма не отправилась (капча/ошибка)."),
        ("защищ", "✓", None),
    ],
    "Обработка ошибок": [
        # «молчит» на искусственно оборванный запрос больше не помечаем — это был
        # ложный ✗ на рабочих формах. Оставляем только реальный дефект: форма
        # показывает «успешно», хотя запрос упал (вводит пользователя в заблуждение).
        ("ложный успех", "✗", "При сбое отправки форма показывает «успешно» – пользователь решит, что заявка ушла, а её нет."),
        ("ошибка на шаге", "✗", "После клика «Далее»/«Оформить заказ» форма пропустила дальше, но на странице осталась ошибка – заказ не оформить."),
        ("корректно", "✓", None),
    ],
    "Автозаполнение полей": [
        ("корректно", "✓", None),
        ("не", "⚠", "Автозаполнение работает некорректно – проверьте вручную."),
    ],
    "Подсказки полей": [
        ("корректно", "✓", None),
        ("нет подсказок", "⚠", "У полей формы нет подсказок (placeholder) – только подписи, если они есть."),
        ("нет", "⚠", "У части полей формы нет подсказок (placeholder)."),
    ],
    "Ошибки валидации": [
        ("без реакции", "✗", "На пустую отправку форма никак не реагирует – ошибка не показывается."),
        ("вручную", "⚠", "Нет автоматической проверки (форма заказа) – откройте форму и проверьте вручную."),
        ("не найд", "–", None),
        ("есть", "✓", None),
        ("нет", "✗", "Пустая форма пытается отправиться – клиентской валидации нет."),
    ],
    "CSRF-защита": [
        ("нет", "✗", "Нет скрытого поля токена сессии (CSRF/sessid) – если на сайте нет другой защиты (SameSite-cookie), форма может быть уязвима к подделке запроса."),
        ("проверить", "⚠", "Не удалось однозначно проверить CSRF-токен – проверьте вручную."),
        ("есть", "✓", None),
    ],
    "Серверная валидация": [
        ("уязвима", "✗", "Сервер принял заведомо невалидные данные (пустые/битые поля) – валидация только на клиенте, легко обойти прямым запросом."),
        ("проверить", "⚠", "Автопроба не дала однозначного ответа (форма блокируется уже после первой отправки, либо сервер не ответил на повтор). Проверьте ВРУЧНУЮ: откройте форму, впишите в e-mail мусор без «@» (или оставьте обязательное поле пустым в обход подсказки) и отправьте – если заявка ушла без ошибки, сервер данные не проверяет."),
        ("защищена", "✓", None),
    ],
    "Защита от спама (пассивно)": [
        ("не обнаружено", "⚠", "Видимых признаков защиты (капча/honeypot) нет – это не доказывает отсутствие лимита на сервере, только что снаружи не видно."),
        ("есть защита", "✓", None),
    ],
    "Защита от спама (активно)": [
        ("не сработала", "⚠", "3 быстрых повторных отправки прошли одинаково успешно – лимит не сработал на этом масштабе (не значит, что его нет вообще)."),
        ("блокирует повтор", "⚠", "Кнопка становится «Отправлено» после отправки – от СЛУЧАЙНОГО двойного клика защищает, но это НЕ доказывает серверную защиту: на части форм кнопка косметическая, и повторы всё равно уходят (в админке появляется несколько заявок с одного теста). Проверьте вручную: перезагрузите страницу и отправьте форму 3–4 раза подряд."),
        ("проверить", "⚠", "Быстрые повторные отправки не дали чёткого сигнала (форма заказа или сбой пробы). Проверьте ВРУЧНУЮ: перезагрузите страницу и отправьте форму 3–4 раза подряд – если все прошли, серверного лимита нет."),
        ("сработала", "✓", None),
    ],
    "Кнопка по заполнению": [
        ("не разблокируется", "✗", "Кнопка отправки заблокирована даже после заполнения обязательных полей – форму нельзя отправить."),
        ("не найд", "–", None),
        ("не блокируется", "⚠", "Кнопка «Отправить» активна даже на пустой форме – по требованию она должна быть НЕактивной, пока не заполнены обязательные поля. Не критично (форма может валидировать по клику), но это знак внимания."),
        ("корректно", "✓", None),
    ],
    # Модалки: без явных правил «Да»/«Нет» падали в дефолт ⚠ «значение не
    # распознано» - пройденная проверка (модалка открылась/закрылась) выглядела
    # как ручная. Теперь «Да» → ✓, «Нет» → ✗, «Проверить» → ⚠.
    "Модалка открывается": [
        ("проверить", "⚠", "Не удалось проверить открытие модалки – проверьте вручную."),
        ("нет", "✗", "Модалка не открылась по клику – форма недоступна пользователю."),
        ("да", "✓", None),
    ],
    "Модалка закрывается": [
        ("проверить", "⚠", "Не удалось проверить закрытие модалки – проверьте вручную."),
        ("нет", "✗", "Модалку не удалось закрыть (крестик/Esc/клик вне окна не сработали)."),
        ("да", "✓", None),
    ],
}


def _матрица_тире(s: str) -> str:
    """Длинное тире «—» → короткое «–» (везде в тексте, который сами пишем)."""
    return (s or "").replace("—", "–")


def _матрица_классифицировать(col: str, val) -> tuple:
    """→ (символ, коммент|None). Пустое исходное значение → ('–', None) - в
    матрице никогда нет пустых ячеек: либо символ, либо «–» (не применимо)."""
    v = str(val or "").strip().lower()
    if v in ("", "-", "—", "–"):
        return "–", None
    for pat, sym, cm in _МАТРИЦА_ПРАВИЛА.get(col, []):
        if pat in v:
            return sym, cm
    if col == "Наличие чек боксов согласия":
        m = re.match(r"\s*(\d+)", v)
        if m:
            n = int(m.group(1))
            return ("✓", None) if n >= 2 else ("✗", f"Чекбоксов согласия {n} (по чек-листу нужно ≥2).")
    if col == "Типы файлов формы":
        return "✓", None
    # Умный дефолт: если для колонки НЕТ явного правила (например, добавили новую
    # проверку и забыли правило) - не пугаем оператора ⚠ на очевидно-успешном
    # значении. «да/есть/ок/успешно/пройдено» → ✓; иначе честно ⚠ «проверьте».
    if re.match(r"^(да|есть|ок|ok|успешно|пройдено|корректно|защищена)\b", v):
        return "✓", None
    return "⚠", f"Значение «{val}» не распознано – проверьте вручную."


def _снять_size_with_cells(path: str) -> None:
    """openpyxl ЖЁСТКО пишет <x:SizeWithCells/> в каждую заметку (см.
    openpyxl/comments/shape_writer.py) - из-за этого Excel ужимает окно заметки
    до размера узкой ячейки матрицы, и длинный комментарий обрезается («не
    влезает»). Убираем этот флаг в уже сохранённом файле: тогда Excel держит
    заметку в её СОБСТВЕННОМ размере (width/height из VML) и текст помещается.
    MoveWithCells оставляем (заметка ездит с ячейкой). Всё в try - на любой
    ошибке файл остаётся как был (просто без этого улучшения)."""
    import zipfile
    import os
    import shutil
    import tempfile
    try:
        with zipfile.ZipFile(path, "r") as z:
            names = z.namelist()
            data = {n: z.read(n) for n in names}
    except Exception:  # noqa: BLE001
        return
    pat = re.compile(rb"<[A-Za-z0-9_]*:?SizeWithCells\s*/>")
    # Гарантируем МИНИМАЛЬНЫЙ размер окна заметки прямо в VML - не полагаясь на
    # то, что openpyxl (в разных версиях по-разному) применил width/height из
    # Comment(). Маленькое окно = обрезанный текст. Большое оставляем как есть.
    pat_size = re.compile(rb"width:(\d+)px;height:(\d+)px")

    def _min_size(m):
        w = max(400, int(m.group(1)))
        h = max(150, int(m.group(2)))
        return ("width:%dpx;height:%dpx" % (w, h)).encode()

    changed = False
    for n in names:
        if n.lower().endswith(".vml"):
            new = pat.sub(b"", data[n])
            new = pat_size.sub(_min_size, new)
            if new != data[n]:
                data[n] = new
                changed = True
    if not changed:
        return
    tmp = None
    try:
        fd, tmp = tempfile.mkstemp(suffix=".xlsx",
                                   dir=os.path.dirname(path) or ".")
        os.close(fd)
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as z:
            for n in names:
                z.writestr(n, data[n])
        shutil.move(tmp, path)
    except Exception:  # noqa: BLE001
        try:
            if tmp and os.path.exists(tmp):
                os.remove(tmp)
        except Exception:  # noqa: BLE001
            pass


def построить_матрицу_проверок(path: str) -> None:
    """Наглядный отчёт поверх консолидированных «Логов»: по листу на домен/город
    («Москва», «Алматы», …, без слова «Матрица» в названии) - проверки строками,
    формы столбцами, ✓/✗/⚠/– по цветам (как в pricing-table). На каждой ✗/⚠ -
    всплывающая подсказка с ПОНЯТНОЙ причиной (без повтора сырого значения).
    Легенда символов - горизонтальной таблицей «Обозначение/Значение» на
    листе «Сводка» (колонки G:H, в стиле её шапки).

    Вызывается ПОСЛЕ консолидировать_форм_строки (нужны уже сведённые строки:
    1 форма = 1 строка) и ПОСЛЕ вставки колонок админки/письма (нужна финальная
    шапка). Идемпотентно - старые листы «по городам» пересоздаются заново.
    При любой ошибке файл не трогает (отчёт остаётся рабочим и без матрицы)."""
    from collections import OrderedDict
    try:
        wb = load_workbook(path)
    except Exception:  # noqa: BLE001
        return
    if "Логи" not in wb.sheetnames:
        return
    ws = wb["Логи"]
    if ws.max_row < 2:
        return
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) if h is not None else "" for h in rows[0]]
    data = rows[1:]
    idx = {h: i for i, h in enumerate(hdr)}
    for нужн in ("Город", "Страница", "Название", "Статус", "Где находится"):
        if нужн not in idx:
            return
    NI, GI, PI, WI, SI = (idx["Название"], idx["Город"], idx["Страница"],
                          idx["Где находится"], idx["Статус"])
    CI = idx.get("Комментарий")
    check_cols = [h for h in hdr
                  if h not in _МАТРИЦА_IDENT and h not in _МАТРИЦА_СКРЫТЬ]
    if not check_cols:
        return

    def _это_форма(r):
        if not r[GI]:
            return False
        # Мобильная вёрстка (mobile_check.py) кладёт МЕТКУ СТРАНИЦЫ в колонку
        # «Город» (у неё нет понятия города) - надёжно исключаем по фиксированному
        # значению «Страница», а не по статусу (её «Проверить» при поломке иначе
        # совпал бы с submit-статусами и подделал бы «город» под имя страницы).
        if str(r[PI] or "").strip().lower() == "мобильная вёрстка":
            return False
        st = str(r[SI] or "").strip().lower()
        if not any(st.startswith(s) for s in _МАТРИЦА_SUBMIT_ST):
            return False
        nm = str(r[NI] or "").strip().lower()
        if not nm or any(nm.startswith(s) for s in _МАТРИЦА_SKIP_NAME):
            return False
        return True

    by_city = OrderedDict()
    for r in data:
        if _это_форма(r):
            by_city.setdefault(r[GI], []).append(r)
    if not by_city:
        return

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter

    FONT = {
        "✓": Font(color="1E7A34", bold=True), "✗": Font(color="B00020", bold=True),
        "⚠": Font(color="8A6100", bold=True), "–": Font(color="9A9A9A"),
    }
    HEAD_FILL = PatternFill("solid", fgColor="EEF3FB")
    HEAD_FONT = Font(bold=True)
    CENTER = Alignment(horizontal="center", vertical="center")
    THIN = Side(style="thin", color="E3E3E3")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _коммент(text):
        # Ширина по самой длинной строке (щедро - не обрезаем), высота - по
        # реальному числу строк С УЧЁТОМ переноса на этой ширине (не только
        # по «\n» - длинное предложение само переносится и должно поместиться).
        # Кириллица в шрифте заметки шире латиницы, поэтому символов-в-строке
        # считаем КОНСЕРВАТИВНО (делим ширину на ~8.5, а не на 7) - иначе строк
        # выходит меньше реального и текст не влезает по высоте.
        text = _матрица_тире(text)
        lines = text.split("\n")
        longest = max((len(s) for s in lines), default=1)
        width = max(300, min(600, int(longest * 8) + 60))
        chars_per_line = max(18, int((width - 30) / 8))
        import math
        wrapped = sum(max(1, math.ceil(len(s) / chars_per_line)) for s in lines)
        height = max(120, 22 * wrapped + 60)
        return Comment(text, "site-checker", width=width, height=height)

    # Названия листов = сами города/домены (без слова «Матрица», без коллизий
    # с уже существующими служебными листами). Прежние листы-матрицы удаляем -
    # пересоздаём каждый прогон заново (идемпотентность).
    _служебные = {"Сводка", "Логи", "Цели"}
    _прежние_матрицы = [s for s in wb.sheetnames if s not in _служебные]
    for sn in _прежние_матрицы:
        del wb[sn]

    matrix_titles = []
    pos = 1
    for город, forms in by_city.items():
        title = str(город)[:31] or "Город"
        n, base = 2, title
        while title in wb.sheetnames or title in matrix_titles:
            suf = f" {n}"
            title = base[: 31 - len(suf)] + suf
            n += 1
        matrix_titles.append(title)
        m = wb.create_sheet(title, pos); pos += 1

        a1 = m.cell(1, 1, "Проверка"); a1.fill = HEAD_FILL; a1.font = HEAD_FONT
        for j, r in enumerate(forms, start=2):
            c = m.cell(1, j, _матрица_тире(str(r[NI]))); c.fill = HEAD_FILL
            c.font = HEAD_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            где = " · ".join(str(r[i]) for i in (PI, WI) if r[i])
            if где:
                c.comment = _коммент(где)
        # Строки-проверки. Если у ВСЕХ форм листа в проверке «–» (нигде не
        # применима - например, ни у одной формы нет выпадающих списков или
        # полей загрузки файлов), строку СКРЫВАЕМ, чтобы не мозолить глаза
        # прочерками. Если хоть у одной формы есть реальный вердикт (✓/✗/⚠) -
        # строку оставляем. Данные при этом остаются в листе «Логи».
        _label_align = Alignment(wrap_text=True, vertical="center")
        row_ptr = 2
        for col in check_cols:
            ci = idx[col]
            вычислено = []
            for r in forms:
                sym, cm = _матрица_классифицировать(col, r[ci])
                # «Статус» ✗: краткая причина ИЗ САМОГО СТАТУСА (парентеза),
                # а НЕ весь сводный «Комментарий» формы - он объединяет пояснения
                # ВСЕХ проверок и раздувался в «полотно» на всю ячейку. Полный
                # текст остаётся в листе «Логи». Для «нет подтверждения» и т.п. -
                # оставляем описательный коммент правила (он уже краткий).
                if col == "Статус" and sym == "✗":
                    _sval = str(r[SI] or "")
                    if _sval.strip().lower().startswith(("ошибк", "не сработ")):
                        _, _st_det = _status_clean_reason(_sval)
                        if _st_det:
                            cm = _st_det
                вычислено.append((sym, cm))
            if all(sym == "–" for sym, _ in вычислено):
                continue                    # нигде не применимо - не показываем
            lc = m.cell(row_ptr, 1, _матрица_тире(col))
            lc.font = Font(bold=True); lc.border = BORDER; lc.alignment = _label_align
            for j, (sym, cm) in enumerate(вычислено, start=2):
                cell = m.cell(row_ptr, j, sym); cell.alignment = CENTER; cell.border = BORDER
                f = FONT.get(sym)
                if f:
                    cell.font = f
                if cm:
                    cell.comment = _коммент(cm)
            row_ptr += 1
        m.freeze_panes = "B2"
        m.column_dimensions["A"].width = 26
        _COL_W = 22   # шире, чем раньше (17) - меньше переносов в названиях форм
        for j in range(2, len(forms) + 2):
            m.column_dimensions[get_column_letter(j)].width = _COL_W
        # Высота шапки - под САМОЕ ДЛИННОЕ название формы на этом листе (с
        # учётом переноса на ширине колонки), чтобы ни у одной формы текст не
        # обрезался по вертикали.
        _max_name_len = max((len(_матрица_тире(str(r[NI]))) for r in forms), default=20)
        _chars_per_line = max(10, _COL_W - 2)
        _lines_needed = max(1, -(-_max_name_len // _chars_per_line))
        m.row_dimensions[1].height = max(34, 16 * _lines_needed + 12)

    # ── Легенда на «Сводке»: горизонтальная таблица справа от Домен/Город/Почта. ──
    if "Сводка" in wb.sheetnames:
        sv = wb["Сводка"]
        sv.cell(1, 7, "Обозначение").fill = HEAD_FILL
        sv.cell(1, 7).font = HEAD_FONT
        sv.cell(1, 8, "Значение").fill = HEAD_FILL
        sv.cell(1, 8).font = HEAD_FONT
        legend = [
            ("✓", "успешно / защищена / консистентно / да / корректно / есть", "1E7A34"),
            ("✗", "не защищена / разнобой / нет / молчит / уязвима / «в админке не найдено»", "B00020"),
            ("⚠", "проверить вручную / под вопросом / Enter=нет / поля не очищены", "8A6100"),
            ("–", "не применимо (списков/чекбоксов на форме нет)", "9A9A9A"),
        ]
        for k, (sym, txt, color) in enumerate(legend, start=2):
            sv.cell(k, 7, sym).font = Font(color=color, bold=True)
            sv.cell(k, 8, _матрица_тире(txt)).alignment = Alignment(wrap_text=True, vertical="center")
        sv.column_dimensions["G"].width = 13
        sv.column_dimensions["H"].width = 70

    order = ["Сводка"] + matrix_titles + \
            [s for s in wb.sheetnames if s not in (["Сводка"] + matrix_titles)]
    try:
        wb._sheets.sort(key=lambda s: order.index(s.title))
    except Exception:  # noqa: BLE001
        pass
    try:
        _atomic_save_wb(wb, path)
        # Заметки не должны ужиматься до ячейки (иначе текст обрезается).
        _снять_size_with_cells(path)
        print(f"   🗂️ Матрица проверок построена: {len(matrix_titles)} лист(ов) "
              f"по домену ({', '.join(matrix_titles)})")
    except Exception as e:  # noqa: BLE001
        print(f"   ⚠️ Матрица проверок не сохранена: {e}")


# --- Уровень 1 (админка): запись реально отправленных форм для сверки ---
# После прогона forms_run сверяет этот список с «Уведомлениями с форм» в админке
# (admin_check.выполнить_проверку). Пишем только реальные отправки форм - без
# целей Метрики и без проверок корзины/оформления заказа.
SUBMITTED_FORMS_FILE = "submitted_forms.json"


def reset_submitted_forms() -> None:
    """Удаляет файл отправок (в начале свежего прогона, вместе с очисткой Excel)."""
    try:
        if os.path.exists(SUBMITTED_FORMS_FILE):
            os.remove(SUBMITTED_FORMS_FILE)
    except Exception:
        pass


def record_submitted_form(rec: dict) -> None:
    """Дописывает одну отправленную форму в submitted_forms.json (список)."""
    import json as _json
    data = []
    try:
        if os.path.exists(SUBMITTED_FORMS_FILE):
            with open(SUBMITTED_FORMS_FILE, encoding="utf-8") as fh:
                data = _json.load(fh)
            if not isinstance(data, list):
                data = []
    except Exception:
        data = []
    data.append(rec)
    try:
        with open(SUBMITTED_FORMS_FILE, "w", encoding="utf-8") as fh:
            _json.dump(data, fh, ensure_ascii=False)
    except Exception:
        pass


# --- Пункт 2.9 (письмо покупателю): запись оформленных заказов для сверки ---
# После прогона forms_run сверяет этот список с почтой покупателя (ПОЧТА):
# order_mail_check заходит в ящик и проверяет, что письмо-подтверждение заказа
# реально пришло. Пишем только успешно оформленные заказы (шаг «проверить»
# с флагом "заказ": True, завершившийся успехом).
PLACED_ORDERS_FILE = "placed_orders.json"


def reset_placed_orders() -> None:
    """Удаляет файл заказов (в начале свежего прогона, вместе с очисткой Excel)."""
    try:
        if os.path.exists(PLACED_ORDERS_FILE):
            os.remove(PLACED_ORDERS_FILE)
    except Exception:
        pass


def record_placed_order(rec: dict) -> None:
    """Дописывает один оформленный заказ в placed_orders.json (список)."""
    import json as _json
    data = []
    try:
        if os.path.exists(PLACED_ORDERS_FILE):
            with open(PLACED_ORDERS_FILE, encoding="utf-8") as fh:
                data = _json.load(fh)
            if not isinstance(data, list):
                data = []
    except Exception:
        data = []
    data.append(rec)
    try:
        with open(PLACED_ORDERS_FILE, "w", encoding="utf-8") as fh:
            _json.dump(data, fh, ensure_ascii=False)
    except Exception:
        pass


def write_summary_sheet(path: str, время_прогона: str = "") -> None:
    """Пересобирает лист «Сводка» в логе: готовое сообщение (сколько форм
    отправлено и на какие домены) + таблица «Домен → Города → Почта для
    проверки заявок». Идемпотентно: читает все строки листа «Логи».

    Вызывается в конце каждого прогона run_test; при прогоне по нескольким
    городам каждый раз пересобирается заново, поэтому после последнего города
    сводка отражает весь прогон целиком.
    """
    from urllib.parse import urlparse
    from openpyxl.styles import Font, Alignment, PatternFill

    try:
        wb = load_workbook(path)
    except Exception:
        return
    ws = wb["Логи"] if "Логи" in wb.sheetnames else wb.worksheets[0]
    headers = [str(c.value or "").strip() for c in ws[1]]

    def idx(name: str) -> int:
        for i, h in enumerate(headers):
            if h.lower() == name.lower():
                return i
        return -1

    i_url, i_st = idx("URL"), idx("Статус")
    i_mail, i_city = idx("Почта получателя"), idx("Город")

    sent = errors = 0
    domains: list[str] = []                 # в порядке появления
    dom_mails: dict[str, set] = {}
    dom_cities: dict[str, set] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v in (None, "") for v in row):
            continue
        url = str(row[i_url] or "").strip() if i_url >= 0 else ""
        st = str(row[i_st] or "").strip().lower() if i_st >= 0 else ""
        mail = str(row[i_mail] or "").strip() if i_mail >= 0 else ""
        city = str(row[i_city] or "").strip() if i_city >= 0 else ""
        if st == "успешно":
            sent += 1
        elif st.startswith("ошибк"):
            errors += 1
        if url:
            p = urlparse(url)
            dom = f"{p.scheme}://{p.netloc}" if p.netloc else url
            if dom not in dom_mails:
                domains.append(dom)
                dom_mails[dom] = set()
                dom_cities[dom] = set()
            if mail:
                dom_mails[dom].add(mail)
            if city:
                dom_cities[dom].add(city)

    # Лист «Сводка»: ТОЛЬКО таблица (Домен · Город(а) · Почта для проверки заявок).
    # Готовое сообщение/счётчики/время - по просьбе убраны из отчёта.
    if "Сводка" in wb.sheetnames:
        del wb["Сводка"]
    sm = wb.create_sheet("Сводка", 0)       # первым листом - чтобы сразу видеть

    hdr_fill = PatternFill("solid", fgColor="EEF3FB")
    for col, title in ((1, "Домен"), (2, "Город(а)"), (3, "Почта для проверки заявок")):
        c = sm.cell(1, col, title)
        c.font = Font(bold=True)
        c.fill = hdr_fill
    rr = 2
    for dom in domains:
        sm.cell(rr, 1, dom)
        sm.cell(rr, 2, ", ".join(sorted(dom_cities[dom])))
        sm.cell(rr, 3, ", ".join(sorted(dom_mails[dom])))
        rr += 1
    sm.column_dimensions["A"].width = 34
    sm.column_dimensions["B"].width = 30
    sm.column_dimensions["C"].width = 46

    # --- Лист «Логи»: жирная шапка (закреплена) + визуальное разделение городов ---
    from openpyxl.styles import Border, Side
    ncol = max(1, len(headers))
    for c in range(1, ncol + 1):
        hc = ws.cell(1, c)
        hc.font = Font(bold=True)
        hc.fill = hdr_fill
    try:
        ws.freeze_panes = "A2"
    except Exception:
        pass
    thick = Side(style="medium", color="7F7F7F")
    prev_city = None
    for r in range(2, ws.max_row + 1):
        cur = str(ws.cell(r, i_city + 1).value or "").strip() if i_city >= 0 else ""
        if prev_city is not None and cur != prev_city:
            for c in range(1, ncol + 1):
                cell = ws.cell(r, c)
                b = cell.border
                cell.border = Border(left=b.left, right=b.right, bottom=b.bottom, top=thick)
        prev_city = cur

    # --- Лист «Цели»: та же жирная шапка (фиолетовая, под «метку цели») ---
    if "Цели" in wb.sheetnames:
        gw = wb["Цели"]
        g_fill = PatternFill("solid", fgColor="EDE7F6")   # мягкий сиреневый
        gcol = max(1, len(GOAL_HEADERS))
        for c in range(1, gcol + 1):
            gc = gw.cell(1, c)
            gc.font = Font(bold=True)
            gc.fill = g_fill
        try:
            gw.freeze_panes = "A2"
        except Exception:
            pass
        # разделение по городам - как в «Логах» (город в колонке 3)
        prev = None
        for r in range(2, gw.max_row + 1):
            cur = str(gw.cell(r, 3).value or "").strip()
            if prev is not None and cur != prev:
                for c in range(1, gcol + 1):
                    cell = gw.cell(r, c)
                    b = cell.border
                    cell.border = Border(left=b.left, right=b.right,
                                         bottom=b.bottom, top=thick)
            prev = cur

    wb.active = wb.sheetnames.index("Сводка")
    _atomic_save_wb(wb, path)


def _значение_формы_для_имени(fc: dict):
    """id/class/data-source/css/name/text для плейсхолдера {значение}, без «название» теста."""
    d = fc or {}
    for k in ("id", "class", "data-source", "css", "name", "text"):
        if k in d and d[k] is not None and str(d[k]).strip() != "":
            return str(d[k]).strip()
    for k, v in d.items():
        if k in ("название", "индекс", "включено"):
            continue
        if v is not None and str(v).strip() != "":
            return str(v).strip()
    return "unknown"


def _playwright_proxy_from_env():
    """Прокси для Playwright из env FORMS_PROXY (http://user:pass@host:port).
    Нужен сайтам, которые режут прямое подключение автоматизации (напр. Метпромко):
    браузер тогда ходит через тот же прокси, что и вкладка «Переменные». Пусто -
    возвращаем None (браузер идёт напрямую, как и раньше для всех проектов)."""
    from urllib.parse import urlparse
    raw = (os.environ.get("FORMS_PROXY") or "").strip()
    if not raw:
        return None
    pr = urlparse(raw if "://" in raw else "http://" + raw)
    if not pr.hostname:
        return None
    server = f"{pr.scheme or 'http'}://{pr.hostname}"
    if pr.port:
        server += f":{pr.port}"
    conf = {"server": server}
    if pr.username:
        conf["username"] = unquote(pr.username)
    if pr.password:
        conf["password"] = unquote(pr.password)
    return conf


# Сквозной счётчик обработанных форм за прогон (все города одного запуска). Строка
# «▶ Форма N: …» - и понятный лог, и надёжный счётчик прогресса: страница считает
# ИМЕННО эти строки (ровно одна на форму/сценарий/модалку), а не разнородные
# «УСПЕШНО/ОШИБКА» (их на одну форму несколько - счётчик раздувался до «108»).
# Процесс форм запускается заново на каждый «Запустить», поэтому счётчик стартует
# с нуля сам; между городами одного запуска он накапливается (это и нужно).
_ПРОГРЕСС_ФОРМ = 0


def _отметить_форму(название: str) -> None:
    global _ПРОГРЕСС_ФОРМ
    _ПРОГРЕСС_ФОРМ += 1
    _имя = (название or "").strip() or "(без названия)"
    print(f"▶ Форма {_ПРОГРЕСС_ФОРМ}: {_имя}", flush=True)


def run_test(ОЧИСТИТЬ_EXCEL=True, stop_flag=None, headless=True,
             город="", почта_получателя="", проба_файлов=False, xss_проба=False,
             валидация_проба=False, лимит_проба=False, проверять_цели=False):
    # проверять_цели: проверка форм САМА цели Метрики НЕ проверяет (это делает
    # «Проверка целей»). Флаг включается только когда движок форм зовёт «Проверка
    # целей» (forms_run --check-goals) - тогда ловим цели заказа для её отчёта.
    # headless=True - браузер работает скрыто (окно не показывается); False - видимый.
    # город / почта_получателя - для прогона по поддоменам (городам): метка города
    # и почта, на которую должна прийти заявка (пишутся в одноимённые колонки лога).
    # Всегда читаем актуальный config.py с диска (после «Сохранить» в редакторе иначе остаётся кэш).
    import importlib
    import time as _time

    import config

    _run_t0 = _time.time()
    # Домен-уровневый «выключатель» ретраев: как только соединение с этим
    # доменом обрывается один раз (см. _goto_with_retry ниже) - дальнейшие
    # переходы этого прогона идут БЕЗ полных повторов, чтобы не терять минуты
    # на заведомо недоступном домене (блокировка IP/прокси, а не разовый сбой).
    _домен_нестабилен = {"n": 0, "предупредили": False}
    importlib.reload(config)
    from config import ТЕЛЕФОН, ПОЧТА, ИМЯ, КОММЕНТАРИЙ, СТРАНИЦЫ, СТРАНИЦЫ_ДЛЯ_ПРОВЕРКИ

    # XSS-проба: подменяем САМО значение имени на payload-маркер. Форма заполняет
    # поле имени токеном ИМЯ (карта «поля») → значение берётся ОТСЮДА; раньше payload
    # клался только в имя_теста и в форму не попадал (тест был вхолостую). Почта и
    # телефон остаются реальными - по ним админка опознаёт нашу заявку.
    if xss_проба:
        ИМЯ = _XSS_PAYLOAD

    # Пункт 2.9: почту покупателя можно переопределить из интерфейса (окружение
    # ORDER_BUYER_EMAIL) - чтобы заказ (и заявки) уходили на РЕАЛЬНУЮ почту, куда
    # придёт письмо-подтверждение: тестовый ящик с IMAP (автопроверка) или своя
    # рабочая почта (проверка вручную). Пусто - остаётся ПОЧТА из config.
    _buyer_email = (os.environ.get("ORDER_BUYER_EMAIL") or "").strip()
    if _buyer_email:
        ПОЧТА = _buyer_email
        print(f"✉️ Почта покупателя из интерфейса: {ПОЧТА} "
              f"(на неё уйдёт заказ и придёт письмо-подтверждение).")

    # Переопределение ссылок под конкретный город (СНГ-домены: другой каталог).
    try:
        from config import URL_ПО_ГОРОДУ as _URL_OVERRIDES
    except ImportError:
        _URL_OVERRIDES = {}
    _city_ov = (_URL_OVERRIDES or {}).get((город or "").strip(), {})
    for _t, _u in _city_ov.items():
        if _t in СТРАНИЦЫ:
            СТРАНИЦЫ[_t] = _u
    if _city_ov:
        print(f"🔁 {город}: переопределены ссылки → {', '.join(_city_ov)}")

    try:
        from config import ФОРМЫ_ЧЕРЕЗ_REQUESTS
    except ImportError:
        ФОРМЫ_ЧЕРЕЗ_REQUESTS = False

    # Сопоставление наших названий форм с типом в админке (Уровень 1). Может
    # отсутствовать - тогда сверка идёт по совпадению названия.
    try:
        from config import АДМИН_ТИПЫ as _АДМИН_ТИПЫ
    except ImportError:
        _АДМИН_ТИПЫ = {}
    _АДМИН_ТИПЫ = _АДМИН_ТИПЫ or {}

    # Выбор форм из интерфейса: если задан список ТОЛЬКО_ФОРМЫ (имена сценариев/
    # форм/модалок), гоняем ТОЛЬКО их. Пусто/не задано - гоняем все формы, как раньше.
    try:
        from config import ТОЛЬКО_ФОРМЫ as _ТОЛЬКО_ФОРМЫ
    except ImportError:
        _ТОЛЬКО_ФОРМЫ = None
    _только_формы = {str(x).strip() for x in (_ТОЛЬКО_ФОРМЫ or []) if str(x).strip()}

    def _форма_выбрана(название) -> bool:
        """True, если форму нужно гнать. Если фильтр не задан - гоним всё."""
        if not _только_формы:
            return True
        return str(название or "").strip() in _только_формы

    try:
        from config import ФОРМАТ_ИМЕНИ_ТЕСТА as _FMT_TEST
    except ImportError:
        _FMT_TEST = "{имя}"
    try:
        from config import ФОРМАТ_ИМЕНИ_АВТО_ПО_УМОЛЧАНИЮ as _FMT_AUTO
    except ImportError:
        _FMT_AUTO = "{значение}. {страница}. {дата}"
    ФОРМАТ_ИМЕНИ_ТЕСТА = (str(_FMT_TEST).strip() or "{имя}")
    ФОРМАТ_ИМЕНИ_АВТО_ПО_УМОЛЧАНИЮ = (
        str(_FMT_AUTO).strip() or "{значение}. {страница}. {дата}"
    )

    телефон_отправки = normalize_phone_for_submit(ТЕЛЕФОН)
    # Город прогона (из forms_run, напр. «Бишкек») имеет приоритет; токен «ГОРОД»/«city»
    # в полях форм (Город доставки и т.п.) берёт именно его. Фоллбэк - config.ГОРОД.
    ГОРОД = (город or "").strip() or (getattr(config, "ГОРОД", "") or "")

    print("=" * 60)
    print("ПРОВЕРКА ФОРМ НА САЙТЕ")
    print("=" * 60)

    ДАТА = datetime.now().strftime("%d.%m.%Y")
    ВРЕМЯ = datetime.now().strftime("%H:%M:%S")
    EXCEL_ФАЙЛ = "log_forms.xlsx"

    def имя_теста_из_конфига(
        страница: str,
        значение_авто,
        название_из_конфига=None,
        *,
        название_контекста=None,
    ) -> str:
        """Собирает имя теста из ИМЯ/ФОРМАТ_* (name_format.build_test_name). Плейсхолдер {название}."""
        return build_test_name(
            имя_конфига=ИМЯ,
            название_из_конфига=название_из_конфига,
            страница=страница,
            значение_авто=значение_авто,
            формат_если_имя=ФОРМАТ_ИМЕНИ_ТЕСТА,
            формат_если_авто=ФОРМАТ_ИМЕНИ_АВТО_ПО_УМОЛЧАНИЮ,
            дата=ДАТА,
            время=ВРЕМЯ,
            название_для_плейсхолдеров=название_контекста,
        )

    def определить_страницу(url):
        for key, value in СТРАНИЦЫ.items():
            if url == value:
                return key
        return "Другая"

    def инициализировать_excel():
        init_excel_log(EXCEL_ФАЙЛ, ОЧИСТИТЬ_EXCEL)
        if ОЧИСТИТЬ_EXCEL:
            reset_submitted_forms()
            reset_placed_orders()

    def записать_в_excel(данные):
        # Постоянные колонки (дата/время/телефон/почта) подставляются здесь,
        # поэтому в местах вызова их можно не дублировать.
        row = dict(данные)
        row.setdefault("дата", ДАТА)
        row.setdefault("время", ВРЕМЯ)
        row.setdefault("город", город)
        row.setdefault("почта_получателя", почта_получателя)
        row.setdefault("телефон", телефон_отправки)
        row.setdefault("почта", ПОЧТА)
        # Статус -> короткое слово; причина неудачи -> в «Комментарий».
        # (Колонка «Комментарий» теперь показывает ПОЧЕМУ не сработало, а при
        # успехе остаётся пустой - тестовый текст комментария в отчёт не пишем.)
        _clean, _reason = _status_clean_reason(str(row.get("статус", "")))
        row["статус"] = _clean
        # «комментарий_готовый» - если задан явно, пишем его как есть (этап падения
        # сценария, «формы нет на домене» и т.п.); иначе - авто-причина по статусу.
        row["комментарий"] = (данные.get("комментарий_готовый") or _reason)
        row.pop("комментарий_готовый", None)
        # Уровень 1 (админка): запоминаем реально отправленные формы. Берём только
        # успешные отправки форм/модалок - не цели Метрики (тип «ЦЕЛЬ», тип_селектора
        # «цель») и не проверки корзины/оформления (тип_селектора «сценарий»).
        try:
            if (row.get("статус") == "Успешно"
                    and str(данные.get("тип_селектора", "")) not in ("сценарий", "цель")
                    and not str(данные.get("тип", "")).upper().startswith("ЦЕЛЬ")):
                _назв = данные.get("название", "")
                record_submitted_form({
                    "город": город,
                    "название": _назв,
                    "админ_тип": _АДМИН_ТИПЫ.get(_назв, ""),
                    "имя": row.get("имя", ""),
                    "почта": ПОЧТА,
                    "телефон": телефон_отправки,
                    "страница": данные.get("страница", ""),
                    "url": данные.get("url", ""),
                    "ts": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
                })
        except Exception:
            pass
        try:
            append_log_row(EXCEL_ФАЙЛ, row)
        except Exception as e:
            print(f"   ⚠️ Ошибка Excel: {e}")

    def _нет_в_текущем_городе(cfg_obj):
        """True, если форма/сценарий помечены «нет_в_городах» и текущий город в списке."""
        lst = (cfg_obj or {}).get("нет_в_городах")
        return bool(lst) and (город or "").strip() in lst

    def _лог_форма_отсутствует(тип_страницы, url, cfg_obj, название):
        """Форма помечена «нет_в_городах» - её на этом домене нет. Такие формы в
        отчёт НЕ выводим (это не проверка и не ошибка, а просто отсутствие формы:
        напр. подписка Хабаровска на других городах). Только отметка в лог-файле
        прогона, чтобы было видно, что форму осознанно пропустили."""
        коммент = (cfg_obj or {}).get("нет_коммент") or "Данной формы нет на сайте в этом домене"
        print(f"   ⏭️ «{название}»: {коммент} - пропускаю (в отчёт не пишу)")

    def отправить_через_requests(url, форма_config, название, probe=False):
        # probe=True (авто-режим): если форму в статическом HTML не видно (её
        # рисует/шлёт JS) - НЕ пишем «не найдена», а возвращаем None, чтобы
        # вызывающий откатился на браузер. probe=False - обычный режим (пишем).
        страница = определить_страницу(url)

        имя_теста = имя_теста_из_конфига(
            страница,
            _значение_формы_для_имени(форма_config),
            название,
        )

        session = requests.Session()

        try:
            hdr = _browser_headers(url)
            response = session.get(url, timeout=10, headers=hdr)
            soup = BeautifulSoup(response.text, "html.parser")

            form = None
            try:
                idx_bs = int(форма_config.get("индекс", 0))
            except (TypeError, ValueError):
                idx_bs = 0
            if "text" in форма_config and str(форма_config.get("text", "")).strip():
                txt = str(форма_config["text"]).strip()
                candidates = [f for f in soup.find_all("form") if txt in f.get_text()]
                if idx_bs < len(candidates):
                    form = candidates[idx_bs]
            else:
                sel_rq = _playwright_form_css_selector(форма_config)
                if sel_rq:
                    for c in _expand_form_selector_fallbacks(sel_rq):
                        candidates = soup.select(c)
                        if idx_bs < len(candidates):
                            form = candidates[idx_bs]
                            break

            if (
                form is not None
                and getattr(form, "name", None) != "form"
            ):
                _pf = form.find_parent("form")
                if _pf is not None:
                    form = _pf

            if not form:
                if probe:
                    return None      # по коду форму не видно - пусть решает браузер
                записать_в_excel(
                    {
                        "тип": "REQUESTS",
                        "страница": страница,
                        "url": url,
                        "тип_селектора": format_form_selector_type(форма_config),
                        "ид": format_form_config_for_log(форма_config),
                        "название": название,
                        "имя": имя_теста,
                        "комментарий": КОММЕНТАРИЙ if КОММЕНТАРИЙ else имя_теста,
                        "статус": "ФОРМА НЕ НАЙДЕНА",
                        "код": "",
                    }
                )
                return False

            data = {}
            for hidden in form.find_all("input", {"type": "hidden"}):
                if hidden.get("name"):
                    data[hidden.get("name")] = hidden.get("value", "")

            _ffmap_rq = _form_field_map_from_config(форма_config)
            _ctx_rq = {
                "имя_теста": имя_теста,
                "телефон": телефон_отправки,
                "почта": ПОЧТА,
                "имя": ИМЯ,
                "комментарий": КОММЕНТАРИЙ,
                "город": ГОРОД,
            }
            if _ffmap_rq:
                for name_attr, tok in _ffmap_rq.items():
                    val = _resolve_form_field_token(tok, **_ctx_rq)
                    if not val:
                        continue
                    el = form.find(attrs={"name": name_attr})
                    if el is not None:
                        data[name_attr] = val
            else:
                for field in form.find_all(["input", "textarea"]):
                    name = field.get("name")
                    if not name or field.get("type") in ["submit", "button", "reset"]:
                        continue
                    if field.get("type") == "hidden":
                        continue

                    name_lower = (name or "").lower()
                    if (
                        name in ("name", "fio")
                        or name_lower in ("fio", "username", "client_name")
                        or (
                            "name" in name_lower
                            and "phone" not in name_lower
                            and "email" not in name_lower
                        )
                    ):
                        data[name] = имя_теста
                    elif "phone" in name_lower or "tel" in name_lower:
                        data[name] = телефон_отправки
                    elif "email" in name_lower:
                        data[name] = ПОЧТА
                    elif field.get("type") == "checkbox":
                        data[name] = field.get("value", "on")
                    elif (
                        field.name == "textarea"
                        or "message" in name_lower
                        or "comment" in name_lower
                    ):
                        data[name] = КОММЕНТАРИЙ if КОММЕНТАРИЙ else имя_теста
                    else:
                        data[name] = имя_теста

            for cb in form.find_all("input", {"type": "checkbox"}):
                cb_name = cb.get("name")
                if cb_name:
                    continue
                cid = cb.get("id")
                if not cid:
                    continue
                if cb.has_attr("required") or cb.get("required") is not None:
                    data[cid] = cb.get("value") or "Y"

            sec = extract_form_security_from_html(response.text)
            for k, v in sec.items():
                if k == "hash" and (not data.get("hash")):
                    data["hash"] = v
                if k == "sessid":
                    data["sessid"] = data.get("sessid") or v

            action = form.get("action") or ""
            submit_url = urljoin(url, action) if action else url
            raw_method = (form.get("method") or "").strip().lower()
            if raw_method:
                method = raw_method
            elif "/ajax/" in action or "form.php" in action:
                method = "post"
            else:
                method = "post"

            post_headers = _ajax_post_headers(url, submit_url)

            if ("/ajax/" in submit_url or "form.php" in submit_url) and not (
                data.get("hash") or ""
            ).strip():
                print(
                    "      ⚠️ Поле hash пустое (на сайте его часто подставляет JS). "
                    "Если заявка не доходит - смотрите вкладку Network при ручной отправке "
                    "или используйте сценарий Playwright для этой формы."
                )

            if method == "post":
                result = session.post(
                    submit_url, data=data, timeout=15, headers=post_headers
                )
            else:
                result = session.get(
                    submit_url, params=data, timeout=15, headers=post_headers
                )

            статус = _interpret_response_status(result)
            if статус.startswith("УСПЕШНО") and result.status_code == 200:
                snippet = re.sub(r"\s+", " ", result.text[:400]).strip()
                print(f"      Ответ (фрагмент): {snippet[:200]}…")

            # Структурные проверки формы ПО КОДУ (без браузера): CSRF, согласие
            # 2.13, выпадающие списки, типы файлов, подсказки полей. Раньше эти
            # колонки у code-проверенных форм были пустыми (в матрице - прочерки);
            # теперь заполняются реальными вердиктами из разметки. Ошибка проверок
            # не должна ронять запись самой отправки - всё под try.
            _структ = {}
            try:
                _куки = csrf_куки_инфо(_куки_из_ответа_requests(response))
                _структ = _html_структурные_проверки(form, response.text, _куки)
            except Exception as _est:  # noqa: BLE001
                print(f"      ⚠️ Структурные проверки по коду не удались: {_est}")

            записать_в_excel(
                {
                    "тип": "REQUESTS",
                    "страница": страница,
                    "url": url,
                    "тип_селектора": format_form_selector_type(форма_config),
                    "ид": format_form_config_for_log(форма_config),
                    "название": название,
                    "имя": имя_теста,
                    "комментарий": КОММЕНТАРИЙ if КОММЕНТАРИЙ else имя_теста,
                    "статус": статус,
                    "код": result.status_code,
                    **_структ,
                }
            )
            print(f"   ✅ {название} - {статус}")
            return True

        except Exception as e:
            записать_в_excel(
                {
                    "тип": "REQUESTS",
                    "страница": страница,
                    "url": url,
                    "тип_селектора": format_form_selector_type(форма_config),
                    "ид": format_form_config_for_log(форма_config),
                    "название": название,
                    "имя": имя_теста,
                    "комментарий": КОММЕНТАРИЙ if КОММЕНТАРИЙ else имя_теста,
                    "статус": f"ОШИБКА ({e})",
                    "код": str(e),
                }
            )
            return False

    def _проверить_форму_авто(url, форма_config, название):
        """Авто-режим (по умолчанию): где форму реально видно по коду и не нужны
        браузерные пробы - проверяем запросом (быстро, без браузера); иначе
        (форму рисует/шлёт JS, либо включены пробы XSS/валидации/лимита) -
        откатываемся на браузер (Playwright, полная проверка)."""
        _нужен_браузер = xss_проба or валидация_проба or лимит_проба
        if not _нужен_браузер:
            rq = отправить_через_requests(url, форма_config, название, probe=True)
            if rq is not None:
                return                     # requests справился - строка записана
            print(f"   ↪ «{название}»: по коду форму не видно - проверяю через браузер")
        отправить_форму_через_playwright(url, форма_config, название)

    def _goto_with_retry(page, url, *, attempts=3, wait_ms=2500):
        """Переход с повтором при обрыве соединения (антибот/лимит сайта).
        Меньше попыток/паузы + умеренный таймаут - прогон быстрее; сценарий/форма
        ещё раз повторятся на верхнем уровне, так что попыток суммарно хватает.

        Если домен уже показал проблему с соединением (не меньше 1 обрыва за этот
        прогон) - дальше НЕ тратим время на полные повторы: одна попытка без
        ожидания. Обычно это значит, что домен целиком недоступен (блокирует
        IP/прокси) - полные ретраи на каждой следующей форме только теряют время."""
        if _домен_нестабилен["n"] >= 1:
            attempts = 1
            if not _домен_нестабилен["предупредили"]:
                _домен_нестабилен["предупредили"] = True
                print("   \u26a1 Домен нестабилен (обрыв соединения) - дальше проверяем "
                      "без повторных попыток, чтобы не терять время.")
        last = None
        for i in range(attempts):
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=30000)
                return
            except Exception as e:  # noqa: BLE001
                last = e
                m = str(e)
                transient = (
                    "ERR_CONNECTION" in m or "ERR_NETWORK" in m
                    or "ERR_ABORTED" in m or "ERR_TIMED_OUT" in m
                    or "Timeout" in m or "net::" in m
                )
                if i < attempts - 1 and transient:
                    print(f"   \u21bb Соединение сброшено сайтом, повтор {i + 2}/{attempts} через {wait_ms} мс\u2026")
                    page.wait_for_timeout(wait_ms)
                    continue
                if transient:
                    _домен_нестабилен["n"] += 1
                raise
        if last:
            raise last

    def _ensure_scenario_page_loaded(page, base_url: str) -> None:
        u = (page.url or "").strip()
        if not u or u == "about:blank" or u.startswith("about:"):
            _goto_with_retry(page, base_url)
            page.wait_for_timeout(2000)
        # Снятие оверлеев (подтверждение города и т.п.) - через ключ «подготовка»
        # в конфиге страницы, см. _run_page_prep (раньше здесь был хардкод mepen.ru).

    def _resolve_scenario_url(page, base_url: str, href) -> str:
        if not href:
            return base_url
        h = str(href).strip()
        if h.startswith("http://") or h.startswith("https://"):
            return h
        return urljoin(page.url or base_url, h)

    _META_STEP_KEYS = frozenset(
        {"действие", "action", "включено", "url", "href", "мс", "ms"}
    )

    def _scenario_placeholder_title(_step: dict, название_сценария: str) -> str:
        """Строка для {название} в сценарии: всегда название текущего сценария (как в редакторе)."""
        return (название_сценария or "").strip()

    def _нормализовать_действие_шага(step: dict) -> str:
        return normalize_step_action(step.get("действие") or step.get("action") or "")

    def _run_page_prep(page, страница_type: str) -> None:
        """
        Best-effort «подготовка» страницы (снять оверлей, навести и т.п.).

        Раньше под это были зашиты site-specific костыли (оверлей города на
        mepen.ru, наведение на карточку на «Листинге»). Теперь шаги берутся из
        конфига страницы (ключ «подготовка»). Любая ошибка гасится - прогон не
        падает из-за подготовки.
        """
        pg = next(
            (
                p
                for p in СТРАНИЦЫ_ДЛЯ_ПРОВЕРКИ
                if isinstance(p, dict) and p.get("тип") == страница_type
            ),
            None,
        )
        steps = prep_steps_from_page(pg) if pg else []
        if not steps:
            return
        for s in steps:
            act = s.get("действие")
            try:
                if act == "пауза":
                    try:
                        ms = int(s.get("мс") or s.get("ms") or 500)
                    except (TypeError, ValueError):
                        ms = 500
                    page.wait_for_timeout(ms)

                elif act == "клик":
                    txt = s.get("текст") or s.get("text")
                    if txt is not None and str(txt).strip():
                        loc = page.get_by_text(str(txt).strip(), exact=True).first
                        if loc.is_visible(timeout=1500):
                            loc.click()
                            page.wait_for_timeout(400)
                            print(f"      🧹 Подготовка: клик по тексту {str(txt).strip()!r}")
                    else:
                        css = s.get("css") or s.get("selector")
                        if css:
                            sel = _normalize_scenario_click_css_selector(str(css).strip())
                            loc = page.locator(sel).first
                            if loc.count() and loc.is_visible(timeout=1500):
                                loc.click()
                                page.wait_for_timeout(400)
                                print(f"      🧹 Подготовка: клик {sel!r}")

                elif act == "наведение":
                    css = s.get("css") or s.get("selector")
                    if css:
                        sel = _normalize_scenario_click_css_selector(str(css).strip())
                        loc = page.locator(sel).first
                        if loc.count():
                            loc.scroll_into_view_if_needed()
                            loc.hover()
                            page.wait_for_timeout(400)
                            print(f"      🧹 Подготовка: наведение {sel!r}")
                # неизвестное действие в «подготовке» - тихо пропускаем
            except Exception as e:
                print(f"      ⚠️ Подготовка ({act}): {e}")

    def _replay_open_step(page, step, base_url):
        """Тихо проигрывает ОДИН шаг ОТКРЫТИЯ формы при переоткрытии (только
        пауза/переход/клик/наведение - то, что открывает модалку/попап). Шаги
        «форма»/«модалка» не трогаем: саму форму мы находим и заполняем отдельно.
        Всё guarded - на ошибке просто идём дальше."""
        try:
            act = _нормализовать_действие_шага(step)
            if act == "пауза":
                try:
                    ms = int(step.get("мс") or step.get("ms") or 500)
                except (TypeError, ValueError):
                    ms = 500
                page.wait_for_timeout(ms)
            elif act == "перейти":
                href = step.get("url") or step.get("href")
                if href:
                    _goto_with_retry(page, _resolve_scenario_url(page, base_url, href))
                    page.wait_for_timeout(1200)
            elif act in ("клик", "наведение"):
                css = step.get("css") or step.get("selector")
                if css:
                    css_norm = _normalize_scenario_click_css_selector(str(css).strip())
                    loc = page.locator(css_norm).first
                    if act == "клик":
                        loc.click(timeout=8000)
                    else:
                        loc.scroll_into_view_if_needed()
                        loc.hover(timeout=8000)
                    page.wait_for_timeout(400)
        except Exception:  # noqa: BLE001
            pass

    def _form_fill_submit_on_page(
        page,
        url_for_excel: str,
        страница: str,
        форма_config: dict,
        название: str,
        *,
        initial_url=None,
        название_контекста=None,
        цели_seen=None,
        безопасная_отправка=False,
        переоткрыть_fn=None,
    ):
        """
        Заполнение и отправка формы на уже открытой странице.
        безопасная_отправка=True (формы «Оформление…» заказа): двойную отправку
        проверяем БЕЗ реального второго клика (смотрим блокировку кнопки), чтобы
        не создать второй заказ. Иначе - реальный двойной клик с подсчётом заявок.
        initial_url: если задан - сначала переход; иначе считаем, что нужная страница уже загружена.
        """
        nctx = название_контекста if название_контекста is not None else название
        имя_теста = имя_теста_из_конфига(
            страница,
            _значение_формы_для_имени(форма_config),
            название,
            название_контекста=nctx,
        )
        # XSS-проба (под галочкой): в поле ИМЕНИ шлём безвредный payload-маркер
        # вместо тест-имени. Почта/телефон остаются реальными - по ним админка
        # опознаёт нашу заявку (_это_наша_заявка по ИЛИ). После отправки смотрим,
        # исполнился ли payload на ответе (Фаза 1).
        if xss_проба:
            имя_теста = _XSS_PAYLOAD
        use_text = "text" in форма_config and str(форма_config.get("text", "")).strip()
        sel = _playwright_form_css_selector(форма_config)
        if not use_text and not sel:
            записать_в_excel(
                {
                    "тип": "PLAYWRIGHT-FORM",
                    "страница": страница,
                    "url": url_for_excel,
                    "тип_селектора": format_form_selector_type(форма_config),
                    "ид": format_form_config_for_log(форма_config),
                    "название": название,
                    "имя": имя_теста,
                    "комментарий": КОММЕНТАРИЙ if КОММЕНТАРИЙ else имя_теста,
                    "статус": "НЕТ СЕЛЕКТОРА ФОРМЫ (id/class/data-source/css/name/text)",
                    "код": "",
                }
            )
            return False

        if initial_url:
            _goto_with_retry(page, initial_url)
            page.wait_for_timeout(2000)
            _run_page_prep(page, страница)
        else:
            page.wait_for_timeout(400)

        log_url = page.url or url_for_excel

        try:
            try:
                idx = int(форма_config.get("индекс", 0))
            except (TypeError, ValueError):
                idx = 0

            if use_text:
                text_val = str(форма_config["text"]).strip()
                loc = page.locator("form").filter(has_text=text_val)
                n_match = loc.count()
                sel_desc = f"form:has-text({text_val!r})"
            else:
                n_match = 0
                loc = page.locator(sel)
                sel_desc = sel
                tried: list[str] = []
                for cand in _expand_form_selector_fallbacks(sel):
                    tried.append(cand)
                    try:
                        loc_try = page.locator(cand)
                        n_try = loc_try.count()
                    except Exception:
                        n_try = 0
                    if n_try > 0:
                        loc = loc_try
                        n_match = n_try
                        sel_desc = cand
                        if cand != sel:
                            print(
                                f"      Подобран селектор: {sel!r} → {cand!r} (найдено {n_try})"
                            )
                        break

            print(
                f"      Селектор: {sel_desc!r} - найдено форм: {n_match}, берём №{idx + 1}"
            )
            if n_match == 0:
                tried_txt = ", ".join(repr(t) for t in _expand_form_selector_fallbacks(sel))
                raise RuntimeError(
                    "Нет элементов по селектору. Перепробованы варианты: "
                    f"{tried_txt}. "
                    "Уточните ключ «css» в конфиге (например div.row, .bx-soa-customer), "
                    "проверьте классы в DevTools (для второй формы - «an-row calculation-order», не только "
                    "«calculation-order»). Если блок подгружается по AJAX - добавьте шаг «перейти»/«пауза» перед формой."
                )
            if idx >= n_match:
                raise RuntimeError(
                    f"индекс={idx} вне диапазона (найдено форм: {n_match}). Уменьшите «индекс» или уточните селектор."
                )

            form = loc.nth(idx)
            # Несколько форм с одним классом (напр. find-form встречается 3 раза):
            # скрытые дубли можно «отправить», но цель на них не срабатывает. Если
            # индекс явно не задан - берём первую ВИДИМУЮ форму.
            if idx == 0 and n_match > 1 and "индекс" not in форма_config:
                for _k in range(n_match):
                    try:
                        if loc.nth(_k).is_visible():
                            form = loc.nth(_k)
                            break
                    except Exception:
                        continue
            form.wait_for(state="visible", timeout=8000)
            try:
                _root_tag = form.evaluate("el => el.tagName.toLowerCase()")
            except Exception:
                _root_tag = ""
            # Селектор по name может указывать на одно поле - поднимаемся к <form>, если есть.
            if _root_tag in ("input", "textarea", "select"):
                anc = form.locator("xpath=ancestor::form[1]")
                if anc.count() > 0:
                    form = anc.first
            try:
                form.evaluate(
                    "(el) => el.scrollIntoView({block: 'center', behavior: 'instant'})"
                )
            except Exception:
                form.scroll_into_view_if_needed()
            page.wait_for_timeout(400)

            # Опционально: расширить область до родительского .row или form (см. ключ «расширить_контейнер»)
            form = _apply_container_expand(form, форма_config)

            _ffmap = _form_field_map_from_config(форма_config)
            _ctx_ff = {
                "имя_теста": имя_теста,
                "телефон": телефон_отправки,
                "почта": ПОЧТА,
                "имя": ИМЯ,
                "комментарий": КОММЕНТАРИЙ,
                "город": ГОРОД,
            }
            if _ffmap:
                print(
                    f"      Заполнение по карте «поля»: {len(_ffmap)} полей "
                    f"(дозаполнение эвристикой: "
                    f"{'да' if форма_config.get('дозаполнить_по_признакам') else 'нет'})"
                )
                for _aname, _tok in _ffmap.items():
                    # «Ссылка на товар» (product-link) - СКРЫТОЕ поле. Заполняем его
                    # напрямую через JS строго по name, иначе обычный fill из-за
                    # одинаковых id у fio/e-mail/phone «протекал» URL-ом в телефон.
                    if _aname == "product-link":
                        _val = page.url or base_url
                        if str(_tok).strip() not in ("URL_ТОВАРА", "URL", "page_url", ""):
                            _val = _resolve_form_field_token(_tok, **_ctx_ff)
                        try:
                            form.evaluate(
                                "(f, v) => { const el = f.querySelector('[name=product-link]');"
                                " if (el) { el.value = v; el.dispatchEvent(new Event('input', {bubbles:true})); } }",
                                _val,
                            )
                        except Exception as _e:  # noqa: BLE001
                            print(f"      ⚠️ product-link: {_e}")
                        continue
                    _val = _resolve_form_field_token(_tok, **_ctx_ff)
                    if _val:
                        _pw_fill_named_field(form, _aname, _val)

            _do_heur = True
            if _ffmap:
                _do_heur = bool(форма_config.get("дозаполнить_по_признакам", False))

            if _do_heur:
                name_filled = False
                for nm in ("fio", "name", "NAME", "username", "client_name"):
                    loc = form.locator(f"input[name='{nm}']")
                    if loc.count() > 0:
                        loc.first.fill(имя_теста, force=True)
                        name_filled = True
                        break

                if not name_filled and имя_теста:
                    for ac in (
                        "input[autocomplete='name']",
                        "input[autocomplete='given-name']",
                    ):
                        loc = form.locator(ac)
                        if loc.count() > 0:
                            loc.first.fill(имя_теста, force=True)
                            name_filled = True
                            break

                if not name_filled and имя_теста:
                    for sub in (
                        "Контактное лицо",
                        "ФИО",
                        "Имя",
                        "имя",
                        "Ваше имя",
                        "Name",
                    ):
                        try:
                            el = form.locator(f'input[placeholder*="{sub}"]')
                            if el.count() > 0:
                                el.first.fill(имя_теста, force=True)
                                name_filled = True
                                break
                        except Exception:
                            pass

                phone_filled = False
                for psel in (
                    "input[autocomplete='tel']",
                    "input[name='phone']",
                    "input[name='PHONE']",
                    "input[name='telephone']",
                    "input[name='tel']",
                    "input[name='mobile']",
                    "input[name='PHONE_MOBILE']",
                    "input[name='form_phone']",
                    "input[name='TEL']",
                    "input[type='tel']",
                    "input[name*='phone']",
                    "input[name*='PHONE']",
                    "input[name*='tel']",
                ):
                    loc = form.locator(psel)
                    if loc.count() > 0:
                        loc.first.fill(телефон_отправки, force=True)
                        phone_filled = True
                        break
                if not phone_filled and телефон_отправки:
                    for sub in ("Телефон", "телефон", "Phone", "Мобильный", "Tel"):
                        try:
                            el = form.locator(f'input[placeholder*="{sub}"]')
                            if el.count() > 0:
                                el.first.fill(телефон_отправки, force=True)
                                phone_filled = True
                                break
                        except Exception:
                            pass
                if not phone_filled and телефон_отправки:
                    print(
                        "      ⚠️ Поле телефона не найдено по типовым name/type - проверьте разметку формы."
                    )

                if ПОЧТА:
                    for loc_sel in (
                        "input[autocomplete='email']",
                        "input[type='email']",
                        "input[name='email']",
                        "input[name='mail']",
                    ):
                        loc = form.locator(loc_sel)
                        if loc.count() > 0:
                            loc.first.fill(ПОЧТА, force=True)
                            break

                if КОММЕНТАРИЙ:
                    for ta_sel in (
                        "textarea[name='ORDER_DESCRIPTION']",
                        "textarea#orderDescription",
                        "textarea[name='ORDER_COMMENT']",
                        "textarea",
                    ):
                        ta = form.locator(ta_sel)
                        if ta.count() > 0:
                            ta.first.fill(КОММЕНТАРИЙ, force=True)
                            break

            # Пункт 2.13: проверяем согласие и политику ДО проставления галочек
            # (иначе не увидим «предустановлены ли изначально»). Отправку не делаем.
            try:
                _c213 = проверка_согласия_2_13(form, page)
                _ok213 = (_c213["чекбоксов"] >= 2 and not _c213["предустановлены"]
                          and _c213["ссылка"] and _c213["валидация"])
                _det213 = (
                    f"чек-боксов согласия: {_c213['чекбоксов']} "
                    f"(нужно ≥2); не предустановлены: "
                    f"{'да' if not _c213['предустановлены'] else 'НЕТ - стоят по умолчанию'}; "
                    f"ссылка на политику: {'да' if _c213['ссылка'] else 'нет'}; "
                    f"без согласия не отправить: {'да' if _c213['валидация'] else 'нет'}")
                записать_в_excel({
                    "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                    "тип_селектора": "согласие", "ид": название,
                    "название": f"Согласие и политика (2.13): {название}",
                    "имя": имя_теста, "статус": "OK" if _ok213 else "Проверить",
                    # Пункт 2.13 - по отдельным колонкам (не одной строкой в «Комментарий»).
                    "согласие_чекбоксы": f"{_c213['чекбоксов']} (нужно ≥2)",
                    "согласие_предустановка": ('да' if not _c213['предустановлены']
                                               else 'НЕТ - стоят по умолчанию'),
                    "согласие_ссылка": 'да' if _c213['ссылка'] else 'нет',
                    "согласие_обязательно": 'да' if _c213['валидация'] else 'нет',
                    "код": "consent213",
                })
                print(f"   📝 Согласие (2.13) «{название}»: "
                      f"{'OK' if _ok213 else 'проверить'} - {_det213}")
            except Exception as _e213:  # noqa: BLE001
                print(f"   ⚠️ Проверка согласия 2.13 не удалась: {_e213}")

            # Аудит полей формы: маска телефона (баг, если нет ограничения) +
            # типы файлов у загрузчика (только вывод). Отправку не трогает.
            try:
                _pf = проверка_полей_форм(form, page)
                # Пункт 2.14: баг, если ЯВНО не работает маска телефона, почта не
                # проверяет формат или пустую форму можно отправить (нет уведомлений
                # о заполнении). None - поле не найдено, не считаем багом.
                _phone_bug = _pf["телефон_ограничен"] is False
                _mail_bug = _pf["почта_ок"] is False
                _date_bug = _pf["дата_ок"] is False
                _req_bug = _pf["обязательность_ок"] is False
                _len_bug = _pf["длина_ок"] is False
                _есть_баг = _phone_bug or _mail_bug or _date_bug or _req_bug or _len_bug
                # Телефон - в «Комментарий»; типы файлов - в отдельную
                # колонку «Типы файлов формы» (пусто, если поля загрузки нет).
                if not _pf["файл_есть"]:
                    _файлы_кол = ""
                elif _pf["файл_любые"]:
                    _файлы_кол = "⚠ ЛЮБЫЕ типы (accept не задан)"
                else:
                    _файлы_кол = ", ".join(_pf["файл_типы"])
                # Дату упоминаем в комментарии, только если поле даты реально
                # есть (дата_ок не None) - иначе на каждой форме шумит «не найдено».
                _дата_txt = (f"дата (формат): {_pf['дата_детали']}; "
                             if _pf["дата_ок"] is not None else "")
                _ком214 = (f"телефон (маска): {_pf['телефон_детали']}; "
                           f"почта (валидация): {_pf['почта_детали']}; "
                           f"{_дата_txt}"
                           f"обязательность/уведомления: {_pf['обязательность_детали']}; "
                           f"ограничение длины: {_pf['длина_детали']}")
                записать_в_excel({
                    "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                    "тип_селектора": "поля", "ид": название,
                    "название": f"Поля формы 2.14 (маска тел./почта/обязательность/длина): {название}",
                    "имя": имя_теста,
                    "статус": "Проверить" if _есть_баг else "OK",
                    "типы_файлов": _файлы_кол,
                    "комментарий_готовый": _ком214,
                    "код": "fields_audit",
                })
                print(f"   🧪 Поля 2.14 «{название}»: {_ком214}"
                      + (f"; файлы: {_файлы_кол}" if _файлы_кол else ""))
            except Exception as _epf:  # noqa: BLE001
                print(f"   ⚠️ Аудит полей формы не удался: {_epf}")

            # Пункт «Все элементы формы присутствуют»: перепись состава формы
            # (чистое чтение DOM, без ввода/отправки) - отдельной строкой отчёта.
            try:
                _cf = состав_формы(form)
                записать_в_excel({
                    "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                    "тип_селектора": "поля", "ид": название,
                    "название": f"Состав формы (поля/кнопки/чекбоксы/радио/списки): {название}",
                    "имя": имя_теста,
                    "статус": "Проверить" if _cf.get("ок") is False else "OK",
                    "комментарий_готовый": _cf["детали"],
                    "код": "form_census",
                })
                print(f"   🧩 Состав формы «{название}»: {_cf['детали']}")
            except Exception as _ecf:  # noqa: BLE001
                print(f"   ⚠️ Перепись состава формы не удалась: {_ecf}")

            # Пункт «Форма стилизована по макету»: консистентность полей + факт
            # стилизации (не браузерный дефолт). Чистое чтение стилей, без ввода.
            try:
                _st = стиль_формы(form)
                записать_в_excel({
                    "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                    "тип_селектора": "поля", "ид": название,
                    "название": f"Стилизация формы (шрифты/цвета/отступы): {название}",
                    "имя": имя_теста,
                    "статус": "Проверить" if _st.get("ок") is False else "OK",
                    "стилизация": _st["состояние"],
                    "комментарий_готовый": _st["детали"],
                    "код": "form_style",
                })
                print(f"   🎨 Стилизация «{название}»: {_st['состояние']} — {_st['детали']}")
            except Exception as _est:  # noqa: BLE001
                print(f"   ⚠️ Проверка стилизации формы не удалась: {_est}")

            # Выпадающие списки: колонка «не найдено/корректно/ошибка»; коммент -
            # только при ошибке (пустой список). Значение select не меняем.
            try:
                _dd = проверка_списков(form)
                записать_в_excel({
                    "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                    "тип_селектора": "поля", "ид": название,
                    "название": f"Выпадающие списки: {название}",
                    "имя": имя_теста,
                    "статус": "Проверить" if _dd["состояние"] == "ошибка" else "OK",
                    "выпадающие_списки": _dd["состояние"],
                    "комментарий_готовый": _dd["коммент"],
                    "код": "dropdowns",
                })
                print(f"   ⬇️ Выпадающие списки «{название}»: {_dd['состояние']}"
                      + (f" — {_dd['коммент']}" if _dd["коммент"] else ""))
            except Exception as _edd:  # noqa: BLE001
                print(f"   ⚠️ Проверка выпадающих списков не удалась: {_edd}")

            # Чекбоксы/радио: реально переключаются? Колонка «не найдено/
            # корректно/ошибка», коммент - только при ошибке. Настоящие клики,
            # но состояние возвращаем как было (согласие/оплату не сбиваем).
            try:
                _cb = проверка_чекбоксов(form)
                записать_в_excel({
                    "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                    "тип_селектора": "поля", "ид": название,
                    "название": f"Чекбоксы/радио (переключаются): {название}",
                    "имя": имя_теста,
                    "статус": "Проверить" if _cb["состояние"] == "ошибка" else "OK",
                    "чекбоксы_радио": _cb["состояние"],
                    "комментарий_готовый": _cb["коммент"],
                    "код": "checkboxes",
                })
                print(f"   ☑️ Чекбоксы/радио «{название}»: {_cb['состояние']}"
                      + (f" — {_cb['коммент']}" if _cb["коммент"] else ""))
            except Exception as _ecb:  # noqa: BLE001
                print(f"   ⚠️ Проверка чекбоксов/радио не удалась: {_ecb}")

            # Автозаполнение личных полей: колонка «корректно/отключено/искажается»,
            # коммент - только при проблеме. Значения полей возвращаются как были.
            try:
                _af = проверка_автозаполнения(form)
                записать_в_excel({
                    "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                    "тип_селектора": "поля", "ид": название,
                    "название": f"Автозаполнение полей: {название}",
                    "имя": имя_теста,
                    "статус": "Проверить" if _af["состояние"] in ("отключено", "искажается") else "OK",
                    "автозаполнение": _af["состояние"],
                    "комментарий_готовый": _af["коммент"] or None,
                    "код": "autofill",
                })
                print(f"   🅰️ Автозаполнение «{название}»: {_af['состояние']}"
                      + (f" — {_af['коммент']}" if _af["коммент"] else ""))
            except Exception as _eaf:  # noqa: BLE001
                print(f"   ⚠️ Проверка автозаполнения не удалась: {_eaf}")

            # Подсказки полей (placeholder/help): соответствуют ли назначению.
            # Флажим только явные противоречия; «нет подсказок» - мягкая заметка.
            try:
                _hp = проверка_подсказок(form)
                записать_в_excel({
                    "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                    "тип_селектора": "поля", "ид": название,
                    "название": f"Подсказки полей (placeholder/help): {название}",
                    "имя": имя_теста,
                    "статус": "Проверить" if _hp["состояние"] == "несоответствие" else "OK",
                    "подсказки": _hp["состояние"],
                    "комментарий_готовый": _hp["коммент"] or None,
                    "код": "placeholders",
                })
                print(f"   💬 Подсказки полей «{название}»: {_hp['состояние']}"
                      + (f" — {_hp['коммент']}" if _hp["коммент"] else ""))
            except Exception as _ehp:  # noqa: BLE001
                print(f"   ⚠️ Проверка подсказок полей не удалась: {_ehp}")

            # Кнопка отправки активна только после заполнения обязательных полей.
            # Чистит поля и ГАРАНТИРОВАННО возвращает (finally) - отправку не ломает.
            try:
                _rb = проверка_кнопки_обязательные(
                    form, page, str(форма_config.get("кнопка_css") or ""))
                записать_в_excel({
                    "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                    "тип_селектора": "поля", "ид": название,
                    "название": f"Кнопка активна по заполнению обязательных: {название}",
                    "имя": имя_теста,
                    "статус": "Проверить" if _rb["состояние"] == "не разблокируется" else "OK",
                    "кнопка_обязательные": _rb["состояние"],
                    "комментарий_готовый": _rb["коммент"] or None,
                    "код": "required_button",
                })
                print(f"   🔘 Кнопка по заполнению «{название}»: {_rb['состояние']}"
                      + (f" — {_rb['коммент']}" if _rb["коммент"] else ""))
            except Exception as _erb:  # noqa: BLE001
                print(f"   ⚠️ Проверка кнопки по заполнению не удалась: {_erb}")

            # Файл-проба (по галочке): грузим безвредный файл с опасным
            # расширением и отправляем - пройдёт ли серверную фильтрацию.
            # Реально ОТПРАВЛЯЕТ форму, поэтому после неё штатную отправку
            # НЕ делаем (форма/кнопка уже исчезли - иначе scroll_into_view
            # к пропавшей кнопке висит 30с). Только когда проба включена И у
            # формы есть поле загрузки.
            if проба_файлов and _pf.get("файл_есть"):
                try:
                    _fp = проба_загрузки_файла(form, page)
                    # Вердикт по правилу «только PDF/DOC/DOCX ≤20 КБ» (охватывает
                    # и опасные типы, и посторонние форматы, и лимит размера).
                    _fp_кол, _fp_дет = фильтр_файлов_вердикт(_fp)
                    _fp_bug = _fp_кол.startswith("✗")
                    записать_в_excel({
                        "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                        "тип_селектора": "сценарий", "ид": название,
                        "название": f"Проба загрузки файла (безопасность): {название}",
                        "имя": имя_теста,
                        "статус": "Ошибка" if _fp_bug else "OK",
                        "типы_файлов": _fp_кол,
                        "комментарий_готовый": _fp_дет,
                        "код": "file_probe",
                    })
                    print(f"   🛡 Файл-проба «{название}»: {_fp_дет}")
                except Exception as _efp:  # noqa: BLE001
                    print(f"   ⚠️ Файл-проба не удалась: {_efp}")
                # Форму проба уже отправила - штатную отправку пропускаем
                # (кнопки/формы на странице больше нет).
                записать_в_excel({
                    "тип": "PLAYWRIGHT-FORM", "страница": страница, "url": log_url,
                    "тип_селектора": format_form_selector_type(форма_config),
                    "ид": format_form_config_for_log(форма_config),
                    "название": название, "имя": имя_теста,
                    "статус": "ЗАПОЛНЕНО (отправлено файл-пробой)",
                    "комментарий_готовый": "штатная отправка пропущена - форму "
                    "отправила файл-проба (см. строку «Проба загрузки файла»)",
                    "код": "browser",
                })
                return True

            _ensure_modal_consent(form, page)  # надёжно: required + согласие по тексту подписи
            page.wait_for_timeout(300)

            # Многошаговые формы (например, Bitrix-чекаут): только заполнить поля,
            # а саму отправку делает отдельный шаг сценария «клик» («Далее»/«Оформить заказ»).
            if not cfg_enabled(форма_config.get("отправлять", True)):
                print(f"   ✍️ {название} - поля заполнены (без отправки).")
                записать_в_excel(
                    {
                        "тип": "PLAYWRIGHT-FORM",
                        "страница": страница,
                        "url": log_url,
                        "тип_селектора": format_form_selector_type(форма_config),
                        "ид": format_form_config_for_log(форма_config),
                        "название": название,
                        "имя": имя_теста,
                        "комментарий": КОММЕНТАРИЙ if КОММЕНТАРИЙ else имя_теста,
                        "статус": "ЗАПОЛНЕНО (без отправки)",
                        "код": "browser",
                    }
                )
                return True

            # Форма заказа? (тип блока «Оформление…» или URL checkout). На заказе
            # НЕ жмём Enter и двойную отправку проверяем безопасно - чтобы не
            # создать второй/лишний заказ на боевом сайте.
            try:
                _act = form.evaluate(
                    "f => { const g = f.tagName==='FORM' ? f : f.querySelector('form');"
                    " return g ? (g.getAttribute('action')||'') : ''; }")
            except Exception:  # noqa: BLE001
                _act = ""
            _is_order = bool(безопасная_отправка) or _ds_похоже_на_заказ(page.url, _act)

            # ── Enter отправляет форму? (пункт чек-листа) ──
            # БЕЗОПАСНО: вешаем перехватчик submit (capture) с preventDefault -
            # если Enter в текстовом поле вызывает отправку, ловим ФАКТ и ОТМЕНЯЕМ
            # её (реальная заявка НЕ уходит). Перехватчик снимаем в finally, чтобы
            # не заблокировать штатную отправку ниже. На форме заказа Enter не жмём.
            _enter_verdict, _enter_ком = None, ""
            if not _is_order:
                _enter_added = False
                try:
                    _ei = form.evaluate(
                        "f => { const g = f.tagName==='FORM' ? f : f.querySelector('form');"
                        " if(!g) return 'noform';"
                        " window.__entSub=false;"
                        " window.__entH=function(e){ window.__entSub=true; e.preventDefault();"
                        " e.stopPropagation(); g.removeEventListener('submit',window.__entH,true); };"
                        " g.addEventListener('submit',window.__entH,true); return 'ok'; }")
                    if _ei != "noform":
                        _enter_added = True
                        _inp = form.locator(
                            "input:not([type='checkbox']):not([type='radio'])"
                            ":not([type='file']):not([type='hidden']):not([type='submit'])"
                            ":not([type='button']):not([type='date']):not([type='range'])"
                            " >> visible=true").first
                        if _inp.count():
                            _inp.press("Enter", timeout=3000)
                            page.wait_for_timeout(350)
                            _sub = bool(form.evaluate("f => !!window.__entSub"))
                            _enter_verdict = "да" if _sub else "нет"
                            if not _sub:
                                _enter_ком = ("Форму нельзя отправить клавишей Enter (только "
                                              "кнопкой). Часто так делают намеренно - проверьте, "
                                              "ожидается ли отправка с клавиатуры.")
                except Exception:  # noqa: BLE001
                    _enter_verdict = None
                finally:
                    if _enter_added:
                        try:
                            form.evaluate(
                                "f => { const g = f.tagName==='FORM' ? f : f.querySelector('form');"
                                " if(g && window.__entH) g.removeEventListener('submit',window.__entH,true); }")
                        except Exception:  # noqa: BLE001
                            pass
                if _enter_verdict:
                    try:
                        записать_в_excel({
                            "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                            "тип_селектора": "поля", "ид": название,
                            "название": f"Enter отправляет форму: {название}",
                            "имя": имя_теста,
                            "статус": "Проверить" if _enter_verdict == "нет" else "OK",
                            "enter_отправляет": _enter_verdict,
                            "комментарий_готовый": _enter_ком or None,
                            "код": "enter_submit",
                        })
                        print(f"   ⏎ Enter отправляет «{название}»: {_enter_verdict}")
                    except Exception:  # noqa: BLE001
                        pass

            # Имена заполненных полей ДО отправки - чтобы после успеха проверить,
            # очистились ли они (пункт «если требуется»).
            try:
                _pre_keys = form.evaluate(
                    "f => [...f.querySelectorAll('input,textarea')]"
                    ".filter(e => { const t=(e.type||'').toLowerCase();"
                    " return !['hidden','submit','button','checkbox','radio','file'].includes(t)"
                    " && (e.offsetWidth||e.offsetHeight||e.getClientRects().length)"
                    " && (e.value||'').trim() && (e.name||e.id); })"
                    ".map(e => e.name || e.id)")
            except Exception:  # noqa: BLE001
                _pre_keys = []

            # Кнопка отправки: по умолчанию стандартные submit-кнопки; если у сайта
            # своя (например button.send у Авиапромсталь) - задаётся ключом «кнопка_css».
            _btn_css = str(форма_config.get("кнопка_css") or "").strip()
            sub = form.locator(
                _btn_css or "button[type='submit'], input[type='submit'], button.btn"
            ).first
            sub.scroll_into_view_if_needed()

            # ── Ошибки валидации отображаются корректно (показ/цвет/текст) ──
            # Нативную валидацию ловим checkValidity() без отправки; кастомную -
            # контролируемым ПУСТЫМ сабмитом под двойной защитой (preventDefault +
            # abort POST), заявка НЕ уходит. Поля восстанавливаются до штатной
            # отправки ниже. Делаем ДО «Обработки ошибок», пока форма не тронута.
            try:
                _vd = проверка_отображения_ошибок(form, page, sub, _is_order)
                записать_в_excel({
                    "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                    "тип_селектора": "поля", "ид": название,
                    "название": f"Ошибки валидации (показ/цвет/текст): {название}",
                    "имя": имя_теста,
                    "статус": ("Проверить" if _vd["состояние"]
                               in ("нет", "без реакции", "проверить вручную") else "OK"),
                    "ошибки_валидации": _vd["состояние"],
                    "комментарий_готовый": _vd["детали"],
                    "код": "validation_display",
                })
                print(f"   🚨 Ошибки валидации «{название}»: {_vd['состояние']} — {_vd['детали']}")
            except Exception as _evd:  # noqa: BLE001
                print(f"   ⚠️ Проверка показа ошибок валидации не удалась: {_evd}")

            # ── Обработка ошибок отправки (пункт чек-листа) ──
            # Нарочно роняем ПЕРВЫЙ запрос отправки (route.abort) и смотрим реакцию
            # формы: показала ошибку / молчит / соврала «успех». Реальная заявка НЕ
            # уходит (запрос отменён). Только AJAX-формы (у навигационных ошибку
            # рисует сервер) и не заказ. Штатную отправку защищаем: в finally снимаем
            # перехват и возвращаем кнопку в рабочее состояние - настоящая отправка
            # ниже уходит как обычно.
            _err_verdict, _err_ком = None, ""
            if not _is_order:
                _err_info = {"n": 0, "ajax": False}
                _err_armed = False
                try:
                    _e_btn_до = _текст_кнопки(sub, 800)
                except Exception:  # noqa: BLE001
                    _e_btn_до = ""
                # Снимок «видно ли подтверждение ДО пробы»: форма часто живёт
                # ВНУТРИ модалки, где панель «Спасибо» уже отрисована (скрыта/
                # видима) как часть разметки. Если детектор увидит её ПОСЛЕ
                # аборта, это не «ложный успех», а разметка модалки - отличить
                # можно только сравнением «было/стало» (иначе ✗ на каждой модалке).
                _подтв_до_err = _подтверждение_видно(page)
                try:
                    def _err_handler(route):
                        try:
                            req = route.request
                            if ((req.method or "").upper() == "POST"
                                    and not _ds_это_трекер(req.url)
                                    and _err_info["n"] == 0):
                                _err_info["n"] += 1
                                _err_info["ajax"] = (req.resource_type in ("fetch", "xhr"))
                                route.abort("failed")
                            else:
                                route.continue_()
                        except Exception:  # noqa: BLE001
                            try:
                                route.continue_()
                            except Exception:  # noqa: BLE001
                                pass
                    page.route("**/*", _err_handler)
                    _err_armed = True
                    try:
                        sub.click(timeout=5000)
                    except Exception:  # noqa: BLE001
                        try:
                            sub.click(timeout=5000, force=True)
                        except Exception:  # noqa: BLE001
                            pass
                    page.wait_for_timeout(1600)
                    if _err_info["n"] == 0 or not _err_info["ajax"]:
                        _err_verdict = None       # запрос не пойман / навигационная форма
                    else:
                        _euved = детект_уведомления_пользователю(
                            page, _e_btn_до, "", кнопка=sub, таймаут_мс=2500)
                        if str(_euved).startswith("Да") and _подтв_до_err:
                            # Подтверждение висело ещё ДО пробы (заранее
                            # отрисованная панель «Спасибо» в модалке) - отличить
                            # реальный «ложный успех» от разметки нельзя.
                            # Не выносим ✗: пропускаем строку (не ложная ошибка).
                            _err_verdict = None
                        elif str(_euved).startswith("Да"):
                            _err_verdict = "ложный успех"
                            _err_ком = ("При УПАВШЕМ запросе форма показала «успешно» - "
                                        "пользователь думает, что заявка ушла, а её нет.")
                        elif _видна_ошибка_отправки(page):
                            _err_verdict = "корректно"
                        else:
                            # Форма молчит на ИСКУССТВЕННО оборванный запрос. Это не
                            # дефект: заявка в реальности уходит (см. «Статус»), а
                            # нативную ошибку при реальном сбое показывает браузер.
                            # Раньше тут был ложный ✗ на рабочих формах - убрано.
                            _err_verdict = None
                except Exception:  # noqa: BLE001
                    _err_verdict = None
                finally:
                    if _err_armed:
                        try:
                            page.unroute("**/*", _err_handler)
                        except Exception:  # noqa: BLE001
                            try:
                                page.unroute("**/*")
                            except Exception:  # noqa: BLE001
                                pass
                    # вернуть кнопку в рабочее состояние для НАСТОЯЩЕЙ отправки ниже
                    try:
                        sub.evaluate("b => { try { b.disabled = false;"
                                     " b.removeAttribute('disabled'); } catch(_){} }")
                    except Exception:  # noqa: BLE001
                        pass
                    page.wait_for_timeout(300)
                if _err_verdict:
                    try:
                        записать_в_excel({
                            "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                            "тип_селектора": "поля", "ид": название,
                            "название": f"Обработка ошибок отправки: {название}",
                            "имя": имя_теста,
                            "статус": "Проверить" if _err_verdict in ("молчит", "ложный успех") else "OK",
                            "обработка_ошибок": _err_verdict,
                            "комментарий_готовый": _err_ком or None,
                            "код": "error_handling",
                        })
                        print(f"   🚑 Обработка ошибок «{название}»: {_err_verdict}"
                              + (f" — {_err_ком}" if _err_ком else ""))
                    except Exception:  # noqa: BLE001
                        pass

            # Пункт 2.7: запомним текст кнопки ДО отправки - чтобы поймать её смену
            # на подтверждение («Отправить» → «Заявка отправлена»).
            try:
                _btn_текст_до = _текст_кнопки(sub, 1000)
            except Exception:
                _btn_текст_до = ""
            _t_отправки = _time.time()   # цели Метрики считаем с момента отправки

            # CSRF: поля уже заполнены, отправка ещё не было - самое время
            # проверить наличие токена (читаем живой DOM, ничего не отправляем).
            _csrf_рез = _найти_csrf_поле(form)
            _csrf_куки = _прочитать_куки_csrf(page)
            _csrf_кол, _csrf_дет = csrf_вердикт(
                bool(_csrf_рез.get("найдено")), bool(_csrf_рез.get("заполнено")),
                bool(_csrf_рез.get("ошибка")), куки=_csrf_куки)
            print(f"   🔑 CSRF-защита «{название}»: {_csrf_кол} - {_csrf_дет}")

            # Данные формы дошли до сервера: снимок значений ПРЯМО перед
            # кликом (позже, чем заполнение - учитывает всё, что могли
            # доделать чужие пробы выше по потоку) + слушаем реальный POST
            # той же единственной легитимной отправки - лишнего запроса не
            # делаем. Регистрируется БЕЗУСЛОВНО (не гейтим по is_order, как
            # двойную отправку) - тут нет второго клика, только наблюдение.
            _дд_поля = _снять_поля_формы(form)
            _дд_запросы_тела = []

            def _дд_on(req):
                try:
                    if (req.method or "").upper() == "POST" and not _ds_это_трекер(req.url):
                        _дд_запросы_тела.append(_тело_запроса_для_поиска(
                            req.headers.get("content-type", ""), req.post_data or ""))
                except Exception:  # noqa: BLE001
                    pass

            try:
                page.on("request", _дд_on)
            except Exception:  # noqa: BLE001
                _дд_on = None

            # Ответ РЕАЛЬНОЙ отправки формы (для честного статуса): ловим статус и
            # тело POST-ответа. Нужен, чтобы «УСПЕШНО» ставить по ФАКТУ (сервер
            # принял / пользователь увидел подтверждение), а не просто потому, что
            # «на странице нет слова ошибка». Берём ПОСЛЕДНИЙ не-трекерный POST.
            _отв_форма = {"текст": None, "статус": None}

            def _отв_on(resp):
                try:
                    rq = resp.request
                    if ((rq.method or "").upper() == "POST"
                            and not _ds_это_трекер(resp.url)
                            and _отв_форма["статус"] is None):  # только ПЕРВЫЙ POST
                        # (реальная отправка; повторный клик двойной отправки идёт
                        # позже и не должен перебивать вердикт настоящей отправки)
                        _отв_форма["статус"] = resp.status
                        try:
                            _отв_форма["текст"] = resp.text()
                        except Exception:  # noqa: BLE001
                            _отв_форма["текст"] = ""
                except Exception:  # noqa: BLE001
                    pass

            try:
                page.on("response", _отв_on)
            except Exception:  # noqa: BLE001
                _отв_on = None

            # Модальные окна работают корректно (если есть): форма могла
            # прийти не из выделенного блока «модалки» (см. _modal_flow_on_page),
            # а из обычного сценария «клик → форма» - и всё равно открыться в
            # модалке/попапе (как «Обратная связь» через .popup form). Ловим
            # это здесь; саму пробу закрытия делаем позже, после отправки.
            _модалка_вокруг = _найти_модалку_вокруг(form)
            # «Модалка открывается»: если форма реально лежит в открытом
            # модальном окне/попапе (нашли контейнер вокруг УЖЕ заполненной
            # формы) - фиксируем «Да» ЗДЕСЬ, до отправки (после отправки попап
            # часто уже закрыт). Раньше эту колонку писал ТОЛЬКО выделенный блок
            # «модалка», а обычные формы, открытые в попапе, шли прочерком.
            _модалка_откр_обычный = "Да" if _модалка_вокруг is not None else ""

            # Защита от спама/ботов - пассивный слой (всегда, без лишних
            # запросов): капча по уже загруженному HTML + honeypot по живому DOM.
            try:
                _honeypot_рез = form.evaluate(_JS_HONEYPOT)
            except Exception:  # noqa: BLE001
                _honeypot_рез = {"найдено": False, "имя": ""}
            _капча_рез = защита_от_спама_из_html(page.content())
            _спам_обн = {
                "капча": _капча_рез["капча"], "капча_какая": _капча_рез["какая"],
                "honeypot": bool(_honeypot_рез.get("найдено")),
                "honeypot_имя": _honeypot_рез.get("имя", ""),
            }
            _спам_кол, _спам_дет = лимит_пассивно_вердикт(_спам_обн)
            print(f"   🚦 Защита от спама (пассивно) «{название}»: {_спам_кол} - {_спам_дет}")

            # ── Двойная отправка (пункт чек-листа): гибрид ──
            # Заказ (безопасная_отправка / URL-признак) - БЕЗОПАСНО: один клик,
            # потом смотрим, заблокировалась ли кнопка (не создаём второй заказ).
            # Лёгкие формы - ТОЧНО: реально жмём второй раз и считаем POST-заявки
            # (одинаковый адрес ≥2 = дубль). Всё в try, чтобы не сломать отправку.
            _ds_verdict, _ds_ком = None, ""
            _ds_safe = _is_order          # заказ - без реального второго клика
            _ds_posts, _ds_on = [], None
            if not _ds_safe:
                try:
                    def _ds_on(req):
                        try:
                            if (req.method or "").upper() == "POST" and not _ds_это_трекер(req.url):
                                _ds_posts.append(req.url.split("?")[0])
                        except Exception:  # noqa: BLE001
                            pass
                    page.on("request", _ds_on)
                except Exception:  # noqa: BLE001
                    _ds_on = None

            # Снимок валидных значений ДО отправки (форма ещё заполнена): нужен
            # активной пробе лимита и серверной валидации, которые бегут ПОСЛЕ
            # отправки, когда форма уже очищена. Без него их вердикт был
            # «неоднозначно / проверьте вручную». Снимаем только если пробы включены.
            _снимок_валидных = None
            if валидация_проба or лимит_проба:
                try:
                    _снимок_валидных = form.evaluate(_JS_RATELIMIT_SNAPSHOT)
                except Exception:  # noqa: BLE001
                    _снимок_валидных = None

            if xss_проба:
                _xss_reset(page)         # чистый флаг перед отправкой payload
            _перекрыто = False   # кнопку перекрыл другой элемент (обход JS-кликом)
            try:
                sub.click(timeout=5000)
            except Exception:
                # Кнопку видно, но её ПЕРЕКРЫВАЕТ другой элемент (баг вёрстки:
                # напр. на «Срочный заказ» декоративная картинка quick_order--img
                # лежит ПОВЕРХ кнопки). force-клик кликает по КООРДИНАТЕ и попадает
                # в оверлей, а не в кнопку - форма НЕ отправляется (POST не уходит,
                # тул писал ложное «нет подтверждения»). JS-клик b.click() шлёт
                # событие ПРЯМО на элемент-кнопку (мимо хит-теста координат):
                # <button type=submit> отправит форму штатно, обработчик сайта
                # сработает. force оставляем запасным вариантом.
                _перекрыто = True
                print("      ↻ Обычный клик перекрыт - отправляю через JS (b.click)")
                try:
                    sub.evaluate("b => b.click()")
                except Exception:  # noqa: BLE001
                    try:
                        sub.click(timeout=5000, force=True)
                    except Exception:  # noqa: BLE001
                        pass

            # Сам тест двойной отправки (guarded - никогда не роняет отправку).
            try:
                if _ds_safe:
                    _locked = False
                    for _ in range(6):                 # ~600 мс наблюдения
                        try:
                            _locked = bool(sub.evaluate(
                                "b => !!(b.disabled || b.getAttribute('aria-disabled')==='true'"
                                " || /disabl|load|sending|wait|process|отправ/i.test(b.className||'')"
                                " || getComputedStyle(b).pointerEvents==='none'"
                                " || b.offsetParent===null)"))
                        except Exception:  # noqa: BLE001
                            _locked = True             # кнопка исчезла = форма ушла
                        if _locked:
                            break
                        page.wait_for_timeout(100)
                    _ds_verdict = "защищена" if _locked else "под вопросом"
                    if not _locked:
                        _ds_ком = ("Кнопка не блокируется при отправке - двойное "
                                   "нажатие теоретически возможно, проверьте вручную.")
                else:
                    # БЫСТРЫЙ второй сабмит - как реальный нетерпеливый пользователь/
                    # спамер: жмём ещё раз СРАЗУ, не дожидаясь, пока форма покажет
                    # «Спасибо» и кнопка заблокируется. Второй клик - через JS
                    # b.click() ПРЯМО на кнопку:
                    #   • он НЕ зависит от того, что кнопку визуально накрыл успех-
                    #     оверлей или она косметически сменила текст (обычный
                    #     координатный клик тут глохнет → флаки «защищена» на
                    #     идентичных формах, как «Оставить заявку» vs «Обратная связь»);
                    #   • по спецификации b.click() НЕ срабатывает на РЕАЛЬНО disabled
                    #     кнопке → форма с СИНХРОННОЙ защитой честно даст 1 POST.
                    # Ждать ~1.6 с, пока кнопка "успокоится", НЕЛЬЗЯ: асинхронная
                    # блокировка/оверлей ПОСЛЕ ответа AJAX не спасает от быстрого
                    # двойного клика - обе заявки уходят (подтверждено в админке).
                    # Перекрытую кнопку (_перекрыто) вторым кликом НЕ трогаем, чтобы
                    # не создать лишнюю реальную заявку - для неё вердикт ниже
                    # станет «под вопросом» (см. блок подсчёта).
                    if not _перекрыто:
                        page.wait_for_timeout(250)     # дать 1-му POST зарегистрироваться
                        # Реальный юзер быстрым дабл-кликом отправит второй раз
                        # ТОЛЬКО если кнопка ещё принимает клики. JS b.click()
                        # игнорит disabled (там он сам no-op по спецификации), но
                        # НЕ уважает pointer-events:none / aria-disabled - поэтому
                        # проверяем их ЯВНО: если форма СИНХРОННО залочила кнопку
                        # так (это защита - реальный клик бы не прошёл), второй раз
                        # НЕ жмём → останется 1 POST → «защищена». СКРЫТОСТЬ/
                        # перекрытие успех-оверлеем тут НЕ проверяем: на (медленном)
                        # сайте оверлея на 250 мс ещё нет, а проверка «спрятана» и
                        # вернула бы прежний флаки (её async-появление ≠ защита от
                        # быстрого дабл-клика).
                        try:
                            _залочена2 = bool(sub.evaluate(
                                "b => { const cs = getComputedStyle(b);"
                                " return !!(b.disabled"
                                "   || b.getAttribute('aria-disabled')==='true'"
                                "   || cs.pointerEvents==='none'); }"))
                        except Exception:  # noqa: BLE001
                            _залочена2 = True   # кнопка исчезла - нажать нельзя
                        if not _залочена2:
                            try:
                                sub.evaluate("b => b.click()")
                            except Exception:  # noqa: BLE001
                                pass    # кнопка disabled/исчезла - второй не ушёл
            except Exception:  # noqa: BLE001
                pass

            page.wait_for_timeout(2000)   # ждём и МЕДЛЕННЫЙ второй POST (иначе _mx=1)

            # Подсчёт отправок (точный режим): один и тот же POST-адрес ≥2 = дубль.
            try:
                if _ds_on is not None:
                    page.remove_listener("request", _ds_on)
                if not _ds_safe:
                    _cnt = {}
                    for _u in _ds_posts:
                        _cnt[_u] = _cnt.get(_u, 0) + 1
                    _mx = max(_cnt.values()) if _cnt else 0
                    if _перекрыто and _mx < 2:
                        # Кнопку вторым кликом дожать НЕ могли (перекрыта): обычный
                        # клик глохнет на оверлее (дал бы ложное «защищена»), а
                        # форсировать = лишняя реальная заявка. Честно - «проверить».
                        _ds_verdict = "под вопросом"
                        _ds_ком = ("Кнопка перекрыта другим элементом - чистую "
                                   "двойную отправку автоматически не проверить "
                                   "(второй клик глохнет на оверлее, а форсировать - "
                                   "создать лишнюю заявку). Проверьте вручную: "
                                   "заполните форму и быстро нажмите отправку 2 раза "
                                   "- если в админку пришло 2 заявки, защиты нет.")
                    elif _mx >= 2:
                        _ds_verdict = "не защищена"
                        _ds_ком = ("Тул отправил форму ДВАЖДЫ - клиентской защиты от "
                                   "повтора нет (кнопка не блокируется, второй клик "
                                   "уходит вторым POST-запросом). На части форм обе "
                                   "заявки реально попадают в админку; на других "
                                   "сервер может отклонить второй по одноразовому "
                                   "токену - проверьте в админке, пришло ли 2 заявки. "
                                   "Смена кнопки на «Отправлено» - косметика.")
                    elif _mx == 1:
                        _ds_verdict = "защищена"
                    else:
                        _ds_verdict = "под вопросом"
                        _ds_ком = ("Не удалось посчитать отправки (форма без POST/"
                                   "переход) - двойную отправку проверьте вручную.")
            except Exception:  # noqa: BLE001
                pass

            # Данные формы дошли до сервера: снимаем слушатель и сверяем
            # снимок «до клика» с реально пойманными телами POST-запросов.
            try:
                if _дд_on is not None:
                    page.remove_listener("request", _дд_on)
            except Exception:  # noqa: BLE001
                pass
            try:
                if _отв_on is not None:
                    page.remove_listener("response", _отв_on)
            except Exception:  # noqa: BLE001
                pass
            _дд_кол, _дд_дет = данные_формы_вердикт(
                bool(_дд_запросы_тела), " ".join(_дд_запросы_тела),
                _дд_поля.get("поля") or {})
            print(f"   📡 Данные дошли до сервера «{название}»: {_дд_кол} - {_дд_дет}")

            html = page.content()
            _form_err = response_indicates_form_error(html)

            # Пункт 2.7: увидел ли пользователь подтверждение заявки (попап/картинка
            # «спасибо», текст успеха или смена текста кнопки). Считаем ДО статуса -
            # это ГЛАВНЫЙ признак, что форма реально сработала для человека.
            try:
                _btn_текст_после = _текст_кнопки(sub, 1000)
            except Exception:
                _btn_текст_после = ""
            # Таймаут увеличен до 12 с: часть форм (особенно Bitrix-попапы)
            # подтверждают МЕДЛЕННО - кнопка становится «Отправлено» / попап
            # «спасибо» появляется через 6-10 с (плюс детекция идёт ПОСЛЕ пробы
            # двойной отправки). На 6 с тул не дожидался и ложно писал «нет
            # подтверждения», хотя заявка уходила. Опрос ВОЗВРАЩАЕТСЯ СРАЗУ, как
            # найдёт «Да», поэтому быстрые формы не замедляются.
            _уведомл_польз = детект_уведомления_пользователю(
                page, _btn_текст_до, _btn_текст_после, кнопка=sub, таймаут_мс=12000)
            _есть_подтверждение = str(_уведомл_польз).startswith("Да")
            # Вердикт по РЕАЛЬНОМУ ответу сервера на отправку (если поймали).
            _отв_вердикт = _ответ_формы_вердикт(
                _отв_форма.get("текст"), _отв_форма.get("статус"))
            # ЯВНЫЙ успех сервера (не голый 200!): JSON success/ok=true или
            # «спасибо/успешно/принят» в теле ответа. Голого 200 недостаточно -
            # у сломанного фронта сервер тоже может вернуть 200 без реальной заявки.
            _явный_успех_ответа = False
            _отв_текст = (_отв_форма.get("текст") or "")
            try:
                _отв_j = json.loads(_отв_текст) if _отв_текст.strip() else None
                if isinstance(_отв_j, dict) and (
                        _отв_j.get("success") is True or _отв_j.get("ok") is True):
                    _явный_успех_ответа = True
            except (ValueError, TypeError):
                pass
            if not _явный_успех_ответа:
                _отв_low = _отв_текст.lower()
                if any(m in _отв_low for m in ("спасибо", "успешно", "принят", "заявка отправлен")):
                    _явный_успех_ответа = True

            _коммент_готовый = None
            if response_indicates_captcha_block(html):
                статус = "ОШИБКА: КАПЧА"
            elif _form_err:
                статус = "ОШИБКА"
                _коммент_готовый = _form_err
            elif "ошибк" in html.lower() and any(
                x in html.lower()
                for x in ("не удалось", "не отправлен", "отклонен", "invalid")
            ):
                статус = "ОШИБКА (сообщение на странице)"
            elif _отв_вердикт == "ошибка":
                # Сервер ответил на отправку ошибкой (не «успех»), даже если на
                # странице нет видимого текста ошибки.
                статус = "ОШИБКА (сервер отклонил отправку)"
                _коммент_готовый = ("Сервер ответил на отправку формы НЕ успехом "
                                    "(ошибка/невалидный ответ) - заявка, скорее всего, не ушла.")
            elif _есть_подтверждение:
                # Пользователь увидел подтверждение - форма реально сработала.
                статус = "УСПЕШНО (Playwright - как ручная отправка)"
            elif _явный_успех_ответа:
                # Подтверждения пользователю нет, но сервер ЯВНО ответил успехом
                # (JSON success / «спасибо» в ответе) - засчитываем успех, но
                # помечаем, что человек подтверждения не видит.
                статус = "УСПЕШНО (Playwright - как ручная отправка)"
                _коммент_готовый = ("Сервер принял заявку, но пользователю НЕ показано "
                                    "подтверждение - человек не поймёт, что заявка ушла.")
            else:
                # Ни подтверждения пользователю, ни ответа сервера об успехе:
                # для посетителя форма выглядит нерабочей (заявка могла не уйти).
                # Раньше тут ставилось «УСПЕШНО» просто из-за отсутствия слова
                # «ошибка» на странице - это и был ложный «успех».
                статус = "НЕТ ПОДТВЕРЖДЕНИЯ (форма не показала успех)"
                _коммент_готовый = ("После отправки нет НИ подтверждения пользователю, "
                                    "ни ответа сервера об успехе - для посетителя форма "
                                    "выглядит нерабочей (заявка могла не уйти). Проверьте "
                                    "форму вручную и приход заявки в админку/на почту.")

            # Защита от XSS (Фаза 1): payload был в имени. Смотрим, исполнился ли
            # он / попал ли сырым HTML на ответ. Пусто, если проба выкл.
            # ВАЖНО: для XSS достаточно, что payload УШЁЛ на сервер (POST с данными),
            # а не что форма показала «успех». Иначе на сломанных формах (статус
            # «НЕТ ПОДТВЕРЖДЕНИЯ») XSS ложно уходил в «Проверить», хотя payload
            # реально отправлен и не исполнился = форма защищена.
            _xss_кол = ""
            if xss_проба:
                _отпр = (str(статус).upper().startswith("УСПЕШНО")
                         or str(_дд_кол).strip().lower().startswith("да"))
                _xss_исп, _xss_сыр = _xss_наблюдение(page, html)
                _xss_ст, _xss_дет = xss_вердикт(_отпр, _xss_исп, _xss_сыр)
                _xss_кол = _xss_ст
                print(f"   🛡️ Защита от XSS «{название}»: {_xss_ст} - {_xss_дет}")

            записать_в_excel(
                {
                    "тип": "PLAYWRIGHT-FORM",
                    "страница": страница,
                    "url": log_url,
                    "тип_селектора": format_form_selector_type(форма_config),
                    "ид": format_form_config_for_log(форма_config),
                    "название": название,
                    "имя": имя_теста,
                    "комментарий": КОММЕНТАРИЙ if КОММЕНТАРИЙ else имя_теста,
                    "комментарий_готовый": _коммент_готовый,
                    "статус": статус,
                    "уведомление": _уведомл_польз,
                    "защита_от_xss": _xss_кол,
                    "csrf_защита": _csrf_кол,
                    "защита_от_спама_пассивно": _спам_кол,
                    "код": "browser",
                }
            )
            print(f"   ✅ {название} - {статус}  ·  уведомление польз.: {_уведомл_польз}")

            # Двойная отправка: отдельная строка отчёта (колонка «Двойная отправка»).
            # Коммент - только когда есть что пояснить (не защищена / под вопросом).
            if _ds_verdict:
                try:
                    записать_в_excel({
                        "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                        "тип_селектора": "поля", "ид": название,
                        "название": f"Двойная отправка (двойной клик): {название}",
                        "имя": имя_теста,
                        "статус": "Проверить" if _ds_verdict == "не защищена" else "OK",
                        "двойная_отправка": _ds_verdict,
                        "комментарий_готовый": _ds_ком or None,
                        "код": "double_submit",
                    })
                    print(f"   🔁 Двойная отправка «{название}»: {_ds_verdict}"
                          + (f" — {_ds_ком}" if _ds_ком else ""))
                except Exception:  # noqa: BLE001
                    pass

            # Данные формы дошли до сервера: отдельная строка отчёта (колонка
            # «Данные дошли до сервера»). Коммент - только когда есть что
            # пояснить (не «Да»).
            if _дд_кол:
                try:
                    записать_в_excel({
                        "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                        "тип_селектора": "поля", "ид": название,
                        "название": f"Данные формы дошли до сервера: {название}",
                        "имя": имя_теста,
                        "статус": "OK" if _дд_кол == "Да" else "Проверить",
                        "данные_дошли": _дд_кол,
                        "комментарий_готовый": _дд_дет if _дд_кол != "Да" else None,
                        "код": "data_delivery",
                    })
                except Exception:  # noqa: BLE001
                    pass

            # Поля очищаются после успешной отправки (пункт «если требуется» -
            # справочно, не ошибка). Читаем те же поля после успеха: пусты - очищены.
            # Только при успехе и пока поля ещё на странице (AJAX-форма не ушла).
            if _pre_keys and str(статус).startswith("УСПЕШНО"):
                try:
                    _after = form.evaluate(
                        "(f, keys) => { const set = new Set(keys.map(String));"
                        " const seen = new Set(); const out = [];"
                        " for (const e of f.querySelectorAll('input,textarea')) {"
                        "   const k = (e.name || e.id || ''); if (!k) continue; seen.add(k);"
                        "   if (set.has(k)) out.push(((e.value||'').trim()) ? 'full' : 'empty'); }"
                        " for (const k of keys) { if (!seen.has(String(k))) out.push('gone'); }"
                        " return out; }",
                        _pre_keys)
                    _present = [s for s in (_after or []) if s != "gone"]
                    _clr = None
                    if _present:
                        _empty = sum(1 for s in _present if s == "empty")
                        if _empty == len(_present):
                            _clr = "очищены"
                        elif _empty == 0:
                            _clr = "не очищены"
                        else:
                            _clr = "очищены частично"
                    if _clr:
                        записать_в_excel({
                            "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                            "тип_селектора": "поля", "ид": название,
                            "название": f"Поля очищаются после отправки: {название}",
                            "имя": имя_теста,
                            "статус": "OK",   # «если требуется» - не ошибка, справочно
                            "поля_очищены": _clr,
                            "комментарий_готовый": (
                                "Поля не очистились после отправки - если форму заполняют "
                                "повторно, стоит очищать (не критично)."
                                if _clr == "не очищены" else None),
                            "код": "fields_cleared",
                        })
                        print(f"   🧹 Поля очищаются «{название}»: {_clr}")
                except Exception:  # noqa: BLE001
                    pass

            # Цель Метрики этой формы (ключ «цель»): ждём немного (летит ajax-ом)
            # и пишем в отчёт «сработала / НЕ сработала». Причину показываем,
            # только если задана в конфиге («цель_причина»).
            _цель_имя = str(форма_config.get("цель") or "").strip()
            if _цель_имя and проверять_цели:
                _t0g = _time.time()
                _найдена = False
                while True:
                    _найдена = any(g["цель"] == _цель_имя for g in _цели_с(_t_отправки))
                    if _найдена or (_time.time() - _t0g) * 1000 >= 8000:
                        break
                    page.wait_for_timeout(500)
                _nm = str(форма_config.get("цель_название") or f"Цель: {название}")
                if _найдена:
                    _г_статус, _г_коммент = "СРАБОТАЛА", None
                else:
                    _г_статус = "НЕ СРАБОТАЛА"
                    _г_коммент = (str(форма_config.get("цель_причина") or "").strip()
                                  or f"Цель «{_цель_имя}» не зафиксирована Метрикой")
                записать_в_excel(
                    {
                        "тип": "ЦЕЛЬ (Метрика)", "страница": страница,
                        "url": page.url or log_url, "тип_селектора": "цель",
                        "ид": _цель_имя, "название": _nm,
                        "имя": имя_теста_из_конфига(страница, "цель", _nm,
                                                    название_контекста=_nm),
                        "комментарий": _г_коммент or "",
                        "комментарий_готовый": _г_коммент,
                        "статус": _г_статус, "код": "ym",
                    }
                )
                print(f"   🎯 Цель «{_цель_имя}» - {_г_статус}")

            # Модальные окна: закрывается (только если форма реально лежала
            # в модалке/попапе - _модалка_вокруг). Делаем ДО переоткрывающих
            # проб (серверная валидация/лимит уводят страницу перезагрузкой).
            if _модалка_вокруг is not None:
                try:
                    _закр_ст2, _закр_способ2 = _проба_закрытия_модалки(page, _модалка_вокруг)
                    записать_в_excel({
                        "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                        "тип_селектора": "модалка", "ид": название,
                        "название": f"Модалка закрывается: {название}",
                        "имя": имя_теста,
                        "статус": "OK" if _закр_ст2 == "Да" else "Проверить",
                        "модалка_открылась": _модалка_откр_обычный,
                        "модалка_закрывается": _закр_ст2,
                        "комментарий_готовый": f"способ: {_закр_способ2}",
                        "код": "modal_close",
                    })
                    print(f"   🪟 Модалка «{название}» закрывается: "
                          f"{_закр_ст2} ({_закр_способ2})")
                except Exception as _emc2:  # noqa: BLE001
                    print(f"   ⚠️ Проба закрытия модалки не удалась: {_emc2}")

            # ── Переоткрытие формы для повторных проб ──
            # Серверная валидация и активный лимит бьют по форме ПОВТОРНО, но
            # большинство форм блокируется после первой отправки (кнопка
            # «Отправлено»), поэтому на том же заходе проба не проходит и вердикт
            # выходил «Проверить». Перед каждой такой пробой пытаемся ЗАНОВО
            # открыть свежую форму: перезагрузка + повтор шагов открытия
            # (переоткрыть_fn) + повторное заполнение из снимка валидных значений.
            # Всё под try: если переоткрыть не вышло - проба идёт по СТАРОЙ форме,
            # ровно как раньше (никакой регрессии, максимум прежнее «Проверить»).
            def _переоткрыть_и_заполнить():
                if переоткрыть_fn is None or not _снимок_валидных or _is_order:
                    return None, None
                try:
                    if not переоткрыть_fn():
                        return None, None
                    page.wait_for_timeout(600)
                    if use_text:
                        _loc = page.locator("form").filter(
                            has_text=str(форма_config.get("text") or "").strip())
                    else:
                        _loc = page.locator(sel)
                    _f = None
                    for _k in range(min(_loc.count(), 6)):
                        try:
                            if _loc.nth(_k).is_visible():
                                _f = _loc.nth(_k)
                                break
                        except Exception:  # noqa: BLE001
                            continue
                    if _f is None:
                        return None, None
                    try:
                        _tag = _f.evaluate("el => el.tagName.toLowerCase()")
                        if _tag in ("input", "textarea", "select"):
                            _anc = _f.locator("xpath=ancestor::form[1]")
                            if _anc.count() > 0:
                                _f = _anc.first
                    except Exception:  # noqa: BLE001
                        pass
                    _f = _apply_container_expand(_f, форма_config)
                    _f.evaluate(_JS_RATELIMIT_RESTORE, _снимок_валидных)
                    _bcss = str(форма_config.get("кнопка_css") or "").strip()
                    _s = _f.locator(
                        _bcss or "button[type='submit'], input[type='submit'], button.btn"
                    ).first
                    try:
                        _s.scroll_into_view_if_needed()
                    except Exception:  # noqa: BLE001
                        pass
                    return _f, _s
                except Exception:  # noqa: BLE001
                    return None, None

            # Серверная валидация (пункт «нельзя отправить неверные данные
            # через DevTools»): под галочкой. Пробуем на СВЕЖЕЙ форме.
            if валидация_проба:
                try:
                    _sv_form, _sv_sub = form, sub
                    _fresh1 = _переоткрыть_и_заполнить()
                    if _fresh1[0] is not None:
                        _sv_form, _sv_sub = _fresh1
                    _srv = проба_серверной_валидации(_sv_form, page, _sv_sub, _is_order,
                                                     снимок=_снимок_валидных)
                    _srv_попытки = _srv.get("попытки") or {}
                    _srv_ст, _srv_verdict_дет = валидация_сервера_вердикт(_srv_попытки)
                    # Комментарий - подробная расшифровка С ИМЕНЕМ ПОЛЯ из пробы
                    # (что именно ПРОШЛО/отклонено); если её нет - краткий вердикт.
                    _srv_дет = _srv.get("детали") or _srv_verdict_дет or ""
                    записать_в_excel({
                        "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                        "тип_селектора": "поля", "ид": название,
                        "название": f"Серверная валидация: {название}",
                        "имя": имя_теста,
                        "статус": "Проверить" if _srv_ст == "УЯЗВИМА" else "OK",
                        "серверная_валидация": _srv_ст,
                        "комментарий_готовый": _srv_дет or None,
                        "код": "server_validation",
                    })
                    print(f"   🛡️ Серверная валидация «{название}»: {_srv_ст} - {_srv_дет}")
                except Exception as _esv:  # noqa: BLE001
                    print(f"   ⚠️ Проба серверной валидации не удалась: {_esv}")

            # Лимит запросов - активный залп (пункт «Ограничено количество
            # запросов»): под галочкой. Тоже на СВЕЖЕЙ форме (иначе повтор не
            # проходит из-за блокировки после первой отправки).
            if лимит_проба:
                try:
                    _rl_form, _rl_sub = form, sub
                    _fresh2 = _переоткрыть_и_заполнить()
                    if _fresh2[0] is not None:
                        _rl_form, _rl_sub = _fresh2
                    _rl = активная_проба_лимита(_rl_form, page, _rl_sub, _is_order,
                                                снимок=_снимок_валидных)
                    _rl_попытки = _rl.get("попытки") or []
                    if _rl_попытки:
                        _rl_ст, _rl_дет = лимит_активно_вердикт(_rl_попытки)
                    else:
                        _rl_ст, _rl_дет = "Проверить", _rl.get("детали") or ""
                    записать_в_excel({
                        "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                        "тип_селектора": "поля", "ид": название,
                        "название": f"Защита от спама (активно): {название}",
                        "имя": имя_теста,
                        "статус": ("OK" if _rl_ст == "Сработала защита"
                                   else "Проверить"),
                        "защита_от_спама_активно": _rl_ст,
                        "комментарий_готовый": _rl_дет or None,
                        "код": "rate_limit",
                    })
                    print(f"   🚦 Защита от спама (активно) «{название}»: {_rl_ст} - {_rl_дет}")
                except Exception as _erl:  # noqa: BLE001
                    print(f"   ⚠️ Активная проба лимита не удалась: {_erl}")

            return True

        except Exception as e:
            print(f"   ❌ Playwright форма: {e}")
            записать_в_excel(
                {
                    "тип": "PLAYWRIGHT-FORM",
                    "страница": страница,
                    "url": log_url,
                    "тип_селектора": format_form_selector_type(форма_config),
                    "ид": format_form_config_for_log(форма_config),
                    "название": название,
                    "имя": имя_теста,
                    "комментарий": КОММЕНТАРИЙ if КОММЕНТАРИЙ else имя_теста,
                    "статус": f"ОШИБКА ({e})",
                    "код": str(e),
                }
            )
            return False

    def _modal_flow_on_page(
        page,
        url_for_excel: str,
        страница: str,
        значение,
        название_теста: str,
        *,
        initial_url=None,
        название_контекста=None,
    ):
        nctx = название_контекста if название_контекста is not None else название_теста
        имя_теста = имя_теста_из_конфига(
            страница,
            значение,
            название_теста,
            название_контекста=nctx,
        )
        if initial_url:
            _goto_with_retry(page, initial_url)
            page.wait_for_timeout(3000)
        else:
            page.wait_for_timeout(400)

        log_url = page.url or url_for_excel

        try:
            # Наведение на карточку и прочая «подготовка» страницы - из конфига
            # (ключ «подготовка»). Раньше тут был хардкод под страницу «Листинг».
            _run_page_prep(page, страница)

            kind, trig_val = _parse_modal_trigger(значение)
            trigger = _find_modal_opener(page, kind, trig_val)
            if not trigger:
                print(f"   ⚠️ Триггер модалки не найден ({kind} → {trig_val!r})")
                return False

            trigger.scroll_into_view_if_needed()
            page.wait_for_timeout(800)
            trigger.click()
            print(f"   🔘 Триггер нажат ({kind})")
            page.wait_for_timeout(600)

            modal = _find_modal_root(page)
            # Модалка «открывается»: _find_modal_root() перебирает стратегии и
            # МОЛЧА съезжает на пустой fallback, если ничего внятного не нашёл -
            # делаем этот факт явным и попадающим в отчёт, а не только в print().
            _модалка_откр = "Да" if _модалка_открылась(modal) else "Нет"
            if _модалка_откр == "Нет":
                print(f"   🪟 Модалка «{название_теста}»: НЕ открылась (пустой/"
                      "невидимый контейнер после клика по триггеру)")
            _fill_modal_fields(
                modal,
                имя_теста,
                телефон_отправки,
                ПОЧТА,
                КОММЕНТАРИЙ,
                page=page,
            )

            scope = _modal_scope_for_fill(page, modal)
            try:
                _ensure_modal_consent(scope, page)
            except Exception as e:
                print(f"      ⚠️ Чекбоксы согласия: {e}")

            _click_modal_submit(scope)
            page.wait_for_timeout(3000)

            html = page.content()
            # Тот же принцип честного статуса, что и у форм: «успех» - только при
            # реальном подтверждении пользователю, а не просто из-за отсутствия
            # слова «ошибка» на странице.
            _увед_м = детект_уведомления_пользователю(page, "", "", таймаут_мс=4000)
            if response_indicates_captcha_block(html):
                статус = "ОШИБКА: КАПЧА"
            elif "ошибк" in html.lower() and any(
                x in html.lower()
                for x in ("не удалось", "не отправлен", "отклонен", "invalid")
            ):
                статус = "ОШИБКА (текст на странице после отправки)"
            elif str(_увед_м).startswith("Да"):
                статус = "УСПЕШНО (клик в браузере)"
            else:
                статус = "НЕТ ПОДТВЕРЖДЕНИЯ (форма не показала успех)"
            print(f"   ✅ {название_теста} - {статус}  ·  уведомление польз.: {_увед_м}")

            записать_в_excel(
                {
                    "тип": "PLAYWRIGHT",
                    "страница": страница,
                    "url": log_url,
                    "тип_селектора": format_modal_selector_type(str(значение)),
                    "ид": format_modal_value_for_log(str(значение)),
                    "название": название_теста,
                    "имя": имя_теста,
                    "комментарий": КОММЕНТАРИЙ if КОММЕНТАРИЙ else имя_теста,
                    "статус": статус,
                    "модалка_открылась": _модалка_откр,
                    "код": "отправлено",
                }
            )

            # Модалка «закрывается»: отдельной строкой, ПОСЛЕ основной (чтобы
            # не исказить её), крестик → Esc → клик вне модалки.
            try:
                _закр_ст, _закр_способ = _проба_закрытия_модалки(page, modal)
                записать_в_excel({
                    "тип": "ПРОВЕРКА", "страница": страница, "url": log_url,
                    "тип_селектора": "модалка", "ид": название_теста,
                    "название": f"Модалка закрывается: {название_теста}",
                    "имя": имя_теста,
                    "статус": "OK" if _закр_ст == "Да" else "Проверить",
                    "модалка_закрывается": _закр_ст,
                    "комментарий_готовый": f"способ: {_закр_способ}",
                    "код": "modal_close",
                })
                print(f"   🪟 Модалка «{название_теста}» закрывается: "
                      f"{_закр_ст} ({_закр_способ})")
            except Exception as _emc:  # noqa: BLE001
                print(f"   ⚠️ Проба закрытия модалки не удалась: {_emc}")

            return True

        except Exception as e:
            print(f"   ❌ Ошибка: {e}")
            записать_в_excel(
                {
                    "тип": "PLAYWRIGHT",
                    "страница": страница,
                    "url": log_url,
                    "тип_селектора": format_modal_selector_type(str(значение)),
                    "ид": format_modal_value_for_log(str(значение)),
                    "название": название_теста,
                    "имя": имя_теста,
                    "комментарий": КОММЕНТАРИЙ if КОММЕНТАРИЙ else имя_теста,
                    "статус": f"ОШИБКА ({e})",
                    "код": str(e),
                }
            )
            return False

    # ── Пул браузера: ОДИН Chromium/контекст на весь прогон (переиспользуется) ──
    # Раньше браузер запускался на каждую форму/сценарий - очень медленно. Теперь
    # запускаем один раз, на каждую форму открываем новую вкладку. При сбое контекста
    # (обрыв соединения и т.п.) - пересоздаём и повторяем.
    # goals - пойманные цели Яндекс.Метрики за прогон: [{"цель", "время"}].
    # Каждый reachGoal уходит запросом на mc.yandex.* с page-url=goal://домен/цель -
    # перехватываем эти запросы, шаг «проверить_цель» ищет цель в списке.
    _pw = {"play": None, "browser": None, "context": None, "goals": []}

    def _поймать_цель(request):
        try:
            body = ""
            try:
                if request.method == "POST":
                    body = request.post_data or ""
            except Exception:  # noqa: BLE001
                body = ""
            for t in _извлечь_цели_из_запроса(request.url, body):
                _pw["goals"].append({"цель": t, "время": _time.time()})
                print(f"      🎯 Метрика: зафиксирована цель «{t}»")
        except Exception:  # noqa: BLE001
            pass

    def _цели_с(ts):
        """Цели, пойманные начиная с момента ts."""
        return [g for g in _pw["goals"] if g["время"] >= ts]

    def _отчёт_сработавших_целей(ts, страница, log_url, seen, контекст=""):
        """АВТООПРЕДЕЛЕНИЕ: пишет в отчёт строку по КАЖДОЙ цели Метрики,
        сработавшей с момента ts (кроме уже записанных - через seen). Так видно,
        какая цель реально фиксируется на форме/кнопке, даже если в конфиге цель
        не задана. Не нужно знать имена целей заранее - прогон их показывает."""
        if not проверять_цели:
            return False          # цели - в «Проверке целей», не в формах
        _есть = False
        for g in _цели_с(ts):
            gнэйм = g["цель"]
            if gнэйм in seen:
                continue
            seen.add(gнэйм)
            _есть = True
            _подпись = f"Сработала цель: {gнэйм}" + (f" - {контекст}" if контекст else "")
            записать_в_excel(
                {
                    "тип": "ЦЕЛЬ (Метрика)",
                    "страница": страница,
                    "url": log_url,
                    "тип_селектора": "цель",
                    "ид": gнэйм,
                    "название": _подпись,
                    "имя": имя_теста_из_конфига(страница, "цель", _подпись,
                                                название_контекста=_подпись),
                    "комментарий": gнэйм,
                    "статус": "ЗАФИКСИРОВАНА (Метрика)",
                    "код": "ym-auto",
                }
            )
            print(f"   🎯 Сработала цель: {gнэйм}" + (f" - {контекст}" if контекст else ""))
        return _есть

    def _shared_context():
        h = _pw
        if h["context"] is None:
            if h["play"] is None:
                h["play"] = sync_playwright().start()
            # Маскируем автоматизацию: иначе часть сайтов (Bitrix) не навешивает свой
            # JS-обработчик отправки на «робота», форма уходит обычным POST и сервер
            # отвечает «Доступ запрещён». Флаг + init-скрипт убирают признак webdriver.
            _launch_kw = dict(
                headless=headless,
                args=["--disable-blink-features=AutomationControlled"],
            )
            _prx = _playwright_proxy_from_env()
            if _prx:
                _launch_kw["proxy"] = _prx
                print(f"🔌 Формы идут через прокси: {_prx['server']}")
            h["browser"] = h["play"].chromium.launch(**_launch_kw)
            h["context"] = h["browser"].new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                # Обычное окно. Перекрытые кнопки (баг вёрстки на части
                # СНГ-доменов) решает принудительный клик в отправке формы,
                # огромное окно для этого не нужно.
                viewport={"width": 1366, "height": 768},
                locale="ru-RU",
            )
            try:
                h["context"].add_init_script(
                    "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
                )
            except Exception:
                pass
            # Цели Метрики ловим ТОЛЬКО когда движок форм вызван «Проверкой целей»
            # (проверять_цели=True). В обычной «Проверке форм» слушатель не вешаем -
            # формы проверяем без целей.
            if проверять_цели:
                try:
                    h["context"].on("request", _поймать_цель)
                except Exception:
                    pass
        return h["context"]

    def _drop_browser():
        h = _pw
        try:
            if h["browser"]:
                h["browser"].close()
        except Exception:
            pass
        h["browser"] = None
        h["context"] = None

    def _close_browser_pool():
        _drop_browser()
        try:
            if _pw["play"]:
                _pw["play"].stop()
        except Exception:
            pass
        _pw["play"] = None

    def _open_page():
        """Новая вкладка в общем контексте; если контекст умер - пересоздать и повторить."""
        try:
            return _shared_context().new_page()
        except Exception:
            _drop_browser()
            return _shared_context().new_page()

    def _with_browser_page(callback):
        page = _open_page()
        try:
            return callback(page)
        finally:
            try:
                page.close()
            except Exception:
                pass

    def отправить_форму_через_playwright(url, форма_config, название):
        """
        Отправка формы через Chromium: как у живого пользователя (JS, hash, AJAX Bitrix).
        Ручная отправка с почтой менеджера совпадает с этим путём; requests без JS - нет.
        """
        страница = определить_страницу(url)
        имя_теста = имя_теста_из_конфига(
            страница,
            _значение_формы_для_имени(форма_config),
            название,
        )

        use_text = "text" in форма_config and str(форма_config.get("text", "")).strip()
        sel = _playwright_form_css_selector(форма_config)
        if not use_text and not sel:
            записать_в_excel(
                {
                    "тип": "PLAYWRIGHT-FORM",
                    "страница": страница,
                    "url": url,
                    "тип_селектора": format_form_selector_type(форма_config),
                    "ид": format_form_config_for_log(форма_config),
                    "название": название,
                    "имя": имя_теста,
                    "комментарий": КОММЕНТАРИЙ if КОММЕНТАРИЙ else имя_теста,
                    "статус": "НЕТ СЕЛЕКТОРА ФОРМЫ (id/class/data-source/css/name/text)",
                    "код": "",
                }
            )
            return False

        print(f"   🌐 Форма через браузер (Playwright): {format_form_config_for_log(форма_config)}")

        def _run_direct(page):
            def _переоткрыть():
                # Свежая форма для повторных проб: просто перезагружаем страницу
                # (прямая форма уже на ней, шагов открытия нет).
                try:
                    _goto_with_retry(page, url)
                    page.wait_for_timeout(1500)
                    _run_page_prep(page, страница)
                    return True
                except Exception:  # noqa: BLE001
                    return False
            return _form_fill_submit_on_page(
                page, url, страница, форма_config, название, initial_url=url,
                переоткрыть_fn=_переоткрыть,
            )
        return _with_browser_page(_run_direct)

    def проверить_кнопку_через_playwright(url, значение, название_теста):
        страница = определить_страницу(url)
        return _with_browser_page(
            lambda page: _modal_flow_on_page(
                page, url, страница, значение, название_теста, initial_url=url
            )
        )

    def run_scenario_playwright(
        base_url: str, шаги: list, *, название_сценария: str = "", тип_блока: str = ""
    ):
        """Один сеанс браузера: цепочка шагов (пауза, переход, клик, форма, модалка).
        тип_блока - тип страницы («Оформление…» = заказ): по нему форм-шаг решает,
        проверять двойную отправку БЕЗОПАСНО (без реального второго клика)."""
        if not шаги:
            print("   ⚠️ Сценарий пуст - пропуск.")
            return
        if ФОРМЫ_ЧЕРЕЗ_REQUESTS:
            print(
                "   ⚠️ Сценарии с «шаги» требуют Playwright. Отключите ФОРМЫ_ЧЕРЕЗ_REQUESTS в config "
                "или уберите блок «шаги» у этой страницы."
            )
            return

        страница = определить_страницу(base_url)
        cap = (название_сценария or "").strip()
        if cap:
            print(f"   📜 Сценарий «{cap}»: {len(шаги)} шаг(ов), базовый URL: {base_url}")
        else:
            print(f"   📜 Сценарий: {len(шаги)} шаг(ов), базовый URL: {base_url}")

        if True:  # общий браузер (пул): новая вкладка вместо запуска нового Chromium
            page = _open_page()
            _тек_шаг_инфо = ""
            _scn_t0 = _time.time()   # цели Метрики считаем с начала сценария
            _scn_цели_seen = set()   # уже записанные цели этого сценария (без дублей)
            try:
                _ensure_scenario_page_loaded(page, base_url)
                _run_page_prep(page, страница)

                for i, step in enumerate(шаги):
                    if stop_flag and stop_flag():
                        print("\n⏸️ Тест остановлен пользователем")
                        return
                    if not isinstance(step, dict):
                        print(f"   ⚠️ Шаг {i + 1}: ожидался словарь, пропуск.")
                        continue
                    if not cfg_enabled(step.get("включено", True)):
                        print(f"   ⏭️ Шаг {i + 1}: отключён в конфиге - пропуск.")
                        continue

                    act = _нормализовать_действие_шага(step)
                    # Запоминаем текущий шаг - чтобы при падении сказать, на чём встали.
                    _шаг_цель = (step.get("css") or step.get("selector")
                                 or step.get("url") or step.get("href") or "")
                    _тек_шаг_инфо = f"шаг {i + 1} «{act}»" + (f" {_шаг_цель}" if _шаг_цель else "")

                    if act == "пауза":
                        try:
                            ms = int(step.get("мс") or step.get("ms") or 500)
                        except (TypeError, ValueError):
                            ms = 500
                        print(f"   ⏳ Шаг {i + 1}: пауза {ms} мс")
                        page.wait_for_timeout(ms)

                    elif act == "перейти":
                        href = step.get("url") or step.get("href")
                        if not href:
                            print(f"   ⚠️ Шаг {i + 1}: «перейти» без url/href - пропуск.")
                            continue
                        dest = _resolve_scenario_url(page, base_url, href)
                        print(f"   🔗 Шаг {i + 1}: переход → {dest}")
                        _goto_with_retry(page, dest)
                        page.wait_for_timeout(1500)

                    elif act == "клик":
                        css = step.get("css") or step.get("selector")
                        if not css:
                            print(
                                f"   ⚠️ Шаг {i + 1}: «клик» без css/selector - пропуск."
                            )
                            continue
                        raw = str(css).strip()
                        css_norm = _normalize_scenario_click_css_selector(raw)
                        if css_norm != raw:
                            print(
                                f"   ↪ Шаг {i + 1}: уточнён селектор клика {raw!r} → {css_norm!r}"
                            )
                        _необяз = bool(
                            step.get("необязательно") or step.get("optional")
                        )
                        print(
                            f"   🖱️ Шаг {i + 1}: клик {css_norm!r}"
                            + (" (необязательный)" if _необяз else "")
                        )
                        try:
                            _to = 6000 if _необяз else 15000
                            page.locator(css_norm).first.click(timeout=_to)
                        except Exception as _e_click:  # noqa: BLE001
                            if _необяз:
                                print(
                                    f"   ↳ необязательный клик пропущен (элемент не найден): {css_norm!r}"
                                )
                            else:
                                raise
                        page.wait_for_timeout(400)

                        # Промежуточная проверка ошибок (пункт «Обработка ошибок»
                        # на многошаговом заказе): если по клику «Далее»/«Оформить
                        # заказ» форма пропустила дальше, но на странице ОСТАЛАСЬ
                        # ошибка валидации/оформления - это дефект (клиент идёт
                        # дальше, но заказ не оформить). Только под флагом шага
                        # «проверить_ошибку» - чтобы не трогать обычные клики.
                        if cfg_enabled(step.get("проверить_ошибку")):
                            try:
                                _pe_html = page.content()
                            except Exception:  # noqa: BLE001
                                _pe_html = ""
                            _pe_err = response_indicates_form_error(_pe_html)
                            if _pe_err:
                                записать_в_excel({
                                    "тип": "ПРОВЕРКА", "страница": страница,
                                    "url": page.url or base_url,
                                    "тип_селектора": "сценарий",
                                    "ид": cap or f"шаг {i + 1}",
                                    "название": f"Обработка ошибок: {cap}",
                                    "статус": "Проверить",
                                    "обработка_ошибок": "ошибка на шаге «Далее»",
                                    "комментарий_готовый": (
                                        "После клика «Далее»/«Оформить заказ» форма "
                                        "пропустила дальше, но на странице осталась "
                                        f"ошибка: {_pe_err}. Пользователь идёт "
                                        "дальше, но заказ не оформить."),
                                    "код": "step_error",
                                })
                                print(f"   🚨 Ошибка на шаге «Далее» «{cap}»: {_pe_err}")

                    elif act == "наведение":
                        css = step.get("css") or step.get("selector")
                        if not css:
                            print(
                                f"   ⚠️ Шаг {i + 1}: «наведение» без css/selector - пропуск."
                            )
                            continue
                        raw = str(css).strip()
                        css_norm = _normalize_scenario_click_css_selector(raw)
                        if css_norm != raw:
                            print(
                                f"   ↪ Шаг {i + 1}: уточнён селектор наведения {raw!r} → {css_norm!r}"
                            )
                        loc = page.locator(css_norm).first
                        loc.scroll_into_view_if_needed()
                        print(f"   ↗ Шаг {i + 1}: наведение {css_norm!r}")
                        loc.hover(timeout=15000)
                        page.wait_for_timeout(400)

                    elif act in ("заполнить_по_метке", "fill_by_label", "поле_по_метке"):
                        # Заполнение поля по ВИДИМОЙ подписи (label) - для форм, где
                        # имена полей рисуются в JS (bx-soa): get_by_label, затем
                        # запасной вариант по placeholder. Не падаем, если поля нет.
                        метка = str(step.get("метка") or step.get("label") or "").strip()
                        _tok = step.get("значение") or step.get("value") or ""
                        _val = _resolve_form_field_token(
                            _tok, имя_теста="", телефон=телефон_отправки,
                            почта=ПОЧТА, имя=ИМЯ, комментарий=КОММЕНТАРИЙ, город=ГОРОД,
                        )
                        if not метка or not str(_val).strip():
                            print(f"   ⚠️ Шаг {i + 1}: «заполнить_по_метке» без метки/значения - пропуск.")
                            continue
                        _filled = False
                        _esc = метка.replace('"', '\\"')
                        for _getter in (
                            lambda: page.get_by_label(метка, exact=False),
                            lambda: page.locator(f'input[placeholder*="{_esc}"], textarea[placeholder*="{_esc}"]'),
                        ):
                            try:
                                _loc = _getter().first
                                _loc.wait_for(state="visible", timeout=7000)
                                _loc.scroll_into_view_if_needed()
                                _loc.fill(str(_val), force=True)
                                _filled = True
                                break
                            except Exception:  # noqa: BLE001
                                continue
                        print(
                            f"   ✏️ Шаг {i + 1}: поле по метке «{метка}» - "
                            f"{'заполнено' if _filled else 'НЕ найдено'}"
                        )

                    elif act == "форма":
                        form_cfg = {
                            k: v
                            for k, v in step.items()
                            if k not in _META_STEP_KEYS
                        }
                        nm = form_cfg.get("название") or f"форма шаг {i + 1}"
                        form_cfg = {**form_cfg, "название": nm}
                        nav = step.get("url") or step.get("href")
                        initial = _resolve_scenario_url(page, base_url, nav) if nav else None
                        print(
                            f"   🌐 Шаг {i + 1}: форма «{nm}»"
                            + (f" (переход: {initial})" if initial else "")
                        )

                        # Колбэк «заново открыть эту форму» для повторных проб
                        # (серверная валидация/лимит). Перезагружаем базу сценария
                        # и повторяем ШАГИ ОТКРЫТИЯ (клики/паузы) до формы. Слайс
                        # шагов фиксируем через дефолт-аргумент (i - переменная цикла).
                        def _переоткрыть_сценарий(_откр=list(шаги[:i]), _init=initial):
                            try:
                                _goto_with_retry(page, base_url)
                                page.wait_for_timeout(800)
                                _run_page_prep(page, страница)
                                for _st in _откр:
                                    _replay_open_step(page, _st, base_url)
                                if _init:
                                    _goto_with_retry(page, _init)
                                    page.wait_for_timeout(1200)
                                return True
                            except Exception:  # noqa: BLE001
                                return False

                        # Переоткрытие уводит страницу перезагрузкой - безопасно
                        # ТОЛЬКО если «форма» последний шаг сценария (после неё
                        # других шагов нет). Иначе не передаём колбэк (проба
                        # пойдёт по старой форме, как раньше).
                        _reopen_cb = (_переоткрыть_сценарий
                                      if i == len(шаги) - 1 else None)

                        _form_fill_submit_on_page(
                            page,
                            base_url,
                            страница,
                            form_cfg,
                            nm,
                            initial_url=initial,
                            название_контекста=_scenario_placeholder_title(
                                step, cap
                            ),
                            цели_seen=_scn_цели_seen,
                            безопасная_отправка=str(тип_блока).startswith("Оформление"),
                            переоткрыть_fn=_reopen_cb,
                        )

                    elif act == "модалка":
                        val = step.get("значение") or step.get("value")
                        nm = (
                            step.get("название_теста")
                            or step.get("название")
                            or f"модалка шаг {i + 1}"
                        )
                        if val is None or str(val).strip() == "":
                            print(
                                f"   ⚠️ Шаг {i + 1}: «модалка» без значения - пропуск."
                            )
                            continue
                        nav = step.get("url") or step.get("href")
                        initial = _resolve_scenario_url(page, base_url, nav) if nav else None
                        print(
                            f"   💬 Шаг {i + 1}: модалка «{nm}»"
                            + (f" (переход: {initial})" if initial else "")
                        )
                        _modal_flow_on_page(
                            page,
                            base_url,
                            страница,
                            val,
                            nm,
                            initial_url=initial,
                            название_контекста=_scenario_placeholder_title(
                                step, cap
                            ),
                        )

                    elif act in ("проверить", "итог", "check"):
                        nm = (
                            step.get("название")
                            or step.get("название_теста")
                            or cap
                            or f"проверка шаг {i + 1}"
                        )
                        успех = str(step.get("успех_текст") or step.get("success_text") or "")
                        ошибка = str(step.get("ошибка_текст") or step.get("error_text") or "")
                        # Подтверждение приходит ajax-ом с разной скоростью
                        # (на части доменов дольше): опрашиваем страницу, пока
                        # не появится признак успеха/ошибки или не выйдет время.
                        try:
                            _ждать_мс = int(step.get("ожидание_мс") or 12000)
                        except (TypeError, ValueError):
                            _ждать_мс = 12000
                        _t0 = _time.time()
                        while True:
                            html = page.content()
                            low = html.lower()
                            _chk_err = response_indicates_form_error(html)
                            if (not успех
                                    or response_indicates_captcha_block(html)
                                    or _chk_err
                                    or (ошибка and ошибка.lower() in low)
                                    or успех.lower() in low
                                    or (_time.time() - _t0) * 1000 >= _ждать_мс):
                                break
                            page.wait_for_timeout(700)
                        log_url = page.url or base_url
                        _коммент_готовый = None
                        if response_indicates_captcha_block(html):
                            статус = "ОШИБКА: КАПЧА"
                        elif _chk_err:
                            # Явная ошибка на странице важнее «признака успеха»: на bx-soa
                            # текст «Заказ сформирован» бывает в скрытом шаблоне и даёт
                            # ложный успех, хотя заказ не оформлен.
                            статус = "ОШИБКА"
                            _коммент_готовый = _chk_err
                        elif ошибка and ошибка.lower() in low:
                            статус = "ОШИБКА (сообщение на странице)"
                        elif успех and успех.lower() in low:
                            статус = "УСПЕШНО (подтверждение на странице)"
                        elif успех:
                            статус = f"ОШИБКА (нет признака успеха «{успех}» на странице)"
                        else:
                            статус = "УСПЕШНО (страница открыта)"
                        имя_лог = имя_теста_из_конфига(
                            страница, "проверка", nm, название_контекста=nm
                        )
                        записать_в_excel(
                            {
                                "тип": "PLAYWRIGHT",
                                "страница": страница,
                                "url": log_url,
                                "тип_селектора": "сценарий",
                                "ид": nm,
                                "название": nm,
                                "имя": имя_лог,
                                "комментарий": КОММЕНТАРИЙ if КОММЕНТАРИЙ else имя_лог,
                                "комментарий_готовый": _коммент_готовый,
                                "статус": статус,
                                "код": "browser",
                            }
                        )
                        print(f"   🔎 Шаг {i + 1}: проверка «{nm}» - {статус}")
                        # Итоговый URL сценария - чтобы «Проверка целей» могла
                        # подтвердить url-цели (оформленный заказ / страница «спасибо»),
                        # на которые обычный прогон целей не попадает.
                        print(f"   🔗 URL сценария: {log_url}")
                        # Пункт 2.9: финальный шаг оформления заказа помечается
                        # «заказ»: True. Если он прошёл успешно - фиксируем заказ,
                        # чтобы forms_run потом проверил письмо-подтверждение покупателю.
                        if cfg_enabled(step.get("заказ", False)) and str(статус).startswith("УСПЕШНО"):
                            record_placed_order({
                                "город": город,
                                "почта": ПОЧТА,
                                "домен": log_url,
                                "название": nm,
                                "ts": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
                            })
                            print(f"   🧾 Заказ зафиксирован для проверки письма покупателю "
                                  f"({ПОЧТА}).")

                    elif act in ("проверить_цель", "check_goal") and not проверять_цели:
                        # Цели проверяет «Проверка целей», а не «Проверка форм» -
                        # шаг пропускаем (формы всё равно кликаются/отправляются).
                        pass
                    elif act in ("проверить_цель", "check_goal"):
                        # Проверка цели Яндекс.Метрики: ждём, пока перехватчик
                        # поймает reachGoal с нужным именем (цели уходят ajax-ом,
                        # поэтому опрашиваем список до «ожидание_мс»).
                        nm = (
                            step.get("название")
                            or step.get("название_теста")
                            or f"цель шаг {i + 1}"
                        )
                        цель = str(step.get("цель") or step.get("goal") or "").strip()
                        try:
                            _ждать_мс = int(step.get("ожидание_мс") or 10000)
                        except (TypeError, ValueError):
                            _ждать_мс = 10000
                        _t0g = _time.time()
                        _найдена = False
                        while цель:
                            _найдена = any(g["цель"] == цель for g in _цели_с(_scn_t0))
                            if _найдена or (_time.time() - _t0g) * 1000 >= _ждать_мс:
                                break
                            page.wait_for_timeout(500)
                        _коммент_готовый = None
                        if not цель:
                            статус = "НЕ СРАБОТАЛА"
                            _коммент_готовый = "В шаге «проверить_цель» не задано имя цели"
                        elif _найдена:
                            статус = "СРАБОТАЛА"
                        else:
                            статус = "НЕ СРАБОТАЛА"
                            _коммент_готовый = (str(step.get("цель_причина") or "").strip()
                                                or f"Цель «{цель}» не зафиксирована Метрикой")
                        имя_лог = имя_теста_из_конфига(
                            страница, "цель", nm, название_контекста=nm
                        )
                        записать_в_excel(
                            {
                                "тип": "ЦЕЛЬ (Метрика)",
                                "страница": страница,
                                "url": page.url or base_url,
                                "тип_селектора": "цель",
                                "ид": цель,
                                "название": nm,
                                "имя": имя_лог,
                                "комментарий": _коммент_готовый or "",
                                "комментарий_готовый": _коммент_готовый,
                                "статус": статус,
                                "код": "ym",
                            }
                        )
                        print(f"   🎯 Шаг {i + 1}: цель «{цель}» - {статус}")

                    elif act in ("проверить_корзину", "check_cart"):
                        # Проверка: реально ли товар попал в корзину. На СНГ-доменах
                        # кнопка «Добавить в корзину» - баг разработчиков: клик ничего
                        # не делает, корзина остаётся пустой. Признак «товар в корзине» -
                        # наличие кнопки оформления (Bitrix рисует её только при товаре).
                        # Если корзина пуста - пишем понятный комментарий и МЯГКО
                        # завершаем сценарий (без «прервался на шаг…»).
                        nm = (
                            step.get("название")
                            or step.get("название_теста")
                            or cap
                            or f"корзина шаг {i + 1}"
                        )
                        товар_css = step.get("признак_товар_css") or step.get("css")
                        коммент = str(
                            step.get("комментарий_провал")
                            or "Кнопка «Добавить в корзину» не работает "
                               "(товар не кладётся в корзину)"
                        )
                        try:
                            ms_wait = int(step.get("ожидание_мс") or step.get("мс") or 6000)
                        except (TypeError, ValueError):
                            ms_wait = 6000
                        есть_товар = True
                        if товар_css:
                            sel = _normalize_scenario_click_css_selector(
                                str(товар_css).strip()
                            )
                            try:
                                page.locator(sel).first.wait_for(
                                    state="visible", timeout=ms_wait
                                )
                                есть_товар = True
                            except Exception:  # noqa: BLE001
                                есть_товар = False
                        if есть_товар:
                            print(
                                f"   🛒 Шаг {i + 1}: корзина не пуста - продолжаем оформление."
                            )
                        else:
                            log_url = page.url or base_url
                            имя_лог = имя_теста_из_конфига(
                                страница, "корзина", nm, название_контекста=nm
                            )
                            записать_в_excel(
                                {
                                    "тип": "PLAYWRIGHT",
                                    "страница": страница,
                                    "url": log_url,
                                    "тип_селектора": "сценарий",
                                    "ид": nm,
                                    "название": nm,
                                    "имя": имя_лог,
                                    "комментарий_готовый": коммент,
                                    "статус": "ОШИБКА (кнопка корзины не работает)",
                                    "код": "cart-broken",
                                }
                            )
                            print(f"   🛒 Шаг {i + 1}: корзина пуста - {коммент}")
                            return

                    else:
                        print(
                            f"   ⚠️ Шаг {i + 1}: неизвестное действие {act!r} "
                            f"(ожидалось: пауза, перейти, клик, наведение, форма, модалка, проверить, проверить_корзину)."
                        )

                def _step_scenario_enabled(s):
                    return isinstance(s, dict) and cfg_enabled(s.get("включено", True))

                had_form_or_modal = any(
                    _нормализовать_действие_шага(s)
                    in ("форма", "модалка", "проверить", "итог", "check",
                        "проверить_цель", "check_goal",
                        "заполнить_по_метке", "fill_by_label", "поле_по_метке")
                    for s in шаги
                    if _step_scenario_enabled(s)
                )
                had_click_only_chain = any(
                    _нормализовать_действие_шага(s) == "клик"
                    for s in шаги
                    if _step_scenario_enabled(s)
                )
                if cap and not had_form_or_modal and had_click_only_chain:
                    log_url = page.url or base_url
                    имя_лог = имя_теста_из_конфига(
                        страница,
                        "клик",
                        cap,
                        название_контекста=cap,
                    )
                    записать_в_excel(
                        {
                            "тип": "PLAYWRIGHT",
                            "страница": страница,
                            "url": log_url,
                            "тип_селектора": "сценарий",
                            "ид": cap,
                            "название": cap,
                            "имя": имя_лог,
                            "комментарий": КОММЕНТАРИЙ if КОММЕНТАРИЙ else имя_лог,
                            "статус": "УСПЕШНО (только клик; tel/mailto - проверьте ОС/браузер)",
                            "код": "click-only",
                        }
                    )
                    print(f"   ✅ «{cap}» - строка в Excel (клик без формы)")

            except Exception as _e_step:  # noqa: BLE001
                # Прокидываем наверх с указанием, на каком шаге упали.
                where = _тек_шаг_инфо or "одном из шагов"
                raise RuntimeError(f"прервался на {where}") from _e_step
            finally:
                try:
                    page.close()
                except Exception:
                    pass

    инициализировать_excel()

    print(f"\n📅 {ДАТА} {ВРЕМЯ}")
    print(f"📁 Лог: {EXCEL_ФАЙЛ}\n")
    print(f"📞 Телефон (в конфиге): {ТЕЛЕФОН}")
    print(f"📞 Телефон для отправки (цифры): {телефон_отправки or '(пусто)'}")
    if not телефон_отправки:
        print(
            "⚠️ ВНИМАНИЕ: ТЕЛЕФОН в config пустой - поля телефона не заполняются, "
            "заявка часто не принимается. Укажите номер в конфиге и сохраните."
        )
    print(f"✉️ Почта: {ПОЧТА}")
    print(f"👤 Имя в config: {ИМЯ if ИМЯ else '(пусто)'}")
    print(
        f"📝 Формат имени теста (ИМЯ задано): {ФОРМАТ_ИМЕНИ_ТЕСТА!r}; "
        f"если ИМЯ пусто: {ФОРМАТ_ИМЕНИ_АВТО_ПО_УМОЛЧАНИЮ!r}"
    )
    print(f"💬 Комментарий: {КОММЕНТАРИЙ if КОММЕНТАРИЙ else '(имя теста)'}")
    print(
        f"📋 Формы на странице: "
        f"{'requests (быстро, без JS)' if ФОРМЫ_ЧЕРЕЗ_REQUESTS else 'Playwright (как в браузере - заявки на почту)'}\n"
    )

    for страница in СТРАНИЦЫ_ДЛЯ_ПРОВЕРКИ:
        if stop_flag and stop_flag():
            print("\n⏸️ Тест остановлен пользователем")
            return

        тип_страницы = страница["тип"]
        url = СТРАНИЦЫ[тип_страницы]

        # Региональные страницы (форма есть только в конкретном городе/стране):
        # «только_города» ограничивает прогон списком городов. Если идёт прогон по
        # городу не из списка - страницу пропускаем (иначе она тянулась бы в КАЖДЫЙ
        # город: Хабаровск/СНГ попадали в отчёт там, где их не выбирали).
        только = страница.get("только_города")
        if только and (город or "").strip() not in только:
            print(f"\n📄 Страница: {тип_страницы} - пропущена (только для: {', '.join(только)})")
            continue

        кроме = страница.get("кроме_городов")
        if кроме and (город or "").strip() in кроме:
            print(f"\n📄 Страница: {тип_страницы} - пропущена для города {город} (есть отдельный блок)")
            continue

        if not cfg_enabled(страница.get("включено", True)):
            print(f"\n{'='*50}")
            print(f"📄 Страница: {тип_страницы} - не в прогоне (настройки сохранены, шаги не выполняются)")
            print(f"{'='*50}")
            continue

        ф_вкл = cfg_enabled(страница.get("формы_включены", True))
        м_вкл = cfg_enabled(страница.get("модалки_включены", True))
        с_вкл = cfg_enabled(страница.get("сценарий_включен", True))
        название_сценария = str(страница.get("название_сценария") or "").strip()
        сценарии_блоки = scenario_blocks_from_page(страница)
        есть_шаги = any((b.get("шаги") or []) for b in сценарии_блоки)

        print(f"\n{'='*50}")
        print(f"📄 Страница: {тип_страницы}")
        print(f"{'='*50}")

        if есть_шаги and с_вкл:
            for sc in сценарии_блоки:
                steps = sc.get("шаги") or []
                if not steps:
                    continue
                if not cfg_enabled(sc.get("включено"), True):
                    nm = sc.get("название") or ""
                    print(
                        f"   ⏭️ Сценарий пропущен (отключён в конфиге): {nm!r}"
                    )
                    continue
                cap = str(sc.get("название") or "").strip() or название_сценария
                if not _форма_выбрана(cap):
                    continue
                if _нет_в_текущем_городе(sc):
                    _лог_форма_отсутствует(тип_страницы, url, sc, cap)
                    continue
                _отметить_форму(cap)
                # До 2 попыток: СНГ-домены часто рвут соединение (анти-бот),
                # повтор обычно проходит.
                _scn_err = None
                for _попытка in (1, 2):
                    try:
                        run_scenario_playwright(url, steps, название_сценария=cap,
                                                тип_блока=тип_страницы)
                        _scn_err = None
                        break
                    except Exception as _e:  # noqa: BLE001
                        _scn_err = _e
                        if _попытка == 1:
                            print(f"   ↻ Сценарий «{cap}» упал ({str(_e)[:80]}), повтор…")
                if _scn_err is not None:
                    # Один упавший сценарий НЕ должен ронять весь прогон -
                    # пишем ошибку в лог и идём к следующей форме.
                    print(f"   ❌ Сценарий «{cap}» прерван ошибкой: {_scn_err}")
                    try:
                        записать_в_excel(
                            {
                                "тип": "PLAYWRIGHT",
                                "страница": тип_страницы,
                                "url": url,
                                "тип_селектора": "сценарий",
                                "ид": cap,
                                "название": cap,
                                "имя": cap,
                                "комментарий": КОММЕНТАРИЙ if КОММЕНТАРИЙ else cap,
                                "комментарий_готовый": f"Сценарий {str(_scn_err)[:200]}"
                                if str(_scn_err).strip() else "Сценарий прервался на одном из шагов",
                                "статус": "ОШИБКА (сценарий прерван)",
                                "код": str(_scn_err)[:300],
                            }
                        )
                    except Exception:  # noqa: BLE001
                        pass
            continue

        if есть_шаги and not с_вкл:
            print(
                "   ⏭️ Сценарий (шаги) отключён для страницы - шаги не выполняются."
            )

        if not ф_вкл:
            print("   ⏭️ Формы для страницы отключены - блок форм пропущен.")
        if not м_вкл:
            print("   ⏭️ Модалки для страницы отключены - блок модалок пропущен.")

        for форма in страница.get("формы") or []:
            if stop_flag and stop_flag():
                print("\n⏸️ Тест остановлен пользователем")
                return
            if not ф_вкл:
                continue
            if not cfg_enabled(форма.get("включено", True)):
                nm = форма.get("название", "")
                print(f"   ⏭️ Форма пропущена (отключена в конфиге): {nm!r}")
                continue
            if not _форма_выбрана(форма.get("название", "")):
                continue
            if _нет_в_текущем_городе(форма):
                _лог_форма_отсутствует(тип_страницы, url, форма, форма.get("название", "?"))
                continue
            _отметить_форму(форма.get("название", ""))
            # До 2 попыток: страница/форма иногда не доходит из-за обрыва соединения
            # (анти-бот СНГ) - повтор обычно проходит.
            _frm_err = None
            for _попытка in (1, 2):
                try:
                    if ФОРМЫ_ЧЕРЕЗ_REQUESTS:
                        отправить_через_requests(url, форма, форма["название"])
                    else:
                        # По умолчанию - авто: по коду где можно, иначе браузер.
                        _проверить_форму_авто(url, форма, форма["название"])
                    _frm_err = None
                    break
                except Exception as _e:  # noqa: BLE001
                    _frm_err = _e
                    if _попытка == 1:
                        print(f"   ↻ Форма «{форма.get('название','?')}» упала ({str(_e)[:70]}), повтор…")
            if _frm_err is not None:
                # Сбой одной формы НЕ должен ронять прогон - пишем ошибку и идём дальше.
                print(f"   ❌ Форма «{форма.get('название','?')}» прервана ошибкой: {_frm_err}")
                try:
                    записать_в_excel(
                        {
                            "тип": "PLAYWRIGHT-FORM",
                            "страница": тип_страницы,
                            "url": url,
                            "тип_селектора": format_form_selector_type(форма),
                            "ид": format_form_config_for_log(форма),
                            "название": форма.get("название", "?"),
                            "имя": форма.get("название", "?"),
                            "комментарий": КОММЕНТАРИЙ if КОММЕНТАРИЙ else форма.get("название", "?"),
                            "статус": "ОШИБКА (форма прервана)",
                            "код": str(_frm_err)[:300],
                        }
                    )
                except Exception:  # noqa: BLE001
                    pass

        for модалка in страница.get("модалки") or []:
            if stop_flag and stop_flag():
                print("\n⏸️ Тест остановлен пользователем")
                return
            if not м_вкл:
                continue
            if not cfg_enabled(модалка.get("включено", True)):
                nm = модалка.get("название_теста", "")
                print(f"   ⏭️ Модалка пропущена (отключена в конфиге): {nm!r}")
                continue
            if not _форма_выбрана(модалка.get("название_теста", "")):
                continue
            _отметить_форму(модалка.get("название_теста", ""))
            проверить_кнопку_через_playwright(
                url, модалка["значение"], модалка["название_теста"]
            )

    # Закрываем общий браузер прогона (пул).
    _close_browser_pool()

    _spent = int(_time.time() - _run_t0)
    _spent_mmss = f"{_spent // 60}:{_spent % 60:02d}"

    try:
        write_summary_sheet(EXCEL_ФАЙЛ, время_прогона=_spent_mmss)
        print("   🧾 Сводка собрана (лист «Сводка»)")
    except Exception as _e:  # noqa: BLE001
        print(f"   ⚠️ Не удалось собрать сводку: {_e}")

    print(f"\n✅ Готово за {_spent_mmss} (мин:сек). Результаты в {EXCEL_ФАЙЛ}")

    # Файл НЕ открываем автоматически: открытый в Excel лог блокируется,
    # и следующий прогон не может его перезаписать (а ещё не даёт скачать копию).
    # В веб-версии результат виден в таблице, рядом есть кнопка «Скачать».

    if stop_flag is None:
        input("\nНажмите Enter для выхода...")


if __name__ == "__main__":
    run_test()
