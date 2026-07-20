"""
Проверка админки Bitrix «Уведомления с форм» (модуль pixana_forms_list).

Идея: после отправки форм тест заходит в админку, открывает список заявок за
сегодня и проверяет, что НАША заявка реально там появилась (форма долетела до
бэкенда, а не только показала «спасибо»).

Сопоставление - по «Тип формы» + времени отправки (без метки в заявке):
админка пишет время до секунды, а движок знает, когда отправил форму.
Логин/пароль берём из локального файла admin.local.json (в git не хранится).
"""
import json
import re
from datetime import datetime
from pathlib import Path


def построить_url_списка(домен: str, дата=None) -> str:
    """URL списка заявок «Уведомления с форм».

    дата задана → фильтр за один день (быстро). дата=None → без фильтра по дате,
    но со SHOWALL (все строки на одной странице, без пагинации) - используется как
    надёжный фолбэк, если фильтр по дате на сервере повёл себя иначе (тогда день
    отсекаем уже в коде). SHOWALL добавляем всегда, чтобы не терять строки на 2-й
    странице (важно для СНГ-админки, куда падают заявки многих городов)."""
    домен = домен.rstrip("/")
    base = f"{домен}/bitrix/admin/pixana_forms_list.php?lang=ru&form_type=all&SHOWALL_1=1"
    if дата is not None:
        d = дата.strftime("%Y-%m-%d")
        base += f"&find_date_from={d}&find_date_to={d}"
    return base


def _текст(html: str) -> str:
    html = html.replace("&nbsp;", " ")
    html = re.sub(r"<[^>]+>", " ", html)
    return re.sub(r"\s+", " ", html).strip()


def _это_форма_входа(html: str) -> bool:
    """True, если полученная страница - форма логина Bitrix (поля USER_LOGIN и
    USER_PASSWORD, при этом таблицы списка заявок нет). По ней и определяем, что
    вход не удался - надёжнее, чем ловить исчезновение поля пароля по таймауту."""
    h = html or ""
    return ("USER_PASSWORD" in h and "USER_LOGIN" in h
            and "adm-list-table" not in h.lower())


# Заголовки колонок админки → наши поля (нестрого, по нормализованному тексту).
_HEADER_FIELD = {
    "id": "id", "ид": "id", "№": "id",
    "дата": "дата_время", "датавремя": "дата_время",
    "типформы": "тип_формы", "форма": "тип_формы",
    "город": "город",
    "имя": "имя", "фио": "имя", "название": "имя",
    "телефон": "телефон", "тел": "телефон", "phone": "телефон",
    "email": "email", "emailадрес": "email", "почта": "email", "eмаил": "email",
    "файл": "файл",
    "вседанные": "все_данные", "данные": "все_данные",
}


def _ячейки_строки(tr_html: str) -> list:
    """Тексты всех ячеек строки (<td>/<th>), в порядке следования."""
    cells = re.findall(r"<t[dh]\b[^>]*>(.*?)</t[dh]>", tr_html, re.S | re.I)
    return [_текст(c) for c in cells]


def разобрать_заявки(html: str) -> list:
    """Разбирает таблицу «Уведомления с форм» по ЗАГОЛОВКАМ колонок (устойчиво к
    лишним столбцам/чекбоксам и смене порядка). Если шапку найти не удалось -
    откат на позиционный разбор (старый формат).
    [{id, дата, время, дата_время, тип_формы, город, имя, телефон, email, все_данные}]."""
    строки = re.findall(r"<tr\b[^>]*>(.*?)</tr>", html, re.S | re.I)

    # 1) ищем строку-заголовок (в ней есть «Тип формы» / «Город»)
    карта, idx_hdr = None, -1
    for i, tr in enumerate(строки):
        тексты = _ячейки_строки(tr)
        joined = _norm(" ".join(тексты))
        if "типформы" in joined or ("город" in joined and "телефон" in joined):
            карта = {}
            for j, h in enumerate(тексты):
                поле = _HEADER_FIELD.get(_norm(h))
                if поле and поле not in карта.values():
                    карта[j] = поле
            idx_hdr = i
            break

    заявки = []
    if карта:
        ncols = max(карта) + 1
        for tr in строки[idx_hdr + 1:]:
            тексты = _ячейки_строки(tr)
            if len(тексты) < ncols:
                continue
            rec = {k: "" for k in ("id", "тип_формы", "город", "имя",
                                   "телефон", "email", "все_данные")}
            дт = ""
            for j, поле in карта.items():
                val = тексты[j] if j < len(тексты) else ""
                if поле == "дата_время":
                    дт = val
                else:
                    rec[поле] = val
            m = re.match(r"(\d{2}\.\d{2}\.\d{4})\s+(\d{2}:\d{2}:\d{2})", дт)
            if not (rec.get("id") or "").strip() and not m:
                continue   # не строка-заявка (итоги/разделители)
            rec["дата"] = m.group(1) if m else ""
            rec["время"] = m.group(2) if m else ""
            rec["дата_время"] = дт
            заявки.append(rec)
        if заявки:
            return заявки

    # 2) фолбэк: старый позиционный разбор по классам adm-list-table
    for s in re.split(r'<tr[^>]*class="[^"]*adm-list-table-row', html)[1:]:
        cells = re.findall(r'<td[^>]*class="[^"]*adm-list-table-cell[^"]*"[^>]*>(.*?)</td>',
                           s, re.S)
        vals = [_текст(c) for c in cells]
        if len(vals) < 7:
            continue
        дт = vals[1]
        m = re.match(r"(\d{2}\.\d{2}\.\d{4})\s+(\d{2}:\d{2}:\d{2})", дт)
        заявки.append({
            "id": vals[0], "дата": m.group(1) if m else "", "время": m.group(2) if m else "",
            "дата_время": дт, "тип_формы": vals[2] if len(vals) > 2 else "",
            "город": vals[3] if len(vals) > 3 else "", "имя": vals[4] if len(vals) > 4 else "",
            "телефон": vals[5] if len(vals) > 5 else "", "email": vals[6] if len(vals) > 6 else "",
            "все_данные": vals[8] if len(vals) > 8 else "",
        })
    return заявки


def _norm(s: str) -> str:
    s = (s or "").lower().replace("ё", "е")
    return re.sub(r"[^a-zа-я0-9]+", "", s)


def _digits(s: str) -> str:
    return re.sub(r"\D", "", s or "")


def _тип_похож(тип_админ: str, наше_название: str) -> bool:
    """Нестрогое совпадение «Тип формы» из админки и нашего названия формы
    (без учёта регистра/скобок/пробелов; подстрока в любую сторону)."""
    a, b = _norm(тип_админ), _norm(наше_название)
    if not a or not b:
        return False
    return a == b or a in b or b in a


def _это_наша_заявка(row: dict, почта: str, телефон: str, имя: str) -> bool:
    """True, если строка админки похожа на НАШУ тестовую заявку (по почте /
    телефону / имени) - так отсеиваем реальные клиентские заявки."""
    if почта and _norm(row.get("email", "")) == _norm(почта):
        return True
    n = _norm(имя)
    if n and n in _norm(row.get("имя", "")):
        return True
    # Телефон: сайт может переформатировать номер (сдвиг кода страны), поэтому
    # сравниваем ХВОСТ цифр - для телефонных форм без имени/почты это единственная зацепка.
    d, rd = _digits(телефон), _digits(row.get("телефон", ""))
    if len(d) >= 7 and len(rd) >= 7 and rd[-7:] == d[-7:]:
        return True
    return False


def _parse_iso(ts: str):
    try:
        return datetime.strptime((ts or "")[:19], "%Y-%m-%dT%H:%M:%S")
    except Exception:
        return None


def _parse_admin_dt(s: str):
    try:
        return datetime.strptime((s or "")[:19], "%d.%m.%Y %H:%M:%S")
    except Exception:
        return None


def _hhmmss(ts: str) -> str:
    d = _parse_iso(ts)
    return d.strftime("%H:%M:%S") if d else ""


def сопоставить(заявки: list, отправки: list, сверять_город: bool = True,
                дата_сегодня: str = "") -> list:
    """Сверяет наши отправленные формы с заявками из админки.

    дата_сегодня («ДД.ММ.ГГГГ») - если задана, среди наших тест-заявок берём
    только сегодняшние (список тянем без фильтра по дате, отсекаем день здесь).
    Для каждой отправки ищет НАШУ (тестовую) заявку (в том же городе, если
    сверять_город); среди кандидатов берёт заявку с совпадающим «Тип формы»,
    при нескольких - ближайшую по времени. Каждая заявка засчитывается один раз.

    сверять_город=False - когда админка обслуживает один город, но подписывает
    его иначе, чем мы (например, Steelgroup у нас, а в админке «Bakı»): тогда
    город не сверяем, полагаясь на тип формы + наши тест-маркеры.
    Возвращает (результаты, свободные_наши_заявки).
    """
    почта = телефон = имя = ""
    for o in отправки:
        почта = почта or (o.get("почта") or "")
        телефон = телефон or (o.get("телефон") or "")
        имя = имя or (o.get("имя") or "")

    наши = [z for z in заявки if _это_наша_заявка(z, почта, телефон, имя)
            and (not дата_сегодня or z.get("дата", "") == дата_сегодня)]
    использованные = set()
    результаты = []

    for o in отправки:
        gn = _norm(o.get("город", "")) if сверять_город else ""
        кандидаты = []
        for i, z in enumerate(наши):
            if i in использованные:
                continue
            if gn and _norm(z.get("город", "")) != gn:
                continue
            кандидаты.append((i, z))

        выбор = _выбрать(кандидаты, o)
        база = {"город": o.get("город", ""), "название": o.get("название", ""),
                "ts": o.get("ts", ""), "страница": o.get("страница", "")}
        if выбор is None:
            if not наши:
                прим = "в админке нет наших тестовых заявок за сегодня"
            else:
                прим = "заявка этого типа в админке не найдена"
            результаты.append({**база, "статус": "НЕ найдено",
                               "заявка": None, "примечание": прим})
        else:
            i, z = выбор
            использованные.add(i)
            результаты.append({**база, "статус": "Есть в админке",
                               "заявка": z, "примечание": ""})

    # Наши тестовые заявки, которые остались без пары (мы их не отправляли ИЛИ
    # не смогли сопоставить тип) - вернём отдельно, чтобы подсказать в логе.
    свободные = [z for i, z in enumerate(наши) if i not in использованные]
    return результаты, свободные


def _выбрать(кандидаты: list, o: dict):
    """Выбирает заявку СТРОГО по совпадению типа формы (по «админ_тип» из конфига,
    иначе по названию формы). При нескольких подходящих - ближайшую по времени.
    Если совпадения типа нет - None (никогда не «угадываем» по одному времени,
    иначе форма, которой в админке нет, ошибочно займёт чужую заявку)."""
    if not кандидаты:
        return None
    ожид = (o.get("админ_тип") or o.get("название") or "")
    похожие = [(i, z) for i, z in кандидаты
               if _тип_похож(z.get("тип_формы", ""), ожид)]
    if not похожие:
        return None
    ts = _parse_iso(o.get("ts", ""))

    def dist(pair):
        zt = _parse_admin_dt(pair[1].get("дата_время", ""))
        if ts and zt:
            return abs((zt - ts).total_seconds())
        return 1e9

    return sorted(похожие, key=dist)[0]


def записать_в_логи(excel_path: str, результаты: list) -> None:
    """Дописывает результат проверки админки ПРЯМО в лист «Логи»: добавляет колонку
    «Статус в админке» и заполняет её у соответствующих строк форм (сопоставление
    по городу + названию формы). Детали (номер заявки / причина) дописывает в
    «Комментарий». Отдельный лист не создаём - вся инфо в одном месте."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_path)
    if "Логи" not in wb.sheetnames:
        return
    ws = wb["Логи"]
    headers = [str(c.value or "").strip() for c in ws[1]]

    def col(name):
        for i, h in enumerate(headers):
            if h.lower() == name.lower():
                return i + 1
        return -1

    i_city, i_name = col("Город"), col("Название")
    i_comment = col("Комментарий")
    if i_city == -1 or i_name == -1:
        return

    # Колонку «Статус в админке» ставим сразу после «Статус» (или в конец).
    i_adm = col("Статус в админке")
    if i_adm == -1:
        i_st = col("Статус")
        i_adm = (i_st + 1) if i_st != -1 else len(headers) + 1
        ws.insert_cols(i_adm)
        # заголовок в стиле остальной шапки «Логов»
        hc = ws.cell(1, i_adm, "Статус в админке")
        hc.font = Font(bold=True)
        hc.fill = PatternFill("solid", fgColor="EEF3FB")
        ws.column_dimensions[get_column_letter(i_adm)].width = 22
        # пересчитать сдвинувшиеся индексы комментария
        headers = [str(c.value or "").strip() for c in ws[1]]
        i_comment = col("Комментарий")

    def norm(v):
        return _norm(str(v or ""))

    # индекс строк по (город, название) - берём каждую строку один раз
    строки = {}
    for r in range(2, ws.max_row + 1):
        key = (norm(ws.cell(r, i_city).value), norm(ws.cell(r, i_name).value))
        строки.setdefault(key, []).append(r)
    занято = set()

    for res in результаты:
        key = (norm(res.get("город", "")), norm(res.get("название", "")))
        rows = [r for r in строки.get(key, []) if r not in занято]
        if not rows:
            continue
        r = rows[0]
        занято.add(r)
        z = res.get("заявка")
        есть = str(res.get("статус", "")).startswith("Есть")
        cell = ws.cell(r, i_adm, "Есть в админке" if есть else "НЕ найдено")
        cell.font = Font(color="1E8E3E" if есть else "C62828", bold=True)
        # детали → в «Комментарий» (объединяем с тем, что уже есть)
        if i_comment != -1:
            если_есть = (f"заявка #{z['id']} в админке "
                         f"({z.get('время','')}, {res.get('домен_кратко','')})") if z else ""
            деталь = если_есть if есть else (res.get("примечание", "")
                                             or "в админке не найдена")
            if деталь:
                prev = str(ws.cell(r, i_comment).value or "").strip()
                ws.cell(r, i_comment,
                        (prev + "; " if prev else "") + деталь)

    __import__("test_all")._atomic_save_wb(wb, excel_path)


def _домен_кратко(домен: str) -> str:
    """https://stalmetural.uz/... → stalmetural.uz (для колонки «Админка»)."""
    m = re.sub(r"^https?://", "", (домен or "").strip()).split("/")[0]
    return m


def _зона_отправки(зоны: list, город: str):
    """Зона (админка) для нашего города: сперва по явному списку городов зоны,
    иначе - зона по умолчанию (с пустым списком «города», обычно РФ)."""
    for z in зоны:
        gs = z.get("города") or []
        if gs and город in gs:
            return z
    for z in зоны:
        if not (z.get("города") or []):
            return z
    return None


def _xss_admin_в_отчёт(excel_path, статус: str, деталь: str, log=print) -> None:
    """Дописывает в «Логи» итоговую строку XSS-в-админке (Фаза 2), заполняя ячейки
    ПО ИМЕНИ колонки - устойчиво к вставленной админкой колонке «Статус в админке»."""
    from openpyxl import load_workbook
    from datetime import datetime as _dt
    wb = load_workbook(excel_path)
    ws = wb["Логи"] if "Логи" in wb.sheetnames else wb.active
    idx = {str(c.value or "").strip(): i + 1 for i, c in enumerate(ws[1])}
    r = ws.max_row + 1
    now = _dt.now()

    def put(name, val):
        if name in idx:
            ws.cell(r, idx[name], val)
    put("Дата", now.strftime("%d.%m.%Y"))
    put("Время", now.strftime("%H:%M:%S"))
    put("Страница", "Админка")
    put("Название", "Защита от XSS в админке (СМУ)")
    put("Статус", "OK" if статус == "Защищена" else
        ("Проверить" if статус == "Проверить" else "Ошибка"))
    put("Защита от XSS", статус)
    put("Комментарий", деталь)
    __import__("test_all")._atomic_save_wb(wb, excel_path)


def выполнить_проверку(проект_дир, зоны, excel_path: str = "log_forms.xlsx",
                       submitted_path: str = "submitted_forms.json",
                       show: bool = False, log=print, xss_проба: bool = False) -> bool:
    """Уровень 1: по каждой АДМИН-ЗОНЕ (РФ / СНГ / Steelgroup - у них разные
    админки, но один логин/пароль) логинится, читает «Уведомления с форм» за
    сегодня и сверяет с нашими отправками. Пишет единый лист «Админка».

    `зоны` - список {домен, города}. города=[] - зона по умолчанию («все
    остальные», обычно РФ). Тихо пропускается без admin.local.json / отправок.
    """
    creds = загрузить_креды(проект_дир)
    if not creds:
        log("ℹ️ Проверка админки пропущена: не заданы логин/пароль "
            "(введите их на странице проверки форм).")
        return False

    p = Path(submitted_path)
    отправки = []
    if p.is_file():
        try:
            отправки = [o for o in (json.loads(p.read_text(encoding="utf-8")) or []) if o]
        except Exception:
            отправки = []
    if not отправки:
        log("ℹ️ Проверка админки: нет отправленных форм для сверки.")
        return False

    if not зоны:
        зоны = [{"домен": "", "города": []}]

    # Раскладываем отправки по зонам (по нашему городу).
    по_зонам = {}
    for o in отправки:
        z = _зона_отправки(зоны, o.get("город", ""))
        if z is None:
            continue
        по_зонам.setdefault(id(z), (z, []))[1].append(o)

    дата = datetime.now()
    from playwright.sync_api import sync_playwright
    все_результаты = []
    итог_есть = итог_нет = 0
    # Фаза 2 XSS: копим по зонам, исполнился/отражён-сырым/присутствует ли наш
    # payload-маркер на странице списка админки (stored XSS).
    _xss_исп = _xss_сыр = _xss_present = False

    with sync_playwright() as pw:
        b = pw.chromium.launch(headless=not show,
                               args=["--disable-blink-features=AutomationControlled"])
        ctx = b.new_context(locale="ru-RU")
        try:
            for z, з_отправки in по_зонам.values():
                домен = z.get("домен") or ""
                кратко = _домен_кратко(домен) or "(основной)"
                log(f"🔎 Админка [{кратко}]: вход и чтение заявок ({len(з_отправки)} форм) …")
                page = ctx.new_page()
                html = ""
                заявки = []
                try:
                    # Входим «как получится» (не блокируемся по факту входа - это
                    # определим ниже по самой странице списка, без гонок таймингов).
                    войти(page, домен, creds["login"], creds["password"])
                    # сперва пробуем с фильтром по дате; если пусто - без даты (SHOWALL)
                    html = _получить_список_html(page, домен, дата)
                    заявки = разобрать_заявки(html)
                    if not заявки:
                        html2 = _получить_список_html(page, домен, None)
                        з2 = разобрать_заявки(html2)
                        заявки, html = (з2, html2) if з2 else (заявки, html2)
                    # Фаза 2 XSS: наша заявка с payload в ФИО сейчас на странице
                    # списка. Смотрим, исполнился ли payload / отрендерен ли сырым
                    # ЗДЕСЬ, в админке (stored XSS). Best-effort, ошибки гасим.
                    if xss_проба:
                        try:
                            import test_all as _t
                            _и, _с = _t._xss_наблюдение(page, html)
                            _xss_исп = _xss_исп or _и
                            _xss_сыр = _xss_сыр or _с
                            if _t._XSS_MARK in (html or ""):
                                _xss_present = True
                        except Exception:  # noqa: BLE001
                            pass
                finally:
                    try:
                        page.close()
                    except Exception:
                        pass

                # Вход определяем ПО СТРАНИЦЕ СПИСКА: есть таблица заявок → вошли;
                # видна форма логина (поле пароля, без таблицы) → вход не удался.
                if not заявки and _это_форма_входа(html):
                    log(f"⚠️ Админка [{кратко}]: не удалось войти - проверьте логин/пароль "
                        f"(они общие для всех зон).")
                    try:
                        with open(f"admin_debug_{кратко}_login.html", "w",
                                  encoding="utf-8") as fh:
                            fh.write(html)
                    except Exception:
                        pass
                    for o in з_отправки:
                        все_результаты.append({
                            "город": o.get("город", ""), "название": o.get("название", ""),
                            "ts": o.get("ts", ""), "статус": "НЕ найдено", "заявка": None,
                            "домен_кратко": кратко,
                            "примечание": "не удалось войти в админку (логин/пароль)",
                        })
                        итог_нет += 1
                    continue

                if not заявки:
                    # вошли, но список пуст/не распознан - сохраняем страницу для диагностики
                    dbg = f"admin_debug_{кратко}.html"
                    try:
                        with open(dbg, "w", encoding="utf-8") as fh:
                            fh.write(html)
                        log(f"   ⚠️ [{кратко}] заявок не распознано - сохранил "
                            f"страницу в {dbg} (пришлите её мне).")
                    except Exception:
                        pass

                # Если в зоне один город (напр. Steelgroup, а в админке он подписан
                # «Bakı») - город не сверяем: тип формы + тест-маркеры и так однозначны.
                города_зоны = {o.get("город", "") for o in з_отправки}
                сверять = len(города_зоны) > 1
                результаты, свободные = сопоставить(
                    заявки, з_отправки, сверять_город=сверять,
                    дата_сегодня=дата.strftime("%d.%m.%Y"))
                for r in результаты:
                    r["домен_кратко"] = кратко
                    все_результаты.append(r)
                е = sum(1 for r in результаты if r["статус"].startswith("Есть"))
                итог_есть += е
                итог_нет += len(результаты) - е
                log(f"   [{кратко}] заявок в админке: {len(заявки)}; "
                    f"найдено {е} из {len(результаты)}.")
                # Предупреждаем только про НЕОЖИДАННЫЕ свободные заявки - тип
                # которых мы вообще не отправляли (возможный пропуск в маппинге).
                # Лишние копии наших же типов (от прошлых прогонов за сегодня) - не шум.
                if свободные:
                    ожид = [(o.get("админ_тип") or o.get("название") or "")
                            for o in з_отправки]
                    неожид = sorted({zz.get("тип_формы", "") for zz in свободные
                                     if not any(_тип_похож(zz.get("тип_формы", ""), e)
                                                for e in ожид)})
                    if неожид:
                        log(f"   ⚠️ [{кратко}] в админке есть наши тест-заявки типов, "
                            f"которых мы не отправляли: "
                            + ", ".join(f"«{t}»" for t in неожид if t))
        finally:
            b.close()

    записать_в_логи(excel_path, все_результаты)
    log(f"✅ Проверка админки: найдено {итог_есть}, НЕ найдено {итог_нет}. "
        f"Смотри колонку «Статус в админке» на листе «Логи».")

    # Фаза 2 XSS: итог по админке (stored XSS). Пишем отдельной строкой в отчёт.
    if xss_проба:
        try:
            if _xss_исп or _xss_сыр:
                _st = "УЯЗВИМА"
                _det = ("payload из заявки ИСПОЛНИЛСЯ / отрендерен сырым HTML в "
                        "админке - это stored XSS: чинить экранирование вывода в админке")
            elif _xss_present:
                _st = "Защищена"
                _det = "payload в админке показан экранированным (как текст) - не исполняется"
            else:
                _st = "Проверить"
                _det = ("заявка с XSS-маркером в списке админки не найдена - "
                        "проверьте, что заявка дошла, либо смотрите вручную")
            _xss_admin_в_отчёт(excel_path, _st, _det, log)
            log(f"🛡️ Защита от XSS (админка): {_st} - {_det}")
        except Exception as e:  # noqa: BLE001
            log(f"⚠️ XSS-проверка в админке не выполнена: {e}")
    return True


def найти_заявку(заявки: list, тип_формы_админ: str, город: str = "",
                 минут_окно: int = 8, после=None):
    """Ищет заявку по «Тип формы» (нестрогое совпадение) + опц. городу + свежести.
    после - datetime: заявка должна быть не старше (минут_окно) от него.
    Возвращает найденную заявку или None."""
    цель = _norm(тип_формы_админ)
    гнорм = _norm(город) if город else ""
    кандидаты = []
    for z in заявки:
        t = _norm(z["тип_формы"])
        if not (t == цель or цель in t or t in цель):
            continue
        if гнорм and _norm(z["город"]) != гнорм:
            continue
        кандидаты.append(z)
    if после is not None and кандидаты:
        def свежесть(z):
            try:
                zt = datetime.strptime(z["дата_время"][:19], "%d.%m.%Y %H:%M:%S")
                return abs((zt - после).total_seconds())
            except Exception:
                return 1e9
        кандидаты = [z for z in кандидаты if свежесть(z) <= минут_окно * 60]
        кандидаты.sort(key=свежесть)
    return кандидаты[0] if кандидаты else None


def загрузить_креды(проект_дир):
    """Логин/пароль админки. Приоритет - переменные окружения ADMIN_LOGIN /
    ADMIN_PASSWORD (их передаёт страница Streamlit, на диск ничего не пишется).
    Фолбэк - локальный файл admin.local.json. None, если нигде нет."""
    import os
    l = (os.environ.get("ADMIN_LOGIN") or "").strip()
    p = os.environ.get("ADMIN_PASSWORD") or ""
    if l and p:
        return {"login": l, "password": p}
    f = Path(проект_дир) / "admin.local.json"
    if not f.is_file():
        return None
    try:
        d = json.loads(f.read_text(encoding="utf-8"))
        login = str(d.get("login") or "")
        # Игнорируем незаполненный шаблон (ВПИШИ_СЮДА… / ВАШ_ЛОГИН…).
        if login and d.get("password") and "ВПИШИ" not in login.upper() \
                and "ВАШ_" not in login.upper():
            return {"login": login, "password": d.get("password")}
    except Exception:
        return None
    return None


def войти(page, домен: str, login: str, password: str) -> None:
    """Логинится в админку Bitrix (поля USER_LOGIN / USER_PASSWORD). Факт входа
    НЕ возвращаем - его надёжнее определить по странице списка (см. вызов);
    здесь после отправки формы просто ждём, пока пропадёт поле пароля."""
    домен = домен.rstrip("/")
    page.goto(f"{домен}/bitrix/admin/index.php?lang=ru",
              wait_until="domcontentloaded", timeout=30000)
    page.wait_for_timeout(800)
    try:
        if page.locator("input[name='USER_LOGIN']").count() > 0:
            page.fill("input[name='USER_LOGIN']", login)
            page.fill("input[name='USER_PASSWORD']", password)
            btn = page.locator("input[name='Login'], button[name='Login'], "
                               "input[type='submit'], button[type='submit']").first
            btn.click(timeout=8000)
            # Ждём завершения входа: либо ушли со страницы логина (поле пароля
            # исчезло), либо просто пауза - без жёсткой привязки к 2 секундам.
            try:
                page.wait_for_selector("input[name='USER_PASSWORD']",
                                       state="detached", timeout=8000)
            except Exception:
                page.wait_for_timeout(1500)
    except Exception:
        pass


def _получить_список_html(page, домен: str, дата=None) -> str:
    """Открывает список «Уведомления с форм» и возвращает HTML.
    дата=None - без фильтра по дате (SHOWALL, все строки)."""
    page.goto(построить_url_списка(домен, дата),
              wait_until="domcontentloaded", timeout=30000)
    page.wait_for_timeout(1500)
    return page.content()


def войти_и_получить(page, домен: str, login: str, password: str, дата=None) -> str:
    """Логин + HTML списка заявок (для отдельного тестера check_admin.py).
    Сначала пробуем с фильтром по дате; если строк не нашли - без фильтра."""
    войти(page, домен, login, password)
    html = _получить_список_html(page, домен, дата)
    if not разобрать_заявки(html):
        html2 = _получить_список_html(page, домен, None)
        if разобрать_заявки(html2):
            return html2
    return html
