"""
pagespeed_report.py - Excel-отчёт по скорости страниц (PageSpeed Insights).

Собирает самостоятельный xlsx (как «Формы» - отдельный файл, скачивается кнопкой):
  • лист «Сводка по типам» - средние оценки desktop/mobile по типам + Δ к
    прошлому периоду (▲/▼), цвет по порогам Google;
  • лист «Детально» - каждая страница: оценки и метрики Lighthouse
    (FCP/LCP/CLS/TBT), цвет по порогам;
  • лист «Рекомендации» - частые замечания Lighthouse.

Стиль близок к остальным отчётам репозитория (openpyxl, цветные ячейки,
подсказки в заголовках). Данные готовит pagespeed_checker / pagespeed_history.
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pagespeed_checker import (
    METRIC_THRESHOLDS, TYPE_LABELS, TYPE_ORDER,
    metric_rating, score_rating,
)
from pagespeed_history import fmt_ts

# ── Палитра (в тон приложению + пороги Google) ───────────────────────────────
INK = "1A1A1A"
MUTED = "5B5853"
LINE = "DEDBD4"
HEADER_BG = "ECEAE4"
TITLE_BG = "F3F2EE"

# заливки оценок (сплошной цвет + белый жирный текст, как бейджи на макете)
FILL_GOOD = "1F9D2F"
FILL_OK = "E08600"
FILL_POOR = "D03B3B"
FILL_NA = "B7B4AD"

# цвет текста метрик по оценке
FG_GOOD = "0F7D28"
FG_OK = "9C5E00"
FG_POOR = "C0392B"
FG_NA = "8A8781"

# дельты
FG_UP = "006300"
FG_DOWN = "C0392B"
FG_FLAT = "8A8781"

_THIN = Side(style="thin", color=LINE)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _font(size=10, bold=False, color=INK):
    return Font(name="Arial", size=size, bold=bold, color=color)


def _fill(color):
    return PatternFill(fill_type="solid", fgColor=color)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


_SCORE_FILL = {"good": FILL_GOOD, "ok": FILL_OK, "poor": FILL_POOR, "na": FILL_NA}
_METRIC_FG = {"good": FG_GOOD, "ok": FG_OK, "poor": FG_POOR, "na": FG_NA}


def _delta_text(v: Optional[float]) -> tuple[str, str]:
    """(текст, цвет) для дельты оценки."""
    if v is None:
        return "–", FG_FLAT
    if v > 0:
        return f"▲ +{v:g}", FG_UP
    if v < 0:
        return f"▼ {v:g}", FG_DOWN
    return "= 0", FG_FLAT


def _score_cell(ws, row, col, score):
    c = ws.cell(row=row, column=col)
    c.border = _BORDER
    c.alignment = _align("center")
    if score is None:
        c.value = "–"
        c.fill = _fill(FILL_NA)
        c.font = _font(bold=True, color="FFFFFF")
    else:
        c.value = score
        c.fill = _fill(_SCORE_FILL[score_rating(score)])
        c.font = _font(bold=True, color="FFFFFF")
    return c


def _metric_cell(ws, row, col, metric, disp, val):
    c = ws.cell(row=row, column=col)
    c.value = disp or "–"
    c.border = _BORDER
    c.alignment = _align("center")
    c.font = _font(color=_METRIC_FG[metric_rating(metric, val)])
    return c


def _delta_cell(ws, row, col, v):
    c = ws.cell(row=row, column=col)
    txt, color = _delta_text(v)
    c.value = txt
    c.border = _BORDER
    c.alignment = _align("center")
    c.font = _font(bold=True, color=color)
    return c


def _header_row(ws, row, headers, comments=None):
    comments = comments or {}
    for i, title in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=title)
        c.fill = _fill(HEADER_BG)
        c.font = _font(bold=True, color=INK)
        c.alignment = _align("center", wrap=True)
        c.border = _BORDER
        if title in comments:
            cm = Comment(comments[title], "Site Checker")
            cm.width, cm.height = 300, 120
            c.comment = cm


def _title_block(ws, title, meta_lines, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=title)
    t.font = _font(size=15, bold=True, color=INK)
    t.alignment = _align("left")
    t.fill = _fill(TITLE_BG)
    r = 2
    for line in meta_lines:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(row=r, column=1, value=line)
        c.font = _font(size=9, color=MUTED)
        c.alignment = _align("left")
        r += 1
    return r + 1   # первая свободная строка после блока


# ── Лист «Сводка по типам» ───────────────────────────────────────────────────
def _build_summary_sheet(wb, project_name, run_ts, prev_ts, agg, deltas, provider_name):
    ws = wb.active
    ws.title = "Сводка по типам"
    ncols = 6
    compared = f"со снятием {fmt_ts(prev_ts)}" if prev_ts else "нет предыдущего периода"
    start = _title_block(ws, "Скорость страниц – сводка по типам", [
        f"Проект: {project_name}   ·   Источник: {provider_name}",
        f"Проверка: {fmt_ts(run_ts)}   ·   Сравнение: {compared}",
    ], ncols)

    headers = ["Тип страницы", "Кол-во", "🖥 Desktop AVG", "Δ Desktop",
               "📱 Mobile AVG", "Δ Mobile"]
    comments = {
        "🖥 Desktop AVG": "Средняя оценка производительности Lighthouse (0–100) по "
                          "страницам этого типа. 90–100 хорошо, 50–89 средне, 0–49 плохо.",
        "Δ Desktop": "Изменение средней оценки к предыдущему периоду. ▲ рост, ▼ падение.",
    }
    hrow = start
    _header_row(ws, hrow, headers, comments)

    by_type = agg.get("by_type", {})
    d_by = deltas.get("by_type", {})
    row = hrow + 1
    ordered = [tc for tc in TYPE_ORDER if tc in by_type] + \
              [tc for tc in by_type if tc not in TYPE_ORDER]
    for tc in ordered:
        b = by_type[tc]
        ws.cell(row=row, column=1, value=TYPE_LABELS.get(tc, tc)).border = _BORDER
        ws.cell(row=row, column=1).font = _font(bold=True)
        ws.cell(row=row, column=1).alignment = _align("left")
        cnt = ws.cell(row=row, column=2, value=b.get("count", 0))
        cnt.border = _BORDER
        cnt.alignment = _align("center")
        _score_cell(ws, row, 3, b.get("desktop_avg"))
        _delta_cell(ws, row, 4, d_by.get(tc, {}).get("desktop"))
        _score_cell(ws, row, 5, b.get("mobile_avg"))
        _delta_cell(ws, row, 6, d_by.get(tc, {}).get("mobile"))
        row += 1

    # Итоговая строка
    overall = agg.get("overall", {})
    d_over = deltas.get("overall", {})
    tot = ws.cell(row=row, column=1, value="Итого")
    tot.font = _font(bold=True)
    tot.border = _BORDER
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).fill = _fill(HEADER_BG)
    ws.cell(row=row, column=2, value=overall.get("count", 0)).alignment = _align("center")
    ws.cell(row=row, column=2).border = _BORDER
    _score_cell(ws, row, 3, overall.get("desktop_avg"))
    _delta_cell(ws, row, 4, d_over.get("desktop"))
    _score_cell(ws, row, 5, overall.get("mobile_avg"))
    _delta_cell(ws, row, 6, d_over.get("mobile"))

    widths = [22, 10, 15, 12, 15, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)


# ── Лист «Детально» ──────────────────────────────────────────────────────────
def _build_detail_sheet(wb, results):
    ws = wb.create_sheet("Детально")
    headers = ["Страница", "Тип",
               "🖥 Оценка", "D FCP", "D LCP", "D CLS", "D TBT",
               "📱 Оценка", "M FCP", "M LCP", "M CLS", "M TBT", "Ошибка"]
    thr = (f"Пороги: FCP ≤{METRIC_THRESHOLDS['fcp'][0]}с/≤{METRIC_THRESHOLDS['fcp'][1]}с · "
           f"LCP ≤{METRIC_THRESHOLDS['lcp'][0]}с/≤{METRIC_THRESHOLDS['lcp'][1]}с · "
           f"CLS ≤{METRIC_THRESHOLDS['cls'][0]}/≤{METRIC_THRESHOLDS['cls'][1]} · "
           f"TBT ≤{int(METRIC_THRESHOLDS['tbt'][0])}мс/≤{int(METRIC_THRESHOLDS['tbt'][1])}мс")
    comments = {"D FCP": thr, "M FCP": thr,
                "🖥 Оценка": "Оценка Lighthouse для десктопа (0–100).",
                "📱 Оценка": "Оценка Lighthouse для мобильных (0–100)."}
    _header_row(ws, 1, headers, comments)

    row = 2
    for r in results:
        d, m = r.desktop, r.mobile
        u = ws.cell(row=row, column=1, value=r.url)
        u.border = _BORDER
        u.font = _font(color=INK)
        u.alignment = _align("left")
        t = ws.cell(row=row, column=2, value=TYPE_LABELS.get(r.type_code, r.type_code))
        t.border = _BORDER
        t.alignment = _align("center")

        _score_cell(ws, row, 3, d.score)
        _metric_cell(ws, row, 4, "fcp", d.fcp_disp, d.fcp_val)
        _metric_cell(ws, row, 5, "lcp", d.lcp_disp, d.lcp_val)
        _metric_cell(ws, row, 6, "cls", d.cls_disp, d.cls_val)
        _metric_cell(ws, row, 7, "tbt", d.tbt_disp, d.tbt_val)

        _score_cell(ws, row, 8, m.score)
        _metric_cell(ws, row, 9, "fcp", m.fcp_disp, m.fcp_val)
        _metric_cell(ws, row, 10, "lcp", m.lcp_disp, m.lcp_val)
        _metric_cell(ws, row, 11, "cls", m.cls_disp, m.cls_val)
        _metric_cell(ws, row, 12, "tbt", m.tbt_disp, m.tbt_val)

        err = " ".join(x for x in (d.error, m.error) if x).strip()
        e = ws.cell(row=row, column=13, value=err)
        e.border = _BORDER
        e.font = _font(color=FG_POOR if err else INK, size=9)
        e.alignment = _align("left", wrap=True)
        row += 1

    widths = [42, 12, 9, 9, 9, 8, 9, 9, 9, 9, 8, 9, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=2, column=1)


# ── Лист «Рекомендации» ──────────────────────────────────────────────────────
def _rec_items_text(items) -> str:
    """Конкретные ресурсы одной рекомендации в столбик: «url – экономия 320 КБ»."""
    lines = []
    for it in items or []:
        url = it.get("url", "")
        info = it.get("info", "")
        lines.append(f"• {url} – {info}" if info else f"• {url}")
    return "\n".join(lines)


def _build_recs_sheet(wb, top_recs, total_pages):
    ws = wb.create_sheet("Рекомендации")
    headers = ["Замечание Lighthouse", "Страниц", "Экономия",
               "Что и где конкретно", "Примеры страниц"]
    comments = {
        "Страниц": "На скольких проверенных страницах встретилось.",
        "Экономия": "Потенциальный выигрыш по оценке Lighthouse (вес/время).",
        "Что и где конкретно": "Конкретные ресурсы (файлы) и сколько на каждом "
                               "можно сэкономить - прямо из отчёта Lighthouse.",
    }
    _header_row(ws, 1, headers, comments)

    row = 2
    for rec in top_recs:
        title = rec.get("title", "")
        cnt = rec.get("pages", 0)
        savings = rec.get("savings", "")
        items_txt = _rec_items_text(rec.get("items"))
        pages_txt = "\n".join(rec.get("example_pages", []))

        a = ws.cell(row=row, column=1, value=title)
        a.font = _font(bold=True)
        b = ws.cell(row=row, column=2, value=cnt)
        b.alignment = _align("center")
        b.font = _font(bold=True, color=FG_POOR if cnt else INK)
        c = ws.cell(row=row, column=3, value=savings or "–")
        c.font = _font(color=FG_OK)
        d = ws.cell(row=row, column=4, value=items_txt or "–")
        d.font = _font(size=9, color=MUTED)
        e = ws.cell(row=row, column=5, value=pages_txt or "–")
        e.font = _font(size=9, color=MUTED)
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = _BORDER
            cell.alignment = _align("left", "top", wrap=True)
        # высота строки под число ресурсов/страниц
        n_lines = max(len(rec.get("items") or []), len(rec.get("example_pages") or []), 1)
        ws.row_dimensions[row].height = max(18, 15 * n_lines + 4)
        row += 1

    if not top_recs:
        c = ws.cell(row=2, column=1, value="Критичных замечаний не найдено 👍")
        c.font = _font()

    for col, w in zip("ABCDE", [46, 9, 16, 52, 40]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = ws.cell(row=2, column=1)


# ── Сборка ───────────────────────────────────────────────────────────────────
def build_workbook(*, project_name, run_ts, prev_ts, results, agg, deltas,
                   top_recs, provider_name="PageSpeed Insights") -> Workbook:
    wb = Workbook()
    _build_summary_sheet(wb, project_name, run_ts, prev_ts, agg, deltas, provider_name)
    _build_detail_sheet(wb, results)
    _build_recs_sheet(wb, top_recs, len(results))
    return wb


def to_bytes(**kwargs) -> bytes:
    """xlsx как bytes - для st.download_button."""
    wb = build_workbook(**kwargs)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_report(path, **kwargs) -> Path:
    """Сохранить xlsx на диск."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(**kwargs).save(path)
    return path
