"""
Страница «Проверка форм» - фоновый прогон отправки форм на сайтах проекта.

Сделана по образцу страницы «Автокликеры»: кнопка стартует отдельный процесс
(forms_run.py) и сразу освобождает интерфейс. Движок открывает реальный Chrome
(Playwright, по умолчанию скрыто), заполняет формы и отправляет, результат
пишется в log_forms.xlsx.

Окружение:
  • Локально (streamlit run app.py) - работает.
  • Облако по ссылке - недоступно (нет браузера и движка на сервере).
  • Свой сервер (в планах) - заработает так же.
"""
import csv
import importlib.util
import json
import os
import random
import re
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

from urllib.parse import urlparse

import streamlit as st

ROOT = Path(__file__).parent.parent
PY = sys.executable
LOG_FILE = ROOT / 'cache' / 'forms.log'
PID_FILE = ROOT / 'cache' / 'forms.pid'


def _load_cities(project: str):
    """Справочник городов проекта: список dict {country, city, url, mail}.
    Пусто, если файла нет. Первый город - основной сайт (Москва)."""
    project = CITIES_FROM.get(project, project)
    f = ROOT / 'forms_tester' / 'projects' / project / 'cities.csv'
    if not f.exists():
        return []
    out = []
    try:
        with open(f, encoding='utf-8', newline='') as fh:
            for row in csv.DictReader(fh):
                city = (row.get('город') or '').strip()
                if city:
                    out.append({
                        'country': (row.get('страна') or 'Россия').strip(),
                        'city': city,
                        'url': (row.get('url') or '').strip(),
                        'mail': (row.get('почта') or '').strip(),
                    })
    except Exception:
        return []
    return out


_COUNTRY_FLAG = {
    'Россия': '🇷🇺', 'Казахстан': '🇰🇿', 'Беларусь': '🇧🇾', 'Кыргызстан': '🇰🇬',
    'Киргизия': '🇰🇬', 'Узбекистан': '🇺🇿', 'Азербайджан': '🇦🇿', 'Армения': '🇦🇲',
    'Украина': '🇺🇦',
    # Steelgroup - отдельный азербайджанский сайт (steelgroup.az), в «Проверке
    # целей» он идёт как smu-az2. Своя группа-«страна», чтобы показывался
    # отдельным доменом (иначе прятался за stalmetural.am под «Арменией»).
    'Steelgroup': '🇦🇿',
}


def _host(url: str) -> str:
    return urlparse((url or '').strip()).netloc


def _main_domains(cities):
    """Основной домен каждой страны: строка справочника с самым «коротким» хостом
    (без поддомена-города: mepen.kz, а не aktau.mepen.kz). Порядок стран - как в csv."""
    best = {}
    for c in cities:
        h = _host(c['url'])
        depth = h.count('.')
        cur = best.get(c['country'])
        if cur is None or depth < cur[0]:
            best[c['country']] = (depth, c)
    out, seen = [], set()
    for c in cities:
        if c['country'] not in seen:
            seen.add(c['country'])
            out.append(best[c['country']][1])
    return out

PROJECTS = {
    'smu': {'name': 'СМУ - Стальметурал', 'domain': 'stalmetural.ru'},
    'imp': {'name': 'ИМП - Инметпром', 'domain': 'inmetprom.ru'},
    'mpe': {'name': 'МПЭ - Мепэн', 'domain': 'mepen.ru'},
    'avia': {'name': 'АПС - Авиапромсталь', 'domain': 'aviastal.ru'},
    'metpromko': {'name': 'МТТ - Метпромко', 'domain': 'metpromko.ru'},
}

# Проекты-варианты берут справочник городов у «родителя» (свой config.py,
# общий cities.csv). Держим в синхроне с CITIES_FROM в forms_run.py.
CITIES_FROM = {}

# Полный текст-подсказка (раньше был большим жёлтым блоком, теперь - в «❓»).
HELP_TEXT = (
    'Проверка открывает реальный браузер (Playwright) на ЭТОМ компьютере: '
    'заполняет формы на сайтах проекта и отправляет заявки. Работает, когда '
    'приложение запущено **локально** (`streamlit run app.py`). В облаке по '
    'ссылке недоступно. После переноса на свой сервер - заработает.'
)


def _read_pid():
    try:
        return int(PID_FILE.read_text().strip())
    except Exception:
        return None


def _pid_alive(pid) -> bool:
    if not pid:
        return False
    if os.name == 'nt':
        try:
            out = subprocess.run(['tasklist', '/FI', f'PID eq {pid}'],
                                 capture_output=True, text=True).stdout
            return str(pid) in out
        except Exception:
            return False
    try:
        os.kill(pid, 0)
        return True
    except Exception:
        return False


def _kill_tree(pid):
    if not pid:
        return
    if os.name == 'nt':
        subprocess.run(['taskkill', '/F', '/T', '/PID', str(pid)],
                       capture_output=True)
    else:
        import signal
        try:
            os.kill(pid, signal.SIGTERM)
        except Exception:
            pass


def _deps_ready() -> tuple[bool, list[str]]:
    """Есть ли в этом окружении движок (его библиотеки + браузер). Возвращает
    (готово, список_чего_нет). На облаке по ссылке тут будет False."""
    missing = []
    for mod, label in (('bs4', 'beautifulsoup4'), ('requests', 'requests'),
                       ('openpyxl', 'openpyxl'), ('playwright', 'playwright')):
        if importlib.util.find_spec(mod) is None:
            missing.append(label)
    return (not missing), missing


def _launch_background(args: list[str], log_path: Path, extra_env: dict | None = None):
    """Запустить процесс в фоне, вывод - в файл. UI не блокируется.
    extra_env - доп. переменные окружения (например, логин/пароль админки);
    они передаются только дочернему процессу и на диск не пишутся."""
    log_path.parent.mkdir(parents=True, exist_ok=True)
    env = dict(os.environ)
    env['PYTHONIOENCODING'] = 'utf-8'
    env['PYTHONUNBUFFERED'] = '1'
    if extra_env:
        env.update({k: v for k, v in extra_env.items() if v})
    creationflags = 0
    if os.name == 'nt':
        creationflags = getattr(subprocess, 'CREATE_NO_WINDOW', 0)
    f = open(log_path, 'a', encoding='utf-8')
    proc = subprocess.Popen(
        [PY, *args], cwd=str(ROOT),
        stdout=f, stderr=subprocess.STDOUT, env=env,
        creationflags=creationflags,
    )
    try:
        PID_FILE.write_text(str(proc.pid), encoding='utf-8')
    except Exception:
        pass
    return proc.pid


def _project_has_admin(project: str) -> bool:
    """True, если у проекта настроены АДМИН_ЗОНЫ (есть проверка админки). Пока
    это только СМУ; ИМП/МПЭ устроены иначе - для них раздел не показываем."""
    p = ROOT / 'forms_tester' / 'projects' / project / 'config.py'
    try:
        spec = importlib.util.spec_from_file_location(f'cfg_adm_{project}', p)
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        return bool(getattr(m, 'АДМИН_ЗОНЫ', None))
    except Exception:
        return False


def _project_uses_proxy(project: str) -> bool:
    """True, если у проекта в config.py задан ИСПОЛЬЗОВАТЬ_ПРОКСИ (сайт режет
    прямое подключение, напр. Метпромко) - тогда галочка «Вкл. Прокси» стартует
    включённой. Остальным проектам флага нет - прокси по умолчанию выключен."""
    p = ROOT / 'forms_tester' / 'projects' / project / 'config.py'
    try:
        spec = importlib.util.spec_from_file_location(f'cfg_prx_{project}', p)
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        return bool(getattr(m, 'ИСПОЛЬЗОВАТЬ_ПРОКСИ', False))
    except Exception:
        return False


def _count_expected(project: str) -> int:
    """Сколько форм ожидается проверить (для шкалы прогресса). Best-effort:
    считаем включённые формы/модалки + шаги-формы в сценариях. Если не вышло - 0."""
    p = ROOT / 'forms_tester' / 'projects' / project / 'config.py'
    try:
        spec = importlib.util.spec_from_file_location(f'cfg_count_{project}', p)
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
    except Exception:
        return 0

    def on(d) -> bool:
        v = d.get('включено', d.get('enabled', True))
        if isinstance(v, bool):
            return v
        return str(v).strip().lower() not in ('false', '0', 'нет', 'off', '')

    total = 0
    try:
        for block in getattr(m, 'СТРАНИЦЫ_ДЛЯ_ПРОВЕРКИ', []) or []:
            if not on(block):
                continue
            for key in ('формы', 'модалки'):
                for item in block.get(key, []) or []:
                    if on(item):
                        total += 1
            # Сценарий движок гонит как ОДНУ единицу (одна строка «▶ Форма»),
            # поэтому считаем 1 на сценарий, а не по шагам - иначе «всего» больше,
            # чем реально печатает прогон. Согласовано с _list_forms.
            _сцены = block.get('сценарии', []) or []
            for sc in _сцены:
                if on(sc):
                    total += 1
            if not _сцены and block.get('шаги'):   # legacy: шаги прямо в блоке = один сценарий
                total += 1
    except Exception:
        return 0
    return total


def _flag_on(v) -> bool:
    """Значение флага «включено» → bool (терпимо к строкам «нет»/«off»/…)."""
    if isinstance(v, bool):
        return v
    if v is None:
        return True
    return str(v).strip().lower() not in ('false', '0', 'нет', 'off', '')


def _list_forms(project: str):
    """Список форм проекта В ПОРЯДКЕ ПРОГОНА: [{'page','name'}, ...].
    Имена = ровно те «названия», что движок пишет в отчёт (сценарии/формы/модалки).
    Учитываем только ВКЛЮЧЁННЫЕ страницы/формы, чтобы не показывать то, что не гоняется."""
    p = ROOT / 'forms_tester' / 'projects' / project / 'config.py'
    try:
        spec = importlib.util.spec_from_file_location(f'cfg_forms_{project}', p)
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
    except Exception:
        return []

    out, seen = [], set()

    def add(page, name):
        name = (name or '').strip()
        if name and name not in seen:
            seen.add(name)
            out.append({'page': page, 'name': name})

    try:
        for block in getattr(m, 'СТРАНИЦЫ_ДЛЯ_ПРОВЕРКИ', []) or []:
            if not _flag_on(block.get('включено', True)):
                continue
            page = str(block.get('тип') or '').strip()
            # Сценарии (клик → форма). Имя = «название» сценария.
            if _flag_on(block.get('сценарий_включен', True)):
                сцены = block.get('сценарии') or []
                for sc in сцены:
                    if _flag_on(sc.get('включено', True)):
                        add(page, sc.get('название') or block.get('название_сценария') or page)
                # legacy: шаги прямо в блоке (без «сценарии»)
                if not сцены and (block.get('шаги')):
                    add(page, block.get('название_сценария') or page)
            # Отдельные формы на странице
            if _flag_on(block.get('формы_включены', True)):
                for f in block.get('формы') or []:
                    if _flag_on(f.get('включено', True)):
                        add(page, f.get('название'))
            # Модалки (кнопка открывает окно)
            if _flag_on(block.get('модалки_включены', True)):
                for mo in block.get('модалки') or []:
                    if _flag_on(mo.get('включено', True)):
                        add(page, mo.get('название_теста'))
    except Exception:
        return out
    return out


# Где искать форму НА странице (для подсказки «?»). Единый словарь с отчётом по
# формам - лежит в forms_tester/form_locations.py (тот же текст едет и в колонку
# «Где находится» отчёта). Если формы нет в словаре, подсказка не показывается.
try:
    from forms_tester.form_locations import FORM_WHERE as _FORM_WHERE
except Exception:
    _FORM_WHERE = {}


def _rows_done(xlsx: Path):
    """Сколько ФОРМ уже записано в лог (строки минус шапка). None - не прочиталось.
    Не считаем служебные строки-проверки (согласие 2.13, cookie/политика/живочат
    2.12) - они не формы и раздували бы прогресс (шкала считает формы)."""
    if not xlsx.exists():
        return 0
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx, read_only=True)
        # Считаем строки ИМЕННО листа «Логи» (не активного и не «Цели»):
        # строки целей на отдельном листе не должны раздувать прогресс.
        ws = wb["Логи"] if "Логи" in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None) or ()
        try:
            i_name = [str(h or '').strip().lower() for h in header].index('название')
        except ValueError:
            i_name = -1
        _skip = ('согласие и политика', 'cookie-уведомление',
                 'ссылка на политику', 'живочат')
        n = 0
        for r in rows:
            if not r or all(v in (None, '') for v in r):
                continue
            nm = str(r[i_name] or '').strip().lower() if 0 <= i_name < len(r) else ''
            if any(nm.startswith(s) for s in _skip):
                continue
            n += 1
        wb.close()
        return max(n, 0)
    except Exception:
        return None


# ── Заголовок + подсказка «❓» ───────────────────────────────────────
# Кнопку скачивания отчёта подсвечиваем зелёным.
st.markdown(
    """
    <style>
    /* Чёрная primary-кнопка «Запустить проверку»: белый текст. Без этого на этой
       странице глобальное правило .stApp p/div красит подпись в тёмный - и текст
       сливается с чёрным фоном. Как на Чек-листе и «Проверке целей». */
    div[data-testid="stButton"] > button[kind="primary"],
    div[data-testid="stButton"] > button[data-testid="stBaseButton-primary"],
    div[data-testid="stButton"] > button[data-testid="baseButton-primary"] {
        background: #1A1A1A !important; border: 1px solid #1A1A1A !important;
        color: #FFFFFF !important;
    }
    div[data-testid="stButton"] > button[kind="primary"] *,
    div[data-testid="stButton"] > button[data-testid="stBaseButton-primary"] *,
    div[data-testid="stButton"] > button[data-testid="baseButton-primary"] * {
        color: #FFFFFF !important;
    }
    div[data-testid="stButton"] > button[kind="primary"]:hover {
        background: #000000 !important; border-color: #000000 !important;
    }
    [data-testid="stDownloadButton"] button {
        background: #1E8E3E !important; border: 1px solid #1E8E3E !important;
    }
    [data-testid="stDownloadButton"] button * { color: #FFFFFF !important; }
    [data-testid="stDownloadButton"] button:hover {
        background: #176D30 !important; border-color: #176D30 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

_th, _qh = st.columns([0.88, 0.12], vertical_alignment='bottom')
with _th:
    st.title('📝 Проверка форм')
with _qh:
    with st.popover('❓', use_container_width=False):
        st.markdown(HELP_TEXT)

# ── Выбор проекта ────────────────────────────────────────────────────
# По умолчанию проект НЕ выбран - чтобы ничего не запускалось случайно.
# ПЕРСИСТ между вкладками: Streamlit при уходе на другую страницу чистит
# состояние виджетов, поэтому раньше проект «сбрасывался» при возврате (и с ним
# исчезали лог/отчёт/прогресс). Держим выбор в ПОСТОЯННОМ ключе session_state
# (его Streamlit не трогает - это не ключ виджета) и восстанавливаем как index
# при каждом заходе. Виджет - без key, чтобы не попасть под чистку. Смена проекта
# сама переключает лог/отчёт - они per-проектные (cache/forms/<pid>/…).
_proj_opts = list(PROJECTS.keys())
_proj_saved = st.session_state.get('fc_project_sel')
_proj_idx = _proj_opts.index(_proj_saved) if _proj_saved in _proj_opts else None
pid_key = st.selectbox('Проект', _proj_opts,
                       format_func=lambda k: PROJECTS[k]['name'],
                       index=_proj_idx, placeholder='- выберите проект -')
st.session_state['fc_project_sel'] = pid_key   # запоминаем для возврата на вкладку
if not pid_key:
    st.info('Выберите проект, чтобы настроить и запустить проверку форм.')
    st.stop()
proj = PROJECTS[pid_key]

# PID и лог прогона - ПЕР-ПРОЕКТНЫЕ (cache/forms/<pid>/…), а не один файл на всё
# приложение. С глобальным файлом проверка, запущенная одним человеком, гасила
# кнопку и роняла страницу у другого (чужая сессия не знала проект → PROJECTS[None]).
# Теперь разные проекты гоняются с разных компьютеров независимо, а один и тот же
# проект честно показывает «идёт», кто бы его ни запустил. Переменные модульные -
# переопределяем здесь, чтобы _read_pid()/_launch_background() читали новые пути.
PID_FILE = ROOT / 'cache' / 'forms' / pid_key / 'run.pid'
LOG_FILE = ROOT / 'cache' / 'forms' / pid_key / 'run.log'

st.markdown(
    f"Будут проверены формы сайта **{proj['name']}** (`{proj['domain']}`): "
    'обратная связь, заявки, расчёты, оформление заказа и т.п. - по настройкам '
    'проекта.'
)

st.divider()

# ── Данные проекта (нужны сценариям, доменам и формам ниже) ──────────
_cities = _load_cities(pid_key)
_chosen_cities = []          # список названий городов для прогона ([] = основной сайт)
_all_names = [c['city'] for c in _cities]
_mains = _main_domains(_cities) if _cities else []   # основной домен каждой страны
_main_by_country = {c['country']: c for c in _mains}
_groups = {}                                         # страны в порядке справочника
for c in _cities:
    _groups.setdefault(c['country'], []).append(c['city'])
_all_forms = _list_forms(pid_key)
_all_form_names = [f['name'] for f in _all_forms]

# ── Шаблоны прогона (пункт 2.6: выбор/сохранение шаблона) ─────────────
# Шаблон запоминает набор: выбранные города + выбранные формы. Можно сохранить
# настройку под именем и переиспользовать (в т.ч. на будущее для автозапуска).
# Хранится в cache/forms/<project>/templates.json (рядом с random_cities.json).
# ВАЖНО: «Применить» выставляет session_state ДО отрисовки виджетов городов/форм,
# поэтому блок стоит здесь (раньше них) - иначе Streamlit не даст менять значения.
_TPL_FILE = ROOT / 'cache' / 'forms' / pid_key / 'templates.json'


def _tpl_load_all() -> dict:
    try:
        return json.loads(_TPL_FILE.read_text(encoding='utf-8')) or {}
    except Exception:
        return {}


def _tpl_save_all(data: dict):
    try:
        _TPL_FILE.parent.mkdir(parents=True, exist_ok=True)
        _TPL_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                             encoding='utf-8')
    except Exception:
        pass


# Режимы выбора городов и форм. Определены ЗДЕСЬ (до _tpl_apply и до блока
# шаблонов), т.к. _tpl_apply сверяется с ними при загрузке шаблона - а сам блок
# шаблонов стоит выше виджетов режимов.
_MODE_OPTIONS = ['Основные домены (по странам)', 'Выбрать города', 'Случайные города']
_FORMS_MODE_OPTIONS = ['Все формы', 'Выбрать формы']


def _tpl_apply(tpl: dict):
    """Проставляет в session_state ВСЕ настройки страницы из шаблона, ВОЗВРАЩАЯ
    ИМЕННО тот режим, что был сохранён (город: основные домены / выбор городов /
    случайные; формы: все / выбор). Логин и пароль админки НЕ трогаем - их всегда
    вводят вручную. Вызывается ДО отрисовки этих виджетов + rerun.

    Раньше загрузка ВСЕГДА переключала на «Выбрать города» и проставляла галочки
    городов - поэтому после загрузки «зачем-то нажимались» галочки городов, даже
    если сохраняли режим «Основные домены». Теперь режим и выбор возвращаются как
    были."""
    # Режимы (город + формы) - вернуть сохранённые.
    _m = tpl.get('mode')
    if _m in _MODE_OPTIONS:
        st.session_state[f'fc_mode_{pid_key}'] = _m
    _fm = tpl.get('forms_mode')
    if _fm in _FORMS_MODE_OPTIONS:
        st.session_state[f'fc_forms_mode_{pid_key}'] = _fm

    # Режим «Основные домены»: какие СТРАНЫ отмечены.
    _countries = tpl.get('countries') or {}
    for c in _mains:
        if c['country'] in _countries:
            st.session_state[f'fc_main_{pid_key}_{c["country"]}'] = \
                bool(_countries[c['country']])

    # Режим «Выбрать города»: какие ГОРОДА отмечены. Ставим init-флаг, чтобы дефолт
    # (основные домены каждой страны) не перетёр наш выбор. Применяем ТОЛЬКО если в
    # шаблоне реально что-то выбрано - иначе (шаблон сохранён в режиме «Основные
    # домены», где список городов пустой) при переключении на «Выбрать города»
    # показался бы пустой список вместо привычного дефолта.
    _city_cb = tpl.get('city_cb')
    if _city_cb and any(_city_cb.values()):
        for nm in _all_names:
            st.session_state[f'fc_cb_{pid_key}_{nm}'] = bool(_city_cb.get(nm, False))
        st.session_state[f'fc_init_osn_{pid_key}'] = True
    elif tpl.get('cities'):                  # старый формат шаблона: только список
        _cset = {c for c in (tpl.get('cities') or []) if c in _all_names}
        for nm in _all_names:
            st.session_state[f'fc_cb_{pid_key}_{nm}'] = (nm in _cset)
        st.session_state[f'fc_init_osn_{pid_key}'] = True

    # Режим «Случайные города»: числа по странам + общее.
    _rnd = tpl.get('random') or {}
    if _rnd:
        _per = _rnd.get('per') or {}
        for c in _groups:
            if c in _per:
                st.session_state[f'fc_rnd_{pid_key}_{c}'] = \
                    max(0, min(int(_per[c] or 0), len(_groups[c])))
        if 'total' in _rnd:
            st.session_state[f'fc_rnd_total_{pid_key}'] = int(_rnd.get('total') or 0)

    # Режим «Выбрать формы»: какие ФОРМЫ отмечены.
    _form_cb = tpl.get('form_cb')
    if _form_cb is not None:
        for nm in _all_form_names:
            st.session_state[f'ff_cb_{pid_key}_{nm}'] = bool(_form_cb.get(nm, True))
        st.session_state[f'ff_init_{pid_key}'] = True
    elif tpl.get('forms') is not None:       # старый формат шаблона: только список
        _fset = {f for f in (tpl.get('forms') or []) if f in _all_form_names}
        for nm in _all_form_names:
            st.session_state[f'ff_cb_{pid_key}_{nm}'] = (nm in _fset)
        st.session_state[f'ff_init_{pid_key}'] = True

    # Настройки запуска (галочки/режимы запуска, админки, почты).
    for _k, _v in (tpl.get('options') or {}).items():
        st.session_state[_k] = _v


def _tpl_collect() -> dict:
    """Собирает ТЕКУЩИЕ настройки страницы из session_state в шаблон. Всё берётся
    из session_state (значения виджетов лежат там с прошлой отрисовки), поэтому
    можно вызывать ПРЯМО по клику «Сохранить» в верхнем блоке - не дожидаясь низа
    страницы и без st.rerun() (rerun сбросил бы галочки, что идут ниже). Логин и
    пароль админки в шаблон НЕ пишем."""
    _opt_keys = [f'fc_admin_on_{pid_key}', f'fc_mail_mode_{pid_key}',
                 f'fc_mail_own_{pid_key}', f'fc_clear_{pid_key}',
                 f'fc_fileprobe_{pid_key}', f'fc_xss_{pid_key}',
                 f'fc_srvval_{pid_key}', f'fc_ratelimit_{pid_key}']
    return {
        'mode': st.session_state.get(f'fc_mode_{pid_key}', _MODE_OPTIONS[0]),
        'forms_mode': st.session_state.get(f'fc_forms_mode_{pid_key}',
                                           _FORMS_MODE_OPTIONS[0]),
        'countries': {c['country']: bool(st.session_state.get(
            f'fc_main_{pid_key}_{c["country"]}', True)) for c in _mains},
        'city_cb': {nm: bool(st.session_state.get(f'fc_cb_{pid_key}_{nm}', False))
                    for nm in _all_names},
        'random': {
            'total': int(st.session_state.get(f'fc_rnd_total_{pid_key}', 0) or 0),
            'per': {c: int(st.session_state.get(f'fc_rnd_{pid_key}_{c}', 0) or 0)
                    for c in _groups}},
        'form_cb': {nm: bool(st.session_state.get(f'ff_cb_{pid_key}_{nm}', True))
                    for nm in _all_form_names},
        'options': {kk: st.session_state[kk] for kk in _opt_keys
                    if kk in st.session_state},
    }


_tpl_open = bool(st.session_state.pop(f'tpl_open_{pid_key}', False))
with st.expander('📁 Проектные шаблоны (сохранить/загрузить настройки страницы)',
                 expanded=_tpl_open):
    st.caption('Как это работает: **1)** настройте страницу как нужно (города, '
               'формы, галочки) → **2)** впишите название → **3)** нажмите '
               '«💾 Сохранить». Потом эти же настройки вернёте кнопкой «Загрузить».')

    # ── СОХРАНИТЬ (идёт ПЕРВЫМ: файл пишется прямо тут, поэтому список «Загрузить»
    # ниже сразу видит новый шаблон - без перезагрузки страницы). ──
    st.markdown('**Сохранить текущие настройки как новый шаблон**')
    _sv1, _sv2 = st.columns([3, 1], vertical_alignment='bottom')
    _tpl_new = _sv1.text_input(
        'Название шаблона', key=f'tpl_name_{pid_key}', label_visibility='collapsed',
        placeholder='Например: СМУ РФ - только оформление заказа')
    _saved_now = ''
    if _sv2.button('💾 Сохранить', use_container_width=True, type='primary',
                   disabled=not (_tpl_new or '').strip(), key=f'tpl_save_{pid_key}'):
        # Сохраняем СРАЗУ и БЕЗ st.rerun(): rerun из этого верхнего блока сбросил
        # бы галочки/режимы, которые отрисовываются НИЖЕ по странице (Streamlit
        # чистит состояние виджетов, не отрисованных в прерванном прогоне) - и
        # настройки «слетали» бы после сохранения. Всё собираем из session_state.
        _nm = _tpl_new.strip()
        _all_t = _tpl_load_all()
        _all_t[_nm] = _tpl_collect()
        _tpl_save_all(_all_t)
        _saved_now = _nm
    if _saved_now:
        st.success(f'✓ Готово! Шаблон «{_saved_now}» сохранён - он уже в списке '
                   '«Загрузить шаблон» ниже. Настройки на странице НЕ изменились, '
                   'города/формы/галочки на месте.')
    st.caption('Шаблон запоминает все настройки страницы (города, формы, галочки '
               'и режимы), кроме логина и пароля админки. Хранится на сервере '
               'проекта **до перезапуска приложения** - после может сброситься.')

    st.divider()

    # ── ЗАГРУЗИТЬ (список читаем ПОСЛЕ сохранения - уже с новым шаблоном). ──
    _tpls = _tpl_load_all()
    if _tpls:
        _tc1, _tc2, _tc3 = st.columns([3, 1, 1], vertical_alignment='bottom')
        _pick = _tc1.selectbox(
            'Загрузить шаблон', list(_tpls.keys()),
            index=None, placeholder='- выберите шаблон -', key=f'tpl_pick_{pid_key}')
        if _tc2.button('Загрузить', use_container_width=True,
                       disabled=not _pick, key=f'tpl_apply_{pid_key}'):
            _tpl_apply(_tpls[_pick])
            st.toast(f'Шаблон «{_pick}» загружен', icon='📥')
            st.rerun()
        if _tc3.button('Удалить', use_container_width=True,
                       disabled=not _pick, key=f'tpl_del_{pid_key}'):
            _tpls.pop(_pick, None)
            _tpl_save_all(_tpls)
            st.rerun()
    else:
        st.caption('Пока сохранённых шаблонов нет - создайте первый выше.')

# (_MODE_OPTIONS / _FORMS_MODE_OPTIONS определены выше - до блока шаблонов.)
# Запоминаются ТОЛЬКО настройки «Случайных городов» (числа по странам).
# Остальные галочки каждый заход начинаются с дефолта - так предсказуемее.
_RND_FILE = ROOT / 'cache' / 'forms' / pid_key / 'random_cities.json'


def _round_robin(total):
    """Раздаёт total по странам по кругу (начиная с России), не больше,
    чем городов в стране."""
    _counts = {k: 0 for k in _groups}
    _left = int(total)
    while _left > 0:
        _gave = False
        for k in _groups:
            if _left <= 0:
                break
            if _counts[k] < len(_groups[k]):
                _counts[k] += 1
                _left -= 1
                _gave = True
        if not _gave:
            break
    return _counts


def _rnd_save():
    """Автосохранение чисел «Случайных городов» для проекта: сколько всего и
    по каждой стране. Подхватываются при следующем заходе."""
    try:
        data = {
            'всего': int(st.session_state.get(f'fc_rnd_total_{pid_key}', 0) or 0),
            'по_странам': {k: int(st.session_state.get(f'fc_rnd_{pid_key}_{k}', 0) or 0)
                           for k in _groups},
        }
        _RND_FILE.parent.mkdir(parents=True, exist_ok=True)
        _RND_FILE.write_text(json.dumps(data, ensure_ascii=False), encoding='utf-8')
    except Exception:
        pass


# Сохранённые числа «Случайных городов» подхватываем сами (раз за сессию).
if not st.session_state.get(f'fc_rnd_loaded_{pid_key}'):
    st.session_state[f'fc_rnd_loaded_{pid_key}'] = True
    if _RND_FILE.exists():
        try:
            _d = json.loads(_RND_FILE.read_text(encoding='utf-8'))
            _by = _d.get('по_странам') or {}
            if _by:
                for k in _groups:
                    if k in _by:
                        st.session_state[f'fc_rnd_{pid_key}_{k}'] = max(
                            0, min(int(_by[k] or 0), len(_groups[k])))
                st.session_state[f'fc_rnd_total_{pid_key}'] = min(
                    sum(int(st.session_state.get(f'fc_rnd_{pid_key}_{k}', 0) or 0)
                        for k in _groups),
                    len(_all_names))
        except Exception:
            pass

# Подсказки «?» на этой странице нужны (в app.py они выключены глобально).
# Английскую техподсказку «Press Enter to apply» у числовых полей прячем -
# значение и так применяется по Enter или клику мимо поля.
st.markdown(
    '<style>[data-testid="stTooltipIcon"], [data-testid="stTooltipHoverTarget"] '
    '{ display: inline-flex !important; }\n'
    '[data-testid="InputInstructions"] { display: none !important; }</style>',
    unsafe_allow_html=True,
)

# ── Домены и поддомены ───────────────────────────────────────────────
# Домен = основной сайт страны (mepen.ru, mepen.kz…); поддомен = город на
# этом домене (spb.mepen.ru). Без справочника городов - только основной сайт.
if _cities:
    _example_sub = _host(_cities[1]['url']) if len(_cities) > 1 else ''
    st.subheader(
        'Домены и поддомены',
        help='Домен - основной сайт страны (например ' + _host(_mains[0]['url']) + '). '
             'Поддомен - город на этом домене (например ' + _example_sub + '). '
             'Заявка с каждого выбранного сайта должна прийти на свою почту из справочника.',
    )
    st.session_state.setdefault(f'fc_mode_{pid_key}', _MODE_OPTIONS[0])
    _mode = st.radio(
        'Что проверяем', _MODE_OPTIONS,
        horizontal=True, label_visibility='collapsed',
        key=f'fc_mode_{pid_key}',
    )

    if _mode == 'Основные домены (по странам)':
        # Одна строка на страну: галочка «Страна - домен». Галочки уже стоят.
        def _mk(country):
            return f'fc_main_{pid_key}_{country}'
        for c in _mains:                       # дефолт: все страны включены
            if _mk(c['country']) not in st.session_state:
                st.session_state[_mk(c['country'])] = True

        # Список доменов СЛЕВА, кнопка-переключатель СПРАВА (в той же строке -
        # без пустого пространства над списком). «Снять все», если всё отмечено,
        # иначе «Выбрать все».
        _all_on = all(st.session_state.get(_mk(c['country']), True) for c in _mains)
        _left, _right = st.columns([4.2, 1.3], vertical_alignment='top')
        with _right:
            if st.button('Снять все' if _all_on else 'Выбрать все',
                         use_container_width=True, key=f'fc_main_toggle_{pid_key}'):
                for c in _mains:
                    st.session_state[_mk(c['country'])] = not _all_on
                st.rerun()
        _sel = []
        with _left:
            for c in _mains:
                _lbl = (f"{_COUNTRY_FLAG.get(c['country'], '🏳')} **{c['country']}** - "
                        f"`{_host(c['url'])}`")
                _hlp = f"Главный сайт страны, город: {c['city']}."
                if c.get('mail'):
                    _hlp += f" Заявка должна прийти на {c['mail']}."
                if st.checkbox(_lbl, key=_mk(c['country']), help=_hlp):
                    _sel.append(c['city'])
        _chosen_cities = _sel
        st.caption(f'Выбрано доменов: **{len(_sel)} / {len(_mains)}**.')

    elif _mode == 'Случайные города':
        # Один экран: сверху общее число (само распределяется по странам),
        # ниже строки стран с числом справа. Правишь число у страны - общее
        # пересчитывается. Основной домен страны всегда идёт первым.
        _tkey = f'fc_rnd_total_{pid_key}'

        def _ckey(country):
            return f'fc_rnd_{pid_key}_{country}'

        def _apply_total():
            for k, v in _round_robin(st.session_state[_tkey]).items():
                st.session_state[_ckey(k)] = v
            _rnd_save()

        def _apply_country():
            st.session_state[_tkey] = sum(
                int(st.session_state.get(_ckey(k), 0) or 0) for k in _groups)
            _rnd_save()

        def _apply_recommended_rnd():
            """Рекомендованный сценарий случайной проверки: Россия - 2 (Москва +
            случайный поддомен), остальные страны - по 1 (их основные домены)."""
            for k in _groups:
                st.session_state[_ckey(k)] = min(2 if k == 'Россия' else 1,
                                                 len(_groups[k]))
            st.session_state[_tkey] = sum(
                int(st.session_state.get(_ckey(k), 0) or 0) for k in _groups)
            _rnd_save()

        if _tkey not in st.session_state:      # первичная инициализация
            st.session_state[_tkey] = min(7, len(_all_names))
            _apply_total()

        # Поле «сколько всего» и кнопка рекомендованного - в одной строке рядом,
        # кнопка выровнена по нижнему краю поля. Звезда ★ текстовая - рисуется
        # цветом текста сайта (эмодзи ⭐ всегда жёлтая).
        # ВАЖНО: значение поля _tkey нельзя менять после его отрисовки, поэтому
        # кнопка работает через колбэк on_click (он выполняется ДО отрисовки полей).
        _tc, _rec, _ = st.columns([2.2, 2.3, 1.5], vertical_alignment='bottom')
        with _tc:
            st.number_input(
                'Сколько всего доменов/поддоменов проверить',
                min_value=0, max_value=len(_all_names), step=1,
                key=_tkey, on_change=_apply_total,
                help='Введи число и нажми Enter (или кликни по пустому месту) - '
                     'оно применится и само распределится по странам ниже, начиная '
                     'с России. Число любой страны можно поправить вручную - общее '
                     'пересчитается.')
        _rec.button('★ Рекомендованный сценарий', use_container_width=True,
                    key=f'fc_rnd_rec_{pid_key}', on_click=_apply_recommended_rnd,
                    help='Поставит рекомендованные числа: Россия - 2, остальные '
                         'страны - по 1 (основной домен каждой страны + случайный '
                         'поддомен России).')

        _counts = {}
        for _country, _names in _groups.items():
            _row_l, _row_n, _row_c = st.columns([3, 1, 0.9], vertical_alignment='center')
            _dom = _host(_main_by_country.get(_country, {}).get('url', ''))
            _row_l.markdown(
                f"{_COUNTRY_FLAG.get(_country, '🏳')} **{_country}** - `{_dom}`")
            if _ckey(_country) not in st.session_state:
                st.session_state[_ckey(_country)] = 0
            _counts[_country] = int(_row_n.number_input(
                _country, min_value=0, max_value=len(_names), step=1,
                key=_ckey(_country), on_change=_apply_country,
                label_visibility='collapsed'))
            _row_c.caption(f'из {len(_names)}')

        # Сборка списка: основной домен страны первым, остальное - случайно.
        _parts = []
        for _country, _names in _groups.items():
            _k = int(_counts.get(_country, 0) or 0)
            if _k <= 0:
                continue
            _mc = _main_by_country.get(_country, {}).get('city')
            _pick = [_mc] if _mc in _names else []
            _pool = [n for n in _names if n not in _pick]
            _extra = min(max(_k - len(_pick), 0), len(_pool))
            if _extra:
                _pick += random.sample(_pool, _extra)
            _pick = _pick[:_k]
            _chosen_cities += _pick
            _parts.append(f"**{_country}**: {', '.join(_pick)}")
        if _chosen_cities:
            st.caption(f'Выбрано {len(_chosen_cities)} - ' + ' · '.join(_parts) +
                       '. Случайные города пересоберутся при запуске.')
        st.caption('Числа запоминаются для этого проекта и подставятся при '
                   'следующем заходе.')

    else:  # Выбрать города - СЕТКА ЧЕКБОКСОВ по странам
        _main_cities = {c['city'] for c in _mains}

        def _ck(city):
            return f'fc_cb_{pid_key}_{city}'
        # один раз ставим дефолт: отмечены ОСНОВНЫЕ домены каждой страны
        # (Москва, aviastal.kz, Минск…), остальные города добираются вручную.
        if not st.session_state.get(f'fc_init_osn_{pid_key}'):
            for nm in _all_names:
                st.session_state[_ck(nm)] = (nm in _main_cities)
            st.session_state[f'fc_init_osn_{pid_key}'] = True

        _b1, _b2, _ = st.columns([1, 1, 4])
        if _b1.button('Выбрать все', use_container_width=True):
            for nm in _all_names:
                st.session_state[_ck(nm)] = True
            st.rerun()
        if _b2.button('Снять все', use_container_width=True):
            for nm in _all_names:
                st.session_state[_ck(nm)] = False
            st.rerun()

        _sel = []
        for _country, _names in _groups.items():
            _dom = _host(_main_by_country.get(_country, {}).get('url', ''))
            st.markdown(
                f"**{_COUNTRY_FLAG.get(_country, '🏳')} {_country}**  ·  {len(_names)} "
                f"<span title='Основной домен: {_dom}' "
                f"style='color:#8A8782;cursor:help;font-size:.85em'>?</span>",
                unsafe_allow_html=True)
            _cols = st.columns(6)
            for _i, _nm in enumerate(_names):
                if _cols[_i % 6].checkbox(_nm, key=_ck(_nm)):
                    _sel.append(_nm)
        _chosen_cities = _sel
        st.caption(f'Выбрано: **{len(_sel)} / {len(_all_names)}** городов.')

    st.divider()

_cities_none = bool(_cities) and not _chosen_cities

# ── Формы ────────────────────────────────────────────────────────────
# Список форм проекта (в порядке прогона) уже собран выше (_all_forms).
# По умолчанию - все; можно выбрать только нужные. Имена совпадают с отчётом.
_chosen_forms = list(_all_form_names)     # по умолчанию - все формы
if _all_forms:
    st.subheader('Формы')
    _fmode = st.radio(
        'Какие формы проверяем',
        ['Все формы', 'Выбрать формы'],
        horizontal=True, label_visibility='collapsed',
        key=f'fc_forms_mode_{pid_key}',
    )
    if _fmode == 'Выбрать формы':
        def _fk(name):
            return f'ff_cb_{pid_key}_{name}'
        # дефолт - все отмечены (один раз на проект)
        if not st.session_state.get(f'ff_init_{pid_key}'):
            for nm in _all_form_names:
                st.session_state[_fk(nm)] = True
            st.session_state[f'ff_init_{pid_key}'] = True

        _fb1, _fb2, _ = st.columns([1, 1, 4])
        if _fb1.button('Выбрать все', use_container_width=True, key=f'ff_all_{pid_key}'):
            for nm in _all_form_names:
                st.session_state[_fk(nm)] = True
            st.rerun()
        if _fb2.button('Снять все', use_container_width=True, key=f'ff_none_{pid_key}'):
            for nm in _all_form_names:
                st.session_state[_fk(nm)] = False
            st.rerun()

        # Группировка по СТРАНИЦЕ сайта (без техназваний: Главная_расчёты и
        # Главная - это одна «Главная»). Всё в карточке с рамкой, между
        # группами тонкая линия. Подсказка «?» - где форма НА странице
        # (шапка/подвал/кнопка-окно), только если она есть в _FORM_WHERE.
        def _pg_group(page):
            return (page or '').split('_')[0] or 'Прочее'

        def _disp_name(nm):
            """Имя формы для показа: без уточнения в скобках на конце
            («Обратная связь (Контакты)» - показываем «Обратная связь»,
            уточнение живёт в подсказке «?»). В отчёт и фильтр идёт полное имя."""
            _b = re.sub(r'\s*\([^()]*\)\s*$', '', nm).strip()
            return _b or nm

        _by_group = {}
        for f in _all_forms:
            _by_group.setdefault(_pg_group(f['page']), []).append(f)
        _sel_f = []
        with st.container(border=True):
            for _gi, (_grp, _items) in enumerate(_by_group.items()):
                if _gi:
                    st.markdown(
                        '<hr style="margin:4px 0 10px 0; border:none; '
                        'border-top:1px solid #ECEAE4">',
                        unsafe_allow_html=True)
                _gcol, _ccol = st.columns([1.1, 4.9], vertical_alignment='top')
                # небольшой отступ сверху, чтобы название страницы стояло
                # ровно по первой строке галочек
                _gcol.markdown(
                    f"<div style='padding-top:.4rem;font-weight:600'>{_grp}</div>",
                    unsafe_allow_html=True)
                _fcols = _ccol.columns(3)
                for _i, _f in enumerate(_items):
                    _hint = _FORM_WHERE.get(_f['name'])
                    if _fcols[_i % 3].checkbox(
                            _disp_name(_f['name']), key=_fk(_f['name']),
                            help=_hint):
                        _sel_f.append(_f['name'])
        _chosen_forms = _sel_f
        st.caption(f'Выбрано форм: **{len(_chosen_forms)} / {len(_all_forms)}**.')
    else:
        st.caption(f'Будут проверены все формы проекта: {len(_all_forms)}.')
# Сохранение шаблона теперь происходит ПРЯМО по клику «Сохранить» в блоке
# «Проектные шаблоны» вверху (через _tpl_collect(), без st.rerun()) - надёжно и
# без сброса галочек. Здесь, внизу, собирать больше нечего.

st.divider()

_forms_all_selected = (len(_chosen_forms) == len(_all_form_names))
_forms_none = bool(_all_forms) and len(_chosen_forms) == 0

# ── Запуск ──────────────────────────────────────────────────────────
# ── Проверка админки: отдельный блок, как у форм (проверять/не проверять) ──
# Логин/пароль вводятся здесь и передаются проверке через окружение - на диск
# ничего не пишется и никуда не отправляется.
_admin_env: dict[str, str] = {}
_admin_on = True
if _project_has_admin(pid_key):
    st.subheader('Проверка админки')
    st.session_state.setdefault(f'fc_admin_on_{pid_key}', True)
    _admin_on = st.checkbox('Проверять, что заявки попали в админку',
                            key=f'fc_admin_on_{pid_key}')
    if _admin_on:
        st.caption(
            'После прогона тест зайдёт в «Уведомления с форм» и в отчёте (лист '
            '«Логи», колонка «Статус в админке») отметит, какие заявки реально '
            'долетели. Логин/пароль нужны только для входа - нигде не сохраняются '
            'и никуда не отправляются.')
        _al = st.text_input('Логин админки', key=f'fc_admin_login_{pid_key}')
        _ap = st.text_input('Пароль админки', type='password',
                            key=f'fc_admin_pass_{pid_key}')
        if _al.strip() and _ap:
            _admin_env = {'ADMIN_LOGIN': _al.strip(), 'ADMIN_PASSWORD': _ap}
        else:
            st.caption('⚠️ Введите логин и пароль - без них админка не проверится.')

# ── Проверка письма о заказе покупателю (пункт 2.9) ──────────────────
# Заказ оформляется на РЕАЛЬНУЮ почту (свою рабочую), чтобы письмо-подтверждение
# реально пришло. Пока один режим - «Своя почта»: заказ уйдёт на неё, письмо
# проверяешь глазами. Автопроверку по IMAP временно убрали.
# Без выбора - как раньше: заказ на адрес из конфига, письмо не проверяем.
_mail_env: dict[str, str] = {}
st.subheader('Проверка письма о заказе')
st.caption('Относится к оформлению заказа через корзину: на какую почту оформить '
           'заказ и как проверить письмо-подтверждение покупателю.')
_MAIL_MODES = ['Не проверять', 'Своя почта (проверю письмо вручную)']
_mail_mode_key = f'fc_mail_mode_{pid_key}'
st.session_state.setdefault(_mail_mode_key, _MAIL_MODES[0])
# Раньше был режим «Тестовая почта» - если он ещё лежит в session_state открытой
# сессии, сбрасываем на «Не проверять», иначе st.radio упадёт (значения нет в списке).
if st.session_state.get(_mail_mode_key) not in _MAIL_MODES:
    st.session_state[_mail_mode_key] = _MAIL_MODES[0]
_mail_mode = st.radio('Как проверяем письмо покупателю', _MAIL_MODES,
                      key=_mail_mode_key, label_visibility='collapsed')

if _mail_mode == _MAIL_MODES[1]:
    st.caption('Заказ оформится на указанную почту. Автопроверку не делаем - '
               'просто открой этот ящик и убедись, что письмо о заказе пришло. '
               'В отчёте (колонка «Письмо покупателю») будет напоминание с адресом.')
    _m_own = st.text_input('Ваша почта (куда придёт письмо о заказе)',
                           key=f'fc_mail_own_{pid_key}', placeholder='ваша@почта.ru')
    if _m_own.strip():
        _mail_env = {'ORDER_BUYER_EMAIL': _m_own.strip()}
    else:
        st.caption('⚠️ Укажите почту - иначе заказ уйдёт на тестовый адрес из конфига.')

st.subheader(
    'Запуск проверки',
    help='Формы проверяются автоматически: по коду (быстро, requests), где это '
         'возможно; где форму по коду не видно (её рисует/шлёт JS) или включены '
         'браузерные пробы ниже - через браузер. Отдельно ничего включать не нужно. '
         'Запуск фоновый - интерфейс сразу свободен, можно уйти в другие вкладки. '
         'Браузер работает скрыто (headless). Заявки отправляются по-настоящему '
         '(формы оформления заказа - без отправки).\n\n'
         'Наличие капчи и honeypot-поля на форме проверяется ВСЕГДА, без галочек '
         'и без единого лишнего запроса - смотри колонки «CSRF-защита», «Данные '
         'дошли до сервера» и «Защита от спама (пассивно)» в отчёте.')

st.session_state.setdefault(f'fc_clear_{pid_key}', True)
clear_log = st.checkbox(
    'Очищать лог Excel перед прогоном', key=f'fc_clear_{pid_key}',
    help='Перед новым прогоном очищает Excel-лог, чтобы результаты не смешивались '
         'со старыми. Снимите - если хотите дописать к прошлому логу.')
st.session_state.setdefault(f'fc_fileprobe_{pid_key}', False)
file_probe = st.checkbox(
    'Проба серверной фильтрации загрузки файлов', key=f'fc_fileprobe_{pid_key}',
    help='Только для форм с полем загрузки файлов. По очереди грузит безвредный '
         'файл 32 типов (в т.ч. опасные: .php/.exe/.js/.bat/.sh…) и разрешённый '
         'тип больше 20 КБ - проверяет, что сервер принимает только PDF/DOC/DOCX '
         'до 20 КБ, а лишнее (посторонний формат ИЛИ больше 20 КБ) отклоняет. '
         'Колонка «Типы файлов формы»: «✗» = принят лишний тип или файл >20 КБ. '
         'ВНИМАНИЕ: каждый принятый тип = отдельная тест-заявка в админке '
         'боевого сайта.')
st.session_state.setdefault(f'fc_xss_{pid_key}', False)
xss_probe = st.checkbox(
    'Проверка защиты от XSS', key=f'fc_xss_{pid_key}',
    help='В поле имени уходит безвредный XSS-маркер; после отправки проверяется, '
         'исполнился ли он на странице-ответе (колонка «Защита от XSS»: Защищена / '
         'УЯЗВИМА). Маркер компьютеру и данным не вредит. ВНИМАНИЕ: создаётся '
         'тест-заявка с маркером в имени - после прогона удалите её в админке.')
st.session_state.setdefault(f'fc_srvval_{pid_key}', False)
srv_val_probe = st.checkbox(
    'Проба серверной валидации', key=f'fc_srvval_{pid_key}',
    help='Проверяет «нельзя отправить неверные данные в обход браузера»: повторно '
         'шлёт форму через JS без клиентских ограничений (обязательность/формат/'
         'длину) с заведомо неверными данными (пустое поле / без имени / без '
         'телефона / e-mail без «@» / слишком длинный текст). Колонка «Серверная '
         'валидация»: Защищена / УЯЗВИМА. Формы оформления заказа пропускаются. '
         'ВНИМАНИЕ: сервер может принять до 5 доп. тест-заявок на форму - после '
         'прогона удалите их в админке.')
st.session_state.setdefault(f'fc_ratelimit_{pid_key}', False)
rate_limit_probe = st.checkbox(
    '🧪 Активная проверка лимита запросов', key=f'fc_ratelimit_{pid_key}',
    help='Отправляет по форме 3 тестовые заявки подряд за несколько секунд - '
         'проверяет, останавливает ли сайт быстрые повторы, то есть реально ли '
         'СРАБАТЫВАЕТ защита от спама/ботов (а не только «есть капча»). ВНИМАНИЕ: '
         'если защиты нет, все 3 попадут в админку как настоящие заявки (удалить '
         'вручную) - каждый запуск это ещё 3 заявки. Число попыток жёстко '
         'ограничено - 3. Формы оформления заказа не трогает.')

_alive = _pid_alive(_read_pid())


@st.cache_resource(show_spinner=False)
def _ensure_browser_forms():
    import browser_setup
    return browser_setup.ensure_browser()


# В облаке при первом заходе доустанавливаем Chromium (до минуты). Локально -
# мгновенно (браузер уже стоит). Пока идёт прогон - не трогаем.
_browser_ok = True
if not _alive:
    with st.spinner('Готовлю браузер (первый запуск в облаке - до минуты)…'):
        _browser_ok, _browser_msg = _ensure_browser_forms()

if _forms_none:
    st.warning('Не выбрано ни одной формы - отметь хотя бы одну, чтобы запустить.')
if _cities_none:
    st.warning('Не выбрано ни одного домена/города - отметь хотя бы один, чтобы запустить.')

# Прокси + проверка доступности сайта (над кнопкой запуска)
_forms_proxy = None
try:
    from site_access import render_proxy_access
    _forms_proxy = render_proxy_access(
        f'forms_{pid_key}',
        default_url=f"https://{proj['domain']}/", pid=pid_key,
        default_on=_project_uses_proxy(pid_key))
except Exception as _e_pa:
    st.caption(f'⚠ Блок прокси/доступа не загрузился: {_e_pa}')

_run_col, _cancel_col = st.columns([3, 1])
with _run_col:
    if st.button('▶ Запустить проверку', use_container_width=True, type='primary',
                 disabled=_alive or _forms_none or _cities_none or not _browser_ok):
        ready, _missing = _deps_ready()
        if not ready:
            # Движка нет в этом окружении (типично для облака по ссылке) -
            # не запускаем, показываем понятную инструкцию ниже. Заодно сбрасываем
            # прогресс и старый лог, чтобы не висел результат прошлого запуска.
            st.session_state['forms_dep_error'] = _missing
            st.session_state.pop('forms_started', None)
            st.session_state.pop('forms_started_ts', None)
            try:
                LOG_FILE.unlink(missing_ok=True)
            except Exception:
                pass
            st.rerun()
        else:
            st.session_state.pop('forms_dep_error', None)
            args = ['forms_run.py', '--project', pid_key]
            if not clear_log:
                args.append('--no-clear-excel')
            if _project_has_admin(pid_key) and not _admin_on:
                args.append('--no-admin')      # админку явно отключили галочкой
            if file_probe:
                args.append('--file-probe')    # проба фильтрации загрузки файлов
            if xss_probe:
                args.append('--xss-probe')     # проба защиты от XSS (маркер в имени)
            if srv_val_probe:
                args.append('--server-validation-probe')  # проба серверной валидации
            if rate_limit_probe:
                args.append('--rate-limit-probe')  # активная проверка лимита запросов
            if _chosen_cities:
                args += ['--cities', ','.join(_chosen_cities)]
            # Фильтр форм: передаём только если выбрано подмножество (не все).
            if _all_forms and not _forms_all_selected and _chosen_forms:
                import json
                _ff = ROOT / 'cache' / 'forms' / pid_key / '_forms_filter.json'
                _ff.parent.mkdir(parents=True, exist_ok=True)
                _ff.write_text(json.dumps(_chosen_forms, ensure_ascii=False),
                               encoding='utf-8')
                args += ['--forms-file', str(_ff)]
            try:
                LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
                LOG_FILE.write_text('', encoding='utf-8')
            except Exception:
                pass
            # Прокси (галочка «Вкл. Прокси» над кнопкой): пробрасываем в движок,
            # чтобы браузер форм ходил через него - для сайтов, режущих прямое
            # подключение (напр. Метпромко). Пусто = браузер идёт напрямую.
            _proxy_env = {'FORMS_PROXY': _forms_proxy} if _forms_proxy else {}
            # Telegram: креды из секретов (те же, что у еженедельной проверки) -
            # forms_run сам отправит отчёт в чат после прогона.
            try:
                import tg_report
                _tg_env = tg_report.runner_env(pid_key)
            except Exception:
                _tg_env = {}
            _launch_background(args, LOG_FILE,
                               extra_env={**_admin_env, **_mail_env, **_proxy_env, **_tg_env})
            st.session_state['forms_started'] = datetime.now().strftime('%H:%M:%S')
            st.session_state['forms_started_ts'] = time.time()
            st.session_state['forms_project'] = pid_key
            st.session_state['forms_cities_n'] = max(1, len(_chosen_cities))
            # Ожидаемое число форм для шкалы прогресса (учитывает выбор форм).
            _per_city = (len(_chosen_forms) if (_all_forms and not _forms_all_selected)
                         else _count_expected(pid_key))
            st.session_state['forms_expected_total'] = _per_city * max(1, len(_chosen_cities))
            st.rerun()
with _cancel_col:
    if st.button('⛔ Отменить', use_container_width=True, disabled=not _alive):
        _kill_tree(_read_pid())
        try:
            PID_FILE.unlink(missing_ok=True)
            with open(LOG_FILE, 'a', encoding='utf-8') as _f:
                _f.write('\n⛔ ОТМЕНЕНО пользователем\n')
        except Exception:
            pass
        st.rerun()

# ── Понятная ошибка: движок не установлен (показываем ТОЛЬКО после клика) ──
if not _alive and st.session_state.get('forms_dep_error'):
    st.error('Не получилось запустить проверку - в этом окружении нет нужных библиотек.')
    st.markdown(
        'Обычно так бывает, если приложение развёрнуто без браузерных зависимостей. '
        'В облаке (Streamlit Cloud) нужны `playwright`, `requests`, `beautifulsoup4` '
        'в `requirements.txt` и системные библиотеки Chromium в `packages.txt` - '
        'после перезапуска приложения браузер доустановится сам.\n\n'
        '**Локально** (на своём компьютере):\n'
        '1. `pip install -r requirements-local.txt`\n'
        '2. `playwright install chromium`\n'
        '3. `streamlit run app.py`, затем снова «Запустить проверку».'
    )
if not _alive and not _browser_ok:
    st.warning(f'Браузер ещё не готов: {_browser_msg}. Если это первый запуск в '
               'облаке - подождите минуту и обновите страницу; иначе проверьте '
               'requirements.txt (playwright) и packages.txt.')

st.divider()

# ── Прогресс ────────────────────────────────────────────────────────
# Всё в этой секции привязано к ВЫБРАННОМУ проекту (pid_key): при смене
# проекта статус/прогресс/лог/скачивание обновляются под него.
st.subheader('Прогресс')

_sel = pid_key
# Факт «идёт проверка» берём из ПЕР-ПРОЕКТНОГО PID/лога - его видят все сессии
# (в т.ч. с другого компьютера). _own = запускали ли ИМЕННО в этой сессии -
# нужно только для секундомера/ETA и подписи «Последний запуск».
_own = (st.session_state.get('forms_project') == _sel)
xlsx = ROOT / 'cache' / 'forms' / _sel / 'log_forms.xlsx'

# Готовность определяем ПО ЛОГУ (движок пишет «✅ ВСЁ ГОТОВО»), а не только по
# «жив ли процесс»: в облаке PID может «висеть» после завершения, из-за чего
# прогресс-бар не переключался на «готово». Живой лог - надёжный признак.
_log_txt = LOG_FILE.read_text(encoding='utf-8', errors='ignore') if LOG_FILE.exists() else ''
_done_by_log = ('✅ ВСЁ ГОТОВО' in _log_txt or 'ОТМЕНЕНО' in _log_txt
                or '✗ Ошибка' in _log_txt)

# Цели тут НЕ трогаем: «Проверка форм» проверяет только формы. Цели Метрики (в
# т.ч. заказ-цели) проверяет и подтягивает «Проверка целей» - она сама прогоняет
# сквозной заказ. Раньше здесь итоги форм писались в cache/forms/<project>/ для
# отчёта целей - убрано, чтобы формы и цели были строго раздельными разделами.


def _forms_done_live(txt: str) -> int:
    """Сколько форм уже отработало - по строкам «▶ Форма N: …», которые движок
    печатает РОВНО ОДИН РАЗ на форму/сценарий/модалку. Раньше считали разнородные
    «УСПЕШНО/ОШИБКА»/«поля заполнены» - их на одну форму приходится несколько, и
    счётчик раздувался (10 форм → «108»)."""
    n = 0
    for ln in txt.splitlines():
        if ln.lstrip().startswith('▶ Форма '):
            n += 1
    return n


if _own and st.session_state.get('forms_started'):
    st.caption(f'Последний запуск: {st.session_state["forms_started"]}')

if _alive and not _done_by_log:
    # Идёт проверка ВЫБРАННОГО проекта (запущена в этой ИЛИ в другой сессии /
    # на другом компьютере - per-проектный PID/лог видят все). Секундомер и ETA
    # считаем только своей сессии (_own); чужой прогон - прогресс по логу.
    if not _own:
        st.info('▶ Проверку этого проекта запустили в другой сессии (на другом '
                'компьютере). Здесь виден живой прогресс по логу; секундомер и '
                'остаток считаются у того, кто запустил.')
    _ts = st.session_state.get('forms_started_ts') if _own else None
    _elapsed = int(time.time() - _ts) if _ts else None
    _mmss = f'{_elapsed // 60}:{_elapsed % 60:02d}' if _elapsed is not None else '…'
    # Прогресс по живому логу (обновляется сразу); подстраховка - строки Excel.
    _done = _forms_done_live(_log_txt)
    _xl = _rows_done(xlsx)
    if _xl and _xl > _done:
        _done = _xl
    _total = (st.session_state.get('forms_expected_total') if _own else 0) \
        or _count_expected(_sel) * (st.session_state.get('forms_cities_n', 1) if _own else 1)
    # Оценка приблизительная (кроме форм в прогон попадают проверки 2.13/2.12,
    # поля, оформление) - реальный счётчик может её превысить. Не показываем
    # «47 из ~40»: подтягиваем «всего» минимум до фактически проверенного.
    if _total and _done > _total:
        _total = _done

    # Финальная стадия: формы отработали, идут доводочные проверки (cookie 2.12,
    # мобильная вёрстка) и СБОРКА отчёта. Раньше прогресс «висел» на 99% с текстом
    # «Идёт проверка», будто зависло, - показываем честно «формирую отчёт».
    _finalizing = any(m in _log_txt for m in (
        '🧹 Отчёт сведён', '🗂️ Матрица проверок построена',
        'Проверка 2.12', 'Мобильная вёрстка'))
    _eta_txt = ''
    if _finalizing:
        st.progress(0.98, text='Почти готово: формирую отчёт…')
    elif _total:
        st.progress(min(_done / max(_total, 1), 0.99),
                    text=f'Проверено форм: {_done} из ~{_total}')
        # Живая оценка остатка по фактическому темпу - честнее фиксированных
        # «20-30 мин» (реальный прогон сильно зависит от числа форм и городов).
        if _done >= 3 and _elapsed:
            _rate = _elapsed / _done                     # секунд на форму
            _rem = int(_rate * max(_total - _done, 0))
            if _rem > 0:
                _eta_txt = (f' Осталось ~{_rem // 60}:{_rem % 60:02d} (мин:сек) '
                            'при текущем темпе.')
    else:
        st.progress(min(0.95, (_elapsed or 0) / 90.0), text='Идёт проверка…')

    st.caption(f'⏳ Идёт… {_mmss}.{_eta_txt} Каждую форму движок реально заполняет '
               'и отправляет несколько раз (обычная отправка, двойной клик, '
               'серверная валидация, антиспам) плюс ~25 проверок - поэтому на '
               'много форм/городов уходит много времени; при блокировках сайта / '
               'капче / сбросах соединения - ещё дольше. Страница обновляется '
               'сама - можно уйти на другие вкладки, прогон не прервётся.')
    with st.expander('Подробный лог', expanded=False):
        st.code('\n'.join(_log_txt.splitlines()[-300:]) or '…', language='text')
    time.sleep(2)
    st.rerun()

elif _alive and _done_by_log:
    # Процесс ещё числится живым (в облаке PID «висит»), но лог говорит «готово».
    # Снимаем зависший PID и перерисовываем - дальше сработает ветка результата.
    try:
        PID_FILE.unlink(missing_ok=True)
    except Exception:
        pass
    st.rerun()

else:
    # Ничего не идёт для выбранного проекта. Лог/результат берём из per-проектных
    # файлов - их видит любая сессия (не только та, что запускала).
    if LOG_FILE.exists() and LOG_FILE.read_text(encoding='utf-8', errors='ignore').strip():
        # Итоговое время прогона = последняя запись лога минус старт (только своя сессия).
        _st_ts = st.session_state.get('forms_started_ts') if _own else None
        _fin_ts = LOG_FILE.stat().st_mtime
        _spent = int(_fin_ts - _st_ts) if _st_ts else None
        _spent_txt = f' · ⏱ заняло {_spent // 60}:{_spent % 60:02d}' if _spent and _spent > 0 else ''
        st.markdown(f'**Статус:** ✅ завершено / остановлено{_spent_txt}')
        with st.expander('Подробный лог', expanded=False):
            st.code('\n'.join(LOG_FILE.read_text(encoding='utf-8', errors='ignore')
                              .splitlines()[-300:]), language='text')
        st.download_button(
            '⬇ Скачать лог (txt)',
            data=LOG_FILE.read_bytes(),
            file_name=f'{_sel}-log.txt',
            mime='text/plain',
            use_container_width=True,
        )
    else:
        st.caption('Лог появится после запуска.')

    # ── Результат: Excel выбранного проекта (имя файла = Проект-Дата) ──
    if xlsx.exists():
        st.divider()
        st.subheader('Результаты (Excel)')
        _date = datetime.fromtimestamp(xlsx.stat().st_mtime).strftime('%d.%m.%Y')
        _fname = f'{_sel.capitalize()}-{_date}.xlsx'   # напр. Mpe-23.06.2026.xlsx
        st.caption(f'Лог проекта {PROJECTS[_sel]["name"]} '
                   '- дата, страница, форма, статус и комментарий с причиной (если не сработало).')
        st.download_button(
            f'⬇ Скачать {_fname}',
            data=xlsx.read_bytes(),
            file_name=_fname,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True,
        )
