"""
Страница «Проверка КП» - проверка вывода данных поддоменов по «Карте
присутствия» (пункт 1.4 чек-листа).

Для выбранного проекта фоново качает главные страницы поддоменов и сверяет с КП
(catalogs/{proj}-kp.csv): город/страна, телефоны (поиск/реклама/общий), почта,
адрес, Telegram, WhatsApp. Результат - Excel «Проверка КП» + «Расхождения»;
если включали сверку с Яндекс.Картами/2ГИС - те же колонки на том же листе.
Сделана по образцу страницы «Проверка форм» (фоновый процесс variables_run.py).
"""
import importlib.util
import json
import os
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

import streamlit as st

import site_access

ROOT = Path(__file__).parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from checklists import page_templates as _tpl
from checklists import ui_widgets as _ui
PY = sys.executable
LOG_FILE = ROOT / 'cache' / 'variables.log'
PID_FILE = ROOT / 'cache' / 'variables.pid'

PROJECTS = {
    'smu': 'СМУ - Стальметурал', 'imp': 'ИМП - Инметпром',
    'mpe': 'МПЭ - Мепэн', 'avia': 'АПС - Авиапромсталь',
    'mpk': 'МПК - Метпромко', 'mpi': 'МПИ - МетПромИнтекс',
    'sm': 'SM - SHOPMET',
}


def _secret(key: str, default: str = '') -> str:
    try:
        if hasattr(st, 'secrets') and key in st.secrets:
            return str(st.secrets[key])
    except Exception:
        pass
    return default


def _sa_json_for_env() -> str:
    """JSON-строка ключа сервисного аккаунта из секретов (TOML-секция, строка-JSON
    или base64) - чтобы прокинуть её в фоновый процесс через GCP_SA_JSON. '' если
    ключ не задан / не разобрался."""
    try:
        v = st.secrets.get('gcp_service_account')
    except Exception:
        v = None
    if v is not None:
        try:
            if isinstance(v, str):
                json.loads(v)          # проверяем, что это валидный JSON
                return v
            return json.dumps(dict(v))
        except Exception:
            pass
    try:
        b = st.secrets.get('gcp_service_account_b64')
    except Exception:
        b = None
    if b:
        try:
            import base64
            return base64.b64decode(''.join(str(b).split())).decode('utf-8')
        except Exception:
            pass
    return ''


def _sa_configured() -> bool:
    """Задан ли ключ сервисного аккаунта в секретах (в любом из форматов)?"""
    for k in ('gcp_service_account', 'gcp_service_account_b64'):
        try:
            if st.secrets.get(k):
                return True
        except Exception:
            pass
    return False


def _load_kp_cities(project: str):
    """Список (город, домен, страна) из КП проекта - для выбора поддоменов."""
    p = ROOT / 'catalogs' / f'{project}-kp.csv'
    if not p.exists():
        return []
    import csv
    out = []
    with p.open(encoding='utf-8') as f:
        for row in csv.DictReader(f):
            if row.get('domain'):
                out.append((row.get('city', ''), row['domain'], row.get('country', '')))
    return out


def _read_pid():
    try:
        return int(PID_FILE.read_text().strip())
    except Exception:
        return None


def _pid_alive(pid) -> bool:
    if not pid:
        return False
    if os.name == 'nt':
        try:
            out = subprocess.run(['tasklist', '/FI', f'PID eq {pid}'],
                                 capture_output=True, text=True).stdout
            return str(pid) in out
        except Exception:
            return False
    try:
        os.kill(pid, 0)
        return True
    except Exception:
        return False


def _kill(pid):
    if not pid:
        return
    if os.name == 'nt':
        subprocess.run(['taskkill', '/F', '/T', '/PID', str(pid)], capture_output=True)
    else:
        import signal
        try:
            os.kill(pid, signal.SIGTERM)
        except Exception:
            pass


def _launch(args, extra_env=None):
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    env = dict(os.environ)
    env['PYTHONIOENCODING'] = 'utf-8'
    env['PYTHONUNBUFFERED'] = '1'
    if extra_env:
        env.update({k: v for k, v in extra_env.items() if v})
    flags = getattr(subprocess, 'CREATE_NO_WINDOW', 0) if os.name == 'nt' else 0
    f = open(LOG_FILE, 'a', encoding='utf-8')
    proc = subprocess.Popen([PY, *args], cwd=str(ROOT), stdout=f,
                            stderr=subprocess.STDOUT, env=env, creationflags=flags)
    try:
        PID_FILE.write_text(str(proc.pid), encoding='utf-8')
    except Exception:
        pass
    return proc.pid


def _deps_ready():
    for mod in ('bs4', 'openpyxl'):  # загрузка на http.client (stdlib)
        if importlib.util.find_spec(mod) is None:
            return False
    return True


# ── Заголовок ────────────────────────────────────────────────────────
st.markdown(
    """<style>
    [data-testid="stDownloadButton"] button { background:#1E8E3E !important;
        border:1px solid #1E8E3E !important; }
    [data-testid="stDownloadButton"] button * { color:#FFF !important; }
    </style>""", unsafe_allow_html=True)

st.title('🗺️ Проверка КП (Карта присутствия)')
st.caption('Пункт 1.4: сверяем данные на поддоменах с «Картой присутствия» (КП) - '
           'город, страна, телефоны (поиск/реклама/общий), почта, адрес, Telegram, '
           'WhatsApp. Скачивает главные страницы и сравнивает с catalogs/*-kp.csv.')

# Персист выбора проекта между вкладками (Streamlit чистит состояние виджета при
# уходе на другую страницу - иначе проект «сбрасывался»). Держим в постоянном
# ключе session_state и восстанавливаем как index; виджет без key.
_kp_opts = list(PROJECTS.keys())
_kp_saved = st.session_state.get('kp_project_sel')
_kp_idx = _kp_opts.index(_kp_saved) if _kp_saved in _kp_opts else None
pid_key = st.selectbox('Проект', _kp_opts,
                       format_func=lambda k: PROJECTS[k], index=_kp_idx,
                       placeholder='- выберите проект -')
st.session_state['kp_project_sel'] = pid_key
if not pid_key:
    st.info('Выберите проект, чтобы запустить проверку КП.')
    st.stop()

_cities = _load_kp_cities(pid_key)
if not _cities:
    st.warning(f'Нет базы КП catalogs/{pid_key}-kp.csv - проверять не с чем.')
    st.stop()

_by_country = {}
for city, dom, country in _cities:
    _by_country.setdefault(country or 'Прочее', []).append(city)
st.caption(f'В КП проекта: **{len(_cities)}** поддоменов, стран: '
           f'{len([c for c in _by_country if c])}.')


# ── Проектные шаблоны ────────────────────────────────────────────────
def _vc_tpl_apply(_tpl_data):
    # Шаблон уже проставил vc_mode + список городов (vc_ms_*) в session_state.
    # Помечаем инициализацию выполненной, иначе дефолт «отметить всё» перетёр
    # бы выбор из шаблона.
    st.session_state[f'vc_init_{pid_key}'] = True


# save_keys - те же ключи, что раньше собирал commit_pending внизу; сохраняем
# ПРЯМО по клику «Сохранить» (значения виджетов уже в session_state).
def _vc_tpl_reset():
    # Сброс к стандартным: охват → «Все домены+поддомены», выбор городов заново
    # (снимаем флаг инициализации и список выбранных городов).
    for _k in (f'vc_mode_{pid_key}', f'vc_init_{pid_key}', f'vc_ms_{pid_key}'):
        st.session_state.pop(_k, None)


_tpl.render_panel(
    'variables', pid_key, on_apply=_vc_tpl_apply, on_reset=_vc_tpl_reset,
    save_keys=lambda: [f'vc_mode_{pid_key}', f'vc_ms_{pid_key}'],
    help_text='Шаблон запоминает охват (все домены / выбранные города) и какие '
              'города отмечены. Хранится на сервере проекта **до перезапуска '
              'приложения** - после может сброситься.')

# ── Источник данных (Карта присутствия) ──────────────────────────────
# Если задана ссылка на Google-таблицу КП - при запуске проверка сама тянет
# свежие данные из таблицы; здесь же можно обновить снапшот вручную.
_kp_url = ''
try:
    import kp_sheets as _ks
    _kp_url = _ks.kp_sheet_url(pid_key)
except Exception:
    _ks = None
_kp_csv = ROOT / 'catalogs' / f'{pid_key}-kp.csv'
with st.expander('📄 Источник данных (Карта присутствия)', expanded=False):
    _upd = (datetime.fromtimestamp(_kp_csv.stat().st_mtime).strftime('%d.%m.%Y %H:%M')
            if _kp_csv.exists() else '–')
    _has_url = bool(_kp_url and _ks)
    _has_key = _sa_configured()
    # Наглядный статус источника: зелёная галочка / красный крестик.
    st.markdown(
        (f"{'✅' if _has_url else '❌'} Ссылка на Google-таблицу КП – "
         f"{'задана' if _has_url else 'не задана'}  \n"
         f"{'✅' if _has_key else '❌'} Ключ сервисного аккаунта – "
         f"{'задан в секретах' if _has_key else 'не задан'}"))
    if _has_url:
        st.caption('КП тянется напрямую из Google-таблицы при каждом запуске проверки – '
                   'свежие правки из таблицы подхватываются автоматически.')
        st.caption('⏳ Поменял(а) данные в Google-таблице – подожди ~2 минуты, '
                   'потом жми «Обновить». Google отдаёт экспорт таблицы с задержкой '
                   '(кэширует на пару минут): нажмёшь сразу – подтянутся ещё старые '
                   'данные.')
        if st.button('↻ Обновить КП из Google сейчас', key=f'kp_refresh_{pid_key}'):
            with st.spinner('Читаю таблицу КП из Google…'):
                _ok, _msg = _ks.refresh_project(pid_key, log=lambda *a, **k: None)
            # Итог кладём в session_state и перерисовываем - иначе st.rerun()
            # стирал плашку и было «непонятно, произошло что-то или нет».
            st.session_state[f'kp_msg_{pid_key}'] = (bool(_ok), str(_msg))
            st.rerun()
        _kp_msg = st.session_state.get(f'kp_msg_{pid_key}')
        if _kp_msg:
            (st.success if _kp_msg[0] else st.error)(
                ('✅ ' if _kp_msg[0] else '❌ ') + f'Обновление КП: {_kp_msg[1]}')
    else:
        st.caption(f'Пока используется зашитый снапшот `{_kp_csv.name}` (обновлён {_upd}). '
                   f'Чтобы тянуть свежее из таблицы, задай секрет '
                   f'`kp_sheet_url_{pid_key}`.')

st.divider()
st.subheader('Что проверяем')

# Города/страны - сетка чекбоксов по странам, как на «Проверке форм».
_COUNTRY_FLAG = {
    'Россия': '🇷🇺', 'Казахстан': '🇰🇿', 'Беларусь': '🇧🇾', 'Кыргызстан': '🇰🇬',
    'Киргизия': '🇰🇬', 'Узбекистан': '🇺🇿', 'Азербайджан': '🇦🇿', 'Армения': '🇦🇲',
    'Украина': '🇺🇦',
}
# Города в порядке КП, без повторов (ключи чекбоксов должны быть уникальны).
_all_city_names, _vc_groups, _seen = [], {}, set()
for _c, _d, _country in _cities:
    if _c in _seen:
        continue
    _seen.add(_c)
    _all_city_names.append(_c)
    _vc_groups.setdefault(_country or 'Прочее', []).append(_c)

_mode = st.radio('Охват', ['Все домены+поддомены', 'Выбрать города'],
                 horizontal=True, label_visibility='collapsed',
                 key=f'vc_mode_{pid_key}')
_chosen = []
if _mode == 'Выбрать города':
    _ms_key = f'vc_ms_{pid_key}'
    # Дефолт (один раз на проект): отмечены все города. Дальше правки сохраняются.
    if not st.session_state.get(f'vc_init_{pid_key}'):
        st.session_state[_ms_key] = list(_all_city_names)
        st.session_state[f'vc_init_{pid_key}'] = True

    # Подпись с флагом+страной для читаемости в раскрывающемся списке -
    # реальное значение (что уходит в --cities) остаётся голым именем города.
    _city_label = {_nm: f"{_COUNTRY_FLAG.get(_country, '🏳')} {_nm}"
                  for _country, _names in _vc_groups.items() for _nm in _names}

    # Строка выбора (чипы) растягивается по высоте, не обрезается/не
    # наезжает на остальной контент - общий CSS-фикс, см. ui_widgets.py.
    _ui.multiselect_grows_css()

    _b1, _b2, _ = st.columns([1, 1, 4])
    if _b1.button('Выбрать все', use_container_width=True, key=f'vc_all_{pid_key}'):
        st.session_state[_ms_key] = list(_all_city_names)
        st.rerun()
    if _b2.button('Снять все', use_container_width=True, key=f'vc_none_{pid_key}'):
        st.session_state[_ms_key] = []
        st.rerun()

    _chosen = st.multiselect(
        'Города', options=_all_city_names, format_func=lambda c: _city_label.get(c, c),
        key=_ms_key, placeholder='Выберите города…', label_visibility='collapsed')
    st.caption(f'Выбрано: **{len(_chosen)} / {len(_all_city_names)}** городов.')
else:
    st.caption(f'Будут проверены все {len(_cities)} доменов и поддоменов проекта.')

st.divider()
st.markdown('**Источники для сверки**')
st.caption('Можно снять «Сайт», если нужны только карты - тогда страницы '
          'сайта вообще не скачиваются, прогон короче.')
_mc1, _mc2, _mc3, _mc4 = st.columns(4)
_check_site = _mc1.checkbox(
    'Сайт (сверка с КП)', value=True, key=f'vc_check_site_{pid_key}',
    help='Скачивает главные страницы поддоменов, сверяет с КП. Это основная '
         'проверка страницы - снимай, только если сейчас нужны исключительно '
         'карты, без затрат времени на сайт.')
_check_yandex_maps = _mc2.checkbox(
    'Яндекс.Карты', key=f'vc_map_yandex_{pid_key}',
    help='Карточка организации на yandex.ru/maps/org/… - сверяем название, '
         'телефон, адрес, сайт с КП. Ссылка берётся из колонки «Карта» под '
         '«Яндекс Бизнес» в таблице КП; если у города её нет - город просто '
         'пропускается (не ошибка).')
_check_2gis_maps = _mc3.checkbox(
    '2ГИС', key=f'vc_map_2gis_{pid_key}',
    help='Карточка организации на 2gis.ru/…/firm/… - те же поля, что и для '
         'Яндекс.Карт. Ссылка - из колонки «Карта» под «2ГИС» в таблице КП.')
_check_google_maps = _mc4.checkbox(
    'Google', key=f'vc_map_google_{pid_key}',
    help='Карточка организации на google.com/maps/place/… - те же поля, что '
         'и для Яндекс.Карт/2ГИС. Ссылка - из колонки «Карта» под «Google» в '
         'таблице КП; если у города её нет - город просто пропускается '
         '(не ошибка).')
if not (_check_site or _check_yandex_maps or _check_2gis_maps or _check_google_maps):
    st.warning('Выбери хотя бы один источник - иначе прогону нечего делать.')
if _check_yandex_maps or _check_2gis_maps or _check_google_maps:
    st.caption('⚠ Карты проверяются в браузере - это заметно дольше обычной '
              'сверки сайта (десятки секунд на карточку).')

# Примерное время: сверка сайта - это секунда на город, а каждая карта
# добавляет браузерную поездку на КАЖДЫЙ город, поэтому оценка меняется в разы.
_est_cities_kp = len(_chosen) if _chosen else len(_cities)
_est_maps = sum((_check_yandex_maps, _check_2gis_maps, _check_google_maps))
if _est_cities_kp:
    # Облако перечитывает файл страницы после обновления кода, но модули из
    # sys.modules не перезагружает: страница уже новая, run_estimate - ещё
    # старый, и импорт новой функции ронял всю страницу ImportError.
    import importlib

    import run_estimate as _re
    if not hasattr(_re, 'estimate_kp_seconds'):
        _re = importlib.reload(_re)
    _lo_kp, _hi_kp = _re.estimate_kp_seconds(_est_cities_kp,
                                             check_site=_check_site, maps=_est_maps)
    st.caption(f'⏱ Примерное время: **{_re.format_estimate(_lo_kp, _hi_kp)}** · '
               f'{_est_cities_kp} городов'
               + (f', карт: {_est_maps}' if _est_maps else '') + '.')

st.divider()
st.subheader('Запуск')
# Прокси проекта - единый механизм (proxy_config.py): БД личного кабинета →
# proxy_url_<pid> → proxy_url → HTTP_PROXY, С УЧЁТОМ use_proxy проекта.
# Раньше здесь брали общий _secret('proxy_url') НАПРЯМУЮ и игнорировали
# use_proxy - подпись «прокси будет использован» показывалась даже проектам
# с выключенным прокси, а сам адрес не учитывал личный кабинет.
from proxy_config import resolve_proxy
if not resolve_proxy(pid_key):
    st.caption('У проекта не настроен прокси («Настройки проекта») - '
               'страницы качаются напрямую.')
else:
    st.caption('Прокси настроен - включить его для этого прогона можно '
               'чек-боксом в сайдбаре слева.')

_alive = _pid_alive(_read_pid())
# Прошлый прогон уже завершён (в логе метка), но pid-файл остался, а на облаке
# os.kill(pid,0) может ложно считать его «живым» (pid переиспользован) - тогда
# кнопка «Запустить» блокировалась до ручного «Отменить». Если в логе есть
# метка завершения - прогон НЕ живой, чистим протухший pid-файл.
if _alive:
    _log_early = (LOG_FILE.read_text(encoding='utf-8', errors='ignore')
                  if LOG_FILE.exists() else '')
    if ('✅ ВСЁ ГОТОВО' in _log_early or 'ОТМЕНЕНО' in _log_early
            or '✗' in _log_early):
        _alive = False
        try:
            PID_FILE.unlink(missing_ok=True)
        except Exception:
            pass
_none_chosen = (_mode == 'Выбрать города' and not _chosen)

# Прокси - чек-бокс дорисовывается в место, оставленное render_account_ui
# в сайдбаре (см. auth.fill_proxy_slot) - теперь с уже известным ЗДЕСЬ pid.
import auth
_effective_proxy = auth.fill_proxy_slot(pid_key)

_nothing_selected = not (_check_site or _check_yandex_maps or _check_2gis_maps
                        or _check_google_maps)

# Карты гоняются Playwright-браузером (Chromium) - на облаке его нет по
# умолчанию, нужно доустановить в рантайме (как для «Проверки форм»/«Проверки
# целей»). Раньше эта проверка тут не звалась вовсе - карты запускались
# «вслепую», падая на первом же прогоне с «Executable doesn't exist».
_check_maps = _check_yandex_maps or _check_2gis_maps or _check_google_maps
_browser_ok, _browser_msg = True, ''
if _check_maps and not _alive:
    with st.spinner('Готовлю браузер для карт (первый запуск в облаке - до минуты)…'):
        import browser_setup
        _browser_ok, _browser_msg = browser_setup.ensure_browser()
    if not _browser_ok:
        st.warning(f'Браузер для карт ещё не готов: {_browser_msg}. Если это '
                  'первый запуск в облаке - подождите минуту и обновите страницу.')

_c1, _c2 = st.columns([3, 1])
with _c1:
    if st.button('▶ Запустить проверку', use_container_width=True,
                 type='primary',
                 disabled=_alive or _none_chosen or _nothing_selected or not _browser_ok):
        if not _deps_ready():
            st.error('В этом окружении нет нужных библиотек (requests/bs4/openpyxl).')
        else:
            args = ['variables_run.py', '--project', pid_key,
                    '--check-site' if _check_site else '--no-check-site']
            if _mode == 'Выбрать города' and _chosen:
                args += ['--cities', ','.join(_chosen)]
            if _check_yandex_maps:
                args += ['--check-yandex-maps']
            if _check_2gis_maps:
                args += ['--check-2gis-maps']
            if _check_google_maps:
                args += ['--check-google-maps']
            # На свежем деплое папки cache/ ещё нет - создаём перед очисткой лога.
            LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
            LOG_FILE.write_text('', encoding='utf-8')
            # Прокидываем в фоновый процесс: прокси + ссылку на КП-таблицу +
            # JSON сервисного аккаунта (для приватных таблиц) - из секретов.
            _env = {}
            if _effective_proxy:
                _env['proxy_url'] = _effective_proxy
            try:
                if _kp_url:
                    _env[f'kp_sheet_url_{pid_key}'] = _kp_url
                _sa_json = _sa_json_for_env()   # секция/строка-JSON/base64 → JSON
                if _sa_json:
                    _env['GCP_SA_JSON'] = _sa_json
            except Exception:
                pass
            # Telegram: креды из секретов (те же, что у форм/целей) - variables_run
            # сам отправит отчёт КП в чат после прогона.
            try:
                import tg_report
                _env.update(tg_report.runner_env(
                    pid_key, PROJECTS.get(pid_key, pid_key)))
            except Exception:
                pass
            _launch(args, extra_env=_env or None)
            st.session_state['vars_started'] = datetime.now().strftime('%H:%M:%S')
            st.session_state['vars_project'] = pid_key
            st.rerun()
with _c2:
    if st.button('⛔ Отменить', use_container_width=True, disabled=not _alive):
        _kill(_read_pid())
        try:
            PID_FILE.unlink(missing_ok=True)
        except Exception:
            pass
        st.rerun()

if _none_chosen:
    st.warning('Выберите хотя бы один город или переключитесь на «Все домены+поддомены».')

# ── Легенда ──────────────────────────────────────────────────────────
with st.expander('Как читать результат'):
    st.markdown(
        '- **✓** - значение на сайте совпадает с КП (для телефона: номер входит в '
        'набор номеров города из КП).\n'
        '- **✗** - расхождение. Наведи курсор на ячейку - во всплывающем '
        'комментарии видно «ожидалось / на сайте».\n'
        '- **⚠** - на сайте не найдено (телефон/почта/адрес/мессенджер).\n'
        '- **–** - в КП этого поля нет (проверять не с чем).\n\n'
        'Все ✗/⚠ также собраны списком на листе «Расхождения». Колонки карт '
        '(Яндекс.Карты/2ГИС) - на том же листе «Проверка КП», тот же принцип; '
        'пустая ячейка у источника значит «не проверяли в этом прогоне» '
        '(галочка была снята) - не путать с «–».')

st.divider()
st.subheader('Прогресс')

_log = LOG_FILE.read_text(encoding='utf-8', errors='ignore') if LOG_FILE.exists() else ''
_done = '✅ ВСЁ ГОТОВО' in _log or 'ОТМЕНЕНО' in _log or '✗' in _log
xlsx = ROOT / 'cache' / 'variables' / pid_key / 'variables.xlsx'

if _alive and not _done:
    # прогресс по строкам «[i/N]»
    import re as _re
    m = None
    for ln in _log.splitlines():
        mm = _re.search(r'\[(\d+)/(\d+)\]', ln)
        if mm:
            m = mm
    if m:
        i, n = int(m.group(1)), int(m.group(2))
        st.progress(min(i / max(n, 1), 0.99), text=f'Проверено {i} из {n} поддоменов')
    else:
        st.progress(0.05, text='Готовлю проверку…')
    with st.expander('Подробный лог', expanded=True):
        st.code('\n'.join(_log.splitlines()[-200:]) or '…', language='text')
    time.sleep(2)
    st.rerun()
else:
    if st.session_state.get('vars_started'):
        st.caption(f'Последний запуск: {st.session_state["vars_started"]}')
    if _log.strip():
        with st.expander('Подробный лог', expanded=False):
            st.code('\n'.join(_log.splitlines()[-200:]), language='text')
    if xlsx.exists():
        _date = datetime.fromtimestamp(xlsx.stat().st_mtime).strftime('%d.%m.%Y')
        st.download_button(
            f'⬇ Скачать «Проверка КП {pid_key.capitalize()}-{_date}.xlsx»',
            data=xlsx.read_bytes(),
            file_name=f'Проверка-КП-{pid_key.capitalize()}-{_date}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True)
        # быстрый предпросмотр - лист «Расхождения» (сайт + карты вместе).
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx, read_only=True)
            if 'Расхождения' in wb.sheetnames:
                ws = wb['Расхождения']
                rows = [[c.value for c in r] for r in ws.iter_rows(values_only=False)]
                # Пустой отчёт пишет одну "плашку-заглушку" вместо строк
                # расхождений - её не считаем расхождением.
                _placeholder = (len(rows) == 2
                                and rows[1][0] == 'Расхождений не найдено 🎉')
                if len(rows) > 1 and not _placeholder:
                    st.caption(f'Найдено расхождений: {len(rows) - 1}. '
                              'Открой лист «Расхождения» в файле.')
                else:
                    st.success('Расхождений не найдено 🎉')
            wb.close()
        except Exception:
            pass
    else:
        st.caption('Отчёт появится после запуска.')


# Сохранение шаблона происходит ПРЯМО по клику «Сохранить» вверху
# (render_panel(save_keys=...)) - не зависит от того, дошла ли отрисовка до низа.
