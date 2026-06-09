"""
Site Checker — Streamlit-приложение для проверки доступности сайтов СМУ/ИМП/МПЭ.

Запуск локально:
    streamlit run app.py

На Streamlit Cloud:
    Привязать репозиторий, указать app.py главным файлом.

Структура страницы:
    Шапка с выбором проекта
    ▶ Профиль проверки (Быстрая / Стандартная / Полная / Свои настройки)
    ▶ Что включить (чек-боксы по 6 типам проверок)
    ▶ Кнопка «Запустить»
    После прогона: прогресс-бар, итоги, скачать xlsx
"""
import asyncio
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import streamlit as st

from sources import (
    list_projects, load_project_config, load_sources,
    build_plan, build_custom_plan, TYPE_LABELS,
)
from profiles import PROFILES, get_profile_kwargs
from history import load_history, save_history
from sitemap import load_product_pathnames, get_cached_products_info
from http_checker import run_batch, STATUS, SPEED
from reporter import build_report, make_report_filename
from metrika_404 import (
    fetch_metrika_emails, save_reports_batch, list_stored_reports,
    load_report, MAILBOX_CONFIG, COUNTRY_LABELS,
    fetch_incremental, get_latest_available_date,
    load_reports_for_date, load_reports_for_period,
)
from telegram_notify import (
    format_summary_message, send_run_notification, check_bot_alive,
)


PROJECT_ROOT = Path(__file__).parent
REPORTS_DIR = PROJECT_ROOT / 'reports'
REPORTS_DIR.mkdir(exist_ok=True)


def get_proxy_url():
    """
    Достать URL прокси для исходящих запросов.
    Источник в порядке приоритета:
      1. Streamlit Secrets (для деплоя): st.secrets["proxy_url"]
      2. Переменная окружения HTTP_PROXY (для локального запуска)
      3. Если ничего нет — работаем напрямую

    Возвращает строку или None.
    """
    try:
        if hasattr(st, 'secrets') and 'proxy_url' in st.secrets:
            return st.secrets['proxy_url']
    except Exception:
        pass
    import os
    return os.environ.get('HTTP_PROXY') or os.environ.get('http_proxy')


def get_metrika_credentials(project_id):
    """
    Достать email и пароль приложения для чтения почты Метрики проекта.
    
    Хранятся в Streamlit Secrets под именами:
      metrika_smu_email, metrika_smu_password
      metrika_imp_email, metrika_imp_password (когда добавим)
      metrika_mpe_email, metrika_mpe_password
    
    Возвращает (email, password) или (None, None) если креды не настроены.
    """
    cfg = MAILBOX_CONFIG.get(project_id)
    if not cfg:
        return None, None
    try:
        email = st.secrets.get(cfg['secret_email']) if hasattr(st, 'secrets') else None
        password = st.secrets.get(cfg['secret_password']) if hasattr(st, 'secrets') else None
        return email, password
    except Exception:
        return None, None


def get_telegram_bot_token():
    """
    Достать токен Telegram-бота из Streamlit Secrets (`telegram_bot_token`).
    Возвращает строку или None.
    """
    try:
        if hasattr(st, 'secrets') and 'telegram_bot_token' in st.secrets:
            return st.secrets['telegram_bot_token']
    except Exception:
        pass
    return None


def get_telegram_recipients(project_id):
    """
    Список chat_id получателей уведомлений для проекта из Secrets:
        telegram_recipients_smu = ["1109083536", "987654321"]
        telegram_recipients_imp = [...]
        telegram_recipients_mpe = [...]
    
    Возвращает список строк (chat_id) или пустой список.
    """
    if not project_id:
        return []
    key = f'telegram_recipients_{project_id}'
    try:
        if hasattr(st, 'secrets') and key in st.secrets:
            val = st.secrets[key]
            # Secrets может быть list или строкой
            if isinstance(val, str):
                # Один chat_id строкой
                return [val.strip()]
            elif isinstance(val, (list, tuple)):
                return [str(v).strip() for v in val if str(v).strip()]
    except Exception:
        pass
    return []


# ── Кастомный CSS: тема Vercel/Anthropic ─────────────────────────


CUSTOM_CSS = """
<style>
    /* Светлая палитра — мягкая корпоративная */
    :root {
        --bg: #F3F2EE;
        --bg-elev: #FFFFFF;
        --bg-elev-2: #ECEAE4;
        --border: #DEDBD4;
        --border-strong: #C9C5BC;
        --text: #1A1A1A;
        --text-soft: #5B5853;
        --text-muted: #8A867F;
        --accent: #1A1A1A;
        --accent-hover: #000000;
        --accent-soft: rgba(26, 26, 26, 0.06);
        --accent-ring: rgba(26, 26, 26, 0.12);
        --ok: #16A34A;
        --ok-soft: rgba(22, 163, 74, 0.08);
        --warn: #D97706;
        --warn-soft: rgba(217, 119, 6, 0.08);
        --err: #DC2626;
        --err-soft: rgba(220, 38, 38, 0.08);
    }

    /* Убираем верхний баннер Streamlit */
    [data-testid="stHeader"] {
        background: transparent;
        height: 0;
    }
    [data-testid="stToolbar"] {
        right: 1rem;
    }

    /* ════════════════════════════════════════════════════════════════
       ОГРАНИЧЕНИЕ ШИРИНЫ КОНТЕНТА.
       layout='wide' даёт всю ширину экрана, но при этом контент
       растекается на широких мониторах. Ограничиваем до 1040px и
       центрируем — карточка «Каталог проекта» с 4 метриками помещается
       целиком (Городов / Категорий / Фильтров / Товаров), ничего не
       обрезается (как было «Фильтров: 13 0...»).
       ════════════════════════════════════════════════════════════════ */
    [data-testid="stMainBlockContainer"],
    [data-testid="stAppViewContainer"] section.main > div.block-container,
    section.main > div.block-container {
        max-width: 1040px !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
        padding-top: 3rem !important;
    }

    /* Базовая типографика */
    .stApp {
        background-color: var(--bg);
    }
    h1, h2, h3 {
        font-weight: 600;
        letter-spacing: -0.02em;
        color: var(--text);
    }
    h1 {
        font-size: 2.4rem !important;
        line-height: 1.1;
        margin-bottom: 0.5rem !important;
    }
    h2 {
        font-size: 1.625rem !important;
        margin-top: 0 !important;
        margin-bottom: 1rem !important;
    }
    h3 {
        font-size: 1.25rem !important;
        margin-bottom: 0.75rem !important;
    }
    /* Базовый размер шрифта (унификация: всё body — ровно 16.5px).
       ВАЖНО: явный color на тексте — иначе streamlit где-то по умолчанию
       рендерит тёмные элементы поверх белого с прозрачным/белым цветом
       (так у меня получилось «все заголовки белые и текста не видно»). */
    .stApp {
        font-size: 16.5px !important;
        color: #1A1A1A !important;
    }
    p, span, div, label, li, .stMarkdown, .stMarkdown *,
    [data-testid="stMarkdownContainer"],
    [data-testid="stMarkdownContainer"] * {
        font-size: 16.5px;
        line-height: 1.55;
        color: #1E212E;
    }
    /* Заголовки — точно тёмные, чтобы их было видно. */
    h1, h2, h3, h4,
    .stMarkdown h1, .stMarkdown h2, .stMarkdown h3, .stMarkdown h4 {
        color: #1A1A1A !important;
    }
    /* Markdown-абзацы — единый размер */
    .stMarkdown p {
        font-size: 16.5px !important;
        line-height: 1.55 !important;
        color: #1A1A1A !important;
    }
    /* Капшен — единая «вторичная» строка чуть меньше и серее */
    [data-testid="stCaptionContainer"],
    [data-testid="stCaptionContainer"] * {
        font-size: 14.5px !important;
        color: #6B7280 !important;
    }
    /* Лейблы виджетов (selectbox, checkbox, radio, expander summary) */
    [data-testid="stWidgetLabel"],
    [data-testid="stWidgetLabel"] *,
    [data-testid="stCheckbox"] label p,
    [data-testid="stRadio"] label p,
    [data-testid="stRadio"] label div,
    [data-testid="stExpander"] summary,
    [data-testid="stExpander"] summary * {
        font-size: 16.5px !important;
        color: #1A1A1A !important;
    }
    /* Кнопки — тот же размер */
    div[data-testid="stButton"] > button,
    div[data-testid="stDownloadButton"] > button,
    div[data-testid="stButton"] > button *,
    div[data-testid="stDownloadButton"] > button * {
        font-size: 16.5px !important;
    }
    /* Поля ввода */
    .stTextInput input,
    .stTextArea textarea,
    .stNumberInput input,
    [data-baseweb="select"] {
        font-size: 16.5px !important;
        color: #1A1A1A !important;
    }
    /* Alert-блоки (st.warning / st.info / st.success / st.error) —
       текст внутри тёмный, не растворяется на цветном фоне. */
    [data-testid="stAlert"],
    [data-testid="stAlert"] *,
    [data-baseweb="notification"],
    [data-baseweb="notification"] * {
        color: #1A1A1A !important;
    }
    /* Текст лейбла на прогресс-баре */
    [data-testid="stProgress"] p,
    [data-testid="stProgress"] label,
    [data-testid="stProgress"] span,
    [data-testid="stProgress"] div {
        color: #1A1A1A !important;
    }

    /* Контейнеры-карточки */
    [data-testid="stVerticalBlockBorderWrapper"] {
        background: var(--bg-elev) !important;
        border: 1px solid var(--border) !important;
        border-radius: 12px !important;
        padding: 24px 28px !important;
        margin-bottom: 16px !important;
        transition: border-color 0.15s, box-shadow 0.15s;
    }

    /* Лейблы шагов в заголовках */
    .step-num {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        width: 26px;
        height: 26px;
        background: var(--accent-soft);
        color: var(--accent);
        border-radius: 7px;
        font-size: 13px;
        font-weight: 700;
        margin-right: 10px;
        vertical-align: 2px;
    }

    /* Поля ввода / селекты */
    [data-baseweb="select"] > div,
    .stTextInput input,
    .stTextArea textarea,
    .stNumberInput input {
        background: var(--bg) !important;
        border: 1px solid var(--border) !important;
        border-radius: 8px !important;
        color: var(--text) !important;
        font-size: 1rem !important;
        transition: border-color 0.15s, box-shadow 0.15s;
    }
    [data-baseweb="select"] > div:hover,
    .stTextInput input:hover,
    .stTextArea textarea:hover,
    .stNumberInput input:hover {
        border-color: var(--border-strong) !important;
    }
    [data-baseweb="select"] > div:focus-within,
    .stTextInput input:focus,
    .stTextArea textarea:focus,
    .stNumberInput input:focus {
        border-color: var(--accent) !important;
        box-shadow: 0 0 0 3px var(--accent-ring) !important;
    }
    /* Текст внутри селекта */
    [data-baseweb="select"] [class*="ValueContainer"],
    [data-baseweb="select"] [class*="SingleValue"],
    [data-baseweb="select"] span {
        color: var(--text) !important;
    }

    /* ════════════════════════════════════════════════════════════════
       ВЫПАДАЮЩИЙ СПИСОК (popover селекта)
       Streamlit использует BaseWeb который ставит inline-стили с тёмным
       фоном (#262730). Перебиваем максимально специфичными селекторами.
       ════════════════════════════════════════════════════════════════ */
    div[data-baseweb="popover"] {
        background: transparent !important;
    }
    div[data-baseweb="popover"] > div,
    div[data-baseweb="popover"] ul,
    div[data-baseweb="popover"] [role="listbox"] {
        background: #FFFFFF !important;
        background-color: #FFFFFF !important;
        border: 1px solid #E1E8F0 !important;
        border-radius: 8px !important;
        box-shadow: 0 8px 24px rgba(30, 33, 46, 0.12) !important;
    }
    /* Каждая опция */
    div[data-baseweb="popover"] li,
    div[data-baseweb="popover"] [role="option"],
    div[data-baseweb="popover"] ul > li {
        background: #FFFFFF !important;
        background-color: #FFFFFF !important;
        color: #1A1A1A !important;
        font-size: 1rem !important;
    }
    div[data-baseweb="popover"] li *,
    div[data-baseweb="popover"] [role="option"] * {
        background: transparent !important;
        background-color: transparent !important;
        color: #1A1A1A !important;
    }
    /* Hover на опции — голубая подсветка */
    div[data-baseweb="popover"] li:hover,
    div[data-baseweb="popover"] [role="option"]:hover,
    div[data-baseweb="popover"] li[aria-selected="true"],
    div[data-baseweb="popover"] [role="option"][aria-selected="true"] {
        background: #EEF3FB !important;
        background-color: #EEF3FB !important;
        color: #1A1A1A !important;
    }
    div[data-baseweb="popover"] li:hover *,
    div[data-baseweb="popover"] [role="option"]:hover *,
    div[data-baseweb="popover"] li[aria-selected="true"] *,
    div[data-baseweb="popover"] [role="option"][aria-selected="true"] * {
        color: #1A1A1A !important;
    }

    /* ════════════════════════════════════════════════════════════════
       РАДИО-КНОПКИ ПРОФИЛЕЙ — карточный стиль
       ════════════════════════════════════════════════════════════════ */
    /* Полностью убираем «коллапснутый» лейбл — на скриншотах он
       выглядел как пустая прозрачная плашка-«Профиль» над карточками. */
    [data-testid="stWidgetLabel"][data-baseweb="form-control-label"][hidden],
    [data-testid="stWidgetLabel"] > div[data-testid="stMarkdownContainer"]:empty,
    [data-testid="stRadio"] > label[data-baseweb="form-control-label"],
    [data-testid="stRadio"] [data-testid="stWidgetLabel"] {
        display: none !important;
        height: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
    }
    [data-testid="stRadio"] > div {
        gap: 8px;
    }
    [data-testid="stRadio"] label {
        background: #FFFFFF !important;
        background-color: #FFFFFF !important;
        border: 1px solid #E1E8F0;
        border-radius: 8px;
        padding: 14px 18px !important;
        margin: 0 !important;
        transition: all 0.15s;
        cursor: pointer;
        /* Flex-выравнивание: точка слева, текст справа, не налезают */
        display: flex !important;
        align-items: flex-start !important;
        gap: 12px !important;
    }
    [data-testid="stRadio"] label:hover {
        border-color: #C7D3E1;
        background: #F7FBFE !important;
        background-color: #F7FBFE !important;
    }
    /* Выбранная карточка профиля */
    [data-testid="stRadio"] label:has(input:checked) {
        border-color: #1A1A1A !important;
        background: #EEF3FB !important;
        background-color: #EEF3FB !important;
    }
    /* Радио-точка (кружок) — фиксированная ширина, не сжимается */
    [data-testid="stRadio"] label > div:first-child {
        flex-shrink: 0 !important;
        margin-top: 2px !important;
    }
    /* Текст внутри label — нормальная ширина, выравнивание по верху */
    [data-testid="stRadio"] label > div:last-child,
    [data-testid="stRadio"] label p {
        flex: 1 !important;
        color: #1A1A1A !important;
        background: transparent !important;
        background-color: transparent !important;
        font-size: 1rem !important;
        line-height: 1.5 !important;
        margin: 0 !important;
    }

    /* Чек-боксы */
    [data-testid="stCheckbox"] {
        padding: 4px 0;
    }
    [data-testid="stCheckbox"] label p {
        font-size: 1rem !important;
        font-weight: 500;
        color: var(--text) !important;
    }
    /* Сам квадратик чек-бокса */
    [data-testid="stCheckbox"] [data-baseweb="checkbox"] > div:first-child {
        background: var(--bg) !important;
        border-color: var(--border-strong) !important;
    }

    /* Текст радио-кнопок */
    [data-testid="stRadio"] label p,
    [data-testid="stRadio"] label div {
        font-size: 1rem !important;
        color: var(--text) !important;
    }

    /* ════════════════════════════════════════════════════════════════
       КНОПКИ — общий слой (любая кнопка по умолчанию белая со светлой
       рамкой). Streamlit в разных версиях рендерит кнопки то через
       <button kind="secondary">, то через data-testid="stBaseButton-…",
       поэтому ловим все варианты сразу, иначе на части кнопок остаётся
       чёрный/тёмный фон по умолчанию (это и было «опять чёрный цвет»
       на кнопке «Обновить из почты»).
       ════════════════════════════════════════════════════════════════ */
    div[data-testid="stButton"] > button,
    div[data-testid="stButton"] > button[kind="secondary"],
    div[data-testid="stButton"] > button[kind="tertiary"],
    div[data-testid="stButton"] > button[data-testid="stBaseButton-secondary"],
    div[data-testid="stButton"] > button[data-testid="baseButton-secondary"] {
        font-weight: 600 !important;
        border-radius: 8px !important;
        border: 1px solid #C7D3E1 !important;
        background: #FFFFFF !important;
        background-color: #FFFFFF !important;
        background-image: none !important;
        color: #1A1A1A !important;
        box-shadow: none !important;
        transition: all 0.15s !important;
        padding: 0.55rem 1.25rem !important;
        font-size: 0.95rem !important;
    }
    /* Светлый hover — ТОЛЬКО для обычных st.button(), не для st.download_button()
       и не для primary. Без этого исключения зелёная «Скачать» превращалась
       в белую при наведении. */
    div[data-testid="stButton"] > button:hover:not(:disabled),
    div[data-testid="stButton"] > button[data-testid="stBaseButton-secondary"]:hover:not(:disabled) {
        border-color: #1A1A1A !important;
        background: #F7FBFE !important;
        background-color: #F7FBFE !important;
        color: #1A1A1A !important;
    }
    /* Текст, иконки и любые вложенные div внутри светлой кнопки */
    .stButton > button p,
    .stButton > button span,
    .stButton > button div,
    button[data-testid="stBaseButton-secondary"] *,
    button[data-testid="baseButton-secondary"] * {
        color: inherit !important;
        background: transparent !important;
        background-color: transparent !important;
    }
    /* Текст / иконки внутри обычной кнопки */
    div[data-testid="stButton"] > button p,
    div[data-testid="stButton"] > button span,
    div[data-testid="stButton"] > button div {
        color: inherit !important;
        background: transparent !important;
        background-color: transparent !important;
    }
    /* disabled-кнопки */
    div[data-testid="stButton"] > button:disabled {
        opacity: 0.45;
        cursor: not-allowed;
        background: #FAFBFC !important;
        background-color: #FAFBFC !important;
        color: #8A93A6 !important;
    }
    /* ════════════════════════════════════════════════════════════════
       TOOLTIP-ПОПАПЫ (от параметра help='…' на виджетах).
       Streamlit рендерит их через BaseWeb с тёмным фоном — это и было
       «опять чёрные модалки». Перебиваем максимально специфично.
       ════════════════════════════════════════════════════════════════ */
    div[data-baseweb="tooltip"],
    [data-baseweb="tooltip"] > div,
    div[role="tooltip"],
    [data-testid="stTooltipContent"],
    [data-testid="stTooltipContent"] > div,
    .stTooltip,
    div[data-baseweb="popover"][role="tooltip"] > div {
        background: #FFFFFF !important;
        background-color: #FFFFFF !important;
        color: #1A1A1A !important;
        border: 1px solid #C7D3E1 !important;
        border-radius: 10px !important;
        box-shadow: 0 6px 20px rgba(30, 33, 46, 0.14) !important;
        padding: 12px 14px !important;
        font-size: 0.95rem !important;
        line-height: 1.5 !important;
        max-width: 380px !important;
    }
    /* Все вложенные элементы в тултипе — тёмный текст, прозрачный фон */
    div[data-baseweb="tooltip"] *,
    div[role="tooltip"] *,
    [data-testid="stTooltipContent"] * {
        color: #1A1A1A !important;
        background: transparent !important;
        background-color: transparent !important;
    }
    /* Стрелочка от тултипа — тоже белая */
    div[data-baseweb="tooltip"] [data-baseweb="popover-arrow"] {
        background: #FFFFFF !important;
        border: 1px solid #C7D3E1 !important;
    }

    /* Все кнопки в ряду — одинаковая высота и выравнивание (st.columns) */
    div[data-testid="stButton"] > button,
    div[data-testid="stDownloadButton"] > button {
        min-height: 44px !important;
        white-space: nowrap !important;
    }
    /* Выравниваем содержимое st.columns по верху, чтобы кнопки на одной линии */
    div[data-testid="stHorizontalBlock"] {
        align-items: stretch !important;
    }

    /* ════════════════════════════════════════════════════════════════
       МОДАЛЬНОЕ ОКНО (например, «Clear caches» из верхнего меню) —
       по умолчанию у Streamlit оно тёмное. Делаем светлым в тон.
       ════════════════════════════════════════════════════════════════ */
    div[role="dialog"],
    div[data-testid="stModal"] > div,
    [data-baseweb="modal"] > div,
    [data-baseweb="modal"] [data-baseweb="modal-content"] {
        background: #FFFFFF !important;
        background-color: #FFFFFF !important;
        color: #1A1A1A !important;
        border-radius: 12px !important;
        border: 1px solid #E1E8F0 !important;
    }
    div[role="dialog"] *,
    [data-baseweb="modal"] * {
        color: #1A1A1A !important;
        background-color: transparent !important;
    }
    div[role="dialog"] code,
    [data-baseweb="modal"] code {
        background: #F7FBFE !important;
        color: #1A1A1A !important;
        padding: 2px 6px;
        border-radius: 4px;
    }

    /* ════════════════════════════════════════════════════════════════
       EXPANDER — шапка тоже бывает тёмной, ставим светлую.
       ════════════════════════════════════════════════════════════════ */
    [data-testid="stExpander"] details > summary,
    [data-testid="stExpander"] summary {
        background: #FFFFFF !important;
        background-color: #FFFFFF !important;
        color: #1A1A1A !important;
        border-radius: 10px !important;
    }
    [data-testid="stExpander"] details[open] > summary {
        border-bottom: 1px solid var(--border) !important;
        border-radius: 10px 10px 0 0 !important;
    }
    [data-testid="stExpander"] details > div {
        background: #FFFFFF !important;
    }
    /* ════════════════════════════════════════════════════════════════
       Кодовые блоки (st.code) — подробный лог
       Streamlit рендерит код через Prism.js, который ставит тёмную тему
       даже на светлой странице. Бьём по всем вложенным элементам.
       ════════════════════════════════════════════════════════════════ */
    [data-testid="stCodeBlock"],
    [data-testid="stCode"],
    .stCodeBlock,
    .stCode,
    pre {
        background: #F7FBFE !important;
        background-color: #F7FBFE !important;
        border: 1px solid #E1E8F0 !important;
        border-radius: 8px !important;
    }
    /* ВСЕ вложенные элементы внутри code-block — светлая тема */
    [data-testid="stCodeBlock"] *,
    [data-testid="stCode"] *,
    .stCodeBlock *,
    .stCode * {
        background: transparent !important;
        background-color: transparent !important;
        color: #1A1A1A !important;
    }
    [data-testid="stCodeBlock"] pre,
    [data-testid="stCode"] pre,
    [data-testid="stCodeBlock"] code,
    [data-testid="stCode"] code,
    pre code {
        background: #F7FBFE !important;
        background-color: #F7FBFE !important;
        color: #1A1A1A !important;
        font-size: 0.875rem !important;
    }
    /* Кнопка копирования в углу code-block — тоже светлая */
    [data-testid="stCodeBlock"] button,
    [data-testid="stCode"] button,
    [data-testid="stCodeBlockCopyButton"] {
        background: var(--bg-elev) !important;
        color: var(--text-muted) !important;
        border: 1px solid var(--border) !important;
    }
    [data-testid="stCodeBlock"] button:hover,
    [data-testid="stCode"] button:hover {
        background: var(--accent-soft) !important;
        color: var(--accent) !important;
    }
    /* Inline-код */
    .stMarkdown code {
        background: #F7FBFE !important;
        color: #1A1A1A !important;
        padding: 2px 6px;
        border-radius: 4px;
        font-size: 0.9rem;
    }

    /* type='primary' — это «Запустить проверку» (синяя). */
    div[data-testid="stButton"] > button[kind="primary"],
    div[data-testid="stButton"] > button[data-testid="stBaseButton-primary"],
    div[data-testid="stButton"] > button[data-testid="baseButton-primary"] {
        background: #1A1A1A !important;
        background-color: #1A1A1A !important;
        background-image: none !important;
        border: 1px solid #1A1A1A !important;
        color: #FFFFFF !important;
        box-shadow: 0 1px 3px rgba(26, 26, 26, 0.20);
        font-weight: 700 !important;
        padding: 0.85rem 1.5rem !important;
    }
    div[data-testid="stButton"] > button[kind="primary"] *,
    div[data-testid="stButton"] > button[data-testid="stBaseButton-primary"] *,
    div[data-testid="stButton"] > button[data-testid="baseButton-primary"] * {
        color: #FFFFFF !important;
        font-weight: 700 !important;
        background: transparent !important;
    }
    div[data-testid="stButton"] > button[kind="primary"]:hover:not(:disabled),
    div[data-testid="stButton"] > button[data-testid="stBaseButton-primary"]:hover:not(:disabled),
    div[data-testid="stButton"] > button[data-testid="baseButton-primary"]:hover:not(:disabled) {
        background: #000000 !important;
        background-color: #000000 !important;
        border-color: #000000 !important;
        color: #FFFFFF !important;
        box-shadow: 0 4px 12px rgba(26, 26, 26, 0.30);
        transform: translateY(-1px);
    }
    div[data-testid="stButton"] > button[kind="primary"]:disabled,
    div[data-testid="stButton"] > button[data-testid="stBaseButton-primary"]:disabled {
        opacity: 0.45;
        cursor: not-allowed;
        background: #1A1A1A !important;
    }

    /* Кнопка скачивания — зелёный градиент. Покрываем ВСЕ комбинации,
       включая type='primary', kind="primary", kind="primaryFormSubmit"
       и data-testid'ы — иначе streamlit рисовал её коралловой как
       primary-кнопку (это и был «красный» цвет на скриншоте). */
    div[data-testid="stDownloadButton"] > button,
    div[data-testid="stDownloadButton"] > button[kind="primary"],
    div[data-testid="stDownloadButton"] > button[kind="secondary"],
    div[data-testid="stDownloadButton"] > button[data-testid="stBaseButton-primary"],
    div[data-testid="stDownloadButton"] > button[data-testid="stBaseButton-secondary"],
    div[data-testid="stDownloadButton"] > button[data-testid="baseButton-primary"],
    div[data-testid="stDownloadButton"] > button[data-testid="baseButton-secondary"] {
        background: linear-gradient(180deg, #22C55E 0%, #16A34A 100%) !important;
        background-color: #16A34A !important;
        background-image: linear-gradient(180deg, #22C55E 0%, #16A34A 100%) !important;
        border: 1px solid #16A34A !important;
        color: #FFFFFF !important;
        font-weight: 700 !important;
        padding: 0.7rem 1.25rem !important;
        border-radius: 8px !important;
        box-shadow: 0 2px 8px rgba(22, 163, 74, 0.20) !important;
        transition: all 0.15s !important;
    }
    div[data-testid="stDownloadButton"] > button *,
    div[data-testid="stDownloadButton"] > button p,
    div[data-testid="stDownloadButton"] > button span,
    div[data-testid="stDownloadButton"] > button div {
        color: #FFFFFF !important;
        font-weight: 700 !important;
        background: transparent !important;
        background-color: transparent !important;
    }
    div[data-testid="stDownloadButton"] > button:hover:not(:disabled) {
        background: linear-gradient(180deg, #34D365 0%, #1FB158 100%) !important;
        background-color: #1FB158 !important;
        background-image: linear-gradient(180deg, #34D365 0%, #1FB158 100%) !important;
        border-color: #1FB158 !important;
        box-shadow: 0 6px 16px rgba(22, 163, 74, 0.30) !important;
        color: #FFFFFF !important;
        transform: translateY(-1px);
    }
    div[data-testid="stDownloadButton"] > button:hover:not(:disabled) * {
        color: #FFFFFF !important;
    }

    /* Метрики — карточный вид */
    [data-testid="stMetric"] {
        background: var(--bg);
        border: 1px solid var(--border);
        border-radius: 10px;
        padding: 14px 18px;
        transition: border-color 0.15s;
    }
    [data-testid="stMetric"]:hover {
        border-color: var(--border-strong);
    }
    [data-testid="stMetricLabel"] {
        color: var(--text-muted) !important;
        font-size: 0.75rem !important;
        text-transform: uppercase !important;
        letter-spacing: 0.05em !important;
        font-weight: 600 !important;
    }
    [data-testid="stMetricValue"] {
        font-size: 1.625rem !important;
        font-weight: 600 !important;
        margin-top: 4px;
        color: var(--text) !important;
        white-space: nowrap !important;
        overflow: visible !important;
    }
    /* Контейнер значения метрики — без обрезки */
    [data-testid="stMetricValue"] > div {
        overflow: visible !important;
        text-overflow: clip !important;
    }

    /* Алерты */
    [data-baseweb="notification"] {
        border-radius: 10px !important;
        border-width: 1px !important;
    }

    /* Прогресс-бар */
    [data-testid="stProgress"] {
        background: transparent !important;
    }
    [data-testid="stProgress"] > div {
        background: transparent !important;
    }
    /* Дорожка прогресса — светлая */
    [data-testid="stProgress"] > div > div {
        background: var(--bg-elev-2) !important;
        background-color: var(--bg-elev-2) !important;
        border-radius: 4px;
    }
    /* Заполнение — синее */
    [data-testid="stProgress"] > div > div > div {
        background: var(--accent) !important;
        background-color: var(--accent) !important;
    }
    /* Текст-метка над прогресс-баром */
    [data-testid="stProgress"] p,
    [data-testid="stProgress"] label,
    [data-testid="stProgress"] span {
        color: var(--text) !important;
        background: transparent !important;
        background-color: transparent !important;
        font-size: 0.95rem !important;
    }

    /* Expander */
    [data-testid="stExpander"] {
        border: 1px solid var(--border) !important;
        border-radius: 10px !important;
        background: var(--bg) !important;
    }
    [data-testid="stExpander"] summary {
        font-weight: 500;
    }

    /* Капшен */
    [data-testid="stCaptionContainer"] {
        color: var(--text-muted) !important;
    }

    /* Разделитель */
    hr {
        margin: 1.5rem 0 !important;
        border-color: var(--border) !important;
    }

    /* Ссылки */
    .stMarkdown a {
        color: var(--accent) !important;
        text-decoration: none;
        font-weight: 500;
    }
    .stMarkdown a:hover {
        text-decoration: underline;
    }

    /* Брендинговая шапка */
    .brand-bar {
        display: flex;
        align-items: center;
        gap: 12px;
        padding-bottom: 16px;
        margin-bottom: 8px;
    }
    .brand-logo {
        width: 40px;
        height: 40px;
        background: linear-gradient(135deg, var(--accent) 0%, var(--accent-hover) 100%);
        border-radius: 9px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 21px;
        box-shadow: 0 2px 8px rgba(26, 26, 26, 0.25);
    }
    .brand-name {
        font-weight: 700;
        font-size: 1.1rem;
        letter-spacing: -0.01em;
        color: var(--text);
    }
    .brand-sub {
        color: var(--text-muted);
        font-size: 0.875rem;
        margin-left: auto;
    }

    /* ════════════════════════════════════════════════════════════════
       ПОСЛЕДНИЙ РУБЕЖ: на скриншотах внутри expander кнопка «Обновить
       из почты» рисовалась чёрной — streamlit добавляет вложенные
       элементы, через которые селектор «> button» не проходит.
       Перебиваем ВСЕ button-элементы кроме primary/download.
       ════════════════════════════════════════════════════════════════ */
    [data-testid="stButton"] button:not([kind="primary"]):not([data-testid="stBaseButton-primary"]):not([data-testid="baseButton-primary"]),
    [data-testid="stExpander"] [data-testid="stButton"] button {
        background: #FFFFFF !important;
        background-color: #FFFFFF !important;
        background-image: none !important;
        color: #1A1A1A !important;
        border: 1px solid #C7D3E1 !important;
    }
    [data-testid="stButton"] button:not([kind="primary"]):not([data-testid="stBaseButton-primary"]):not([data-testid="baseButton-primary"]) *,
    [data-testid="stExpander"] [data-testid="stButton"] button * {
        color: #1A1A1A !important;
        background: transparent !important;
        background-color: transparent !important;
    }
    [data-testid="stButton"] button:not([kind="primary"]):not([data-testid="stBaseButton-primary"]):not([data-testid="baseButton-primary"]):hover:not(:disabled),
    [data-testid="stExpander"] [data-testid="stButton"] button:hover:not(:disabled) {
        background: #F7FBFE !important;
        background-color: #F7FBFE !important;
        color: #1A1A1A !important;
        border-color: #1A1A1A !important;
    }
    [data-testid="stButton"] button:not([kind="primary"]):not([data-testid="stBaseButton-primary"]):not([data-testid="baseButton-primary"]):hover:not(:disabled) * {
        color: #1A1A1A !important;
    }

    /* ════════════════════════════════════════════════════════════════
       FILE UPLOADER — приводим к светлой палитре.
       Дефолт у Streamlit — тёмный drop-zone, который выпадает из
       светлой темы и торчит чёрным окошком. Перекрашиваем под --bg-elev.
       ════════════════════════════════════════════════════════════════ */
    [data-testid="stFileUploaderDropzone"],
    [data-testid="stFileUploader"] section,
    [data-testid="stFileUploader"] [data-testid="stFileDropzone"] {
        background: var(--bg-elev) !important;
        background-color: var(--bg-elev) !important;
        border: 1px dashed var(--border-strong) !important;
        color: var(--text-soft) !important;
    }
    [data-testid="stFileUploaderDropzone"]:hover {
        border-color: var(--accent) !important;
        background-color: var(--accent-soft) !important;
    }
    /* Текст и иконка внутри зоны */
    [data-testid="stFileUploaderDropzone"] *,
    [data-testid="stFileUploader"] section * {
        color: var(--text-soft) !important;
    }
    [data-testid="stFileUploaderDropzone"] small,
    [data-testid="stFileUploader"] section small {
        color: var(--text-muted) !important;
    }
    [data-testid="stFileUploaderDropzone"] svg,
    [data-testid="stFileUploader"] section svg {
        color: var(--text-muted) !important;
        fill: var(--text-muted) !important;
    }
    /* Кнопка Browse files внутри dropzone */
    [data-testid="stFileUploaderDropzone"] button,
    [data-testid="stFileUploader"] section button {
        background: var(--bg) !important;
        background-color: var(--bg) !important;
        color: var(--text) !important;
        border: 1px solid var(--border-strong) !important;
    }
    [data-testid="stFileUploaderDropzone"] button:hover,
    [data-testid="stFileUploader"] section button:hover {
        background-color: var(--bg-elev-2) !important;
        border-color: var(--accent) !important;
        color: var(--accent) !important;
    }
    [data-testid="stFileUploaderDropzone"] button *,
    [data-testid="stFileUploader"] section button * {
        color: inherit !important;
    }
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


# ── Session state defaults ────────────────────────────────────────


def init_session():
    """Установить дефолты в session_state при первом заходе."""
    defaults = {
        'project_id': None,
        'sources': None,
        'sources_project_id': None,    # для какого проекта загружены sources
        'profile': 'standard',
        'check_main': True,
        'check_catalog': True,
        'check_categories': True,
        'check_filters': True,
        'check_products': True,
        'check_text': True,
        'is_running': False,
        'run_results': None,            # последние результаты прогона
        'run_report_path': None,        # путь к скачиваемому файлу
        'run_started_at': None,
        'run_finished_at': None,
        'custom_urls_text': '',         # содержимое textarea в custom-режиме
        'custom_save_list': False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


init_session()


# ── Утилиты ────────────────────────────────────────────────────────


def format_duration(sec: int) -> str:
    """Секунды → 'X сек' / 'X,Y мин' / 'H ч M мин'."""
    if sec < 60:
        return f'{sec} сек'
    if sec < 3600:
        m = sec / 60
        if m < 10:
            return f'{m:.1f} мин'.replace('.', ',')
        return f'{int(m)} мин'
    h = sec // 3600
    m = (sec % 3600) // 60
    return f'{h} ч {m} мин' if m else f'{h} ч'


def reset_run_state():
    """Сброс баннера прошлого результата при изменении настроек."""
    st.session_state.run_results = None
    st.session_state.run_report_path = None


@st.cache_data(ttl=3600, show_spinner='Загружается каталог проекта…')
def cached_load_sources(project_id: str):
    """Кеш каталога проекта в session (1 час)."""
    cfg = load_project_config(project_id)
    src = load_sources(cfg)
    return cfg, src


def _build_metrika_only_xlsx(reports, sheet_title: str = '404 из Метрики') -> bytes:
    """
    Сформировать простой xlsx по списку Report404 без данных Site Checker.

    Используется для кнопок «Скачать отчёт за вчера» и «Скачать за период» —
    когда пользователю нужен только агрегат Метрики, без запуска проверки сайта.

    Сортировка: свежие даты сверху, в пределах даты — по убыванию просмотров.
    Возвращает bytes (готовые для st.download_button).
    """
    from io import BytesIO as _BIO
    from openpyxl import Workbook as _Wb
    from openpyxl.styles import (
        Font as _Font, PatternFill as _Fill, Alignment as _Align,
        Border as _Border, Side as _Side,
    )

    # Заметные границы (раньше были слишком светлые — пользователь жаловался
    # «почему в отчёте нет границ»). Те же цвета что и в reporter.py.
    _side_light = _Side(style='thin', color='C7D0DA')
    _side_strong = _Side(style='thin', color='A8B2BD')
    _border_light = _Border(
        top=_side_light, left=_side_light, bottom=_side_light, right=_side_light,
    )
    _border_strong = _Border(
        top=_side_strong, left=_side_strong, bottom=_side_strong, right=_side_strong,
    )

    buf = _BIO()
    wb = _Wb()
    ws = wb.active
    ws.title = sheet_title[:31]  # лимит Excel на имя листа
    ws.sheet_view.showGridLines = False

    headers = [
        ('Дата', 14),
        ('Страна', 18),
        ('URL страницы', 60),
        ('Просмотры', 12),
        ('Посетители', 12),
        ('Реферер', 38),
        ('Заголовок страницы', 40),
    ]
    for i, (label, width) in enumerate(headers, 1):
        from openpyxl.utils import get_column_letter as _gcl
        ws.column_dimensions[_gcl(i)].width = width
        c = ws.cell(row=1, column=i)
        c.value = label
        c.font = _Font(bold=True, size=10, color='71717A')
        c.fill = _Fill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
        c.alignment = _Align(horizontal='left', vertical='center', indent=1)
        c.border = _border_strong
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    # Плоский список с сортировкой
    flat = []
    for rep in reports:
        for p in rep.pages:
            flat.append((rep, p))
    flat.sort(key=lambda rp: (rp[0].report_date, -rp[1].views), reverse=False)
    flat.sort(key=lambda rp: rp[0].report_date, reverse=True)

    from datetime import datetime as _dt2
    for row_idx, (rep, p) in enumerate(flat, start=2):
        try:
            date_str = _dt2.strptime(rep.report_date, '%Y-%m-%d').strftime('%d.%m.%Y')
        except ValueError:
            date_str = rep.report_date

        c1 = ws.cell(row=row_idx, column=1)
        c1.value = date_str
        c1.alignment = _Align(horizontal='left', vertical='center', indent=1)

        c2 = ws.cell(row=row_idx, column=2)
        c2.value = f'{rep.country_code} — {rep.country_name}'
        c2.alignment = _Align(horizontal='left', vertical='center', indent=1)

        url_cell = ws.cell(row=row_idx, column=3)
        if p.page_url:
            url_cell.value = p.page_url
            url_cell.hyperlink = p.page_url
            url_cell.font = _Font(
                name='Consolas', size=10, color='1A56E8', underline='single',
            )
        else:
            url_cell.value = '—'
            url_cell.font = _Font(size=10, italic=True, color='9CA3AF')
        url_cell.alignment = _Align(
            horizontal='left', vertical='center', indent=1, wrap_text=True,
        )

        c4 = ws.cell(row=row_idx, column=4)
        c4.value = p.views
        c4.alignment = _Align(horizontal='right', vertical='center', indent=1)

        c5 = ws.cell(row=row_idx, column=5)
        c5.value = p.visitors
        c5.alignment = _Align(horizontal='right', vertical='center', indent=1)

        c6 = ws.cell(row=row_idx, column=6)
        c6.value = p.referer or '—'
        c6.alignment = _Align(
            horizontal='left', vertical='center', indent=1, wrap_text=True,
        )

        c7 = ws.cell(row=row_idx, column=7)
        c7.value = p.page_title
        c7.alignment = _Align(
            horizontal='left', vertical='center', indent=1, wrap_text=True,
        )

        # Границы и базовый шрифт на все ячейки строки
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _border_light
            if cell.font.color is None or cell.font.color.rgb is None:
                cell.font = _Font(size=10, color='1E212E')
        ws.row_dimensions[row_idx].height = 22

    if flat:
        ws.auto_filter.ref = f'A1:G{len(flat) + 1}'

    wb.save(buf)
    return buf.getvalue()


# ── Шапка ──────────────────────────────────────────────────────────


st.title('Доступность сайтов')
st.caption('Ежедневная проверка — СМУ · ИМП · МПЭ')


# ── Шаг 1: выбор проекта ───────────────────────────────────────────


# ── Шаг 1: выбор проекта ───────────────────────────────────────────


with st.container(border=True):
    st.markdown('<h3><span class="step-num">1</span>Какой сайт проверяем</h3>', unsafe_allow_html=True)

    projects = list_projects()
    project_options = ['— выберите —'] + [p['name'] for p in projects] + ['Свой список URL']
    project_to_id = {p['name']: p['id'] for p in projects}

    # Подсчёт текущего индекса для select-box
    current_label = '— выберите —'
    if st.session_state.project_id == '__custom__':
        current_label = 'Свой список URL'
    elif st.session_state.project_id:
        for p in projects:
            if p['id'] == st.session_state.project_id:
                current_label = p['name']
                break

    selected_label = st.selectbox(
        'Проект',
        project_options,
        index=project_options.index(current_label),
        label_visibility='collapsed',
    )

    if selected_label == 'Свой список URL':
        new_pid = '__custom__'
    elif selected_label == '— выберите —':
        new_pid = None
    else:
        new_pid = project_to_id[selected_label]

    # Если проект сменился — сброс результата прошлого прогона
    if new_pid != st.session_state.project_id:
        reset_run_state()
        st.session_state.project_id = new_pid


# ── Дальше расходится логика: обычный проект ИЛИ custom-режим ──────


is_custom = (st.session_state.project_id == '__custom__')
is_project = (st.session_state.project_id is not None and not is_custom)


# ═══════════════════════════════════════════════════════════════════
# CUSTOM MODE — свой список URL
# ═══════════════════════════════════════════════════════════════════


if is_custom:
    with st.container(border=True):
        st.markdown('<h3>Список URL для проверки</h3>', unsafe_allow_html=True)
        st.caption(
            'Вставьте ссылки – по одной на строку. Можно загрузить из файла (.txt или .csv). '
            'Если протокол не указан, добавится https://. Строки после символа # игнорируются.'
        )

        uploaded = st.file_uploader(
            'Загрузить .txt / .csv',
            type=['txt', 'csv'],
            label_visibility='collapsed',
        )
        if uploaded:
            try:
                text = uploaded.read().decode('utf-8', errors='replace')
                # Для CSV — берём первое поле каждой строки
                if uploaded.name.lower().endswith('.csv'):
                    lines = []
                    for line in text.splitlines():
                        cells = line.split(',') if ',' in line else line.split(';')
                        first = cells[0].strip().strip('"').strip("'")
                        lines.append(first)
                    text = '\n'.join(lines)
                # Дописываем к существующему
                existing = st.session_state.custom_urls_text.strip()
                st.session_state.custom_urls_text = (existing + '\n' + text) if existing else text
            except Exception as e:
                st.error(f'Не удалось прочитать файл: {e}')

        custom_text = st.text_area(
            'URLs',
            value=st.session_state.custom_urls_text,
            height=240,
            label_visibility='collapsed',
            placeholder='https://example.com/page1\nhttps://example.com/page2\nexample.com/page3',
            key='custom_urls_input',
        )
        if custom_text != st.session_state.custom_urls_text:
            st.session_state.custom_urls_text = custom_text
            reset_run_state()

        # Парсим URL'ы из текста
        custom_plan = build_custom_plan(custom_text.split('\n'))
        valid_count = len(custom_plan.tasks)

        col1, col2 = st.columns([3, 1])
        with col1:
            if valid_count == 0:
                st.info('Введите URL\'ы')
            else:
                st.success(f'Готово к проверке: {valid_count} URL')
        with col2:
            save_list = st.checkbox(
                'Сохранить список',
                value=st.session_state.custom_save_list,
                help='Список останется в сессии и появится при следующем открытии',
            )
            st.session_state.custom_save_list = save_list

        check_text_custom = st.checkbox(
            'Искать битые переменные в текстах',
            value=st.session_state.check_text,
            help='{{переменная}}, %name%, undefined и [object Object]',
        )
        st.session_state.check_text = check_text_custom


# ═══════════════════════════════════════════════════════════════════
# PROJECT MODE — обычный режим с профилями и каталогом
# ═══════════════════════════════════════════════════════════════════


elif is_project:
    # Загружаем каталог из cache
    try:
        cfg, src = cached_load_sources(st.session_state.project_id)
        st.session_state.sources = src
        st.session_state.sources_project_id = st.session_state.project_id
    except Exception as e:
        st.error(f'Не удалось загрузить каталог: {e}')
        st.stop()

    # ─── Метрики проекта в одной карточке ─────────────────────
    stats = src.stats
    # Достаём кеш товаров (если был хоть один прогон с галкой «Товары»)
    products_info = get_cached_products_info(st.session_state.project_id)

    with st.container(border=True):
        st.markdown(
            f'<p style="color:var(--text-muted);font-size:0.875rem;'
            f'margin-bottom:0.75rem;text-transform:uppercase;letter-spacing:0.05em;'
            f'font-weight:600">Каталог проекта</p>',
            unsafe_allow_html=True,
        )
        c1, c2, c3, c4 = st.columns(4)
        c1.metric('Городов', stats['subdomains_count'])
        c2.metric('Категорий', f'{stats["categories_count"]:,}'.replace(',', ' '))
        if stats['has_filters']:
            c3.metric('Фильтров', f'{stats["filters_count"]:,}'.replace(',', ' '))
        else:
            c3.metric('Фильтров', 'нет')

        # Товары — динамически из sitemap. Если ещё не загружали — показываем «—»
        if products_info and products_info['count'] > 0:
            fmt_count = f'{products_info["count"]:,}'.replace(',', ' ')
            from datetime import datetime as _dt
            d = _dt.fromtimestamp(products_info['fetched_at_ms'] / 1000)
            label_suffix = '' if products_info['is_fresh'] else ' (устарел)'
            c4.metric(
                'Товаров',
                fmt_count,
                help=f'По данным sitemap.xml от {d.strftime("%d.%m.%Y %H:%M")}{label_suffix}. '
                     f'Обновится автоматически при следующем прогоне с галкой «Карточки товаров».',
            )
        else:
            c4.metric(
                'Товаров',
                '—',
                help='Запустите проверку с галкой «Карточки товаров» — '
                     'приложение загрузит sitemap.xml и покажет здесь число.',
            )

        # Главный город — маленькой подписью внизу карточки
        st.markdown(
            f'<p style="color:var(--text-muted);font-size:0.85rem;margin-top:0.5rem;margin-bottom:0">'
            f'Главный город (всегда в выборке): <strong>{cfg.get("mandatory_city", "Москва")}</strong>'
            f'</p>',
            unsafe_allow_html=True,
        )

    # ─── Шаг 2: профиль в карточке ───────────────────────────
    with st.container(border=True):
        st.markdown('<h3><span class="step-num">2</span>Что и сколько проверять</h3>', unsafe_allow_html=True)

        profile_labels = {pid: f'{p["label"]} — {p["description"]}' for pid, p in PROFILES.items()}
        profile_choices = list(profile_labels.keys()) + ['custom']

        def profile_format(pid):
            if pid == 'custom':
                return 'Свои настройки — задать вручную'
            return profile_labels[pid]

        new_profile = st.radio(
            label='profile_radio',
            options=profile_choices,
            index=profile_choices.index(st.session_state.profile) if st.session_state.profile in profile_choices else 0,
            format_func=profile_format,
            label_visibility='collapsed',
        )
        if new_profile != st.session_state.profile:
            st.session_state.profile = new_profile
            reset_run_state()

        # Если выбран профиль — подставляем его значения. Если «Свои» — даём слайдеры
        if st.session_state.profile == 'custom':
            st.markdown('<p style="color:var(--text-muted);font-size:0.875rem;margin-top:1rem;margin-bottom:0.5rem">Параметры выборки</p>', unsafe_allow_html=True)
            col1, col2 = st.columns(2)
            with col1:
                random_subs = st.number_input(
                    'Случайных городов',
                    min_value=0, max_value=stats['subdomains_count'] - 1,
                    value=5, step=1,
                    help='Москва добавится автоматически',
                )
                cats_per_sub = st.number_input(
                    'Категорий на каждый город',
                    min_value=0, max_value=50, value=5, step=1,
                )
            with col2:
                if stats['has_filters']:
                    filters_per_sub = st.number_input(
                        'Фильтров на каждый город',
                        min_value=0, max_value=50, value=5, step=1,
                    )
                else:
                    filters_per_sub = 0
                    st.markdown('_У проекта нет фильтров_')
                products_per_sub = st.number_input(
                    'Товаров на каждый город',
                    min_value=0, max_value=50, value=3, step=1,
                )
        else:
            kw = get_profile_kwargs(st.session_state.profile)
            random_subs = kw['random_subdomains_count']
            cats_per_sub = kw['categories_per_subdomain']
            filters_per_sub = kw['filters_per_subdomain'] if stats['has_filters'] else 0
            products_per_sub = kw['products_per_subdomain']

    # ─── Шаг 3: чек-боксы в карточке ────────────────────────
    with st.container(border=True):
        st.markdown('<h3><span class="step-num">3</span>Какие пункты включить</h3>', unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1:
            check_main = st.checkbox('🏠 Главные страницы', value=st.session_state.check_main, help='Пункт 1.1 чек-листа')
            check_catalog = st.checkbox('📁 Страница /catalog/', value=st.session_state.check_catalog, help='Пункт 1.2')
            check_categories = st.checkbox('📂 Категории', value=st.session_state.check_categories, help='Пункт 1.3')
        with c2:
            if stats['has_filters']:
                check_filters = st.checkbox('🏷️ Фильтры', value=st.session_state.check_filters, help='Пункт 1.4')
            else:
                check_filters = False
                st.markdown('<span style="color:var(--text-muted)">🏷️ Фильтры _(нет в каталоге проекта)_</span>', unsafe_allow_html=True)
            check_products = st.checkbox('🛒 Карточки товаров', value=st.session_state.check_products, help='Пункт 1.5 — из sitemap.xml')
            check_text = st.checkbox('🔤 Битые переменные', value=st.session_state.check_text, help='Пункт 1.6 — {{city}}, %price%, undefined и т.д.')

        # сохраняем в state
        for key, val in [
            ('check_main', check_main), ('check_catalog', check_catalog),
            ('check_categories', check_categories), ('check_filters', check_filters),
            ('check_products', check_products), ('check_text', check_text),
        ]:
            if val != st.session_state[key]:
                st.session_state[key] = val
                reset_run_state()

    # ═══════════════════════════════════════════════════════════════════
    # БЛОК 404 ИЗ МЕТРИКИ — отдельная самодостаточная карточка.
    # Работает БЕЗ запуска полной проверки сайта: можно просто скачать
    # отчёт Метрики за вчера или за период.
    # ═══════════════════════════════════════════════════════════════════
    metrika_pid = st.session_state.project_id
    m_email, m_password = get_metrika_credentials(metrika_pid)
    metrika_creds_ok = bool(m_email and m_password)

    if metrika_creds_ok:
        # Эта секция максимально сжата: данные Метрики автоматически
        # подтягиваются в основной xlsx-отчёт при запуске проверки сайта,
        # поэтому отдельные кнопки «Скачать за вчера» / «Обновить» здесь
        # больше не торчат. Если нужен только агрегат за период без
        # запуска проверки — раскрываете expander.
        with st.container(border=True):
            with st.expander(
                '📥 Сводный отчёт Метрики за период (7 / 14 / 30 дней)',
                expanded=False,
            ):
                st.caption(
                    'При обычном запуске проверки сайта данные Метрики уже '
                    'автоматически подтягиваются в xlsx-отчёт. Этот блок — на '
                    'случай если нужен отдельный сводный xlsx за период, без '
                    'запуска проверки.'
                )
                period_options = {'7 дней': 7, '2 недели': 14, '30 дней': 30}
                pc1, pc2 = st.columns([1, 2])
                with pc1:
                    period_label = st.selectbox(
                        'Период',
                        list(period_options.keys()),
                        label_visibility='collapsed',
                        key='metrika_period_select',
                        help='За сколько прошлых дней собрать единый файл. '
                             'Сегодняшний день не входит — Метрика присылает '
                             'отчёт за прошедший день утром следующего.',
                    )
                selected_days = period_options[period_label]
                period_reports = load_reports_for_period(metrika_pid, selected_days)
                with pc2:
                    if period_reports:
                        period_buf = _build_metrika_only_xlsx(
                            period_reports,
                            sheet_title=f'404 за {selected_days} дн.',
                        )
                        total_pages_in_period = sum(
                            r.total_pages for r in period_reports
                        )
                        unique_dates = len({r.report_date for r in period_reports})
                        st.download_button(
                            label=f'⬇ Скачать ({unique_dates} дн., '
                                  f'{total_pages_in_period} стр.)',
                            data=period_buf,
                            file_name=f'metrika-404-{metrika_pid}-{selected_days}d.xlsx',
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            use_container_width=True,
                            key='dl_metrika_period',
                        )
                    else:
                        st.button(
                            'Нет данных за период',
                            use_container_width=True,
                            disabled=True,
                            key='dl_metrika_period_disabled',
                            help='Чтобы данные появились — запустите обычную '
                                 'проверку сайта (галочка автоматически '
                                 'обновит почту), или включите чекбокс '
                                 '«Перечитать всё заново» ниже и нажмите '
                                 '«Обновить из почты».',
                        )

                # Одна кнопка «Обновить из почты». Под капотом всегда работает
                # upgrade_if_better=True — это значит: новые письма сохраняются,
                # а уже существующие отчёты ПЕРЕЗАПИСЫВАЮТСЯ, если новый разбор
                # извлёк больше URL'ов или страниц (например, в Метрике
                # добавили колонку «Адрес страницы»). Поэтому отдельная галка
                # «Перечитать всё заново» — лишняя.
                st.divider()
                refresh_clicked = st.button(
                    '🔄 Обновить из почты',
                    use_container_width=True,
                    key='btn_refresh_metrika',
                    help='Зайти в ящик Метрики и забрать письма за последние '
                         '14 дней. Новые сохраняются. Уже существующие — '
                         'перезаписываются, если в свежем разборе появилось '
                         'больше URL\'ов (полезно после смены шаблона рассылки '
                         'в Метрике, когда там добавили «Адрес страницы»).',
                )

                if refresh_clicked:
                    metrika_proxy = get_proxy_url()
                    log_messages_m = []
                    progress_m = st.progress(0, text='Подключаюсь к почте Яндекса…')

                    def on_log_m(level, msg):
                        log_messages_m.append(msg)

                    def on_progress_m(done, total):
                        if total > 0:
                            progress_m.progress(
                                min(1.0, done / total),
                                text=f'Обрабатываю письмо {done} из {total}…',
                            )

                    try:
                        summary = fetch_incremental(
                            project_id=metrika_pid,
                            email_addr=m_email,
                            password=m_password,
                            folder=MAILBOX_CONFIG[metrika_pid]['folder'],
                            proxy_url=metrika_proxy,
                            lookback_days=14,
                            log=on_log_m,
                            progress=on_progress_m,
                            force_refresh=False,
                            upgrade_if_better=True,
                        )
                        progress_m.empty()
                        fetched = summary['fetched']
                        upgraded = summary.get('upgraded', 0)
                        skipped = summary['skipped']
                        total_letters = summary.get('total_in_letters', 0)
                        parts = []
                        if fetched:
                            parts.append(f'новых: **{fetched}**')
                        if upgraded:
                            parts.append(f'обновлено: **{upgraded}**')
                        if skipped:
                            parts.append(f'без изменений: {skipped}')
                        if fetched or upgraded:
                            st.success('✅ Готово · ' + ' · '.join(parts))
                            st.rerun()
                        elif total_letters > 0:
                            st.info(
                                f'ℹ Просмотрено {total_letters} писем — всё актуально.'
                            )
                        else:
                            st.warning(
                                '⚠ За период писем от Метрики в ящике нет. '
                                'Проверьте рассылку в Метрике.'
                            )
                        with st.expander('Подробный лог', expanded=False):
                            st.code(
                                '\n'.join(log_messages_m[-100:]) or '(лог пуст)',
                                language='text',
                            )
                    except Exception as e:
                        progress_m.empty()
                        import traceback
                        st.error(f'❌ Не удалось обновить: {type(e).__name__}: {e}')
                        with st.expander('Подробный лог (с трассировкой)', expanded=True):
                            st.code(
                                '\n'.join(log_messages_m[-100:])
                                + '\n\n--- TRACEBACK ---\n'
                                + traceback.format_exc(),
                                language='text',
                            )

    # ─── Оценка плана + кнопка запуска в одной карточке ───────
    selected_cities_count = 1 + random_subs  # Москва + случайные
    per_sub = (
        (1 if check_main else 0) +
        (1 if check_catalog else 0) +
        (cats_per_sub if check_categories else 0) +
        (filters_per_sub if check_filters else 0) +
        (products_per_sub if check_products else 0)
    )
    total_checks = selected_cities_count * per_sub
    # Реалистичная оценка:
    #   - 5 сек на запрос (с учётом тяжёлых страниц)
    #   - +20% буфер на возможные ретраи
    #   - +30% если проект через прокси (двойной хоп через сервер)
    base_per_request_sec = 5
    proxy_overhead = 1.30 if cfg.get('use_proxy') else 1.0
    retry_buffer = 1.20
    estimated_sec = max(1, int((total_checks / 6) * base_per_request_sec * proxy_overhead * retry_buffer))


# ═══════════════════════════════════════════════════════════════════
# КНОПКА ЗАПУСКА (общая для обоих режимов)
# ═══════════════════════════════════════════════════════════════════


if is_project or is_custom:
    can_run = False
    if is_custom:
        can_run = valid_count > 0
    elif is_project:
        can_run = total_checks > 0

    # Карточка с превью + кнопкой
    with st.container(border=True):
        if is_project:
            if total_checks == 0:
                st.warning('Не выбрано ни одного пункта для проверки')
            else:
                st.markdown(
                    f'<p style="color:var(--text-muted);font-size:0.875rem;'
                    f'margin-bottom:0.5rem;text-transform:uppercase;letter-spacing:0.05em;'
                    f'font-weight:600">Готов к запуску</p>'
                    f'<p style="font-size:1.05rem;margin-bottom:0.5rem">'
                    f'Будет проверено <strong style="color:var(--accent)">{selected_cities_count} городов</strong> × '
                    f'<strong style="color:var(--accent)">{per_sub} страниц</strong> = '
                    f'<strong style="color:var(--accent)">{total_checks} проверок</strong>'
                    f'</p>'
                    f'<p style="color:var(--text-soft);font-size:0.9rem;margin-bottom:1rem">'
                    f'Примерно <strong>{format_duration(estimated_sec)}</strong>. '
                    f'На больших каталогах или при медленных серверах может быть в 1,5–2 раза дольше.'
                    f'</p>',
                    unsafe_allow_html=True,
                )
        elif is_custom and valid_count > 0:
            # Custom-режим — без прокси, но с буфером на ретраи
            est_sec = max(1, int((valid_count / 6) * 5 * 1.20))
            st.markdown(
                f'<p style="color:var(--text-muted);font-size:0.875rem;'
                f'margin-bottom:0.5rem;text-transform:uppercase;letter-spacing:0.05em;'
                f'font-weight:600">Готов к запуску</p>'
                f'<p style="font-size:1.05rem;margin-bottom:0.5rem">'
                f'Будет проверено <strong style="color:var(--accent)">{valid_count} URL</strong>'
                f'</p>'
                f'<p style="color:var(--text-soft);font-size:0.9rem;margin-bottom:1rem">'
                f'Примерно <strong>{format_duration(est_sec)}</strong>. '
                f'Точное время зависит от скорости сайтов.'
                f'</p>',
                unsafe_allow_html=True,
            )

        if st.button(
            '▶ Запустить проверку',
            type='primary',
            disabled=not can_run,
            use_container_width=True,
            key='btn_run',
        ):
            st.session_state.is_running = True
            reset_run_state()
            st.rerun()


# ═══════════════════════════════════════════════════════════════════
# ВЫПОЛНЕНИЕ ПРОГОНА
# ═══════════════════════════════════════════════════════════════════


async def run_check_async(project_id, plan, options, on_progress):
    """Async обёртка вокруг run_batch."""
    return await run_batch(
        plan.tasks,
        concurrency=options.get('concurrency', 6),
        timeout_ms=options.get('timeout_ms', 120000),
        max_attempts=options.get('max_attempts', 3),
        retry_delay_ms=options.get('retry_delay_ms', 2500),
        check_text=options.get('check_text', True),
        on_progress=on_progress,
        proxy_url=options.get('proxy_url'),
    )


if st.session_state.is_running:
    with st.container(border=True):
        st.markdown('<h3>⏳ Идёт проверка</h3>', unsafe_allow_html=True)

        # Крупное предупреждение чтобы контент-менеджер случайно не закрыл вкладку
        st.warning(
            '⚠ **Не закрывайте вкладку до окончания проверки.** '
            'Можно переключаться на другие вкладки, но эту нужно держать открытой — '
            'если её закрыть, прогон оборвётся и отчёт не сохранится.'
        )

        # Места для прогресс-бара и счётчика
        progress_bar = st.progress(0, text='Подготовка…')
        metrics_row = st.empty()
        log_expander = st.expander('Подробный лог', expanded=False)
        log_area = log_expander.empty()
        log_messages = []

    def append_log(msg: str):
        log_messages.append(msg)
        log_area.code('\n'.join(log_messages[-100:]), language='text')

    # ─── Формируем план ────────────────────────────────────────
    started_ms = int(time.time() * 1000)
    st.session_state.run_started_at = started_ms

    # Прокси применяется только когда проект явно его разрешает (use_proxy: true).
    # Custom-режим всегда без прокси.
    # Для каждого проекта в projects/{id}.json — флаг use_proxy.
    proxy_url = None

    try:
        if is_custom:
            plan = build_custom_plan(st.session_state.custom_urls_text.split('\n'))
            project_id_for_report = 'custom'
            project_name_for_report = 'Свой список URL'
            check_text_opt = st.session_state.check_text
            append_log(f'Запуск custom-прогона: {len(plan.tasks)} URL')
            append_log('Прокси: не используется (custom-режим)')

        else:
            cfg = load_project_config(st.session_state.project_id)
            src = st.session_state.sources

            # Решение про прокси на основе use_proxy в конфиге
            if cfg.get('use_proxy'):
                proxy_url = get_proxy_url()
                if proxy_url:
                    append_log(f'Прокси: включён для проекта {cfg["name"]}')
                else:
                    append_log(f'⚠ Прокси нужен для {cfg["name"]}, но не настроен в Streamlit Secrets')
            else:
                append_log(f'Прокси: не используется (для {cfg["name"]} не требуется)')

            # ─── Автоматическое обновление почты Метрики (если креды настроены) ──
            metrika_email_auto, metrika_password_auto = get_metrika_credentials(st.session_state.project_id)
            if metrika_email_auto and metrika_password_auto:
                append_log('Обновляю 404-отчёты из почты Метрики (последние 3 дня)…')
                try:
                    metrika_proxy_auto = get_proxy_url()
                    summary = fetch_incremental(
                        project_id=st.session_state.project_id,
                        email_addr=metrika_email_auto,
                        password=metrika_password_auto,
                        folder=MAILBOX_CONFIG[st.session_state.project_id]['folder'],
                        proxy_url=metrika_proxy_auto,
                        lookback_days=3,
                        log=lambda lvl, msg: append_log(msg),
                    )
                    if summary['fetched'] > 0:
                        append_log(f'Получено новых отчётов Метрики: {summary["fetched"]}')
                    else:
                        append_log('Новых отчётов Метрики нет (всё уже было в кеше)')
                except Exception as e:
                    append_log(f'⚠ Не удалось обновить почту: {e}. Продолжаю без свежих 404-данных.')

            # Загружаем sitemap если нужны товары
            if st.session_state.check_products and not src.products:
                append_log(f'Загружаю sitemap из {cfg.get("sitemap_url")}…')
                try:
                    sm = asyncio.run(load_product_pathnames(
                        cfg,
                        [c for c in src.categories],
                        [f for f in src.filters],
                        log=lambda lvl, msg: append_log(msg),
                        proxy_url=proxy_url,
                    ))
                    src.products = sm.get('pathnames', [])
                    append_log(f'Из sitemap: {len(src.products)} товаров')
                    if sm.get('warning'):
                        append_log(f'⚠ {sm["warning"]}')
                except Exception as e:
                    append_log(f'⚠ Не удалось загрузить sitemap: {e}')

            # Загружаем историю ротации
            history = load_history(st.session_state.project_id)
            recent_paths = set(history.keys())
            append_log(f'История ротации: {len(recent_paths)} URL за последние 7 дней')

            plan = build_plan(
                src,
                random_subdomains_count=random_subs,
                categories_per_subdomain=cats_per_sub,
                filters_per_subdomain=filters_per_sub,
                products_per_subdomain=products_per_sub,
                check_main=check_main,
                check_catalog=check_catalog,
                check_categories=check_categories,
                check_filters=check_filters,
                check_products=check_products,
                mandatory_city=cfg.get('mandatory_city', 'Москва'),
                rotation_history=recent_paths,
            )
            check_text_opt = check_text
            project_id_for_report = st.session_state.project_id
            project_name_for_report = cfg['name']

            cities_str = ', '.join(s.city for s in plan.selected_subdomains)
            append_log(f'Города: {cities_str}')
            append_log(f'Всего проверок: {len(plan.tasks)}')

        # ─── Запуск прогона ────────────────────────────────────
        total = len(plan.tasks)
        counters = {'ok': 0, 'warn': 0, 'err': 0}

        def on_progress(result, done, total_n):
            if result.is_ok:
                counters['ok'] += 1
            elif result.is_warning:
                counters['warn'] += 1
            else:
                counters['err'] += 1
            # Не обновляем UI на каждый чек — Streamlit сам выводит в конце
            # (обновление UI в async-функции из Streamlit делать тяжело)

        results = asyncio.run(run_check_async(
            project_id_for_report,
            plan,
            options={
                'concurrency': 6,
                'timeout_ms': 120000,
                'max_attempts': 3,
                'retry_delay_ms': 2500,
                'check_text': check_text_opt,
                'proxy_url': proxy_url,
            },
            on_progress=on_progress,
        ))

        finished_ms = int(time.time() * 1000)
        st.session_state.run_finished_at = finished_ms

        # Сохраняем историю ротации (для обычных прогонов)
        if is_project:
            checked_paths = list(set(urlparse(r.url).path for r in results))
            save_history(st.session_state.project_id, checked_paths)

        # ─── Формируем xlsx-отчёт ──────────────────────────────
        append_log('Формирую xlsx-отчёт…')
        report_filename = make_report_filename(
            project_id_for_report, started_ms, REPORTS_DIR,
        )
        report_path = REPORTS_DIR / report_filename

        # ─── Подгружаем данные Метрики для листа «404 из Метрики» ───
        metrika_reports_for_xlsx = None
        metrika_data_date = None
        metrika_is_stale = False
        if not is_custom:
            metrika_email_check, metrika_password_check = get_metrika_credentials(st.session_state.project_id)
            if metrika_email_check and metrika_password_check:
                latest_d = get_latest_available_date(st.session_state.project_id)
                if latest_d:
                    metrika_data_date = latest_d
                    metrika_reports_for_xlsx = load_reports_for_date(
                        st.session_state.project_id, latest_d,
                    )
                    # Если самый свежий день — это НЕ вчера, ставим флаг stale
                    from datetime import datetime as _dtt, timedelta as _tdd
                    yesterday = (_dtt.now().date() - _tdd(days=1)).strftime('%Y-%m-%d')
                    metrika_is_stale = latest_d != yesterday
                    append_log(
                        f'В отчёт добавлю 404 из Метрики за {latest_d}: '
                        f'{len(metrika_reports_for_xlsx)} стран, '
                        f'{sum(r.total_pages for r in metrika_reports_for_xlsx)} страниц.'
                        + (' (данные не самые свежие)' if metrika_is_stale else '')
                    )

        build_report(
            project_name=project_name_for_report,
            started_at_ms=started_ms,
            finished_at_ms=finished_ms,
            selected_subdomains=plan.selected_subdomains,
            results=results,
            output_path=report_path,
            metrika_reports=metrika_reports_for_xlsx,
            metrika_data_date=metrika_data_date,
            metrika_is_stale=metrika_is_stale,
        )

        # ─── Автоматическая отправка в Telegram ─────────────────────────
        tg_token = get_telegram_bot_token()
        tg_recipients = get_telegram_recipients(
            st.session_state.project_id if not is_custom else None
        )
        if tg_token and tg_recipients:
            append_log(f'Отправляю уведомление в Telegram ({len(tg_recipients)} получателей)…')
            try:
                # Собираем топ-проблем — самые срочные URL
                problems_for_tg = []
                for r in results:
                    if r.is_error:
                        status_text = {
                            'not_found': '404 Не найдена',
                            'client_error': 'Ошибка на сайте',
                            'server_error': 'Сервер не отвечает',
                            'timeout': 'Нет ответа',
                            'network_error': 'Нет соединения',
                        }.get(r.status, r.status)
                        problems_for_tg.append({
                            'city': r.city or '—',
                            'url': r.url,
                            'status': status_text,
                        })
                # Топ-5 по приоритету (ошибки идут первыми)
                problems_for_tg = problems_for_tg[:5]

                # Считаем метрика-страницы
                metrika_pages_total = 0
                if metrika_reports_for_xlsx:
                    metrika_pages_total = sum(r.total_pages for r in metrika_reports_for_xlsx)

                text_issues_total = sum(
                    len(r.text_issues) for r in results if r.has_text_issues
                )

                # Структурные проблемы в контенте + пустые разделы
                content_bugs_total = sum(
                    getattr(r, 'content_bugs', 0) or 0 for r in results
                )
                content_bug_pages_total = sum(
                    1 for r in results if getattr(r, 'has_content_bugs', False)
                )
                empty_sections_list = [
                    {'city': r.city or '—', 'url': r.url}
                    for r in results
                    if getattr(r, 'content', None) is not None
                    and getattr(r.content, 'page_kind', '') == 'empty'
                ]

                from datetime import datetime as _dtt
                started_display = _dtt.fromtimestamp(started_ms / 1000).strftime('%d.%m.%Y %H:%M')
                duration_sec = (finished_ms - started_ms) // 1000

                summary_text = format_summary_message(
                    project_name=project_name_for_report,
                    started_at=started_display,
                    duration_sec=duration_sec,
                    total_checks=len(results),
                    ok_count=sum(1 for r in results if r.is_ok),
                    warn_count=sum(1 for r in results if r.is_warning),
                    err_count=sum(1 for r in results if r.is_error),
                    text_issues_count=text_issues_total,
                    metrika_pages_count=metrika_pages_total,
                    metrika_data_date=metrika_data_date,
                    top_problems=problems_for_tg,
                    content_bugs_count=content_bugs_total,
                    content_bug_pages=content_bug_pages_total,
                    empty_sections=empty_sections_list,
                )

                tg_proxy = get_proxy_url()
                tg_result = send_run_notification(
                    bot_token=tg_token,
                    recipients=tg_recipients,
                    project_name=project_name_for_report,
                    summary_text=summary_text,
                    report_file=report_path,
                    proxy_url=tg_proxy,
                    log=lambda lvl, msg: append_log(msg),
                )

                if tg_result['sent'] > 0:
                    append_log(
                        f'✓ Telegram: отправлено {tg_result["sent"]} получателям'
                        + (f', не доставлено {tg_result["failed"]}' if tg_result['failed'] else '')
                    )
                else:
                    append_log(
                        f'⚠ Telegram: не удалось отправить никому. '
                        f'Ошибки: {tg_result["errors"]}'
                    )
            except Exception as e:
                append_log(f'⚠ Telegram-отправка упала: {e}. Продолжаю.')
        elif tg_token and not tg_recipients:
            append_log(
                f'Telegram-уведомление пропущено: '
                f'для проекта {st.session_state.project_id} нет получателей в Secrets '
                f'(ключ telegram_recipients_{st.session_state.project_id}).'
            )
        # Если нет токена — вообще ничего не пишем в лог, это нормальный сценарий

        st.session_state.run_results = results
        st.session_state.run_report_path = str(report_path)
        st.session_state.is_running = False

        append_log(f'Готово. Отчёт: {report_filename}')

        progress_bar.progress(1.0, text='Готово')
        st.rerun()

    except Exception as e:
        st.error(f'Ошибка: {e}')
        st.session_state.is_running = False
        append_log(f'❌ Ошибка: {e}')


# ═══════════════════════════════════════════════════════════════════
# РЕЗУЛЬТАТ ПРОГОНА (после завершения)
# ═══════════════════════════════════════════════════════════════════


if st.session_state.run_results and not st.session_state.is_running:
    results = st.session_state.run_results
    total = len(results)
    ok_count = sum(1 for r in results if r.is_ok)
    warn_count = sum(1 for r in results if r.is_warning)
    err_count = total - ok_count - warn_count
    text_issues_count = sum(len(r.text_issues) for r in results if r.has_text_issues)
    duration = (st.session_state.run_finished_at - st.session_state.run_started_at) // 1000

    # ─── Главная карточка результатов ─────────────────────────
    with st.container(border=True):
        # Зелёная плашка с заголовком
        any_problems = err_count > 0 or warn_count > 0 or text_issues_count > 0
        if not any_problems:
            st.markdown(
                f'<div style="background:linear-gradient(180deg, rgba(80, 227, 194, 0.10) 0%, transparent 100%);'
                f'border-left:3px solid var(--ok);padding:14px 18px;border-radius:8px;margin-bottom:1rem">'
                f'<p style="margin:0;font-size:1.1rem"><strong style="color:var(--ok)">✓ Все проверки прошли успешно</strong></p>'
                f'<p style="margin:6px 0 0 0;color:var(--text-soft)">'
                f'Проверено {total} страниц за {format_duration(duration)}. Проблем не найдено.</p>'
                f'</div>',
                unsafe_allow_html=True,
            )
        else:
            problems_summary = []
            if err_count > 0:
                problems_summary.append(f'{err_count} {"страница не работает" if err_count == 1 else "страниц не работают"}')
            if warn_count > 0:
                problems_summary.append(f'{warn_count} с предупреждениями')
            if text_issues_count > 0:
                problems_summary.append(f'{text_issues_count} битых переменных')
            st.markdown(
                f'<div style="background:linear-gradient(180deg, rgba(245, 166, 35, 0.10) 0%, transparent 100%);'
                f'border-left:3px solid var(--warn);padding:14px 18px;border-radius:8px;margin-bottom:1rem">'
                f'<p style="margin:0;font-size:1.1rem"><strong style="color:var(--warn)">Найдены проблемы</strong></p>'
                f'<p style="margin:6px 0 0 0;color:var(--text-soft)">'
                f'{", ".join(problems_summary)}. Проверено {total} страниц за {format_duration(duration)}.</p>'
                f'</div>',
                unsafe_allow_html=True,
            )

        # Метрики
        c1, c2, c3, c4 = st.columns(4)
        c1.metric('Всего', total)
        c2.metric('✅ Работает', ok_count)
        c3.metric('⚠ Предупреждений', warn_count)
        c4.metric('❌ Не работает', err_count, delta_color='inverse')

        # ─── ВЫДЕЛЕННАЯ кнопка скачивания ─────────────────────
        if st.session_state.run_report_path:
            report_path = Path(st.session_state.run_report_path)
            if report_path.exists():
                st.markdown('<div style="margin-top:1rem"></div>', unsafe_allow_html=True)
                with open(report_path, 'rb') as f:
                    st.download_button(
                        label=f'📥 Скачать полный отчёт ({report_path.name})',
                        data=f.read(),
                        file_name=report_path.name,
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        use_container_width=True,
                        type='primary',
                    )
                st.caption(f'В отчёте: все проверки в формате xlsx с фильтрами по статусу')

        # ─── Индикатор Telegram-отправки (если настроен) ──────────
        tg_token = get_telegram_bot_token()
        if tg_token and not is_custom and st.session_state.project_id:
            tg_recipients = get_telegram_recipients(st.session_state.project_id)
            if tg_recipients:
                st.markdown(
                    f'<p style="color:var(--text-muted);font-size:0.85rem;'
                    f'margin-top:0.75rem;text-align:center">'
                    f'📱 Уведомление с отчётом отправлено в Telegram '
                    f'({len(tg_recipients)} {"получатель" if len(tg_recipients) == 1 else "получателей"})'
                    f'</p>',
                    unsafe_allow_html=True,
                )

    # ─── Список проблем в отдельной карточке ──────────────
    problems = [r for r in results if r.is_error or r.is_warning or r.has_text_issues]
    if problems:
        with st.container(border=True):
            st.markdown(
                f'<h3 style="margin-bottom:1rem">Список проблем '
                f'<span style="color:var(--text-muted);font-weight:400;font-size:0.95rem">({len(problems)})</span>'
                f'</h3>',
                unsafe_allow_html=True,
            )
            status_labels = {
                'ok': 'Работает',
                'redirect': 'Перенаправление',
                'not_found': 'Страница не найдена',
                'client_error': 'Ошибка на сайте',
                'server_error': 'Сервер не отвечает',
                'timeout': 'Нет ответа',
                'network_error': 'Нет соединения',
            }
            for r in problems[:50]:
                if r.is_error:
                    emoji = '❌'
                    color = 'var(--err)'
                elif r.is_warning:
                    emoji = '⚠️'
                    color = 'var(--warn)'
                else:
                    emoji = '🔤'
                    color = 'var(--warn)'

                status_text = status_labels.get(r.status, r.status)
                extra = ''
                if r.has_text_issues:
                    extra = f' · <span style="color:var(--warn)">{len(r.text_issues)} битых переменных</span>'

                city_part = f'[{r.city}] ' if r.city else ''
                st.markdown(
                    f'<div style="padding:8px 0;border-bottom:1px solid var(--border);font-size:0.92rem">'
                    f'{emoji} <strong>{city_part}</strong>{r.type_label}: '
                    f'<a href="{r.url}" target="_blank" style="color:var(--accent);text-decoration:none">{r.url}</a> '
                    f'— <span style="color:{color}">{status_text}</span>{extra}'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            if len(problems) > 50:
                st.caption(f'... и ещё {len(problems) - 50}. Все детали — в xlsx-отчёте.')



# ── Футер ──────────────────────────────────────────────────────────


st.divider()
st.caption(
    'Если приложение не открывалось дольше недели, при первом заходе '
    'может появиться сообщение «Yes, get this app back up» — нажмите на эту '
    'кнопку и подождите 30–60 секунд, пока сервис проснётся.'
)
st.caption(
    '_Проверка работоспособности сайтов · '
    'Site Checker v2.0 (Python + Streamlit)_'
)
